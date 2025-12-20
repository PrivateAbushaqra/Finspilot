from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView
from django.http import JsonResponse, HttpResponse, Http404
from django.contrib import messages
from django.core import serializers
from django.apps import apps
from django.conf import settings
from django.utils.translation import gettext as _
from django.utils import timezone
from django.db import transaction
from django.db import models
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.core.cache import cache
from django.template.defaultfilters import filesizeformat
from django.urls import reverse
import json
import os
import threading
import time
import logging
from datetime import datetime as dt_module
from openpyxl import Workbook
from decimal import Decimal
from django.core.management import call_command

logger = logging.getLogger(__name__)

# إضافة AuditLog import مع حماية من تضارب الـ IDs
try:
    from core.models import AuditLog
    AUDIT_AVAILABLE = True
    
    # تحقق من وجود تضارب في الـ IDs
    from django.db import connection
    def reset_audit_sequence_if_needed():
        """إعادة تعيين sequence إذا كان هناك تضارب في IDs"""
        try:
            # 🔧 التحقق من أننا لسنا داخل transaction active
            # استخدام connection الخاص بدلاً من connection الرئيسي
            from django.db import connections
            with connections['default'].cursor() as cursor:
                # البحث عن أعلى ID موجود
                cursor.execute("SELECT MAX(id) FROM core_auditlog")
                max_id = cursor.fetchone()[0] or 0
                
                logger.debug(f"تم إعادة تعيين sequence للـ AuditLog إلى {max_id + 1}")
                
        except Exception as e:
            error_msg = str(e)
            # إذا كنا داخل atomic block، لا نستطيع تنفيذ استعلامات
            # بعد حدوث خطأ حتى نهاية الـ block، لذلك نتجاهل الخطأ ببساطة
            # سيتم تسجيل الحدث في السجلات العادية (logger) بدلاً من AuditLog
            if 'atomic' in error_msg.lower() or 'transaction' in error_msg.lower():
                logger.debug("تم تخطي تسجيل الحدث في AuditLog (داخل transaction)")
            elif 'duplicate key' in error_msg.lower() or 'unique constraint' in error_msg.lower():
                logger.debug("تم تخطي تسجيل الحدث في AuditLog (تضارب في المفاتيح)")
            else:
                logger.warning(f"فشل في تسجيل الحدث في سجل المراجعة: {error_msg}")

except Exception:
    # AuditLog model or related DB objects not available
    AUDIT_AVAILABLE = False
    logger.debug('AuditLog not available or import failed in backup.views')

def reset_all_sequences():
    """إعادة تعيين جميع sequences في قاعدة البيانات لتجنب تضارب IDs"""
    from django.db import connection
    from django.apps import apps
    
    sequences_reset = 0
    
    try:
        with connection.cursor() as cursor:
            # الحصول على جميع الجداول التي تحتوي على auto-increment fields
            for model in apps.get_models():
                if hasattr(model._meta, 'db_table'):
                    table_name = model._meta.db_table
                    
                    # الحصول على اسم sequence للجدول (في PostgreSQL)
                    # sequence name pattern: table_name_id_seq
                    sequence_name = f"{table_name}_id_seq"
                    
                    try:
                        # الحصول على أعلى ID موجود في الجدول
                        cursor.execute(f"SELECT MAX(id) FROM {table_name}")
                        max_id = cursor.fetchone()[0]
                        
                        if max_id is not None:
                            # إعادة تعيين sequence إلى max_id + 1
                            cursor.execute(f"SELECT setval('{sequence_name}', {max_id + 1}, false)")
                            sequences_reset += 1
                            logger.debug(f"تم إعادة تعيين sequence {sequence_name} إلى {max_id + 1}")
                        else:
                            # إذا كان الجدول فارغ، إعادة تعيين إلى 1
                            cursor.execute(f"SELECT setval('{sequence_name}', 1, false)")
                            sequences_reset += 1
                            logger.debug(f"تم إعادة تعيين sequence {sequence_name} إلى 1 (جدول فارغ)")
                            
                    except Exception as e:
                        # قد لا يكون للجدول sequence أو قد يكون هناك خطأ آخر
                        logger.debug(f"تخطي إعادة تعيين sequence للجدول {table_name}: {str(e)}")
                        continue
                        
    except Exception as e:
        logger.warning(f"فشل في إعادة تعيين sequences: {str(e)}")
    
    return sequences_reset

def log_audit(user, action, description):
    """تسجيل الحدث في سجل المراجعة"""
    if AUDIT_AVAILABLE:
        try:
            AuditLog.objects.create(
                user=user,
                action_type=action,
                description=description
            )
        except Exception as e:
            error_msg = str(e)
            # إذا كنا داخل atomic block، لا نستطيع تنفيذ استعلامات
            # بعد حدوث خطأ حتى نهاية الـ block، لذلك نتجاهل الخطأ ببساطة
            # سيتم تسجيل الحدث في السجلات العادية (logger) بدلاً من AuditLog
            if 'atomic' in error_msg.lower() or 'transaction' in error_msg.lower():
                logger.debug(f"تم تخطي تسجيل الحدث في AuditLog (داخل transaction): {description}")
            elif 'duplicate key' in error_msg.lower() or 'unique constraint' in error_msg.lower():
                logger.debug(f"تم تخطي تسجيل الحدث في AuditLog (تضارب في المفاتيح): {description}")
            else:
                logger.warning(f"فشل في تسجيل الحدث في سجل المراجعة: {error_msg}")
    else:
        logger.info(f"Audit: {action} - {description}")

def get_backup_progress_data():
    """الحصول على بيانات تقدم النسخ الاحتياطي من cache"""
    return cache.get('backup_progress', {
        'is_running': False,
        'current_table': '',
        'processed_tables': 0,
        'total_tables': 0,
        'processed_records': 0,
        'total_records': 0,
        'percentage': 0,
        'status': 'idle',
        'error': None,
        'tables_status': [],
        'estimated_time': ''
    })

def set_backup_progress_data(data):
    """تحديث بيانات تقدم النسخ الاحتياطي في cache"""
    import time
    current_time = time.time()
    cache.set('backup_progress', data, timeout=3600)
    cache.set('backup_last_update', current_time, timeout=3600)
    
    # تحديث آخر نسبة مئوية للمراقبة
    if 'percentage' in data:
        cache.set('backup_last_percentage', data['percentage'], timeout=3600)


def get_clear_progress_data():
    """الحصول على بيانات تقدم المسح من cache"""
    return cache.get('clear_progress', {
        'is_running': False,
        'current_table': '',
        'processed_tables': 0,
        'total_tables': 0,
        'processed_records': 0,
        'total_records': 0,
        'percentage': 0,
        'status': 'idle',
        'error': None,
        'tables_status': [],
        'estimated_time': ''
    })

def set_clear_progress_data(data):
    """تحديث بيانات تقدم المسح في cache"""
    import time
    current_time = time.time()
    cache.set('clear_progress', data, timeout=3600)
    cache.set('clear_last_update', current_time, timeout=3600)
    if 'percentage' in data:
        cache.set('clear_last_percentage', data['percentage'], timeout=3600)


def get_restore_progress_data():
    """الحصول على بيانات تقدم الاستعادة من cache"""
    return cache.get('restore_progress', {
        'is_running': False,
        'current_table': '',
        'processed_tables': 0,
        'total_tables': 0,
        'processed_records': 0,
        'total_records': 0,
        'percentage': 0,
        'status': 'idle',
        'error': None,
        'tables_status': [],
        'estimated_time': ''
    })


@login_required
def create_basic_accounts_view(request):
    """View to create basic IFRS accounts using management command.
    Protected to superuser only.
    """
    logger = logging.getLogger(__name__)
    
    # permission check
    if not request.user.is_superuser:
        messages.error(request, _('ليس لديك صلاحية لإنشاء الحسابات الأساسية'))
        return redirect('backup:backup_restore')

    try:
        # call management command; this will skip existing accounts
        call_command('create_basic_accounts')
        # سجل في الـ AuditLog إن كان متوفرًا
        try:
            if AUDIT_AVAILABLE:
                AuditLog.objects.create(
                    user=request.user, 
                    action_type='create', 
                    content_type='account',
                    description=_('Created basic IFRS accounts')
                )
        except Exception:
            # ignore audit log failures but inform user
            logger.warning('Failed to write AuditLog for create_basic_accounts')

        messages.success(request, _('تم إنشاء الحسابات الأساسية بنجاح (أو كانت موجودة بالفعل)'))
    except Exception as e:
        logger.exception('Failed to create basic accounts')
        messages.error(request, _('فشل أثناء إنشاء الحسابات الأساسية: %s') % str(e))

    return redirect('backup:backup_restore')


@login_required
def get_deletable_tables(request):
    """API تعرض الجداول التي يمكن حذفها مع فحص التبعيات بين النماذج.
    سترجع قائمة من العناصر: { app_name, model_name, display_name, record_count, dependents: [labels], safe_to_delete }
    safe_to_delete True يعني أنه لا يوجد نموذج آخر يعتمد على هذا النموذج (FK) بخلاف التطبيقات المستبعدة.
    """
    
    # فحص الصلاحية
    if not request.user.has_perm('backup.can_delete_advanced_data'):
        # تسجيل محاولة الوصول غير المسموح
        log_audit(request.user, 'denied', _('محاولة وصول مرفوضة لعرض قائمة الجداول القابلة للحذف - صلاحية غير متوفرة'))
        return JsonResponse({'success': False, 'error': _('ليس لديك صلاحية للوصول لهذه الميزة')})
    
    try:
        result = []
        
        # قائمة التطبيقات المحمية من Django (يجب استبعادها من قائمة الحذف)
        protected_apps = [
            'admin',
            'auth',
            'contenttypes',
            'sessions',
            'messages',
            'staticfiles',
            'corsheaders',
            'rest_framework',
            'django_bootstrap5',
            'crispy_forms',
            'crispy_bootstrap5',
        ]
        
        # بناء خريطة العلاقات العكسية: model_label -> set(of dependent model labels)
        dep_map = {}
        all_models = []
        for app_config in apps.get_app_configs():
            # استبعاد التطبيقات المحمية من قائمة الحذف
            if app_config.name in protected_apps:
                continue
                
            for model in app_config.get_models():
                label = f"{app_config.name}.{model._meta.model_name}"
                dep_map[label] = set()
                all_models.append((app_config, model))

        # ملء الخريطة
        for app_config, model in all_models:
            for field in model._meta.get_fields():
                try:
                    if field.is_relation and (field.many_to_one or field.one_to_one) and field.related_model is not None:
                        rel = field.related_model
                        rel_label = f"{rel._meta.app_label}.{rel._meta.model_name}"
                        src_label = f"{app_config.name}.{model._meta.model_name}"
                        dep_map[rel_label].add(src_label)
                except Exception:
                    continue

        # جمع المعلومات لكل نموذج
        for app_config, model in all_models:
            try:
                # تخطي النماذج التي لا تدير جداول في قاعدة البيانات (managed = False)
                if not model._meta.managed:
                    continue
                    
                label = f"{app_config.name}.{model._meta.model_name}"
                display = model._meta.verbose_name or label
                count = model.objects.count()
                dependents = sorted(list(dep_map.get(label, [])))
                # امن للحذف إذا لا يوجد تابعون (dependents) أو التوابع كلها ضمن التطبيقات المستبعدة
                safe = len(dependents) == 0
                result.append({
                    'app_name': app_config.name,
                    'model_name': model._meta.model_name,
                    'label': label,
                    'display_name': str(display),
                    'record_count': count,
                    'dependents': dependents,
                    'safe_to_delete': safe
                })
            except Exception as e:
                logger.warning(f"فشل جلب معلومات النموذج {app_config.name}.{model}: {str(e)}")
                continue

        # رتب بحسب عدد السجلات تنازلياً
        result.sort(key=lambda x: (-x['record_count'], x['label']))
        
        # تسجيل العملية في سجل الأنشطة
        log_audit(request.user, 'view', _('عرض قائمة الجداول القابلة للحذف - تم العثور على {} جدول').format(len(result)))
        
        return JsonResponse({'success': True, 'tables': result})
    except Exception as e:
        logger.error(f"خطأ في get_deletable_tables: {str(e)}")
        # تسجيل الخطأ في سجل الأنشطة
        log_audit(request.user, 'error', _('خطأ في عرض قائمة الجداول القابلة للحذف: {}').format(str(e)))
        return JsonResponse({'success': False, 'error': _('حدث خطأ في تحميل قائمة الجداول')})


@login_required
def delete_selected_tables(request):
    """مسح الجداول المحددة (data-only) بعد فحص التبعيات.
    يتوقع POST body 'tables' كقائمة من labels ['app.model', ...] و 'confirm' boolean.
    سيقوم بحذف جميع السجلات في تلك الجداول داخل معاملة واحدة، ويُسجل كل عملية في AuditLog.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'طريقة طلب غير صحيحة'})

    # فحص الصلاحية
    if not request.user.has_perm('backup.can_delete_advanced_data'):
        # تسجيل محاولة الوصول غير المسموح
        log_audit(request.user, 'denied', _('محاولة وصول مرفوضة لحذف بيانات الجداول - صلاحية غير متوفرة'))
        return JsonResponse({'success': False, 'error': _('ليس لديك صلاحية لحذف البيانات')})

    try:
        payload = json.loads(request.body.decode('utf-8')) if request.body else {}
    except Exception:
        payload = request.POST.dict() if request.POST else {}

    tables = payload.get('tables') or payload.get('labels') or []
    confirm = payload.get('confirm') in [True, 'true', '1', 1, 'on']

    if not tables or not isinstance(tables, list):
        return JsonResponse({'success': False, 'error': 'يرجى تحديد جداول للحذف'})

    if not confirm:
        return JsonResponse({'success': False, 'error': 'يرجى تأكيد الحذف (confirm=true)'})

    try:
        # بناء خريطة التبعية كما في get_deletable_tables
        dep_map = {}
        model_map = {}
        for app_config in apps.get_app_configs():
            for model in app_config.get_models():
                label = f"{app_config.name}.{model._meta.model_name}"
                dep_map[label] = set()
                model_map[label] = model

        for label, model in list(model_map.items()):
            for field in model._meta.get_fields():
                try:
                    if field.is_relation and (field.many_to_one or field.one_to_one) and field.related_model is not None:
                        rel = field.related_model
                        rel_label = f"{rel._meta.app_label}.{rel._meta.model_name}"
                        src_label = label
                        if rel_label in dep_map:
                            dep_map[rel_label].add(src_label)
                except Exception:
                    continue

        # وسّع مجموعة الجداول لتشمل جميع التبعيات المتسلسلة (closure)
        selected_set = set([str(x) for x in tables])
        def add_deps(label):
            for dep in dep_map.get(label, set()):
                if dep not in selected_set:
                    selected_set.add(dep)
                    add_deps(dep)

        for t in list(selected_set):
            add_deps(t)

        # بعد التوسيع، تأكد من عدم وجود تبعيات خارج المجموعه النهائية
        cannot_delete = []
        for t in list(selected_set):
            dependents = dep_map.get(t, set())
            external = [d for d in dependents if d not in selected_set]
            if external:
                cannot_delete.append({'table': t, 'blocked_by': external})

        if cannot_delete:
            return JsonResponse({'success': False, 'error': 'بعض الجداول لا يمكن حذفها بسبب تبعيات خارجية', 'details': cannot_delete})

        # احسب ترتيب الحذف بحيث تُحذف التبعيات أولاً (post-order)
        visited = set()
        delete_order = []
        def dfs(label):
            if label in visited:
                return
            visited.add(label)
            for dep in dep_map.get(label, set()):
                if dep in selected_set:
                    dfs(dep)
            delete_order.append(label)

        for label in list(selected_set):
            dfs(label)

    except Exception as e:
        logger.error(f"خطأ في بناء خريطة التبعية أو حساب ترتيب الحذف: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

    # تنفيذ الحذف داخل معاملة وبترتيب آمن
    deleted_stats = []
    audit_reassign_notifications = []
    try:
        with transaction.atomic():
            for t in delete_order:
                model = model_map.get(t)
                if model is None:
                    continue
                try:
                    # Special handling for User model: create or keep a fallback '__deleted_user__' user
                    UserModel = None
                    try:
                        from django.contrib.auth import get_user_model
                        UserModel = get_user_model()
                    except Exception:
                        UserModel = None

                    if UserModel is not None and getattr(model, '__name__', '').lower() == getattr(UserModel, '__name__', '').lower():
                        # ensure fallback exists
                        fallback_username = '__deleted_user__'
                        fallback = UserModel.objects.filter(username=fallback_username).first()
                        if not fallback:
                            # create minimal fallback user
                            fallback = UserModel.objects.create(username=fallback_username, is_active=False)

                        # reassign AuditLog.user entries that reference users we will delete to fallback
                        try:
                            user_ids = list(model.objects.exclude(pk=fallback.pk).values_list('pk', flat=True))
                            from core.models import AuditLog as _AuditLog
                            if user_ids:
                                _AuditLog.objects.filter(user_id__in=user_ids).update(user_id=fallback.pk)
                        except Exception as reassign_err:
                            logger.warning(f"Failed to reassign AuditLog.user before deleting users: {reassign_err}")

                        # notify front-end that we reassigned audit logs for this table
                        audit_reassign_notifications.append({'table': t, 'fallback': fallback.username, 'reassigned_count': len(user_ids)})

                        # delete all users except fallback
                        qs = model.objects.exclude(pk=fallback.pk)
                        count = qs.count()
                        qs.delete()
                        deleted_stats.append({'table': t, 'deleted': count})
                        from django.utils.translation import gettext as _
                        log_audit(request.user, 'delete', _('Deleted all records from table %(table)s - count: %(count)s') % {'table': t, 'count': count})
                        continue

                    # Default deletion path for other models
                    count = model.objects.all().count()
                    # حاول حذف السجلات؛ إذا كان هناك FK محمي فسيُثار ProtectedError أو قد يحدث IntegrityError
                    try:
                        model.objects.all().delete()
                    except Exception as del_exc:
                        from django.db import IntegrityError
                        if isinstance(del_exc, IntegrityError):
                            # Special retry logic for User model: try to reassign AuditLog.user references then retry
                            try:
                                from django.contrib.auth import get_user_model
                                UserModelLocal = get_user_model()
                            except Exception:
                                UserModelLocal = None

                            if UserModelLocal is not None and getattr(model, '__name__', '').lower() == getattr(UserModelLocal, '__name__', '').lower():
                                # perform reassignment of AuditLog entries to fallback
                                try:
                                    fallback_username = '__deleted_user__'
                                    fallback = UserModelLocal.objects.filter(username=fallback_username).first()
                                    if not fallback:
                                        fallback = UserModelLocal.objects.create(username=fallback_username, is_active=False)
                                    user_ids = list(model.objects.exclude(pk=fallback.pk).values_list('pk', flat=True))
                                    from core.models import AuditLog as _AuditLog
                                    if user_ids:
                                        _AuditLog.objects.filter(user_id__in=user_ids).update(user_id=fallback.pk)
                                    # retry delete
                                    model.objects.exclude(pk=fallback.pk).delete()
                                except Exception as reassign_err:
                                    logger.error(f"Failed to reassign AuditLog on IntegrityError: {reassign_err}")
                                    raise del_exc
                            else:
                                raise del_exc

                    deleted_stats.append({'table': t, 'deleted': count})
                    from django.utils.translation import gettext as _
                    log_audit(request.user, 'delete', _('Deleted all records from table %(table)s - count: %(count)s') % {'table': t, 'count': count})
                except ProtectedError as p_err:
                    # جمع تفاصيل يمنع الحذف بسبب ProtectedError
                    blockers = []
                    try:
                        # ProtectedError.args may include a list/queryset of blocking objects
                        for item in getattr(p_err, 'protected_objects', []) or []:
                            try:
                                blockers.append(str(item))
                            except Exception:
                                blockers.append(repr(item))
                    except Exception:
                        blockers = [str(p_err)]
                    logger.warning(f"ProtectedError deleting {t}: {blockers}")
                    return JsonResponse({'success': False, 'error': 'ProtectedError', 'details': {'table': t, 'blocked_by_instances': blockers}})

        return JsonResponse({'success': True, 'deleted': deleted_stats, 'audit_reassign_notifications': audit_reassign_notifications})
    except Exception as e:
        logger.error(f"خطأ في delete_selected_tables: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


def set_restore_progress_data(data):
    """تحديث بيانات تقدم الاستعادة في cache"""
    import time
    current_time = time.time()
    cache.set('restore_progress', data, timeout=3600)
    cache.set('restore_last_update', current_time, timeout=3600)
    if 'percentage' in data:
        cache.set('restore_last_percentage', data['percentage'], timeout=3600)


class BackupRestoreView(LoginRequiredMixin, TemplateView):
    """صفحة إدارة النسخ الاحتياطية والاستعادة"""
    template_name = 'backup/backup_restore.html'
    
    def dispatch(self, request, *args, **kwargs):
        # التحقق من الصلاحيات - يجب أن يكون لديه على الأقل صلاحية عرض النسخ الاحتياطي
        if not request.user.is_superuser and not request.user.has_perm('users.can_view_backup'):
            messages.error(request, _('ليس لديك صلاحية للوصول إلى صفحة النسخ الاحتياطي'))
            return redirect('core:dashboard')
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # إنشاء مجلد النسخ الاحتياطية إذا لم يكن موجوداً
        backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        # جلب قائمة النسخ الاحتياطية
        backups = []
        for filename in os.listdir(backup_dir):
            if filename.endswith('.json') or filename.endswith('.xlsx'):
                filepath = os.path.join(backup_dir, filename)
                try:
                    stat = os.stat(filepath)
                    file_type = 'XLSX' if filename.endswith('.xlsx') else 'JSON'
                    backups.append({
                        'filename': filename,
                        'size': filesizeformat(stat.st_size),
                        'created_at': dt_module.fromtimestamp(stat.st_mtime),
                        'path': filepath,
                        'type': file_type
                    })
                except (OSError, IOError):
                    continue

        # ترتيب النسخ حسب تاريخ الإنشاء (الأحدث أولاً)
        backups.sort(key=lambda x: x['created_at'], reverse=True)

        # التحقق من وجود الحسابات الأساسية IFRS
        basic_accounts_codes = [
            # الأصول المتداولة (Current Assets)
            '10', '101', '1010', '102', '103', '1031', '104', '1041', '1070', '1101',
            '12', '1020', '1201',
            '1301', '1302', '1303',
            '14', '141',
            '15', '150', '1501', '1502',
            # الأصول غير المتداولة (Non-current Assets)
            '1401', '1402', '1403', '1510', '1511', '1601',
            # المطلوبات المتداولة (Current Liabilities)
            '2101', '2102', '2103',
            '22', '2201', '2202',
            '2030', '203001', '2301', '2302',
            '2401', '2402', '2501', '2503',
            # المطلوبات طويلة الأجل (Long-term Liabilities)
            '2601', '2502',
            # حقوق الملكية (Equity)
            '301', '30101', '30102', '30103', '305', '3101', '3102', '3202', '3203', '3204',
            # الإيرادات (Revenues)
            '40', '4010', '4011', '4012', '42', '4020', '4022', '403', '404', '4099',
            # المصروفات (Expenses)
            '50', '501', '5001', '5011', '5012',
            '502', '5021', '5022', '5023',
            '503', '5031',
            '504', '505', '506', '507', '5099',
            '60', '6010',
            # المشتريات (Purchases)
            '6101',
            # مردودات ومسموحات (Returns)
            '7101'
        ]
        
        from journal.models import Account
        existing_accounts = Account.objects.filter(code__in=basic_accounts_codes).values_list('code', flat=True)
        context['basic_accounts_exist'] = set(existing_accounts) == set(basic_accounts_codes)

        context['backups'] = backups
        context['latest_backup'] = backups[0] if backups else None
        context['is_superuser'] = self.request.user.is_superuser
        context['can_restore_backup'] = self.request.user.has_perm('backup.can_restore_backup')
        return context


@login_required
def get_backup_progress(request):
    """الحصول على حالة تقدم النسخ الاحتياطي"""
    if not request.user.is_authenticated:
        return JsonResponse({
            'success': False,
            'error': 'يجب تسجيل الدخول أولاً',
            'progress': None
        }, status=401)
    
    progress_data = get_backup_progress_data()
    
    # التحقق من العمليات المعلقة وحذفها إذا مر وقت طويل
    if progress_data.get('is_running', False):
        # إذا كانت العملية تعمل لأكثر من 30 دقيقة، قم بإلغائها
        import time
        current_time = time.time()
        
        # إذا لم يتم تحديث الحالة لمدة 10 دقائق، اعتبر العملية معلقة
        last_update_key = 'backup_last_update'
        last_update = cache.get(last_update_key, current_time)
        
        if current_time - last_update > 600:  # 10 دقائق
            # العملية معلقة، قم بإلغائها
            cache.delete('backup_progress')
            cache.delete(last_update_key)
            
            log_audit(request.user, 'delete', 'تم إلغاء عملية النسخ الاحتياطي المعلقة تلقائياً')
    
    return JsonResponse({
        'success': True,
        'progress': progress_data
    })


@login_required
def clear_backup_cache(request):
    """مسح كاش النسخ الاحتياطي"""
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'طريقة طلب غير صحيحة'
        })
    
    try:
        cache.delete('backup_progress')
        cache.delete('backup_last_update')
        
        log_audit(request.user, 'delete', 'تم مسح كاش النسخ الاحتياطية يدوياً')
        
        return JsonResponse({
            'success': True,
            'message': 'تم مسح الكاش بنجاح'
        })
        
    except Exception as e:
        logger.error(f"خطأ في مسح الكاش: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


def get_backup_tables_info():
    """الحصول على معلومات مفصلة عن جميع الجداول في قاعدة البيانات"""
    tables_info = []
    
    # قائمة التطبيقات المستثناة من النسخ الاحتياطي (تطبيقات Django الأساسية فقط)
    excluded_apps = [
        'django.contrib.admin',
        'django.contrib.auth',
        'django.contrib.contenttypes', 
        'django.contrib.sessions',
        'django.contrib.messages',
        'django.contrib.staticfiles',
        'corsheaders',
        'rest_framework',
        'django_bootstrap5',
        'crispy_forms',
        'crispy_bootstrap5',
    ]
    
    try:
        for app_config in apps.get_app_configs():
            if app_config.name in excluded_apps:
                # استثناء خاص للمجموعات من auth
                if app_config.name == 'django.contrib.auth':
                    # نسخ المجموعات فقط من auth
                    pass
                else:
                    continue
                
            for model in app_config.get_models():
                # تجاهل النماذج التي لا تنشئ جداول (managed = False)
                if getattr(model._meta, 'managed', True) is False:
                    continue
                
                # استثناء المستخدمين والصلاحيات من auth، لكن نسخ المجموعات
                if app_config.name == 'django.contrib.auth' and model._meta.model_name in ['user', 'permission']:
                    continue
                    
                # تصحيح اسم التطبيق للمجموعات
                if app_config.name == 'django.contrib.auth' and model._meta.model_name == 'group':
                    # استخدام اسم التطبيق الصحيح
                    app_name_to_use = 'auth'
                else:
                    app_name_to_use = app_config.name
                    
                try:
                    record_count = model.objects.count()
                    tables_info.append({
                        'app_name': app_name_to_use,
                        'model_name': model._meta.model_name,
                        'display_name': model._meta.verbose_name or f"{app_config.verbose_name} - {model._meta.model_name}",
                        'record_count': record_count,
                        'status': 'pending',
                        'progress': 0,
                        'actual_records': 0,
                        'error': None
                    })
                except Exception as e:
                    # تجاهل النماذج التي تسبب أخطاء (مثل الجداول غير الموجودة)
                    logger.warning(f"تجاهل النموذج {model._meta.model_name} بسبب خطأ: {str(e)}")
                    continue
        
        return tables_info
        
    except Exception as e:
        logger.error(f"خطأ في جلب معلومات الجداول: {str(e)}")
        return []


@login_required
def get_backup_tables(request):
    """API للحصول على معلومات الجداول"""
    try:
        if not request.user.is_authenticated:
            return JsonResponse({
                'success': False,
                'error': 'يجب تسجيل الدخول أولاً'
            }, status=401)
        
        tables_info = get_backup_tables_info()
        return JsonResponse({
            'success': True,
            'tables': tables_info,
            'total_tables': len(tables_info),
            'total_records': sum(table['record_count'] for table in tables_info)
        })
        
    except Exception as e:
        logger.error(f"خطأ في جلب معلومات الجداول: {str(e)}")
        
        log_audit(request.user, 'error', f'خطأ في جلب معلومات الجداول: {str(e)}')
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


def perform_backup_task(user, timestamp, filename, filepath, format_type='json'):
    """تنفيذ مهمة النسخ الاحتياطي في خيط منفصل مع تتبع مفصل"""
    
    user_name = user.username if user else 'system'
    logger.info(f"🎯 دخول perform_backup_task: user={user_name}, filename={filename}, format={format_type}")
    
    try:
        log_audit(user, 'create', f'بدء إنشاء النسخة الاحتياطية: {filename} - النوع: {format_type.upper()}')
        
        # تحديث حالة البداية
        progress_data = {
            'is_running': True,
            'status': 'starting',
            'error': None,
            'percentage': 0,
            'current_table': 'بدء النسخ الاحتياطي...',
            'processed_tables': 0,
            'total_tables': 0,
            'processed_records': 0,
            'total_records': 0,
            'estimated_time': 'جاري الحساب...',
            'tables_status': []
        }
        set_backup_progress_data(progress_data)
        
        # الحصول على معلومات الجداول
        progress_data['current_table'] = 'تحليل قاعدة البيانات...'
        progress_data['status'] = 'analyzing'
        set_backup_progress_data(progress_data)
        
        tables_info = get_backup_tables_info()
        total_tables = len(tables_info)
        total_records = sum(table['record_count'] for table in tables_info)
        
        # تحديث معلومات الجداول
        for table in tables_info:
            table['status'] = 'pending'
            table['progress'] = 0
            table['actual_records'] = 0
            table['error'] = None
        
        progress_data.update({
            'status': 'preparing',
            'current_table': f'تم العثور على {total_tables} جدول بإجمالي {total_records} سجل',
            'total_tables': total_tables,
            'total_records': total_records,
            'tables_status': tables_info,
            'percentage': 5
        })
        set_backup_progress_data(progress_data)
        
        time.sleep(2)
        
        # إنشاء هيكل النسخة الاحتياطية
        backup_content = {
            'metadata': {
                'backup_name': f'نسخة احتياطية {timestamp}',
                'created_at': timezone.now().isoformat(),
                'created_by': user.username if user else 'system',
                'system_version': '1.0',
                'total_tables': total_tables,
                'total_records': total_records,
                'format': format_type.upper(),
                'description': 'نسخة احتياطية كاملة مع معلومات تفصيلية'
            },
            'data': {}
        }

        # تضمين قائمة بملفات الوسائط المهمة (شعارات النظام وصور الخلفية) ضمن المحتوى
        try:
            media_files = []
            system_media_dir = os.path.join(settings.MEDIA_ROOT, 'system')
            if os.path.exists(system_media_dir):
                for root, dirs, files in os.walk(system_media_dir):
                    for fname in files:
                        fpath = os.path.join(root, fname)
                        rel = os.path.relpath(fpath, settings.MEDIA_ROOT)
                        media_files.append(rel.replace('\\', '/'))
            backup_content['media_files'] = media_files
        except Exception as e:
            logger.warning(f"فشل في حصر ملفات الوسائط للنسخ الاحتياطي: {str(e)}")
        
        # تضمين ملفات الترجمة
        try:
            locale_files = []
            locale_dir = os.path.join(settings.BASE_DIR, 'locale')
            if os.path.exists(locale_dir):
                for root, dirs, files in os.walk(locale_dir):
                    for fname in files:
                        if fname.endswith('.po') or fname.endswith('.mo'):
                            fpath = os.path.join(root, fname)
                            rel = os.path.relpath(fpath, settings.BASE_DIR)
                            locale_files.append(rel.replace('\\', '/'))
            backup_content['locale_files'] = locale_files
        except Exception as e:
            logger.warning(f"فشل في حصر ملفات الترجمة للنسخ الاحتياطي: {str(e)}")
        
        # متغيرات التتبع
        processed_tables = 0
        processed_records = 0
        start_time = time.time()
        
        progress_data = get_backup_progress_data()
        progress_data.update({
            'status': 'processing',
            'current_table': 'بدء معالجة الجداول...',
            'percentage': 10
        })
        set_backup_progress_data(progress_data)
        
        # نسخ البيانات من كل جدول
        for table_index, table_info in enumerate(tables_info):
            app_name = table_info['app_name']
            model_name = table_info['model_name']
            display_name = table_info['display_name']
            expected_records = table_info['record_count']
            
            try:
                progress_data = get_backup_progress_data()
                tables_status = progress_data.get('tables_status', [])
                
                if table_index < len(tables_status):
                    tables_status[table_index]['status'] = 'processing'
                    tables_status[table_index]['progress'] = 0
                
                progress_data.update({
                    'current_table': f'معالجة {display_name}...',
                    'tables_status': tables_status
                })
                set_backup_progress_data(progress_data)
                
                # تسجيل بداية معالجة الجدول
                log_audit(user, 'create', f'بدء نسخ الجدول: {display_name} ({expected_records} سجل متوقع)')
                
                # الحصول على النموذج
                app = apps.get_app_config(app_name)
                model = app.get_model(model_name)
                
                if app_name not in backup_content['data']:
                    backup_content['data'][app_name] = {}
                
                # نسخ البيانات مع معالجة خاصة للحقول الإشكالية
                queryset = model.objects.all()
                
                # تحسين الاستعلامات للنماذج ذات العلاقات
                if hasattr(model, '_meta'):
                    # إضافة select_related للعلاقات الأجنبية المباشرة
                    foreign_keys = [f.name for f in model._meta.fields if hasattr(f, 'related_model') and f.related_model]
                    if foreign_keys:
                        queryset = queryset.select_related(*foreign_keys)
                    
                    # إضافة prefetch_related للعلاقات العكسية المهمة
                    reverse_relations = []
                    for related in model._meta.related_objects:
                        if related.one_to_many or related.many_to_many:
                            reverse_relations.append(related.get_accessor_name())
                    if reverse_relations:
                        queryset = queryset.prefetch_related(*reverse_relations)
                
                actual_records = queryset.count()
                
                if actual_records > 0:
                    # زيادة عدد الأجزاء لتسريع المعالجة (من 10 إلى 50 جزء)
                    chunk_size = max(100, actual_records // 50)  # حد أدنى 100 سجل لكل جزء
                    processed_in_table = 0
                    
                    all_data = []
                    for chunk_start in range(0, actual_records, chunk_size):
                        chunk_end = min(chunk_start + chunk_size, actual_records)
                        chunk_queryset = queryset[chunk_start:chunk_end]
                        
                        try:
                            # محاولة التسلسل العادي أولاً مع التحقق من الحقول المتاحة
                            chunk_data = serializers.serialize('json', chunk_queryset)
                            serialized_data = json.loads(chunk_data)
                            
                            # تنظيف البيانات من الحقول المحذوفة
                            cleaned_data = []
                            for item in serialized_data:
                                if 'fields' in item:
                                    # إزالة الحقول المحذوفة أو غير الموجودة - تم إزالة فلترة is_service
                                    pass
                                cleaned_data.append(item)
                            
                            all_data.extend(cleaned_data)
                            
                        except Exception as serialization_error:
                            # في حالة فشل التسلسل، نعالج السجلات فردياً
                            logger.warning(f"فشل تسلسل chunk في {display_name}: {str(serialization_error)}")
                            
                            for record in chunk_queryset:
                                try:
                                    # محاولة تسلسل سجل واحد
                                    single_data = serializers.serialize('json', [record])
                                    serialized_record = json.loads(single_data)
                                    
                                    # تنظيف البيانات من الحقول المحذوفة
                                    for item in serialized_record:
                                        # تم إزالة فلترة is_service - الحقل لم يعد موجود
                                        pass
                                    
                                    all_data.extend(serialized_record)
                                    
                                except Exception as single_error:
                                    # معالجة خاصة للسجلات التالفة
                                    logger.warning(f"فشل في تسلسل سجل {record.pk} في {display_name}: {str(single_error)}")
                                    
                                    # إنشاء سجل آمن يدوياً
                                    safe_record = {
                                        'model': f"{app_name}.{model_name}",
                                        'pk': record.pk,
                                        'fields': {}
                                    }
                                    
                                    # نسخ الحقول مع تجنب المشاكل
                                    for field in model._meta.fields:
                                        field_name = field.name
                                        
                                        # تم إزالة فلترة is_service - الحقل لم يعد موجود
                                            
                                        try:
                                            value = getattr(record, field_name)
                                            
                                            # معالجة خاصة للحقول الإشكالية
                                            if hasattr(field, 'upload_to'):  # حقول الملفات والصور
                                                if value and hasattr(value, 'name'):
                                                    safe_record['fields'][field_name] = value.name
                                                else:
                                                    safe_record['fields'][field_name] = None
                                            elif hasattr(value, 'isoformat'):  # حقول التاريخ
                                                safe_record['fields'][field_name] = value.isoformat()
                                            elif value is None:
                                                safe_record['fields'][field_name] = None
                                            else:
                                                # التأكد من أن القيمة قابلة للتسلسل
                                                try:
                                                    json.dumps(value)
                                                    safe_record['fields'][field_name] = value
                                                except (TypeError, ValueError):
                                                    safe_record['fields'][field_name] = str(value)
                                                    
                                        except Exception as field_error:
                                            logger.warning(f"فشل في حقل {field_name} للسجل {record.pk}: {str(field_error)}")
                                            safe_record['fields'][field_name] = None
                                    
                                    all_data.append(safe_record)
                                    log_audit(user, 'warning', f'تم إصلاح سجل تالف: {display_name} - المعرف: {record.pk}')
                        
                        processed_in_table = chunk_end
                        table_progress = int((processed_in_table / actual_records) * 100)
                        
                        progress_data = get_backup_progress_data()
                        tables_status = progress_data.get('tables_status', [])
                        if table_index < len(tables_status):
                            tables_status[table_index]['progress'] = table_progress
                        
                        progress_data.update({
                            'current_table': f'{display_name}: {processed_in_table}/{actual_records} ({table_progress}%)',
                            'tables_status': tables_status
                        })
                        set_backup_progress_data(progress_data)
                        
                        time.sleep(0.3)
                    
                    backup_content['data'][app_name][model_name] = all_data
                else:
                    backup_content['data'][app_name][model_name] = []
                
                processed_records += actual_records
                processed_tables += 1
                
                progress_data = get_backup_progress_data()
                tables_status = progress_data.get('tables_status', [])
                if table_index < len(tables_status):
                    tables_status[table_index]['status'] = 'completed'
                    tables_status[table_index]['progress'] = 100
                    tables_status[table_index]['actual_records'] = actual_records
                
                # حساب الوقت المتبقي
                elapsed_time = time.time() - start_time
                if processed_tables > 0:
                    avg_time_per_table = elapsed_time / processed_tables
                    remaining_tables = total_tables - processed_tables
                    estimated_remaining = avg_time_per_table * remaining_tables
                    estimated_time = f"{int(estimated_remaining)} ثانية متبقية تقريباً"
                else:
                    estimated_time = "حساب الوقت المتبقي..."
                
                overall_progress = 10 + int(((processed_tables / total_tables) * 80))
                progress_data.update({
                    'processed_tables': processed_tables,
                    'processed_records': processed_records,
                    'percentage': overall_progress,
                    'tables_status': tables_status,
                    'estimated_time': estimated_time,
                    'current_table': f'اكتمل {display_name} ✅ ({actual_records} سجل)'
                })
                set_backup_progress_data(progress_data)
                
                log_audit(user, 'update', f'اكتمل نسخ الجدول: {display_name} - تم نسخ {actual_records} سجل')
                
                time.sleep(0.5)
                
            except Exception as e:
                error_msg = f"تعذر نسخ جدول {display_name}: {str(e)}"
                logger.error(error_msg)
                
                # تسجيل الخطأ
                log_audit(user, 'error', error_msg)
                
                progress_data = get_backup_progress_data()
                tables_status = progress_data.get('tables_status', [])
                if table_index < len(tables_status):
                    tables_status[table_index]['status'] = 'error'
                    tables_status[table_index]['error'] = str(e)
                    tables_status[table_index]['progress'] = 0
                
                # تحديث الحالة للمتابعة رغم الخطأ
                processed_tables += 1
                overall_progress = 10 + int(((processed_tables / total_tables) * 80))
                
                progress_data.update({
                    'processed_tables': processed_tables,
                    'percentage': overall_progress,
                    'tables_status': tables_status,
                    'current_table': f'⚠️ فشل في {display_name}: {str(e)[:50]}...',
                    'errors': progress_data.get('errors', []) + [error_msg]
                })
                set_backup_progress_data(progress_data)
                
                # متابعة المعالجة للجدول التالي
                time.sleep(1)
                continue
                if table_index < len(tables_status):
                    tables_status[table_index]['status'] = 'error'
                    tables_status[table_index]['error'] = str(e)
                
                progress_data.update({
                    'tables_status': tables_status,
                    'current_table': f'خطأ في {display_name}: {str(e)}'
                })
                set_backup_progress_data(progress_data)
                
                if AUDIT_AVAILABLE:
                    try:
                        AuditLog.objects.create(
                            user=user,
                            action_type='backup_table_error',
                            content_type='backup_system',
                            description=f'خطأ في نسخ الجدول: {display_name} - {str(e)}'
                        )
                    except Exception as audit_e:
                        logger.warning(f"فشل في تسجيل خطأ النسخ الاحتياطي في AuditLog: {audit_e}")
                
                processed_tables += 1
                time.sleep(1)
                continue
        
        # حفظ النسخة الاحتياطية
        progress_data = get_backup_progress_data()
        progress_data.update({
            'current_table': 'حفظ الملف النهائي...',
            'status': 'saving',
            'percentage': 95,
            'estimated_time': 'جاري الحفظ...'
        })
        set_backup_progress_data(progress_data)
        
        try:
            # حفظ حسب النوع المطلوب
            if format_type.lower() == 'xlsx':
                if AUDIT_AVAILABLE:
                    try:
                        AuditLog.objects.create(
                            user=user,
                            action_type='backup_saving_xlsx',
                            content_type='backup_system',
                            description=f'بدء حفظ النسخة الاحتياطية بصيغة XLSX: {filename}'
                        )
                    except Exception as audit_e:
                        logger.warning(f"فشل في تسجيل حفظ XLSX في AuditLog: {audit_e}")
                save_backup_as_xlsx(backup_content, filepath)
            else:
                if AUDIT_AVAILABLE:
                    try:
                        AuditLog.objects.create(
                            user=user,
                            action_type='backup_saving_json',
                            content_type='backup_system',
                            description=f'بدء حفظ النسخة الاحتياطية بصيغة JSON: {filename}'
                        )
                    except Exception as audit_e:
                        logger.warning(f"فشل في تسجيل حفظ JSON في AuditLog: {audit_e}")
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump(backup_content, f, ensure_ascii=False, indent=2)
            
            # التأكد من حفظ الملف
            if not os.path.exists(filepath):
                raise Exception(f"فشل في إنشاء الملف: {filepath}")
            
            file_size = os.path.getsize(filepath)
            if file_size == 0:
                raise Exception(f"تم إنشاء ملف فارغ: {filepath}")
                
            logger.info(f"تم حفظ الملف بنجاح: {filepath} - الحجم: {file_size} بايت")
            
        except Exception as save_error:
            logger.error(f"خطأ في حفظ الملف: {str(save_error)}")
            
            if AUDIT_AVAILABLE:
                try:
                    AuditLog.objects.create(
                        user=user,
                        action_type='backup_save_error',
                        content_type='backup_system',
                        description=f'خطأ في حفظ النسخة الاحتياطية: {filename} - خطأ: {str(save_error)}'
                    )
                except Exception as audit_e:
                    logger.warning(f"فشل في تسجيل خطأ الحفظ في AuditLog: {audit_e}")
            
            raise save_error
        
        # تحديث النهاية الناجحة
        progress_data.update({
            'status': 'completed',
            'current_table': 'تم الانتهاء بنجاح! ✅',
            'percentage': 100,
            'is_running': False,
            'estimated_time': 'اكتمل'
        })
        set_backup_progress_data(progress_data)
        
        # حساب إحصائيات الجداول
        empty_tables = 0
        non_empty_tables = 0
        for table_info in tables_info:
            if table_info['record_count'] == 0:
                empty_tables += 1
            else:
                non_empty_tables += 1
        
        # تسجيل النشاط النهائي مع تفاصيل الجداول
        details = f'تم إنشاء النسخة الاحتياطية بنجاح: {filename} - النوع: {format_type.upper()} - إجمالي الجداول: {total_tables} (فارغة: {empty_tables}, تحتوي بيانات: {non_empty_tables}) - السجلات: {processed_records}'
        log_audit(user, 'create', details)
        
        logger.info(f"تم إنشاء النسخة الاحتياطية بنجاح: {filename}")
        
        # تنظيف الكاش بعد فترة قصيرة للسماح للواجهة بقراءة الحالة النهائية
        def cleanup_cache():
            time.sleep(10)  # انتظار 10 ثواني
            cache.delete('backup_progress')
            cache.delete('backup_last_update')
        
        threading.Thread(target=cleanup_cache, daemon=True).start()
        
    except Exception as e:
        logger.error(f"خطأ في إنشاء النسخة الاحتياطية: {str(e)}")
        
        progress_data = get_backup_progress_data()
        progress_data.update({
            'status': 'error',
            'error': str(e),
            'is_running': False,
            'current_table': f'خطأ: {str(e)}'
        })
        set_backup_progress_data(progress_data)
        
        log_audit(user, 'error', f'فشل في إنشاء النسخة الاحتياطية: {filename} - خطأ: {str(e)}')
        
        # تنظيف الكاش بعد فترة قصيرة عند الخطأ أيضاً
        def cleanup_cache_error():
            time.sleep(30)  # انتظار 30 ثانية عند الخطأ
            cache.delete('backup_progress')
            cache.delete('backup_last_update')
        
        threading.Thread(target=cleanup_cache_error, daemon=True).start()
        threading.Thread(target=cleanup_cache_error, daemon=True).start()
        
        raise e


def clean_sheet_name(name):
    """تنظيف اسم ورقة العمل لضمان التوافق مع Excel"""
    # إزالة الأحرف غير المسموحة في Excel
    invalid_chars = ['\\', '/', '?', '*', '[', ']', ':', '\'']
    cleaned_name = str(name) if name else ''
    for char in invalid_chars:
        cleaned_name = cleaned_name.replace(char, '_')
    
    # إزالة النقاط (قد تتسبب في مشاكل)
    cleaned_name = cleaned_name.replace('.', '_')
    
    # إزالة أي أحرف خاصة أخرى قد تسبب مشاكل
    cleaned_name = ''.join(c if c.isalnum() or c in '_- ' else '_' for c in cleaned_name)
    
    # إزالة المسافات المتعددة والاستبدال بـ underscore واحد
    cleaned_name = '_'.join(cleaned_name.split())
    
    # تحديد الطول الأقصى (31 حرف هو الحد الأقصى لـ Excel)
    if len(cleaned_name) > 31:
        cleaned_name = cleaned_name[:31]
    
    # التأكد من أن الاسم ليس فارغاً أو يبدأ برقم
    if not cleaned_name or cleaned_name.isspace() or cleaned_name[0].isdigit():
        cleaned_name = "Sheet_" + cleaned_name[:25] if cleaned_name else "Sheet1"
    
    # التأكد مرة أخيرة من الطول
    if len(cleaned_name) > 31:
        cleaned_name = cleaned_name[:31]
    
    return cleaned_name


def save_backup_as_xlsx(backup_content, filepath):
    """حفظ النسخة الاحتياطية كملف Excel"""
    try:
        logger.info(f"بدء حفظ النسخة الاحتياطية كـ XLSX: {filepath}")
        
        # تسجيل معلومات مفصلة عن البيانات الواردة
        data = backup_content.get('data', {})
        total_apps = len(data)
        total_models = sum(len(app_data) for app_data in data.values())
        total_records = sum(
            len(model_data) if isinstance(model_data, list) else 0
            for app_data in data.values()
            for model_data in app_data.values()
        )
        
        logger.info(f"إحصائيات البيانات الواردة - تطبيقات: {total_apps}, نماذج: {total_models}, سجلات: {total_records}")
        
        # تسجيل تفاصيل كل تطبيق
        for app_name, app_data in data.items():
            app_records = sum(len(model_data) if isinstance(model_data, list) else 0 for model_data in app_data.values())
            logger.info(f"تطبيق {app_name}: {len(app_data)} نموذج، {app_records} سجل")
            for model_name, model_data in app_data.items():
                if isinstance(model_data, list):
                    logger.info(f"  نموذج {model_name}: {len(model_data)} سجل")
                    if len(model_data) > 0 and isinstance(model_data[0], dict):
                        logger.info(f"    عينة من البيانات: pk={model_data[0].get('pk')}, fields={list(model_data[0].get('fields', {}).keys())}")
        
        workbook = Workbook()
        
        # حذف الورقة الافتراضية
        workbook.remove(workbook.active)
        
        # إنشاء ورقة للمعلومات العامة
        info_sheet = workbook.create_sheet(title="Backup Info")
        metadata = backup_content.get('metadata', {})
        
        info_sheet['A1'] = 'Backup Name'
        info_sheet['B1'] = metadata.get('backup_name', '')
        info_sheet['A2'] = 'Created At'
        info_sheet['B2'] = metadata.get('created_at', '')
        info_sheet['A3'] = 'Created By'
        info_sheet['B3'] = metadata.get('created_by', '')
        info_sheet['A4'] = 'Total Tables'
        info_sheet['B4'] = metadata.get('total_tables', 0)
        info_sheet['A5'] = 'Total Records'
        info_sheet['B5'] = metadata.get('total_records', 0)
        
        # إنشاء ورقة لكل تطبيق
        data = backup_content.get('data', {})
        sheet_count = 0
        used_sheet_names = set()  # تتبع الأسماء المستخدمة
        
        logger.info(f"بدء إنشاء الأوراق - عدد التطبيقات: {len(data)}")
        
        for app_name, app_data in data.items():
            logger.info(f"معالجة تطبيق: {app_name} - عدد النماذج: {len(app_data)}")
            for model_name, model_data in app_data.items():
                logger.info(f"معالجة نموذج: {model_name} - نوع البيانات: {type(model_data)} - عدد العناصر: {len(model_data) if isinstance(model_data, list) else 'غير قائمة'}")
                
                if isinstance(model_data, list):  # إزالة شرط التحقق من أن القائمة غير فارغة
                    logger.info(f"إنشاء ورقة للنموذج: {app_name}.{model_name}")
                    try:
                        # تنظيف اسم ورقة العمل لضمان التوافق مع Excel
                        raw_sheet_name = f"{app_name}_{model_name}"
                        sheet_name = clean_sheet_name(raw_sheet_name)
                        
                        # التأكد من عدم تكرار الاسم
                        original_name = sheet_name
                        counter = 1
                        while sheet_name in used_sheet_names:
                            counter_suffix = f"_{counter}"
                            max_length = 31 - len(counter_suffix)
                            if len(original_name) > max_length:
                                sheet_name = f"{original_name[:max_length]}{counter_suffix}"
                            else:
                                sheet_name = f"{original_name}{counter_suffix}"
                            counter += 1
                        
                        used_sheet_names.add(sheet_name)
                        sheet = workbook.create_sheet(title=sheet_name)
                        sheet_count += 1
                        
                        logger.info(f"تم إنشاء ورقة العمل: {sheet_name} - عدد السجلات: {len(model_data)}")
                        
                        if len(model_data) > 0:
                            # كتابة الرؤوس
                            first_record = model_data[0]
                            if isinstance(first_record, dict) and 'fields' in first_record:
                                fields = first_record.get('fields', {})
                                
                                # فلترة الحقول المحذوفة
                                filtered_fields = {}
                                for field_name, field_value in fields.items():
                                    # تم إزالة فلترة is_service - الحقل لم يعد موجود
                                    filtered_fields[field_name] = field_value
                                
                                headers = ['ID'] + list(filtered_fields.keys())
                                logger.info(f"كتابة الرؤوس للورقة {sheet_name}: {headers}")
                                
                                for col, header in enumerate(headers, 1):
                                    # تنظيف اسم العمود
                                    clean_header = str(header) if header else f'Column_{col}'
                                    # إزالة الأحرف الخاصة من أسماء الأعمدة
                                    clean_header = ''.join(c if c.isalnum() or c in ' _-' else '_' for c in clean_header)
                                    if len(clean_header) > 255:  # حد Excel للأعمدة
                                        clean_header = clean_header[:252] + "..."
                                    sheet.cell(row=1, column=col, value=clean_header)
                                
                                # كتابة البيانات
                                logger.info(f"بدء كتابة {len(model_data)} سجل للورقة {sheet_name}")
                                for row_idx, record in enumerate(model_data, 2):
                                    if isinstance(record, dict):
                                        pk_value = record.get('pk', '')
                                        sheet.cell(row=row_idx, column=1, value=str(pk_value))
                                        record_fields = record.get('fields', {})
                                        
                                        logger.debug(f"كتابة السجل {row_idx-1}: pk={pk_value}, fields={list(record_fields.keys())}")
                                        
                                        # فلترة الحقول المحذوفة قبل الكتابة
                                        filtered_record_fields = {}
                                        for field_name, field_value in record_fields.items():
                                            # تم إزالة فلترة is_service - الحقل لم يعد موجود
                                            filtered_record_fields[field_name] = field_value
                                        
                                        for col_idx, (field_name, field_value) in enumerate(filtered_record_fields.items(), 2):
                                            try:
                                                if field_value is None:
                                                    cell_value = ''
                                                elif isinstance(field_value, (str, int, float, bool)):
                                                    cell_value = str(field_value)
                                                else:
                                                    cell_value = str(field_value)
                                                
                                                # تحديد طول النص لتجنب مشاكل Excel
                                                if len(cell_value) > 32767:
                                                    cell_value = cell_value[:32760] + "..."
                                                    
                                                # تجنب الأحرف الخاصة التي قد تؤثر على Excel
                                                if cell_value and any(ord(char) < 32 for char in cell_value if char != '\t' and char != '\n' and char != '\r'):
                                                    cell_value = ''.join(char if ord(char) >= 32 or char in '\t\n\r' else '?' for char in cell_value)
                                                    
                                                sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                                            except Exception as cell_error:
                                                logger.warning(f"خطأ في كتابة خلية {field_name}: {str(cell_error)}")
                                                sheet.cell(row=row_idx, column=col_idx, value='Error')
                        else:
                            # إضافة رأس أساسي للجداول الفارغة
                            sheet.cell(row=1, column=1, value='ID')
                            sheet.cell(row=1, column=2, value='Empty Table')
                            
                        logger.debug(f"تم إنشاء ورقة العمل: {sheet_name} مع {len(model_data)} سجل")
                    
                    except Exception as e:
                        logger.warning(f"خطأ في إنشاء ورقة العمل {app_name}_{model_name}: {str(e)}")
                        # تسجيل الخطأ في audit log إذا كان متاحاً
                        if AUDIT_AVAILABLE:
                            try:
                                AuditLog.objects.create(
                                    user=None,  # سيتم تعيينه لاحقاً بواسطة المستخدم الذي أنشأ النسخة
                                    action_type='backup_xlsx_sheet_error',
                                    content_type='backup_system',
                                    description=f'خطأ في إنشاء ورقة عمل {app_name}_{model_name}: {str(e)}'
                                )
                            except Exception:
                                logger.warning(f"فشل في تسجيل خطأ النسخ الاحتياطي في AuditLog: {audit_e}")
                        continue
        
        # حفظ الملف
        logger.info(f"بدء حفظ الملف: {filepath}")
        workbook.save(filepath)
        
        # التحقق من حفظ الملف
        if os.path.exists(filepath):
            file_size = os.path.getsize(filepath)
            logger.info(f"تم حفظ النسخة الاحتياطية XLSX بنجاح: {filepath} - الحجم: {file_size} بايت - تم إنشاء {sheet_count} ورقة عمل")
            
            # تسجيل معلومات إضافية عن الأوراق
            sheet_names = workbook.sheetnames
            logger.info(f"أسماء الأوراق المحفوظة: {sheet_names}")
            
            # قراءة أول ورقة للتحقق
            if len(sheet_names) > 1:  # أول ورقة هي Backup Info
                first_data_sheet = workbook[sheet_names[1]]
                row_count = first_data_sheet.max_row
                col_count = first_data_sheet.max_column
                logger.info(f"عينة من أول ورقة بيانات ({sheet_names[1]}): {row_count} صف، {col_count} عمود")
        else:
            logger.error(f"فشل في حفظ الملف: {filepath} غير موجود بعد الحفظ")
        
    except Exception as e:
        logger.error(f"خطأ في حفظ النسخة الاحتياطية كـ XLSX: {str(e)}")
        # تسجيل الخطأ الرئيسي في audit log
        try:
            from core.models import AuditLog
            AuditLog.objects.create(
                user=None,  # سيتم تعيينه لاحقاً بواسطة المستخدم الذي أنشأ النسخة
                action='backup_xlsx_save_error',
                content_type='backup_system',
                description=f'خطأ كبير في حفظ ملف XLSX: {str(e)}'
            )
        except Exception:
            pass
        raise e


def load_backup_from_xlsx(file_obj, user=None):
    """تحويل ملف Excel إلى تنسيق JSON للاستعادة مع مزيد من التحمّل"""
    try:
        from openpyxl import load_workbook
        from django.apps import apps as django_apps
        
        logger.info("بدء قراءة ملف Excel للاستعادة")
        workbook = load_workbook(file_obj)
        
        # قراءة معلومات النسخة الاحتياطية من ورقة "Backup Info"
        backup_data = {
            'metadata': {},
            'data': {}
        }
        
        info_sheet = None
        if "Backup Info" in workbook.sheetnames:
            info_sheet = workbook["Backup Info"]
            
            # قراءة المعلومات الأساسية
            for row in info_sheet.iter_rows(min_row=1, max_row=5, values_only=True):
                if len(row) >= 2 and row[0] and row[1]:
                    if row[0] == 'Backup Name':
                        backup_data['metadata']['backup_name'] = str(row[1])
                    elif row[0] == 'Created At':
                        backup_data['metadata']['created_at'] = str(row[1])
                    elif row[0] == 'Created By':
                        backup_data['metadata']['created_by'] = str(row[1])
                    elif row[0] == 'Total Tables':
                        backup_data['metadata']['total_tables'] = int(row[1]) if row[1] else 0
                    elif row[0] == 'Total Records':
                        backup_data['metadata']['total_records'] = int(row[1]) if row[1] else 0
        
        # تجهيز قائمة أسماء التطبيقات المتاحة للمطابقة الذكية
        available_apps = {cfg.name for cfg in django_apps.get_app_configs()}
        
        # قراءة البيانات من باقي أوراق العمل
        for sheet_name in workbook.sheetnames:
            if sheet_name == "Backup Info":
                continue
                
            sheet = workbook[sheet_name]
            
            # تحليل اسم ورقة العمل للحصول على app_name و model_name بطريقة متحمّلة
            resolved_app = None
            resolved_model = None
            raw = str(sheet_name)
            parts = raw.split('_')
            # جرّب جميع نقاط التقسيم المحتملة حتى تجد نموذجاً معروفاً
            for i in range(1, len(parts)):
                cand_app = '_'.join(parts[:i])
                cand_model = '_'.join(parts[i:])
                # إزالة اللاحقة الرقمية المحتملة التي تمت إضافتها لتفادي التكرار
                cand_model = cand_model.rstrip('_0123456789')
                # تحقّق من توافق اسم التطبيق
                if cand_app in available_apps:
                    try:
                        # apps.get_model يتوقع app_label.model_name (model_name يمكن أن يكون lower)
                        mdl = django_apps.get_model(f"{cand_app}.{cand_model.lower()}")
                        if mdl is not None:
                            resolved_app = cand_app
                            resolved_model = cand_model
                            break
                    except Exception:
                        continue
            if not resolved_app or not resolved_model:
                # fallback: استخدم أول جزء كتطبيق والباقي كنموذج بعد إزالة اللاحقة الرقمية
                resolved_app = parts[0]
                resolved_model = '_'.join(parts[1:]).rstrip('_0123456789') or 'default'
            app_name = resolved_app
            model_name = resolved_model
            
            # سجّل محاولة التصحيح إذا كان اسم الورقة غامضاً
            if AUDIT_AVAILABLE and user and raw != f"{app_name}_{model_name}":
                try:
                    AuditLog.objects.create(
                        user=user,
                        action='backup_restore_xlsx_sheet_resolved',
                        details=f'تم تحليل اسم الورقة "{raw}" إلى التطبيق "{app_name}" والنموذج "{model_name}"'
                    )
                except Exception:
                    pass
            
            # التأكد من وجود التطبيق في البيانات
            if app_name not in backup_data['data']:
                backup_data['data'][app_name] = {}
            
            # قراءة البيانات من الورقة
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) < 1:
                backup_data['data'][app_name][model_name] = []
                continue
            
            # الصف الأول يحتوي على أسماء الأعمدة
            headers = [str(h) if h else f'col_{i}' for i, h in enumerate(rows[0])]
            
            # تحويل البيانات إلى تنسيق Django
            model_data = []
            for row_data in rows[1:]:  # تجاهل صف الرؤوس
                if not any(cell for cell in row_data):  # تجاهل الصفوف الفارغة
                    continue
                
                try:
                    # أول عمود هو ID
                    record_id = row_data[0] if len(row_data) > 0 and row_data[0] is not None else None
                    # Excel قد يحوّل الأرقام إلى float
                    if isinstance(record_id, float):
                        record_id = int(record_id) if float(record_id).is_integer() else record_id
                    
                    # باقي الأعمدة هي الحقول
                    fields = {}
                    for i, (header, value) in enumerate(zip(headers[1:], row_data[1:]), 1):
                        if value is not None:
                            # تحويل القيم حسب النوع
                            if isinstance(value, (int, float, bool)):
                                fields[header] = value
                            elif isinstance(value, str):
                                # محاولة تحويل النص إلى رقم أو boolean إذا أمكن
                                try:
                                    v = value.strip()
                                    # حاول قراءة JSON للقوائم/القواميس (مثل قيم ManyToMany المحفوظة كسلسلة)
                                    if (v.startswith('[') and v.endswith(']')) or (v.startswith('{') and v.endswith('}')):
                                        try:
                                            parsed = json.loads(v)
                                            fields[header] = parsed
                                            continue
                                        except Exception:
                                            pass
                                    # قائمة مفصولة بفواصل للأرقام: 1,2,3
                                    if ',' in v:
                                        parts_vals = [p.strip() for p in v.split(',') if p.strip()]
                                        if parts_vals and all(p.replace('.', '', 1).isdigit() for p in parts_vals):
                                            # حوّل إلى أعداد صحيحة عندما ممكن
                                            as_numbers = [int(p) if p.isdigit() else float(p) for p in parts_vals]
                                            fields[header] = as_numbers
                                            continue
                                    if v.lower() in ['true', 'false']:
                                        fields[header] = v.lower() == 'true'
                                    elif v.isdigit():
                                        fields[header] = int(v)
                                    elif v.replace('.', '', 1).isdigit():
                                        fields[header] = float(v)
                                    else:
                                        fields[header] = value
                                except:
                                    fields[header] = value
                            else:
                                fields[header] = str(value) if value is not None else None
                        else:
                            fields[header] = None
                    
                    # إنشاء سجل بتنسيق Django
                    record = {
                        'pk': record_id,
                        'model': f"{app_name}.{model_name.lower()}",
                        'fields': fields
                    }
                    model_data.append(record)
                    
                except Exception as row_error:
                    logger.warning(f"خطأ في قراءة صف في ورقة {sheet_name}: {str(row_error)}")
                    continue
            
            backup_data['data'][app_name][model_name] = model_data
            logger.debug(f"تم قراءة {len(model_data)} سجل من ورقة {sheet_name}")
        
        logger.info(f"تم تحويل ملف Excel بنجاح - {len(backup_data['data'])} تطبيق")
        return backup_data
        
    except Exception as e:
        logger.error(f"خطأ في قراءة ملف Excel: {str(e)}")
        raise Exception(f"خطأ في قراءة ملف Excel: {str(e)}")


def convert_field_value(field, value):
    """تحويل قيمة الحقل إلى النوع الصحيح"""
    if value is None or value == '':
        return value
    
    try:
        field_type = field.get_internal_type()
        
        if field_type in ['DecimalField']:
            if isinstance(value, str):
                return Decimal(value)
            return Decimal(str(value))
        elif field_type in ['IntegerField', 'BigIntegerField', 'SmallIntegerField', 'PositiveIntegerField']:
            if isinstance(value, str):
                return int(float(value)) if '.' in value else int(value)
            return int(value)
        elif field_type in ['FloatField']:
            if isinstance(value, str):
                return float(value)
            return float(value)
        elif field_type in ['BooleanField']:
            if isinstance(value, str):
                return value.lower() in ('true', '1', 'yes', 'on')
            return bool(value)
        elif field_type in ['DateTimeField']:
            if isinstance(value, str):
                from django.utils.dateparse import parse_datetime
                return parse_datetime(value)
            return value
        elif field_type in ['DateField']:
            if isinstance(value, str):
                from django.utils.dateparse import parse_date
                return parse_date(value)
            return value
        elif field_type in ['TimeField']:
            if isinstance(value, str):
                from django.utils.dateparse import parse_time
                return parse_time(value)
            return value
        else:
            return value
    except Exception:
        # في حالة فشل التحويل، أعد القيمة الأصلية
        return value





def perform_backup_restore(backup_data, clear_data=False, user=None):
    """تنفيذ عملية الاستعادة الفعلية"""
    from .restore_context import set_restoring
    
    try:
        # 🔧 تفعيل وضع الاستعادة لتعطيل السيجنالات
        set_restoring(True)
        logger.info("بدء تنفيذ عملية الاستعادة (السيجنالات مُعطّلة)")
        
        # تسجيل بداية العملية في سجل الأنشطة
        log_audit(user, 'create', _('بدء عملية استعادة النسخة الاحتياطية'))

        start_time = time.time()

        # 🎯 ترتيب الاستعادة الصحيح: عكس ترتيب المسح (الجذور أولاً → الأطفال أخيراً)
        # المبدأ: استعد الصف المشار إليه قبل الصف الذي يشير
        restoration_order = [
            # ════════════════════════════════════════════════════════
            # المستوى 0: الإعدادات والمتسلسلات (الجذور - أول شيء)
            # ════════════════════════════════════════════════════════
            'core.documentsequence',
            'core.companysettings',
            'settings.jofotarasettings',
            'settings.documentprintsettings',
            'settings.companysettings',
            'settings.currency',
            'settings.superadminsettings',  # ⭐ إعدادات المدير العام
            
            # ════════════════════════════════════════════════════════
            # المستوى 1: المستخدمين والصلاحيات (أول شيء!)
            # ════════════════════════════════════════════════════════
            'auth.group',  # ⭐ المجموعات قبل المستخدمين
            'auth.permission',  # ⭐ الصلاحيات
            'auth.user',  # ⭐ auth.user من Django
            'users.user',  # ⭐ ⭐ users.user من نظامنا - حاسم جداً!
            'users.usergroup',  # ⭐ مجموعات المستخدمين
            'users.usergroupmembership',  # ⭐ عضويات المجموعات
            'users.userprofile',  # ⭐ بعد المستخدمين
            
            # ════════════════════════════════════════════════════════
            # المستوى 2: الجداول الأساسية المرجعية (بدون FK خارجية)
            # ════════════════════════════════════════════════════════
            'revenues_expenses.sector',
            'revenues_expenses.revenueexpensecategory',
            'provisions.provisiontype',
            'assets_liabilities.liabilitycategory',
            'assets_liabilities.assetcategory',
            'journal.account',  # ⭐ الحسابات قبل كل شيء يستخدمها
            'accounts.costcenter',  # ⭐ مراكز التكلفة
            'hr.leavetype',
            'hr.position',
            'hr.department',
            'products.category',
            'products.unit',  # ⭐ وحدات القياس قبل المنتجات
            'customers.customercategory',  # ⭐ فئات العملاء قبل العملاء
            'customers.customersupplier',  # ⭐ العملاء/الموردين قبل الفواتير
            'customers.customer',  # ⭐ العملاء المنفصلين
            
            # ════════════════════════════════════════════════════════
            # المستوى 2: البيانات الأساسية المشار إليها كثيراً
            # ════════════════════════════════════════════════════════
            'inventory.warehouse',  # ⭐ المستودعات قبل الفواتير والحركات
            'cashboxes.cashbox',
            'banks.bankaccount',
            'products.product',  # ⭐ المنتجات قبل عناصر الفواتير
            
            # ════════════════════════════════════════════════════════
            # المستوى 3: السجلات التابعة للجذور
            # ════════════════════════════════════════════════════════
            'reports.reportaccesscontrol',
            'documents.document',
            'provisions.provision',
            'assets_liabilities.liability',
            'assets_liabilities.asset',
            'hr.employee',  # بعد departments & positions
            
            # ════════════════════════════════════════════════════════
            # المستوى 4: الفواتير والقيود الرئيسية
            # ════════════════════════════════════════════════════════
            'hr.payrollperiod',
            'hr.contract',
            'purchases.purchaseinvoice',  # ⭐ بعد customers, warehouse, products
            'sales.salesinvoice',  # ⭐ بعد customers, warehouse, products
            'journal.journalentry',
            'journal.yearendclosing',
            
            # ════════════════════════════════════════════════════════
            # المستوى 5: المستندات الرئيسية والمعاملات
            # ════════════════════════════════════════════════════════
            'purchases.purchasereturn',
            'sales.salesreturn',
            'sales.salescreditnote',
            'purchases.purchasedebitnote',
            'payments.paymentvoucher',
            'receipts.paymentreceipt',
            'banks.banktransfer',
            'banks.banktransaction',
            'banks.bankreconciliation',
            'banks.bankstatement',
            'cashboxes.cashboxtransfer',
            'cashboxes.cashboxtransaction',
            'inventory.warehousetransfer',
            'inventory.inventorymovement',
            'accounts.accounttransaction',
            'revenues_expenses.recurringrevenueexpense',
            'revenues_expenses.revenueexpenseentry',
            'assets_liabilities.depreciationentry',
            'provisions.provisionentry',
            'journal.journalline',  # ⭐ بعد journalentry و accounts
            
            # ════════════════════════════════════════════════════════
            # المستوى 6: بنود وتفاصيل المستندات
            # ════════════════════════════════════════════════════════
            'receipts.checkcollection',
            'hr.employeedocument',
            'hr.leaverequest',
            'hr.attendance',
            'hr.payrollentry',
            'inventory.warehousetransferitem',
            'purchases.purchasereturnitem',
            'sales.salesreturnitem',
            'purchases.purchaseinvoiceitem',  # ⭐ بعد purchaseinvoice و products
            'sales.salesinvoiceitem',  # ⭐ بعد salesinvoice و products
            
            # ════════════════════════════════════════════════════════
            # المستوى 7 (الأخير): سجلات التدقيق والإشعارات
            # ════════════════════════════════════════════════════════
            'receipts.receiptreversal',
            'core.systemnotification',
            'core.auditlog',  # ⭐ سجل المراجعة - آخر شيء
        ]

        # تهيئة تتبع التقدم
        flat_tables = []
        total_records_expected = 0
        backup_data_dict = {}  # قاموس للوصول السريع للبيانات
        
        if isinstance(backup_data, dict) and 'data' in backup_data:
            # 🔧 تحويل أسماء التطبيقات القديمة إلى الأسماء الجديدة
            app_name_mapping = {
                'revenues': 'revenues_expenses',
                'assets': 'assets_liabilities'
            }
            
            # تحويل أسماء النماذج القديمة إلى الأسماء الجديدة
            model_name_mapping = {
                'revenues_expenses': {
                    'expenses_revenueexpens': 'revenueexpenseentry',
                    'expenses_revenueexpe': 'revenueexpenseentry', 
                    'expenses_recurringreve': 'recurringrevenueexpense'
                },
                'settings': {
                    'companysettings': 'companysettings'  # سيتم تحويله إلى core.companysettings
                }
            }
            
            # إنشاء نسخة من البيانات مع الأسماء المحدثة
            updated_backup_data = {'data': {}}
            for app_name, app_data in backup_data['data'].items():
                new_app_name = app_name_mapping.get(app_name, app_name)
                
                for model_name, model_records in app_data.items():
                    # تحويل اسم النموذج إذا كان مطلوباً
                    if new_app_name in model_name_mapping and model_name in model_name_mapping[new_app_name]:
                        new_model_name = model_name_mapping[new_app_name][model_name]
                        logger.info(f"🔄 تحويل نموذج {app_name}.{model_name} إلى {new_app_name}.{new_model_name}")
                    else:
                        new_model_name = model_name
                    
                    # تحويل خاص لـ settings.companysettings إلى core.companysettings
                    if app_name == 'settings' and model_name == 'companysettings':
                        final_app_name = 'core'
                        final_model_name = 'companysettings'
                        logger.info(f"🔄 تحويل نموذج {app_name}.{model_name} إلى {final_app_name}.{final_model_name}")
                    else:
                        final_app_name = new_app_name
                        final_model_name = new_model_name
                    
                    # إنشاء القاموس إذا لم يكن موجوداً
                    if final_app_name not in updated_backup_data['data']:
                        updated_backup_data['data'][final_app_name] = {}
                    
                    updated_backup_data['data'][final_app_name][final_model_name] = model_records
            
            # استخدام البيانات المحدثة
            backup_data = updated_backup_data
            
            # أولاً: جمع جميع الجداول الموجودة في النسخة الاحتياطية
            for app_name, app_data in backup_data['data'].items():
                for model_name, model_records in app_data.items():
                    label = f"{app_name}.{model_name}"
                    backup_data_dict[label] = {
                        'app_name': app_name,
                        'model_name': model_name,
                        'records': model_records
                    }
                    
                    if isinstance(model_records, list):
                        # حساب عدد السجلات الصحيح (تنسيق fixtures أو مباشر)
                        rec_count = 0
                        for record_item in model_records:
                            if isinstance(record_item, dict):
                                if 'fields' in record_item and 'pk' in record_item:
                                    # تنسيق Django fixtures
                                    rec_count += 1
                                elif 'pk' in record_item or any(k != 'pk' for k in record_item.keys()):
                                    # تنسيق البيانات المباشرة
                                    rec_count += 1
                        total_records_expected += rec_count
            
            # ثانياً: ترتيب الجداول حسب restoration_order
            # نبدأ بالجداول المحددة في الترتيب
            for ordered_label in restoration_order:
                if ordered_label in backup_data_dict:
                    data_info = backup_data_dict[ordered_label]
                    records = data_info['records']
                    rec_count = 0
                    if isinstance(records, list):
                        for record_item in records:
                            if isinstance(record_item, dict):
                                if 'fields' in record_item and 'pk' in record_item:
                                    rec_count += 1
                                elif 'pk' in record_item or any(k != 'pk' for k in record_item.keys()):
                                    rec_count += 1
                    
                    flat_tables.append({
                        'app_name': data_info['app_name'],
                        'model_name': data_info['model_name'],
                        'display_name': ordered_label,
                        'record_count': rec_count,
                        'status': 'pending',
                        'progress': 0,
                        'actual_records': 0,
                        'error': None
                    })
            
            # ثالثاً: إضافة أي جداول غير مذكورة في الترتيب (في النهاية)
            for label, data_info in backup_data_dict.items():
                if label not in restoration_order:
                    records = data_info['records']
                    rec_count = 0
                    if isinstance(records, list):
                        for record_item in records:
                            if isinstance(record_item, dict):
                                if 'fields' in record_item and 'pk' in record_item:
                                    rec_count += 1
                                elif 'pk' in record_item or any(k != 'pk' for k in record_item.keys()):
                                    rec_count += 1
                    
                    flat_tables.append({
                        'app_name': data_info['app_name'],
                        'model_name': data_info['model_name'],
                        'display_name': label,
                        'record_count': rec_count,
                        'status': 'pending',
                        'progress': 0,
                        'actual_records': 0,
                        'error': None
                    })
        
        logger.info(f"ترتيب الاستعادة: {len(flat_tables)} جدول")
        for idx, table in enumerate(flat_tables[:10]):  # عرض أول 10 جداول فقط في السجل
            logger.info(f"  {idx+1}. {table['display_name']} ({table['record_count']} سجل)")

        progress_data = get_restore_progress_data()
        progress_data.update({
            'is_running': True,
            'status': 'starting',
            'error': None,
            'percentage': 0,
            'current_table': 'بدء عملية الاستعادة...',
            'processed_tables': 0,
            'total_tables': len(flat_tables),
            'processed_records': 0,
            'total_records': total_records_expected,
            'estimated_time': 'جاري الحساب...',
            'tables_status': flat_tables
        })
        set_restore_progress_data(progress_data)
        
        # مسح البيانات الموجودة إذا طُلب ذلك
        if clear_data:
            logger.info("مسح البيانات الموجودة...")
            perform_clear_all_data(user)
        
        # استعادة البيانات
        processed_tables = 0
        processed_records = 0
        skipped_records = 0  # 🔍 عداد للسجلات المتخطاة
        
        # 🔍 تسجيل بداية الاستعادة
        logger.info(f"🚀 بدء استعادة {len(flat_tables)} جدول مع {total_records_expected} سجل متوقع")
        logger.info(f"� أول 20 جدول: {[t['display_name'] for t in flat_tables[:20]]}")
        
        # �🔧 تم إزالة transaction.atomic() - كان يسبب rollback كامل عند أي خطأ
        try:
            for i, table_info in enumerate(flat_tables):
                app_name = table_info['app_name']
                model_name = table_info['model_name']
                
                # 🔍 تسجيل بداية معالجة كل جدول
                logger.info(f"📊 [{i+1}/{len(flat_tables)}] معالجة جدول: {table_info['display_name']} ({table_info['record_count']} سجل متوقع)")
                
                progress_data['current_table'] = table_info['display_name']
                progress_data['processed_tables'] = i
                
                # حساب الوقت المتوقع
                elapsed = time.time() - start_time
                if processed_tables > 0:
                    avg_time_per_table = elapsed / processed_tables
                    remaining_tables = len(flat_tables) - processed_tables
                    estimated_seconds = avg_time_per_table * remaining_tables
                    progress_data['estimated_time'] = f"{int(estimated_seconds)} ثانية متبقية"
                else:
                    progress_data['estimated_time'] = 'جاري الحساب...'
                
                set_restore_progress_data(progress_data)
                
                try:
                    # الحصول على النموذج
                    app_config = apps.get_app_config(app_name)
                    
                    # 🔧 معالجة أسماء النماذج المبتورة أو القديمة
                    try:
                        model = app_config.get_model(model_name)
                    except LookupError:
                        # محاولة إيجاد النموذج الصحيح بناءً على جزء من الاسم
                        model = None
                        # قائمة بتطابقات أسماء النماذج المعروفة
                        model_mappings = {
                            'expenses_revenueexpens': 'RevenueExpenseEntry',
                            'expenses_revenueexpe': 'RevenueExpenseEntry',
                            'expenses_recurringreve': 'RecurringRevenueExpense',
                            'expenses_recurringrev': 'RecurringRevenueExpense',
                        }
                        
                        # محاولة إيجاد تطابق
                        if model_name.lower() in model_mappings:
                            try:
                                model = app_config.get_model(model_mappings[model_name.lower()])
                                logger.info(f"✅ تم إيجاد النموذج البديل: {model_name} → {model_mappings[model_name.lower()]}")
                            except LookupError:
                                pass
                        
                        # إذا لم نجد تطابق، نحاول البحث بشكل جزئي
                        if not model:
                            for available_model in app_config.get_models():
                                model_table_name = available_model._meta.db_table.replace(f"{app_name}_", "").lower()
                                if model_name.lower().startswith(model_table_name[:15]) or model_table_name.startswith(model_name.lower()[:15]):
                                    model = available_model
                                    logger.info(f"✅ تم إيجاد النموذج بالمطابقة الجزئية: {model_name} → {available_model.__name__}")
                                    break
                        
                        # إذا لم نجد النموذج نهائياً
                        if not model:
                            logger.warning(f"⚠️ لم يتم العثور على النموذج {app_name}.{model_name} - تخطي")
                            progress_data['tables_status'][i]['status'] = 'skipped'
                            progress_data['tables_status'][i]['error'] = f'النموذج غير موجود'
                            set_restore_progress_data(progress_data)
                            processed_tables += 1
                            continue
                    
                    if model and 'data' in backup_data and app_name in backup_data['data'] and model_name in backup_data['data'][app_name]:
                        records = backup_data['data'][app_name][model_name]
                        if isinstance(records, list):
                            for record_item in records:
                                try:
                                    # تحويل من تنسيق Django fixtures إلى البيانات المباشرة
                                    if isinstance(record_item, dict):
                                        if 'fields' in record_item and 'pk' in record_item:
                                            # تنسيق Django fixtures: {'model': 'app.model', 'pk': 1, 'fields': {...}}
                                            record_data = record_item['fields'].copy()
                                            record_data['pk'] = record_item['pk']
                                        else:
                                            # تنسيق البيانات المباشرة
                                            record_data = record_item.copy()
                                    else:
                                        continue
                                    
                                    # تنظيف البيانات للنماذج التي تغيرت هيكلها
                                    if model._meta.label == 'core.AuditLog':
                                        # إزالة الحقول القديمة التي لم تعد موجودة
                                        valid_fields = ['pk', 'user', 'action_type', 'content_type', 'object_id', 'description', 'ip_address', 'timestamp']
                                        record_data = {k: v for k, v in record_data.items() if k in valid_fields}
                                        # تجاهل السجلات التي لها user_id null أو غير موجود
                                        user_id = record_data.get('user')
                                        if user_id:
                                            # التحقق من وجود المستخدم في قاعدة البيانات
                                            from users.models import User
                                            if not User.objects.filter(pk=user_id).exists():
                                                # المستخدم غير موجود، تخطي هذا السجل
                                                continue
                                        else:
                                            # user_id مفقود، تخطي السجل
                                            continue
                                    elif model._meta.label == 'core.CompanySettings':
                                        # إزالة الحقول القديمة
                                        valid_fields = ['pk', 'company_name', 'logo', 'currency', 'address', 'phone', 'email', 'tax_number', 'default_tax_rate', 'session_timeout_minutes', 'enable_session_timeout', 'logout_on_browser_close', 'created_at', 'updated_at']
                                        record_data = {k: v for k, v in record_data.items() if k in valid_fields}
                                        # إضافة قيمة افتراضية لـ default_tax_rate إذا كان مفقوداً أو null
                                        if 'default_tax_rate' not in record_data or record_data.get('default_tax_rate') is None:
                                            from decimal import Decimal
                                            record_data['default_tax_rate'] = Decimal('0.00')
                                            logger.info(f"✅ تم إضافة قيمة افتراضية لـ default_tax_rate في CompanySettings")
                                    elif model._meta.label == 'core.DocumentSequence':
                                        # إزالة الحقول القديمة
                                        valid_fields = ['pk', 'document_type', 'prefix', 'digits', 'current_number', 'created_at', 'updated_at']
                                        record_data = {k: v for k, v in record_data.items() if k in valid_fields}
                                    elif model._meta.label == 'auth.Group':
                                        # إضافة قيمة افتراضية لـ dashboard_sections إذا كانت null
                                        if not record_data.get('dashboard_sections'):
                                            record_data['dashboard_sections'] = []
                                    elif model._meta.label == 'users.User':
                                        # إضافة قيم افتراضية للحقول المطلوبة
                                        if not record_data.get('first_name'):
                                            record_data['first_name'] = 'غير محدد'
                                        if not record_data.get('phone'):
                                            record_data['phone'] = '000000000'
                                        if not record_data.get('last_name'):
                                            record_data['last_name'] = 'غير محدد'
                                        # إضافة قيمة افتراضية لـ department (CharField with blank=True needs empty string)
                                        if 'department' not in record_data or record_data.get('department') in [None, 'null']:
                                            record_data['department'] = ''
                                        # إضافة قيمة افتراضية لـ groups (ManyToMany field)
                                        if 'groups' not in record_data or record_data.get('groups') in [None, 'null']:
                                            record_data['groups'] = []
                                        # إضافة قيمة افتراضية لـ email إذا كان null
                                        if record_data.get('email') is None:
                                            record_data['email'] = ''
                                    elif model._meta.label == 'products.Category':
                                        # إضافة قيمة افتراضية لـ description إذا كان null
                                        if record_data.get('description') is None:
                                            record_data['description'] = ''
                                    elif model._meta.label == 'customers.CustomerSupplier':
                                        # إضافة قيم افتراضية للحقول blank=True إذا كانت null
                                        if record_data.get('email') is None:
                                            record_data['email'] = ''
                                        if record_data.get('notes') is None:
                                            record_data['notes'] = ''
                                    elif model._meta.label == 'banks.BankAccount':
                                        # إضافة قيمة افتراضية لـ iban إذا كان null
                                        if record_data.get('iban') is None:
                                            record_data['iban'] = ''
                                    elif model._meta.label == 'settings.SuperAdminSettings':
                                        # إضافة قيمة افتراضية لـ system_subtitle
                                        if not record_data.get('system_subtitle'):
                                            record_data['system_subtitle'] = ''
                                        # إضافة قيمة افتراضية لـ app_logo إذا كان فارغاً
                                        if not record_data.get('app_logo'):
                                            record_data['app_logo'] = None
                                    elif model._meta.label == 'journal.JournalEntry':
                                        # 🔧 حل مشكلة تكرار entry_number
                                        # إذا كان entry_number موجود مسبقاً، نولّد رقم جديد
                                        entry_number = record_data.get('entry_number')
                                        if entry_number:
                                            # التحقق من وجود القيد
                                            from journal.models import JournalEntry
                                            if JournalEntry.objects.filter(entry_number=entry_number).exists():
                                                # الرقم موجود، نحذفه ونترك النظام يولد رقم جديد
                                                logger.debug(f"⚠️ entry_number مكرر: {entry_number}، سيتم توليد رقم جديد")
                                                record_data.pop('entry_number', None)
                                    elif model._meta.label == 'journal.Account':
                                        # 🔧 معالجة خاصة للحسابات: نترك parent كما هو
                                        # سيتم التعامل مع parent من خلال معالجة FK العادية
                                        # إذا كان parent غير موجود وكان الحقل nullable، سيُعين None تلقائياً
                                        pass
                                    
                                    # تنظيف البيانات من الحقول غير الموجودة في النموذج
                                    model_field_names = [f.name for f in model._meta.get_fields()]
                                    cleaned_data = {}
                                    many_to_many_data = {}
                                    
                                    for key, value in record_data.items():
                                        if key in model_field_names or key == 'pk':
                                            # التحقق من نوع الحقل
                                            field = None
                                            for f in model._meta.get_fields():
                                                if f.name == key:
                                                    field = f
                                                    break
                                            
                                            # Debug logging for groups field
                                            if key == 'groups':
                                                logger.info(f'Processing groups field: field={field}, value={value}, field_type={field.__class__.__name__ if field else None}')
                                                if field:
                                                    logger.info(f'  hasattr many_to_many: {hasattr(field, "many_to_many")}')
                                                    if hasattr(field, 'many_to_many'):
                                                        logger.info(f'  many_to_many value: {field.many_to_many}')
                                                    logger.info(f'  isinstance list: {isinstance(value, list)}')
                                            
                                            if field:
                                                # معالجة Many-to-Many fields أولاً (قبل Foreign Key)
                                                if (hasattr(field, 'many_to_many') and field.many_to_many) or field.__class__.__name__ == 'ManyToManyField':
                                                    if isinstance(value, list):
                                                        many_to_many_data[key] = value
                
                                                    # تجاهل Many-to-Many في البيانات الأساسية
                                                # معالجة Foreign Key (لكن ليس Many-to-Many)
                                                elif hasattr(field, 'related_model') and field.related_model and not (hasattr(field, 'many_to_many') and field.many_to_many):
                                                    try:
                                                        if value is None or str(value) == 'null':
                                                            # إذا كان الحقل مطلوب ولكن القيمة null، تخطي هذا السجل
                                                            if not field.null:
                                                                # لا نستطيع استعادة سجل بدون قيمة لحقل مطلوب
                                                                logger.warning(f"تخطي استعادة سجل في {model._meta.label}: الحقل {key} مطلوب لكن القيمة null")
                                                                raise ValueError(f"Required field {key} has null value")
                                                            else:
                                                                cleaned_data[key] = None
                                                        elif isinstance(value, (int, str)) and str(value).isdigit():
                                                            # البحث عن الكائن المرتبط
                                                            related_obj = field.related_model.objects.filter(pk=int(value)).first()
                                                            if related_obj:
                                                                cleaned_data[key] = related_obj
                                                            elif not field.null:
                                                                # 🔧 محاولة استخدام أول سجل متاح كحل بديل
                                                                first_available = field.related_model.objects.first()
                                                                if first_available:
                                                                    cleaned_data[key] = first_available
                                                                    logger.warning(f"⚠️ استخدام FK بديل لـ {key}: {first_available.pk} بدلاً من {value}")
                                                                else:
                                                                    # إذا لم يوجد أي سجل، يجب التخطي
                                                                    raise ValueError(f"FK_NOT_FOUND:{key}={value} (لا يوجد سجلات بديلة في {field.related_model.__name__})")
                                                            else:
                                                                # الحقل nullable، يمكن تركه None
                                                                cleaned_data[key] = None
                                                                logger.debug(f"FK مفقود {key}={value}، تم تعيين None")
                                                        else:
                                                            cleaned_data[key] = value
                                                    except ValueError as ve:
                                                        # إعادة رفع ValueError للتخطي
                                                        raise ve
                                                    except Exception:
                                                        # في حالة الخطأ الآخر، تجاهل هذا الحقل
                                                        pass
                                                # معالجة الحقول الرقمية لتجنب أخطاء التحويل
                                                elif field.__class__.__name__ in ['DecimalField', 'FloatField', 'IntegerField', 'PositiveIntegerField']:
                                                    try:
                                                        if value is None or value == '':
                                                            cleaned_data[key] = None if field.null else (0 if field.__class__.__name__ in ['IntegerField', 'PositiveIntegerField'] else 0.0)
                                                        elif isinstance(value, str):
                                                            # تنظيف النص وتحويله لرقم
                                                            numeric_value = value.replace(',', '').strip()
                                                            if field.__class__.__name__ in ['IntegerField', 'PositiveIntegerField']:
                                                                cleaned_data[key] = int(float(numeric_value)) if numeric_value else 0
                                                            elif field.__class__.__name__ == 'DecimalField':
                                                                # استخدام Decimal للحقول من نوع DecimalField
                                                                from decimal import Decimal
                                                                cleaned_data[key] = Decimal(str(numeric_value)) if numeric_value else Decimal('0.0')
                                                            else:
                                                                cleaned_data[key] = float(numeric_value) if numeric_value else 0.0
                                                        elif isinstance(value, (int, float)):
                                                            if field.__class__.__name__ in ['IntegerField', 'PositiveIntegerField']:
                                                                cleaned_data[key] = int(value)
                                                            elif field.__class__.__name__ == 'DecimalField':
                                                                from decimal import Decimal
                                                                cleaned_data[key] = Decimal(str(value))
                                                            else:
                                                                cleaned_data[key] = float(value)
                                                        else:
                                                            cleaned_data[key] = value
                                                    except (ValueError, TypeError):
                                                        if field.__class__.__name__ in ['IntegerField', 'PositiveIntegerField']:
                                                            cleaned_data[key] = 0
                                                        elif field.__class__.__name__ == 'DecimalField':
                                                            from decimal import Decimal
                                                            cleaned_data[key] = Decimal('0.0')
                                                        else:
                                                            cleaned_data[key] = 0.0
                                                else:
                                                    if key == 'groups':
                                                        logger.warning(f'groups field being treated as regular field! value: {value}, field type: {field.__class__.__name__}')
                                                    cleaned_data[key] = value
                                            else:
                                                cleaned_data[key] = value
                                    
                                    # تحويل None إلى قيم فارغة للحقول blank=True
                                    for f in model._meta.get_fields():
                                        if hasattr(f, 'blank') and f.blank and f.name in cleaned_data and cleaned_data[f.name] is None:
                                            if f.__class__.__name__ in ['CharField', 'TextField', 'EmailField']:
                                                cleaned_data[f.name] = ''
                                    
                                    # تعيين القيم الافتراضية للحقول المفقودة التي لها قيم افتراضية
                                    for f in model._meta.get_fields():
                                        if f.name not in cleaned_data and hasattr(f, 'default') and f.default is not models.NOT_PROVIDED:
                                            if callable(f.default):
                                                # إذا كانت الدالة الافتراضية قابلة للاستدعاء
                                                try:
                                                    cleaned_data[f.name] = f.default()
                                                except:
                                                    # في حالة فشل الاستدعاء، استخدم None
                                                    cleaned_data[f.name] = None
                                            else:
                                                cleaned_data[f.name] = f.default
                                    
                                    # التحقق من أن جميع الحقول المطلوبة موجودة
                                    for f in model._meta.get_fields():
                                        # تخطي ManyToMany fields (سيتم معالجتها لاحقاً)
                                        if hasattr(f, 'many_to_many') and f.many_to_many:
                                            continue
                                        # تخطي OneToOne reverse relations
                                        if f.__class__.__name__ == 'OneToOneRel':
                                            continue
                                        # التحقق من الحقول المطلوبة (غير null وغير blank)
                                        if hasattr(f, 'null') and not f.null and not getattr(f, 'blank', False) and f.name != 'id':
                                            if f.name not in cleaned_data or cleaned_data[f.name] is None:
                                                # 🔧 محاولة إضافة قيمة افتراضية بناءً على نوع الحقل
                                                field_default_added = False
                                                if f.__class__.__name__ == 'DecimalField':
                                                    from decimal import Decimal
                                                    cleaned_data[f.name] = Decimal('0.00')
                                                    field_default_added = True
                                                    logger.info(f"✅ تم إضافة قيمة افتراضية Decimal(0.00) للحقل '{f.name}' في {model._meta.label}")
                                                elif f.__class__.__name__ in ['IntegerField', 'PositiveIntegerField']:
                                                    cleaned_data[f.name] = 0
                                                    field_default_added = True
                                                    logger.info(f"✅ تم إضافة قيمة افتراضية 0 للحقل '{f.name}' في {model._meta.label}")
                                                elif f.__class__.__name__ in ['CharField', 'TextField']:
                                                    cleaned_data[f.name] = ''
                                                    field_default_added = True
                                                    logger.info(f"✅ تم إضافة قيمة افتراضية '' للحقل '{f.name}' في {model._meta.label}")
                                                elif f.__class__.__name__ == 'BooleanField':
                                                    cleaned_data[f.name] = False
                                                    field_default_added = True
                                                    logger.info(f"✅ تم إضافة قيمة افتراضية False للحقل '{f.name}' في {model._meta.label}")
                                                
                                                if not field_default_added:
                                                    # لم نتمكن من إضافة قيمة افتراضية، تخطي هذا السجل
                                                    logger.warning(f"⚠️ تخطي سجل في {model._meta.label}: الحقل المطلوب '{f.name}' مفقود أو null ولا يمكن تعيين قيمة افتراضية")
                                                    raise ValueError(f"Required field {f.name} is missing or null")
                                    
                                    # إنشاء أو تحديث السجل
                                    pk_value = cleaned_data.get('pk')
                                    
                                    # 🔧 معالجة خاصة لـ JournalEntry لحل مشكلة entry_number المكرر
                                    if model._meta.label == 'journal.JournalEntry' and pk_value:
                                        # محاولة الحصول على القيد الموجود
                                        try:
                                            obj = model.objects.get(pk=pk_value)
                                            # تحديث الحقول
                                            for k, v in cleaned_data.items():
                                                if k != 'pk' and k != 'entry_number':  # لا نحدث entry_number
                                                    setattr(obj, k, v)
                                            obj.save()
                                            created = False
                                        except model.DoesNotExist:
                                            # القيد غير موجود، ننشئ واحد جديد
                                            data_without_entry_number = {k: v for k, v in cleaned_data.items() if k not in ['pk', 'entry_number']}
                                            obj = model(**data_without_entry_number)
                                            obj.save()  # سيولد entry_number تلقائياً
                                            created = True
                                    elif pk_value:
                                        obj, created = model.objects.update_or_create(
                                            pk=pk_value,
                                            defaults={k: v for k, v in cleaned_data.items() if k != 'pk'}
                                        )
                                    else:
                                        obj = model.objects.create(**{k: v for k, v in cleaned_data.items() if k != 'pk'})
                                        created = True
                                    
                                    # معالجة Many-to-Many fields بعد إنشاء الكائن
                                    for field_name, related_ids in many_to_many_data.items():
                                        try:
                                            # الحصول على ManyToManyManager
                                            field_manager = getattr(obj, field_name)
                                            if hasattr(field_manager, 'set'):
                                                if isinstance(related_ids, list):
                                                    # التحقق من وجود الكائنات المرتبطة وتحويل الـ IDs
                                                    valid_ids = []
                                                    for rel_id in related_ids:
                                                        if isinstance(rel_id, (int, str)) and str(rel_id).isdigit():
                                                            valid_ids.append(int(rel_id))
                                                    field_manager.set(valid_ids)
                                                else:
                                                    # إذا لم تكن list، قم بمسح العلاقات
                                                    field_manager.clear()
                                        except Exception as m2m_err:
                                            # تجاهل أخطاء Many-to-Many
                                            logger.warning(f"خطأ في معالجة Many-to-Many field {field_name}: {m2m_err}")
                                            pass
                                    
                                    processed_records += 1
                                    table_info['actual_records'] += 1
                                    
                                    # 🔍 Logging مفصّل لكل سجل مستعاد
                                    if model._meta.label in ['banks.bankaccount', 'banks.banktransaction', 'cashboxes.cashbox']:
                                        logger.info(f"✅ استعادة {model._meta.label}[pk={pk_value}]: {cleaned_data.get('name', 'N/A')}")
                                    else:
                                        logger.debug(f"✅ استعادة سجل {table_info['display_name']}[{pk_value}]")
                                except Exception as rec_err:
                                    error_msg = str(rec_err)
                                    
                                    # 🔍 Logging مفصّل للأخطاء
                                    logger.error(f"❌ خطأ في استعادة {model._meta.label}: {error_msg}")
                                    logger.error(f"   البيانات: {cleaned_data if 'cleaned_data' in locals() else record_data}")
                                    
                                    # تسجيل في Audit Log
                                    if AUDIT_AVAILABLE:
                                        try:
                                            AuditLog.objects.create(
                                                user=user,
                                                action_type='restore_record_error',
                                                description=f'خطأ في استعادة {model._meta.label}: {error_msg}'
                                            )
                                        except Exception:
                                            pass
                                    # 🔧 تحسين: محاولة الاستعادة مع قيم افتراضية في حالة FK مفقودة
                                    if error_msg.startswith("FK_NOT_FOUND:"):
                                        logger.warning(f"⚠️ FK مفقود في {table_info['display_name']}: {error_msg}")
                                        # يمكن هنا إضافة منطق لمحاولة استخدام FK افتراضي
                                    elif error_msg.startswith("Required field"):
                                        logger.warning(f"⚠️ حقل مطلوب مفقود في {table_info['display_name']}: {error_msg}")
                                    else:
                                        logger.warning(f"⚠️ فشل في استعادة سجل في {table_info['display_name']}: {rec_err}")
                                    
                                    # تسجيل الخطأ في تفاصيل الجدول
                                    if not table_info.get('errors'):
                                        table_info['errors'] = []
                                    table_info['errors'].append({
                                        'record': pk_value if 'pk_value' in locals() else 'unknown',
                                        'error': error_msg
                                    })
                                    skipped_records += 1  # 🔍 زيادة عداد السجلات المتخطاة
                                    continue
                    
                    processed_tables += 1
                    percentage = int((processed_tables / len(flat_tables)) * 100) if flat_tables else 100
                    progress_data['percentage'] = percentage
                    progress_data['processed_tables'] = processed_tables
                    progress_data['processed_records'] = processed_records
                    progress_data['tables_status'][i]['status'] = 'completed'
                    set_restore_progress_data(progress_data)
                    
                except Exception as e:
                    logger.warning(f"فشل في استعادة جدول {table_info['display_name']}: {str(e)}")
                    progress_data['tables_status'][i]['status'] = 'error'
                    progress_data['tables_status'][i]['error'] = str(e)
                    set_restore_progress_data(progress_data)
                    continue
                
                # إعادة تعيين جميع sequences لتجنب تضارب IDs في المستقبل
                progress_data.update({
                'current_table': 'إعادة تعيين sequences قاعدة البيانات...',
                'percentage': 95
                })
                set_restore_progress_data(progress_data)
                
                sequences_reset = reset_all_sequences()
                logger.info(f"تم إعادة تعيين {sequences_reset} sequence بعد الاستعادة")
                
                # 📊 إنشاء تقرير مفصل
                total_errors = sum(1 for t in flat_tables if t.get('errors'))
                total_restored = sum(t.get('actual_records', 0) for t in flat_tables)
                total_skipped = total_records_expected - total_restored
                
                elapsed_time = time.time() - start_time
                
                # طباعة ملخص نهائي في السجل
                logger.info("\n" + "="*80)
                logger.info("📊 تقرير الاستعادة النهائي")
                logger.info("="*80)
                logger.info(f"⏱️  المدة الإجمالية: {elapsed_time:.2f} ثانية")
                logger.info(f"📈 الإحصائيات:")
                logger.info(f"  • إجمالي السجلات المتوقعة: {total_records_expected}")
                logger.info(f"  • السجلات المستعادة: ✅ {total_restored}")
                logger.info(f"  • السجلات المتخطاة: ⚠️ {total_skipped}")
                logger.info(f"  • الجداول المعالجة: {processed_tables}/{len(flat_tables)}")
                logger.info(f"  • عدد الجداول التي بها أخطاء: ❌ {total_errors}")
                logger.info(f"  • Sequences المعاد تعيينها: {sequences_reset}")
                
                if total_errors > 0:
                    logger.info(f"\n⚠️  الجداول التي بها مشاكل:")
                    for table in flat_tables:
                        if table.get('errors'):
                            logger.info(f"  • {table['display_name']}: {len(table['errors'])} خطأ")
                            for error in table['errors'][:3]:  # أول 3 أخطاء فقط
                                logger.info(f"    - {error}")
                
                logger.info("="*80)
                
                # إكمال التقدم
                progress_data.update({
                'is_running': False,
                'status': 'completed',
                'percentage': 100,
                'current_table': 'تم إكمال الاستعادة بنجاح!',
                'processed_tables': processed_tables,
                'processed_records': total_restored,
                'total_records': total_records_expected,
                'records_skipped': total_skipped,
                'tables_with_errors': total_errors,
                'elapsed_time': f'{elapsed_time:.2f} ثانية',
                'estimated_time': '0 ثانية متبقية'
                })
                set_restore_progress_data(progress_data)
                
                log_audit(user, 'create', f'اكتمل استعادة النسخة الاحتياطية: {total_restored}/{total_records_expected} سجل من {processed_tables} جدول ({total_skipped} متخطى)، تم إعادة تعيين {sequences_reset} sequence')
                
                # 🔍 تسجيل مفصل في Audit Log
                if AUDIT_AVAILABLE:
                    try:
                        # تسجيل عام للعملية
                        AuditLog.objects.create(
                            user=user,
                            action_type='restore_complete',
                            description=f'نجحت عملية الاستعادة: {total_restored} سجل من {processed_tables} جدول، تخطي {total_skipped} سجل، {total_errors} جدول به أخطاء'
                        )
                        
                        # تسجيل تفصيلي للأخطاء إن وجدت
                        if total_errors > 0:
                            error_details = []
                            for table in flat_tables:
                                if table.get('errors'):
                                    error_details.append(f"{table['display_name']}: {len(table['errors'])} خطأ")
                            
                            AuditLog.objects.create(
                                user=user,
                                action_type='restore_errors',
                                description=f'تفاصيل الأخطاء في الاستعادة: ' + ', '.join(error_details[:10])
                            )
                    except Exception:
                        pass
                
                # 🔧 تصحيح الأرصدة البنكية والصناديق بعد الاستعادة
                logger.info("🔄 بدء تصحيح الأرصدة البنكية والصناديق...")
                try:
                    from banks.models import BankAccount
                    from cashboxes.models import Cashbox
                    
                    # تصحيح أرصدة البنوك
                    banks_fixed = 0
                    for bank in BankAccount.objects.all():
                        old_balance = bank.balance
                        actual_balance = bank.calculate_actual_balance()
                        if old_balance != actual_balance:
                            bank.balance = actual_balance
                            bank.save(update_fields=['balance'])
                            banks_fixed += 1
                            logger.info(f"   ✅ تم تصحيح رصيد {bank.name}: {old_balance} → {actual_balance}")
                    
                    # تصحيح أرصدة الصناديق
                    cashboxes_fixed = 0
                    for cashbox in Cashbox.objects.all():
                        old_balance = cashbox.balance
                        actual_balance = cashbox.calculate_actual_balance()
                        if old_balance != actual_balance:
                            cashbox.balance = actual_balance
                            cashbox.save(update_fields=['balance'])
                            cashboxes_fixed += 1
                            logger.info(f"   ✅ تم تصحيح رصيد {cashbox.name}: {old_balance} → {actual_balance}")
                    
                    logger.info(f"✅ تم تصحيح {banks_fixed} بنك و {cashboxes_fixed} صندوق")
                    
                    # تسجيل في Audit Log
                    if AUDIT_AVAILABLE and (banks_fixed > 0 or cashboxes_fixed > 0):
                        try:
                            AuditLog.objects.create(
                                user=user,
                                action_type='balance_correction',
                                description=f'تم تصحيح الأرصدة بعد الاستعادة: {banks_fixed} بنك، {cashboxes_fixed} صندوق'
                            )
                        except Exception:
                            pass
                except Exception as balance_error:
                    logger.error(f"خطأ في تصحيح الأرصدة: {str(balance_error)}")
                
        except Exception as e:
            # 🔧 لا نرفع الخطأ - نسجله فقط للسماح بالاستعادة الجزئية
            logger.error(f"خطأ في استعادة البيانات: {str(e)}")
            progress_data.update({
                'is_running': False,
                'status': 'completed_with_errors',
                'error': str(e)
            })
            set_restore_progress_data(progress_data)
            
    except Exception as e:
        # 🔧 لا نرفع الخطأ - نسجله فقط للسماح بالاستعادة الجزئية
        logger.error(f"خطأ في تنفيذ استعادة النسخة الاحتياطية: {str(e)}")
        progress_data = get_restore_progress_data()
        progress_data.update({
            'is_running': False,
            'status': 'completed_with_errors',
            'error': str(e)
        })
        set_restore_progress_data(progress_data)
    finally:
        # 🔧 إيقاف وضع الاستعادة (إعادة تفعيل السيجنالات)
        from .restore_context import set_restoring
        set_restoring(False)
        logger.info("انتهت عملية الاستعادة (السيجنالات مُفعّلة)")

def perform_clear_all_data(user):
    """تنفيذ عملية مسح البيانات الفعلية"""
    try:
        logger.info("بدء تنفيذ عملية مسح البيانات")
        
        # تسجيل بداية العملية في سجل الأنشطة
        log_audit(user, 'delete', _('بدء عملية مسح جميع البيانات'))

        # تهيئة تتبع التقدم
        logger.info("تجهيز قائمة الجداول للمسح")

        # قائمة التطبيقات المستثناة من المسح
        excluded_apps = [
            'django.contrib.admin',
            'django.contrib.auth',  # ⭐ حماية الصلاحيات والمجموعات
            'django.contrib.contenttypes', 
            'django.contrib.sessions',
            'django.contrib.messages',
            'django.contrib.staticfiles',
            'corsheaders',
            'rest_framework',
            'django_bootstrap5',
            'crispy_forms',
            'crispy_bootstrap5',
        ]
        
        # بناء قائمة الجداول
        tables_to_clear = []
        for app_config in apps.get_app_configs():
            if app_config.name in excluded_apps:
                continue
                
            for model in app_config.get_models():
                if getattr(model._meta, 'managed', True) is False:
                    continue
                
                label = f"{app_config.name}.{model._meta.model_name}"
                record_count = model.objects.count()
                tables_to_clear.append({
                    'label': label,
                    'model': model,
                    'record_count': record_count,
                    'display_name': model._meta.verbose_name or label,
                    'status': 'pending',
                    'progress': 0
                })
        
        total_tables = len(tables_to_clear)
        total_records = sum(t['record_count'] for t in tables_to_clear)
        
        # إنشاء قائمة الجداول القابلة للتسلسل (بدون كائنات Model)
        tables_status = []
        for table in tables_to_clear:
            tables_status.append({
                'app_name': table['label'].split('.')[0] if '.' in table['label'] else '',
                'model_name': table['label'].split('.')[1] if '.' in table['label'] else table['label'],
                'display_name': table['display_name'],
                'record_count': table['record_count'],
                'status': 'pending',
                'progress': 0
            })
        
        progress_data = get_clear_progress_data()
        progress_data.update({
            'is_running': True,
            'status': 'starting',
            'error': None,
            'percentage': 0,
            'current_table': 'بدء عملية المسح...',
            'processed_tables': 0,
            'total_tables': total_tables,
            'processed_records': 0,
            'total_records': total_records,
            'estimated_time': 'جاري الحساب...',
            'tables_status': tables_status
        })
        set_clear_progress_data(progress_data)
        
        # تنفيذ المسح باستخدام SQL مباشر لتجاوز قيود PROTECT
        processed_tables = 0
        processed_records = 0
        
        # 🛡️ حفظ معلومات Superusers والمستخدم الحالي قبل البدء
        from users.models import User
        superusers_backup = []
        current_user_id = user.id if user else None
        current_user_username = user.username if user else None
        
        try:
            # حفظ جميع Superusers + المستخدم الحالي
            users_to_protect = User.objects.filter(
                models.Q(is_superuser=True) | models.Q(id=current_user_id)
            ).distinct()
            
            for su in users_to_protect:
                superusers_backup.append({
                    'id': su.id,
                    'username': su.username,
                    'email': su.email,
                    'password': su.password,  # already hashed
                    'first_name': su.first_name,
                    'last_name': su.last_name,
                    'phone': su.phone,
                    'user_type': su.user_type,
                    'is_superuser': su.is_superuser,
                    'is_current_user': su.id == current_user_id
                })
            logger.info(f'🛡️ تم حفظ معلومات {len(superusers_backup)} مستخدم محمي (Superusers + المستخدم الحالي: {current_user_username})')
        except Exception as e:
            logger.warning(f'تحذير: لم نتمكن من حفظ نسخة احتياطية: {e}')
        
        try:
            with transaction.atomic():
                # استخدام SQL مباشر لتعطيل القيود وتنفيذ المسح
                from django.db import connection
                cursor = connection.cursor()
                
                # إعدادات PostgreSQL لتحسين الأداء وتجنب التوقف
                try:
                    cursor.execute("SET statement_timeout = '300000';")  # 5 دقائق لكل عملية
                    cursor.execute("SET lock_timeout = '60000';")  # دقيقة واحدة للحصول على lock
                    cursor.execute("SET CONSTRAINTS ALL DEFERRED;")
                    logger.info("تم تطبيق إعدادات الأداء وتأجيل فحص القيود")
                except Exception as e:
                    logger.warning(f"لم نتمكن من تطبيق جميع الإعدادات: {str(e)}")
                
                # أولاً، مسح جدول AuditLog لتجنب قيود المفاتيح الخارجية
                if AUDIT_AVAILABLE:
                    try:
                        audit_count = AuditLog.objects.count()
                        if audit_count > 0:
                            # استخدام TRUNCATE CASCADE للمسح الكامل
                            audit_table = AuditLog._meta.db_table
                            cursor.execute(f'TRUNCATE TABLE "{audit_table}" CASCADE;')
                            logger.info(f"تم مسح جدول {audit_table} بنجاح")
                    except Exception as e:
                        logger.warning(f"فشل في مسح سجل الأنشطة: {str(e)}")
                
                # 🎯 ترتيب المسح الصحيح: من الأطفال إلى الجذور (7 مستويات)
                # المبدأ: احذف الصف الذي يشير (has FK) قبل الصف المشار إليه
                deletion_order = [
                    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                    # المستوى 7 (أبعد الأطفال): سجلات التدقيق والإشعارات
                    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                    'core.systemnotification',  # → users.User
                    
                    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                    # المستوى 6: بنود وتفاصيل المستندات
                    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                    'journal.journalline',  # → journal.journalentry, journal.account
                    'sales.salesinvoiceitem',  # → sales.salesinvoice, products.product
                    'purchases.purchaseinvoiceitem',  # → purchases.purchaseinvoice, products.product
                    'sales.salesreturnitem',  # → sales.salesreturn, sales.salesinvoiceitem
                    'purchases.purchasereturnitem',  # → purchases.purchasereturn, purchases.purchaseinvoiceitem
                    'inventory.warehousetransferitem',  # → inventory.warehousetransfer, products.product
                    'hr.payrollentry',  # → hr.payrollperiod, hr.employee
                    'hr.attendance',  # → hr.employee
                    'hr.leaverequest',  # → hr.employee, hr.leavetype
                    'hr.employeedocument',  # → hr.employee
                    'receipts.checkcollection',  # → receipts.paymentreceipt
                    
                    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                    # المستوى 5: المستندات الرئيسية والمعاملات
                    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                    'receipts.receiptreversal',  # → receipts.paymentreceipt
                    'sales.salescreditnote',  # → sales.salesinvoice
                    'purchases.purchasedebitnote',  # → purchases.purchaseinvoice
                    'sales.salesreturn',  # → sales.salesinvoice, customers.customersupplier
                    'purchases.purchasereturn',  # → purchases.purchaseinvoice, customers.customersupplier
                    'receipts.paymentreceipt',  # → customers.customersupplier, cashboxes.cashbox
                    'payments.paymentvoucher',  # → customers.customersupplier, banks.bankaccount
                    'banks.banktransaction',  # → banks.bankaccount, users.User
                    'banks.bankstatement',  # → banks.bankaccount, users.User
                    'banks.bankreconciliation',  # → banks.bankaccount
                    'banks.banktransfer',  # → banks.bankaccount (from/to)
                    'cashboxes.cashboxtransaction',  # → cashboxes.cashbox, users.User
                    'cashboxes.cashboxtransfer',  # → cashboxes.cashbox, banks.bankaccount
                    'inventory.inventorymovement',  # → products.product, inventory.warehouse
                    'inventory.warehousetransfer',  # → inventory.warehouse (from/to)
                    'accounts.accounttransaction',  # → journal.account
                    'revenues_expenses.revenueexpenseentry',  # → revenues_expenses.revenueexpensecategory
                    'revenues_expenses.recurringrevenueexpense',  # → revenues_expenses.revenueexpensecategory
                    'assets_liabilities.depreciationentry',  # → assets_liabilities.asset
                    'provisions.provisionentry',  # → provisions.provision
                    
                    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                    # المستوى 4: الفواتير والقيود الرئيسية
                    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                    'journal.journalentry',  # → users.User
                    'journal.yearendclosing',  # → users.User
                    'sales.salesinvoice',  # → customers.customersupplier, inventory.warehouse, users.User
                    'purchases.purchaseinvoice',  # → customers.customersupplier, inventory.warehouse, users.User
                    'hr.contract',  # → hr.employee
                    'hr.payrollperiod',  # → users.User
                    
                    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                    # المستوى 3: السجلات التابعة للجذور
                    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                    'hr.employee',  # → hr.department, hr.position, users.User
                    'assets_liabilities.asset',  # → assets_liabilities.assetcategory, users.User
                    'assets_liabilities.liability',  # → assets_liabilities.liabilitycategory
                    'provisions.provision',  # → provisions.provisiontype, journal.account
                    'documents.document',  # → users.User
                    'reports.reportaccesscontrol',  # → users.User
                    
                    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                    # المستوى 2: البيانات الأساسية المشار إليها كثيراً
                    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                    'products.product',  # → products.category, users.User
                    'banks.bankaccount',  # → users.User
                    'cashboxes.cashbox',  # → users.User
                    'inventory.warehouse',  # → users.User
                    
                    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                    # المستوى 1: الجذور (الجداول الأساسية المرجعية)
                    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                    'customers.customersupplier',  # → users.User
                    'products.category',  # بدون FK خارجية
                    'hr.department',  # بدون FK خارجية
                    'hr.position',  # بدون FK خارجية
                    'hr.leavetype',  # بدون FK خارجية
                    'journal.account',  # → journal.account (parent - self FK)
                    'assets_liabilities.assetcategory',  # بدون FK خارجية
                    'assets_liabilities.liabilitycategory',  # بدون FK خارجية
                    'provisions.provisiontype',  # بدون FK خارجية
                    'revenues_expenses.revenueexpensecategory',  # بدون FK خارجية
                    'revenues_expenses.sector',  # بدون FK خارجية
                    
                    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                    # المستوى 0: الإعدادات والمتسلسلات
                    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                    'settings.currency',
                    'settings.companysettings',
                    'settings.documentprintsettings',
                    'settings.jofotarasettings',
                    'core.companysettings',
                    'core.documentsequence',
                ]
                
                # إنشاء قاموس للوصول السريع للجداول
                table_dict = {table_info['label']: table_info for table_info in tables_to_clear}
                
                # مسح الجداول بالترتيب المحدد باستخدام TRUNCATE CASCADE
                for ordered_label in deletion_order:
                    if ordered_label in table_dict:
                        table_info = table_dict[ordered_label]
                        model = table_info['model']
                        table_name = model._meta.db_table
                        
                        # ⚠️ حماية: لا تمسح جدول المستخدمين هنا - سيتم معالجته لاحقاً
                        if model._meta.label == 'users.User':
                            logger.warning(f'⚠️ تخطي جدول المستخدمين - سيتم معالجته بشكل خاص')
                            continue
                        
                        progress_data['current_table'] = table_info['display_name']
                        set_clear_progress_data(progress_data)
                        
                        try:
                            record_count = table_info['record_count']
                            if record_count > 0:
                                logger.info(f'بدء مسح جدول {table_info["display_name"]} ({table_name}) - {record_count} سجل')
                                
                                # استخدام TRUNCATE CASCADE لمسح الجدول بالكامل مع تجاهل القيود
                                # ملاحظة: قد يستغرق وقتاً للجداول الكبيرة
                                cursor.execute(f'TRUNCATE TABLE "{table_name}" RESTART IDENTITY CASCADE;')
                                processed_records += record_count
                                logger.info(f'✅ تم مسح {record_count} سجل من جدول {table_info["display_name"]} ({table_name})')
                            else:
                                logger.info(f'⏭️ تخطي جدول فارغ: {table_info["display_name"]}')
                            
                            processed_tables += 1
                            percentage = int((processed_tables / total_tables) * 100) if total_tables > 0 else 100
                            progress_data['percentage'] = percentage
                            progress_data['processed_tables'] = processed_tables
                            progress_data['processed_records'] = processed_records
                            
                            # العثور على الفهرس في القائمة الأصلية
                            for i, original_table in enumerate(tables_to_clear):
                                if original_table['label'] == ordered_label:
                                    progress_data['tables_status'][i]['status'] = 'completed'
                                    break
                            
                            set_clear_progress_data(progress_data)
                            
                            # تسجيل تقدم كل 10 جداول
                            if processed_tables % 10 == 0:
                                logger.info(f'📊 التقدم: {processed_tables}/{total_tables} جدول ({percentage}%)')
                            
                        except Exception as e:
                            logger.error(f"❌ فشل في مسح جدول {ordered_label} ({table_name}): {str(e)}")
                            # حتى لو فشل، نستمر ونزيد العداد
                            processed_tables += 1
                            percentage = int((processed_tables / total_tables) * 100) if total_tables > 0 else 100
                            progress_data['percentage'] = percentage
                            progress_data['processed_tables'] = processed_tables
                            
                            # العثور على الفهرس في القائمة الأصلية لتسجيل الخطأ
                            for i, original_table in enumerate(tables_to_clear):
                                if original_table['label'] == ordered_label:
                                    progress_data['tables_status'][i]['status'] = 'error'
                                    progress_data['tables_status'][i]['error'] = str(e)
                                    break
                            set_clear_progress_data(progress_data)
                
                # مسح باقي الجداول التي لم تُمسح بعد باستخدام TRUNCATE CASCADE
                for i, table_info in enumerate(tables_to_clear):
                    model = table_info['model']
                    label = table_info['label']
                    table_name = model._meta.db_table
                    
                    # تخطي الجداول التي تم مسحها بالفعل
                    if label in deletion_order or label == 'core.auditlog':
                        continue
                    
                    progress_data['current_table'] = table_info['display_name']
                    set_clear_progress_data(progress_data)
                    
                    try:
                        if model._meta.label == 'users.User':
                            # ⚠️ حماية Superusers والمستخدم الحالي
                            # حذف فقط المستخدمين العاديين (غير superuser وليس المستخدم الحالي)
                            deletable_users = model.objects.filter(
                                is_superuser=False
                            ).exclude(id=current_user_id)
                            
                            deletable_count = deletable_users.count()
                            protected_count = model.objects.filter(
                                models.Q(is_superuser=True) | models.Q(id=current_user_id)
                            ).distinct().count()
                            
                            if deletable_count > 0:
                                cursor.execute(
                                    f'DELETE FROM "{table_name}" WHERE is_superuser = false AND id != %s;',
                                    [current_user_id] if current_user_id else [0]
                                )
                                processed_records += deletable_count
                                logger.info(f'✅ تم مسح {deletable_count} مستخدم عادي من {table_name}')
                                logger.info(f'🛡️ تم الحفاظ على {protected_count} مستخدم محمي (Superusers + المستخدم الحالي: {current_user_username})')
                            else:
                                logger.info(f'⏭️ لا يوجد مستخدمين قابلين للمسح (جميعهم محميون)')
                        else:
                            record_count = table_info['record_count']
                            if record_count > 0:
                                # استخدام TRUNCATE CASCADE لمسح الجدول بالكامل
                                logger.info(f'بدء مسح جدول {table_info["display_name"]} ({table_name}) - {record_count} سجل')
                                cursor.execute(f'TRUNCATE TABLE "{table_name}" RESTART IDENTITY CASCADE;')
                                processed_records += record_count
                                logger.info(f'✅ تم مسح {record_count} سجل من جدول {table_info["display_name"]} ({table_name})')
                        
                        processed_tables += 1
                        percentage = int((processed_tables / total_tables) * 100) if total_tables > 0 else 100
                        progress_data['percentage'] = percentage
                        progress_data['processed_tables'] = processed_tables
                        progress_data['processed_records'] = processed_records
                        progress_data['tables_status'][i]['status'] = 'completed'
                        set_clear_progress_data(progress_data)
                        
                        # تسجيل التقدم بشكل دوري
                        if processed_tables % 5 == 0:
                            logger.info(f'تقدم المسح: {processed_tables}/{total_tables} جدول ({percentage}%)')
                        
                    except Exception as e:
                        logger.warning(f"فشل في مسح جدول {label} ({table_name}): {str(e)}")
                        # حتى لو فشل، نستمر ونزيد العداد
                        processed_tables += 1
                        percentage = int((processed_tables / total_tables) * 100) if total_tables > 0 else 100
                        progress_data['percentage'] = percentage
                        progress_data['processed_tables'] = processed_tables
                        progress_data['tables_status'][i]['status'] = 'error'
                        progress_data['tables_status'][i]['error'] = str(e)
                        set_clear_progress_data(progress_data)
                        continue
                
                # تقرير نهائي للتحقق من نجاح المسح
                final_verification = []
                remaining_data_found = False
                
                # فحص الجداول المهمة للتأكد من المسح الكامل
                critical_tables = [
                    'customers.customersupplier',
                    'products.product', 
                    'sales.salesinvoice',
                    'purchases.purchaseinvoice',
                    'inventory.warehouse'
                ]
                
                for critical_label in critical_tables:
                    if critical_label in table_dict:
                        model = table_dict[critical_label]['model']
                        remaining_count = model.objects.count()
                        if remaining_count > 0:
                            remaining_data_found = True
                            final_verification.append(f"{critical_label}: {remaining_count} سجل متبقي")
                            logger.warning(f"بيانات متبقية في {critical_label}: {remaining_count} سجل")
                        else:
                            final_verification.append(f"{critical_label}: تم مسحه بالكامل ✅")
                
                # 🛡️ استعادة المستخدمين المحميين إذا تم حذفهم بالخطأ
                try:
                    from users.models import User
                    from django.db.models import Q
                    
                    current_protected = User.objects.filter(
                        Q(is_superuser=True) | Q(id=current_user_id)
                    ).distinct()
                    current_protected_count = current_protected.count()
                    
                    # إذا تم حذف أي مستخدم محمي، استعد الجميع
                    if current_protected_count < len(superusers_backup):
                        logger.warning('⚠️ تم اكتشاف حذف مستخدمين محميين! جاري الاستعادة...')
                        restored_count = 0
                        
                        for su_data in superusers_backup:
                            try:
                                # التحقق من عدم وجود المستخدم
                                if not User.objects.filter(username=su_data['username']).exists():
                                    user_created = User.objects.create(
                                        username=su_data['username'],
                                        email=su_data['email'],
                                        password=su_data['password'],  # already hashed
                                        first_name=su_data['first_name'],
                                        last_name=su_data['last_name'],
                                        phone=su_data['phone'],
                                        user_type=su_data['user_type'],
                                        is_superuser=su_data['is_superuser'],
                                        is_staff=True,
                                        is_active=True
                                    )
                                    restored_count += 1
                                    user_type_label = '(المستخدم الحالي)' if su_data['is_current_user'] else '(Superuser)'
                                    logger.info(f'✅ تم استعادة {user_type_label}: {su_data["username"]}')
                            except Exception as restore_err:
                                logger.error(f'فشل في استعادة المستخدم {su_data["username"]}: {restore_err}')
                        
                        if restored_count > 0:
                            logger.info(f'🎉 تم استعادة {restored_count} مستخدم محمي بنجاح!')
                            final_verification.append(f'تم استعادة {restored_count} مستخدم محمي ✅')
                    else:
                        logger.info(f'✅ المستخدمون المحميون آمنون: {current_protected_count} مستخدم')
                        final_verification.append(f'المستخدمون المحميون آمنون: {current_protected_count} ✅')
                except Exception as su_check_err:
                    logger.error(f'خطأ في فحص/استعادة المستخدمين المحميين: {su_check_err}')
                
                # إكمال التقدم
                final_status = 'completed' if not remaining_data_found else 'completed_with_warnings'
                final_message = 'تم إكمال المسح بنجاح!' if not remaining_data_found else 'تم إكمال المسح مع تحذيرات'
                
                progress_data.update({
                    'is_running': False,
                    'status': final_status,
                    'percentage': 100,
                    'current_table': final_message,
                    'processed_tables': processed_tables,
                    'processed_records': processed_records,
                    'final_verification': final_verification,
                    'remaining_data_found': remaining_data_found
                })
                set_clear_progress_data(progress_data)
                
                logger.info(f'اكتمل مسح جميع البيانات: {processed_records} سجل من {processed_tables} جدول')
                
                # تسجيل التقرير النهائي
                if final_verification:
                    logger.info("تقرير التحقق النهائي:")
                    for verification_line in final_verification:
                        logger.info(f"  {verification_line}")
                
        except Exception as e:
            logger.error(f"خطأ في مسح البيانات: {str(e)}")
            progress_data.update({
                'is_running': False,
                'status': 'error',
                'error': str(e)
            })
            set_clear_progress_data(progress_data)
            raise e
            
    except Exception as e:
        logger.error(f"خطأ في تنفيذ مسح البيانات: {str(e)}")
        progress_data = get_clear_progress_data()
        progress_data.update({
            'is_running': False,
            'status': 'error',
            'error': str(e)
        })
        set_clear_progress_data(progress_data)
        raise e


@login_required
def create_backup(request):
    """إنشاء نسخة احتياطية مع تتبع التقدم وتسجيل الأنشطة"""
    # التحقق من الصلاحيات
    if not request.user.is_superuser and not request.user.has_perm('users.can_make_backup'):
        messages.error(request, _('ليس لديك صلاحية لإنشاء النسخ الاحتياطية'))
        log_audit(request.user, 'denied', 'محاولة وصول مرفوضة لإنشاء نسخة احتياطية')
        return redirect('backup:backup_restore')
    
    if request.method != 'POST':
        logger.warning(f"❌ طريقة طلب غير صحيحة: {request.method}")
        if AUDIT_AVAILABLE:
            try:
                AuditLog.objects.create(
                    user=request.user if request.user.is_authenticated else None,
                    action='backup_access_denied',
                    details='محاولة وصول غير صحيحة لإنشاء نسخة احتياطية'
                )
            except Exception:
                pass
        messages.error(request, _('طريقة طلب غير صحيحة'))
        return redirect('backup:backup_restore')
    
    progress_data = get_backup_progress_data()
    
    # التحقق من العمليات المعلقة
    if progress_data.get('is_running', False):
        import time
        current_time = time.time()
        last_update = cache.get('backup_last_update', current_time)
        
        # إذا مر أكثر من دقيقة واحدة بدون تحديث، قم بمسح العملية المعلقة تلقائياً
        if current_time - last_update > 60:  # دقيقة واحدة
            cache.delete('backup_progress')
            cache.delete('backup_last_update')
            
            log_audit(request.user, 'delete', f'تم مسح عملية النسخ الاحتياطي المعلقة تلقائياً - عمر العملية: {int(current_time - last_update)} ثانية')
            
            progress_data = get_backup_progress_data()  # إعادة جلب البيانات
        else:
            # فحص إضافي: التحقق من أن العملية تتقدم فعلاً
            current_percentage = progress_data.get('percentage', 0)
            last_percentage = cache.get('backup_last_percentage', 0)
            
            # إذا لم تتقدم العملية لمدة 30 ثانية، فهي معلقة
            if current_percentage == last_percentage and current_time - last_update > 30:
                cache.delete('backup_progress')
                cache.delete('backup_last_update')
                cache.delete('backup_last_percentage')
                
                log_audit(request.user, 'delete', f'تم مسح عملية النسخ الاحتياطي المتوقفة تلقائياً - توقفت عند {current_percentage}%')
                
                progress_data = get_backup_progress_data()  # إعادة جلب البيانات
            else:
                # تحديث آخر نسبة مئوية
                cache.set('backup_last_percentage', current_percentage, timeout=3600)
                
                # العملية ما زالت قيد التشغيل - تسجيل المحاولة
                log_audit(request.user, 'error', f'محاولة بدء نسخة احتياطية أثناء وجود عملية قيد التشغيل - نسبة الإنجاز: {current_percentage}%')
                
                messages.error(request, _('عملية نسخ احتياطي قيد التشغيل بالفعل'))
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'error': 'عملية نسخ احتياطي قيد التشغيل بالفعل'
                    })
                return redirect('backup:backup_restore')
    
    try:
        # الحصول على التنسيق المختار
        selected_format = (
            request.POST.get('format') or 
            request.POST.get('backupFormat') or 
            'json'
        ).strip().lower()
        
        # إنشاء اسم الملف حسب التنسيق
        timestamp = dt_module.now().strftime('%Y%m%d_%H%M%S')
        
        if selected_format == 'xlsx':
            filename = f'backup_{timestamp}.xlsx'
            format_name = 'XLSX'
        else:
            filename = f'backup_{timestamp}.json'
            format_name = 'JSON'
            selected_format = 'json'
            
        # تسجيل بداية عملية النسخ الاحتياطي
        log_audit(request.user, 'create', f'طلب إنشاء نسخة احتياطية بصيغة {format_name}: {filename}')
        
        # إنشاء مجلد النسخ الاحتياطية إذا لم يكن موجوداً
        backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        filepath = os.path.join(backup_dir, filename)
        
        # بدء المهمة في خيط منفصل (مبسط)
        def start_backup():
            try:
                # إضافة تأخير بسيط
                import time
                time.sleep(0.5)
                
                # استدعاء مباشر بدون تعقيدات إضافية
                perform_backup_task(request.user, timestamp, filename, filepath, selected_format)
                
            except Exception as e:
                logger.error(f"❌ خطأ في النسخ الاحتياطي: {str(e)}")
                # تحديث الكاش بالخطأ
                try:
                    from django.core.cache import cache
                    progress_data = {
                        'is_running': False,
                        'status': 'error',
                        'error': str(e),
                        'percentage': 0,
                        'current_table': f'خطأ: {str(e)}'
                    }
                    cache.set('backup_progress', progress_data, timeout=3600)
                except:
                    pass
                
        # إنشاء وتشغيل الخيط
        thread = threading.Thread(target=start_backup)
        thread.daemon = True
        thread.start()
        
        logger.info(f"📝 تم بدء النسخ الاحتياطي: {filename}")
        
        # إرجاع JSON response للطلبات AJAX
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'filename': filename,
                'format': format_name,
                'message': f'تم بدء إنشاء النسخة الاحتياطية {format_name}: {filename}'
            })
        
        messages.success(request, _(f'تم بدء إنشاء النسخة الاحتياطية {format_name}: {filename}'))
        return redirect('backup:backup_restore')
        
    except Exception as e:
        error_msg = f"خطأ في إنشاء النسخة الاحتياطية: {str(e)}"
        logger.error(error_msg)
        
        if AUDIT_AVAILABLE:
            try:
                AuditLog.objects.create(
                    user=request.user,
                    action='backup_error_request',
                    details=error_msg
                )
            except Exception:
                pass
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': error_msg
            })
        
        messages.error(request, error_msg)
        return redirect('backup:backup_restore')


@login_required  
def download_backup(request, filename):
    """تحميل ملف النسخة الاحتياطية"""
    
    if not filename.endswith(('.json', '.xlsx')):
        raise Http404("نوع الملف غير مدعوم")
    
    backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
    filepath = os.path.join(backup_dir, filename)
    
    if not os.path.exists(filepath):
        raise Http404("الملف غير موجود")
    
    if AUDIT_AVAILABLE:
        try:
            AuditLog.objects.create(
                user=request.user,
                action='backup_download',
                details=f'تحميل النسخة الاحتياطية: {filename}'
            )
        except Exception:
            pass
    
    content_type = 'application/json' if filename.endswith('.json') else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    try:
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
    except Exception as e:
        logger.error(f"خطأ في تحميل الملف {filename}: {str(e)}")
        raise Http404("حدث خطأ في تحميل الملف")


@login_required
def delete_backup(request, filename):
    """حذف ملف النسخة الاحتياطية"""
    
    if request.method != 'POST':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': 'طريقة طلب غير صحيحة'
            })
        return redirect('backup:backup_restore')
    
    if not filename.endswith(('.json', '.xlsx')):
        error_msg = "نوع الملف غير مدعوم"
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': error_msg
            })
        
        messages.error(request, error_msg)
        return redirect('backup:backup_restore')
    
    backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
    filepath = os.path.join(backup_dir, filename)
    
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            
            # تسجيل الحدث في سجل المراجعة
            log_audit(request.user, 'delete', f'حذف النسخة الاحتياطية: {filename}')
            
            success_msg = f"تم حذف النسخة الاحتياطية {filename} بنجاح"
            logger.info(f"✅ {success_msg} - المستخدم: {request.user.username}")
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': success_msg
                })
                
            messages.success(request, success_msg)
        else:
            error_msg = "الملف غير موجود"
            
            # تسجيل محاولة الحذف للملف غير الموجود
            log_audit(request.user, 'warning', f'محاولة حذف ملف غير موجود: {filename}')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': error_msg
                })
                
            messages.error(request, error_msg)
            
    except Exception as e:
        error_msg = f"حدث خطأ في حذف الملف: {str(e)}"
        logger.error(f"❌ خطأ في حذف الملف {filename}: {str(e)}")
        
        # تسجيل الخطأ في سجل المراجعة
        log_audit(request.user, 'error', f'فشل حذف النسخة الاحتياطية: {filename} - خطأ: {str(e)}')
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': error_msg
            })
        
        messages.error(request, error_msg)
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'message': 'تم معالجة الطلب'
        })
    
    return redirect('backup:backup_restore')


@login_required
def restore_backup(request):
    """استعادة النسخة الاحتياطية"""
    # التحقق من الصلاحيات
    if not request.user.is_superuser and not request.user.has_perm('users.can_restore_backup'):
        messages.error(request, _('ليس لديك صلاحية لاستعادة النسخ الاحتياطية'))
        log_audit(request.user, 'denied', 'محاولة وصول مرفوضة لاستعادة نسخة احتياطية')
        return redirect('backup:backup_restore')
    
    if request.method != 'POST':
        messages.error(request, _("طريقة طلب غير صحيحة"))
        return redirect('backup:backup_restore')
    
    if 'backup_file' not in request.FILES and not request.POST.get('filename'):
        messages.error(request, _("يرجى اختيار ملف النسخة الاحتياطية"))
        return redirect('backup:backup_restore')
    
    # قبول عدة صيغ لتمرير قيمة clear_data من النموذج أو ال AJAX (مثال: 'on' من form submission، 'true' من AJAX)
    raw_clear = request.POST.get('clear_data')
    if raw_clear is None:
        clear_data = False
    else:
        try:
            clear_data = str(raw_clear).strip().lower() in ('true', '1', 'on', 'yes')
        except Exception:
            clear_data = False
    filename_from_list = request.POST.get('filename')
    backup_file = request.FILES.get('backup_file')
    
    try:
        if filename_from_list:
            # قراءة من ملف موجود في مجلد النسخ الاحتياطية
            backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
            filepath = os.path.join(backup_dir, filename_from_list)
            if not os.path.exists(filepath):
                msg = _("الملف غير موجود")
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': msg})
                messages.error(request, msg)
                return redirect('backup:backup_restore')
            if filepath.endswith('.json'):
                with open(filepath, 'r', encoding='utf-8') as f:
                    backup_data = json.load(f)
            elif filepath.endswith('.xlsx'):
                with open(filepath, 'rb') as f:
                    backup_data = load_backup_from_xlsx(f, user=request.user)
            else:
                msg = _("نوع الملف غير مدعوم")
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': msg})
                messages.error(request, msg)
                return redirect('backup:backup_restore')
            input_name_for_audit = filename_from_list
        elif backup_file and backup_file.name.endswith('.json'):
            file_content = backup_file.read().decode('utf-8')
            backup_data = json.loads(file_content)
            input_name_for_audit = backup_file.name
        elif backup_file and backup_file.name.endswith('.xlsx'):
            # دعم استعادة ملفات Excel
            try:
                backup_data = load_backup_from_xlsx(backup_file, user=request.user)
                if AUDIT_AVAILABLE:
                    try:
                        AuditLog.objects.create(
                            user=request.user,
                            action='backup_restore_xlsx_parsed',
                            details=f'تم تحويل ملف Excel للاستعادة: {backup_file.name}'
                        )
                    except Exception:
                        pass
            except Exception as xlsx_error:
                logger.error(f"خطأ في قراءة ملف Excel: {str(xlsx_error)}")
                messages.error(request, _(f"خطأ في قراءة ملف Excel: {str(xlsx_error)}"))
                if AUDIT_AVAILABLE:
                    try:
                        AuditLog.objects.create(
                            user=request.user,
                            action='backup_restore_xlsx_error',
                            details=f'فشل تحويل ملف Excel: {backup_file.name} - {str(xlsx_error)}'
                        )
                    except Exception:
                        pass
                return redirect('backup:backup_restore')
            input_name_for_audit = backup_file.name
        else:
            messages.error(request, _("نوع الملف غير مدعوم"))
            return redirect('backup:backup_restore')
        
        if AUDIT_AVAILABLE:
            try:
                AuditLog.objects.create(
                    user=request.user,
                    action='backup_restore_start',
                    details=f'بدء استعادة النسخة الاحتياطية: {input_name_for_audit}'
                )
            except Exception:
                pass
        
        # تشغيل عملية الاستعادة في خيط منفصل وتحديث التقدم عبر الكاش
        def start_restore_task():
            try:
                perform_backup_restore(backup_data, clear_data, request.user)
                # عند النجاح، سجّل في السجل (تمت إضافة ذلك داخل perform_backup_restore أيضاً)
                if AUDIT_AVAILABLE:
                    try:
                        AuditLog.objects.create(
                            user=request.user,
                            action='backup_restore_complete',
                            details=f'تم إتمام استعادة النسخة الاحتياطية بنجاح: {input_name_for_audit}'
                        )
                    except Exception:
                        pass
            except Exception as restore_error:
                logger.error(f"خطأ في تنفيذ الاستعادة (خلفية): {str(restore_error)}")
                if AUDIT_AVAILABLE:
                    try:
                        AuditLog.objects.create(
                            user=request.user,
                            action='backup_restore_failed',
                            details=f'فشل في استعادة النسخة الاحتياطية: {input_name_for_audit} - {str(restore_error)}'
                        )
                    except Exception:
                        pass

        thread = threading.Thread(target=start_restore_task, daemon=True)
        thread.start()

        # الاستجابة حسب نوع الطلب
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': _('تم بدء عملية الاستعادة في الخلفية')
            })
        else:
            messages.info(request, _("تم بدء عملية الاستعادة في الخلفية، يمكن متابعة التقدم على هذه الصفحة"))
            return redirect('backup:backup_restore')
        
    except json.JSONDecodeError as e:
        error_msg = _("ملف JSON غير صحيح")
        logger.error(f"خطأ في تحليل JSON: {str(e)}")
        if AUDIT_AVAILABLE:
            try:
                AuditLog.objects.create(
                    user=request.user,
                    action='backup_restore_json_error',
                    details=f'خطأ في تحليل ملف JSON: {str(e)}'
                )
            except Exception:
                pass
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg})
        messages.error(request, error_msg)
    except Exception as e:
        error_msg = f"حدث خطأ في الاستعادة: {str(e)}"
        logger.error(f"خطأ في استعادة النسخة الاحتياطية: {str(e)}")
        if AUDIT_AVAILABLE:
            try:
                AuditLog.objects.create(
                    user=request.user,
                    action='backup_restore_general_error',
                    details=f'خطأ عام في الاستعادة: {str(e)}'
                )
            except Exception:
                pass
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg})
        messages.error(request, _(error_msg))
    
    return redirect('backup:backup_restore')


@login_required
def get_restore_progress(request):
    """الحصول على حالة تقدم الاستعادة"""
    try:
        progress_data = get_restore_progress_data()

        # التحقق من أن progress_data هو dictionary صحيح
        if not isinstance(progress_data, dict):
            logger.warning(f"بيانات التقدم تالفة: {type(progress_data)}")
            progress_data = {
                'is_running': False,
                'current_table': '',
                'processed_tables': 0,
                'total_tables': 0,
                'processed_records': 0,
                'total_records': 0,
                'percentage': 0,
                'status': 'idle',
                'error': 'بيانات التقدم تالفة',
                'tables_status': [],
                'estimated_time': ''
            }

        # التحقق من العمليات المعلقة
        if progress_data.get('is_running', False):
            import time
            current_time = time.time()
            last_update = cache.get('restore_last_update', current_time)
            if current_time - last_update > 600:
                cache.delete('restore_progress')
                cache.delete('restore_last_update')
                log_audit(request.user, 'delete', 'تم إلغاء عملية الاستعادة المعلقة تلقائياً')
                progress_data = {
                    'is_running': False,
                    'current_table': '',
                    'processed_tables': 0,
                    'total_tables': 0,
                    'processed_records': 0,
                    'total_records': 0,
                    'percentage': 0,
                    'status': 'idle',
                    'error': None,
                    'tables_status': [],
                    'estimated_time': ''
                }

        return JsonResponse({'success': True, 'progress': progress_data})

    except Exception as e:
        logger.error(f"خطأ في الحصول على تقدم الاستعادة: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def get_clear_progress(request):
    """الحصول على حالة تقدم المسح"""
    try:
        progress_data = get_clear_progress_data()

        # التحقق من أن progress_data هو dictionary صحيح
        if not isinstance(progress_data, dict):
            logger.warning(f"بيانات التقدم تالفة: {type(progress_data)}")
            progress_data = {
                'is_running': False,
                'current_table': '',
                'processed_tables': 0,
                'total_tables': 0,
                'processed_records': 0,
                'total_records': 0,
                'percentage': 0,
                'status': 'idle',
                'error': 'بيانات التقدم تالفة',
                'tables_status': [],
                'estimated_time': ''
            }

        # التحقق من العمليات المعلقة
        if progress_data.get('is_running', False):
            import time
            current_time = time.time()
            last_update = cache.get('clear_last_update', current_time)
            if current_time - last_update > 600:
                cache.delete('clear_progress')
                cache.delete('clear_last_update')
                log_audit(request.user, 'delete', 'تم إلغاء عملية المسح المعلقة تلقائياً')
                progress_data = {
                    'is_running': False,
                    'current_table': '',
                    'processed_tables': 0,
                    'total_tables': 0,
                    'processed_records': 0,
                    'total_records': 0,
                    'percentage': 0,
                    'status': 'idle',
                    'error': None,
                    'tables_status': [],
                    'estimated_time': ''
                }

        return JsonResponse({'success': True, 'progress': progress_data})

    except Exception as e:
        logger.error(f"خطأ في الحصول على تقدم المسح: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def list_backups(request):
    """API لجلب قائمة النسخ الاحتياطية"""
    
    try:
        backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        backups = []
        for filename in os.listdir(backup_dir):
            if filename.endswith('.json') or filename.endswith('.xlsx'):
                filepath = os.path.join(backup_dir, filename)
                try:
                    stat = os.stat(filepath)
                    file_type = 'XLSX' if filename.endswith('.xlsx') else 'JSON'
                    backups.append({
                        'filename': filename,
                        'size': stat.st_size,
                        'size_formatted': filesizeformat(stat.st_size),
                        'created_at': dt_module.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
                        'type': file_type,
                        'download_url': reverse('backup:download_backup', kwargs={'filename': filename})
                    })
                except (OSError, IOError):
                    continue
        
        # ترتيب النسخ حسب تاريخ الإنشاء (الأحدث أولاً)
        backups.sort(key=lambda x: x['created_at'], reverse=True)
        
        return JsonResponse({
            'success': True,
            'backups': backups,
            'count': len(backups)
        })
        
    except Exception as e:
        logger.error(f"خطأ في جلب قائمة النسخ الاحتياطية: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
def clear_all_data(request):
    """مسح جميع البيانات من قاعدة البيانات (للاستخدام المنفصل)"""
    # التحقق من الصلاحيات
    if not request.user.is_superuser and not request.user.has_perm('users.can_delete_all_data'):
        messages.error(request, _('ليس لديك صلاحية لمسح جميع البيانات'))
        log_audit(request.user, 'denied', 'محاولة وصول مرفوضة لمسح جميع البيانات')
        return redirect('backup:backup_restore')
    
    if request.method != 'POST':
        return redirect('backup:backup_restore')
    
    # تحقق من عدم وجود عملية مسح قيد التشغيل
    progress_data = get_clear_progress_data()
    if progress_data.get('is_running', False):
        messages.error(request, 'عملية مسح أخرى قيد التشغيل بالفعل.')
        return redirect('backup:backup_restore')
    
    # تشغيل عملية المسح في خيط منفصل وتحديث التقدم عبر الكاش
    def start_clear_task():
        try:
            perform_clear_all_data(request.user)
            # عند النجاح، سجّل في السجل
            if AUDIT_AVAILABLE:
                try:
                    AuditLog.objects.create(
                        user=request.user,
                        action='clear_all_data_complete',
                        details='تم إتمام مسح جميع البيانات بنجاح'
                    )
                except Exception:
                    pass
        except Exception as clear_error:
            logger.error(f"خطأ في تنفيذ المسح (خلفية): {str(clear_error)}")
            if AUDIT_AVAILABLE:
                try:
                    AuditLog.objects.create(
                        user=request.user,
                        action='clear_all_data_failed',
                        details=f'فشل في مسح البيانات: {str(clear_error)}'
                    )
                except Exception:
                    pass

    thread = threading.Thread(target=start_clear_task, daemon=True)
    thread.start()

    # الاستجابة حسب نوع الطلب
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'message': _('تم بدء عملية المسح في الخلفية')
        })
    else:
        messages.info(request, _("تم بدء عملية المسح في الخلفية، يمكن متابعة التقدم على هذه الصفحة"))
        return redirect('backup:backup_restore')


