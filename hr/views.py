from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.http import JsonResponse, HttpResponse
from django.urls import reverse_lazy
from django.db.models import Q, Sum, Count
from django.utils import timezone
from django.contrib.contenttypes.models import ContentType
from datetime import datetime, date, timedelta
import csv
import io
from decimal import Decimal
from django.contrib.auth import get_user_model

from .models import (
    Department, Position, Employee, Contract, Attendance, 
    LeaveType, LeaveRequest, PayrollPeriod, PayrollEntry, EmployeeDocument,
    EmployeeDeduction
)
from .forms import (
    EmployeeForm, ContractForm, AttendanceForm, LeaveRequestForm,
    PayrollPeriodForm, PayrollEntryForm, AttendanceUploadForm,
    DepartmentForm, PositionForm, EmployeeDocumentForm, EmployeeDeductionForm
)
from core.models import AuditLog
from core.utils import get_client_ip

User = get_user_model()


def create_hr_audit_log(request, action_type, content_type_model, object_id, description):
    """Helper function to create audit log for HR module"""
    try:
        AuditLog.objects.create(
            user=request.user,
            action_type=action_type,
            content_type=ContentType.objects.get_for_model(content_type_model),
            object_id=object_id,
            description=description,
            ip_address=get_client_ip(request),
        )
    except Exception:
        pass


# HRMixin removed - now everyone can access HR resources


# Employee Views
class EmployeeListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = "hr.view_hr_employee"
    model = Employee
    template_name = 'hr/employee_list.html'
    context_object_name = 'employees'
    paginate_by = 20

    def get_queryset(self):
        queryset = Employee.objects.select_related('department', 'position', 'user')
        
        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(employee_id__icontains=search) |
                Q(email__icontains=search)
            )
        
        # Filtering
        department = self.request.GET.get('department')
        if department:
            queryset = queryset.filter(department_id=department)
            
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
            
        return queryset.order_by('first_name', 'last_name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['departments'] = Department.objects.filter(is_active=True)
        context['status_choices'] = Employee.STATUS_CHOICES
        return context


class EmployeeDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    permission_required = "hr.view_hr_employee"
    model = Employee
    template_name = 'hr/employee_detail.html'
    context_object_name = 'employee'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employee = self.get_object()
        
        # Contracts
        context['contracts'] = employee.contracts.all()[:5]
        
        # Recent attendance
        context['recent_attendance'] = employee.attendances.all()[:10]
        
        # Recent leave requests
        context['recent_leaves'] = employee.leave_requests.all()[:5]
        
        # Recent payrolls
        context['recent_payrolls'] = employee.payroll_entries.all()[:5]
        
        # Leave balance
        context['leave_balance'] = employee.get_current_leave_balance()
        
        # Active additional deductions
        context['active_deductions'] = employee.deductions.filter(is_active=True)
        
        return context


class EmployeeCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = "hr.add_hr_employee"
    model = Employee
    form_class = EmployeeForm
    template_name = 'hr/employee_form.html'
    success_url = reverse_lazy('hr:employee_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        
        # Log activity
        create_hr_audit_log(
            self.request,
            'CREATE',
            Employee,
            self.object.id,
            f'Created new employee: {self.object.full_name}'
        )
        
        messages.success(self.request, _('Employee created successfully.'))
        
        # If it's an AJAX request, return JSON with employee ID
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'employee_id': self.object.id,
                'message': str(_('Employee created successfully.'))
            })
        
        return response
    
    def form_invalid(self, form):
        # If it's an AJAX request, return JSON with errors
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'errors': form.errors,
                'message': 'Error in entered data'
            }, status=400)
        
        return super().form_invalid(form)


class EmployeeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = "hr.change_hr_employee"
    model = Employee
    form_class = EmployeeForm
    template_name = 'hr/employee_form.html'
    success_url = reverse_lazy('hr:employee_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Log activity
        create_hr_audit_log(
            self.request,
            'UPDATE',
            Employee,
            self.object.id,
            f'Updated employee data: {self.object.full_name}'
        )
        
        messages.success(self.request, _('Employee updated successfully.'))
        return response


class EmployeeDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = "hr.delete_hr_employee"
    model = Employee
    template_name = 'hr/employee_confirm_delete.html'
    success_url = reverse_lazy('hr:employee_list')

    def delete(self, request, *args, **kwargs):
        employee = self.get_object()
        employee_name = employee.full_name
        
        # Log activity
        create_hr_audit_log(
            request,
            'DELETE',
            Employee,
            employee.id,
            f'Deleted employee: {employee_name}'
        )
        
        response = super().delete(request, *args, **kwargs)
        messages.success(request, _('Employee deleted successfully.'))
        return response


# Attendance Views
class AttendanceListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = "hr.view_hr_attendance"
    model = Attendance
    template_name = 'hr/attendance_list.html'
    context_object_name = 'attendances'
    paginate_by = 20

    def get_queryset(self):
        queryset = Attendance.objects.select_related('employee', 'employee__department')
        
        # Filter by date
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
            
        # Filter by employee
        employee = self.request.GET.get('employee')
        if employee:
            queryset = queryset.filter(employee_id=employee)
            
        return queryset.order_by('-date', 'employee__first_name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['employees'] = Employee.objects.filter(status='active')
        return context


@login_required
def attendance_upload(request):
    """Upload attendance CSV file"""
    if request.method == 'POST':
        form = AttendanceUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']
            
            # Check file type
            if not csv_file.name.endswith('.csv'):
                messages.error(request, _('File must be CSV format.'))
                return render(request, 'hr/attendance_upload.html', {'form': form})
            
            # Read file
            data_set = csv_file.read().decode('UTF-8')
            io_string = io.StringIO(data_set)
            next(io_string)  # Skip header
            
            created_count = 0
            error_count = 0
            
            for column in csv.reader(io_string, delimiter=',', quotechar='"'):
                try:
                    employee_id = column[0]
                    date_str = column[1]
                    check_in = column[2] if column[2] else None
                    check_out = column[3] if column[3] else None
                    
                    # Find employee
                    employee = Employee.objects.get(employee_id=employee_id)
                    
                    # Convert date
                    attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    
                    # Create attendance record
                    attendance, created = Attendance.objects.get_or_create(
                        employee=employee,
                        date=attendance_date,
                        defaults={
                            'check_in_time': check_in,
                            'check_out_time': check_out,
                            'is_manual_entry': False,
                            'created_by': request.user
                        }
                    )
                    
                    if created:
                        created_count += 1
                    
                except Exception as e:
                    error_count += 1
                    continue
            
            # Log activity
            create_hr_audit_log(
                request,
                'UPLOAD',
                Attendance,
                None,
                f'Uploaded attendance file: {created_count} new records, {error_count} errors'
            )
            
            if created_count > 0:
                messages.success(request, _(f'Successfully uploaded {created_count} attendance records.'))
            if error_count > 0:
                messages.warning(request, _(f'Failed to upload {error_count} records.'))
                
            return redirect('hr:attendance_list')
    else:
        form = AttendanceUploadForm()
    
    return render(request, 'hr/attendance_upload.html', {'form': form})


# Leave Views
class LeaveRequestListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = "hr.view_hr_leaverequest"
    model = LeaveRequest
    template_name = 'hr/leave_request_list.html'
    context_object_name = 'leave_requests'
    paginate_by = 20

    def get_queryset(self):
        queryset = LeaveRequest.objects.select_related('employee', 'leave_type')
        
        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
            
        return queryset.order_by('-created_at')


@login_required
@permission_required("hr.approve_hr_leaverequest", raise_exception=True)
def approve_leave_request(request, pk):
    """Approve leave request"""
    leave_request = get_object_or_404(LeaveRequest, pk=pk)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'approve':
            leave_request.status = 'approved'
            leave_request.approved_by = request.user
            leave_request.approval_date = timezone.now()
            leave_request.save()
            
            # Log activity
            create_hr_audit_log(
                request,
                'APPROVE',
                LeaveRequest,
                leave_request.id,
                f'Approved leave request: {leave_request}'
            )
            
            messages.success(request, _('Leave request approved successfully.'))
            
        elif action == 'reject':
            leave_request.status = 'rejected'
            leave_request.rejection_reason = request.POST.get('rejection_reason', '')
            leave_request.save()
            
            # Log activity
            create_hr_audit_log(
                request,
                'REJECT',
                LeaveRequest,
                leave_request.id,
                f'Rejected leave request: {leave_request}'
            )
            
            messages.success(request, _('Leave request rejected.'))
    
    return redirect('hr:leave_request_list')


# Payroll Views
class PayrollPeriodListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = "hr.view_hr_payrollperiod"
    model = PayrollPeriod
    template_name = 'hr/payroll_period_list.html'
    context_object_name = 'payroll_periods'
    paginate_by = 20


@login_required
def process_payroll(request, pk):
    """Process payroll for a specific period"""
    payroll_period = get_object_or_404(PayrollPeriod, pk=pk)
    
    if payroll_period.is_processed:
        messages.warning(request, _('Payroll period already processed.'))
        return redirect('hr:payroll_period_list')
    
    if request.method == 'POST':
        # Get all active employees
        active_employees = Employee.objects.filter(status='active')
        
        created_count = 0
        
        for employee in active_employees:
            # Calculate basic salary and allowances
            basic_salary = employee.basic_salary
            allowances = employee.allowances
            
            # Calculate overtime hours
            overtime_hours = Attendance.objects.filter(
                employee=employee,
                date__range=[payroll_period.start_date, payroll_period.end_date]
            ).aggregate(total_overtime=Sum('overtime_hours'))['total_overtime'] or 0
            
            # Calculate overtime amount (1.5x multiplier)
            hourly_rate = basic_salary / Decimal('160')  # 160 hours per month
            overtime_amount = overtime_hours * hourly_rate * Decimal('1.5')
            
            # Calculate social security deduction
            social_security_deduction = employee.basic_salary * employee.social_security_rate / 100
            
            # Calculate unpaid leave deduction
            unpaid_leave_days = LeaveRequest.objects.filter(
                employee=employee,
                start_date__range=[payroll_period.start_date, payroll_period.end_date],
                status='approved',
                leave_type__is_paid=False
            ).aggregate(total_days=Sum('days_count'))['total_days'] or 0
            
            daily_rate = basic_salary / 30
            unpaid_leave_deduction = unpaid_leave_days * daily_rate
            
            # Create payroll entry
            payroll_entry = PayrollEntry.objects.create(
                payroll_period=payroll_period,
                employee=employee,
                basic_salary=basic_salary,
                allowances=allowances,
                overtime_amount=overtime_amount,
                social_security_deduction=social_security_deduction,
                unpaid_leave_deduction=unpaid_leave_deduction,
                created_by=request.user
            )
            
            created_count += 1
        
        # Update period status
        payroll_period.is_processed = True
        payroll_period.processed_date = timezone.now()
        payroll_period.processed_by = request.user
        payroll_period.save()
        
        # Log activity
        create_hr_audit_log(
            request,
            'PROCESS',
            PayrollPeriod,
            payroll_period.id,
            f'Processed payroll for period: {payroll_period.name} ({created_count} employees)'
        )
        
        messages.success(request, _(f'Successfully processed payroll for {created_count} employees.'))
        return redirect('hr:payroll_period_list')
    
    # Preview employees
    employees = Employee.objects.filter(status='active')
    return render(request, 'hr/process_payroll.html', {
        'payroll_period': payroll_period,
        'employees': employees
    })


@login_required
def create_payroll_journal_entries(request, pk):
    """Create accounting journal entries for payroll"""
    payroll_period = get_object_or_404(PayrollPeriod, pk=pk)
    
    if not payroll_period.is_processed:
        messages.error(request, _('Payroll period must be processed first.'))
        return redirect('hr:payroll_period_list')
    
    from journal.models import JournalEntry, JournalEntryLine, Account
    from decimal import Decimal
    
    # Find required accounts
    try:
        salaries_expense_account = Account.objects.get(name__icontains='salaries')
        social_security_payable_account = Account.objects.get(name__icontains='social security')
        salaries_payable_account = Account.objects.get(name__icontains='salaries payable')
    except Account.DoesNotExist:
        messages.error(request, _('Required accounts not found. Please create salary related accounts first.'))
        return redirect('hr:payroll_period_list')
    
    # Calculate totals
    payroll_entries = payroll_period.payroll_entries.all()
    total_gross_salary = sum(entry.gross_salary for entry in payroll_entries)
    total_social_security = sum(entry.social_security_deduction for entry in payroll_entries)
    total_net_salary = sum(entry.net_salary for entry in payroll_entries)
    
    # Create journal entry
    journal_entry = JournalEntry.objects.create(
        date=payroll_period.end_date,
        description=f'Payroll entry for period: {payroll_period.name}',
        reference=f'PAYROLL-{payroll_period.id}',
        created_by=request.user
    )
    
    # Journal line: Salaries expense (debit)
    JournalEntryLine.objects.create(
        journal_entry=journal_entry,
        account=salaries_expense_account,
        debit=total_gross_salary,
        credit=Decimal('0'),
        description='Total salaries'
    )
    
    # Journal line: Social security payable (credit)
    if total_social_security > 0:
        JournalEntryLine.objects.create(
            journal_entry=journal_entry,
            account=social_security_payable_account,
            debit=Decimal('0'),
            credit=total_social_security,
            description='Social security payable'
        )
    
    # Journal line: Salaries payable (credit)
    JournalEntryLine.objects.create(
        journal_entry=journal_entry,
        account=salaries_payable_account,
        debit=Decimal('0'),
        credit=total_net_salary,
        description='Net salaries payable'
    )
    
    # Link journal entry with payroll entries
    for payroll_entry in payroll_entries:
        payroll_entry.journal_entry = journal_entry
        payroll_entry.save()
    
    # Log activity
    create_hr_audit_log(
        request,
        'CREATE',
        PayrollPeriod,
        journal_entry.id,
        f'Created payroll journal entry for period: {payroll_period.name}'
    )
    
    messages.success(request, _('Payroll journal entries created successfully.'))
    return redirect('hr:payroll_period_list')


# Dashboard
@login_required
@permission_required("hr.view_hr_menu", raise_exception=True)
def hr_dashboard(request):
    """Human Resources Dashboard"""
    
    context = {
        'total_employees': Employee.objects.filter(status='active').count(),
        'total_departments': Department.objects.filter(is_active=True).count(),
        'pending_leaves': LeaveRequest.objects.filter(status='pending').count(),
        'today_attendances': Attendance.objects.filter(date=date.today()).count(),
        
        # New employees this month
        'new_employees': Employee.objects.filter(
            hire_date__year=date.today().year,
            hire_date__month=date.today().month
        ).count(),
        
        # Birthdays this month
        'birthdays_this_month': Employee.objects.filter(
            birth_date__month=date.today().month,
            status='active'
        ).count(),
        
        # Pending leave requests
        'recent_leave_requests': LeaveRequest.objects.filter(
            status='pending'
        ).select_related('employee', 'leave_type')[:5],
        
        # Today's attendance statistics
        'today_attendance_stats': {
            'present': Attendance.objects.filter(date=date.today(), attendance_type='present').count(),
            'absent': Attendance.objects.filter(date=date.today(), attendance_type='absent').count(),
            'late': Attendance.objects.filter(date=date.today(), attendance_type='late').count(),
        }
    }
    
    return render(request, 'hr/dashboard.html', context)


# Department Views
class DepartmentListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = "hr.view_hr_department"
    model = Department
    template_name = 'hr/department_list.html'
    context_object_name = 'departments'
    paginate_by = 20


class DepartmentDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    permission_required = "hr.view_hr_department"
    model = Department
    template_name = 'hr/department_detail.html'
    context_object_name = 'department'


class DepartmentCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = "hr.add_hr_department"
    model = Department
    form_class = DepartmentForm
    template_name = 'hr/department_form.html'
    success_url = reverse_lazy('hr:department_list')


class DepartmentUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = "hr.change_hr_department"
    model = Department
    form_class = DepartmentForm
    template_name = 'hr/department_form.html'
    success_url = reverse_lazy('hr:department_list')


class DepartmentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = "hr.delete_hr_department"
    model = Department
    template_name = 'hr/department_confirm_delete.html'
    success_url = reverse_lazy('hr:department_list')


# Position Views
class PositionListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = "hr.view_hr_position"
    model = Position
    template_name = 'hr/position_list.html'
    context_object_name = 'positions'
    paginate_by = 20


class PositionDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    permission_required = "hr.view_hr_position"
    model = Position
    template_name = 'hr/position_detail.html'
    context_object_name = 'position'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        position = self.get_object()
        
        # Add employee statistics to context
        context['active_employees_count'] = position.employees.filter(status='active').count()
        context['inactive_employees_count'] = position.employees.filter(status='inactive').count()
        
        return context


class PositionCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = "hr.add_hr_position"
    model = Position
    form_class = PositionForm
    template_name = 'hr/position_form.html'
    success_url = reverse_lazy('hr:position_list')


class PositionUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = "hr.change_hr_position"
    model = Position
    form_class = PositionForm
    template_name = 'hr/position_form.html'
    success_url = reverse_lazy('hr:position_list')


class PositionDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = "hr.delete_hr_position"
    model = Position
    template_name = 'hr/position_confirm_delete.html'
    success_url = reverse_lazy('hr:position_list')


# Contract Views
class ContractListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = "hr.view_hr_contract"
    model = Contract
    template_name = 'hr/contract_list.html'
    context_object_name = 'contracts'
    paginate_by = 20


class ContractDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    permission_required = "hr.view_hr_contract"
    model = Contract
    template_name = 'hr/contract_detail.html'
    context_object_name = 'contract'


class ContractCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = "hr.add_hr_contract"
    model = Contract
    form_class = ContractForm
    template_name = 'hr/contract_form.html'
    success_url = reverse_lazy('hr:contract_list')

    def get_initial(self):
        initial = super().get_initial()
        employee_id = self.request.GET.get('employee')
        if employee_id:
            try:
                from .models import Employee
                employee = Employee.objects.get(pk=employee_id)
                initial['employee'] = employee
            except Employee.DoesNotExist:
                pass
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        
        # Log activity
        from core.signals import log_user_activity
        log_user_activity(
            self.request,
            'create',
            form.instance,
            f'Created new contract: {form.instance.contract_number} for employee {form.instance.employee.full_name}'
        )
        
        return response


class ContractUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = "hr.change_hr_contract"
    model = Contract
    form_class = ContractForm
    template_name = 'hr/contract_form.html'
    success_url = reverse_lazy('hr:contract_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Log activity
        from core.signals import log_user_activity
        log_user_activity(
            self.request,
            'update',
            form.instance,
            f'Updated contract: {form.instance.contract_number} for employee {form.instance.employee.full_name}'
        )
        
        return response


class ContractDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = "hr.delete_hr_contract"
    model = Contract
    template_name = 'hr/contract_confirm_delete.html'
    success_url = reverse_lazy('hr:contract_list')

    def delete(self, request, *args, **kwargs):
        contract = self.get_object()
        # Log activity before deletion
        from core.signals import log_user_activity
        log_user_activity(
            request,
            'delete',
            contract,
            f'Deleted contract: {contract.contract_number} for employee {contract.employee.full_name}'
        )
        return super().delete(request, *args, **kwargs)


# Additional Attendance Views
class AttendanceDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    permission_required = "hr.view_hr_attendance"
    model = Attendance
    template_name = 'hr/attendance_detail.html'
    context_object_name = 'attendance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        attendance = self.get_object()
        
        # Calculate current month statistics
        month_attendances = attendance.employee.attendances.filter(
            date__month=attendance.date.month,
            date__year=attendance.date.year
        )
        
        context['month_present'] = month_attendances.filter(attendance_type='present').count()
        context['month_absent'] = month_attendances.filter(attendance_type='absent').count()
        context['month_late'] = month_attendances.filter(attendance_type='late').count()
        context['month_total_hours'] = month_attendances.aggregate(
            total=Sum('worked_hours')
        )['total'] or 0
        
        return context


class AttendanceCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = "hr.add_hr_attendance"
    model = Attendance
    form_class = AttendanceForm
    template_name = 'hr/attendance_form.html'
    success_url = reverse_lazy('hr:attendance_list')

    def get_initial(self):
        initial = super().get_initial()
        employee_id = self.request.GET.get('employee')
        if employee_id:
            try:
                from .models import Employee
                employee = Employee.objects.get(pk=employee_id)
                initial['employee'] = employee
            except Employee.DoesNotExist:
                pass
        return initial
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        
        # Log activity
        create_hr_audit_log(
            self.request,
            'CREATE',
            Attendance,
            self.object.id,
            f'Created attendance record for employee {self.object.employee.full_name} on {self.object.date}'
        )
        
        messages.success(
            self.request, 
            _('Attendance record created successfully')
        )
        return response


class AttendanceUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = "hr.change_hr_attendance"
    model = Attendance
    form_class = AttendanceForm
    template_name = 'hr/attendance_form.html'
    success_url = reverse_lazy('hr:attendance_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Log activity
        create_hr_audit_log(
            self.request,
            'UPDATE',
            Attendance,
            self.object.id,
            f'Updated attendance record for employee {self.object.employee.full_name} on {self.object.date}'
        )
        
        messages.success(
            self.request, 
            _('Attendance record updated successfully')
        )
        return response


class AttendanceDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = "hr.delete_hr_attendance"
    model = Attendance
    template_name = 'hr/attendance_confirm_delete.html'
    success_url = reverse_lazy('hr:attendance_list')
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        employee_name = self.object.employee.full_name
        attendance_date = self.object.date
        
        # Log activity before deletion
        create_hr_audit_log(
            request,
            'DELETE',
            Attendance,
            self.object.id,
            f'Deleted attendance record for employee {employee_name} on {attendance_date}'
        )
        
        response = super().delete(request, *args, **kwargs)
        
        messages.success(
            request, 
            _('Attendance record deleted successfully')
        )
        return response


# Additional Leave Views
class LeaveRequestDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    permission_required = "hr.view_hr_leaverequest"
    model = LeaveRequest
    template_name = 'hr/leave_request_detail.html'
    context_object_name = 'leave_request'


class LeaveRequestCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = "hr.add_hr_leaverequest"
    model = LeaveRequest
    form_class = LeaveRequestForm
    template_name = 'hr/leave_request_form.html'
    success_url = reverse_lazy('hr:leave_request_list')

    def get_initial(self):
        initial = super().get_initial()
        employee_id = self.request.GET.get('employee')
        if employee_id:
            try:
                from .models import Employee
                employee = Employee.objects.get(pk=employee_id)
                initial['employee'] = employee
            except Employee.DoesNotExist:
                pass
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        
        # Log activity
        from core.signals import log_user_activity
        log_user_activity(
            self.request,
            'create',
            form.instance,
            f'Created leave request for employee {form.instance.employee.full_name} of type {form.instance.leave_type.name}'
        )
        
        return response


class LeaveRequestUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = "hr.change_hr_leaverequest"
    model = LeaveRequest
    form_class = LeaveRequestForm
    template_name = 'hr/leave_request_form.html'
    success_url = reverse_lazy('hr:leave_request_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Log activity
        from core.signals import log_user_activity
        log_user_activity(
            self.request,
            'update',
            form.instance,
            f'Updated leave request for employee {form.instance.employee.full_name}'
        )
        
        return response


class LeaveRequestDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = "hr.delete_hr_leaverequest"
    model = LeaveRequest
    template_name = 'hr/leave_request_confirm_delete.html'
    success_url = reverse_lazy('hr:leave_request_list')

    def delete(self, request, *args, **kwargs):
        leave_request = self.get_object()
        # Log activity before deletion
        from core.signals import log_user_activity
        log_user_activity(
            request,
            'delete',
            leave_request,
            f'Deleted leave request for employee {leave_request.employee.full_name}'
        )
        return super().delete(request, *args, **kwargs)


# Leave Type Views
class LeaveTypeListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = "hr.view_hr_leavetype"
    model = LeaveType
    template_name = 'hr/leave_type_list.html'
    context_object_name = 'leave_types'
    paginate_by = 20


class LeaveTypeDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    permission_required = "hr.view_hr_leavetype"
    model = LeaveType
    template_name = 'hr/leave_type_detail.html'
    context_object_name = 'leave_type'


class LeaveTypeCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = "hr.add_hr_leavetype"
    model = LeaveType
    fields = ['name', 'days_per_year', 'max_days', 'is_paid', 'is_active', 'notes']
    template_name = 'hr/leave_type_form.html'
    success_url = reverse_lazy('hr:leave_type_list')


class LeaveTypeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = "hr.change_hr_leavetype"
    model = LeaveType
    fields = ['name', 'days_per_year', 'max_days', 'is_paid', 'is_active', 'notes']
    template_name = 'hr/leave_type_form.html'
    success_url = reverse_lazy('hr:leave_type_list')


class LeaveTypeDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = "hr.delete_hr_leavetype"
    model = LeaveType
    template_name = 'hr/leave_type_confirm_delete.html'
    success_url = reverse_lazy('hr:leave_type_list')


# Additional Payroll Views
class PayrollPeriodDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    permission_required = "hr.view_hr_payrollperiod"
    model = PayrollPeriod
    template_name = 'hr/payroll_period_detail.html'
    context_object_name = 'payroll_period'


class PayrollPeriodCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = "hr.add_hr_payrollperiod"
    model = PayrollPeriod
    form_class = PayrollPeriodForm
    template_name = 'hr/payroll_period_form.html'
    success_url = reverse_lazy('hr:payroll_period_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class PayrollPeriodUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = "hr.change_hr_payrollperiod"
    model = PayrollPeriod
    form_class = PayrollPeriodForm
    template_name = 'hr/payroll_period_form.html'
    success_url = reverse_lazy('hr:payroll_period_list')


class PayrollPeriodDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = "hr.delete_hr_payrollperiod"
    model = PayrollPeriod
    template_name = 'hr/payroll_period_confirm_delete.html'
    success_url = reverse_lazy('hr:payroll_period_list')


# Payroll Entry Views
class PayrollEntryListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = "hr.view_hr_payrollentry"
    model = PayrollEntry
    template_name = 'hr/payroll_entry_list.html'
    context_object_name = 'payroll_entries'
    paginate_by = 20


class PayrollEntryDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    permission_required = "hr.view_hr_payrollentry"
    model = PayrollEntry
    template_name = 'hr/payroll_entry_detail.html'
    context_object_name = 'payroll_entry'


class PayrollEntryCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = "hr.add_hr_payrollentry"
    model = PayrollEntry
    form_class = PayrollEntryForm
    template_name = 'hr/payroll_entry_form.html'
    success_url = reverse_lazy('hr:payroll_entry_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class PayrollEntryUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = "hr.change_hr_payrollentry"
    model = PayrollEntry
    form_class = PayrollEntryForm
    template_name = 'hr/payroll_entry_form.html'
    success_url = reverse_lazy('hr:payroll_entry_list')


class PayrollEntryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = "hr.delete_hr_payrollentry"
    model = PayrollEntry
    template_name = 'hr/payroll_entry_confirm_delete.html'
    success_url = reverse_lazy('hr:payroll_entry_list')


# Employee Document Views
class EmployeeDocumentListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = "hr.view_hr_employeedocument"
    model = EmployeeDocument
    template_name = 'hr/employee_document_list.html'
    context_object_name = 'employee_documents'
    paginate_by = 20


class EmployeeDocumentDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    permission_required = "hr.view_hr_employeedocument"
    model = EmployeeDocument
    template_name = 'hr/employee_document_detail.html'
    context_object_name = 'employee_document'


class EmployeeDocumentCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = "hr.add_hr_employeedocument"
    model = EmployeeDocument
    form_class = EmployeeDocumentForm
    template_name = 'hr/employee_document_form.html'
    success_url = reverse_lazy('hr:employee_document_list')

    def get_initial(self):
        initial = super().get_initial()
        employee_id = self.kwargs.get('employee_id') or self.request.GET.get('employee')
        if employee_id:
            try:
                employee = Employee.objects.get(pk=employee_id)
                initial['employee'] = employee
            except Employee.DoesNotExist:
                pass
        return initial

    def form_valid(self, form):
        # Set the user who uploaded
        form.instance.uploaded_by = self.request.user
        
        response = super().form_valid(form)
        
        # Log activity
        create_hr_audit_log(
            request=self.request,
            action_type="create",
            content_type_model=EmployeeDocument,
            object_id=self.object.id,
            description=_("Created employee document: %(title)s for employee %(employee)s") % {
                'title': self.object.title,
                'employee': self.object.employee.full_name
            }
        )
        
        messages.success(self.request, _('Employee document created successfully.'))
        return response


class EmployeeDocumentUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = "hr.change_hr_employeedocument"
    model = EmployeeDocument
    form_class = EmployeeDocumentForm
    template_name = 'hr/employee_document_form.html'
    success_url = reverse_lazy('hr:employee_document_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Log activity
        create_hr_audit_log(
            request=self.request,
            action_type="update",
            content_type_model=EmployeeDocument,
            object_id=self.object.id,
            description=_("Updated employee document: %(title)s for employee %(employee)s") % {
                'title': self.object.title,
                'employee': self.object.employee.full_name
            }
        )
        
        messages.success(self.request, _('Employee document updated successfully.'))
        return response


class EmployeeDocumentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = "hr.delete_hr_employeedocument"
    model = EmployeeDocument
    template_name = 'hr/employee_document_confirm_delete.html'
    success_url = reverse_lazy('hr:employee_document_list')

    def delete(self, request, *args, **kwargs):
        # Save data before deletion for logging
        obj = self.get_object()
        employee_name = obj.employee.full_name
        document_title = obj.title
        
        response = super().delete(request, *args, **kwargs)
        
        # Log activity
        create_hr_audit_log(
            request=request,
            action_type="delete",
            content_type_model=EmployeeDocument,
            object_id=None,  # Cannot get ID after deletion
            description=_("Deleted employee document: %(title)s for employee %(employee)s") % {
                'title': document_title,
                'employee': employee_name
            }
        )
        
        messages.success(request, _('Employee document deleted successfully.'))
        return response


# Report Views
@login_required
def hr_reports(request):
    """Human Resources Reports"""
    if not (request.user.is_superuser or request.user.has_perm('hr.can_view_hr')):
        messages.error(request, _('You do not have permission to access HR module.'))
        return redirect('core:dashboard')
    
    # Quick statistics for reports
    context = {
        'total_employees': Employee.objects.count(),
        'active_employees': Employee.objects.filter(status='active').count(),
        'total_departments': Department.objects.count(),
        'pending_leaves': LeaveRequest.objects.filter(status='pending').count(),
    }
    
    # Log activity
    create_hr_audit_log(
        request=request,
        action_type="view",
        content_type_model=Employee,
        object_id=None,
        description=_("Viewed HR reports dashboard")
    )
    
    return render(request, 'hr/reports.html', context)


@login_required
def employee_report(request):
    """Employee Report"""
    employees = Employee.objects.select_related('department', 'position').order_by('department__name', 'first_name')
    
    # Statistics
    total_employees = employees.count()
    active_employees = employees.filter(status='active').count()
    inactive_employees = employees.filter(status='inactive').count()
    
    context = {
        'employees': employees,
        'total_employees': total_employees,
        'active_employees': active_employees,
        'inactive_employees': inactive_employees,
        'title': _('Employee Report')
    }
    
    # Log activity
    create_hr_audit_log(
        request=request,
        action_type="view",
        content_type_model=Employee,
        object_id=None,
        description=_("Viewed employee report")
    )
    
    return render(request, 'hr/reports/employee_report.html', context)


@login_required
def attendance_report(request):
    """Attendance Report - Fixed"""
    from datetime import datetime, timedelta
    
    # Filter by date
    selected_month = int(request.GET.get('month', datetime.now().month))
    selected_year = int(request.GET.get('year', datetime.now().year))
    selected_department = request.GET.get('department', '')
    selected_employee = request.GET.get('employee', '')
    
    # Build query
    attendances = Attendance.objects.select_related('employee', 'employee__department')
    
    if selected_month != 0:
        attendances = attendances.filter(date__month=selected_month)
    
    attendances = attendances.filter(date__year=selected_year).order_by('-date', 'employee__first_name')
    
    if selected_department:
        attendances = attendances.filter(employee__department_id=selected_department)
    
    if selected_employee:
        attendances = attendances.filter(employee_id=selected_employee)
    
    # Statistics
    total_records = attendances.count()
    present_count = attendances.filter(attendance_type='present').count()
    absent_count = attendances.filter(attendance_type='absent').count()
    late_count = attendances.filter(attendance_type='late').count()
    
    # Additional data for filters - month names
    months = [
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ]
    years = list(range(2020, datetime.now().year + 2))
    departments = Department.objects.all()
    employees = Employee.objects.filter(status='active').order_by('first_name')
    
    context = {
        'attendances': attendances,  # Display all records
        'total_records': total_records,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'months': months,
        'years': years,
        'departments': departments,
        'employees': employees,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'selected_department': selected_department,
        'selected_employee': selected_employee,
    }
    
    # Log activity
    create_hr_audit_log(
        request,
        "view",
        Attendance,
        None,
        _("Viewed attendance report")
    )
    
    return render(request, 'hr/reports/attendance_report.html', context)


@login_required
def payroll_report(request, pk):
    """Payroll Sheet Report"""
    payroll_period = get_object_or_404(PayrollPeriod, pk=pk)
    payroll_entries = payroll_period.payroll_entries.select_related('employee')
    
    context = {
        'payroll_period': payroll_period,
        'payroll_entries': payroll_entries,
        'total_gross': sum(entry.gross_salary for entry in payroll_entries),
        'total_deductions': sum(entry.total_deductions for entry in payroll_entries),
        'total_net': sum(entry.net_salary for entry in payroll_entries),
    }
    
    return render(request, 'hr/payroll_report.html', context)


@login_required
def payroll_summary_report(request):
    """Payroll Summary Report"""
    from django.db.models import Sum, Count, Avg
    from datetime import datetime
    
    # Filters
    selected_month = int(request.GET.get('month', datetime.now().month))
    selected_year = int(request.GET.get('year', datetime.now().year))
    selected_department = request.GET.get('department', '')
    
    # Build employee query
    employees = Employee.objects.filter(status='active').select_related('department')
    
    if selected_department:
        employees = employees.filter(department_id=selected_department)
    
    # Create dummy payroll_entries from employee data
    payroll_entries = []
    total_gross = 0
    total_deductions = 0
    total_net = 0
    
    for emp in employees:
        basic_salary = float(emp.basic_salary)
        allowances = float(emp.allowances)
        overtime_amount = 0  # Can be calculated from Attendance
        gross_salary = basic_salary + allowances + overtime_amount
        
        # Calculate deductions using the same method as Employee model
        withholding_tax = float(emp.withholding_tax_amount)
        social_security = float(emp.social_security_amount)
        additional_deductions = float(emp.additional_deduction_amount)
        total_deductions_for_emp = withholding_tax + social_security + additional_deductions
        
        net_salary = gross_salary - total_deductions_for_emp
        
        # Create dummy object
        class PayrollEntry:
            def __init__(self, employee, basic_salary, allowances, overtime_amount, gross_salary, total_deductions, net_salary):
                self.employee = employee
                self.basic_salary = basic_salary
                self.allowances = allowances
                self.overtime_amount = overtime_amount
                self.gross_salary = gross_salary
                self.total_deductions = total_deductions
                self.net_salary = net_salary
        
        entry = PayrollEntry(emp, basic_salary, allowances, overtime_amount, gross_salary, total_deductions_for_emp, net_salary)
        payroll_entries.append(entry)
        
        total_gross += gross_salary
        total_deductions += total_deductions_for_emp
        total_net += net_salary
    
    # Data for filters
    months = [
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ]
    years = list(range(datetime.now().year - 5, datetime.now().year + 1))
    departments = Department.objects.filter(is_active=True)
    
    context = {
        'payroll_entries': payroll_entries,
        'total_employees': len(payroll_entries),
        'total_gross': total_gross,
        'total_deductions': total_deductions,
        'total_net': total_net,
        'months': months,
        'years': years,
        'departments': departments,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'selected_department': selected_department,
        'title': _('Payroll Summary Report')
    }
    
    # Log activity
    create_hr_audit_log(
        request,
        "view",
        Employee,
        None,
        _("Viewed payroll summary report")
    )
    
    return render(request, 'hr/reports/payroll_summary_report.html', context)


@login_required
def leave_balance_report(request):
    """Leave Balance Report"""
    from django.db.models import Sum, Q
    
    employees = Employee.objects.filter(status='active').select_related('department')
    
    # Get leave types
    leave_types = LeaveType.objects.filter(is_active=True)
    annual_leave_type = leave_types.filter(code='ANNUAL').first()
    sick_leave_type = leave_types.filter(code='SICK').first()
    emergency_leave_type = leave_types.filter(code='EMERGENCY').first()
    
    employee_balances = []
    for employee in employees:
        # Calculate annual leaves
        annual_used = LeaveRequest.objects.filter(
            employee=employee,
            leave_type=annual_leave_type,
            status='approved'
        ).aggregate(total=Sum('days_count'))['total'] or 0
        annual_total = annual_leave_type.days_per_year if annual_leave_type else 21
        
        # Calculate sick leaves
        sick_used = LeaveRequest.objects.filter(
            employee=employee,
            leave_type=sick_leave_type,
            status='approved'
        ).aggregate(total=Sum('days_count'))['total'] or 0
        sick_total = sick_leave_type.days_per_year if sick_leave_type else 14
        
        # Calculate emergency leaves
        emergency_used = LeaveRequest.objects.filter(
            employee=employee,
            leave_type=emergency_leave_type,
            status='approved'
        ).aggregate(total=Sum('days_count'))['total'] or 0
        emergency_total = emergency_leave_type.days_per_year if emergency_leave_type else 7
        
        # Totals
        total_used = annual_used + sick_used + emergency_used
        total_available = annual_total + sick_total + emergency_total
        total_remaining = max(0, total_available - total_used)
        
        employee_balances.append({
            'employee': employee,
            'annual_leave_used': int(annual_used),
            'annual_leave_total': annual_total,
            'annual_leave_remaining': max(0, annual_total - annual_used),
            'annual_leave_used_percentage': min(100, (annual_used / annual_total * 100) if annual_total > 0 else 0),
            
            'sick_leave_used': int(sick_used),
            'sick_leave_total': sick_total,
            'sick_leave_remaining': max(0, sick_total - sick_used),
            'sick_leave_used_percentage': min(100, (sick_used / sick_total * 100) if sick_total > 0 else 0),
            
            'emergency_leave_used': int(emergency_used),
            'emergency_leave_total': emergency_total,
            'emergency_leave_remaining': max(0, emergency_total - emergency_used),
            'emergency_leave_used_percentage': min(100, (emergency_used / emergency_total * 100) if emergency_total > 0 else 0),
            
            'total_leave_used': int(total_used),
            'total_leave_remaining': int(total_remaining),
        })
    
    context = {
        'employee_balances': employee_balances,
        'title': _('Leave Balance Report')
    }
    
    # Log activity
    create_hr_audit_log(
        request,
        "view",
        Employee,
        None,
        _("Viewed leave balance report")
    )
    
    return render(request, 'hr/reports/leave_balance_report.html', context)


@login_required
def expiring_documents(request):
    """Expiring Documents"""
    from datetime import timedelta
    
    warning_date = date.today() + timedelta(days=30)
    
    expiring_docs = EmployeeDocument.objects.filter(
        expiry_date__lte=warning_date,
        expiry_date__gte=date.today()
    ).select_related('employee')
    
    return render(request, 'hr/expiring_documents.html', {'expiring_docs': expiring_docs})


# API Views
@login_required
def get_positions_by_department(request, department_id):
    """Get positions by department (AJAX)"""
    positions = Position.objects.filter(department_id=department_id, is_active=True)
    data = [{'id': pos.id, 'title': pos.title} for pos in positions]
    return JsonResponse(data, safe=False)


@login_required
def get_employee_info(request, employee_id):
    """Get employee info (AJAX)"""
    try:
        employee = Employee.objects.get(id=employee_id)
        data = {
            'full_name': employee.full_name,
            'department': employee.department.name,
            'position': employee.position.title,
            'basic_salary': float(employee.basic_salary),
            'allowances': float(employee.allowances),
        }
        return JsonResponse(data)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Employee not found'}, status=404)


@login_required
def get_employee_leave_balance(request, employee_id):
    """Get employee leave balance (AJAX)"""
    try:
        employee = Employee.objects.get(id=employee_id)
        balance = employee.get_current_leave_balance()
        return JsonResponse(balance)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Employee not found'}, status=404)


# Report Views
@login_required
def department_report(request):
    """Employees by Department Report"""
    if not request.user.is_superuser and not request.user.has_perm('hr.can_view_hr'):
        messages.error(request, _('You do not have permission to access HR module.'))
        return redirect('core:dashboard')
    
    from django.db.models import Count, Avg, Sum
    
    departments = Department.objects.filter(is_active=True).prefetch_related('employees').annotate(
        employee_count=Count('employees', filter=Q(employees__status='active')),
        avg_salary=Avg('employees__basic_salary', filter=Q(employees__status='active')),
        total_salary=Sum('employees__basic_salary', filter=Q(employees__status='active'))
    ).order_by('name')
    
    total_employees = Employee.objects.filter(status='active').count()
    
    context = {
        'departments': departments,
        'total_employees': total_employees,
        'title': _('Department Report')
    }
    
    # Log activity
    create_hr_audit_log(
        request,
        "view",
        Employee,
        None,
        _("Viewed department report")
    )
    
    return render(request, 'hr/reports/department_report.html', context)


@login_required
def new_hires_report(request):
    """New Hires Report"""
    from datetime import datetime, timedelta
    
    # Filter by date
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if not from_date:
        from_date = (datetime.now().date() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not to_date:
        to_date = datetime.now().date().strftime('%Y-%m-%d')
    
    new_hires = Employee.objects.filter(
        hire_date__range=[from_date, to_date]
    ).select_related('department', 'position').order_by('-hire_date', 'first_name')
    
    total_new_hires = new_hires.count()
    
    context = {
        'new_hires': new_hires,
        'total_new_hires': total_new_hires,
        'from_date': from_date,
        'to_date': to_date,
        'title': _('New Hires Report')
    }
    
    # Log activity
    create_hr_audit_log(
        request,
        "view",
        Employee,
        None,
        _("Viewed new hires report")
    )
    
    return render(request, 'hr/reports/new_hires_report.html', context)


@login_required
def late_arrivals_report(request):
    """Late Arrivals Report"""
    from datetime import datetime
    from django.db.models import Count
    
    # Filter by date
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if not from_date:
        from_date = datetime.now().replace(day=1).strftime('%Y-%m-%d')
    if not to_date:
        to_date = datetime.now().strftime('%Y-%m-%d')
    
    late_arrivals = Attendance.objects.filter(
        attendance_type='late',
        date__range=[from_date, to_date]
    ).select_related('employee', 'employee__department').order_by('-date', 'employee__first_name')
    
    # Calculate statistics
    total_late = late_arrivals.count()
    by_employee = late_arrivals.values('employee__first_name', 'employee__last_name').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    context = {
        'late_arrivals': late_arrivals,
        'total_late': total_late,
        'by_employee': by_employee,
        'from_date': from_date,
        'to_date': to_date,
    }
    
    # Log activity
    create_hr_audit_log(
        request,
        "view",
        Attendance,
        None,
        _("Viewed late arrivals report")
    )
    
    return render(request, 'hr/reports/late_arrivals_report.html', context)


@login_required
def overtime_report(request):
    """Overtime Report"""
    from datetime import datetime
    from django.db.models import Sum
    
    # Filter by date
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if not from_date:
        from_date = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
    if not to_date:
        to_date = datetime.now().strftime('%Y-%m-%d')
    
    # Filter by employee
    employee_id = request.GET.get('employee')
    
    # Get overtime records
    overtime_attendances = Attendance.objects.filter(
        date__range=[from_date, to_date],
        overtime_hours__gt=0
    ).select_related('employee', 'employee__department').order_by('-date', 'employee__first_name')
    
    if employee_id:
        overtime_attendances = overtime_attendances.filter(employee_id=employee_id)
    
    # Calculate totals
    total_overtime = overtime_attendances.aggregate(total=Sum('overtime_hours'))['total'] or 0
    total_records = overtime_attendances.count()
    
    context = {
        'overtime_records': overtime_attendances,
        'total_overtime': float(total_overtime),
        'total_records': total_records,
        'from_date': from_date,
        'to_date': to_date,
        'employees': Employee.objects.filter(status='active').order_by('first_name'),
        'selected_employee': employee_id,
    }
    
    # Log activity
    create_hr_audit_log(
        request,
        "view", 
        Attendance,
        None,
        _("Viewed overtime report")
    )
    
    return render(request, 'hr/reports/overtime_report.html', context)


@login_required
def leave_summary_report(request):
    """Leave Summary Report"""
    from datetime import datetime, timedelta
    from django.db.models import Count, Sum
    
    # Get filter parameters
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    employee_id = request.GET.get('employee')
    
    # Set default dates if not provided
    if not from_date:
        from_date = (datetime.now().date() - timedelta(days=90)).strftime('%Y-%m-%d')
    if not to_date:
        to_date = datetime.now().date().strftime('%Y-%m-%d')
    
    # Base queryset
    leave_requests = LeaveRequest.objects.filter(
        start_date__range=[from_date, to_date]
    )
    
    # Filter by employee if selected
    if employee_id:
        leave_requests = leave_requests.filter(employee_id=employee_id)
    
    # Statistics
    total_requests = leave_requests.count()
    approved_requests = leave_requests.filter(status='approved').count()
    pending_requests = leave_requests.filter(status='pending').count()
    rejected_requests = leave_requests.filter(status='rejected').count()
    
    # Get all employees for filter dropdown
    employees = Employee.objects.filter(status='active').order_by('first_name', 'last_name')
    
    context = {
        'from_date': from_date,
        'to_date': to_date,
        'employee_id': employee_id,
        'leave_requests': leave_requests.order_by('-start_date'),
        'total_requests': total_requests,
        'approved_requests': approved_requests,
        'pending_requests': pending_requests,
        'rejected_requests': rejected_requests,
        'employees': employees,
        'title': _('Leave Summary Report')
    }
    
    # Log activity
    create_hr_audit_log(
        request,
        "view",
        Employee,
        None,
        _("Viewed leave summary report")
    )
    
    return render(request, 'hr/reports/leave_summary_report.html', context)


@login_required
def upcoming_leaves_report(request):
    """Upcoming Leaves Report"""
    from datetime import datetime, timedelta
    
    # Get filter parameters
    employee_id = request.GET.get('employee')
    
    # Get upcoming leaves in the next month
    today = datetime.now().date()
    next_month = today + timedelta(days=30)
    
    upcoming_leaves = LeaveRequest.objects.filter(
        start_date__range=[today, next_month],
        status__in=['approved', 'pending']
    ).select_related('employee', 'employee__department', 'leave_type')
    
    # Filter by employee if selected
    if employee_id:
        upcoming_leaves = upcoming_leaves.filter(employee_id=employee_id)
    
    upcoming_leaves = upcoming_leaves.order_by('-start_date')
    
    # Get all employees for filter dropdown
    employees = Employee.objects.filter(status='active').order_by('first_name', 'last_name')
    
    context = {
        'upcoming_leaves': upcoming_leaves,
        'today': today,
        'next_month': next_month,
        'total_leaves': upcoming_leaves.count(),
        'employees': employees,
        'employee_id': employee_id,
        'title': _('Upcoming and Pending Leaves Report')
    }
    
    # Log activity
    create_hr_audit_log(
        request,
        "view",
        LeaveRequest,
        None,
        _("Viewed upcoming and pending leaves report")
    )
    
    return render(request, 'hr/reports/upcoming_leaves_report.html', context)


@login_required
def salary_breakdown_report(request):
    """Salary Breakdown Report"""
    from django.db.models import Sum, Avg, Count
    
    # Filter by department
    department_id = request.GET.get('department')
    
    employees = Employee.objects.filter(status='active').select_related('department')
    if department_id:
        employees = employees.filter(department_id=department_id)
    
    # Calculate statistics by department
    salary_breakdown = []
    departments = Department.objects.filter(is_active=True)
    
    for dept in departments:
        dept_employees = employees.filter(department=dept)
        employee_count = dept_employees.count()
        
        if employee_count > 0:
            total_basic = sum(float(e.basic_salary) for e in dept_employees)
            total_allowances = sum(float(e.allowances) for e in dept_employees)
            total_salary = total_basic + total_allowances
            avg_salary = total_salary / employee_count
            
            salary_breakdown.append({
                'department_name': dept.name,
                'employee_count': employee_count,
                'total_basic_salary': total_basic,
                'total_allowances': total_allowances,
                'total_salary': total_salary,
                'average_salary': avg_salary
            })
    
    # Order by total salary
    salary_breakdown.sort(key=lambda x: x['total_salary'], reverse=True)
    
    context = {
        'salary_breakdown': salary_breakdown,
        'departments': departments,
        'selected_department': department_id,
        'title': _('Salary Breakdown Report')
    }
    
    # Log activity
    create_hr_audit_log(
        request,
        "view",
        None,
        None,
        _("Viewed salary breakdown report")
    )
    
    return render(request, 'hr/reports/salary_breakdown_report.html', context)


@login_required
def payroll_comparison_report(request):
    """Payroll Comparison Report"""
    from datetime import datetime
    from django.db.models import Sum, Count, Q
    from calendar import month_name
    import json
    
    # Filter by year
    year = int(request.GET.get('year', datetime.now().year))
    
    # Monthly statistics
    monthly_data = []
    for month in range(1, 13):
        # Get payroll periods for this month
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year}-{month:02d}-31"
        else:
            from calendar import monthrange
            last_day = monthrange(year, month)[1]
            end_date = f"{year}-{month:02d}-{last_day:02d}"
        
        # Get processed payroll periods in this month
        periods = PayrollPeriod.objects.filter(
            start_date__gte=start_date,
            start_date__lte=end_date,
            is_processed=True
        )
        
        # Calculate from payroll entries
        payroll_entries = PayrollEntry.objects.filter(
            payroll_period__in=periods
        )
        
        total_employees = payroll_entries.values('employee').distinct().count()
        total_salary = payroll_entries.aggregate(total=Sum('basic_salary'))['total'] or 0
        total_allowances = payroll_entries.aggregate(total=Sum('allowances'))['total'] or 0
        total_gross = payroll_entries.aggregate(total=Sum('gross_salary'))['total'] or 0
        
        # If no payroll entries, use employee data
        if total_employees == 0:
            employees_in_month = Employee.objects.filter(
                hire_date__lte=end_date,
                status='active'
            ).exclude(
                Q(termination_date__lt=start_date) & ~Q(termination_date=None)
            )
            
            total_employees = employees_in_month.count()
            total_salary = employees_in_month.aggregate(total=Sum('basic_salary'))['total'] or 0
            total_allowances = employees_in_month.aggregate(total=Sum('allowances'))['total'] or 0
            total_gross = float(total_salary) + float(total_allowances)
        
        monthly_data.append({
            'month': month,
            'month_name': str(_(month_name[month])),
            'total_employees': total_employees,
            'total_salary': float(total_salary),
            'total_allowances': float(total_allowances),
            'total_payroll': float(total_gross),
        })
    
    years = list(range(datetime.now().year - 5, datetime.now().year + 1))
    
    context = {
        # Provide list for template iteration and JSON for JavaScript
        'monthly_data': monthly_data,
        'monthly_data_json': json.dumps(monthly_data),
        'year': year,
        'years': years,
        'title': _('Payroll Comparison Report')
    }
    
    # Log activity
    create_hr_audit_log(
        request,
        "view",
        PayrollPeriod,
        None,
        _("Viewed payroll comparison report for year %(year)s") % {'year': year}
    )
    
    return render(request, 'hr/reports/payroll_comparison_report.html', context)


@login_required
def contract_expiry_report(request):
    """Contract Expiry Report"""
    if not request.user.is_superuser and not request.user.has_perm('hr.can_view_hr'):
        messages.error(request, _('You do not have permission to access HR module.'))
        return redirect('core:dashboard')
    
    from datetime import datetime, timedelta
    today = datetime.now().date()
    next_three_months = today + timedelta(days=90)
    
    contracts = Contract.objects.filter(
        end_date__lte=next_three_months,
        end_date__gte=today,
        status='active'
    ).select_related('employee', 'employee__department').order_by('end_date', 'employee__first_name')
    
    # Add days_remaining and is_active for each contract
    expiring_contracts = []
    for contract in contracts:
        # Calculate remaining days
        days_remaining = (contract.end_date - today).days
        
        # Add property to contract
        contract.days_remaining = days_remaining
        contract.is_active = contract.status == 'active'
        
        expiring_contracts.append(contract)
    
    total_expiring = len(expiring_contracts)
    
    context = {
        'expiring_contracts': expiring_contracts,
        'total_expiring': total_expiring,
        'title': _('Contract Expiry Report')
    }
    
    # Log activity
    create_hr_audit_log(
        request,
        'view',
        Contract,
        None,
        _('Viewed contract expiry report')
    )
    
    return render(request, 'hr/reports/contract_expiry_report.html', context)


@login_required
def contract_types_report(request):
    """Contract Types Report"""
    from django.db.models import Count
    
    # Contract types statistics
    contract_types_data = []
    total_contracts = Contract.objects.filter(status='active').count()
    
    for contract_type, label in Contract.CONTRACT_TYPES:
        count = Contract.objects.filter(contract_type=contract_type, status='active').count()
        percentage = (count / total_contracts * 100) if total_contracts > 0 else 0
        
        contract_types_data.append({
            'type': label,
            'type_code': contract_type,
            'count': count,
            'percentage': percentage
        })
    
    # Order by count
    contract_types_data.sort(key=lambda x: x['count'], reverse=True)
    
    context = {
        'contract_types_data': contract_types_data,
        'total_contracts': total_contracts,
        'title': _('Contract Types Report')
    }
    
    # Log activity
    create_hr_audit_log(
        request,
        "view",
        Contract,
        None,
        _("Viewed contract types report")
    )
    
    return render(request, 'hr/reports/contract_types_report.html', context)


@login_required
def probation_report(request):
    """Probation Period Report"""
    if not request.user.is_superuser and not request.user.has_perm('hr.can_view_hr'):
        messages.error(request, _('You do not have permission to access HR module.'))
        return redirect('core:dashboard')
    
    from datetime import datetime, timedelta
    
    from django.utils import timezone
    today = timezone.now().date()
    # Only include probation contracts that have not yet ended
    probation_contracts = Contract.objects.filter(
        contract_type='probation',
        status='active',
        end_date__gte=today
    ).select_related('employee', 'employee__department').order_by('-end_date', 'employee__first_name')
    
    # Add days_remaining for each contract
    # today is already set above
    probation_list = []
    
    for contract in probation_contracts:
        if contract.end_date:
            days_remaining = (contract.end_date - today).days
            contract.days_remaining = days_remaining
            probation_list.append(contract)
    
    total_probation = len(probation_list)
    
    context = {
        'probation_contracts': probation_list,
        'total_probation': total_probation,
        'title': _('Probation Period Report')
    }
    
    # Log activity
    create_hr_audit_log(
        request,
        'view',
        Contract,
        None,
        _('Viewed probation report')
    )
    
    return render(request, 'hr/reports/probation_report.html', context)


@login_required
def headcount_report(request):
    """Total Employee Count Report"""
    from django.db.models import Count, Q
    
    # Employee statistics
    total_employees = Employee.objects.count()
    total_active = Employee.objects.filter(status='active').count()
    total_inactive = Employee.objects.exclude(status='active').count()
    
    # Count by department
    departments = Department.objects.filter(is_active=True).annotate(
        active_count=Count('employees', filter=Q(employees__status='active')),
        inactive_count=Count('employees', filter=~Q(employees__status='active'))
    )
    
    # Create headcount_data for template
    headcount_data = []
    for dept in departments:
        total_count = dept.active_count + dept.inactive_count
        percentage = (total_count / total_employees * 100) if total_employees > 0 else 0
        
        headcount_data.append({
            'department_name': dept.name,
            'active_count': dept.active_count,
            'inactive_count': dept.inactive_count,
            'total_count': total_count,
            'percentage': percentage
        })
    
    # Order by total count
    headcount_data.sort(key=lambda x: x['total_count'], reverse=True)
    
    context = {
        'total_employees': total_active,
        'headcount_data': headcount_data,
        'title': _('Headcount Report')
    }
    
    # Log activity
    create_hr_audit_log(
        request,
        "view",
        Employee,
        None,
        _("Viewed headcount report")
    )
    
    return render(request, 'hr/reports/headcount_report.html', context)


@login_required
def turnover_report(request):
    """Employee Turnover Report"""
    if not request.user.is_superuser and not request.user.has_perm('hr.can_view_hr'):
        messages.error(request, _('You do not have permission to access HR module.'))
        return redirect('core:dashboard')
    
    from datetime import datetime
    current_year = datetime.now().year
    
    # Use new data - terminated employees
    terminated_employees = Employee.objects.filter(
        is_terminated=True,
        termination_date__year=current_year
    ).select_related('department', 'position').order_by('-termination_date', 'first_name')
    
    total_terminated = terminated_employees.count()
    total_active = Employee.objects.filter(status='active', is_terminated=False).count()
    turnover_rate = (total_terminated / (total_active + total_terminated) * 100) if (total_active + total_terminated) > 0 else 0
    
    context = {
        'terminated_employees': terminated_employees,
        'total_terminated': total_terminated,
        'total_active': total_active,
        'turnover_rate': round(turnover_rate, 2),
        'current_year': current_year,
        'title': _('Employee Turnover Report')
    }
    
    # Log activity
    create_hr_audit_log(
        request,
        'view',
        Employee,
        None,
        _('Viewed turnover report')
    )
    
    return render(request, 'hr/reports/turnover_report.html', context)


@login_required
def anniversary_report(request):
    """Work Anniversary Report"""
    if not request.user.is_superuser and not request.user.has_perm('hr.can_view_hr'):
        messages.error(request, _('You do not have permission to access HR module.'))
        return redirect('core:dashboard')
    
    from datetime import datetime
    from dateutil.relativedelta import relativedelta
    
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    anniversary_employees = Employee.objects.filter(
        hire_date__month=current_month,
        status='active'
    ).select_related('department', 'position').order_by('hire_date', 'first_name')
    
    # Add years_of_service and anniversary_date
    anniversaries = []
    for emp in anniversary_employees:
        # Calculate years of service
        years_of_service = current_year - emp.hire_date.year
        
        # Anniversary date for this year
        anniversary_date = emp.hire_date.replace(year=current_year)
        
        emp.years_of_service = years_of_service
        emp.anniversary_date = anniversary_date
        anniversaries.append(emp)
    
    context = {
        'anniversaries': anniversaries,
        'current_month': datetime.now().strftime('%B'),
        'total_anniversaries': len(anniversaries),
        'title': _('Work Anniversary Report')
    }
    
    # Log activity
    create_hr_audit_log(
        request,
        'view',
        Employee,
        None,
        _('Viewed anniversary report')
    )
    
    return render(request, 'hr/reports/anniversary_report.html', context)


# Export Functions
@login_required
def export_employees_excel(request):
    """Export employee data to Excel"""
    if not request.user.is_superuser and not request.user.has_perm('hr.can_view_hr'):
        messages.error(request, _('You do not have permission to access HR module.'))
        return redirect('core:dashboard')
    
    import openpyxl
    from django.http import HttpResponse
    from datetime import datetime
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Employees'
    
    # Headers
    headers = [
        'Employee ID', 'First Name', 'Last Name', 'Email', 'Phone',
        'Department', 'Position', 'Hire Date', 'Status', 'Basic Salary'
    ]
    
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)
    
    # Data
    employees = Employee.objects.select_related('department', 'position').all()
    for row_num, employee in enumerate(employees, 2):
        worksheet.cell(row=row_num, column=1, value=employee.employee_id)
        worksheet.cell(row=row_num, column=2, value=employee.first_name)
        worksheet.cell(row=row_num, column=3, value=employee.last_name)
        worksheet.cell(row=row_num, column=4, value=employee.email)
        worksheet.cell(row=row_num, column=5, value=employee.phone)
        worksheet.cell(row=row_num, column=6, value=str(employee.department))
        worksheet.cell(row=row_num, column=7, value=str(employee.position))
        worksheet.cell(row=row_num, column=8, value=employee.hire_date)
        worksheet.cell(row=row_num, column=9, value=employee.get_status_display())
        worksheet.cell(row=row_num, column=10, value=float(employee.basic_salary))
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=employees_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    workbook.save(response)
    
    # Log activity
    create_hr_audit_log(
        request,
        'export',
        Employee,
        None,
        _('Exported employees data to Excel')
    )
    
    return response


@login_required
def export_attendance_excel(request):
    """Export attendance data to Excel"""
    if not request.user.is_superuser and not request.user.has_perm('hr.can_view_hr'):
        messages.error(request, _('You do not have permission to access HR module.'))
        return redirect('core:dashboard')
    
    import openpyxl
    from django.http import HttpResponse
    from datetime import datetime, timedelta
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Attendance'
    
    # Headers
    headers = [
        'Employee ID', 'Employee Name', 'Date', 'Check In', 'Check Out',
        'Status', 'Worked Hours', 'Overtime Hours'
    ]
    
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)
    
    # Data (last 30 days)
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=30)
    
    attendance_records = Attendance.objects.filter(
        date__range=[start_date, end_date]
    ).select_related('employee')
    
    for row_num, attendance in enumerate(attendance_records, 2):
        worksheet.cell(row=row_num, column=1, value=attendance.employee.employee_id)
        worksheet.cell(row=row_num, column=2, value=f"{attendance.employee.first_name} {attendance.employee.last_name}")
        worksheet.cell(row=row_num, column=3, value=attendance.date)
        worksheet.cell(row=row_num, column=4, value=attendance.check_in_time)
        worksheet.cell(row=row_num, column=5, value=attendance.check_out_time)
        worksheet.cell(row=row_num, column=6, value=attendance.get_attendance_type_display())
        worksheet.cell(row=row_num, column=7, value=float(attendance.worked_hours or 0))
        worksheet.cell(row=row_num, column=8, value=float(attendance.overtime_hours or 0))
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=attendance_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    workbook.save(response)
    
    # Log activity
    create_hr_audit_log(
        request,
        'export',
        Attendance,
        None,
        _('Exported attendance data to Excel')
    )
    
    return response


@login_required
def export_payroll_excel(request):
    """Export payroll data to Excel"""
    if not request.user.is_superuser and not request.user.has_perm('hr.can_view_hr'):
        messages.error(request, _('You do not have permission to access HR module.'))
        return redirect('core:dashboard')
    
    import openpyxl
    from django.http import HttpResponse
    from datetime import datetime
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Payroll'
    
    # Headers
    headers = [
        'Employee ID', 'Employee Name', 'Period', 'Basic Salary', 'Allowances',
        'Overtime', 'Bonuses', 'Deductions', 'Social Security', 'Income Tax',
        'Gross Salary', 'Net Salary'
    ]
    
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)
    
    # Data (last 3 months)
    payroll_entries = PayrollEntry.objects.select_related(
        'employee', 'payroll_period'
    ).order_by('-payroll_period__start_date')[:100]  # Last 100 entries
    
    for row_num, entry in enumerate(payroll_entries, 2):
        worksheet.cell(row=row_num, column=1, value=entry.employee.employee_id)
        worksheet.cell(row=row_num, column=2, value=f"{entry.employee.first_name} {entry.employee.last_name}")
        worksheet.cell(row=row_num, column=3, value=entry.payroll_period.period_name)
        worksheet.cell(row=row_num, column=4, value=float(entry.basic_salary))
        worksheet.cell(row=row_num, column=5, value=float(entry.allowances))
        worksheet.cell(row=row_num, column=6, value=float(entry.overtime_amount))
        worksheet.cell(row=row_num, column=7, value=float(entry.bonuses))
        worksheet.cell(row=row_num, column=8, value=float(entry.deductions))
        worksheet.cell(row=row_num, column=9, value=float(entry.social_security_deduction))
        worksheet.cell(row=row_num, column=10, value=float(entry.income_tax))
        worksheet.cell(row=row_num, column=11, value=float(entry.gross_salary))
        worksheet.cell(row=row_num, column=12, value=float(entry.net_salary))
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=payroll_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    workbook.save(response)
    
    # Log activity
    create_hr_audit_log(
        request,
        'export',
        PayrollEntry,
        None,
        _('Exported payroll data to Excel')
    )
    
    return response


# Employee Deduction Views
@login_required
def employee_deduction_add(request, employee_id):
    """Add new deduction for employee"""
    employee = get_object_or_404(Employee, pk=employee_id)
    
    if request.method == 'POST':
        form = EmployeeDeductionForm(request.POST)
        if form.is_valid():
            deduction = form.save(commit=False)
            deduction.employee = employee
            deduction.created_by = request.user
            deduction.save()
            
            # Log activity
            create_hr_audit_log(
                request,
                'create',
                EmployeeDeduction,
                deduction.id,
                _(f'Added deduction "{deduction.name}" for employee {employee.full_name}')
            )
            
            return JsonResponse({
                'success': True,
                'message': _('Deduction added successfully'),
                'deduction': {
                    'id': deduction.id,
                    'name': deduction.name,
                    'type': deduction.get_deduction_type_display(),
                    'value': str(deduction.value),
                    'calculated_amount': str(deduction.calculated_amount),
                    'notes': deduction.notes,
                    'is_active': deduction.is_active,
                }
            })
        else:
            return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    
    return JsonResponse({'success': False, 'message': _('Invalid request')}, status=400)


@login_required
def employee_deduction_edit(request, deduction_id):
    """Edit employee deduction"""
    deduction = get_object_or_404(EmployeeDeduction, pk=deduction_id)
    
    if request.method == 'POST':
        form = EmployeeDeductionForm(request.POST, instance=deduction)
        if form.is_valid():
            deduction = form.save()
            
            # Log activity
            create_hr_audit_log(
                request,
                'update',
                EmployeeDeduction,
                deduction.id,
                _(f'Updated deduction "{deduction.name}" for employee {deduction.employee.full_name}')
            )
            
            return JsonResponse({
                'success': True,
                'message': _('Deduction updated successfully'),
                'deduction': {
                    'id': deduction.id,
                    'name': deduction.name,
                    'type': deduction.get_deduction_type_display(),
                    'value': str(deduction.value),
                    'calculated_amount': str(deduction.calculated_amount),
                    'notes': deduction.notes,
                    'is_active': deduction.is_active,
                }
            })
        else:
            return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    
    return JsonResponse({'success': False, 'message': _('Invalid request')}, status=400)


@login_required
def employee_deduction_delete(request, deduction_id):
    """Delete employee deduction"""
    if request.method == 'POST':
        deduction = get_object_or_404(EmployeeDeduction, pk=deduction_id)
        employee_name = deduction.employee.full_name
        deduction_name = deduction.name
        
        deduction.delete()
        
        # Log activity
        create_hr_audit_log(
            request,
            'delete',
            EmployeeDeduction,
            deduction_id,
            _(f'Deleted deduction "{deduction_name}" for employee {employee_name}')
        )
        
        return JsonResponse({'success': True, 'message': _('Deduction deleted successfully')})
    
    return JsonResponse({'success': False, 'message': _('Invalid request')}, status=400)


@login_required
def employee_deduction_list(request, employee_id):
    """Employee deductions list"""
    employee = get_object_or_404(Employee, pk=employee_id)
    deductions = employee.deductions.all().order_by('-created_at')
    
    deductions_data = []
    for deduction in deductions:
        deductions_data.append({
            'id': deduction.id,
            'name': deduction.name,
            'type': deduction.get_deduction_type_display(),
            'type_value': deduction.deduction_type,
            'value': str(deduction.value),
            'calculated_amount': str(deduction.calculated_amount),
            'notes': deduction.notes,
            'is_active': deduction.is_active,
        })
    
    return JsonResponse({'success': True, 'deductions': deductions_data})
