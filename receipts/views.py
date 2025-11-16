from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db import transaction
from django.db.models import Q, Sum
from django.utils import timezone
from django.core.paginator import Paginator
from decimal import Decimal
import logging

logger = logging.getLogger(__name__)

from .models import PaymentReceipt, CheckCollection, ReceiptReversal
from customers.models import CustomerSupplier
from cashboxes.models import Cashbox, CashboxTransaction
from accounts.models import AccountTransaction
from journal.services import JournalService
from journal.models import JournalEntry


def process_cheque_errors_warnings():
    """
    Automatically process cheque errors and warnings according to IFRS 9
    """
    from datetime import datetime
    from django.utils import timezone

    # الحصول على جميع الشيكات
    cheques = PaymentReceipt.objects.filter(payment_type='check').select_related('customer')

    processed_errors = []
    processed_warnings = []

    for cheque in cheques:
        # معالجة الأخطاء: الشيكات المرتدة بدون قيد يومية
        if cheque.check_status == 'bounced':
            # فحص وجود قيد يومية
            journal_exists = JournalEntry.objects.filter(
                reference_type='check_bounced',
                reference_id=cheque.id
            ).exists()

            if not journal_exists:
                try:
                    # إنشاء قيد يومية تلقائي
                    collection_date = timezone.now().date()
                    JournalService.create_check_bounced_entry(cheque, collection_date)

                    # تحديث سبب الارتداد إذا لم يكن محدد
                    if not cheque.bounce_reason:
                        cheque.bounce_reason = _('Bounce detected during audit - journal entry created automatically')
                        cheque.save()

                    processed_errors.append({
                        'cheque': cheque,
                        'action': _('Automatically created journal entry for bounced check'),
                        'details': _('Journal entry transferring the amount from Accounts Receivable to Checks in collection for amount %(amount)s') % {'amount': cheque.amount}
                    })

                except Exception as e:
                    logger.error(_('Error creating journal entry for bounced check %(num)s: %(error)s') % {'num': cheque.check_number, 'error': e})

        # معالجة التحذيرات: الشيكات المحصلة
        elif cheque.check_status == 'collected':
            collection = CheckCollection.objects.filter(
                receipt=cheque,
                status='collected'
            ).first()

            if collection:
                days_difference = (collection.collection_date - cheque.check_due_date).days

                if days_difference > 0:
                    # تحصيل متأخر
                    processed_warnings.append({
                        'cheque': cheque,
                        'type': _('Late collection'),
                        'days_late': days_difference,
                        'action': _('Warning recorded - please follow up collection risks'),
                        'ifrs_note': _('May affect revenue timing under IFRS 9')
                    })

                elif days_difference < 0:
                    # تحصيل مبكر - فحص الفاتورة
                    from sales.models import SalesInvoice
                    try:
                        invoice = SalesInvoice.objects.filter(
                            customer=cheque.customer,
                            total_amount=cheque.amount,
                            date__lte=cheque.check_date
                        ).first()

                        if invoice:
                            processed_warnings.append({
                                'cheque': cheque,
                                'type': _('Early collection'),
                                'days_early': abs(days_difference),
                                'action': _('Check linked with invoice and revenue reviewed'),
                                'ifrs_note': _('Revenue reviewed - no IFRS 9 impact'),
                                'invoice': invoice.invoice_number
                            })
                        else:
                            processed_warnings.append({
                                'cheque': cheque,
                                'type': _('Early collection'),
                                'days_early': abs(days_difference),
                                'action': _('No related invoice found'),
                                'ifrs_note': _('Please ensure there is no premature revenue recognition')
                            })

                    except Exception as e:
                        logger.error(_('Error checking linked invoice for check %(num)s: %(error)s') % {'num': cheque.check_number, 'error': e})

    return {
        'processed_errors': processed_errors,
        'processed_warnings': processed_warnings
    }


def create_receipt_journal_entry(receipt, user):
    """Create journal entry for receipt voucher"""
    try:
        # Create journal entry using JournalService
        JournalService.create_receipt_voucher_entry(receipt, user)
    except Exception as e:
        print(f"Error creating journal entry for receipt voucher: {e}")
        # Don't stop the operation if journal entry creation fails
        pass


@login_required
def receipt_list(request):
    """Receipt vouchers list"""
    if not (request.user.has_perm('receipts.can_view_receipts') or request.user.has_perm('receipts.view_paymentreceipt')):
        from django.core.exceptions import PermissionDenied
        messages.error(request, _('You do not have permission to view receipt vouchers'))
        raise PermissionDenied(_('You do not have permission to view receipt vouchers'))
    
    receipts = PaymentReceipt.objects.all().select_related(
        'customer', 'cashbox', 'created_by'
    ).order_by('-date', '-receipt_number')
    
    # Filter by customer
    customer_id = request.GET.get('customer')
    if customer_id:
        receipts = receipts.filter(customer_id=customer_id)
    
    # Filter by payment type
    payment_type = request.GET.get('payment_type')
    if payment_type:
        receipts = receipts.filter(payment_type=payment_type)
    
    # Filter by date
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        receipts = receipts.filter(date__gte=date_from)
    if date_to:
        receipts = receipts.filter(date__lte=date_to)
    
    # Filter by status
    status = request.GET.get('status')
    if status == 'active':
        receipts = receipts.filter(is_active=True, is_reversed=False)
    elif status == 'reversed':
        receipts = receipts.filter(is_reversed=True)
    
    # Pagination
    paginator = Paginator(receipts, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Helper data
    customers = CustomerSupplier.objects.filter(
        type__in=['customer', 'both'], 
        is_active=True
    ).order_by('name')
    
    context = {
        'receipts': page_obj,
        'customers': customers,
        'page_title': _('Receipt Vouchers'),
    }
    return render(request, 'receipts/receipt_list.html', context)


@login_required
def receipt_add(request):
    """Add new receipt voucher"""
    if not (request.user.has_perm('receipts.can_add_receipts') or request.user.has_perm('receipts.add_paymentreceipt')):
        from django.core.exceptions import PermissionDenied
        messages.error(request, _('You do not have permission to add receipt vouchers'))
        raise PermissionDenied(_('You do not have permission to add receipt vouchers'))
    
    if request.method == 'POST':
        receipt_number = request.POST.get('receipt_number', '').strip()
        customer_id = request.POST.get('customer')
        payment_type = request.POST.get('payment_type')
        amount = request.POST.get('amount')
        date = request.POST.get('date')
        description = request.POST.get('description', '')
        notes = request.POST.get('notes', '')
        
        # Cash box data for cash payment
        cashbox_id = request.POST.get('cashbox')
        
        # Bank transfer data
        bank_account_id = request.POST.get('bank_account')
        bank_transfer_reference = request.POST.get('bank_transfer_reference', '')
        bank_transfer_date_raw = request.POST.get('bank_transfer_date', '').strip() if request.POST.get('bank_transfer_date') else ''
        bank_transfer_date = bank_transfer_date_raw if bank_transfer_date_raw else None  # تحويل السلسلة الفارغة إلى None
        bank_transfer_notes = request.POST.get('bank_transfer_notes', '')
        
        # Check data
        check_number = request.POST.get('check_number', '')
        check_date_raw = request.POST.get('check_date', '').strip() if request.POST.get('check_date') else ''
        check_date = check_date_raw if check_date_raw else None  # تحويل السلسلة الفارغة إلى None
        check_due_date_raw = request.POST.get('check_due_date', '').strip() if request.POST.get('check_due_date') else ''
        check_due_date = check_due_date_raw if check_due_date_raw else None  # تحويل السلسلة الفارغة إلى None
        bank_name = request.POST.get('bank_name', '')
        check_cashbox_id = request.POST.get('check_cashbox')
        
        # Validate basic data
        if not all([receipt_number, customer_id, payment_type, amount, date]):
            messages.error(request, _('All fields are required'))
            return redirect('receipts:receipt_add')
        
        # Check for duplicate receipt number
        if PaymentReceipt.objects.filter(receipt_number=receipt_number).exists():
            messages.error(request, _('Receipt number already exists, please choose another'))
            return redirect('receipts:receipt_add')
        
        try:
            amount = Decimal(amount)
            if amount <= 0:
                messages.error(request, _('Amount must be greater than zero'))
                return redirect('receipts:receipt_add')
            
            customer = get_object_or_404(
                CustomerSupplier, 
                id=customer_id, 
                type__in=['customer', 'both']
            )
            
            with transaction.atomic():
                # Create receipt voucher
                receipt = PaymentReceipt.objects.create(
                    receipt_number=receipt_number,
                    customer=customer,
                    payment_type=payment_type,
                    amount=amount,
                    date=date,
                    description=description,
                    notes=notes,
                    cashbox_id=cashbox_id if payment_type == 'cash' else None,
                    bank_account_id=bank_account_id if payment_type == 'bank_transfer' else None,
                    bank_transfer_reference=bank_transfer_reference if payment_type == 'bank_transfer' else '',
                    bank_transfer_date=bank_transfer_date if payment_type == 'bank_transfer' else None,
                    bank_transfer_notes=bank_transfer_notes if payment_type == 'bank_transfer' else '',
                    check_number=check_number if payment_type == 'check' else '',
                    check_date=check_date if payment_type == 'check' else None,
                    check_due_date=check_due_date if payment_type == 'check' else None,
                    bank_name=bank_name if payment_type == 'check' else '',
                    check_cashbox_id=check_cashbox_id if payment_type == 'check' else None,
                    created_by=request.user
                )
                
                # إضافة الحركة إلى حساب العميل
                AccountTransaction.create_transaction(
                    customer_supplier=customer,
                    transaction_type='receipt',  # تصحيح نوع المعاملة
                    direction='credit',  # دائن - يقلل من رصيد العميل
                    amount=amount,
                    reference_type='receipt',
                    reference_id=receipt.id,
                    description=f'{_("Receipt voucher")} {receipt.receipt_number} - {description}',
                    notes=f'{_("Payment type")}: {payment_type}',
                    user=request.user,
                    date=date
                )
                
                # For cash payment: add amount to cash box
                if payment_type == 'cash' and cashbox_id:
                    cashbox = get_object_or_404(Cashbox, id=cashbox_id)
                    
                    # Add cash box transaction - الرصيد سيتم تحديثه تلقائياً عبر signal
                    CashboxTransaction.objects.create(
                        cashbox=cashbox,
                        transaction_type='deposit',
                        date=date,
                        amount=amount,
                        description=f'{_("Receipt voucher")} {receipt.receipt_number} {_("from")} {customer.name}',
                        reference_type='receipt',
                        reference_id=receipt.id,
                        created_by=request.user
                    )
                
                # For bank transfer: add amount to bank account
                if payment_type == 'bank_transfer' and bank_account_id:
                    from banks.models import BankAccount, BankTransaction
                    bank_account = get_object_or_404(BankAccount, id=bank_account_id)
                    
                    # Add bank transaction
                    BankTransaction.objects.create(
                        bank=bank_account,
                        transaction_type='deposit',
                        date=bank_transfer_date if bank_transfer_date else date,
                        amount=amount,
                        reference_number=bank_transfer_reference,
                        description=f'{_("Receipt voucher")} {receipt.receipt_number} {_("from")} {customer.name} - {bank_transfer_notes}',
                        created_by=request.user
                    )
                    
                    # تسجيل النشاط
                    try:
                        from core.signals import log_user_activity
                        log_user_activity(
                            request,
                            'create',
                            receipt,
                            _('Bank transfer for receipt voucher %(number)s - Bank: %(account)s - Amount: %(amount)s') % {
                                'number': receipt.receipt_number,
                                'account': bank_account.name,
                                'amount': amount
                            }
                        )
                    except Exception as e:
                        logger.error(_('Error logging bank transfer activity for receipt %(num)s: %(error)s') % {'num': receipt.receipt_number, 'error': e})
                
                # For checks: update check status only (no cashbox update until collection)
                if payment_type == 'check':
                    receipt.check_status = 'pending'
                    receipt.save()
                
                # القيد المحاسبي يتم إنشاؤه تلقائياً بواسطة الـ signal في journal/signals.py
                # create_receipt_journal_entry(receipt, request.user)
                
                # تسجيل النشاط في سجل التدقيق
                from core.models import AuditLog
                payment_type_display = dict(PaymentReceipt.PAYMENT_TYPES).get(payment_type, payment_type)
                AuditLog.objects.create(
                    user=request.user,
                    action_type='create',
                    content_type='PaymentReceipt',
                    object_id=receipt.id,
                    description=_('Create receipt voucher %(number)s - Customer: %(customer)s - Amount: %(amount)s - Payment type: %(ptype)s') % {
                        'number': receipt.receipt_number,
                        'customer': customer.name,
                        'amount': amount,
                        'ptype': payment_type_display
                    }
                )
                
                messages.success(request, _('Receipt voucher {} created successfully').format(receipt.receipt_number))
                return redirect('receipts:receipt_detail', receipt_id=receipt.id)
        
        except Exception as e:
            messages.error(request, _('Error occurred while creating receipt voucher: {}').format(str(e)))
    
    # البيانات المساعدة للنموذج
    customers = CustomerSupplier.objects.filter(
        type__in=['customer', 'both'], 
        is_active=True
    ).order_by('name')
    cashboxes = Cashbox.objects.filter(is_active=True).order_by('name')
    
    # Get bank accounts
    from banks.models import BankAccount
    bank_accounts = BankAccount.objects.filter(is_active=True).order_by('name')
    
    context = {
        'customers': customers,
        'cashboxes': cashboxes,
        'bank_accounts': bank_accounts,
        'page_title': _('Add Receipt Voucher'),
        'today': timezone.now().date(),
    }
    return render(request, 'receipts/receipt_add.html', context)


@login_required
def receipt_detail(request, receipt_id):
    """Receipt voucher detail"""
    if not (request.user.has_perm('receipts.can_view_receipts') or request.user.has_perm('receipts.view_paymentreceipt')):
        from django.core.exceptions import PermissionDenied
        messages.error(request, _('You do not have permission to view receipt voucher details'))
        raise PermissionDenied(_('You do not have permission to view receipt voucher details'))
    
    receipt = get_object_or_404(PaymentReceipt, id=receipt_id)
    
    # Get related account movements
    account_movements = AccountTransaction.objects.filter(
        reference_type='receipt',
        reference_id=receipt.id
    ).order_by('-created_at')
    
    # Get check collections (if any)
    collections = CheckCollection.objects.filter(receipt=receipt).order_by('-collection_date')
    
    context = {
        'receipt': receipt,
        'account_movements': account_movements,
        'collections': collections,
        'page_title': f'{_("Receipt Voucher Details")} - {receipt.receipt_number}',
    }
    return render(request, 'receipts/receipt_detail.html', context)


@login_required
def receipt_edit(request, receipt_id):
    """Edit receipt voucher"""
    if not (request.user.has_perm('receipts.can_edit_receipts') or request.user.has_perm('receipts.change_paymentreceipt')):
        from django.core.exceptions import PermissionDenied
        messages.error(request, _('You do not have permission to edit receipt vouchers'))
        raise PermissionDenied(_('You do not have permission to edit receipt vouchers'))
    
    receipt = get_object_or_404(PaymentReceipt, id=receipt_id)
    
    # Check edit permission
    if receipt.is_reversed:
        messages.error(request, _('Cannot edit a reversed receipt voucher'))
        return redirect('receipts:receipt_detail', receipt_id=receipt_id)
    
    if request.method == 'POST':
        description = request.POST.get('description', '')
        notes = request.POST.get('notes', '')
        
        try:
            receipt.description = description
            receipt.notes = notes
            receipt.save()
            
            messages.success(request, _('Receipt voucher updated successfully'))
            return redirect('receipts:receipt_detail', receipt_id=receipt_id)
            
        except Exception as e:
            messages.error(request, _('Error occurred while updating: %(error)s') % {'error': str(e)})
    
    return redirect('receipts:receipt_detail', receipt_id=receipt_id)


@login_required
def receipt_reverse(request, receipt_id):
    """Reverse receipt voucher"""
    if not (request.user.has_perm('receipts.can_delete_receipts') or request.user.has_perm('receipts.delete_paymentreceipt')):
        from django.core.exceptions import PermissionDenied
        messages.error(request, _('You do not have permission to reverse receipt vouchers'))
        raise PermissionDenied(_('You do not have permission to reverse receipt vouchers'))
    
    receipt = get_object_or_404(PaymentReceipt, id=receipt_id)
    
    # Check reversal permission
    if not receipt.can_be_reversed:
        messages.error(request, _('This voucher cannot be reversed'))
        return redirect('receipts:receipt_detail', receipt_id=receipt_id)
    
    if request.method == 'POST':
        reason = request.POST.get('reason', '')
        notes = request.POST.get('notes', '')
        
        if not reason:
            messages.error(request, _('Reversal reason is required'))
            return redirect('receipts:receipt_detail', receipt_id=receipt_id)
        
        try:
            with transaction.atomic():
                # تحديث السند كمعكوس
                receipt.is_reversed = True
                receipt.reversed_by = request.user
                receipt.reversed_at = timezone.now()
                receipt.reversal_reason = reason
                receipt.save()
                
                # إنشاء سجل العكس
                reversal = ReceiptReversal.objects.create(
                    original_receipt=receipt,
                    reversal_date=timezone.now().date(),
                    reason=reason,
                    notes=notes,
                    created_by=request.user
                )
                
                # عكس حركة حساب العميل
                AccountTransaction.create_transaction(
                    customer_supplier=receipt.customer,
                    transaction_type='receipt',
                    direction='debit',  # مدين - يزيد من رصيد العميل (عكس الدفع)
                    amount=receipt.amount,
                    reference_type='receipt_reversal',
                    reference_id=reversal.id,
                    description=f'{_("Reversal of receipt voucher")} {receipt.receipt_number} - {reason}',
                    notes=f'{_("Reversal")}: {notes}',
                    user=request.user,
                    date=timezone.now().date()
                )
                
                # For cash payment: subtract amount from cash box
                if receipt.payment_type == 'cash' and receipt.cashbox:
                    cashbox = receipt.cashbox
                    if cashbox.balance >= receipt.amount:
                        # Add cash box transaction - الرصيد سيتم تحديثه تلقائياً عبر signal
                        CashboxTransaction.objects.create(
                            cashbox=cashbox,
                            transaction_type='withdrawal',
                            date=timezone.now().date(),
                            amount=receipt.amount,  # استخدم القيمة الموجبة، signal سيتعامل معها
                            description=f'{_("Reversal of receipt voucher")} {receipt.receipt_number} - {reason}',
                            reference_type='receipt',
                            reference_id=receipt.id,
                            created_by=request.user
                        )
                    else:
                        messages.warning(request, _('Warning: Insufficient cash box balance, reversal applied to account only'))
                
                messages.success(request, f'{_("Receipt voucher")} {receipt.receipt_number} {_("has been reversed successfully")}')
                return redirect('receipts:receipt_detail', receipt_id=receipt_id)
                
        except Exception as e:
            messages.error(request, f'{_("Error occurred while reversing voucher")}: {str(e)}')
    
    return redirect('receipts:receipt_detail', receipt_id=receipt_id)


@login_required
def check_list(request):
    """Checks list"""
    if not (request.user.has_perm('receipts.can_view_receipts') or request.user.has_perm('receipts.view_paymentreceipt')):
        from django.core.exceptions import PermissionDenied
        messages.error(request, _('You do not have permission to view receipt vouchers'))
        raise PermissionDenied(_('You do not have permission to view receipt vouchers'))
    
    checks = PaymentReceipt.objects.filter(
        payment_type='check'
    ).select_related('customer', 'created_by', 'check_cashbox').order_by('-check_due_date')
    
    # فلترة حسب الحالة
    status = request.GET.get('status')
    if status:
        checks = checks.filter(check_status=status)
    
    # فلترة حسب تاريخ الاستحقاق
    due_date_from = request.GET.get('due_date_from')
    due_date_to = request.GET.get('due_date_to')
    if due_date_from:
        checks = checks.filter(check_due_date__gte=due_date_from)
    if due_date_to:
        checks = checks.filter(check_due_date__lte=due_date_to)
    
    # التقسيم لصفحات
    paginator = Paginator(checks, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'checks': page_obj,
        'page_title': _('Checks'),
    }
    return render(request, 'receipts/check_list.html', context)


@login_required
def check_list_export_excel(request):
    """Export checks list to Excel"""
    # محاولة استيراد openpyxl
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from django.http import HttpResponse
        return HttpResponse(_('OpenPyXL is not available'), status=500)

    # الحصول على نفس البيانات المعروضة في القائمة
    checks = PaymentReceipt.objects.filter(
        payment_type='check'
    ).select_related('customer', 'created_by', 'check_cashbox').order_by('-check_due_date')

    # فلترة حسب الحالة
    status = request.GET.get('status')
    if status:
        checks = checks.filter(check_status=status)

    # فلترة حسب تاريخ الاستحقاق
    due_date_from = request.GET.get('due_date_from')
    due_date_to = request.GET.get('due_date_to')
    if due_date_from:
        checks = checks.filter(check_due_date__gte=due_date_from)
    if due_date_to:
        checks = checks.filter(check_due_date__lte=due_date_to)

    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = str(_('Checks List'))

    # تنسيق العناوين
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # العناوين
    headers = [
        str(_('Receipt Number')),
        str(_('Check Number')),
        str(_('Customer')),
        str(_('Amount')),
        str(_('Check Date')),
        str(_('Due Date')),
        str(_('Bank Name')),
        str(_('Check Cashbox')),
        str(_('Status')),
        str(_('ECL'))
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # البيانات
    row = 2
    for check in checks:
        ws.cell(row=row, column=1, value=check.receipt_number)
        ws.cell(row=row, column=2, value=check.check_number)
        ws.cell(row=row, column=3, value=check.customer.name)
        ws.cell(row=row, column=4, value=float(check.amount))
        ws.cell(row=row, column=5, value=check.check_date.strftime('%Y-%m-%d') if check.check_date else '')
        ws.cell(row=row, column=6, value=check.check_due_date.strftime('%Y-%m-%d') if check.check_due_date else '')
        ws.cell(row=row, column=7, value=check.bank_name)
        ws.cell(row=row, column=8, value=check.check_cashbox.name if check.check_cashbox else '')
        ws.cell(row=row, column=9, value=str(check.get_check_status_display()))
        ws.cell(row=row, column=10, value=float(check.expected_credit_loss) if check.expected_credit_loss else 0)
        row += 1

    # تعديل عرض الأعمدة
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # إنشاء الاستجابة
    from django.http import HttpResponse
    from django.utils import timezone
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"checks_list_{timezone.now().date()}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)

    # تسجيل النشاط
    try:
        from core.signals import log_export_activity
        log_export_activity(request, str(_('Checks List')), filename, 'Excel')
    except Exception:
        pass

    return response


@login_required
def check_collect(request, receipt_id):
    """Check collection with automatic handling of errors and warnings per IFRS 9"""
    receipt = get_object_or_404(PaymentReceipt, id=receipt_id, payment_type='check')
    
    if request.method == 'POST':
        collection_date = request.POST.get('collection_date')
        status = request.POST.get('status')  # collected أو bounced
        cashbox_id = request.POST.get('cashbox')
        notes = request.POST.get('notes', '')
        bounce_reason = request.POST.get('bounce_reason', '')  # سبب الارتداد
        
        if not all([collection_date, status]):
            messages.error(request, _('All fields are required'))
            return redirect('receipts:receipt_detail', receipt_id=receipt_id)
        
        try:
            with transaction.atomic():
                # إنشاء سجل التحصيل
                collection = CheckCollection.objects.create(
                    receipt=receipt,
                    collection_date=collection_date,
                    status=status,
                    cashbox_id=cashbox_id if status == 'collected' else None,
                    notes=notes,
                    created_by=request.user
                )
                
                # تحديث حالة الشيك
                receipt.check_status = status
                if status == 'bounced' and bounce_reason:
                    receipt.bounce_reason = bounce_reason
                receipt.save()
                
                # حساب خسائر الائتمان المتوقعة (ECL) وفق IFRS 9
                try:
                    ecl_amount, ecl_method = receipt.calculate_expected_credit_loss()
                    receipt.expected_credit_loss = ecl_amount
                    receipt.ecl_calculation_date = timezone.now().date()
                    receipt.ecl_calculation_method = ecl_method
                    receipt.save()
                    
                    # إضافة ملاحظة ECL في سجل التحصيل
                    if ecl_amount > 0:
                        collection.notes += '\n' + (_('💰 ECL calculated: %(amount)s (%(method)s)') % {'amount': ecl_amount, 'method': ecl_method})
                        collection.save()
                        
                        # تسجيل في السجل للمراجعة
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.info(_('ECL calculated for check %(num)s: %(amount)s (%(method)s)') % {'num': receipt.check_number, 'amount': ecl_amount, 'method': ecl_method})
                        
                except Exception as e:
                    logger.error(_('Error calculating ECL for check %(num)s: %(error)s') % {'num': receipt.check_number, 'error': e})
                
                # إذا تم التحصيل بنجاح
                if status == 'collected' and cashbox_id:
                    cashbox = get_object_or_404(Cashbox, id=cashbox_id)
                    
                    # Add cash box transaction - الرصيد سيتم تحديثه تلقائياً عبر signal
                    CashboxTransaction.objects.create(
                        cashbox=cashbox,
                        transaction_type='deposit',
                        date=collection_date,
                        amount=receipt.amount,
                        description=f'{_("Check collection")} {receipt.check_number} - {_("voucher")} {receipt.receipt_number}',
                        reference_type='receipt',
                        reference_id=receipt.id,
                        created_by=request.user
                    )
                    
                    # معالجة التحذيرات تلقائياً - IFRS 9
                    from datetime import datetime
                    collection_date_obj = datetime.strptime(collection_date, '%Y-%m-%d').date()
                    
                    if collection_date_obj > receipt.check_due_date:
                        # تحصيل بعد تاريخ الاستحقاق - حساب عدد الأيام المتأخرة
                        days_late = (collection_date_obj - receipt.check_due_date).days
                        
                        # إضافة تنبيه في السجل
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.warning(_('Check %(num)s collected past due date by %(days)s days. Due date: %(due)s, Collection date: %(coll)s. May affect revenue timing under IFRS 9.') % {
                            'num': receipt.check_number,
                            'days': days_late,
                            'due': receipt.check_due_date,
                            'coll': collection_date
                        })
                        
                        # إضافة ملاحظة في سجل التحصيل
                        collection.notes += '\n' + (_('⚠️ IFRS 9 warning: collected %(days)s days after due date (%(due)s)') % {'days': days_late, 'due': receipt.check_due_date})
                        collection.save()
                        
                        # توصية بمتابعة العميل
                        collection.notes += '\n' + _('📋 Recommendation: follow up the customer and monitor collection risks')
                        collection.save()
                    
                    elif collection_date_obj < receipt.check_due_date:
                        # تحصيل قبل تاريخ الاستحقاق - التحقق من حالة الفاتورة
                        days_early = (receipt.check_due_date - collection_date_obj).days
                        
                        # البحث عن الفاتورة المرتبطة بالشيك
                        from sales.models import SalesInvoice
                        try:
                            # افتراض أن الشيك مرتبط بفاتورة مبيعات
                            invoice = SalesInvoice.objects.filter(
                                customer=receipt.customer,
                                total_amount=receipt.amount,
                                date__lte=receipt.check_date
                            ).first()
                            
                            if invoice:
                                # التحقق من حالة الفاتورة (افتراض أن هناك حقل is_completed)
                                is_invoice_complete = getattr(invoice, 'is_completed', True)  # افتراض أنها مكتملة إذا لم يكن الحقل موجود
                                
                                if not is_invoice_complete:
                                    # الفاتورة غير مكتملة - تسجيل كدفعة مقدمة
                                    try:
                                        JournalService.create_check_early_collection_entry(
                                            receipt, collection_date_obj, is_invoice_complete=False, user=request.user
                                        )
                                    except Exception as je:
                                        logger.error(_('Error creating early collection journal entry for check %(num)s: %(error)s') % {'num': receipt.check_number, 'error': je})
                                    
                                    # إضافة تنبيه
                                    logger.info(_('Recorded check %(num)s as an advance payment due to incomplete invoice') % {'num': receipt.check_number})
                                    
                                    collection.notes += '\n' + _('ℹ️ Amount recorded as an advance from customers (invoice not completed)')
                                    collection.save()
                                else:
                                    # الفاتورة مكتملة - اعتراف طبيعي
                                    try:
                                        JournalService.create_check_early_collection_entry(
                                            receipt, collection_date_obj, is_invoice_complete=True, user=request.user
                                        )
                                    except Exception as je:
                                        logger.error(_('Error creating early collection journal entry for check %(num)s: %(error)s') % {'num': receipt.check_number, 'error': je})
                                    
                                    # إضافة ملاحظة
                                    collection.notes += '\n' + _('✅ Revenue reviewed - no IFRS 9 impact (invoice complete)')
                                    collection.save()
                            else:
                                # لم يتم العثور على فاتورة مرتبطة - اعتراف طبيعي
                                try:
                                    JournalService.create_check_early_collection_entry(
                                        receipt, collection_date_obj, is_invoice_complete=True, user=request.user
                                    )
                                except Exception as e:
                                    # Error searching for linked invoice - treat as normal receipt
                                    logger.error(_('Error searching for linked invoice: %(error)s') % {'error': e})
                                collection.notes += '\n' + _('⚠️ No related invoice found - please ensure no premature revenue recognition')
                                collection.save()
                        except Exception as e:
                            # Error searching for the linked invoice - treat as normal collection
                            logger.error(_('Error searching for linked invoice: %(error)s') % {'error': e})
                            try:
                                JournalService.create_check_early_collection_entry(
                                    receipt, collection_date_obj, is_invoice_complete=True, user=request.user
                                )
                            except Exception as je:
                                logger.error(_('Error creating early collection journal entry for check %(num)s: %(error)s') % {'num': receipt.check_number, 'error': je})
                
                # إذا ارتد الشيك - معالجة الأخطاء تلقائياً IFRS 9 متوافق
                elif status == 'bounced':
                    # إنشاء القيد اليومية للشيك المرتد
                    from datetime import datetime
                    collection_date_obj = datetime.strptime(collection_date, '%Y-%m-%d').date()
                    
                    try:
                        JournalService.create_check_bounced_entry(
                            receipt, collection_date_obj, user=request.user
                        )
                    except Exception as e:
                        logger.error(_('Error creating bounced check journal entry for check %(num)s: %(error)s') % {'num': receipt.check_number, 'error': e})
                    
                    # إضافة تنبيه في السجل
                    logger.warning(_('Check bounce %(num)s - automatic journal entry created to transfer the amount from checks in collection to accounts receivable under IFRS 9. Bounce reason: %(reason)s') % {
                        'num': receipt.check_number,
                        'reason': bounce_reason or _('Unspecified')
                    })
                    
                    # إضافة ملاحظة في سجل التحصيل
                    collection.notes += '\n' + _('❌ Bounced check journal entry created under IFRS 9')
                    if bounce_reason:
                        collection.notes += '\n' + _('📝 Bounce reason: %(reason)s') % {'reason': bounce_reason}
                    collection.save()
                
                status_text = _('Collected') if status == 'collected' else _('Bounced')
                messages.success(request, _('Recorded %(status)s for check %(num)s with automatic processing') % {'status': status_text, 'num': receipt.check_number})
                return redirect('receipts:receipt_detail', receipt_id=receipt_id)
                
        except Exception as e:
            messages.error(request, _('Error occurred while collecting the check: %(error)s') % {'error': str(e)})
    
    # البيانات المساعدة
    cashboxes = Cashbox.objects.filter(is_active=True).order_by('name')
    
    # أسباب الارتداد المحتملة
    bounce_reasons = [
        _('Insufficient funds'),
        _('Invalid signature'),
        _('Bank stop'),
        _('Invalid date'),
        _('Incorrect account number'),
        _('Forged check'),
        _('Other reasons')
    ]
    
    context = {
        'receipt': receipt,
        'cashboxes': cashboxes,
        'bounce_reasons': bounce_reasons,
        'page_title': f'{_("Check Collection")} - {receipt.check_number}',
        'today': timezone.now().date(),
    }
    return render(request, 'receipts/check_collect.html', context)


@login_required
def get_customer_balance(request, customer_id):
    """Get customer balance (Ajax)"""
    try:
        customer = get_object_or_404(
            CustomerSupplier, 
            id=customer_id, 
            type__in=['customer', 'both']
        )
        
        return JsonResponse({
            'balance': float(customer.balance),
            'customer_name': customer.name
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
