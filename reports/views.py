from datetime import date, timedelta
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.translation import gettext as _

from accounts.models import AccountTransaction
from customers.models import CustomerSupplier
from core.signals import log_view_activity, log_export_activity
from journal.models import Account, JournalEntry


@login_required
def reports_index(request):
    """صفحة رئيسية للتقارير"""
    if not request.user.has_perm('users.can_access_reports'):
        raise PermissionDenied("ليس لديك صلاحية الوصول للتقارير")
    
    # تسجيل النشاط
    log_view_activity(request, 'view', None, 'عرض صفحة التقارير الرئيسية')
    
    context = {
        'title': 'التقارير المالية',
    }
    return render(request, 'reports/index.html', context)


def _parse_date(value, default):
    try:
        # Expecting YYYY-MM-DD from input[type=date]
        parts = [int(x) for x in str(value).split('-')]
        if len(parts) == 3:
            return date(parts[0], parts[1], parts[2])
    except Exception:
        pass
    return default


def clean_numeric_value(value):
    """
    Clean numeric value by removing commas, spaces, and converting to float.
    Handles both string and numeric inputs.
    """
    if value is None or value == '':
        return 0.0

    # Convert to string and clean
    str_value = str(value).strip()

    # Remove commas and spaces
    str_value = str_value.replace(',', '').replace(' ', '')

    # Handle empty string after cleaning
    if not str_value or str_value == '-':
        return 0.0

    try:
        # Try to convert to float
        return float(str_value)
    except (ValueError, TypeError):
        # If conversion fails, return 0
        return 0.0


@login_required
def customer_statement(request):
    """
    Customer Statement report: select customer and date range, view transactions with opening/closing balances.
    Supports export to CSV/XLSX (XLSX best-effort; falls back to CSV if libs missing).
    """
    # Permission gate: allow superuser or user_type admin/superadmin implicitly; else require explicit permission
    user = request.user
    has_perm = (
        getattr(user, 'is_superuser', False) or
        getattr(user, 'user_type', None) in ['superadmin', 'admin'] or
        user.has_reports_permission()
    )
    if not has_perm:
        raise PermissionDenied

    customers = CustomerSupplier.objects.filter(type__in=['customer', 'both'], is_active=True).order_by('name')

    # Defaults: last 30 days
    today = date.today()
    default_start = today - timedelta(days=30)
    default_end = today

    customer_id = request.GET.get('customer')
    start_raw = request.GET.get('start_date')
    end_raw = request.GET.get('end_date')
    export = request.GET.get('export')  # csv or xlsx

    # Defaults: last 30 days
    today = date.today()
    default_start = today - timedelta(days=30)
    default_end = today

    customer_id = request.GET.get('customer')
    start_raw = request.GET.get('start_date')
    end_raw = request.GET.get('end_date')
    export = request.GET.get('export')  # csv or xlsx

    start_date = _parse_date(start_raw, default_start)
    end_date = _parse_date(end_raw, default_end)

    selected_customer = None
    statement_rows = []
    opening_balance = Decimal('0.000')
    totals_debit = Decimal('0.000')
    totals_credit = Decimal('0.000')
    closing_balance = Decimal('0.000')

    if customer_id:
        selected_customer = customers.filter(pk=customer_id).first()

    view_logged = False

    if selected_customer:
        # Opening balance = net of all transactions before start_date
        before_qs = AccountTransaction.objects.filter(
            customer_supplier=selected_customer,
            date__lt=start_date,
        )
        before_debit = before_qs.filter(direction='debit').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        before_credit = before_qs.filter(direction='credit').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        opening_balance = (before_debit - before_credit).quantize(Decimal('0.001'))

        # Transactions in period
        period_qs = AccountTransaction.objects.filter(
            customer_supplier=selected_customer,
            date__gte=start_date,
            date__lte=end_date,
        ).order_by('date', 'created_at', 'id')

        running = opening_balance
        for t in period_qs:
            debit = t.amount if t.direction == 'debit' else Decimal('0')
            credit = t.amount if t.direction == 'credit' else Decimal('0')
            running = (running + debit - credit).quantize(Decimal('0.001'))
            statement_rows.append({
                'date': t.date,
                'number': t.transaction_number,
                'type': t.get_transaction_type_display(),
                'description': t.description or '',
                'debit': debit,
                'credit': credit,
                'running_balance': running,
            })

        totals_debit = period_qs.filter(direction='debit').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        totals_credit = period_qs.filter(direction='credit').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        closing_balance = (opening_balance + totals_debit - totals_credit).quantize(Decimal('0.001'))

        # Audit log for viewing a report
        try:
            desc = _("Viewing customer statement for %(customer)s from %(start)s to %(end)s") % {
                'customer': selected_customer.name,
                'start': start_date,
                'end': end_date,
            }
            class ReportObj:
                id = 0
                pk = 0
                def __str__(self):
                    return str(_('Customer Statement'))
            log_view_activity(request, 'view', ReportObj(), desc)
            view_logged = True
        except Exception:
            pass

        # Handle export
        if export and export.lower() in ('csv', 'xlsx'):
            filename_base = f"customer_statement_{selected_customer.sequence_number or selected_customer.pk}_{start_date}_{end_date}"
            if export.lower() == 'xlsx':
                # Best-effort XLSX using openpyxl; fallback to CSV
                try:
                    import io
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.title = str(_('Statement'))
                    headers = [
                        _('Date'), _('Number'), _('Type'), _('Description'), _('Debit'), _('Credit'), _('Running Balance')
                    ]
                    ws.append(headers)
                    # Opening row
                    ws.append([str(start_date), '-', _('Opening Balance'), '', '', '', clean_numeric_value(opening_balance)])
                    for r in statement_rows:
                        ws.append([
                            str(r['date']), r['number'], r['type'], r['description'], clean_numeric_value(r['debit']), clean_numeric_value(r['credit']), clean_numeric_value(r['running_balance'])
                        ])
                    # Totals and closing
                    ws.append(['', '', _('Totals'), '', clean_numeric_value(totals_debit), clean_numeric_value(totals_credit), ''])
                    ws.append(['', '', _('Closing Balance'), '', '', '', clean_numeric_value(closing_balance)])
                    bio = io.BytesIO()
                    wb.save(bio)
                    bio.seek(0)
                    response = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = f'attachment; filename={filename_base}.xlsx'
                    try:
                        log_export_activity(request, str(_('Customer Statement')), f'{filename_base}.xlsx', 'XLSX')
                    except Exception:
                        pass
                    return response
                except Exception:
                    # Fallback to CSV
                    export = 'csv'

            if export.lower() == 'csv':
                import csv
                response = HttpResponse(content_type='text/csv; charset=utf-8')
                response.write('\ufeff')  # BOM for Excel + Arabic
                response['Content-Disposition'] = f'attachment; filename={filename_base}.csv'
                writer = csv.writer(response)
                writer.writerow([_('Customer Statement')])
                writer.writerow([_('Customer'), selected_customer.name])
                writer.writerow([_('From'), start_date, _('To'), end_date])
                writer.writerow([])
                writer.writerow([_('Date'), _('Number'), _('Type'), _('Description'), _('Debit'), _('Credit'), _('Running Balance')])
                writer.writerow([start_date, '-', _('Opening Balance'), '', '', '', opening_balance])
                for r in statement_rows:
                    writer.writerow([r['date'], r['number'], r['type'], r['description'], r['debit'], r['credit'], r['running_balance']])
                writer.writerow(['', '', _('Totals'), '', totals_debit, totals_credit, ''])
                writer.writerow(['', '', _('Closing Balance'), '', '', '', closing_balance])
                try:
                    log_export_activity(request, str(_('Customer Statement')), f'{filename_base}.csv', 'CSV')
                except Exception:
                    pass
                return response

    context = {
        'customers': customers,
        'selected_customer': selected_customer,
        'start_date': start_date,
        'end_date': end_date,
        'statement_rows': statement_rows,
        'opening_balance': opening_balance,
        'totals_debit': totals_debit,
        'totals_credit': totals_credit,
        'closing_balance': closing_balance,
    }

    if not view_logged:
        try:
            class ReportObj:
                id = 0
                pk = 0
                def __str__(self):
                    return str(_('Customer Statement'))
            log_view_activity(request, 'view', ReportObj(), str(_('Viewing customer statement page')))
        except Exception:
            pass

    return render(request, 'reports/customer_statement.html', context)


@login_required
def sales_by_salesperson(request):
    """
    Sales Report By Sales Person: select salesperson and date range, view sales invoices created by that person.
    Supports export to CSV/XLSX.
    """
    # Permission gate: allow superuser or user_type admin/superadmin implicitly; else require explicit permission
    user = request.user
    has_perm = (
        getattr(user, 'is_superuser', False) or
        getattr(user, 'user_type', None) in ['superadmin', 'admin'] or
        user.has_reports_permission()
    )
    if not has_perm:
        raise PermissionDenied

    from django.contrib.auth import get_user_model
    from sales.models import SalesInvoice
    User = get_user_model()

    # All active users can be selected; النتائج قد تكون فارغة إذا لم ينشئ المستخدم فواتير
    sales_users = User.objects.filter(
        is_active=True
    ).order_by('first_name', 'last_name', 'username')

    # Defaults: last 90 days + next 30 days (لتغطية الفواتير المستقبلية أيضاً)
    today = date.today()
    default_start = today - timedelta(days=90)
    default_end = today + timedelta(days=30)

    salesperson_id = request.GET.get('salesperson')
    start_raw = request.GET.get('start_date')
    end_raw = request.GET.get('end_date')
    export = request.GET.get('export')

    start_date = _parse_date(start_raw, default_start)
    end_date = _parse_date(end_raw, default_end)

    selected_salesperson = None
    sales_invoices = []
    totals_subtotal = Decimal('0.000')
    totals_tax = Decimal('0.000')
    totals_discount = Decimal('0.000')
    totals_total = Decimal('0.000')
    invoice_count = 0
    returns_count = 0

    view_logged = False

    if salesperson_id:
        try:
            selected_salesperson = User.objects.get(pk=salesperson_id)
        except User.DoesNotExist:
            selected_salesperson = None

    if selected_salesperson:
        # Get sales invoices created by this salesperson in the date range
        invoices_qs = SalesInvoice.objects.filter(
            created_by=selected_salesperson,
            date__gte=start_date,
            date__lte=end_date,
        ).order_by('-date', '-invoice_number').select_related('customer')

        # Get sales returns created by this salesperson in the date range
        from sales.models import SalesReturn
        returns_qs = SalesReturn.objects.filter(
            created_by=selected_salesperson,
            date__gte=start_date,
            date__lte=end_date,
        ).order_by('-date', '-return_number').select_related('customer', 'original_invoice')

        for invoice in invoices_qs:
            sales_invoices.append({
                'type': 'invoice',
                'date': invoice.date,
                'document_number': invoice.invoice_number,
                'customer': invoice.customer.name if invoice.customer else _('Cash Customer'),
                'payment_type': invoice.get_payment_type_display(),
                'subtotal': invoice.subtotal,
                'tax_amount': invoice.tax_amount,
                'discount_amount': invoice.discount_amount,
                'total_amount': invoice.total_amount,
                'original_invoice': None,
            })

        for return_invoice in returns_qs:
            sales_invoices.append({
                'type': 'return',
                'date': return_invoice.date,
                'document_number': return_invoice.return_number,
                'customer': return_invoice.customer.name if return_invoice.customer else _('Cash Customer'),
                'payment_type': 'مردود',
                'subtotal': -return_invoice.subtotal,  # سالب لأنه مردود
                'tax_amount': -return_invoice.tax_amount,
                'discount_amount': Decimal('0'),  # المردودات لا تحتوي على خصم عادة
                'total_amount': -return_invoice.total_amount,
                'original_invoice': return_invoice.original_invoice.invoice_number if return_invoice.original_invoice else None,
            })

        # Sort by date descending
        sales_invoices.sort(key=lambda x: x['date'], reverse=True)

        # Calculate totals (including returns as negative values)
        totals_subtotal = invoices_qs.aggregate(total=Sum('subtotal'))['total'] or Decimal('0')
        totals_tax = invoices_qs.aggregate(total=Sum('tax_amount'))['total'] or Decimal('0')
        totals_discount = invoices_qs.aggregate(total=Sum('discount_amount'))['total'] or Decimal('0')
        totals_total = invoices_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        
        # Subtract returns from totals
        returns_subtotal = returns_qs.aggregate(total=Sum('subtotal'))['total'] or Decimal('0')
        returns_tax = returns_qs.aggregate(total=Sum('tax_amount'))['total'] or Decimal('0')
        returns_total = returns_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        
        totals_subtotal -= returns_subtotal
        totals_tax -= returns_tax
        totals_total -= returns_total
        
        invoice_count = invoices_qs.count()
        returns_count = returns_qs.count()

        # Audit log for viewing a report
        try:
            desc = _("Viewing sales report for %(salesperson)s from %(start)s to %(end)s") % {
                'salesperson': selected_salesperson.get_full_name() or selected_salesperson.username,
                'start': start_date,
                'end': end_date,
            }
            class ReportObj:
                id = 0
                pk = 0
                def __str__(self):
                    return str(_('Sales Report By Sales Person'))
            log_view_activity(request, 'view', ReportObj(), desc)
            view_logged = True
        except Exception:
            pass

        # Handle export
        if export and export.lower() in ('csv', 'xlsx'):
            filename_base = f"sales_by_salesperson_{selected_salesperson.username}_{start_date}_{end_date}"
            if export.lower() == 'xlsx':
                try:
                    import io
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.title = str(_('Sales Report'))
                    headers = [
                        _('Type'), _('Date'), _('Document Number'), _('Customer'), _('Payment Type'), 
                        _('Subtotal'), _('Tax Amount'), _('Discount Amount'), _('Total Amount'), _('Original Invoice')
                    ]
                    ws.append(headers)
                    for inv in sales_invoices:
                        ws.append([
                            'فاتورة' if inv['type'] == 'invoice' else 'مردود',
                            str(inv['date']), inv['document_number'], inv['customer'], inv['payment_type'],
                            clean_numeric_value(inv['subtotal']), clean_numeric_value(inv['tax_amount']), clean_numeric_value(inv['discount_amount']), 
                            clean_numeric_value(inv['total_amount']), inv['original_invoice'] or ''
                        ])
                    # Totals row
                    ws.append(['', '', '', _('Totals'), '', clean_numeric_value(totals_subtotal), clean_numeric_value(totals_tax), clean_numeric_value(totals_discount), clean_numeric_value(totals_total), ''])
                    ws.append(['', '', '', _('Invoice Count'), invoice_count, '', '', '', '', ''])
                    ws.append(['', '', '', _('Returns Count'), returns_count, '', '', '', '', ''])
                    bio = io.BytesIO()
                    wb.save(bio)
                    bio.seek(0)
                    response = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = f'attachment; filename={filename_base}.xlsx'
                    try:
                        log_export_activity(request, str(_('Sales Report By Sales Person')), f'{filename_base}.xlsx', 'XLSX')
                    except Exception:
                        pass
                    return response
                except Exception:
                    export = 'csv'

            if export.lower() == 'csv':
                import csv
                response = HttpResponse(content_type='text/csv; charset=utf-8')
                response.write('\ufeff')
                response['Content-Disposition'] = f'attachment; filename={filename_base}.csv'
                writer = csv.writer(response)
                writer.writerow([_('Sales Report By Sales Person')])
                writer.writerow([_('Sales Person'), selected_salesperson.get_full_name() or selected_salesperson.username])
                writer.writerow([_('From'), start_date, _('To'), end_date])
                writer.writerow([])
                writer.writerow([_('Date'), _('Type'), _('Document Number'), _('Customer'), _('Payment Type'), _('Subtotal'), _('Tax Amount'), _('Discount Amount'), _('Total Amount'), _('Original Invoice')])
                for inv in sales_invoices:
                    writer.writerow([
                        inv['date'], 
                        'فاتورة' if inv['type'] == 'invoice' else 'مردود',
                        inv['document_number'], 
                        inv['customer'], 
                        inv['payment_type'], 
                        inv['subtotal'], 
                        inv['tax_amount'], 
                        inv['discount_amount'], 
                        inv['total_amount'],
                        inv['original_invoice'] or ''
                    ])
                writer.writerow(['', '', '', _('Totals'), '', totals_subtotal, totals_tax, totals_discount, totals_total, ''])
                writer.writerow(['', '', '', _('Invoice Count'), invoice_count, '', '', '', '', ''])
                writer.writerow(['', '', '', _('Returns Count'), returns_count, '', '', '', '', ''])
                try:
                    log_export_activity(request, str(_('Sales Report By Sales Person')), f'{filename_base}.csv', 'CSV')
                except Exception:
                    pass
                return response

    context = {
        'sales_users': sales_users,
        'selected_salesperson': selected_salesperson,
        'start_date': start_date,
        'end_date': end_date,
        'sales_invoices': sales_invoices,
        'totals_subtotal': totals_subtotal,
        'totals_tax': totals_tax,
        'totals_discount': totals_discount,
        'totals_total': totals_total,
        'invoice_count': invoice_count,
        'returns_count': returns_count,
    }

    if not view_logged:
        try:
            class ReportObj:
                id = 0
                pk = 0
                def __str__(self):
                    return str(_('Sales Report By Sales Person'))
            log_view_activity(request, 'view', ReportObj(), str(_('Viewing sales report by salesperson page')))
        except Exception:
            pass

    return render(request, 'reports/sales_by_salesperson.html', context)


@login_required
def trial_balance(request):
    """
    Trial Balance report: view all accounts with their balances as of a specific date.
    Supports export to PDF and Excel.
    """
    # Permission gate
    user = request.user
    has_perm = (
        getattr(user, 'is_superuser', False) or
        getattr(user, 'user_type', None) in ['superadmin', 'admin'] or
        user.has_reports_permission()
    )
    if not has_perm:
        raise PermissionDenied

    # Defaults
    today = date.today()
    as_of_date = _parse_date(request.GET.get('as_of_date'), today)
    account_type_filter = request.GET.get('account_type', '')
    export = request.GET.get('export', '')  # pdf or excel

    # Get all active accounts for trial balance
    # IFRS compliant: Include ALL accounts that have journal entries
    # This includes both leaf accounts and parent accounts with direct entries
    from django.db.models import Exists, OuterRef, Q

    all_accounts = Account.objects.filter(is_active=True).order_by('code')
    
    # Get accounts that have journal entries (either direct or through children)
    accounts = []
    for account in all_accounts:
        # Check if this account has direct journal entries
        from journal.models import JournalLine
        has_entries = JournalLine.objects.filter(account=account).exists()
        
        # Include account if it has entries AND doesn't have active children
        # OR if it has entries AND has children (parent with direct entries - shouldn't happen but handle it)
        has_active_children = account.children.filter(is_active=True).exists()
        
        if has_entries and not has_active_children:
            # Leaf account with entries - include it
            balance = account.get_balance(as_of_date=as_of_date)
            if balance != 0:
                accounts.append(account)
        elif has_entries and has_active_children:
            # Parent account with direct entries - SHOULD NOT HAPPEN but include anyway
            # Calculate balance from direct entries only (not including children)
            lines = JournalLine.objects.filter(account=account)
            if as_of_date:
                lines = lines.filter(journal_entry__entry_date__lte=as_of_date)
            
            from django.db.models import Sum
            debit_total = lines.aggregate(total=Sum('debit'))['total'] or Decimal('0')
            credit_total = lines.aggregate(total=Sum('credit'))['total'] or Decimal('0')
            
            if account.account_type in ['asset', 'expense', 'purchases']:
                balance = debit_total - credit_total
            else:
                balance = credit_total - debit_total
            
            if balance != 0:
                accounts.append(account)

    if account_type_filter:
        accounts = [acc for acc in accounts if acc.account_type == account_type_filter]

    # Calculate balances
    trial_balance_data = []
    total_debit = Decimal('0')
    total_credit = Decimal('0')

    for account in accounts:
        # For parent accounts with direct entries, calculate from direct entries only
        from journal.models import JournalLine
        has_active_children = account.children.filter(is_active=True).exists()
        
        if has_active_children:
            # Parent account - get balance from direct entries only (not children)
            lines = JournalLine.objects.filter(account=account)
            if as_of_date:
                lines = lines.filter(journal_entry__entry_date__lte=as_of_date)
            
            from django.db.models import Sum
            debit_total = lines.aggregate(total=Sum('debit'))['total'] or Decimal('0')
            credit_total = lines.aggregate(total=Sum('credit'))['total'] or Decimal('0')
            
            if account.account_type in ['asset', 'expense', 'purchases']:
                balance = debit_total - credit_total
            else:
                balance = credit_total - debit_total
        else:
            # Leaf account - use get_balance method
            balance = account.get_balance(as_of_date=as_of_date)

        # IFRS-compliant trial balance logic:
        # Accounts with normal debit balances: Assets, Expenses, Purchases
        # Accounts with normal credit balances: Liabilities, Equity, Revenue
        
        if account.account_type in ['asset', 'expense', 'purchases']:
            # Normal debit balance accounts
            if balance >= 0:
                debit_balance = balance
                credit_balance = Decimal('0')
            else:
                # Negative balance = credit side
                debit_balance = Decimal('0')
                credit_balance = -balance
        else:  # liability, equity, revenue
            # Normal credit balance accounts
            if balance >= 0:
                debit_balance = Decimal('0')
                credit_balance = balance
            else:
                # Negative balance = debit side
                debit_balance = -balance
                credit_balance = Decimal('0')

        trial_balance_data.append({
            'account': account,
            'debit_balance': debit_balance,
            'credit_balance': credit_balance,
        })
        total_debit += debit_balance
        total_credit += credit_balance

    # Handle export
    if export == 'pdf':
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        from io import BytesIO

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        styles = getSampleStyleSheet()
        elements.append(Paragraph(_('Trial Balance Report'), styles['Heading1']))
        elements.append(Paragraph(f"{_('As of Date')}: {as_of_date}", styles['Normal']))

        data = [[_('Account Code'), _('Account Name'), _('Debit Balance'), _('Credit Balance')]]
        for item in trial_balance_data:
            data.append([
                item['account'].code,
                item['account'].name,
                f"{item['debit_balance']:.3f}",
                f"{item['credit_balance']:.3f}",
            ])
        data.append([_('Total'), '', f"{total_debit:.3f}", f"{total_credit:.3f}"])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="trial_balance_{as_of_date}.pdf"'
        try:
            log_export_activity(request, str(_('Trial Balance Report')), f'trial_balance_{as_of_date}.pdf', 'PDF')
        except Exception:
            pass
        return response

    elif export == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.title = _('Trial Balance')

        # Headers
        ws['A1'] = _('Trial Balance Report')
        ws['A2'] = f"{_('As of Date')}: {as_of_date}"
        ws['A4'] = _('Account Code')
        ws['B4'] = _('Account Name')
        ws['C4'] = _('Debit Balance')
        ws['D4'] = _('Credit Balance')

        # Data
        row = 5
        for item in trial_balance_data:
            ws[f'A{row}'] = item['account'].code
            ws[f'B{row}'] = item['account'].name
            ws[f'C{row}'] = clean_numeric_value(item['debit_balance'])
            ws[f'D{row}'] = clean_numeric_value(item['credit_balance'])
            row += 1

        # Totals
        ws[f'A{row}'] = _('Total')
        ws[f'C{row}'] = clean_numeric_value(total_debit)
        ws[f'D{row}'] = clean_numeric_value(total_credit)

        # Styling
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}4'].font = Font(bold=True)
            ws[f'{col}4'].alignment = Alignment(horizontal='center')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="trial_balance_{as_of_date}.xlsx"'
        wb.save(response)
        try:
            log_export_activity(request, str(_('Trial Balance Report')), f'trial_balance_{as_of_date}.xlsx', 'Excel')
        except Exception:
            pass
        return response

    context = {
        'as_of_date': as_of_date,
        'account_type_filter': account_type_filter,
        'trial_balance_data': trial_balance_data,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'account_types': Account.ACCOUNT_TYPES,
    }

    try:
        class ReportObj:
            id = 0
            pk = 0
            def __str__(self):
                return str(_('Trial Balance Report'))
        log_view_activity(request, 'view', ReportObj(), str(_('Viewing trial balance report page')))
    except Exception:
        pass

    return render(request, 'reports/trial_balance.html', context)


@login_required
def balance_sheet(request):
    """
    الميزانية العمومية - Balance Sheet
    """
    user = request.user
    has_perm = (
        getattr(user, 'is_superuser', False) or
        getattr(user, 'user_type', None) in ['superadmin', 'admin'] or
        user.has_reports_permission()
    )
    if not has_perm:
        raise PermissionDenied

    # فلاتر التاريخ
    today = date.today()
    as_of_date = _parse_date(request.GET.get('as_of_date'), today)

    # حساب الأصول
    assets = Account.objects.filter(account_type='asset', is_active=True)
    total_assets = Decimal('0')
    asset_accounts = []
    for account in assets:
        balance = account.get_balance(as_of_date)
        if balance != 0:
            asset_accounts.append({
                'account': account,
                'balance': balance
            })
            total_assets += balance

    # حساب المطلوبات
    liabilities = Account.objects.filter(account_type='liability', is_active=True)
    total_liabilities = Decimal('0')
    liability_accounts = []
    for account in liabilities:
        balance = account.get_balance(as_of_date)
        if balance != 0:
            liability_accounts.append({
                'account': account,
                'balance': balance
            })
            total_liabilities += balance

    # حساب حقوق الملكية
    equities = Account.objects.filter(account_type='equity', is_active=True)
    total_equity = Decimal('0')
    equity_accounts = []
    for account in equities:
        balance = account.get_balance(as_of_date)
        if balance != 0:
            equity_accounts.append({
                'account': account,
                'balance': balance
            })
            total_equity += balance

    # تسجيل النشاط
    try:
        class ReportObj:
            id = 0
            pk = 0
            def __str__(self):
                return str(_('Balance Sheet Report'))
        log_view_activity(request, 'view', ReportObj(), str(_('Viewing balance sheet report')))
    except Exception:
        pass

    context = {
        'as_of_date': as_of_date,
        'asset_accounts': asset_accounts,
        'total_assets': total_assets,
        'liability_accounts': liability_accounts,
        'total_liabilities': total_liabilities,
        'equity_accounts': equity_accounts,
        'total_equity': total_equity,
        'total_liabilities_equity': total_liabilities + total_equity,
    }

    return render(request, 'reports/balance_sheet.html', context)


@login_required
def income_statement(request):
    """
    قائمة الدخل - Income Statement
    """
    user = request.user
    has_perm = (
        getattr(user, 'is_superuser', False) or
        getattr(user, 'user_type', None) in ['superadmin', 'admin'] or
        user.has_reports_permission()
    )
    if not has_perm:
        raise PermissionDenied

    # فلاتر التاريخ
    today = date.today()
    start_date = _parse_date(request.GET.get('start_date'), today.replace(day=1))
    end_date = _parse_date(request.GET.get('end_date'), today)

    # حساب الإيرادات
    revenues = Account.objects.filter(account_type='revenue', is_active=True)
    total_revenues = Decimal('0')
    revenue_accounts = []
    for account in revenues:
        balance = account.get_balance(end_date) - account.get_balance(start_date - timedelta(days=1))
        if balance != 0:
            revenue_accounts.append({
                'account': account,
                'balance': balance
            })
            total_revenues += balance

    # حساب المصاريف
    expenses = Account.objects.filter(account_type='expense', is_active=True)
    total_expenses = Decimal('0')
    expense_accounts = []
    for account in expenses:
        balance = account.get_balance(end_date) - account.get_balance(start_date - timedelta(days=1))
        if balance != 0:
            expense_accounts.append({
                'account': account,
                'balance': balance
            })
            total_expenses += balance

    # حساب المبيعات
    sales = Account.objects.filter(account_type='sales', is_active=True)
    total_sales = Decimal('0')
    sales_accounts = []
    for account in sales:
        balance = account.get_balance(end_date) - account.get_balance(start_date - timedelta(days=1))
        if balance != 0:
            sales_accounts.append({
                'account': account,
                'balance': balance
            })
            total_sales += balance

    # حساب المشتريات
    purchases = Account.objects.filter(account_type='purchases', is_active=True)
    total_purchases = Decimal('0')
    purchase_accounts = []
    for account in purchases:
        balance = account.get_balance(end_date) - account.get_balance(start_date - timedelta(days=1))
        if balance != 0:
            purchase_accounts.append({
                'account': account,
                'balance': balance
            })
            total_purchases += balance

    # حساب صافي الربح
    gross_profit = total_sales - total_purchases
    net_profit = total_revenues - total_expenses + gross_profit

    # تسجيل النشاط
    try:
        class ReportObj:
            id = 0
            pk = 0
            def __str__(self):
                return str(_('Income Statement Report'))
        log_view_activity(request, 'view', ReportObj(), str(_('Viewing income statement report')))
    except Exception:
        pass

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'sales_accounts': sales_accounts,
        'total_sales': total_sales,
        'purchase_accounts': purchase_accounts,
        'total_purchases': total_purchases,
        'gross_profit': gross_profit,
        'revenue_accounts': revenue_accounts,
        'total_revenues': total_revenues,
        'expense_accounts': expense_accounts,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
    }

    return render(request, 'reports/income_statement.html', context)


@login_required
def cash_flow(request):
    """
    التدفقات النقدية - Cash Flow Statement
    """
    user = request.user
    has_perm = (
        getattr(user, 'is_superuser', False) or
        getattr(user, 'user_type', None) in ['superadmin', 'admin'] or
        user.has_reports_permission()
    )
    if not has_perm:
        raise PermissionDenied

    # فلاتر التاريخ
    today = date.today()
    start_date = _parse_date(request.GET.get('start_date'), today.replace(day=1))
    end_date = _parse_date(request.GET.get('end_date'), today)

    # حساب التدفقات النقدية من العمليات التشغيلية
    # حساب التغيير في أرصدة الحسابات النقدية خلال الفترة
    
    # تحديد الحسابات النقدية (الأصول النقدية)
    # الحسابات التي تحتوي على نقد أو بنوك أو صناديق
    cash_accounts = Account.objects.filter(
        is_active=True,
        account_type='asset'
    ).filter(
        Q(name__icontains='نقد') |
        Q(name__icontains='صندوق') |
        Q(name__icontains='بنك') |
        Q(code__startswith='101') |  # صناديق
        Q(code__startswith='102')    # بنوك
    )
    
    # حساب أرصدة نهاية الفترة
    cash_end_balance = Decimal('0')
    for account in cash_accounts.distinct():
        cash_end_balance += account.get_balance(end_date)
    
    # حساب أرصدة بداية الفترة (يوم قبل start_date)
    start_minus_one = start_date - timedelta(days=1)
    cash_start_balance = Decimal('0')
    for account in cash_accounts.distinct():
        cash_start_balance += account.get_balance(start_minus_one)
    
    # التدفق النقدي من العمليات التشغيلية = التغيير في النقد
    operating_cash_flow = cash_end_balance - cash_start_balance

    # حساب التدفقات النقدية من الاستثمارات
    investment_cash_flow = Decimal('0')

    # حساب التدفقات النقدية من التمويل
    financing_cash_flow = Decimal('0')

    # تسجيل النشاط
    try:
        class ReportObj:
            id = 0
            pk = 0
            def __str__(self):
                return str(_('Cash Flow Statement Report'))
        log_view_activity(request, 'view', ReportObj(), str(_('Viewing cash flow statement report')))
    except Exception:
        pass

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'operating_cash_flow': operating_cash_flow,
        'investment_cash_flow': investment_cash_flow,
        'financing_cash_flow': financing_cash_flow,
        'net_cash_flow': operating_cash_flow + investment_cash_flow + financing_cash_flow,
    }

    return render(request, 'reports/cash_flow.html', context)


@login_required
def financial_ratios(request):
    """
    المؤشرات المالية - Financial Ratios
    """
    user = request.user
    has_perm = (
        getattr(user, 'is_superuser', False) or
        getattr(user, 'user_type', None) in ['superadmin', 'admin'] or
        user.has_reports_permission()
    )
    if not has_perm:
        raise PermissionDenied

    # فلاتر التاريخ
    today = date.today()
    as_of_date = _parse_date(request.GET.get('as_of_date'), today)

    # حساب المؤشرات
    ratios = {}

    # هامش الربح (Profit Margin)
    # نحتاج إلى حساب صافي الربح والإيرادات
    revenues = Account.objects.filter(account_type__in=['revenue', 'sales'], is_active=True)
    total_revenues = Decimal('0')
    for account in revenues:
        total_revenues += account.get_balance(as_of_date)

    expenses = Account.objects.filter(account_type='expense', is_active=True)
    total_expenses = Decimal('0')
    for account in expenses:
        total_expenses += account.get_balance(as_of_date)

    purchases = Account.objects.filter(account_type='purchases', is_active=True)
    total_purchases = Decimal('0')
    for account in purchases:
        total_purchases += account.get_balance(as_of_date)

    net_profit = total_revenues - total_expenses - total_purchases

    if total_revenues > 0:
        ratios['profit_margin'] = (net_profit / total_revenues) * 100
    else:
        ratios['profit_margin'] = Decimal('0')

    # السيولة (Liquidity Ratio) - الأصول المتداولة / المطلوبات المتداولة
    # افتراضيًا، جميع الأصول والمطلوبات
    assets = Account.objects.filter(account_type='asset', is_active=True)
    total_assets = Decimal('0')
    for account in assets:
        total_assets += account.get_balance(as_of_date)

    liabilities = Account.objects.filter(account_type='liability', is_active=True)
    total_liabilities = Decimal('0')
    for account in liabilities:
        total_liabilities += account.get_balance(as_of_date)

    if total_liabilities > 0:
        ratios['liquidity_ratio'] = total_assets / total_liabilities
    else:
        ratios['liquidity_ratio'] = Decimal('0')

    # نسبة الدين إلى رأس المال (Debt to Equity)
    equities = Account.objects.filter(account_type='equity', is_active=True)
    total_equity = Decimal('0')
    for account in equities:
        total_equity += account.get_balance(as_of_date)

    if total_equity > 0:
        ratios['debt_to_equity'] = total_liabilities / total_equity
    else:
        ratios['debt_to_equity'] = Decimal('0')

    # العائد على الأصول (Return on Assets)
    if total_assets > 0:
        ratios['return_on_assets'] = (net_profit / total_assets) * 100
    else:
        ratios['return_on_assets'] = Decimal('0')

    # تسجيل النشاط
    try:
        class ReportObj:
            id = 0
            pk = 0
            def __str__(self):
                return str(_('Financial Ratios Report'))
        log_view_activity(request, 'view', ReportObj(), str(_('Viewing financial ratios report')))
    except Exception:
        pass

    context = {
        'as_of_date': as_of_date,
        'ratios': ratios,
    }

    return render(request, 'reports/financial_ratios.html', context)


@login_required
def aging_report(request):
    """
    Accounts Receivable Aging Report: Display receivables classified by aging periods
    """
    # Permission gate: allow superuser or user_type admin/superadmin implicitly; else require explicit permission
    user = request.user
    has_perm = (
        getattr(user, 'is_superuser', False) or
        getattr(user, 'user_type', None) in ['superadmin', 'admin'] or
        user.has_reports_permission()
    )
    if not has_perm:
        raise PermissionDenied

    from sales.models import SalesInvoice
    from purchases.models import PurchaseInvoice
    from receipts.models import PaymentReceipt
    from payments.models import PaymentVoucher
    from datetime import date, timedelta

    # فلاتر
    customer_filter = request.GET.get('customer', '')
    supplier_filter = request.GET.get('supplier', '')
    account_filter = request.GET.get('account', '')
    start_date = _parse_date(request.GET.get('start_date'), date.today() - timedelta(days=365))
    end_date = _parse_date(request.GET.get('end_date'), date.today())

    # حساب المديونيات للعملاء
    customers_data = []
    customers = CustomerSupplier.objects.filter(type__in=['customer', 'both'], is_active=True)
    if customer_filter:
        customers = customers.filter(id=customer_filter)

    for customer in customers:
        # مجموع الفواتير الآجلة
        sales_invoices = SalesInvoice.objects.filter(
            customer=customer,
            payment_type='credit',
            date__lte=end_date
        ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')

        # مجموع المدفوعات
        payments = PaymentReceipt.objects.filter(
            customer=customer,
            date__lte=end_date,
            is_active=True,
            is_reversed=False
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        outstanding = sales_invoices - payments
        if outstanding <= 0:
            continue

        # تصنيف حسب الأيام
        aging_0_30 = Decimal('0')
        aging_31_60 = Decimal('0')
        aging_61_90 = Decimal('0')
        aging_over_90 = Decimal('0')

        # افتراضياً، الفواتير القديمة أولاً
        invoices = SalesInvoice.objects.filter(
            customer=customer,
            payment_type='credit',
            date__lte=end_date
        ).order_by('date')

        remaining_outstanding = outstanding
        for invoice in invoices:
            if remaining_outstanding <= 0:
                break
            days = (date.today() - invoice.date).days
            invoice_outstanding = min(remaining_outstanding, invoice.total_amount)
            if days <= 30:
                aging_0_30 += invoice_outstanding
            elif days <= 60:
                aging_31_60 += invoice_outstanding
            elif days <= 90:
                aging_61_90 += invoice_outstanding
            else:
                aging_over_90 += invoice_outstanding
            remaining_outstanding -= invoice_outstanding

        customers_data.append({
            'customer': customer,
            'outstanding': outstanding,
            'aging_0_30': aging_0_30,
            'aging_31_60': aging_31_60,
            'aging_61_90': aging_61_90,
            'aging_over_90': aging_over_90,
        })

    # حساب المديونيات للموردين
    suppliers_data = []
    suppliers = CustomerSupplier.objects.filter(type__in=['supplier', 'both'], is_active=True)
    if supplier_filter:
        suppliers = suppliers.filter(id=supplier_filter)

    for supplier in suppliers:
        # مجموع الفواتير الآجلة
        purchase_invoices = PurchaseInvoice.objects.filter(
            supplier=supplier,
            payment_type='credit',
            date__lte=end_date
        ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')

        # مجموع المدفوعات
        payments = PaymentVoucher.objects.filter(
            supplier=supplier,
            date__lte=end_date,
            is_active=True,
            is_reversed=False
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        outstanding = purchase_invoices - payments
        if outstanding <= 0:
            continue

        # تصنيف حسب الأيام
        aging_0_30 = Decimal('0')
        aging_31_60 = Decimal('0')
        aging_61_90 = Decimal('0')
        aging_over_90 = Decimal('0')

        invoices = PurchaseInvoice.objects.filter(
            supplier=supplier,
            payment_type='credit',
            date__lte=end_date
        ).order_by('date')

        remaining_outstanding = outstanding
        for invoice in invoices:
            if remaining_outstanding <= 0:
                break
            days = (date.today() - invoice.date).days
            invoice_outstanding = min(remaining_outstanding, invoice.total_amount)
            if days <= 30:
                aging_0_30 += invoice_outstanding
            elif days <= 60:
                aging_31_60 += invoice_outstanding
            elif days <= 90:
                aging_61_90 += invoice_outstanding
            else:
                aging_over_90 += invoice_outstanding
            remaining_outstanding -= invoice_outstanding

        suppliers_data.append({
            'supplier': supplier,
            'outstanding': outstanding,
            'aging_0_30': aging_0_30,
            'aging_31_60': aging_31_60,
            'aging_61_90': aging_61_90,
            'aging_over_90': aging_over_90,
        })

    # إجماليات
    total_customers_outstanding = sum(c['outstanding'] for c in customers_data)
    total_suppliers_outstanding = sum(s['outstanding'] for s in suppliers_data)

    # إجماليات الفترات للعملاء
    total_customers_aging_0_30 = sum(c['aging_0_30'] for c in customers_data)
    total_customers_aging_31_60 = sum(c['aging_31_60'] for c in customers_data)
    total_customers_aging_61_90 = sum(c['aging_61_90'] for c in customers_data)
    total_customers_aging_over_90 = sum(c['aging_over_90'] for c in customers_data)

    # إجماليات الفترات للموردين
    total_suppliers_aging_0_30 = sum(s['aging_0_30'] for s in suppliers_data)
    total_suppliers_aging_31_60 = sum(s['aging_31_60'] for s in suppliers_data)
    total_suppliers_aging_61_90 = sum(s['aging_61_90'] for s in suppliers_data)
    total_suppliers_aging_over_90 = sum(s['aging_over_90'] for s in suppliers_data)

    # تسجيل النشاط
    try:
        class ReportObj:
            id = 0
            pk = 0
            def __str__(self):
                return str(_('Accounts Receivable Aging Report'))
        log_view_activity(request, 'view', ReportObj(), str(_('Accounts Receivable Aging Report')))
    except Exception:
        pass

    context = {
        'customers_data': customers_data,
        'suppliers_data': suppliers_data,
        'total_customers_outstanding': total_customers_outstanding,
        'total_suppliers_outstanding': total_suppliers_outstanding,
        'total_customers_aging_0_30': total_customers_aging_0_30,
        'total_customers_aging_31_60': total_customers_aging_31_60,
        'total_customers_aging_61_90': total_customers_aging_61_90,
        'total_customers_aging_over_90': total_customers_aging_over_90,
        'total_suppliers_aging_0_30': total_suppliers_aging_0_30,
        'total_suppliers_aging_31_60': total_suppliers_aging_31_60,
        'total_suppliers_aging_61_90': total_suppliers_aging_61_90,
        'total_suppliers_aging_over_90': total_suppliers_aging_over_90,
        'customers': CustomerSupplier.objects.filter(type__in=['customer', 'both'], is_active=True),
        'suppliers': CustomerSupplier.objects.filter(type__in=['supplier', 'both'], is_active=True),
        'start_date': start_date,
        'end_date': end_date,
        'customer_filter': customer_filter,
        'supplier_filter': supplier_filter,
    }

    return render(request, 'reports/aging_report.html', context)


@login_required
def inventory_report(request):
    """
    تقرير أرصدة المخزون: عرض أرصدة المنتجات في المستودعات مع فلترة
    """
    # Permission gate: allow superuser or user_type admin/superadmin implicitly; else require explicit permission
    user = request.user
    has_perm = (
        getattr(user, 'is_superuser', False) or
        getattr(user, 'user_type', None) in ['superadmin', 'admin'] or
        user.has_reports_permission()
    )
    if not has_perm:
        raise PermissionDenied

    from inventory.models import Warehouse, InventoryMovement
    from products.models import Product
    from django.db.models import Sum

    # فلاتر
    warehouse_filter = request.GET.get('warehouse', '')
    product_filter = request.GET.get('product', '')
    as_of_date = _parse_date(request.GET.get('as_of_date'), date.today())

    # الحصول على المنتجات والمستودعات
    products = Product.objects.filter(is_active=True, product_type='physical')
    if product_filter:
        products = products.filter(id=product_filter)

    warehouses = Warehouse.objects.filter(is_active=True)
    if warehouse_filter:
        warehouses = warehouses.filter(id=warehouse_filter)

    # حساب الأرصدة
    inventory_data = []
    for product in products:
        product_data = {
            'product': product,
            'warehouses': {},
            'total_stock': 0,
            'total_value': 0,
        }
        
        for warehouse in warehouses:
            # حساب الرصيد في المستودع حتى التاريخ المحدد
            incoming = InventoryMovement.objects.filter(
                product=product,
                warehouse=warehouse,
                movement_type='in',
                date__lte=as_of_date
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            outgoing = InventoryMovement.objects.filter(
                product=product,
                warehouse=warehouse,
                movement_type='out',
                date__lte=as_of_date
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            stock = incoming - outgoing
            value = stock * product.calculate_weighted_average_cost()
            
            if stock != 0:
                product_data['warehouses'][warehouse.name] = {
                    'stock': stock,
                    'value': value,
                }
                product_data['total_stock'] += stock
                product_data['total_value'] += value
        
        if product_data['total_stock'] != 0:
            inventory_data.append(product_data)

    # إجماليات
    total_value = sum(p['total_value'] for p in inventory_data)
    total_products = len(inventory_data)

    # تسجيل النشاط
    try:
        class ReportObj:
            id = 0
            pk = 0
            def __str__(self):
                return str(_('Inventory Balances Report'))
        log_view_activity(request, 'view', ReportObj(), str(_('View Inventory Ralance Report')))
    except Exception:
        pass

    context = {
        'inventory_data': inventory_data,
        'warehouses': Warehouse.objects.filter(is_active=True),
        'products': Product.objects.filter(is_active=True, product_type='physical'),
        'as_of_date': as_of_date,
        'warehouse_filter': warehouse_filter,
        'product_filter': product_filter,
        'total_value': total_value,
        'total_products': total_products,
    }

    return render(request, 'reports/inventory_report.html', context)
