from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, TemplateView
from django.http import HttpResponse, JsonResponse
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.db import transaction, IntegrityError
from django.utils import timezone
from decimal import Decimal, ROUND_HALF_UP
from datetime import date, datetime
import json
from .models import SalesInvoice, SalesInvoiceItem, SalesReturn, SalesReturnItem
from .models import SalesCreditNote
from products.models import Product
from customers.models import CustomerSupplier
from settings.models import CompanySettings, Currency
from core.models import DocumentSequence
import openpyxl
from core.models import AuditLog
from django.views.generic import TemplateView
import time


def get_product_stock_in_warehouse(product, warehouse):
    """
    حساب المخزون المتوفر للمنتج في مستودع معين
    """
    try:
        from inventory.models import InventoryMovement
        incoming = InventoryMovement.objects.filter(
            product=product,
            warehouse=warehouse,
            movement_type='in'
        ).aggregate(total=Sum('quantity'))['total'] or Decimal('0')
        
        outgoing = InventoryMovement.objects.filter(
            product=product,
            warehouse=warehouse,
            movement_type='out'
        ).aggregate(total=Sum('quantity'))['total'] or Decimal('0')
        
        return incoming - outgoing
    except ImportError:
        # في حالة عدم وجود نموذج المخزون
        return product.current_stock if hasattr(product, 'current_stock') else Decimal('0')


def get_or_create_user_pos_cashboxes(user):
    """
    الحصول على صناديق نقطة البيع للمستخدم (النقد والبطاقة).
    ملاحظة مهمة: الصناديق تُنشأ تلقائياً عند إنشاء مستخدم POS في users/signals.py
    هذه الدالة تبحث فقط عن الصناديق الموجودة ولا تنشئ صناديق إضافية.
    """
    from cashboxes.models import Cashbox
    from users.signals import _build_pos_cashbox_names
    
    # الحصول على أسماء الصناديق الموحدة من users/signals
    expected_cash_name, expected_card_name = _build_pos_cashbox_names(user.username)
    
    # البحث عن الصناديق بناءً على الأسماء الموحدة
    cash_cashbox = Cashbox.objects.filter(
        responsible_user=user,
        is_active=True,
        name__iexact=expected_cash_name
    ).first()

    card_cashbox = Cashbox.objects.filter(
        responsible_user=user,
        is_active=True,
        name__iexact=expected_card_name
    ).first()

    return cash_cashbox, card_cashbox


def get_product_stock(request, product_id, warehouse_id):
    """
    API endpoint للحصول على مخزون المنتج في مستودع معين
    """
    try:
        product = get_object_or_404(Product, id=product_id)
        from inventory.models import Warehouse
        warehouse = get_object_or_404(Warehouse, id=warehouse_id)
        stock = get_product_stock_in_warehouse(product, warehouse)
        
        # التعديل: إرسال is_service مع المخزون
        return JsonResponse({
            'stock': str(stock),
            'is_service': product.is_service # <-- هذا هو التعديل المضاف
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

from core.models import DocumentSequence
from accounts.services import create_sales_invoice_transaction, create_sales_return_transaction, delete_transaction_by_reference
from journal.services import JournalService
import json
import logging
from django.utils.translation import gettext_lazy as _


def create_sales_invoice_journal_entry(invoice, user):
    """Create journal entry for sales invoice"""
    try:
        # إنشاء قيد الإيراد (المبيعات)
        sales_entry = JournalService.create_sales_invoice_entry(invoice, user)
        if sales_entry:
            print(f"✅ تم إنشاء قيد المبيعات {sales_entry.entry_number} للفاتورة {invoice.invoice_number}")
        else:
            print(f"⚠️ لم يتم إنشاء قيد المبيعات للفاتورة {invoice.invoice_number}")
        
        # إنشاء قيد تكلفة البضاعة المباعة (COGS)
        # يتم إنشاؤه هنا مرة واحدة فقط بعد حفظ الفاتورة وعناصرها
        cogs_entry = JournalService.create_cogs_entry(invoice, user)
        if cogs_entry:
            print(f"✅ تم إنشاء قيد COGS {cogs_entry.entry_number} للفاتورة {invoice.invoice_number}")
        else:
            print(f"⚠️ لم يتم إنشاء قيد COGS للفاتورة {invoice.invoice_number}")
            
    except Exception as e:
        print(f"❌ خطأ في إنشاء القيد المحاسبي لفاتورة المبيعات {invoice.invoice_number}: {e}")
        import traceback
        traceback.print_exc()
        # لا نوقف العملية في حالة فشل إنشاء القيد المحاسبي
        # لكن نسجل الخطأ بشكل واضح

def create_sales_return_journal_entry(sales_return, user):
    """Create journal entries for sales return"""
    try:
        # القيد الأول: عكس قيد الإيراد
        JournalService.create_sales_return_entry(sales_return, user)
        
        # القيد الثاني: عكس قيد COGS (إرجاع البضاعة للمخزون)
        JournalService.create_sales_return_cogs_entry(sales_return, user)
    except Exception as e:
        print(f"خطأ في إنشاء القيود المحاسبية لمردود المبيعات: {e}")
        # لا نوقف العملية في حالة فشل إنشاء القيد المحاسبي
        pass


def create_sales_invoice_account_transaction(invoice, user):
    """Create customer account transaction when creating sales invoice"""
    try:
        from accounts.models import AccountTransaction
        import uuid
        
        # إذا كانت طريقة الدفع ليست نقداً، نسجل حركة في حساب العميل
        if invoice.payment_type != 'cash' and invoice.customer:
            # توليد رقم الحركة
            transaction_number = f"SALE-{uuid.uuid4().hex[:8].upper()}"
            
            # حساب الرصيد السابق
            last_transaction = AccountTransaction.objects.filter(
                customer_supplier=invoice.customer
            ).order_by('-created_at').first()
            
            previous_balance = last_transaction.balance_after if last_transaction else Decimal('0')
            
            # إنشاء حركة مدينة للعميل (زيادة الذمم المدينة)
            new_balance = previous_balance + invoice.total_amount
            
            AccountTransaction.objects.create(
                transaction_number=transaction_number,
                date=invoice.date,
                customer_supplier=invoice.customer,
                transaction_type='sales_invoice',
                direction='debit',
                amount=invoice.total_amount,
                reference_type='sales_invoice',
                reference_id=invoice.id,
                description=f'مبيعات - فاتورة رقم {invoice.invoice_number}',
                balance_after=new_balance,
                created_by=user
            )
            
    except ImportError:
        # في حالة عدم وجود نموذج الحسابات
        pass
    except Exception as e:
        print(f"خطأ في إنشاء حركة الحساب: {e}")
        # لا نوقف العملية في حالة فشل تسجيل الحركة المالية
        pass


def create_sales_return_account_transaction(return_invoice, user):
    """Create customer account transaction when creating sales return"""
    try:
        from accounts.models import AccountTransaction
        import uuid
        
        # إذا كان هناك عميل، نسجل حركة دائنة (تقليل الذمم المدينة)
        if return_invoice.customer:
            # توليد رقم الحركة
            transaction_number = f"SRET-{uuid.uuid4().hex[:8].upper()}"
            
            # حساب الرصيد السابق
            last_transaction = AccountTransaction.objects.filter(
                customer_supplier=return_invoice.customer
            ).order_by('-created_at').first()
            
            previous_balance = last_transaction.balance_after if last_transaction else Decimal('0')
            
            # إنشاء حركة دائنة للعميل (تقليل الذمم المدينة)
            new_balance = previous_balance - return_invoice.total_amount
            
            AccountTransaction.objects.create(
                transaction_number=transaction_number,
                date=return_invoice.date,
                customer_supplier=return_invoice.customer,
                transaction_type='sales_return',
                direction='credit',
                amount=return_invoice.total_amount,
                reference_type='sales_return',
                reference_id=return_invoice.id,
                description=f'مردود مبيعات - فاتورة رقم {return_invoice.return_number}',
                balance_after=new_balance,
                created_by=user
            )
            
    except ImportError:
        # في حالة عدم وجود نموذج الحسابات
        pass
    except Exception as e:
        print(f"خطأ في إنشاء حركة الحساب: {e}")
        # لا نوقف العملية في حالة فشل تسجيل الحركة المالية
        pass


def create_sales_return_inventory_movements(return_invoice, user):
    """Create inventory movements for sales return"""
    try:
        from inventory.models import InventoryMovement, Warehouse
        import uuid
        
        # الحصول على المستودع الافتراضي
        default_warehouse = Warehouse.objects.filter(is_active=True).first()
        if not default_warehouse:
            # إنشاء مستودع افتراضي إذا لم يكن موجوداً
            default_warehouse = Warehouse.objects.create(
                name='المستودع الرئيسي',
                code='MAIN',
                is_active=True
            )
        
        # إنشاء حركة مخزون واردة لكل عنصر في المردود (فقط للكميات الأكبر من صفر)
        for item in return_invoice.items.all():
            # تخطي المنتجات ذات الكمية صفر أو السالبة
            if item.quantity <= 0:
                continue
                
            # توليد رقم الحركة
            movement_number = f"RETURN-IN-{uuid.uuid4().hex[:8].upper()}"
            
            InventoryMovement.objects.create(
                movement_number=movement_number,
                date=return_invoice.date,
                product=item.product,
                warehouse=default_warehouse,
                movement_type='in',
                reference_type='sales_return',
                reference_id=return_invoice.id,
                quantity=item.quantity,
                unit_cost=item.unit_price,
                notes=f'مردود مبيعات - رقم {return_invoice.return_number}',
                created_by=user
            )
            
    except ImportError:
        # في حالة عدم وجود نموذج المخزون
        pass
    except Exception as e:
        print(f"خطأ في إنشاء حركة المخزون: {e}")
        # لا نوقف العملية في حالة فشل تسجيل الحركة المخزونية
        pass


class SalesInvoiceListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = SalesInvoice
    template_name = 'sales/invoice_list.html'
    context_object_name = 'invoices'
    paginate_by = 10
    
    def test_func(self):
        return self.request.user.has_sales_permission()
    
    def get_queryset(self):
        queryset = SalesInvoice.objects.all()
        
        # Search functionality
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(invoice_number__icontains=search) |
                Q(customer__name__icontains=search)
            )
        
        # Date filter
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        
        # Payment type filter
        payment_type = self.request.GET.get('payment_type')
        if payment_type:
            queryset = queryset.filter(payment_type=payment_type)
        
        # Cashbox filter (للفواتير النقدية فقط)
        cashbox_id = self.request.GET.get('cashbox')
        if cashbox_id:
            queryset = queryset.filter(cashbox_id=cashbox_id)
        
        # Apply ordering
        order_by = self.request.GET.get('order_by', '-date')
        if order_by.startswith('-'):
            queryset = queryset.order_by(order_by, '-id')
        else:
            queryset = queryset.order_by(order_by, 'id')
        
        return queryset.select_related('customer', 'created_by', 'cashbox')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Statistics
        invoices = SalesInvoice.objects.all()
        context['total_invoices'] = invoices.count()
        context['paid_invoices'] = invoices.filter(payment_type='cash').count()
        context['pending_invoices'] = invoices.filter(payment_type='credit').count()
        context['total_sales_amount'] = invoices.aggregate(
            total=Sum('total_amount')
        )['total'] or 0
        
        # This month's invoices
        current_month = timezone.now().replace(day=1)
        month_invoices = invoices.filter(date__gte=current_month).count()
        context['month_invoices'] = month_invoices
        
        # Active customers (customers with invoices)
        active_customers = CustomerSupplier.objects.filter(
            Q(type='customer') | Q(type='both'),
            salesinvoice__isnull=False
        ).distinct().count()
        context['active_customers'] = active_customers
        
        # Cashboxes list for filter
        try:
            from cashboxes.models import Cashbox
            context['cashboxes'] = Cashbox.objects.filter(is_active=True).order_by('name')
        except ImportError:
            context['cashboxes'] = []
        
        # Currency and company settings
        try:
            company_settings = CompanySettings.objects.first()
            if company_settings and company_settings.base_currency:
                context['base_currency'] = company_settings.base_currency
            else:
                context['base_currency'] = Currency.objects.filter(is_active=True).first()
        except:
            pass
        
        context['active_currencies'] = Currency.objects.filter(is_active=True)
        
        # Current ordering
        context['current_order'] = self.request.GET.get('order_by', '-date')
        
        return context


@login_required
def export_invoices_to_xlsx(request):
    """Export invoice list to Excel file"""
    from core.utils import get_client_ip
    
    # بناء queryset بنفس التصفية المستخدمة في SalesInvoiceListView
    queryset = SalesInvoice.objects.all()
    
    # Search functionality
    search = request.GET.get('search')
    if search:
        queryset = queryset.filter(
            Q(invoice_number__icontains=search) |
            Q(customer__name__icontains=search)
        )
    
    # Date filter
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)
    
    # Payment type filter
    payment_type = request.GET.get('payment_type')
    if payment_type:
        queryset = queryset.filter(payment_type=payment_type)
    
    # Cashbox filter
    cashbox_id = request.GET.get('cashbox')
    if cashbox_id:
        queryset = queryset.filter(cashbox_id=cashbox_id)
    
    # Apply ordering
    order_by = request.GET.get('order_by', '-date')
    if order_by.startswith('-'):
        queryset = queryset.order_by(order_by, '-id')
    else:
        queryset = queryset.order_by(order_by, 'id')
    
    # تحميل البيانات المرتبطة
    invoices = queryset.select_related('customer', 'created_by', 'cashbox')
    
    # إنشاء ملف Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "فواتير المبيعات"
    
    # إضافة العناوين
    headers = [
        'رقم الفاتورة',
        'التاريخ',
        'العميل',
        'نوع الدفع',
        'المبلغ الإجمالي',
        'الحالة',
        'الصندوق',
        'أنشأ بواسطة',
        'تاريخ الإنشاء'
    ]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # إضافة البيانات
    for row_num, invoice in enumerate(invoices, 2):
        ws.cell(row=row_num, column=1, value=invoice.invoice_number)
        ws.cell(row=row_num, column=2, value=invoice.date.strftime('%Y-%m-%d') if invoice.date else '')
        ws.cell(row=row_num, column=3, value=invoice.customer.name if invoice.customer else '')
        ws.cell(row=row_num, column=4, value=invoice.get_payment_type_display())
        ws.cell(row=row_num, column=5, value=float(invoice.total_amount))
        ws.cell(row=row_num, column=6, value=invoice.get_status_display())
        ws.cell(row=row_num, column=7, value=invoice.cashbox.name if invoice.cashbox else '')
        ws.cell(row=row_num, column=8, value=invoice.created_by.get_full_name() if invoice.created_by else '')
        ws.cell(row=row_num, column=9, value=invoice.created_at.strftime('%Y-%m-%d %H:%M') if invoice.created_at else '')
    
    # إنشاء الاستجابة
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="فواتير_المبيعات.xlsx"'
    
    # حفظ الملف
    wb.save(response)
    
    # تسجيل في سجل الأنشطة
    AuditLog.objects.create(
        user=request.user,
        action_type='export',
        model_name='SalesInvoice',
        object_id=None,  # تصدير قائمة
        description=f'تصدير قائمة الفواتير إلى Excel ({invoices.count()} فاتورة)',
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', ''),
        changes={}
    )
    
    return response


@login_required
def sales_invoice_create(request):
    """Create new sales invoice"""
    
    def get_invoice_create_context(request, form_data=None):
        """Prepare invoice creation page context with entered data if any"""
        user = request.user
        context = {
            'customers': CustomerSupplier.objects.filter(type__in=['customer', 'both']),
            'today_date': date.today().isoformat(),
        }
        
        # إضافة البيانات المُدخلة إذا كانت موجودة
        if form_data:
            context.update(form_data)
            # إنشاء قائمة المنتجات الموجودة لعرضها في القالب
            existing_products = []
            products = form_data.get('products', [])
            quantities = form_data.get('quantities', [])
            prices = form_data.get('prices', [])
            tax_rates = form_data.get('tax_rates', [])
            for i, product_id in enumerate(products):
                if product_id:
                    try:
                        product = Product.objects.get(id=product_id)
                        existing_products.append({
                            'id': product.id,
                            'code': product.code,
                            'name': product.name,
                            'quantity': quantities[i] if i < len(quantities) else '',
                            'price': prices[i] if i < len(prices) else '',
                            'tax_rate': tax_rates[i] if i < len(tax_rates) else '',
                            'current_stock': product.current_stock,
                            'sale_price': product.sale_price,
                        })
                    except Product.DoesNotExist:
                        pass
            context['existing_products'] = existing_products
            context['form_data_json'] = json.dumps(form_data)
        
        # إضافة المستودعات
        try:
            from inventory.models import Warehouse
            context['warehouses'] = Warehouse.objects.filter(is_active=True)
            # إضافة المستودع الافتراضي للمستخدم
            try:
                context['default_warehouse'] = user.default_sales_warehouse
            except AttributeError:
                context['default_warehouse'] = None
        except ImportError:
            context['warehouses'] = []
            context['default_warehouse'] = None

        # إضافة الصناديق النقدية
        try:
            from cashboxes.models import Cashbox
            context['cashboxes'] = Cashbox.objects.filter(is_active=True)
            
            # تحديد الصندوق الافتراضي حسب نوع المستخدم
            try:
                if user.has_perm('users.can_access_pos'):
                    # مستخدم POS: استخدام الصندوق المرتبط به (responsible_user)
                    pos_cashbox = Cashbox.objects.filter(responsible_user=user, is_active=True).first()
                    context['default_cashbox'] = pos_cashbox or user.default_cashbox
                else:
                    # مستخدم عادي: استخدام الصندوق الافتراضي المحفوظ
                    context['default_cashbox'] = user.default_cashbox
            except AttributeError:
                context['default_cashbox'] = None
        except ImportError:
            context['cashboxes'] = []
            context['default_cashbox'] = None
        
        products = Product.objects.filter(is_active=True).select_related('category')
        products_data = []
        for product in products:
            products_data.append({
                'id': product.id,
                'code': product.code,
                'name': product.name,
                'price': float(product.sale_price),
                'tax_rate': float(product.tax_rate),
                'category': product.category.name if product.category else ''
            })

        context['products_json'] = json.dumps(products_data)
        context['products'] = products  # إضافة المنتجات للـ modal
        
        # إضافة الفئات للـ modal
        try:
            from products.models import Category
            context['categories'] = Category.objects.filter(is_active=True).order_by('name')
        except ImportError:
            context['categories'] = []

        # التحقق من صلاحيات المستخدم
        context['can_edit_invoice_number'] = user.is_superuser or user.is_staff
        context['can_edit_date'] = user.is_superuser or user.is_staff
        
        # التحقق من صلاحية الوصول لنقطة البيع
        context['has_pos_access'] = user.has_perm('users.can_access_pos')
        # إذا كان المستخدم يمكنه الوصول لنقطة البيع، لا يمكنه تغيير الصندوق
        context['can_change_cashbox'] = not user.has_perm('users.can_access_pos')

        # اسم المستخدم المنشئ (الاسم الأول + الاسم الأخير) لعرضه في القالب
        try:
            creator_full = f"{user.first_name or ''} {user.last_name or ''}".strip()
        except Exception:
            creator_full = user.username
        context['creator_full_name'] = creator_full
        # صلاحية تغيير منشئ الفاتورة وقائمة المستخدمين
        try:
            can_change_creator = user.is_superuser
        except Exception:
            can_change_creator = False
        context['can_change_creator'] = can_change_creator
        if can_change_creator:
            try:
                from django.contrib.auth import get_user_model
                U = get_user_model()
                context['all_users'] = U.objects.all()
            except Exception:
                context['all_users'] = []

        # إضافة رقم الفاتورة المتوقع للعرض
        try:
            sequence = DocumentSequence.objects.get(document_type='sales_invoice')
            # استخدم معاينة الرقم التالي لتعكس أعلى رقم مستخدم فعلياً
            if hasattr(sequence, 'peek_next_number'):
                context['next_invoice_number'] = sequence.peek_next_number()
            else:
                context['next_invoice_number'] = sequence.get_formatted_number()
        except DocumentSequence.DoesNotExist:
            # إضافة رسالة خطأ بدلاً من إنشاء رقم تلقائي
            context['sequence_error'] = _('Sales invoice sequence not configured. Please configure it in settings first.')
            context['next_invoice_number'] = None

        # Currency settings
        try:
            company_settings = CompanySettings.objects.first()
            if company_settings and company_settings.base_currency:
                context['base_currency'] = company_settings.base_currency
            else:
                context['base_currency'] = Currency.objects.filter(is_active=True).first()
        except:
            pass

        return context
    
    if request.method == 'POST':
        # جمع جميع البيانات المُدخلة لإعادة عرضها في حالة الأخطاء
        form_data = {
            'customer_id': request.POST.get('customer'),
            'warehouse_id': request.POST.get('warehouse'),
            'payment_type': request.POST.get('payment_type'),
            'cashbox_id': request.POST.get('cashbox'),
            'notes': request.POST.get('notes', ''),
            'discount_amount': request.POST.get('discount', '0'),
            'manual_invoice': request.POST.get('invoice_number'),
            'invoice_date': request.POST.get('date', date.today().isoformat()),
            'inclusive_tax': 'inclusive_tax' in request.POST,
            'creator_user_id': request.POST.get('creator_user'),
            'set_default_warehouse': request.POST.get('set_default_warehouse') == 'on',
            'set_default_cashbox': request.POST.get('set_default_cashbox') == 'on',
            'products': request.POST.getlist('products[]'),
            'quantities': request.POST.getlist('quantities[]'),
            'prices': request.POST.getlist('prices[]'),
            'tax_amounts': request.POST.getlist('tax_amounts[]'),
        }
        
        try:
            # helper to robustly parse decimals from various client locales (commas, arabic separators, spaces)
            def parse_decimal_input(val, name='value', default=Decimal('0')):
                try:
                    if val is None or val == '':
                        return default
                    s = str(val).strip()
                    # Arabic decimal separators and common comma thousands
                    s = s.replace('\u066b', '.')  # Arabic decimal separator if present
                    s = s.replace('\u066c', ',')  # Arabic thousands separator if present
                    # Replace comma with dot for decimal point, remove spaces
                    s = s.replace(',', '.')
                    s = s.replace(' ', '')
                    return Decimal(s)
                except Exception:
                    # Log parsing error in AuditLog for visibility
                    try:
                        from core.signals import log_user_activity
                        dummy = SalesInvoice()
                        log_user_activity(request, 'error', dummy, _('Failed to parse numeric value for field %(name)s: %(val)s') % {'name': name, 'val': val})
                    except Exception:
                        pass
                    from decimal import Decimal as DecimalClass
                    return DecimalClass('0')
            # سنحاول عدة مرات لتجنب تعارض الأرقام في حال السباق
            max_attempts = 5
            attempt = 0
            allow_manual = True
            while attempt < max_attempts:
                attempt += 1
                try:
                    with transaction.atomic():
                        user = request.user
                        customer_id = request.POST.get('customer')
                        warehouse_id = request.POST.get('warehouse')
                        payment_type = request.POST.get('payment_type')
                        # الحصول على الصندوق فقط للدفع النقدي (الشيكات تُعالج من خلال سند القبض)
                        #cashbox_id = request.POST.get('cashbox') if payment_type == 'cash' else None
                        # ✅ السماح للصندوق بالتعيين إذا كان الدفع نقداً أو بطاقة (لكي تتمكن البطاقة من حمل صندوقها وبناء قيدها)
                        if payment_type in ['cash', 'card']:
                            cashbox_id = request.POST.get('cashbox')
                        else:
                            cashbox_id = None
                        set_default_cashbox = request.POST.get('set_default_cashbox') == 'on' and payment_type == 'cash'
                        notes = request.POST.get('notes', '')
                        discount_amount = Decimal(request.POST.get('discount', '0'))

                        # التحقق من صلاحية تعديل رقم الفاتورة
                        manual_invoice = request.POST.get('invoice_number') if allow_manual else None
                        if manual_invoice and (user.is_superuser or user.is_staff):
                            invoice_number = manual_invoice
                        else:
                            invoice_number = None

                        # توليد رقم الفاتورة إذا لم يكن محدد
                        if not invoice_number:
                            try:
                                sequence = DocumentSequence.objects.get(document_type='sales_invoice')
                                invoice_number = sequence.get_next_number()
                            except DocumentSequence.DoesNotExist:
                                # إضافة رسالة خطأ بدلاً من إنشاء رقم تلقائي
                                errors.append(_('Sales invoice sequence not configured. Please configure it in settings first.'))

                        # التحقق من صلاحية تعديل التاريخ
                        if user.is_superuser or user.is_staff:
                            invoice_date = request.POST.get('date', date.today())
                        else:
                            invoice_date = date.today()

                        # التحقق من البيانات المطلوبة
                        errors = []
                        if not customer_id:
                            errors.append(_('Please select a customer'))
                        if not payment_type:
                            errors.append(_('Please select a payment method'))
                        if not warehouse_id:
                            errors.append(_('Please select a warehouse'))
                        
                        if errors:
                            # سجل محاولة فاشلة في سجل النشاط لتتبع أخطاء الإدخال
                            try:
                                from core.signals import log_user_activity
                                dummy = SalesInvoice()
                                log_user_activity(
                                    request,
                                    'error',
                                    dummy,
                                    _('Failed to create invoice: Required fields are missing')
                                )
                            except Exception:
                                pass

                            for error in errors:
                                messages.error(request, error)
                            # جمع البيانات المُدخلة لإعادة عرضها
                            form_data = {
                                'customer_id': customer_id,
                                'warehouse_id': warehouse_id,
                                'payment_type': payment_type,
                                'cashbox_id': cashbox_id,
                                'notes': notes,
                                'discount_amount': str(discount_amount) if discount_amount else '',
                                'manual_invoice': manual_invoice,
                                'invoice_date': invoice_date.isoformat() if hasattr(invoice_date, 'isoformat') else invoice_date,
                            }
                            context = get_invoice_create_context(request, form_data)
                            return render(request, 'sales/invoice_add.html', context)

                        customer = get_object_or_404(CustomerSupplier, id=customer_id)

                        # الحصول على المستودع إذا تم تحديده
                        warehouse = None
                        if warehouse_id:
                            try:
                                from inventory.models import Warehouse
                                warehouse = Warehouse.objects.get(id=warehouse_id)
                            except (ImportError, Warehouse.DoesNotExist):
                                warehouse = None

                        # الحصول على الصندوق فقط للدفع النقدي (الشيكات تُعالج من خلال سند القبض)
                        cashbox = None
                        #if payment_type == 'cash':
                        if payment_type in ['cash', 'card']:
                            # 🔧 إعطاء الأولوية للصندوق المُختار من المستخدم
                            if cashbox_id:
                                try:
                                    from cashboxes.models import Cashbox
                                    cashbox = Cashbox.objects.get(id=cashbox_id, is_active=True)
                                except (ImportError, Cashbox.DoesNotExist):
                                    messages.error(request, _('The selected cashbox does not exist or is not active'))
                                    context = get_invoice_create_context(request, form_data)
                                    return render(request, 'sales/invoice_add.html', context)
                            # إذا لم يتم اختيار صندوق، استخدم صندوق المستخدم التلقائي (POS)
                            elif user.has_perm('users.can_access_pos'):
                                try:
                                    from cashboxes.models import Cashbox
                                    cashbox = Cashbox.objects.filter(responsible_user=user).first()
                                    if not cashbox:
                                        # إذا لم يكن له صندوق، سيتم إنشاؤه تلقائياً في signals.py
                                        pass
                                except ImportError:
                                    pass
                        products = request.POST.getlist('products[]')
                        quantities = request.POST.getlist('quantities[]')
                        prices = request.POST.getlist('prices[]')
                        tax_amounts = request.POST.getlist('tax_amounts[]')

                        # التحقق من وجود منتجات
                        if not products or not any(p for p in products if p):
                            # سجل محاولة فاشلة في سجل النشاط لتتبع أخطاء الإدخال
                            try:
                                from core.signals import log_user_activity
                                dummy = SalesInvoice()
                                log_user_activity(
                                    request,
                                    'error',
                                    dummy,
                                    _('Failed to create invoice: No items added')
                                )
                            except Exception:
                                pass

                            messages.error(request, _('Please add at least one product'))
                            context = get_invoice_create_context(request, form_data)
                            return render(request, 'sales/invoice_add.html', context)

                        # حساب المجاميع أولاً قبل إنشاء أي شيء (للتحقق من الحد الائتماني)
                        subtotal = Decimal('0')
                        total_tax_amount = Decimal('0')
                        
                        # determine inclusive_tax flag
                        inclusive_tax_flag = 'inclusive_tax' in request.POST

                        # حساب المجاميع المؤقتة
                        stock_warnings = {}  # قاموس التحذيرات للكميات الزائدة
                        for i, product_id in enumerate(products):
                            if product_id and i < len(quantities) and i < len(prices):
                                try:
                                    product = Product.objects.get(id=product_id)
                                    quantity = parse_decimal_input(quantities[i], name='quantity', default=Decimal('0'))
                                    unit_price = parse_decimal_input(prices[i], name='price', default=Decimal('0'))
                                    tax_rate = product.tax_rate or Decimal('0')

                                    if quantity <= 0 or unit_price < 0:
                                        continue

                                    # فحص المخزون في المستودع المختار أو المخزون العام
                                    try:
                                        if warehouse:
                                            available_stock = get_product_stock_in_warehouse(product, warehouse)
                                        else:
                                            available_stock = product.current_stock
                                    except Exception as e:
                                        # في حالة فشل حساب المخزون، استخدم المخزون العام
                                        available_stock = product.current_stock
                                        print(f"خطأ في حساب المخزون للمنتج {product.name}: {e}")

                                    # تحذير إذا المنتج غير متوفر في المستودع
                                    # تم التعديل: إضافة شرط التحقق من النوع (ليس خدمة)
                                    if not product.is_service:
                                        if available_stock <= 0:
                                            stock_warnings[str(product.id)] = _('Warning: Product "%(product)s" is not available in the selected warehouse.') % {'product': product.name}
                                        elif quantity > available_stock:
                                            stock_warnings[str(product.id)] = _('Warning: Requested quantity (%(quantity)s) exceeds available stock (%(available)s) for product "%(product)s" in the selected warehouse.') % {
                                                'quantity': quantity,
                                                'available': available_stock,
                                                'product': product.name
                                            }

                                    line_subtotal = quantity * unit_price
                                    line_tax_amount = line_subtotal * (tax_rate / 100) if tax_rate > 0 else Decimal('0')
                                    
                                    subtotal += line_subtotal
                                    total_tax_amount += line_tax_amount
                                except (Product.DoesNotExist, ValueError, TypeError):
                                    continue

                        # حساب الإجمالي النهائي
                        if inclusive_tax_flag:
                            final_total = (subtotal + total_tax_amount - discount_amount).quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
                        else:
                            final_total = (subtotal - discount_amount).quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)

                        # فحص الحد الائتماني قبل إنشاء أي شيء
                        if payment_type == 'credit' and customer and customer.credit_limit > 0:
                            current_balance = customer.current_balance
                            available_credit = customer.credit_limit - abs(current_balance) if current_balance < 0 else customer.credit_limit
                            
                            if final_total > available_credit:
                                # رسالة تحذير مع اقتراحات
                                error_message = _(
                                    'لا يمكن إنشاء الفاتورة لأن المبلغ الإجمالي (%(total)s) يتجاوز الحد الائتماني المتاح للعميل (%(available)s).\n\nاقتراحات:\n1. زيادة الحد الائتماني للعميل\n2. تحصيل المستحقات المتأخرة من العميل\n3. تحويل الفاتورة إلى دفع نقدي'
                                ) % {
                                    'total': f"{final_total:.3f}",
                                    'available': f"{available_credit:.3f}"
                                }
                                
                                # التأكد من عدم وجود رسائل سابقة من نفس النوع
                                existing_messages = [msg for msg in messages.get_messages(request) if msg.message == error_message]
                                if not existing_messages:
                                    messages.warning(request, error_message)
                                
                                # تسجيل المحاولة الفاشلة في سجل الأنشطة
                                try:
                                    from core.signals import log_user_activity
                                    log_user_activity(
                                        request,
                                        'error',
                                        customer,
                                        _('Failed to create sales invoice: Credit limit exceeded - Amount %(total)s > Available %(available)s') % {
                                            'total': f"{final_total:.3f}",
                                            'available': f"{available_credit:.3f}"
                                        }
                                    )
                                except Exception:
                                    pass
                                
                                context = get_invoice_create_context(request, form_data)
                                return render(request, 'sales/invoice_add.html', context)

                        # فحص تحذيرات المخزون
                        skip_stock_check = request.POST.get('skip_stock_check') == 'true'
                        if stock_warnings and not skip_stock_check:
                            # إضافة التحذيرات إلى الرسائل
                            for warning_msg in stock_warnings.values():
                                messages.warning(request, warning_msg)
                            
                            # إضافة التحذيرات إلى form_data لعرضها في القالب
                            form_data['stock_warnings'] = stock_warnings
                            form_data['show_skip_button'] = True
                            
                            # تسجيل التحذير في سجل الأنشطة
                            try:
                                from core.signals import log_user_activity
                                dummy = SalesInvoice()
                                log_user_activity(
                                    request,
                                    'warning',
                                    dummy,
                                    _('Warning: Requested quantities exceed available stock in %(count)s products') % {'count': len(stock_warnings)}
                                )
                            except Exception:
                                pass
                            
                            context = get_invoice_create_context(request, form_data)
                            return render(request, 'sales/invoice_add.html', context)

                        # إذا وصلنا هنا، الحد الائتماني مسموح والمخزون مسموح أو تم تخطيه - يمكننا إنشاء الفاتورة
                        # إعادة تعيين المجاميع لإعادة الحساب أثناء الإنشاء الفعلي
                        subtotal = Decimal('0')
                        total_tax_amount = Decimal('0')

                        # إنشاء الفاتورة
                        invoice = SalesInvoice.objects.create(
                            invoice_number=invoice_number,
                            date=invoice_date,
                            customer=customer,
                            warehouse=warehouse,
                            payment_type=payment_type,
                            cashbox=cashbox,
                            discount_amount=discount_amount,
                            notes=notes,
                            created_by=user,
                            inclusive_tax=inclusive_tax_flag,
                            subtotal=0,  # سيتم تحديثها لاحقاً
                            tax_amount=0,  # سيتم تحديثها لاحقاً
                            total_amount=0  # سيتم تحديثها لاحقاً
                        )

                        # --- بداية التعديل (الوسط) ---
                        try:
                            from journal.services import JournalService
                            JournalService.create_sales_invoice_entry(invoice, user)
                        except Exception as e:
                            print(f"Error creating journal: {e}")
                        # --- نهاية التعديل ---

                        # إذا كان المستخدم له صلاحية تغيير منشئ الفاتورة، تحقق من وجود حقل creator_user_id
                        try:
                            if user.is_superuser:
                                creator_user_id = request.POST.get('creator_user')
                                if creator_user_id:
                                    from django.contrib.auth import get_user_model
                                    U = get_user_model()
                                    try:
                                        chosen = U.objects.get(id=creator_user_id)
                                        invoice.created_by = chosen
                                        invoice.save()

                                        # سجل التغيير في السجل
                                        from core.signals import log_user_activity
                                        log_user_activity(
                                            request,
                                            'update',
                                            invoice,
                                            _('Invoice creator changed to %(name)s by %(user)s') % {
                                                'name': f"{chosen.first_name or ''} {chosen.last_name or chosen.username}",
                                                'user': user.username
                                            }
                                        )
                                    except Exception:
                                        pass
                        except Exception:
                            pass

                        # إضافة عناصر الفاتورة
                        for i, product_id in enumerate(products):
                            if product_id and i < len(quantities) and i < len(prices):
                                try:
                                    product = Product.objects.get(id=product_id)
                                    # parse quantity/price/tax robustly to accept '1.5' or '1,5' etc.
                                    quantity = parse_decimal_input(quantities[i], name='quantity', default=Decimal('0'))
                                    unit_price_input = parse_decimal_input(prices[i], name='price', default=Decimal('0'))
                                    tax_rate = product.tax_rate or Decimal('0')
                                    
                                    # ✅ إذا كانت الضريبة شاملة والسعر المُدخل يحتوي على ضريبة، نستخرج السعر الأساسي
                                    if inclusive_tax_flag and tax_rate > 0:
                                        # السعر المُدخل شامل الضريبة، نحوله إلى سعر بدون ضريبة
                                        # السعر الأساسي = السعر المُدخل ÷ (1 + نسبة الضريبة)
                                        unit_price = unit_price_input / (Decimal('1') + (tax_rate / Decimal('100')))
                                    else:
                                        # السعر المُدخل لا يحتوي على ضريبة (الحالة العادية)
                                        unit_price = unit_price_input

                                    # Safeguard: if submitted unit_price differs from product.sale_price
                                    # and appears to be a cost/last-purchase/zero value, prefer product.sale_price
                                    try:
                                        submitted_price = unit_price
                                        product_sale = Decimal(str(product.sale_price))
                                        product_cost = Decimal(str(product.cost_price)) if product.cost_price is not None else None
                                        last_purchase = Decimal(str(product.get_last_purchase_price() or 0))

                                        # If submitted price is 0 or equals cost or equals last purchase price
                                        # but differs from the official sale price, correct it to sale_price.
                                        suspicious = False
                                        if submitted_price == 0:
                                            suspicious = True
                                        elif product_cost is not None and submitted_price == product_cost and submitted_price != product_sale:
                                            suspicious = True
                                        elif last_purchase is not None and submitted_price == last_purchase and submitted_price != product_sale:
                                            suspicious = True

                                        if suspicious and submitted_price != product_sale:
                                            # Log this correction to the AuditLog (via helper)
                                            try:
                                                from core.signals import log_user_activity
                                                # create a small description with field change
                                                desc = _('Unit price changed for product %(code)s during invoice creation: from %(old)s to official sale price %(new)s') % {
                                                    'code': product.code,
                                                    'old': str(submitted_price),
                                                    'new': str(product_sale)
                                                }
                                                # attach note to invoice instance for logging context
                                                log_user_activity(request, 'update', product, desc)
                                            except Exception:
                                                pass

                                            # enforce sale price
                                            unit_price = product_sale
                                    except Exception:
                                        # if any error in price checks, proceed with submitted price
                                        pass

                                    # حساب مبلغ الضريبة لهذا السطر
                                    line_subtotal = quantity * unit_price
                                    line_tax_amount = line_subtotal * (tax_rate / Decimal('100'))
                                    line_total = line_subtotal + line_tax_amount

                                    # إنشاء عنصر الفاتورة
                                    SalesInvoiceItem.objects.create(
                                        invoice=invoice,
                                        product=product,
                                        quantity=quantity,
                                        unit_price=unit_price,
                                        tax_rate=tax_rate,
                                        tax_amount=line_tax_amount.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP),
                                        total_amount=line_total.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
                                    )

                                    # إضافة إلى المجاميع
                                    subtotal += line_subtotal
                                    total_tax_amount += line_tax_amount

                                except (Product.DoesNotExist, ValueError, TypeError):
                                    continue

                        # تحديث مجاميع الفاتورة
                        invoice.subtotal = subtotal.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
                        # if invoice.inclusive_tax is False and user had permission to unset it, zero tax amounts
                        if invoice.inclusive_tax:
                            invoice.tax_amount = total_tax_amount.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
                            invoice.total_amount = (subtotal + total_tax_amount - discount_amount).quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
                        else:
                            invoice.tax_amount = Decimal('0').quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
                            invoice.total_amount = (subtotal - discount_amount).quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)

                        # حفظ الفاتورة (تم فحص الحد الائتماني مسبقاً)
                        invoice.save()

                        # حفظ المستودع الافتراضي إذا تم تحديده
                        set_default_warehouse = request.POST.get('set_default_warehouse')
                        if set_default_warehouse and warehouse:
                            old_default = user.default_sales_warehouse
                            user.default_sales_warehouse = warehouse
                            user.save()
                            
                            # تسجيل التغيير في سجل الأنشطة
                            try:
                                from core.signals import log_user_activity
                                log_user_activity(
                                    request,
                                    'update',
                                    user,
                                    _('Default sales warehouse updated from %(old)s to %(new)s') % {
                                        'old': old_default.name if old_default else _('Not specified'),
                                        'new': warehouse.name
                                    }
                                )
                            except Exception:
                                pass

                        # إذا تم إدخال رقم يدوي يطابق البادئة، نقوم بدفع عداد التسلسل للأمام لتجنب التكرار لاحقاً
                        try:
                            if manual_invoice:
                                seq = DocumentSequence.objects.get(document_type='sales_invoice')
                                if manual_invoice.startswith(seq.prefix):
                                    tail = manual_invoice[len(seq.prefix):]
                                    if tail.isdigit():
                                        seq.advance_to_at_least(int(tail))
                        except Exception:
                            pass

                        # إنشاء حركة حساب للعميل
                        create_sales_invoice_account_transaction(invoice, request.user)

                        # إنشاء القيد المحاسبي
                        create_sales_invoice_journal_entry(invoice, request.user)

                        # حفظ الصندوق النقدي كافتراضي إذا تم طلب ذلك
                        if set_default_cashbox and cashbox:
                            old_default = user.default_cashbox
                            user.default_cashbox = cashbox
                            user.save()
                            
                            # تسجيل التغيير في سجل الأنشطة
                            try:
                                from core.signals import log_user_activity
                                log_user_activity(
                                    request,
                                    'update',
                                    user,
                                    _('Default cashbox updated from %(old)s to %(new)s') % {
                                        'old': old_default.name if old_default else _('Not specified'),
                                        'new': cashbox.name
                                    }
                                )
                            except Exception:
                                pass

                        # تسجيل نشاط صريح لإنشاء الفاتورة (بالإضافة للإشارات العامة)
                        try:
                            from core.signals import log_user_activity
                            activity_desc = _('Created sales invoice number %(number)s') % {'number': invoice.invoice_number}
                            if payment_type == 'cash' and cashbox:
                                activity_desc += _(' - Cash payment from cashbox: %(cashbox)s') % {'cashbox': cashbox.name}
                            log_user_activity(
                                request,
                                'create',
                                invoice,
                                activity_desc
                            )
                        except Exception:
                            pass

                        messages.success(
                            request,
                            _('تم إنشاء فاتورة المبيعات رقم %(number)s بنجاح') % {'number': invoice.invoice_number}
                        )
                        return redirect('sales:invoice_detail', pk=invoice.pk)
                except IntegrityError as ie:
                    # على الأرجح تعارض في رقم الفاتورة، أعد المحاولة برقم جديد
                    if 'invoice_number' in str(ie):
                        # عطّل الرقم اليدوي في المحاولة القادمة وأعد المحاولة
                        allow_manual = False
                        if attempt < max_attempts:
                            continue
                        else:
                            raise
                    else:
                        raise

            # إذا وصلنا هنا ولم نرجع، نبلغ عن فشل عام
            messages.error(request, _('تعذر إنشاء الفاتورة بعد محاولات متعددة، يرجى المحاولة لاحقاً'))
            context = get_invoice_create_context(request, form_data)
            return render(request, 'sales/invoice_add.html', context)
        except Exception as e:
            messages.error(request, _('حدث خطأ أثناء حفظ الفاتورة: %(error)s') % {'error': str(e)})
            context = get_invoice_create_context(request, form_data)
            return render(request, 'sales/invoice_add.html', context)
    
    # GET request - عرض نموذج إنشاء الفاتورة
    try:
        context = get_invoice_create_context(request)
        
        # سجل عرض صفحة إنشاء الفاتورة في سجل النشاط
        try:
            from core.signals import log_user_activity
            dummy = SalesInvoice()
            log_user_activity(
                request,
                'view',
                dummy,
                _('Viewed sales invoice creation page')
            )
        except Exception:
            pass

        return render(request, 'sales/invoice_add.html', context)
    except Exception as e:
        # سجل الاستثناء في اللوق وAuditLog ثم أعِد توجيه المستخدم بدل أن ترفع 500
        import logging
        logger = logging.getLogger(__name__)
        logger.exception("Error rendering sales invoice add page: %s", e)

        try:
            from core.signals import log_user_activity
            # استخدم كائن SalesInvoice وهمي لوصف الحدث
            dummy = SalesInvoice()
            log_user_activity(request, 'view', dummy, _('خطأ عند عرض صفحة إنشاء الفاتورة: %(error)s') % {'error': str(e)})
        except Exception:
            # لا نرمي أي استثناء إضافي أثناء تسجيل الخطأ
            pass

        messages.error(request, _('حدث خطأ أثناء تحميل صفحة إنشاء الفاتورة. تم تسجيل الخطأ لدى النظام.'))
        return redirect('sales:invoice_list')


class SalesInvoiceDetailView(LoginRequiredMixin, DetailView):
    model = SalesInvoice
    template_name = 'sales/invoice_detail.html'
    context_object_name = 'invoice'
    
    def get_queryset(self):
        # تحسين الاستعلام لتضمين البيانات المرتبطة
        return SalesInvoice.objects.select_related(
            'customer', 'created_by'
        ).prefetch_related(
            'items__product__category'
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # إضافة عناصر الفاتورة إلى السياق
        context['invoice_items'] = self.object.items.select_related('product__category').all()
        
        # إضافة القيود المحاسبية المرتبطة
        from journal.models import JournalEntry
        from django.db.models import Q
        # البحث عن القيود المرتبطة بالفاتورة من خلال sales_invoice أو reference_id
        context['journal_entries'] = JournalEntry.objects.filter(
            Q(sales_invoice=self.object) | 
            Q(reference_type__in=['sales_invoice', 'sales_invoice_cogs'], reference_id=self.object.id)
        ).select_related('created_by').distinct()
        
        # Currency settings
        try:
            company_settings = CompanySettings.objects.first()
            if company_settings and company_settings.base_currency:
                context['base_currency'] = company_settings.base_currency
            else:
                context['base_currency'] = Currency.objects.filter(is_active=True).first()
        except:
            pass
        
        return context


class SalesInvoiceUpdateView(LoginRequiredMixin, UpdateView):
    model = SalesInvoice
    template_name = 'sales/invoice_edit.html'
    fields = ['invoice_number', 'date', 'customer', 'payment_type', 'notes']
    
    def get_form(self, form_class=None):
        """Override to add CSS classes to form fields"""
        form = super().get_form(form_class)
        
        # Add Bootstrap classes to all form fields
        for field_name, field in form.fields.items():
            if field_name == 'notes':
                field.widget.attrs.update({
                    'class': 'form-control',
                    'rows': '3',
                    'placeholder': 'أدخل ملاحظات إضافية...'
                })
            elif field_name == 'date':
                field.widget.attrs.update({
                    'class': 'form-control',
                    'type': 'date'
                })
            elif field_name in ['customer', 'payment_type']:
                field.widget.attrs.update({
                    'class': 'form-select'
                })
            else:
                field.widget.attrs.update({
                    'class': 'form-control'
                })
        
        return form
    
    def get_success_url(self):
        return reverse_lazy('sales:invoice_detail', kwargs={'pk': self.object.pk})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from inventory.models import Warehouse
        context['warehouses'] = Warehouse.objects.filter(is_active=True)
        
        # Add products for item addition
        from products.models import Product
        context['products'] = Product.objects.filter(
            is_active=True, 
            product_type__in=['physical', 'service']
        ).select_related('category').order_by('name')
        
        # Add all users for created_by field if user has permission
        if self.request.user.can_change_invoice_creator():
            from django.contrib.auth import get_user_model
            User = get_user_model()
            context['all_users'] = User.objects.filter(is_active=True).order_by('first_name', 'last_name', 'username')
            context['can_change_creator'] = True
        else:
            context['can_change_creator'] = False
        
        # Add decimal places from currency settings
        try:
            company_settings = CompanySettings.objects.first()
            if company_settings and company_settings.base_currency:
                context['decimal_places'] = company_settings.base_currency.decimal_places
            else:
                context['decimal_places'] = 3  # Default to 3 if not set
        except:
            context['decimal_places'] = 3  # Default to 3 if error
        
        return context
    
    def form_valid(self, form):
        """Handle valid form submission"""
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"🔄 بدء معالجة تعديل الفاتورة {form.instance.invoice_number}")
        
        # Store old values for comparison
        old_values = {}
        try:
            old_invoice = SalesInvoice.objects.get(pk=form.instance.pk)
            old_values['invoice_number'] = old_invoice.invoice_number
            old_values['date'] = old_invoice.date
            old_values['customer_id'] = old_invoice.customer_id
            old_values['customer_name'] = old_invoice.customer.name if old_invoice.customer else 'نقدي'
            old_values['payment_type'] = old_invoice.payment_type
            old_values['discount_amount'] = old_invoice.discount_amount
            old_values['notes'] = old_invoice.notes
            old_values['total_amount'] = old_invoice.total_amount
        except SalesInvoice.DoesNotExist:
            pass
        
        # Track if journal entries need updating
        needs_journal_update = False
        
        # Handle discount amount
        discount_amount = self.request.POST.get('discount_amount')
        if discount_amount is not None:
            logger.info(f"  📥 قيمة الخصم المرسلة: '{discount_amount}'")
            # Use robust decimal parsing to handle various client locales
            def parse_decimal_input(val, name='value'):
                from decimal import Decimal as DecimalClass
                default = DecimalClass('0')
                try:
                    if val is None or val == '':
                        return default
                    s = str(val).strip()
                    # Arabic decimal separators and common comma thousands
                    s = s.replace('\u066b', '.')  # Arabic decimal separator if present
                    s = s.replace('\u066c', ',')  # Arabic thousands separator if present
                    # Replace comma with dot for decimal point, remove spaces
                    s = s.replace(',', '.')
                    s = s.replace(' ', '')
                    result = DecimalClass(s)
                    logger.info(f"  🔢 قيمة الخصم المحللة: {result}")
                    return result
                except Exception as e:
                    logger.error(f"  ❌ خطأ في تحليل قيمة الخصم '{val}': {e}")
                    # Log parsing error in AuditLog for visibility
                    try:
                        from core.signals import log_user_activity
                        dummy = SalesInvoice()
                        log_user_activity(self.request, 'error', dummy, _('فشل تحليل قيمة رقمية للحقل %(name)s: %(val)s') % {'name': name, 'val': val})
                    except Exception:
                        pass
                    return DecimalClass('0')
            
            new_discount = parse_decimal_input(discount_amount, name='discount_amount')
            logger.info(f"  💰 قيمة الخصم الجديدة: {new_discount}")
            form.instance.discount_amount = new_discount
            if 'discount_amount' in old_values and old_values['discount_amount'] != new_discount:
                needs_journal_update = True
                logger.info(f"  💰 تغيير مبلغ الخصم من {old_values['discount_amount']} إلى {new_discount}")
                # Log discount change in AuditLog
                try:
                    from core.signals import log_user_activity
                    log_user_activity(self.request, 'update', form.instance, _('تعديل مبلغ الخصم من %(old)s إلى %(new)s') % {'old': old_values['discount_amount'], 'new': new_discount})
                except Exception as e:
                    logger.error(f"  ❌ خطأ في تسجيل تعديل الخصم: {e}")
        
        # Handle warehouse selection
        warehouse_id = self.request.POST.get('warehouse')
        if warehouse_id:
            from inventory.models import Warehouse
            try:
                warehouse = Warehouse.objects.get(id=warehouse_id, is_active=True)
                old_warehouse = form.instance.warehouse
                form.instance.warehouse = warehouse
                if old_warehouse and old_warehouse.id != warehouse.id:
                    logger.info(f"  📦 تغيير المستودع من {old_warehouse.name} إلى {warehouse.name}")
            except Warehouse.DoesNotExist:
                logger.warning(f"  ⚠️ المستودع {warehouse_id} غير موجود")
                pass  # Keep existing warehouse if invalid
        
        # Handle created_by change if user has permission
        created_by_id = self.request.POST.get('created_by')
        old_creator_info = None
        new_creator_info = None
        
        if created_by_id and self.request.user.can_change_invoice_creator():
            from django.contrib.auth import get_user_model
            User = get_user_model()
            try:
                new_creator = User.objects.get(id=created_by_id, is_active=True)
                old_creator = form.instance.created_by
                if old_creator.id != new_creator.id:
                    # حفظ المعلومات للتسجيل لاحقاً
                    old_creator_info = {
                        'name': old_creator.get_full_name() or old_creator.username,
                        'id': old_creator.id
                    }
                    new_creator_info = {
                        'name': new_creator.get_full_name() or new_creator.username,
                        'id': new_creator.id
                    }
                    form.instance.created_by = new_creator
            except User.DoesNotExist:
                pass  # Keep existing creator if invalid
        
        # detect change to inclusive_tax and log activity
        try:
            old = SalesInvoice.objects.get(pk=form.instance.pk)
            old_inclusive = old.inclusive_tax
        except SalesInvoice.DoesNotExist:
            old_inclusive = None

        # Handle item changes from the form
        from decimal import Decimal, ROUND_HALF_UP
        item_changes = []
        index = 0
        while f'item_changes[{index}][item_id]' in self.request.POST:
            try:
                item_id = int(self.request.POST.get(f'item_changes[{index}][item_id]'))
                quantity = Decimal(self.request.POST.get(f'item_changes[{index}][quantity]'))
                unit_price = Decimal(self.request.POST.get(f'item_changes[{index}][unit_price]'))
                item_changes.append({
                    'item_id': item_id,
                    'quantity': quantity,
                    'unit_price': unit_price
                })
            except (ValueError, TypeError):
                pass
            index += 1
        
        # Process item changes if any
        if item_changes:
            logger.info(f"🔄 معالجة {len(item_changes)} تغيير على العناصر")
            
            for change in item_changes:
                try:
                    item = SalesInvoiceItem.objects.get(pk=change['item_id'], invoice=form.instance)
                    
                    old_quantity = item.quantity
                    old_price = item.unit_price
                    
                    # Check for actual changes
                    if old_quantity == change['quantity'] and old_price == change['unit_price']:
                        continue
                    
                    # Update item
                    item.quantity = change['quantity']
                    item.unit_price = change['unit_price']
                    
                    # Recalculate amounts
                    line_subtotal = item.quantity * item.unit_price
                    line_tax_amount = line_subtotal * (item.tax_rate / Decimal('100'))
                    line_total = line_subtotal + line_tax_amount
                    
                    item.tax_amount = line_tax_amount.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
                    item.total_amount = line_total.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
                    item.save()
                    
                    logger.info(f"  ✅ تحديث العنصر {item.product.name}: الكمية {old_quantity}→{change['quantity']}, السعر {old_price}→{change['unit_price']}")
                    
                    # Update inventory movements if quantity changed
                    if old_quantity != change['quantity']:
                        from inventory.models import InventoryMovement
                        quantity_diff = change['quantity'] - old_quantity
                        
                        # Create inventory movement for the difference (only if not zero)
                        if quantity_diff != 0:
                            InventoryMovement.objects.create(
                                product=item.product,
                                warehouse=form.instance.warehouse,
                                movement_type='out' if quantity_diff > 0 else 'in',
                                quantity=abs(quantity_diff),
                                date=form.instance.date,
                                reference_type='sales_invoice',
                                reference_id=form.instance.id,
                                notes=f'تعديل فاتورة المبيعات {form.instance.invoice_number} - تغيير الكمية'
                            )
                        
                        # Update product stock
                        item.product.current_stock -= quantity_diff
                        item.product.save(update_fields=['current_stock'])
                        
                        logger.info(f"  📦 تحديث المخزون: {item.product.name} بمقدار {quantity_diff}")
                    
                    # Log item change activity
                    try:
                        from core.signals import log_user_activity
                        change_details = []
                        if old_quantity != change['quantity']:
                            change_details.append(f'الكمية من {old_quantity} إلى {change["quantity"]}')
                        if old_price != change['unit_price']:
                            change_details.append(f'سعر الوحدة من {old_price} إلى {change["unit_price"]}')
                        
                        if change_details:
                            changes_text = '، '.join(change_details)
                            log_user_activity(
                                self.request,
                                'update',
                                item,
                                f'تحديث عنصر {item.product.name} في فاتورة المبيعات {form.instance.invoice_number}: {changes_text}'
                            )
                    except Exception as e:
                        logger.error(f"❌ خطأ في تسجيل نشاط العنصر: {e}")
                        
                except SalesInvoiceItem.DoesNotExist:
                    logger.warning(f"  ⚠️ العنصر {change['item_id']} غير موجود")
                except Exception as e:
                    logger.error(f"  ❌ خطأ في تحديث العنصر {change['item_id']}: {e}")

        response = super().form_valid(form)
        
        # Log all changes
        changes = []
        if 'date' in old_values and old_values['date'] != form.instance.date:
            changes.append(f"التاريخ من {old_values['date']} إلى {form.instance.date}")
        if 'customer_id' in old_values and old_values['customer_id'] != form.instance.customer_id:
            new_customer_name = form.instance.customer.name if form.instance.customer else 'نقدي'
            changes.append(f"العميل من {old_values['customer_name']} إلى {new_customer_name}")
        if 'payment_type' in old_values and old_values['payment_type'] != form.instance.payment_type:
            changes.append(f"نوع الدفع من {old_values['payment_type']} إلى {form.instance.payment_type}")
        if 'discount_amount' in old_values and old_values['discount_amount'] != form.instance.discount_amount:
            changes.append(f"مبلغ الخصم من {old_values['discount_amount']} إلى {form.instance.discount_amount}")
        if 'notes' in old_values and old_values['notes'] != form.instance.notes:
            changes.append(f"تحديث الملاحظات")
        
        response = super().form_valid(form)
        
        # Check if discount changed
        discount_changed = 'discount_amount' in old_values and old_values['discount_amount'] != form.instance.discount_amount
        if discount_changed:
            needs_journal_update = True
            logger.info(f"  💰 تغيير مبلغ الخصم من {old_values['discount_amount']} إلى {form.instance.discount_amount}")
            # Update totals when discount changes
            form.instance.update_totals()
            form.instance.refresh_from_db()
        
        # Update invoice totals and journal entries if items were changed or discount changed
        if item_changes or discount_changed:
            # Update invoice totals
            form.instance.update_totals()
            form.instance.refresh_from_db()
            
            logger.info(f"  📊 إعادة حساب مجاميع الفاتورة: الإجمالي {form.instance.total_amount}")
            
            # Update or create journal entries
            try:
                # Delete existing journal entries for this invoice
                from journal.models import JournalEntry
                JournalEntry.objects.filter(
                    reference_type__in=['sales_invoice', 'sales_invoice_cogs'],
                    reference_id=form.instance.id
                ).delete()
                
                # Create new journal entries
                create_sales_invoice_journal_entry(form.instance, self.request.user)
                
                logger.info(f"  📝 إعادة إنشاء القيود المحاسبية")
            except Exception as e:
                logger.error(f"❌ خطأ في إنشاء القيود المحاسبية: {e}")
            
            # Update customer balance if total changed
            if form.instance.customer:
                try:
                    old_total = old_values.get('total_amount', form.instance.total_amount)
                    if old_total != form.instance.total_amount:
                        total_diff = form.instance.total_amount - old_total
                        form.instance.customer.balance += total_diff
                        form.instance.customer.save(update_fields=['balance'])
                        
                        logger.info(f"  👤 تحديث رصيد العميل {form.instance.customer.name}: {total_diff}")
                except Exception as e:
                    logger.error(f"❌ خطأ في تحديث رصيد العميل: {e}")
        
        # Log main invoice changes
        if changes:
            try:
                from core.signals import log_user_activity
                changes_text = '، '.join(changes)
                log_user_activity(
                    self.request,
                    'update',
                    self.object,
                    f'تحديث فاتورة المبيعات {self.object.invoice_number}: {changes_text}'
                )
            except Exception as e:
                logger.error(f"❌ خطأ في تسجيل التحديثات: {e}")
        
        # Log creator change after saving
        if old_creator_info and new_creator_info:
            try:
                from core.signals import log_user_activity
                log_user_activity(
                    self.request,
                    'update',
                    self.object,
                    f'تغيير منشئ الفاتورة {self.object.invoice_number} من {old_creator_info["name"]} إلى {new_creator_info["name"]}'
                )
            except Exception:
                pass

        try:
            new_inclusive = self.object.inclusive_tax
            if old_inclusive is not None and old_inclusive != new_inclusive:
                from core.signals import log_user_activity
                log_user_activity(
                    self.request,
                    'update',
                    self.object,
                    f'تغيير خيار شامل ضريبة من {old_inclusive} إلى {new_inclusive} لفاتورة {self.object.invoice_number}'
                )
        except Exception:
            pass

        logger.info(f"✅ تم حفظ الفاتورة {self.object.invoice_number} بنجاح")
        
        messages.success(self.request, 'تم تحديث فاتورة المبيعات بنجاح')
        return response
    
    def form_invalid(self, form):
        """Handle form validation errors"""
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"❌ فشل حفظ الفاتورة. الأخطاء: {form.errors}")
        
        messages.error(self.request, 'حدث خطأ في حفظ البيانات. يرجى مراجعة الحقول والمحاولة مرة أخرى.')
        for field, errors in form.errors.items():
            field_name = form.fields[field].label if field in form.fields else field
            for error in errors:
                messages.error(self.request, f'{field_name}: {error}')
        return super().form_invalid(form)


class SalesInvoiceDeleteView(LoginRequiredMixin, DeleteView):
    model = SalesInvoice
    template_name = 'sales/invoice_delete.html'
    success_url = reverse_lazy('sales:invoice_list')
    
    def delete(self, request, *args, **kwargs):
        from django.db import transaction
        
        invoice = self.get_object()
        invoice_number = invoice.invoice_number
        
        # استخدام transaction لضمان التنفيذ الكامل أو الإلغاء
        with transaction.atomic():
            # حذف حركات المخزون المرتبطة (حذفها سيُرجع الكميات تلقائياً)
            try:
                from inventory.models import InventoryMovement
                inventory_movements = InventoryMovement.objects.filter(
                    reference_type='sales_invoice',
                    reference_id=invoice.id
                )
                if inventory_movements.exists():
                    movement_count = inventory_movements.count()
                    # حفظ تفاصيل الحركات للتسجيل
                    movements_details = []
                    for movement in inventory_movements:
                        movements_details.append({
                            'product': movement.product.name,
                            'quantity': movement.quantity,
                            'warehouse': movement.warehouse.name
                        })
                    
                    # حذف الحركات - سيؤدي لإرجاع الكميات تلقائياً
                    inventory_movements.delete()
                    
                    # تسجيل في سجل الأنشطة
                    try:
                        from core.signals import log_user_activity
                        details_text = ', '.join([f"{m['product']} ({m['quantity']}) من {m['warehouse']}" for m in movements_details])
                        log_user_activity(
                            request,
                            'delete',
                            invoice,
                            _('تم حذف %(count)s حركة مخزون للفاتورة %(number)s - إرجاع الكميات: %(details)s') % {
                                'count': movement_count,
                                'number': invoice_number,
                                'details': details_text
                            }
                        )
                    except Exception:
                        pass
                    
                    print(f"✅ تم حذف {movement_count} حركة مخزون للفاتورة {invoice_number} - تم إرجاع الكميات للمخزون")
                else:
                    print(f"⚠️ لا توجد حركات مخزون للفاتورة {invoice_number} (فاتورة قديمة أو خدمية)")
            except ImportError:
                pass
            except Exception as e:
                print(f"❌ خطأ في حذف حركات المخزون: {e}")
                messages.error(request, _('حدث خطأ في إرجاع الكميات للمخزون'))
                raise  # إلغاء العملية كاملة
            
            # حذف القيود المحاسبية المرتبطة
            try:
                from journal.models import JournalEntry
                from django.db.models import Q
                
                # حذف جميع القيود المرتبطة بالفاتورة (المبيعات و COGS)
                # البحث عن القيود بطرق متعددة للتأكد من حذف الكل
                journal_entries = JournalEntry.objects.filter(
                    Q(sales_invoice=invoice) | 
                    Q(reference_type='sales_invoice', reference_id=invoice.id) |
                    Q(reference_type='sales_invoice_cogs', reference_id=invoice.id)
                ).distinct()
                
                if journal_entries.exists():
                    entry_count = journal_entries.count()
                    entry_numbers = ', '.join([entry.entry_number for entry in journal_entries])
                    
                    # طباعة تفاصيل القيود قبل الحذف للتصحيح
                    print(f"🔍 القيود المراد حذفها للفاتورة {invoice_number}:")
                    for entry in journal_entries:
                        print(f"   - {entry.entry_number}: reference_type={entry.reference_type}, reference_id={entry.reference_id}")
                    
                    journal_entries.delete()
                    
                    # تسجيل في سجل الأنشطة
                    try:
                        from core.signals import log_user_activity
                        log_user_activity(
                            request,
                            'delete',
                            invoice,
                            _('تم حذف %(count)s قيد محاسبي للفاتورة %(number)s: %(entries)s') % {
                                'count': entry_count,
                                'number': invoice_number,
                                'entries': entry_numbers
                            }
                        )
                    except Exception:
                        pass
                    
                    print(f"✅ تم حذف {entry_count} قيد محاسبي للفاتورة {invoice_number}: {entry_numbers}")
                else:
                    print(f"⚠️ لا توجد قيود محاسبية للفاتورة {invoice_number}")
            except Exception as e:
                print(f"❌ خطأ في حذف القيود المحاسبية: {e}")
                messages.error(request, _('حدث خطأ في حذف القيود المحاسبية'))
                raise
            
            # حذف حركات حساب العميل
            try:
                delete_transaction_by_reference('sales_invoice', invoice.id)
                print(f"✅ تم حذف حركات حساب العميل للفاتورة {invoice_number}")
            except Exception as e:
                print(f"⚠️ خطأ في حذف حركات حساب العميل: {e}")
        
            # تسجيل النشاط قبل الحذف
            from core.signals import log_activity
            log_activity(
                user=request.user,
                action_type='DELETE',
                obj=invoice,
                description=f'تم حذف فاتورة مبيعات رقم: {invoice_number} - تم إرجاع الكميات للمخزون وعكس القيود المحاسبية',
                request=request
            )
            
            # حذف الفاتورة نفسها (يجب أن يكون داخل transaction)
            invoice.delete()
        
        messages.success(request, _('تم حذف فاتورة المبيعات رقم %(number)s بنجاح وإرجاع الكميات للمخزون') % {'number': invoice_number})
        return redirect(self.success_url)


class SalesReturnListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = SalesReturn
    template_name = 'sales/return_list.html'
    context_object_name = 'returns'
    paginate_by = 10
    
    def test_func(self):
        return self.request.user.has_sales_permission()
    
    def get_queryset(self):
        queryset = SalesReturn.objects.all()
        
        # Search functionality
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(return_number__icontains=search) |
                Q(customer__name__icontains=search)
            )
        
        # Date filter
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        
        # Customer filter
        customer = self.request.GET.get('customer')
        if customer:
            queryset = queryset.filter(customer_id=customer)
        
        # Apply ordering
        order_by = self.request.GET.get('order_by', '-date')
        if order_by.startswith('-'):
            queryset = queryset.order_by(order_by, '-id')
        else:
            queryset = queryset.order_by(order_by, 'id')
        
        return queryset.select_related('customer', 'original_invoice', 'created_by')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Statistics
        from django.utils import timezone
        from datetime import datetime, timedelta
        
        returns = SalesReturn.objects.all()
        context['total_returns'] = returns.count()
        context['total_returns_amount'] = returns.aggregate(
            total=Sum('total_amount')
        )['total'] or 0
        
        # Monthly and daily stats
        today = timezone.now().date()
        month_start = today.replace(day=1)
        context['monthly_returns'] = returns.filter(date__gte=month_start).count()
        context['daily_returns'] = returns.filter(date=today).count()
        
        # Customers for filter
        context['customers'] = CustomerSupplier.objects.filter(
            type__in=['customer', 'both']
        )
        
        # Currency settings
        try:
            company_settings = CompanySettings.objects.first()
            if company_settings and company_settings.base_currency:
                context['base_currency'] = company_settings.base_currency
            else:
                context['base_currency'] = Currency.objects.filter(is_active=True).first()
        except:
            pass
        
        # Current ordering
        context['current_order'] = self.request.GET.get('order_by', '-date')
        
        return context


class SalesReturnDetailView(LoginRequiredMixin, DetailView):
    model = SalesReturn
    template_name = 'sales/return_detail.html'
    context_object_name = 'return'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get return items
        context['items'] = self.object.items.select_related('product').all()
        
        # Calculate subtotal without tax for each item
        items_with_subtotal = []
        for item in context['items']:
            item.subtotal = item.quantity * item.unit_price
            items_with_subtotal.append(item)
        context['items'] = items_with_subtotal
        
        # إضافة القيود المحاسبية المرتبطة
        from journal.models import JournalEntry
        # البحث عن القيود المرتبطة بالمردود
        context['journal_entries'] = JournalEntry.objects.filter(sales_return=self.object).select_related('created_by').distinct()
        
        # Currency settings
        try:
            company_settings = CompanySettings.objects.first()
            if company_settings and company_settings.base_currency:
                context['base_currency'] = company_settings.base_currency
            else:
                context['base_currency'] = Currency.objects.filter(is_active=True).first()
        except:
            pass
        
        return context


class SalesReturnUpdateView(LoginRequiredMixin, UpdateView):
    model = SalesReturn
    template_name = 'sales/return_edit.html'
    fields = ['return_number', 'date', 'customer', 'notes']
    
    def get_success_url(self):
        return reverse_lazy('sales:return_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        # تسجيل النشاط
        try:
            from core.signals import log_user_activity
            log_user_activity(
                self.request,
                'update',
                self.object,
                f'تحديث مردود مبيعات رقم {self.object.return_number}'
            )
        except Exception:
            pass
        
        messages.success(self.request, f'تم تحديث المردود بنجاح')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Currency settings
        try:
            company_settings = CompanySettings.objects.first()
            if company_settings and company_settings.base_currency:
                context['base_currency'] = company_settings.base_currency
            else:
                context['base_currency'] = Currency.objects.filter(is_active=True).first()
        except:
            pass
        return context


@login_required
def sales_return_create(request):
    """Create new sales return"""
    
    # دالة مساعدة لتحويل القيم النصية إلى Decimal بشكل آمن
    def parse_decimal_input(val, name='value', default=Decimal('0')):
        try:
            if val is None or val == '':
                return default
            s = str(val).strip()
            # Arabic decimal separators and common comma thousands
            s = s.replace('\u066b', '.')  # Arabic decimal separator if present
            s = s.replace('\u066c', ',')  # Arabic thousands separator if present
            # Replace comma with dot for decimal point, remove spaces
            s = s.replace(',', '.')
            s = s.replace(' ', '')
            return Decimal(s)
        except Exception:
            # Log parsing error in AuditLog for visibility
            try:
                from core.signals import log_user_activity
                dummy = SalesReturn()
                log_user_activity(request, 'error', dummy, _('فشل تحليل قيمة رقمية للحقل %(name)s: %(val)s') % {'name': name, 'val': val})
            except Exception:
                pass
            return default
    
    def get_return_create_context(request, form_data=None):
        """Prepare return creation page context with entered data if any"""
        user = request.user
        context = {
            'customers': CustomerSupplier.objects.filter(type__in=['customer', 'both']),
            'today_date': date.today().isoformat(),
        }
        
        # إضافة البيانات المُدخلة إذا كانت موجودة
        if form_data:
            context.update(form_data)
        
        # إضافة الفواتير المتاحة للمرتجعات
        context['invoices'] = SalesInvoice.objects.filter(
            customer__isnull=False
        ).select_related('customer').order_by('-date')[:50]  # حدّث لأحدث 50 فاتورة
        
        # توليد رقم المرتجع المقترح
        try:
            from core.models import DocumentSequence
            sequence = DocumentSequence.objects.get(document_type='sales_return')
            context['suggested_return_number'] = sequence.get_formatted_number()
        except DocumentSequence.DoesNotExist:
            # التراجع إلى الترقيم البسيط
            last_return = SalesReturn.objects.order_by('-id').first()
            if last_return:
                try:
                    number = int(last_return.return_number.split('-')[-1]) + 1
                    context['suggested_return_number'] = f"RETURN-{number:06d}"
                except (ValueError, IndexError):
                    context['suggested_return_number'] = f"RETURN-{SalesReturn.objects.count() + 1:06d}"
            else:
                context['suggested_return_number'] = "RETURN-000001"
        
        return context
    
    if request.method == 'POST':
        # جمع جميع البيانات المُدخلة لإعادة عرضها في حالة الأخطاء
        form_data = {
            'original_invoice_id': request.POST.get('original_invoice'),
            'return_reason': request.POST.get('return_reason'),
            'notes': request.POST.get('notes', ''),
            'products': request.POST.getlist('products[]'),
            'quantities': request.POST.getlist('quantities[]'),
            'prices': request.POST.getlist('prices[]'),
            'tax_amounts': request.POST.getlist('tax_amounts[]'),
        }
        
        try:
            # سنحاول عدة مرات لتجنب تعارض الأرقام في حال السباق
            max_attempts = 5
            attempt = 0
            allow_manual = True
            while attempt < max_attempts:
                attempt += 1
                try:
                    with transaction.atomic():
                        user = request.user
                        return_reason = request.POST.get('return_reason')
                        notes = request.POST.get('notes', '')
                        
                        # توليد رقم المرتجع إذا لم يكن محدد
                        return_number = None
                        manual_return_number = request.POST.get('return_number')
                        
                        # في المحاولة الأولى فقط، نسمح بالرقم اليدوي
                        if allow_manual and manual_return_number and (user.is_superuser or user.is_staff):
                            # التحقق من عدم وجود رقم مكرر
                            if not SalesReturn.objects.filter(return_number=manual_return_number).exists():
                                return_number = manual_return_number

                        # إذا لم يكن هناك رقم يدوي أو كان مكرراً، استخدم التسلسل التلقائي
                        if not return_number:
                            try:
                                sequence = DocumentSequence.objects.select_for_update().get(document_type='sales_return')
                                return_number = sequence.get_next_number()
                            except DocumentSequence.DoesNotExist:
                                # توليد بديل إذا لم يوجد تسلسل - مع قفل للحماية من التعارضات
                                last_return = SalesReturn.objects.select_for_update().order_by('-id').first()
                                if last_return:
                                    # استخراج الرقم من آخر مرتجع
                                    import re
                                    match = re.search(r'(\d+)$', last_return.return_number)
                                    if match:
                                        number = int(match.group(1)) + 1
                                    else:
                                        number = SalesReturn.objects.count() + 1
                                    return_number = f"SRET-{number:06d}"
                                else:
                                    return_number = "SRET-000001"

                        # التحقق من البيانات المطلوبة
                        errors = []
                        original_invoice_id = request.POST.get('original_invoice')
                        if not original_invoice_id:
                            errors.append(_('يرجى اختيار الفاتورة الأصلية'))
                        if not return_reason:
                            errors.append(_('يرجى تحديد سبب المرتجع'))
                        
                        if errors:
                            for error in errors:
                                messages.error(request, error)
                            # جمع البيانات المُدخلة لإعادة عرضها
                            form_data = {
                                'original_invoice_id': original_invoice_id,
                                'return_reason': return_reason,
                                'notes': notes,
                            }
                            context = get_return_create_context(request, form_data)
                            return render(request, 'sales/return_add.html', context)

                        # الحصول على الفاتورة الأصلية
                        original_invoice = get_object_or_404(SalesInvoice, id=original_invoice_id)
                        
                        # جلب العميل من الفاتورة الأصلية
                        customer = original_invoice.customer

                        # إنشاء المردود (لاحظ: SalesReturn لا يحتوي على حقل warehouse)
                        sales_return = SalesReturn.objects.create(
                            return_number=return_number,
                            date=date.today(),
                            original_invoice=original_invoice,
                            customer=customer,
                            notes=notes,
                            created_by=user,
                            subtotal=0,  # سيتم تحديثها لاحقاً
                            tax_amount=0,  # سيتم تحديثها لاحقاً
                            total_amount=0  # سيتم تحديثها لاحقاً
                        )

                        # معالجة بيانات المنتجات المرتجعة
                        return_products = request.POST.getlist('products[]')
                        return_quantities = request.POST.getlist('quantities[]')
                        return_prices = request.POST.getlist('prices[]')
                        return_tax_amounts = request.POST.getlist('tax_amounts[]')

                        # التحقق من وجود منتجات للإرجاع
                        if not return_products or not any(p for p in return_products if p):
                            # سجل محاولة فاشلة في سجل النشاط لتتبع أخطاء الإدخال
                            try:
                                from core.signals import log_user_activity
                                dummy = SalesReturn()
                                log_user_activity(
                                    request,
                                    'error',
                                    dummy,
                                    _('فشل إنشاء مردود: لا توجد عناصر مضافة')
                                )
                            except Exception:
                                pass

                            messages.error(request, _('يرجى إضافة منتج واحد على الأقل للإرجاع'))
                            context = get_return_create_context(request, form_data)
                            return render(request, 'sales/return_add.html', context)

                        # حساب المجاميع أولاً قبل إنشاء أي شيء (للتحقق من الحد الائتماني)
                        subtotal = Decimal('0')
                        total_tax_amount = Decimal('0')
                        
                        # حساب المجاميع المؤقتة
                        for i, product_id in enumerate(return_products):
                            if product_id and i < len(return_quantities) and i < len(return_prices):
                                try:
                                    product = Product.objects.get(id=product_id)
                                    quantity = parse_decimal_input(return_quantities[i], name='quantity', default=Decimal('0'))
                                    unit_price = parse_decimal_input(return_prices[i], name='price', default=Decimal('0'))
                                    
                                    # البحث عن tax_rate من الفاتورة الأصلية للمنتج نفسه
                                    tax_rate = Decimal('0')
                                    try:
                                        original_item = original_invoice.items.filter(product=product).first()
                                        if original_item:
                                            tax_rate = original_item.tax_rate or Decimal('0')
                                        else:
                                            tax_rate = product.tax_rate or Decimal('0')
                                    except:
                                        tax_rate = product.tax_rate or Decimal('0')

                                    if quantity <= 0 or unit_price < 0:
                                        continue

                                    line_subtotal = quantity * unit_price
                                    line_tax_amount = line_subtotal * (tax_rate / 100) if tax_rate > 0 else Decimal('0')
                                    
                                    subtotal += line_subtotal
                                    total_tax_amount += line_tax_amount
                                except (Product.DoesNotExist, ValueError, TypeError):
                                    continue

                        # حساب الإجمالي النهائي
                        final_total = (subtotal + total_tax_amount).quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)

                        # إذا كان هناك حد ائتماني، تحقق من أن المبلغ الإجمالي لا يتجاوز الحد المتاح
                        if customer and customer.credit_limit and customer.credit_limit > 0:
                            current_balance = customer.current_balance
                            available_credit = customer.credit_limit - abs(current_balance) if current_balance < 0 else customer.credit_limit
                            
                            if final_total > available_credit:
                                # رسالة تحذير مع اقتراحات
                                error_message = _(
                                    'لا يمكن إنشاء المردود لأن المبلغ الإجمالي (%(total)s) يتجاوز الحد الائتماني المتاح للعميل (%(available)s).\n\nاقتراحات:\n1. زيادة الحد الائتماني للعميل\n2. تحصيل المستحقات المتأخرة من العميل\n3. تحويل المردود إلى دفع نقدي'
                                ) % {
                                    'total': f"{final_total:.3f}",
                                    'available': f"{available_credit:.3f}"
                                }
                                
                                # التأكد من عدم وجود رسائل سابقة من نفس النوع
                                existing_messages = [msg for msg in messages.get_messages(request) if msg.message == error_message]
                                if not existing_messages:
                                    messages.warning(request, error_message)
                                
                                # تسجيل المحاولة الفاشلة في سجل الأنشطة
                                try:
                                    from core.signals import log_user_activity
                                    log_user_activity(
                                        request,
                                        'error',
                                        customer,
                                        _('فشل في إنشاء مردود مبيعات: تجاوز الحد الائتماني - المبلغ %(total)s > المتاح %(available)s') % {
                                            'total': f"{final_total:.3f}",
                                            'available': f"{available_credit:.3f}"
                                        }
                                    )
                                except Exception:
                                    pass
                                
                                context = get_return_create_context(request, form_data)
                                return render(request, 'sales/return_add.html', context)

                        # إذا وصلنا هنا، الحد الائتماني مسموح - يمكننا إنشاء المردود
                        # إعادة تعيين المجاميع لإعادة الحساب أثناء الإنشاء الفعلي
                        subtotal = Decimal('0')
                        total_tax_amount = Decimal('0')

                        # إضافة عناصر المرتجع
                        for i, product_id in enumerate(return_products):
                            if product_id and i < len(return_quantities) and i < len(return_prices):
                                try:
                                    product = Product.objects.get(id=product_id)
                                    # parse quantity/price/tax robustly to accept '1.5' or '1,5' etc.
                                    quantity = parse_decimal_input(return_quantities[i], name='quantity', default=Decimal('0'))
                                    unit_price = parse_decimal_input(return_prices[i], name='price', default=Decimal('0'))
                                    
                                    # البحث عن tax_rate من الفاتورة الأصلية للمنتج نفسه
                                    tax_rate = Decimal('0')
                                    try:
                                        original_item = original_invoice.items.filter(product=product).first()
                                        if original_item:
                                            tax_rate = original_item.tax_rate or Decimal('0')
                                        else:
                                            tax_rate = product.tax_rate or Decimal('0')
                                    except:
                                        tax_rate = product.tax_rate or Decimal('0')

                                    # حساب مبلغ الضريبة لهذا السطر
                                    line_subtotal = quantity * unit_price
                                    line_tax_amount = line_subtotal * (tax_rate / Decimal('100'))
                                    line_total = line_subtotal + line_tax_amount

                                    # إنشاء عنصر المردود
                                    SalesReturnItem.objects.create(
                                        return_invoice=sales_return,
                                        product=product,
                                        quantity=quantity,
                                        unit_price=unit_price,
                                        tax_rate=tax_rate,
                                        tax_amount=line_tax_amount.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP),
                                        total_amount=line_total.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
                                    )

                                    # إضافة إلى المجاميع
                                    subtotal += line_subtotal
                                    total_tax_amount += line_tax_amount

                                except (Product.DoesNotExist, ValueError, TypeError):
                                    continue

                        # تحديث مجاميع المردود
                        sales_return.subtotal = subtotal.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
                        sales_return.tax_amount = total_tax_amount.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
                        sales_return.total_amount = (subtotal + total_tax_amount).quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)

                        # حفظ المردود
                        sales_return.save()

                        # إنشاء القيود المحاسبية
                        create_sales_return_journal_entry(sales_return, user)

                        # إنشاء حركة حساب للعميل
                        create_sales_return_account_transaction(sales_return, user)

                        # حركات المخزون تُنشأ تلقائياً عبر signals
                        
                        # تسجيل النشاط
                        try:
                            from core.signals import log_user_activity
                            log_user_activity(
                                request,
                                'create',
                                sales_return,
                                f'إنشاء مردود مبيعات {sales_return.return_number} للفاتورة {sales_return.original_invoice.invoice_number}'
                            )
                        except Exception:
                            pass
                        
                        messages.success(request, f'تم إنشاء مردود المبيعات {sales_return.return_number} بنجاح')
                        return redirect('sales:return_detail', pk=sales_return.pk)
                except IntegrityError as ie:
                    # على الأرجح تعارض في رقم المرتجع، أعد المحاولة برقم جديد
                    if 'return_number' in str(ie):
                        # عطّل الرقم اليدوي في المحاولة القادمة وأعد المحاولة
                        allow_manual = False
                        if attempt < max_attempts:
                            continue
                        else:
                            raise
                    else:
                        raise

            # إذا وصلنا هنا ولم نرجع، نبلغ عن فشل عام
            messages.error(request, _('تعذر إنشاء المردود بعد محاولات متعددة، يرجى المحاولة لاحقاً'))
            context = get_return_create_context(request, form_data)
            return render(request, 'sales/return_add.html', context)
        except Exception as e:
            messages.error(request, _('حدث خطأ أثناء حفظ المردود: %(error)s') % {'error': str(e)})
            context = get_return_create_context(request, form_data)
            return render(request, 'sales/return_add.html', context)
    
    # GET request - عرض نموذج إنشاء المردود
    context = get_return_create_context(request)
    return render(request, 'sales/return_add.html', context)


class SalesCreditNoteListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = SalesCreditNote
    template_name = 'sales/creditnote_list.html'
    context_object_name = 'creditnotes'
    paginate_by = 25

    def test_func(self):
        return (
            self.request.user.has_perm('sales.can_view_credit_notes') or
            self.request.user.is_superuser
        )

    def get_queryset(self):
        queryset = SalesCreditNote.objects.select_related('customer', 'created_by').all()
        
        # فلترة حسب العميل
        customer_id = self.request.GET.get('customer')
        if customer_id:
            queryset = queryset.filter(customer_id=customer_id)
        
        # فلترة حسب التاريخ
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        
        return queryset.order_by('-date', '-id')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['customers'] = CustomerSupplier.objects.filter(type__in=['customer', 'both'])
        
        # Currency settings
        try:
            company_settings = CompanySettings.objects.first()
            if company_settings and company_settings.base_currency:
                context['base_currency'] = company_settings.base_currency
            else:
                context['base_currency'] = Currency.objects.filter(is_active=True).first()
        except:
            pass
        
        return context


def sales_creditnote_create(request):
    if not (
        request.user.has_perm('sales.can_add_sales') or
        request.user.is_superuser
    ):
        messages.error(request, _('ليس لديك صلاحية لإنشاء إشعار دائن'))
        return redirect('/')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                customer_id = request.POST.get('customer')
                if not customer_id:
                    messages.error(request, _('يرجى اختيار عميل'))
                    return redirect('sales:creditnote_add')

                customer = get_object_or_404(CustomerSupplier, id=customer_id)

                # توليد الرقم
                try:
                    seq = DocumentSequence.objects.get(document_type='credit_note')
                    note_number = seq.get_next_number()
                except DocumentSequence.DoesNotExist:
                    # إضافة رسالة خطأ بدلاً من إنشاء رقم تلقائي
                    messages.error(request, _('Credit note sequence not configured. Please configure it in settings first.'))
                    return redirect('sales:creditnote_add')

                credit = SalesCreditNote.objects.create(
                    note_number=note_number,
                    date=request.POST.get('date', date.today()),
                    customer=customer,
                    subtotal=Decimal(request.POST.get('subtotal', '0') or '0'),
                    notes=request.POST.get('notes', ''),
                    created_by=request.user
                )

                # إنشاء القيد المحاسبي
                try:
                    from journal.services import JournalService
                    JournalService.create_sales_credit_note_entry(credit, request.user)
                    messages.success(request, f'تم إنشاء القيد المحاسبي لإشعار الدائن رقم {credit.note_number}')
                except Exception as e:
                    error_msg = f"Error creating journal entry for credit note: {str(e)}"
                    logging.getLogger(__name__).error(error_msg)
                    messages.warning(request, f'تم إنشاء إشعار الدائن ولكن حدث خطأ في القيد المحاسبي: {str(e)}')

                try:
                    from core.signals import log_user_activity
                    log_user_activity(request, 'create', credit, _('إنشاء إشعار دائن رقم %(number)s') % {'number': credit.note_number})
                except Exception:
                    pass

                messages.success(request, _('تم إنشاء إشعار دائن رقم %(number)s') % {'number': credit.note_number})
                return redirect('sales:creditnote_detail', pk=credit.pk)
        except Exception as e:
            messages.error(request, _('حدث خطأ أثناء حفظ الإشعار: %(error)s') % {'error': str(e)})
            return redirect('sales:creditnote_add')

    context = {
        'customers': CustomerSupplier.objects.filter(Q(type='customer')|Q(type='both')),
        'today_date': date.today().isoformat(),
    }
    try:
        seq = DocumentSequence.objects.get(document_type='credit_note')
        context['next_note_number'] = seq.peek_next_number() if hasattr(seq, 'peek_next_number') else seq.get_formatted_number()
    except DocumentSequence.DoesNotExist:
        # إضافة رسالة خطأ بدلاً من إنشاء رقم تلقائي
        context['sequence_error'] = _('Credit note sequence not configured. Please configure it in settings first.')
        context['next_note_number'] = None

    return render(request, 'sales/creditnote_add.html', context)


class SalesCreditNoteDetailView(LoginRequiredMixin, DetailView):
    model = SalesCreditNote
    template_name = 'sales/creditnote_detail.html'
    context_object_name = 'creditnote'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        try:
            company_settings = CompanySettings.objects.first()
            if company_settings and company_settings.base_currency:
                context['base_currency'] = company_settings.base_currency
            else:
                context['base_currency'] = Currency.objects.filter(is_active=True).first()
        except:
            pass
        
        # إضافة القيود المحاسبية المرتبطة
        from journal.models import JournalEntry
        context['journal_entries'] = JournalEntry.objects.filter(
            reference_type='credit_note',
            reference_id=self.object.id
        ).select_related('created_by')
        
        return context


class SalesCreditNoteUpdateView(LoginRequiredMixin, UpdateView):
    model = SalesCreditNote
    template_name = 'sales/creditnote_edit.html'
    fields = ['note_number', 'date', 'customer', 'subtotal', 'tax_amount', 'total_amount', 'notes']
    
    def get_success_url(self):
        return reverse_lazy('sales:creditnote_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        # تسجيل النشاط
        try:
            from core.signals import log_user_activity
            log_user_activity(
                self.request,
                'update',
                self.object,
                _('تحديث إشعار دائن رقم %(number)s') % {'number': self.object.note_number}
            )
        except Exception:
            pass
        
        messages.success(self.request, _('تم تحديث إشعار الدائن بنجاح'))
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['customers'] = CustomerSupplier.objects.filter(type__in=['customer', 'both'])
        
        # Currency settings
        try:
            company_settings = CompanySettings.objects.first()
            if company_settings and company_settings.base_currency:
                context['base_currency'] = company_settings.base_currency
            else:
                context['base_currency'] = Currency.objects.filter(is_active=True).first()
        except:
            pass
        
        return context


class SalesCreditNoteDeleteView(LoginRequiredMixin, DeleteView):
    model = SalesCreditNote
    template_name = 'sales/creditnote_delete.html'
    success_url = reverse_lazy('sales:creditnote_list')
    
    def delete(self, request, *args, **kwargs):
        creditnote = self.get_object()
        note_number = creditnote.note_number
        
        # تسجيل النشاط قبل الحذف
        try:
            from core.signals import log_user_activity
            log_user_activity(
                request,
                'delete',
                creditnote,
                _('حذف إشعار دائن رقم %(number)s') % {'number': note_number}
            )
        except Exception:
            pass
        
        messages.success(request, _('تم حذف إشعار الدائن رقم %(number)s بنجاح') % {'number': note_number})
        return super().delete(request, *args, **kwargs)


class SalesReturnDeleteView(LoginRequiredMixin, DeleteView):
    model = SalesReturn
    template_name = 'sales/return_delete.html'
    success_url = reverse_lazy('sales:return_list')
    
    def delete(self, request, *args, **kwargs):
        return_invoice = self.get_object()
        return_number = return_invoice.return_number
        
        # حذف حركات المخزون المرتبطة (فقط إذا كانت موجودة)
        try:
            from inventory.models import InventoryMovement
            inventory_movements = InventoryMovement.objects.filter(
                reference_type='sales_return',
                reference_id=return_invoice.id
            )
            if inventory_movements.exists():
                movement_count = inventory_movements.count()
                inventory_movements.delete()
                print(f"تم حذف {movement_count} حركة مخزون للمردود {return_number}")
            else:
                print(f"لا توجد حركات مخزون للمردود {return_number} (مردود قديم)")
        except ImportError:
            pass
        except Exception as e:
            print(f"خطأ في حذف حركات المخزون: {e}")
        
        # تسجيل النشاط قبل الحذف
        from core.signals import log_activity
        log_activity(
            user=request.user,
            action_type='DELETE',
            obj=return_invoice,
            description=f'تم حذف مردود مبيعات رقم: {return_number}',
            request=request
        )
        
        messages.success(request, f'تم حذف مردود المبيعات رقم {return_number} بنجاح')
        return super().delete(request, *args, **kwargs)


def print_sales_invoice(request, pk):
    """Print sales invoice"""
    invoice = get_object_or_404(SalesInvoice, pk=pk)
    
    # تسجيل نشاط الطباعة
    from core.signals import log_print_activity
    log_print_activity(request, 'sales_invoice', invoice.pk)
    
    context = {
        'invoice': invoice,
        'items': invoice.items.all(),
        'company_settings': CompanySettings.objects.first(),
        'is_print': True,  # لتخصيص العرض للطباعة
    }
    
    return render(request, 'sales/invoice_detail.html', context)


def print_pos_invoice(request, pk):
    """Print POS invoice (for POS printers)"""
    invoice = get_object_or_404(SalesInvoice, pk=pk)
    
    # تسجيل نشاط الطباعة
    from core.signals import log_print_activity
    log_print_activity(request, 'pos_invoice', invoice.pk)
    
    context = {
        'invoice': invoice,
        'items': invoice.items.all(),
        'company_settings': CompanySettings.objects.first(),
        'is_pos_print': True,  # لتخصيص العرض لطباعة POS
    }
    
    return render(request, 'sales/pos_invoice_print.html', context)


@login_required
def pos_view(request):
    """Point of Sale screen"""
    # التحقق من صلاحية الوصول لنقطة البيع
    # السماح للمستخدمين من نوع pos_user بالوصول تلقائياً
    is_pos_user_type = hasattr(request.user, 'user_type') and request.user.user_type == 'pos_user'
    has_pos_permission = request.user.has_perm('sales.can_access_pos') or request.user.has_perm('sales.can_view_pos')
    
    if not (is_pos_user_type or has_pos_permission or request.user.is_superuser):
        messages.error(request, _('You do not have access to the point of sale.'))
        return redirect('core:dashboard')
    
    # إنشاء عميل "Cash Customer" إذا لم يكن موجوداً
    cash_customer, created = CustomerSupplier.objects.get_or_create(
        name='Cash Customer',
        type='customer',
        defaults={
            'city': 'نقطة البيع',
            'is_active': True,
            'credit_limit': 0,
            'balance': 0,
        }
    )
    
    # تسجيل إنشاء العميل في سجل الأنشطة إذا تم إنشاؤه الآن
    if created:
        from core.models import AuditLog
        AuditLog.objects.create(
            user=request.user,
            action_type='create',
            content_type='CustomerSupplier',
            object_id=cash_customer.id,
            description=f'إنشاء عميل نقطة البيع تلقائياً: {cash_customer.name}'
        )

    # التحقق من أن مستخدم POS لديه شفت مفتوح
    if is_pos_user_type and not request.user.is_superuser:
        from .models import POSShift
        current_shift = POSShift.current_shift_for_user(request.user)
        if not current_shift:
            return render(request, 'sales/pos_no_shift.html', {
                'user_name': request.user.get_full_name() or request.user.username,
                'can_manage_pos_shifts': request.user.has_perm('sales.can_manage_pos_shifts') or request.user.has_perm('users.can_access_pos') or request.user.is_superuser,
            })
    else:
        current_shift = None

    # إدارة الصناديق حسب نوع المستخدم
    user_is_admin = request.user.is_superuser or request.user.is_staff
    
    if not user_is_admin:
        # المستخدم العادي: الحصول على صندوقي النقد والبطاقة الخاصين به
        try:
            pos_cashbox, pos_card_cashbox = get_or_create_user_pos_cashboxes(request.user)
            selected_cashbox = pos_cashbox
        except Exception as e:
            messages.error(request, f'خطأ في إعداد صندوق نقطة البيع: {str(e)}')
            return redirect('core:dashboard')
    else:
        # المستخدم Admin/SuperAdmin: لا يتم إنشاء صندوق تلقائي
        selected_cashbox = None
    
    context = {
        'user_name': request.user.get_full_name() or request.user.username,
        'products': Product.objects.filter(is_active=True),
        'customers': CustomerSupplier.objects.filter(
            type__in=['customer', 'both'], 
            is_active=True
        ),
        'payment_methods': SalesInvoice.POS_PAYMENT_METHODS,
        'cash_customer': cash_customer,
        'user_is_admin': user_is_admin,
        'selected_cashbox': selected_cashbox,
        'pos_cashbox': selected_cashbox if not user_is_admin else None,
        'pos_card_cashbox': pos_card_cashbox if not user_is_admin else None,
        'current_shift': current_shift,
        'can_manage_pos_shifts': request.user.has_perm('sales.can_manage_pos_shifts') or request.user.has_perm('users.can_access_pos') or request.user.is_superuser,
    }
    
    # إعدادات العملة
    try:
        # الحصول على العملة الأساسية
        base_currency = Currency.get_base_currency()
        if not base_currency:
            # إذا لم توجد عملة أساسية، استخدم العملة النشطة الأولى
            base_currency = Currency.objects.filter(is_active=True).first()
        context['base_currency'] = base_currency
    except Exception as e:
        # في حالة وجود خطأ، استخدم عملة افتراضية
        context['base_currency'] = None
    
    # إضافة الصناديق للمستخدمين Admin
    if user_is_admin:
        try:
            from cashboxes.models import Cashbox
            context['cashboxes'] = Cashbox.objects.filter(is_active=True)
        except ImportError:
            context['cashboxes'] = []
    
    return render(request, 'sales/pos.html', context)


@login_required
def pos_shifts_view(request):
    """عرض وإدارة شفتات نقطة البيع"""
    if not (request.user.has_perm('sales.can_manage_pos_shifts') or request.user.is_superuser):
        messages.error(request, _('You do not have permission to manage POS shifts.'))
        return redirect('sales:pos')

    from .models import POSShift
    from cashboxes.models import Cashbox
    from django.contrib.auth import get_user_model
    
    User = get_user_model()

    shifts = POSShift.objects.select_related('user', 'opened_by', 'closed_by').order_by('-opened_at')
    
    shift_rows = []
    for shift in shifts:
        # الحصول على صناديق المستخدم للشفت الحالي
        cash_box, card_box = get_or_create_user_pos_cashboxes(shift.user)
        cashboxes = [box for box in [cash_box, card_box] if box]
        
        shift_rows.append({
            'shift': shift,
            'cashboxes': cashboxes,
        })
    
    # الحصول على قائمة مستخدمي POS
    pos_users = User.objects.filter(user_type='pos_user', is_active=True).order_by('username')

    return render(request, 'sales/pos_shifts.html', {
        'shift_rows': shift_rows,
        'pos_users': pos_users,
        'can_manage_pos_shifts': True,
    })


@login_required
@require_http_methods(['POST'])
def open_pos_shift(request):
    """فتح شفت جديد لمستخدم POS"""
    if not (request.user.has_perm('sales.can_manage_pos_shifts') or request.user.is_superuser):
        messages.error(request, _('You do not have permission to open POS shifts.'))
        return redirect('sales:pos_shifts')

    from .models import POSShift
    from django.contrib.auth import get_user_model
    
    User = get_user_model()
    
    # الحصول على معرف مستخدم POS من البيانات المرسلة
    user_id = request.POST.get('pos_user_id')
    
    if not user_id:
        messages.error(request, _('Please select a POS user.'))
        return redirect('sales:pos_shifts')
    
    try:
        pos_user = User.objects.get(id=user_id, user_type='pos_user', is_active=True)
    except User.DoesNotExist:
        messages.error(request, _('Selected POS user not found.'))
        return redirect('sales:pos_shifts')
    
    # التحقق من عدم وجود شفت مفتوح بالفعل
    current_shift = POSShift.current_shift_for_user(pos_user)
    if current_shift:
        messages.info(request, _(f'User {pos_user.username} already has an open POS shift.'))
        return redirect('sales:pos_shifts')

    # الكود الجديد يبدأ هنا
    # نقوم بإنشاء الشفت فقط
    new_shift = POSShift.objects.create(
        user=pos_user,
        opened_by=request.user,
        status='open',
    )

    # تحذير: إذا كان النظام يقوم بإنشاء صناديق تلقائية عند فتح الشفت،
    # تأكد من عدم وجود دالة مثل ensure_shift_cashboxes(new_shift) 
    # في مكان آخر في ملف models.py الخاص بالـ POSShift أو في Signals.
    
    messages.success(request, _(f'POS shift opened for {pos_user.username}.'))
    return redirect('sales:pos_shifts')

@login_required
@require_http_methods(['POST'])
def close_pos_shift(request, pk=None):
    """Close an open POS shift."""
    if not (request.user.has_perm('sales.can_manage_pos_shifts') or request.user.is_superuser):
        messages.error(request, _('You do not have permission to close POS shifts.'))
        return redirect('sales:pos')

    from .models import POSShift
    if pk:
        shift = get_object_or_404(POSShift, pk=pk)
    else:
        shift = POSShift.current_shift_for_user(request.user)
        if not shift:
            messages.error(request, _('No open POS shift found.'))
            return redirect('sales:pos')

    if shift.status != 'open':
        messages.info(request, _('This shift is already closed.'))
        return redirect('sales:pos_shifts')

    shift.status = 'closed'
    shift.closed_by = request.user
    shift.closed_at = timezone.now()
    shift.save(update_fields=['status', 'closed_by', 'closed_at'])

    messages.success(request, _('POS shift closed successfully.'))
    return redirect('sales:pos_shifts')


@login_required
@require_http_methods(['POST'])
def pos_create_invoice(request):
    """Create invoice from point of sale"""
    try:
        # التحقق من صلاحية نقطة البيع
        if not hasattr(request.user, 'has_pos_permission') or not request.user.has_pos_permission():
            return JsonResponse({'success': False, 'message': _('You do not have access to the point of sale.')})
        
        # التحقق من أن الطلب AJAX
        if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': _('Invalid request.')})
        
        # التحقق من وجود البيانات
        if not request.body:
            return JsonResponse({'success': False, 'message': _('No data sent.')})
            
        data = json.loads(request.body)
        
        # التحقق من وجود البيانات المطلوبة
        if not data.get('items'):
            return JsonResponse({'success': False, 'message': _('No items in the invoice.')})
        
        if not data.get('total'):
            return JsonResponse({'success': False, 'message': 'الإجمالي غير صحيح'})

        payment_method = data.get('payment_method')
        if payment_method not in ['cash', 'card']:
            return JsonResponse({'success': False, 'message': 'يجب اختيار طريقة الدفع نقداً أو بطاقة'})

        with transaction.atomic():
            # توليد رقم الفاتورة باستخدام معاينة أولاً
            try:
                sequence = DocumentSequence.objects.get(document_type='pos_invoice')
                invoice_number = sequence.peek_next_number()
            except DocumentSequence.DoesNotExist:
                # في حالة عدم وجود تسلسل نقطة البيع، إنشاء واحد جديد
                try:
                    sequence = DocumentSequence.objects.create(
                        document_type='pos_invoice',
                        prefix='POS-',
                        digits=6,
                        current_number=1
                    )
                    invoice_number = sequence.peek_next_number()
                except Exception as seq_error:
                    # في حالة فشل إنشاء التسلسل، استخدام رقم بسيط
                    print(f"فشل في إنشاء تسلسل pos_invoice: {seq_error}")
                    last_invoice = SalesInvoice.objects.filter(
                        invoice_number__startswith='POS-'
                    ).order_by('-id').first()
                    if last_invoice:
                        try:
                            last_number = int(last_invoice.invoice_number.split('-')[-1])
                            number = last_number + 1
                        except:
                            number = 1
                    else:
                        number = 1
                    invoice_number = f"POS-{number:06d}"
            
            # إنشاء الفاتورة - استخدام Cash Customer دائماً
            try:
                cash_customer = CustomerSupplier.objects.get(name='Cash Customer', type='customer')
            except CustomerSupplier.DoesNotExist:
                # في حالة عدم وجود العميل، إنشاء واحد
                cash_customer = CustomerSupplier.objects.create(
                    name='Cash Customer',
                    type='customer',
                    city='نقطة البيع',
                    is_active=True,
                    credit_limit=0,
                    balance=0,
                )
            
            # التحقق من وجود شفت مفتوح لمستخدم POS
            if hasattr(request.user, 'user_type') and request.user.user_type == 'pos_user':
                from .models import POSShift
                current_shift = POSShift.current_shift_for_user(request.user)
                if not current_shift:
                    return JsonResponse({'success': False, 'message': _('No open POS shift.')})

            # تحديد الصندوق حسب نوع المستخدم
            user_is_admin = request.user.is_superuser or request.user.is_staff
            selected_cashbox = None
            
            if not user_is_admin:
                # المستخدم العادي: استخدام صندوق الكاش أو صندوق البطاقة حسب طريقة الدفع
                try:
                    from cashboxes.models import Cashbox
                    pos_cashbox, pos_card_cashbox = get_or_create_user_pos_cashboxes(request.user)
                    selected_cashbox = pos_card_cashbox if payment_method == 'card' else pos_cashbox
                    if not selected_cashbox:
                        raise Exception(_('Cashbox for selected payment method not found'))
                except Exception as e:
                    return JsonResponse({'success': False, 'message': f'خطأ في تحديد صندوق نقطة البيع: {str(e)}'})
            else:
                # المستخدم Admin: يجب تحديد الصندوق من البيانات المرسلة
                cashbox_id = data.get('cashbox_id')
                if not cashbox_id:
                    return JsonResponse({'success': False, 'message': 'يجب تحديد الصندوق للمستخدمين الإداريين'})
                
                try:
                    from cashboxes.models import Cashbox
                    selected_cashbox = Cashbox.objects.get(id=cashbox_id, is_active=True)
                except Cashbox.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'الصندوق المحدد غير موجود'})
                except Exception as e:
                    return JsonResponse({'success': False, 'message': f'خطأ في تحديد الصندوق: {str(e)}'})
            
            # التحقق من عدم تكرار رقم الفاتورة
            max_attempts = 10
            attempt = 0
            while attempt < max_attempts:
                try:
                    invoice = SalesInvoice.objects.create(
                        invoice_number=invoice_number,
                        customer=cash_customer,  # استخدام Cash Customer دائماً
                        date=date.today(),
                        payment_type='cash',  # دفع نقدي دائماً
                        pos_payment_method=payment_method,
                        cashbox=selected_cashbox,  # ربط الفاتورة بالصندوق المحدد
                        notes=data.get('notes', ''),
                        created_by=request.user,
                        subtotal=Decimal(str(data.get('subtotal', 0))),
                        tax_amount=Decimal(str(data.get('tax_amount', 0))),
                        discount_amount=Decimal(str(data.get('discount_amount', 0))),
                        total_amount=Decimal(str(data.get('total', 0))),
                    )
                    break
                except Exception as e:
                    if 'UNIQUE constraint failed' in str(e):
                        # في حالة تكرار رقم الفاتورة، توليد رقم جديد
                        attempt += 1
                        try:
                            sequence = DocumentSequence.objects.get(document_type='pos_invoice')
                            invoice_number = sequence.get_next_number()
                        except:
                            # رقم بديل
                            import time
                            timestamp = int(time.time())
                            invoice_number = f"POS-{timestamp}"
                        
                        if attempt >= max_attempts:
                            raise Exception('فشل في إنشاء رقم فاتورة فريد')
                    else:
                        raise e
            
            # إضافة عناصر الفاتورة
            for item_data in data.get('items', []):
                product = Product.objects.get(id=item_data['product_id'])
                
                SalesInvoiceItem.objects.create(
                    invoice=invoice,
                    product=product,
                    quantity=Decimal(str(item_data['quantity'])),
                    unit_price=Decimal(str(item_data['unit_price'])),
                    total_amount=Decimal(str(item_data['total'])),
                    tax_rate=Decimal(str(item_data.get('tax_rate', 0))),
                    tax_amount=Decimal(str(item_data.get('tax_amount', 0))),
                )
                
                # حركات المخزون تُنشأ تلقائياً عبر signals
            
            # إنشاء حركة حساب للعميل
            try:
                from accounts.models import AccountTransaction
                import uuid

                # إذا كان هناك عميل وطريقة الدفع ليست نقداً
                if invoice.customer and invoice.payment_type != 'cash':
                    # حساب الإجمالي
                    total_amount = invoice.total_amount

                    # توليد رقم الحركة
                    transaction_number = f"POS-SALE-{uuid.uuid4().hex[:8].upper()}"

                    # حساب الرصيد السابق
                    last_transaction = AccountTransaction.objects.filter(
                        customer_supplier=invoice.customer
                    ).order_by('-created_at').first()

                    previous_balance = last_transaction.balance_after if last_transaction else Decimal('0')

                    # إنشاء حركة مدينة للعميل (زيادة الذمم المدينة)
                    new_balance = previous_balance + total_amount

                    AccountTransaction.objects.create(
                        transaction_number=transaction_number,
                        date=date.today(),
                        customer_supplier=invoice.customer,
                        transaction_type='sales_invoice',
                        direction='debit',
                        amount=total_amount,
                        reference_type='sales_invoice',
                        reference_id=invoice.id,
                        description=f'مبيعات - فاتورة رقم {invoice.invoice_number}',
                        balance_after=new_balance,
                        created_by=request.user
                    )

            except ImportError:
                # في حالة عدم وجود نموذج الحسابات
                pass
            except Exception as account_error:
                print(f"خطأ في إنشاء حركة الحساب: {account_error}")
                # لا نوقف العملية في حالة فشل تسجيل الحركة المالية
                pass
            
        #    # إنشاء القيد المحاسبي للفاتورة
        #    try:
        #        create_sales_invoice_journal_entry(invoice, request.user)
        #    except Exception as journal_error:
        #        print(f"خطأ في إنشاء القيد المحاسبي: {journal_error}")
        #       # لا نوقف العملية في حالة فشل إنشاء القيد المحاسبي
        #       pass
        #    
            
            # تحديث التسلسل بعد نجاح جميع العمليات
            try:
                if 'sequence' in locals() and sequence:
                    # استخراج الرقم من invoice_number
                    if invoice_number.startswith(sequence.prefix):
                        used_number = int(invoice_number[len(sequence.prefix):])
                        sequence.advance_to_at_least(used_number)
            except Exception as seq_update_error:
                print(f"خطأ في تحديث التسلسل: {seq_update_error}")
                # لا نوقف العملية في حالة فشل تحديث التسلسل
                pass
            
            # تسجيل إنشاء الفاتورة في سجل الأنشطة (خارج transaction للتأكد من عدم فشل البيانات الأساسية)
            try:
                from core.models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    action_type='create',
                    content_type='SalesInvoice',
                    object_id=invoice.id,
                    description=f'إنشاء فاتورة نقطة البيع رقم {invoice.invoice_number} - المبلغ: {invoice.total_amount}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except Exception as audit_error:
                print(f"خطأ في تسجيل الأنشطة: {audit_error}")
                # لا نوقف العملية في حالة فشل تسجيل الأنشطة
                pass
            
            # تأكد من إرجاع JSON response بشكل صحيح
            response_data = {
                'success': True, 
                'message': 'تم إنشاء الفاتورة بنجاح',
                'invoice_id': invoice.id,
                'invoice_number': invoice.invoice_number
            }
            return JsonResponse(response_data, safe=True, status=200)
            
    except json.JSONDecodeError as e:
        return JsonResponse({
            'success': False, 
            'message': _('Error parsing sent data')
        }, status=400)
    except Product.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'message': _('Product not found')
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False, 
            'message': f'{_("An error occurred while creating the invoice:")} {str(e)}'
        }, status=500)


@login_required
def pos_get_product(request, product_id):
    """Get product data for point of sale"""
    if not request.user.has_pos_permission():
        return JsonResponse({'success': False, 'message': _('You do not have access to the point of sale.')})
    
    try:
        product = Product.objects.get(id=product_id, is_active=True)
        
        # حساب المخزون الحالي باستخدام property مع معالجة الأخطاء
        try:
            current_stock = product.current_stock
            if current_stock is None:
                current_stock = 0
        except Exception as stock_error:
            # في حالة خطأ في حساب المخزون، استخدم 0
            import sys
            print(f"خطأ في حساب المخزون للمنتج {product_id}: {stock_error}", file=sys.stderr)
            current_stock = 0
        
        # الحصول على نسبة الضريبة (استخدام القيمة المحددة للمنتج فقط)
        tax_rate = float(product.tax_rate or 0)
        
        # Calculate displayed price as sale_price / (1 + tax_rate/100)
        displayed_price = float(product.sale_price)
        if tax_rate > 0:
            displayed_price = displayed_price / (1 + tax_rate / 100)
        
        return JsonResponse({
            'success': True,
            'product': {
                'id': product.id,
                'name': product.name,
                'price': displayed_price,
                'stock': float(current_stock),
                'tax_rate': tax_rate,
                'barcode': product.barcode or '',
                'track_inventory': True,  # افتراض أن جميع المنتجات تتبع المخزون
                'is_service': getattr(product, 'is_service', False), # إضافة هذه السطر لإرسال حالة الخدمة
            }
        })
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'message': _('Product not found.')})
    except Exception as e:
        # تسجيل الخطأ بشكل مفصل
        import sys
        import traceback
        print(f"خطأ في pos_get_product للمنتج {product_id}:", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)
        return JsonResponse({'success': False, 'message': f'حدث خطأ: {str(e)}'})


@login_required
def pos_search_products(request):
    """Search for products in point of sale"""
    if not request.user.has_pos_permission():
        return JsonResponse({'success': False, 'message': _('You do not have access to the point of sale.')})
    
    query = request.GET.get('q', '').strip()
    if len(query) < 2:
        return JsonResponse({'products': []})
    
    try:
        products = Product.objects.filter(
            Q(name__icontains=query) | Q(barcode__icontains=query) | Q(code__icontains=query),
            is_active=True
        )[:20]
        
        products_data = []
        for product in products:
            # حساب المخزون الحالي
            current_stock = product.current_stock
            
            products_data.append({
                'id': product.id,
                'name': product.name,
                'price': float(product.sale_price),
                'stock': float(current_stock) if current_stock is not None else 0,
                'tax_rate': float(product.tax_rate or 0),
                'barcode': product.barcode or '',
                'code': product.code or '',
                'track_inventory': True,  # افتراض أن جميع المنتجات تتبع المخزون
            })
        
        return JsonResponse({'products': products_data})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'حدث خطأ في البحث: {str(e)}'})


class SalesReportView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'sales/sales_report.html'
    
    def test_func(self):
        # التحقق من صلاحية عرض تقارير المبيعات
        return (
            self.request.user.is_superuser or 
            self.request.user.has_perm('reports.can_view_sales_reports') or
            self.request.user.has_sales_permission()
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        from django.utils import timezone
        from datetime import datetime, timedelta
        from django.db.models import Sum, Count
        
        # الحصول على التواريخ من الطلب
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        # التواريخ الافتراضية (بداية السنة الحالية إلى اليوم)
        today = timezone.now().date()
        if not start_date:
            # بداية السنة الحالية
            start_date = today.replace(month=1, day=1).strftime('%Y-%m-%d')
        if not end_date:
            end_date = today.strftime('%Y-%m-%d')
            
        # تحويل التواريخ لاستخدامها في الاستعلامات
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # فواتير المبيعات في الفترة المحددة
        sales_invoices = SalesInvoice.objects.filter(
            date__gte=start_date_obj,
            date__lte=end_date_obj
        ).select_related('customer').order_by('-date')
        
        # إحصائيات المبيعات
        stats = {
            'total_invoices': sales_invoices.count(),
            'total_amount': sales_invoices.aggregate(total=Sum('total_amount'))['total'] or 0,
            'total_tax': sales_invoices.aggregate(total=Sum('tax_amount'))['total'] or 0,
            'total_discount': sales_invoices.aggregate(total=Sum('discount_amount'))['total'] or 0,
        }
        
        # مبيعات حسب العملاء (استبعاد الفواتير بدون عميل)
        customer_sales = sales_invoices.filter(customer__isnull=False).values('customer__name').annotate(
            total_amount=Sum('total_amount'),
            invoice_count=Count('id')
        ).order_by('-total_amount')[:10]  # أفضل 10 عملاء
        
        # مبيعات حسب اليوم
        daily_sales = sales_invoices.extra({'date': "date(date)"}).values('date').annotate(
            total_amount=Sum('total_amount'),
            invoice_count=Count('id')
        ).order_by('date')
        
        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'sales_invoices': sales_invoices,
            'stats': stats,
            'customer_sales': customer_sales,
            'daily_sales': daily_sales,
        })
        
        return context


class SalesStatementView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """View sales statement"""
    template_name = 'sales/sales_statement.html'
    
    def test_func(self):
        return self.request.user.has_perm('sales.can_view_sales_statement') or self.request.user.is_superuser

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # فترة البحث الافتراضية (الشهر الحالي)
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
        
@login_required
@require_POST
def send_invoice_to_jofotara(request, pk):
    """Send sales invoice to JoFotara - Supports both AJAX and Form POST"""
    print(f"DEBUG: I am inside the send_invoice_to_jofotara function for PK: {pk}")
    
    # التحقق من نوع الطلب
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    try:
        # Get the invoice
        invoice = get_object_or_404(SalesInvoice, pk=pk)
        print(f"DEBUG: Fetching status for: {invoice.invoice_number}")
        
        # Check if user has permission
        if not request.user.has_perm('sales.can_send_to_jofotara'):
            error_msg = 'ليس لديك صلاحية إرسال الفواتير إلى JoFotara'
            if is_ajax:
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('sales:invoice_list')
        
        # Import the utility function
        from settings.utils import send_sales_invoice_to_jofotara
        
        import time
        from django.contrib import messages
        from django.utils import timezone
        from settings.utils import get_invoice_status_from_jofotara

        # Send the invoice
        result = send_sales_invoice_to_jofotara(invoice, request.user)
        print(f"DEBUG: FULL SERVER RESPONSE: {result}")
        
        # --- محاولة جلب الحالة إذا كان الـ QR مفقوداً ---
        if result.get('success') and not result.get('qr_code'):
            time.sleep(1.5) # تأخير بسيط للسماح للسيرفر بمعالجة الطلب
            try:
                status_result = get_invoice_status_from_jofotara(invoice.invoice_number)
                
                if status_result.get('success') and status_result.get('qr_code'):
                    result['qr_code'] = status_result['qr_code']
                    result['verification_url'] = status_result.get('verification_url')
                else:
                    if not is_ajax:
                        messages.warning(request, "تم إرسال الفاتورة بنجاح، لكن رمز QR غير جاهز حالياً. يرجى الانتظار دقيقة ثم التحديث.")
            except Exception as e:
                print(f"DEBUG: Status check failed: {e}")
        
        # --- استكمال الحفظ (الجزء المضاف لإتمام العملية) ---
        if result.get('success'):
            if 'uuid' in result:
                invoice.jofotara_uuid = result.get('uuid')
                invoice.jofotara_sent_at = timezone.now()
                invoice.jofotara_verification_url = result.get('verification_url')
                
                qr_val = (result.get('qr_code') or "").strip()
                
                if qr_val:
                    invoice.jofotara_qr_code = qr_val if qr_val.startswith('data:image') else f"data:image/png;base64,{qr_val}"
                else:
                    invoice.jofotara_qr_code = None 
                
                invoice.is_posted_to_tax = True
                invoice.save()
        
        # استكمال الحفظ
        if result['success']:
            if 'uuid' in result:
                invoice.jofotara_uuid = result.get('uuid')
                invoice.jofotara_sent_at = timezone.now()
                invoice.jofotara_verification_url = result.get('verification_url')
                
                qr_val = (result.get('qr_code') or "").strip()
                
                if qr_val:
                    invoice.jofotara_qr_code = qr_val if qr_val.startswith('data:image') else f"data:image/png;base64,{qr_val}"
                else:
                    invoice.jofotara_qr_code = None 
                
                invoice.is_posted_to_tax = True
                invoice.save()
                
                # تسجيل في سجل الأنشطة
                try:
                    from core.models import AuditLog
                    AuditLog.objects.create(
                        user=request.user,
                        action_type='post_to_tax',
                        content_type='SalesInvoice',
                        object_id=invoice.id,
                        description=f'تم إرسال الفاتورة {invoice.invoice_number} إلى JoFotara بنجاح',
                        ip_address=request.META.get('REMOTE_ADDR')
                    )
                except Exception:
                    pass
            
            msg = f'تم إرسال الفاتورة {invoice.invoice_number} إلى JoFotara بنجاح.'
            if not invoice.jofotara_qr_code:
                msg += " (لم يتم استلام رمز QR)."
            
            if not is_ajax:
                messages.success(request, msg)
        else:
            # فشل الإرسال
            invoice.is_posted_to_tax = False
            invoice.save()
            error_msg = f'فشل إرسال الفاتورة {invoice.invoice_number}: {result.get("error", "خطأ غير معروف")}'
            if not is_ajax:
                messages.error(request, error_msg)
        
        # الرد بناءً على نوع الطلب
        if is_ajax:
            return JsonResponse(result)
        return redirect('sales:invoice_list')
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error sending invoice to JoFotara: {str(e)}")
        if is_ajax:
            return JsonResponse({'success': False, 'error': f'خطأ في النظام: {str(e)}'})
        messages.error(request, f'خطأ في النظام: {str(e)}')
        return redirect('sales:invoice_list')

@login_required
@require_POST
def send_creditnote_to_jofotara(request, pk):
    """Send credit note to JoFotara"""
    # للطلبات AJAX، نتحقق من الـ header ونعيد JSON response
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    if not is_ajax:
        return JsonResponse({
            'success': False,
            'error': 'هذه الدالة تستخدم للطلبات AJAX فقط'
        }, status=400)
    
    try:
        # Get the credit note
        credit_note = get_object_or_404(SalesCreditNote, pk=pk)
        
        # Check if user has permission to send credit notes
        if not request.user.has_perm('sales.can_send_to_jofotara'):
            return JsonResponse({
                'success': False,
                'error': 'ليس لديك صلاحية إرسال الإشعارات الدائنة إلى JoFotara'
            })
        
        # Import the utility function
        from settings.utils import send_credit_note_to_jofotara
        
        # Send the credit note
        result = send_credit_note_to_jofotara(credit_note, request.user)
        
        if result['success']:
            # Update credit note with JoFotara data
            if 'uuid' in result:
                credit_note.jofotara_uuid = result['uuid']
                credit_note.jofotara_sent_at = timezone.now()
                credit_note.jofotara_verification_url = result.get('verification_url')
                credit_note.jofotara_qr_code = result.get('qr_code')  # حفظ QR Code
                credit_note.is_posted_to_tax = True  # تم الترحيل بنجاح
                credit_note.save()
                
                # تسجيل في سجل الأنشطة
                try:
                    from core.models import AuditLog
                    AuditLog.objects.create(
                        user=request.user,
                        action_type='post_to_tax',
                        content_type='SalesCreditNote',
                        object_id=credit_note.id,
                        description=f'تم إرسال الإشعار الدائن {credit_note.note_number} إلى JoFotara بنجاح - UUID: {result["uuid"]}',
                        ip_address=request.META.get('REMOTE_ADDR')
                    )
                except Exception:
                    pass
            
            # تحقق من وجود QR Code في الاستجابة
            if not result.get('qr_code'):
                messages.warning(request, f'تم إرسال الإشعار الدائن {credit_note.note_number} إلى JoFotara لكن لم يتم استلام رمز QR. يرجى التحقق من إعدادات JoFotara.')
            else:
                messages.success(request, f'تم إرسال الإشعار الدائن {credit_note.note_number} إلى JoFotara بنجاح وحفظ رمز QR')
        else:
            # فشل الإرسال
            credit_note.is_posted_to_tax = False
            credit_note.save()
            
            # تسجيل الفشل في سجل الأنشطة
            try:
                from core.models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    action_type='error',
                    content_type='SalesCreditNote',
                    object_id=credit_note.id,
                    description=f'فشل إرسال الإشعار الدائن {credit_note.note_number} إلى JoFotara: {result.get("error", "خطأ غير معروف")}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except Exception:
                pass
            
            messages.error(request, f'فشل في إرسال الإشعار الدائن: {result.get("error", "خطأ غير معروف")}')
        
        return JsonResponse(result)
        
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.error(f"Error sending credit note to JoFotara: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'خطأ في النظام: {str(e)}'
        })


@login_required
@require_POST
def send_return_to_jofotara(request, pk):
    """Send sales return to JoFotara"""
    # للطلبات AJAX، نتحقق من الـ header ونعيد JSON response
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    if not is_ajax:
        return JsonResponse({
            'success': False,
            'error': 'هذه الدالة تستخدم للطلبات AJAX فقط'
        }, status=400)
    
    try:
        # Get the return
        sales_return = get_object_or_404(SalesReturn, pk=pk)
        
        # Check if user has permission to send returns
        if not request.user.has_perm('sales.can_send_to_jofotara'):
            return JsonResponse({
                'success': False,
                'error': 'ليس لديك صلاحية إرسال المرتجعات إلى JoFotara'
            })
        
        # Import the utility function
        from settings.utils import send_return_to_jofotara as send_return_api
        
        # Send the return
        result = send_return_api(sales_return, request.user)
        
        if result['success']:
            # Update return with JoFotara data
            if 'uuid' in result:
                sales_return.jofotara_uuid = result['uuid']
                sales_return.jofotara_sent_at = timezone.now()
                sales_return.jofotara_verification_url = result.get('verification_url')
                sales_return.jofotara_qr_code = result.get('qr_code')  # حفظ QR Code
                sales_return.is_posted_to_tax = True  # تم الترحيل بنجاح
                sales_return.save()
                
                # تسجيل في سجل الأنشطة
                try:
                    from core.models import AuditLog
                    AuditLog.objects.create(
                        user=request.user,
                        action_type='post_to_tax',
                        content_type='SalesReturn',
                        object_id=sales_return.id,
                        description=f'تم إرسال المرتجع {sales_return.return_number} إلى JoFotara بنجاح - UUID: {result["uuid"]}',
                        ip_address=request.META.get('REMOTE_ADDR')
                    )
                except Exception:
                    pass
            
            # تحقق من وجود QR Code في الاستجابة
            if not result.get('qr_code'):
                messages.warning(request, f'تم إرسال المرتجع {sales_return.return_number} إلى JoFotara لكن لم يتم استلام رمز QR. يرجى التحقق من إعدادات JoFotara.')
            else:
                messages.success(request, f'تم إرسال المرتجع {sales_return.return_number} إلى JoFotara بنجاح وحفظ رمز QR')
        else:
            # فشل الإرسال
            sales_return.is_posted_to_tax = False
            sales_return.save()
            
            # تسجيل الفشل في سجل الأنشطة
            try:
                from core.models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    action_type='error',
                    content_type='SalesReturn',
                    object_id=sales_return.id,
                    description=f'فشل إرسال المرتجع {sales_return.return_number} إلى JoFotara: {result.get("error", "خطأ غير معروف")}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except Exception:
                pass
            
            messages.error(request, f'فشل في إرسال المرتجع: {result.get("error", "خطأ غير معروف")}')
        
        return JsonResponse(result)
        
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.error(f"Error sending return to JoFotara: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'خطأ في النظام: {str(e)}'
        })


@login_required
@csrf_exempt
def get_invoices_for_returns(request):
    """API endpoint لجلب الفواتير المتاحة للمرتجعات مع البحث"""
    print(f"API called by user: {request.user}")
    print(f"Request method: {request.method}")
    print(f"Request GET params: {request.GET}")
    
    try:
        search_term = request.GET.get('search', '').strip()
        print(f"Search term: '{search_term}'")
        
        # جلب الفواتير مع إمكانية البحث
        invoices = SalesInvoice.objects.select_related('customer').all()
        print(f"Total invoices in database: {invoices.count()}")
        
        if search_term:
            invoices = invoices.filter(
                Q(invoice_number__icontains=search_term) |
                Q(customer__name__icontains=search_term)
            )
            print(f"Filtered invoices count: {invoices.count()}")
        
        # تحديد الحد الأقصى للنتائج
        invoices = invoices[:20]
        
        invoices_data = []
        for invoice in invoices:
            print(f"Processing invoice: {invoice.invoice_number}")
            # جلب عناصر الفاتورة
            items_data = []
            for item in invoice.items.select_related('product').all():
                items_data.append({
                    'product_id': item.product.id,
                    'product_name': item.product.name,
                    'product_code': item.product.code,
                    'quantity': float(item.quantity),
                    'unit_price': float(item.unit_price),
                    'tax_rate': float(item.tax_rate),
                    'tax_amount': float(item.tax_amount),
                    'total': float(item.total_amount)
                })
            
            invoices_data.append({
                'id': invoice.id,
                'invoice_number': invoice.invoice_number,
                'customer': invoice.customer.name if invoice.customer else '',
                'customer_id': invoice.customer.id if invoice.customer else None,
                'date': invoice.date.strftime('%Y-%m-%d'),
                'total': float(invoice.total_amount),
                'payment_type': invoice.get_payment_type_display(),
                'items': items_data
            })
        
        print(f"Returning {len(invoices_data)} invoices")
        return JsonResponse({
            'success': True,
            'invoices': invoices_data
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error fetching invoices for returns: {str(e)}")
        print(f"Error in API: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'خطأ في النظام: {str(e)}'
        })


def get_invoice_items(request, invoice_id):
    """Fetch invoice items via AJAX for returns"""
    try:
        print(f"طلب جلب عناصر الفاتورة رقم: {invoice_id}")
        
        invoice = SalesInvoice.objects.get(id=invoice_id)
        items = invoice.items.select_related('product').all()
        
        print(f"تم العثور على {items.count()} عنصر في الفاتورة")
        
        items_data = []
        returnable_items = []
        for item in items:
            # Calculate total returned quantity for this product in returns for this invoice
            total_returned = SalesReturnItem.objects.filter(
                return_invoice__original_invoice=invoice,
                product=item.product
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            remaining_quantity = item.quantity - total_returned
            
            print(f"المنتج {item.product.name}: الكمية الأصلية={item.quantity}, المرتجع={total_returned}, المتبقي={remaining_quantity}")
            
            item_data = {
                'id': item.id,
                'product_id': item.product.id,
                'product_name': item.product.name,
                'product_code': item.product.code,
                'original_quantity': float(item.quantity),
                'returned_quantity': float(total_returned),
                'remaining_quantity': float(remaining_quantity),
                'unit_price': float(item.unit_price),
                'tax_rate': float(item.tax_rate),
            }
            
            items_data.append(item_data)  # All items for display
            
            if remaining_quantity > 0:  # Only returnable items for return form
                returnable_items.append(item_data)
        
        print(f"عدد العناصر الكلية: {len(items_data)}, المتاحة للإرجاع: {len(returnable_items)}")
        
        return JsonResponse({
            'success': True,
            'all_items': items_data,  # All invoice items for display
            'returnable_items': returnable_items,  # Items available for return
            'invoice_number': invoice.invoice_number,
            'customer_name': invoice.customer.name,
            'invoice_date': invoice.date.strftime('%Y-%m-%d'),
        })
        
    except SalesInvoice.DoesNotExist:
        print(f"الفاتورة رقم {invoice_id} غير موجودة")
        return JsonResponse({
            'success': False,
            'message': 'الفاتورة غير موجودة'
        })
    except Exception as e:
        print(f"خطأ في جلب عناصر الفاتورة: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })


class SalesCreditNoteReportView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """Credit notes statement"""
    model = SalesCreditNote
    template_name = 'sales/creditnote_report.html'
    context_object_name = 'creditnotes'
    paginate_by = 50
    
    def test_func(self):
        # التحقق من صلاحية عرض تقارير المبيعات
        return (
            self.request.user.is_superuser or 
            self.request.user.has_perm('reports.can_view_sales_reports') or
            self.request.user.has_sales_permission()
        )
    
    def get(self, request, *args, **kwargs):
        # تسجيل النشاط
        try:
            from core.signals import log_user_activity
            log_user_activity(
                request,
                'view',
                None,
                _('عرض كشف مذكرات الدائن')
            )
        except Exception:
            pass
        
        return super().get(request, *args, **kwargs)
    
    def get_queryset(self):
        queryset = SalesCreditNote.objects.select_related('customer', 'created_by').all()
        
        # فلترة حسب العميل
        customer_id = self.request.GET.get('customer')
        if customer_id:
            queryset = queryset.filter(customer_id=customer_id)
        
        # فلترة حسب التاريخ
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        
        return queryset.order_by('-date', '-note_number')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # إضافة إجماليات
        queryset = self.get_queryset()
        context['total_amount'] = queryset.aggregate(
            total=Sum('total_amount')
        )['total'] or 0
        
        context['total_count'] = queryset.count()
        
        # إضافة قائمة العملاء للفلترة
        from customers.models import CustomerSupplier
        context['customers'] = CustomerSupplier.objects.filter(
            type__in=['customer', 'both']
        ).order_by('name')
        
        # إضافة قيم الفلترة الحالية
        context['filters'] = {
            'date_from': self.request.GET.get('date_from', ''),
            'date_to': self.request.GET.get('date_to', ''),
            'customer': self.request.GET.get('customer', ''),
        }
        
        try:
            company_settings = CompanySettings.objects.first()
            if company_settings and company_settings.base_currency:
                context['base_currency'] = company_settings.base_currency
        except Exception:
            pass
            
        return context


# AJAX Views for Invoice Item Management
@login_required
@require_POST
def invoice_add_item(request, invoice_id):
    """Add new item to sales invoice via AJAX"""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        invoice = get_object_or_404(SalesInvoice, pk=invoice_id)
        
        # Check permissions
        if not request.user.has_sales_permission():
            return JsonResponse({'success': False, 'message': _('ليس لديك صلاحية لتعديل فواتير المبيعات')})
        
        # Handle both JSON and form-data POST requests
        if request.content_type == 'application/json':
            import json
            body = json.loads(request.body)
            product_id = body.get('product_id')
            quantity = body.get('quantity')
            unit_price = body.get('unit_price')
            tax_rate = body.get('tax_rate')
        else:
            product_id = request.POST.get('product_id')
            quantity = request.POST.get('quantity')
            unit_price = request.POST.get('unit_price')
            tax_rate = request.POST.get('tax_rate')
        
        # التحقق من الحقول المطلوبة (tax_rate يمكن أن يكون صفر)
        if not product_id or not quantity or not unit_price or (tax_rate is None or tax_rate == ''):
            logger.error(f"إضافة منتج فشلت - بيانات ناقصة: product_id={product_id}, quantity={quantity}, unit_price={unit_price}, tax_rate={tax_rate}")
            return JsonResponse({'success': False, 'message': _('جميع الحقول مطلوبة')})
        
        product = get_object_or_404(Product, pk=product_id)
        
        quantity = Decimal(quantity)
        unit_price = Decimal(unit_price)
        tax_rate = Decimal(tax_rate)
        
        # Calculate amounts
        line_subtotal = quantity * unit_price
        line_tax_amount = line_subtotal * (tax_rate / Decimal('100'))
        line_total = line_subtotal + line_tax_amount
        
        # Create the item
        item = SalesInvoiceItem.objects.create(
            invoice=invoice,
            product=product,
            quantity=quantity,
            unit_price=unit_price,
            tax_rate=tax_rate,
            tax_amount=line_tax_amount.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP),
            total_amount=line_total.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
        )
        
        # Update invoice totals
        invoice.update_totals()
        
        # Update journal entries for the invoice
        try:
            # Delete existing journal entries for this invoice
            from journal.models import JournalEntry
            JournalEntry.objects.filter(
                reference_type__in=['sales_invoice', 'sales_invoice_cogs'],
                reference_id=invoice.id
            ).delete()
            
            # Create new journal entries with updated amounts
            create_sales_invoice_journal_entry(invoice, request.user)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f'تحذير: لم يتمكن من تحديث القيود المحاسبية: {str(e)}')
            # لا نوقف العملية إذا فشل إنشاء القيود
            pass
        
        # Log activity
        try:
            from core.signals import log_user_activity
            log_user_activity(
                request,
                'create',
                item,
                f'إضافة منتج {product.name} لفاتورة المبيعات {invoice.invoice_number}'
            )
        except Exception:
            pass
        
        return JsonResponse({
            'success': True, 
            'message': _('تم إضافة المنتج بنجاح'),
            'item': {
                'id': item.id,
                'product_name': product.name,
                'quantity': float(item.quantity),
                'unit_price': float(item.unit_price),
                'tax_rate': float(item.tax_rate),
                'tax_amount': float(item.tax_amount),
                'total_amount': float(item.total_amount)
            }
        })
        
    except Exception as e:
        logger.error(f"خطأ في إضافة منتج للفاتورة {invoice_id}: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': _('حدث خطأ أثناء إضافة المنتج')})


@login_required
@require_POST
def invoice_update_item(request, invoice_id, item_id):
    """Update sales invoice item via AJAX"""
    try:
        invoice = get_object_or_404(SalesInvoice, pk=invoice_id)
        item = get_object_or_404(SalesInvoiceItem, pk=item_id, invoice=invoice)
        
        # Check permissions
        if not request.user.has_sales_permission():
            return JsonResponse({'success': False, 'message': 'ليس لديك صلاحية لتعديل فواتير المبيعات'})
        
        # Handle both JSON and form-data POST requests
        if request.content_type == 'application/json':
            import json
            body = json.loads(request.body)
            quantity = body.get('quantity')
            unit_price = body.get('unit_price')
        else:
            quantity = request.POST.get('quantity')
            unit_price = request.POST.get('unit_price')
        
        if not all([quantity, unit_price]):
            return JsonResponse({'success': False, 'message': 'الكمية وسعر الوحدة مطلوبان'})
        
        old_quantity = item.quantity
        old_price = item.unit_price
        
        try:
            quantity = Decimal(quantity)
            unit_price = Decimal(unit_price)
        except:
            return JsonResponse({'success': False, 'message': 'الكمية وسعر الوحدة يجب أن تكون أرقام صحيحة'})
        
        # Check if there were changes
        has_changes = (quantity != old_quantity) or (unit_price != old_price)
        
        if not has_changes:
            # No changes needed
            return JsonResponse({
                'success': True, 
                'message': 'لا توجد تغييرات لحفظها',
                'item': {
                    'tax_amount': float(item.tax_amount),
                    'total_amount': float(item.total_amount)
                }
            })
        
        # Calculate new amounts
        line_subtotal = quantity * unit_price
        line_tax_amount = line_subtotal * (item.tax_rate / Decimal('100'))
        line_total = line_subtotal + line_tax_amount
        
        # Update the item
        item.quantity = quantity
        item.unit_price = unit_price
        item.tax_amount = line_tax_amount.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
        item.total_amount = line_total.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
        item.save()
        
        # Update invoice totals
        invoice.update_totals()
        
        # Refresh invoice from database to ensure we have latest data
        invoice.refresh_from_db()
        
        # Update or create journal entries for the invoice
        try:
            # Delete existing journal entries for this invoice
            from journal.models import JournalEntry
            JournalEntry.objects.filter(
                reference_type__in=['sales_invoice', 'sales_invoice_cogs'],
                reference_id=invoice.id
            ).delete()
            
            # Create new journal entries with updated amounts
            create_sales_invoice_journal_entry(invoice, request.user)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f'تحذير: لم يتمكن من تحديث القيود المحاسبية: {str(e)}')
            # لا نوقف العملية إذا فشل إنشاء القيود
            pass
        
        # Log activity with detailed changes
        try:
            from core.signals import log_user_activity
            change_details = []
            if quantity != old_quantity:
                change_details.append(f'الكمية من {old_quantity} إلى {quantity}')
            if unit_price != old_price:
                change_details.append(f'سعر الوحدة من {old_price} إلى {unit_price}')
            
            changes_text = '، '.join(change_details)
            log_user_activity(
                request,
                'update',
                item,
                f'تحديث عنصر {item.product.name} في فاتورة المبيعات {invoice.invoice_number}: {changes_text}'
            )
        except Exception:
            pass
        
        return JsonResponse({
            'success': True, 
            'message': 'تم تحديث العنصر وحفظ التغييرات بنجاح',
            'item': {
                'id': item.id,
                'product_name': item.product.name,
                'quantity': str(item.quantity),
                'unit_price': str(item.unit_price),
                'tax_amount': str(item.tax_amount),
                'total_amount': str(item.total_amount)
            },
            'invoice': {
                'subtotal': str(invoice.subtotal),
                'tax_amount': str(invoice.tax_amount),
                'total_amount': str(invoice.total_amount),
                'discount_amount': str(invoice.discount_amount)
            }
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'خطأ في تحديث عنصر الفاتورة: {str(e)}')
        return JsonResponse({'success': False, 'message': f'خطأ: {str(e)}'})


@login_required
@require_POST
def invoice_delete_item(request, invoice_id, item_id):
    """Delete item from sales invoice via AJAX"""
    try:
        invoice = get_object_or_404(SalesInvoice, pk=invoice_id)
        item = get_object_or_404(SalesInvoiceItem, pk=item_id, invoice=invoice)
        
        # Check permissions
        if not request.user.has_sales_permission():
            return JsonResponse({'success': False, 'message': 'ليس لديك صلاحية لتعديل فواتير المبيعات'})
        
        product_name = item.product.name
        
        # Delete the item
        item.delete()
        
        # Update invoice totals
        invoice.update_totals()
        
        # Update journal entries for the invoice
        try:
            # Delete existing journal entries for this invoice
            from journal.models import JournalEntry
            JournalEntry.objects.filter(
                reference_type__in=['sales_invoice', 'sales_invoice_cogs'],
                reference_id=invoice.id
            ).delete()
            
            # Create new journal entries with updated amounts
            create_sales_invoice_journal_entry(invoice, request.user)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f'تحذير: لم يتمكن من تحديث القيود المحاسبية: {str(e)}')
            # لا نوقف العملية إذا فشل إنشاء القيود
            pass
        
        # Log activity
        try:
            from core.signals import log_user_activity
            log_user_activity(
                request,
                'delete',
                invoice,
                f'حذف منتج {product_name} من فاتورة المبيعات {invoice.invoice_number}'
            )
        except Exception:
            pass
        
        return JsonResponse({
            'success': True, 
            'message': 'تم حذف العنصر بنجاح',
            'invoice': {
                'subtotal': str(invoice.subtotal),
                'tax_amount': str(invoice.tax_amount),
                'total_amount': str(invoice.total_amount),
                'discount_amount': str(invoice.discount_amount)
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'خطأ: {str(e)}'})


@login_required
@require_http_methods(["GET", "POST"])
def invoice_post_to_tax(request, pk):
    """
    Post sales invoice to tax authority (JoFotara) and update QR
    """
    invoice = get_object_or_404(SalesInvoice, pk=pk)
    
    # Check permissions
    if not (request.user.is_superuser or request.user.has_perm('sales.can_post_sales')):
        messages.error(request, _('You do not have permission to post invoices to tax authority'))
        return redirect('sales:invoice_detail', pk=pk)
    
    # استيراد الوظائف المطلوبة
    import time
    from django.utils import timezone
    from settings.utils import get_invoice_status_from_jofotara
    from sales.utils import send_sales_invoice_to_jofotara # الدالة التي ترسل الفاتورة

    # 1. عملية الإرسال الفعلي للسيرفر
    result = send_sales_invoice_to_jofotara(invoice, request.user)
    
    # 2. إذا نجح الإرسال ولكن لا يوجد QR، حاول الاستعلام
    if result.get('success') and not result.get('qr_code'):
        time.sleep(1.5)
        status_result = get_invoice_status_from_jofotara(invoice.invoice_number)
        if status_result.get('success') and status_result.get('qr_code'):
            result['qr_code'] = status_result['qr_code']
            result['verification_url'] = status_result.get('verification_url')

    # 3. الحفظ في قاعدة البيانات
    if result.get('success'):
        invoice.is_posted_to_tax = True
        invoice.jofotara_sent_at = timezone.now()
        invoice.jofotara_uuid = result.get('uuid')
        
        qr_val = (result.get('qr_code') or "").strip()
        if qr_val:
            invoice.jofotara_qr_code = qr_val if qr_val.startswith('data:image') else f"data:image/png;base64,{qr_val}"
        
        invoice.save()
        messages.success(request, _('Invoice has been successfully posted to tax authority'))
    else:
        messages.error(request, f"خطأ في الاتصال بالضريبة: {result.get('error', 'Unknown Error')}")
    
    return redirect('sales:invoice_detail', pk=pk)


@login_required
@require_http_methods(["GET", "POST"])
def return_post_to_tax(request, pk):
    """
    Post sales return to tax authority (JoFotara)
    """
    sales_return = get_object_or_404(SalesReturn, pk=pk)
    
    # Check permissions
    if not (request.user.is_superuser or request.user.has_perm('sales.can_post_sales_returns')):
        messages.error(request, _('You do not have permission to post returns to tax authority'))
        return redirect('sales:return_list')
    
    # Check if already posted
    if sales_return.is_posted_to_tax:
        messages.warning(request, _('This return has already been posted to tax authority'))
        return redirect('sales:return_list')
    
    try:
        with transaction.atomic():
            # Update return status
            sales_return.is_posted_to_tax = True
            sales_return.jofotara_sent_at = timezone.now()
            sales_return.save(update_fields=['is_posted_to_tax', 'jofotara_sent_at'])
            
            # Log activity
            try:
                from core.signals import log_user_activity
                log_user_activity(
                    request,
                    'update',
                    sales_return,
                    f'تم ترحيل مرتجع المبيعات {sales_return.return_number} إلى إدارة الضريبة'
                )
            except Exception:
                pass
            
            messages.success(request, _('Sales return has been successfully posted to tax authority'))
            
    except Exception as e:
        messages.error(request, f'خطأ في ترحيل المرتجع: {str(e)}')
    
    return redirect('sales:return_list')


@login_required
@require_http_methods(["GET", "POST"])
def creditnote_post_to_tax(request, pk):
    """
    Post credit note to tax authority (JoFotara)
    """
    creditnote = get_object_or_404(SalesCreditNote, pk=pk)
    
    # Check permissions
    if not (request.user.is_superuser or request.user.has_perm('sales.can_post_credit_notes')):
        messages.error(request, _('You do not have permission to post credit notes to tax authority'))
        return redirect('sales:creditnote_list')
    
    # Check if already posted
    if creditnote.is_posted_to_tax:
        messages.warning(request, _('This credit note has already been posted to tax authority'))
        return redirect('sales:creditnote_list')
    
    try:
        with transaction.atomic():
            # Update credit note status
            creditnote.is_posted_to_tax = True
            creditnote.jofotara_sent_at = timezone.now()
            creditnote.save(update_fields=['is_posted_to_tax', 'jofotara_sent_at'])
            
            # Log activity
            try:
                from core.signals import log_user_activity
                log_user_activity(
                    request,
                    'update',
                    creditnote,
                    f'تم ترحيل إشعار دائن {creditnote.note_number} إلى إدارة الضريبة'
                )
            except Exception:
                pass
            
            messages.success(request, _('Credit note has been successfully posted to tax authority'))
            
    except Exception as e:
        messages.error(request, f'خطأ في ترحيل الإشعار الدائن: {str(e)}')
    
    return redirect('sales:creditnote_list')

class SalesReturnStatementView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'sales/sales_return_statement.html'
    
    def test_func(self):
        return self.request.user.has_perm('sales.can_view_sales_returns_statement') or self.request.user.is_superuser

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # فترة البحث الافتراضية (الشهر الحالي)
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
        
        # تطبيق الفلاتر من الطلب
        if self.request.GET.get('start_date'):
            start_date = datetime.strptime(self.request.GET.get('start_date'), '%Y-%m-%d').date()
        if self.request.GET.get('end_date'):
            end_date = datetime.strptime(self.request.GET.get('end_date'), '%Y-%m-%d').date()
        
        # جلب مردودات المبيعات مرتبة حسب التاريخ
        from sales.models import SalesReturn # تأكد من الاستيراد
        sales_returns = SalesReturn.objects.filter(
            date__range=[start_date, end_date]
        ).order_by('date', 'return_number')
        
        # حساب الرصيد التراكمي
        running_balance = Decimal('0')
        statement_data = []
        
        for return_item in sales_returns:
            debit_amount = return_item.subtotal
            running_balance += debit_amount
            
            statement_data.append({
                'date': return_item.date,
                'document_number': return_item.return_number,
                'description': 'مردود مبيعات',
                'debit': debit_amount,
                'balance': running_balance,
                'return_item': return_item
            })
        
        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'statement_data': statement_data,
            'final_balance': running_balance,
        })
        
        return context