#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
اختبار أزرار النسخ الاحتياطي في واجهة المستخدم
اختبار شامل للتأكد من عمل الأزرار بشكل صحيح وإنتاج ملفات قابلة للاستخدام
"""

import os
import sys
import json
import time
import requests
from datetime import datetime
from pathlib import Path

# إضافة مجلد المشروع إلى المسار
BASE_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(BASE_DIR))

# إعداد Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'finspilot.settings')
import django
django.setup()

from django.contrib.auth import get_user_model
from django.test import Client
from django.urls import reverse
from backup.views import perform_backup_restore, load_backup_from_xlsx
import openpyxl

class BackupUIButtonTester:
    """فئة لاختبار أزرار النسخ الاحتياطي في الواجهة"""

    def __init__(self):
        self.client = Client()
        self.base_url = 'http://127.0.0.1:8000'
        self.session = requests.Session()
        self.test_results = []
        self.backup_files_created = []

    def login(self):
        """تسجيل الدخول للحصول على الجلسة"""
        try:
            # الحصول على نموذج المستخدم
            User = get_user_model()

            # إنشاء مستخدم اختبار إذا لم يكن موجوداً
            test_user, created = User.objects.get_or_create(
                username='test_backup_user',
                defaults={
                    'email': 'test@example.com',
                    'is_staff': True,
                    'is_superuser': True
                }
            )

            if created:
                test_user.set_password('testpass123')
                test_user.save()
                print("✅ تم إنشاء مستخدم اختبار")

            # تسجيل الدخول مباشرة باستخدام requests
            login_url = f"{self.base_url}/ar/accounts/login/"
            
            # الحصول على صفحة تسجيل الدخول أولاً للحصول على CSRF token
            response = self.session.get(login_url)
            if 'csrfmiddlewaretoken' in response.text:
                import re
                csrf_match = re.search(r'name="csrfmiddlewaretoken" value="([^"]+)"', response.text)
                if csrf_match:
                    csrf_token = csrf_match.group(1)
                else:
                    raise Exception("لم يتم العثور على CSRF token في صفحة تسجيل الدخول")
            else:
                raise Exception("لم يتم العثور على CSRF token في صفحة تسجيل الدخول")

            # إرسال بيانات تسجيل الدخول
            login_data = {
                'username': 'test_backup_user',
                'password': 'testpass123',
                'csrfmiddlewaretoken': csrf_token
            }
            
            headers = {
                'Referer': login_url
            }
            
            response = self.session.post(login_url, data=login_data, headers=headers)
            
            if response.status_code == 200 and 'login' not in response.url.lower():
                print("✅ تم تسجيل الدخول بنجاح")
                return True
            else:
                raise Exception(f"فشل في تسجيل الدخول - HTTP {response.status_code}")

        except Exception as e:
            print(f"❌ خطأ في تسجيل الدخول: {str(e)}")
            return False

    def test_backup_button(self, format_type):
        """اختبار زر النسخ الاحتياطي لتنسيق معين"""
        print(f"\n🔄 اختبار زر النسخ الاحتياطي {format_type.upper()}...")

        start_time = time.time()

        try:
            # إرسال طلب POST لإنشاء النسخة الاحتياطية
            url = f"{self.base_url}/ar/backup/create-backup/"
            headers = {
                'X-Requested-With': 'XMLHttpRequest',
                'Referer': f"{self.base_url}/ar/backup/"
            }

            # الحصول على CSRF token من الصفحة
            page_response = self.session.get(f"{self.base_url}/ar/backup/")
            if 'csrfmiddlewaretoken' in page_response.text:
                import re
                csrf_match = re.search(r'name="csrfmiddlewaretoken" value="([^"]+)"', page_response.text)
                if csrf_match:
                    csrf_token = csrf_match.group(1)
                else:
                    raise Exception("لم يتم العثور على CSRF token")
            else:
                raise Exception("لم يتم العثور على CSRF token في الصفحة")

            data = {
                'format': format_type,
                'csrfmiddlewaretoken': csrf_token
            }

            response = self.session.post(url, data=data, headers=headers)

            if response.status_code != 200:
                raise Exception(f"HTTP {response.status_code}: {response.text}")

            result = response.json()

            if not result.get('success'):
                raise Exception(f"فشل في إنشاء النسخة الاحتياطية: {result.get('error', 'خطأ غير معروف')}")

            filename = result.get('filename')
            if not filename:
                raise Exception("لم يتم إرجاع اسم الملف")

            print(f"✅ تم بدء إنشاء النسخة الاحتياطية: {filename}")

            # انتظار انتهاء العملية (مراقبة وجود الملف)
            self.wait_for_backup_completion(filename)

            # التحقق من إنشاء الملف
            backup_dir = BASE_DIR / 'media' / 'backups'
            filepath = backup_dir / filename

            if not filepath.exists():
                raise Exception(f"لم يتم إنشاء ملف النسخة الاحتياطية: {filepath}")

            file_size = filepath.stat().st_size
            print(f"✅ تم إنشاء الملف: {filename} (الحجم: {file_size:,} بايت)")

            # التحقق من صحة الملف
            self.validate_backup_file(filepath, format_type)

            duration = time.time() - start_time

            self.backup_files_created.append({
                'filename': filename,
                'format': format_type,
                'filepath': str(filepath),
                'size': file_size,
                'duration': duration
            })

            self.test_results.append({
                'test': f'زر النسخ الاحتياطي {format_type.upper()}',
                'status': 'نجح',
                'duration': f"{duration:.2f} ثانية",
                'filename': filename,
                'file_size': f"{file_size:,} بايت",
                'details': f"تم إنشاء النسخة الاحتياطية بنجاح"
            })

            return True

        except Exception as e:
            duration = time.time() - start_time
            error_msg = f"فشل في اختبار زر {format_type.upper()}: {str(e)}"
            print(f"❌ {error_msg}")

            self.test_results.append({
                'test': f'زر النسخ الاحتياطي {format_type.upper()}',
                'status': 'فشل',
                'duration': f"{duration:.2f} ثانية",
                'error': str(e),
                'details': error_msg
            })

            return False

    def wait_for_backup_completion(self, filename, timeout=120):
        """انتظار انتهاء عملية النسخ الاحتياطي عن طريق فحص وجود الملف"""
        print("⏳ انتظار انتهاء عملية النسخ الاحتياطي...")

        backup_dir = BASE_DIR / 'media' / 'backups'
        filepath = backup_dir / filename
        
        start_time = time.time()
        
        while time.time() - start_time < timeout:
            if filepath.exists() and filepath.stat().st_size > 1000:  # ملف موجود وحجمه أكبر من 1KB
                # انتظار قليل إضافي للتأكد من اكتمال الكتابة
                time.sleep(2)
                final_size = filepath.stat().st_size
                time.sleep(1)
                
                # التحقق من أن الملف لم يعد يتغير (اكتمل الكتابة)
                if filepath.stat().st_size == final_size:
                    print("✅ انتهت عملية النسخ الاحتياطي بنجاح")
                    return True
            
            time.sleep(2)

        # حتى لو انتهت المهلة، تحقق من وجود الملف
        if filepath.exists() and filepath.stat().st_size > 1000:
            print("⚠️ انتهت مهلة الانتظار لكن الملف موجود")
            return True
            
        raise Exception(f"انتهت مهلة الانتظار ({timeout} ثانية) ولم يتم إنشاء الملف")

    def validate_backup_file(self, filepath, format_type):
        """التحقق من صحة ملف النسخة الاحتياطية"""
        print(f"🔍 التحقق من صحة ملف {format_type.upper()}: {filepath.name}")

        try:
            if format_type.lower() == 'json':
                # التحقق من ملف JSON
                with open(filepath, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                if 'metadata' not in data:
                    raise Exception("ملف JSON لا يحتوي على metadata")

                if 'data' not in data:
                    raise Exception("ملف JSON لا يحتوي على data")

                metadata = data['metadata']
                total_tables = metadata.get('total_tables', 0)
                total_records = metadata.get('total_records', 0)

                print(f"✅ ملف JSON صالح: {total_tables} جدول، {total_records:,} سجل")

            elif format_type.lower() == 'xlsx':
                # التحقق من ملف Excel
                workbook = openpyxl.load_workbook(filepath, read_only=True)

                if "Backup Info" not in workbook.sheetnames:
                    raise Exception("ملف Excel لا يحتوي على ورقة Backup Info")

                info_sheet = workbook["Backup Info"]

                # قراءة بعض المعلومات الأساسية
                rows = list(info_sheet.iter_rows(values_only=True))
                if len(rows) < 5:
                    raise Exception("ورقة Backup Info لا تحتوي على معلومات كافية")

                total_tables = rows[3][1] if len(rows) > 3 and len(rows[3]) > 1 else 0
                total_records = rows[4][1] if len(rows) > 4 and len(rows[4]) > 1 else 0

                print(f"✅ ملف Excel صالح: {total_tables} جدول، {total_records:,} سجل، {len(workbook.sheetnames)} ورقة")

            else:
                raise Exception(f"تنسيق غير مدعوم: {format_type}")

        except Exception as e:
            raise Exception(f"ملف النسخة الاحتياطية غير صالح: {str(e)}")

    def test_backup_restore(self, backup_file_info):
        """اختبار استعادة ملف النسخة الاحتياطية"""
        print(f"\n🔄 اختبار استعادة الملف: {backup_file_info['filename']}")

        start_time = time.time()

        try:
            filepath = backup_file_info['filepath']
            format_type = backup_file_info['format']

            # تحميل بيانات النسخة الاحتياطية
            if format_type.lower() == 'json':
                with open(filepath, 'r', encoding='utf-8') as f:
                    backup_data = json.load(f)
            elif format_type.lower() == 'xlsx':
                backup_data = load_backup_from_xlsx(open(filepath, 'rb'))
            else:
                raise Exception(f"تنسيق غير مدعوم: {format_type}")

            # محاولة استعادة البيانات (بدون مسح البيانات الموجودة للسلامة)
            perform_backup_restore(backup_data, clear_data=False, user=None)

            duration = time.time() - start_time

            self.test_results.append({
                'test': f'استعادة {format_type.upper()}',
                'status': 'نجح',
                'duration': f"{duration:.2f} ثانية",
                'filename': backup_file_info['filename'],
                'details': f"تمت الاستعادة بنجاح من ملف {format_type.upper()}"
            })

            print(f"✅ تم اختبار الاستعادة بنجاح للملف: {backup_file_info['filename']}")

            return True

        except Exception as e:
            duration = time.time() - start_time
            error_msg = f"فشل في اختبار الاستعادة: {str(e)}"
            print(f"❌ {error_msg}")

            self.test_results.append({
                'test': f'استعادة {format_type.upper()}',
                'status': 'فشل',
                'duration': f"{duration:.2f} ثانية",
                'filename': backup_file_info['filename'],
                'error': str(e),
                'details': error_msg
            })

            return False

    def run_all_tests(self):
        """تشغيل جميع اختبارات الأزرار"""
        print("🚀 بدء اختبار أزرار النسخ الاحتياطي في الواجهة")
        print("=" * 60)

        # تسجيل الدخول
        if not self.login():
            print("❌ فشل في تسجيل الدخول - إيقاف الاختبار")
            return False

        # اختبار زر JSON
        json_success = self.test_backup_button('json')

        # انتظار قليل بين الاختبارات
        time.sleep(5)

        # اختبار زر XLSX
        xlsx_success = self.test_backup_button('xlsx')

        # اختبار الاستعادة للملفات المنشأة
        print("\n🔄 اختبار قابلية الاستعادة...")
        for backup_info in self.backup_files_created:
            self.test_backup_restore(backup_info)

        # ملخص النتائج
        total_tests = len(self.test_results)
        passed_tests = len([r for r in self.test_results if r['status'] == 'نجح'])
        failed_tests = total_tests - passed_tests

        print("\n" + "=" * 60)
        print("📊 ملخص نتائج الاختبار:")
        print(f"إجمالي الاختبارات: {total_tests}")
        print(f"الناجحة: {passed_tests}")
        print(f"الفاشلة: {failed_tests}")

        if failed_tests == 0:
            print("✅ جميع اختبارات الأزرار نجحت!")
        else:
            print("❌ فشل في بعض الاختبارات")

        return failed_tests == 0

    def save_results(self):
        """حفظ نتائج الاختبار في ملف TXT"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"ui_backup_buttons_test_results_{timestamp}.txt"
        filepath = BASE_DIR / 'test_result' / filename

        # التأكد من وجود مجلد test_result
        filepath.parent.mkdir(exist_ok=True)

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write("اختبار أزرار النسخ الاحتياطي في الواجهة\n")
            f.write("=" * 60 + "\n\n")
            f.write(f"تاريخ الاختبار: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"الخادم: {self.base_url}\n\n")

            # إحصائيات عامة
            total_tests = len(self.test_results)
            passed_tests = len([r for r in self.test_results if r['status'] == 'نجح'])
            failed_tests = total_tests - passed_tests

            f.write("إحصائيات الاختبار:\n")
            f.write(f"- إجمالي الاختبارات: {total_tests}\n")
            f.write(f"- الناجحة: {passed_tests}\n")
            f.write(f"- الفاشلة: {failed_tests}\n\n")

            # تفاصيل كل اختبار
            f.write("تفاصيل الاختبارات:\n")
            f.write("-" * 40 + "\n")

            for result in self.test_results:
                f.write(f"\nاختبار: {result['test']}\n")
                f.write(f"الحالة: {result['status']}\n")
                f.write(f"المدة: {result['duration']}\n")

                if 'filename' in result:
                    f.write(f"الملف: {result['filename']}\n")
                if 'file_size' in result:
                    f.write(f"حجم الملف: {result['file_size']}\n")
                if 'error' in result:
                    f.write(f"الخطأ: {result['error']}\n")

                f.write(f"التفاصيل: {result['details']}\n")
                f.write("-" * 40 + "\n")

            # معلومات الملفات المنشأة
            if self.backup_files_created:
                f.write("\nالملفات المنشأة:\n")
                f.write("-" * 40 + "\n")

                for backup in self.backup_files_created:
                    f.write(f"\nالملف: {backup['filename']}\n")
                    f.write(f"التنسيق: {backup['format'].upper()}\n")
                    f.write(f"الحجم: {backup['size']:,} بايت\n")
                    f.write(f"المدة: {backup['duration']:.2f} ثانية\n")
                    f.write(f"المسار: {backup['filepath']}\n")

        print(f"💾 تم حفظ النتائج في: {filepath}")
        return str(filepath)


def main():
    """الدالة الرئيسية"""
    tester = BackupUIButtonTester()

    try:
        success = tester.run_all_tests()
        results_file = tester.save_results()

        print(f"\n📄 تم حفظ تفاصيل النتائج في: {results_file}")

        if success:
            print("🎉 جميع اختبارات الأزرار نجحت!")
            return 0
        else:
            print("⚠️ فشل في بعض الاختبارات")
            return 1

    except Exception as e:
        print(f"❌ خطأ عام في الاختبار: {str(e)}")
        return 1


if __name__ == '__main__':
    sys.exit(main())