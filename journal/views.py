from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import Q, Sum
from django.db import transaction
from django.utils.translation import gettext_lazy as _
from django.utils.safestring import mark_safe
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from decimal import Decimal
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import Account, JournalEntry, JournalLine
from .forms import (AccountForm, JournalEntryForm, JournalLineFormSet, 
                   JournalSearchForm, TrialBalanceForm)
from .services import JournalService
from core.signals import log_view_activity
from core.signals import log_view_activity


@login_required
def journal_dashboard(request):
    """لوحة تحكم القيود المحاسبية"""
    # إحصائيات عامة
    total_entries = JournalEntry.objects.count()
    total_accounts = Account.objects.filter(is_active=True).count()
    recent_entries = JournalEntry.objects.select_related().order_by('-created_at')[:5]
    
    # إحصائيات الحسابات حسب النوع
    account_types_count = {}
    for account_type, type_name in Account.ACCOUNT_TYPES:
        count = Account.objects.filter(account_type=account_type, is_active=True).count()
        account_types_count[type_name] = count
    
    # القيود الحديثة مع التفاصيل
    recent_entries_with_totals = []
    for entry in recent_entries:
        entry_dict = {
            'entry': entry,
            'total_debit': entry.lines.aggregate(total=Sum('debit'))['total'] or 0,
            'total_credit': entry.lines.aggregate(total=Sum('credit'))['total'] or 0,
        }
        recent_entries_with_totals.append(entry_dict)
    
    context = {
        'total_entries': total_entries,
        'total_accounts': total_accounts,
        'recent_entries': recent_entries_with_totals,
        'account_types_count': account_types_count,
    }
    return render(request, 'journal/dashboard.html', context)


@login_required
@permission_required('journal.view_journalaccount', raise_exception=True)
def account_list(request):
    """قائمة الحسابات"""
    accounts = Account.objects.filter(is_active=True)
    
    # البحث
    search = request.GET.get('search')
    if search and search != 'None' and search.strip():
        accounts = accounts.filter(
            Q(name__icontains=search) | 
            Q(code__icontains=search) |
            Q(description__icontains=search)
        )
    
    # فلترة النوع
    account_type = request.GET.get('type')
    if account_type and account_type != 'None' and account_type.strip():
        accounts = accounts.filter(account_type=account_type)
    
    # الترتيب
    order_by = request.GET.get('order_by', 'code')
    direction = request.GET.get('direction', 'asc')
    
    if direction == 'desc':
        order_by = '-' + order_by
    
    # ترتيب حسب الحقل المطلوب مع الترتيب الهرمي
    if order_by in ['code', '-code', 'name', '-name', 'account_type', '-account_type', 'balance', '-balance', 'parent__name', '-parent__name', 'parent__code', '-parent__code']:
        # ترتيب حسب الحقل المطلوب
        accounts = accounts.order_by(order_by)
    else:
        # الترتيب الافتراضي: الحسابات الأب أولاً، ثم الحسابات الفرعية
        parent_accounts = accounts.filter(parent__isnull=True).order_by('code')
        child_accounts = accounts.filter(parent__isnull=False).order_by('code')
        # دمج الاستعلامات
        accounts = list(parent_accounts) + list(child_accounts)
    
    # الترقيم
    total_accounts = accounts.count()  # عدد الحسابات الكلي
    paginator = Paginator(accounts, 30)  # زيادة عدد الحسابات في الصفحة إلى 30
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'account_types': Account.ACCOUNT_TYPES,
        'current_type': account_type,
        'search_query': search,
        'order_by': request.GET.get('order_by', 'code'),
        'direction': direction,
        'total_accounts': total_accounts,
    }
    
    # سجل النشاط: فتح صفحة قائمة الحسابات
    try:
        class Obj:
            id = 0
            pk = 0
            def __str__(self):
                return str(_('Accounts List'))
        log_view_activity(request, 'view', Obj(), str(_('Viewing accounts list')))
    except Exception:
        pass
    
    return render(request, 'journal/account_list.html', context)

@login_required
def export_accounts_excel(request):
    """تصدير قائمة الحسابات إلى Excel"""
    accounts = Account.objects.filter(is_active=True)
    
    # البحث
    search = request.GET.get('search')
    if search and search != 'None' and search.strip():
        accounts = accounts.filter(
            Q(name__icontains=search) | 
            Q(code__icontains=search) |
            Q(description__icontains=search)
        )
    
    # فلترة النوع
    account_type = request.GET.get('type')
    if account_type and account_type != 'None' and account_type.strip():
        accounts = accounts.filter(account_type=account_type)
    
    # الترتيب
    order_by = request.GET.get('order_by', 'code')
    direction = request.GET.get('direction', 'asc')
    
    if direction == 'desc':
        order_by = '-' + order_by
    
    # ترتيب حسب الحقل المطلوب مع الترتيب الهرمي
    if order_by in ['code', '-code', 'name', '-name', 'account_type', '-account_type', 'balance', '-balance', 'parent__name', '-parent__name']:
        accounts = accounts.order_by(order_by)
    else:
        accounts = accounts.order_by('parent__code', 'code')
    
    # إنشاء ملف Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _("قائمة الحسابات")
    
    # تنسيق العناوين
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # إضافة العناوين
    headers = [
        _("كود الحساب"),
        _("اسم الحساب"),
        _("النوع"),
        _("الحساب الرئيسي"),
        _("الرصيد"),
        _("الحالة")
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
    
    # إضافة البيانات
    for row_num, account in enumerate(accounts, 2):
        ws.cell(row=row_num, column=1, value=account.code).border = border
        ws.cell(row=row_num, column=2, value=account.name).border = border
        ws.cell(row=row_num, column=3, value=str(account.get_account_type_display())).border = border
        ws.cell(row=row_num, column=4, value=account.parent.name if account.parent else "").border = border
        ws.cell(row=row_num, column=5, value=float(account.get_balance())).border = border
        ws.cell(row=row_num, column=6, value=_("نشط") if account.is_active else _("غير نشط")).border = border
    
    # إنشاء الاستجابة
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="accounts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # حفظ الملف
    wb.save(response)
    
    # سجل النشاط: تصدير الحسابات
    try:
        class Obj:
            id = 0
            pk = 0
            def __str__(self):
                return str(_('Accounts Export'))
        log_view_activity(request, 'export', Obj(), str(_('Export account list to Excel')))
    except Exception:
        pass
    
    return response

@login_required
@permission_required('journal.add_journalaccount', raise_exception=True)
def account_create(request):
    """إنشاء حساب جديد"""
    is_modal = request.GET.get('modal') == '1'

    # تسجيل فتح نموذج الإنشاء (بما في ذلك منبثق)
    try:
        class Obj:
            id = 0
            pk = 0
            def __str__(self):
                return str(_('Create Account'))
        log_view_activity(request, 'view', Obj(), str(_('Open account creation screen')))
    except Exception:
        pass

    if request.method == 'POST':
        form = AccountForm(request.POST)
        if form.is_valid():
            account = form.save()

            # تحديث رصيد parent إذا كان موجوداً
            if account.parent:
                account.parent.update_account_balance()

            # سجل النشاط عند الإنشاء
            try:
                from core.models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    action_type='create',
                    content_type='Account',
                    object_id=account.pk,
                    description=f"تم إنشاء الحساب: {account.code} - {account.name}"
                )
            except Exception:
                pass

            if is_modal or request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'id': account.id,
                    'code': account.code,
                    'name': account.name,
                    'account_type': account.get_account_type_display(),
                    'display': f"{account.code} - {account.name}"
                })
            else:
                messages.success(request, _('Account created successfully'))
                return redirect('journal:account_list')
        else:
            if is_modal or request.headers.get('x-requested-with') == 'XMLHttpRequest':
                # أعد النموذج مع الأخطاء كـ HTML جزئي
                html = render(request, 'journal/account_form_modal.html', {'form': form, 'title': _('Create new account')}).content.decode('utf-8')
                return JsonResponse({'success': False, 'html': html}, status=400)
    else:
        form = AccountForm()

    # عرض النموذج
    if is_modal:
        return render(request, 'journal/account_form_modal.html', {'form': form, 'title': _('Create new account')})
    return render(request, 'journal/account_form.html', {'form': form, 'title': _('Create new account')})


@login_required
def account_edit(request, pk):
    """تعديل حساب"""
    account = get_object_or_404(Account, pk=pk)
    old_parent = account.parent  # حفظ parent القديم
    
    if request.method == 'POST':
        form = AccountForm(request.POST, instance=account)
        if form.is_valid():
            # حفظ التغييرات
            account = form.save()
            
            # إذا تغير parent، نحتاج لتحديث أرصدة كل من parent القديم والجديد
            new_parent = account.parent
            if old_parent != new_parent:
                from journal.services import JournalService
                # تحديث parent القديم إذا كان موجوداً
                if old_parent:
                    old_parent.update_account_balance()
                # تحديث parent الجديد إذا كان موجوداً
                if new_parent:
                    new_parent.update_account_balance()
                # تحديث الحساب نفسه
                account.update_account_balance()
            
            # سجل النشاط
            try:
                from core.models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    action_type='update',
                    content_type='Account',
                    object_id=account.pk,
                    description=f"تم تعديل الحساب: {account.code} - {account.name}"
                )
            except Exception:
                pass
            
            messages.success(request, _('Account updated successfully'))
            return redirect('journal:account_list')
    else:
        form = AccountForm(instance=account)
    
    return render(request, 'journal/account_form.html', {
        'form': form, 
        'title': _('Edit Account'),
        'account': account
    })


@login_required
def account_delete(request, pk):
    """حذف حساب"""
    account = get_object_or_404(Account, pk=pk)
    
    # التحقق من الصلاحيات
    if not (request.user.is_superuser or request.user.has_perm('users.can_delete_accounts')):
        messages.error(request, _('You do not have permission to delete accounts'))
        return redirect('journal:account_list')
    
    # التحقق من وجود حركات على الحساب
    if account.journal_lines.exists():
        messages.error(request, _('Cannot delete account because it has transactions'))
        return redirect('journal:account_list')
    
    if request.method == 'POST':
        account_name = account.name
        
        # تحديث رصيد parent قبل الحذف إذا كان موجوداً
        if account.parent:
            account.parent.update_account_balance()
        
        account.delete()
        
        # تسجيل النشاط
        try:
            from core.models import AuditLog
            AuditLog.objects.create(
                user=request.user,
                action_type='delete',
                content_type='Account',
                description=f'تم حذف الحساب: {account_name}'
            )
        except Exception as e:
            pass
            
        messages.success(request, _('Account deleted successfully'))
        return redirect('journal:account_list')
    
    return redirect('journal:account_list')


@login_required
def account_detail(request, pk):
    """تفاصيل الحساب"""
    account = get_object_or_404(Account, pk=pk)
    
    # الحصول على حركات الحساب
    lines = JournalLine.objects.filter(account=account).order_by('-created_at')
    
    # إذا كان الحساب رئيسي وليس له حركات مباشرة، أظهر حركات الحسابات الفرعية
    if not lines.exists() and account.children.filter(is_active=True).exists():
        # الحصول على جميع الحسابات الفرعية النشطة
        child_accounts = account.children.filter(is_active=True)
        lines = JournalLine.objects.filter(account__in=child_accounts).select_related('account', 'journal_entry').order_by('-created_at')
    
    # البحث في الحركات
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if date_from:
        lines = lines.filter(journal_entry__entry_date__gte=date_from)
    if date_to:
        lines = lines.filter(journal_entry__entry_date__lte=date_to)
    
    # حساب الإجماليات
    totals = lines.aggregate(
        total_debit=Sum('debit'),
        total_credit=Sum('credit')
    )
    
    # الترقيم
    paginator = Paginator(lines, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    try:
        balance = account.get_balance()
    except Exception as e:
        print(f"ERROR calculating balance: {e}")
        balance = Decimal('0')
    
    balance = balance or Decimal('0')
    
    context = {
        'account': account,
        'page_obj': page_obj,
        'totals': totals,
        'date_from': date_from,
        'date_to': date_to,
        'balance': balance
    }
    
    # تسجيل النشاط في سجل الأنشطة
    log_view_activity(request, 'view', account, str(_('Viewing account details: {}')).format(account.name))
    
    return render(request, 'journal/account_detail.html', context)


@login_required
def journal_entry_list(request):
    """قائمة القيود المحاسبية"""
    entries = JournalEntry.objects.all().order_by('-entry_date', '-created_at')
    
    # البحث والتصفية
    form = JournalSearchForm(request.GET)
    if form.is_valid():
        if form.cleaned_data['date_from']:
            entries = entries.filter(entry_date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data['date_to']:
            entries = entries.filter(entry_date__lte=form.cleaned_data['date_to'])
        if form.cleaned_data['account']:
            entries = entries.filter(lines__account=form.cleaned_data['account']).distinct()
        if form.cleaned_data['entry_number']:
            entries = entries.filter(entry_number__icontains=form.cleaned_data['entry_number'])
    
    # الترتيب
    order_by = request.GET.get('order_by', 'entry_date')
    direction = request.GET.get('direction', 'desc')
    
    if direction == 'asc':
        entries = entries.order_by(order_by)
    else:
        entries = entries.order_by(f'-{order_by}')
    
    # حساب الإحصائيات
    total_amount = entries.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    
    # الترقيم
    paginator = Paginator(entries, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'entries': page_obj,
        'form': form,
        'total_amount': total_amount,
        'order_by': order_by,
        'direction': direction,
    }
    
    # تسجيل النشاط في سجل الأنشطة
    log_view_activity(request, 'view', None, str(_('Viewing journal entries list')))
    
    return render(request, 'journal/entry_list.html', context)


@login_required
def journal_entry_create(request):
    """إنشاء قيد محاسبي جديد"""
    # سجل النشاط: فتح صفحة إنشاء القيد
    try:
        class Obj:
            id = 0
            pk = 0
            def __str__(self):
                return str(_('Create Journal Entry'))
        log_view_activity(request, 'view', Obj(), str(_('Viewing create journal entry page')))
    except Exception:
        pass
    
    if request.method == 'POST':
        # طباعة البيانات المستلمة للتشخيص
        print("\n" + "="*80)
        print("📥 بيانات POST المستلمة:")
        print("="*80)
        for key, value in request.POST.items():
            if 'account' in key or 'debit' in key or 'credit' in key:
                print(f"  {key} = {value}")
        print("="*80 + "\n")
        
        # إنشاء كائن مؤقت بدون حفظه لربط النماذج
        temp_entry = JournalEntry(created_by=request.user)
        form = JournalEntryForm(request.POST, instance=temp_entry, user=request.user)
        formset = JournalLineFormSet(request.POST, instance=temp_entry, prefix='form')

        # التحقق من صحة النماذج
        form_valid = form.is_valid()
        formset_valid = formset.is_valid()
        
        print(f"✓ form_valid: {form_valid}")
        print(f"✓ formset_valid: {formset_valid}")
        if not formset_valid:
            print(f"❌ formset.errors: {formset.errors}")
            print(f"❌ formset.non_form_errors: {formset.non_form_errors()}")
        
        if form_valid and formset_valid:
            try:
                with transaction.atomic():
                    # حفظ القيد أولاً بقيمة إجمالية مبدئية
                    entry = form.save(commit=False)
                    entry.created_by = request.user
                    entry.total_amount = Decimal('0')
                    
                    # التحقق من البنود قبل الحفظ - فحص إذا كان القيد يمس حسابين بنكيين
                    bank_accounts_touched = []
                    cashbox_accounts_touched = []
                    for form_line in formset:
                        if form_line.cleaned_data and not form_line.cleaned_data.get('DELETE', False):
                            account = form_line.cleaned_data.get('account')
                            if account and account.code.startswith('102'):  # حساب بنكي
                                bank_accounts_touched.append(account)
                            elif account and account.code.startswith('101'):  # حساب صندوق
                                cashbox_accounts_touched.append(account)
                    
                    # إذا كان القيد يمس حسابين بنكيين أو أكثر، منع الحفظ تماماً
                    if len(bank_accounts_touched) >= 2:
                        error_msg = mark_safe(
                            _('❌ خطأ: لا يمكن إنشاء قيد محاسبي يمس حسابين بنكيين أو أكثر. '
                              'للقيام بتحويل بين حسابات بنكية، يجب استخدام صفحة التحويلات البنكية '
                              '<a href="/ar/banks/transfers/add/" class="alert-link" target="_blank">من هنا</a>. '
                              '<br><br>عند التحويل من خلال صفحة البنوك، سيتم إنشاء القيد المحاسبي تلقائياً '
                              'وضمان انعكاس التحويل بشكل صحيح على جميع الأنظمة.')
                        )
                        messages.error(request, error_msg)
                        # تسجيل في سجل الأنشطة
                        try:
                            from core.models import AuditLog
                            AuditLog.objects.create(
                                user=request.user,
                                action_type='error',
                                content_type='JournalEntry',
                                description=f'محاولة إنشاء قيد محاسبي يمس {len(bank_accounts_touched)} حسابات بنكية - تم منع الحفظ'
                            )
                        except Exception:
                            pass
                        # إعادة عرض النموذج مع البيانات المدخلة
                        context = {
                            'form': form,
                            'formset': formset,
                            'accounts': Account.objects.filter(is_active=True).order_by('code'),
                            'title': _('Create new journal entry')
                        }
                        return render(request, 'journal/entry_create.html', context)
                    
                    # إذا كان القيد يمس حسابين أو أكثر من حسابات الصناديق، منع الحفظ تماماً
                    if len(cashbox_accounts_touched) >= 2:
                        error_msg = mark_safe(
                            _('❌ خطأ: لا يمكن إنشاء قيد محاسبي يمس حسابين أو أكثر من حسابات الصناديق. '
                              'للقيام بتحويل بين الصناديق، يجب استخدام صفحة التحويلات بين الصناديق '
                              '<a href="/ar/cashboxes/" class="alert-link" target="_blank">من هنا</a>. '
                              '<br><br>عند التحويل من خلال صفحة الصناديق، سيتم إنشاء القيد المحاسبي تلقائياً '
                              'وضمان انعكاس التحويل بشكل صحيح على جميع الأنظمة.')
                        )
                        messages.error(request, error_msg)
                        # تسجيل في سجل الأنشطة
                        try:
                            from core.models import AuditLog
                            AuditLog.objects.create(
                                user=request.user,
                                action_type='error',
                                content_type='JournalEntry',
                                description=f'محاولة إنشاء قيد محاسبي يمس {len(cashbox_accounts_touched)} حسابات صناديق - تم منع الحفظ'
                            )
                        except Exception:
                            pass
                        # إعادة عرض النموذج مع البيانات المدخلة
                        context = {
                            'form': form,
                            'formset': formset,
                            'accounts': Account.objects.filter(is_active=True).order_by('code'),
                            'title': _('Create new journal entry')
                        }
                        return render(request, 'journal/entry_create.html', context)
                    
                    entry.save()

                    # حفظ البنود يدوياً - استبعاد البنود الفارغة
                    formset.instance = entry
                    lines_to_save = []
                    
                    for form in formset:
                        if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                            account = form.cleaned_data.get('account')
                            debit = form.cleaned_data.get('debit') or Decimal('0')
                            credit = form.cleaned_data.get('credit') or Decimal('0')
                            
                            # احفظ فقط البنود التي لها حساب وقيمة
                            if account and (debit > 0 or credit > 0):
                                line = form.save(commit=False)
                                line.journal_entry = entry
                                lines_to_save.append(line)
                    
                    # حفظ جميع البنود الصالحة
                    for line in lines_to_save:
                        line.save()

                    # التحقق من وجود بنود صالحة
                    if not lines_to_save:
                        # لا توجد بنود صالحة - أضف خطأ
                        messages.error(request, _('يجب إدخال بنود القيد المحاسبي'))
                        # حذف القيد الفارغ
                        entry.delete()
                        return redirect('journal:entry_create')

                    # حساب الإجمالي
                    total_debit = entry.lines.aggregate(total=Sum('debit'))['total'] or Decimal('0')
                    entry.total_amount = total_debit
                    entry.save()

                # تسجيل العملية في سجل الأنشطة
                try:
                    from core.models import AuditLog
                    AuditLog.objects.create(
                        user=request.user,
                        action_type='create',
                        content_type='JournalEntry',
                        object_id=entry.pk,
                        description=f'تم إنشاء قيد محاسبي رقم {entry.entry_number} بإجمالي {entry.total_amount}'
                    )
                except Exception:
                    pass

                messages.success(request, _('Entry created successfully'))
                return redirect('journal:entry_detail', pk=entry.pk)
                
            except Exception as e:
                # خطأ غير متوقع أثناء الحفظ
                try:
                    from core.models import AuditLog
                    AuditLog.objects.create(
                        user=request.user,
                        action_type='error',
                        content_type='JournalEntry',
                        description=f'استثناء غير متوقع أثناء إنشاء القيد: {str(e)}'
                    )
                except Exception:
                    pass
                messages.error(request, _('An unexpected error occurred while creating the entry. Please try again.'))
        else:
            # هناك أخطاء في التحقق - إضافة رسائل خطأ واضحة
            if not form_valid:
                messages.error(request, _('Please correct the errors in the basic entry information'))
            
            if not formset_valid:
                # عرض أخطاء الـ formset بشكل واضح
                if formset.non_form_errors():
                    for error in formset.non_form_errors():
                        messages.error(request, error)
                else:
                    messages.error(request, _('Please correct the errors in the entry lines'))
            
            # تسجيل أخطاء التحقق
            try:
                from core.models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    action_type='error',
                    content_type='JournalEntry',
                    description='خطأ تحقق في نموذج القيد أو البنود عند الإنشاء'
                )
            except Exception:
                pass
        
        # في حالة الأخطاء، سيتم عرض النموذج مرة أخرى مع البيانات المدخلة
        # form و formset يحتفظان بالبيانات المدخلة من request.POST
    else:
        # طلب GET - إنشاء نماذج فارغة
        form = JournalEntryForm(user=request.user)
        temp_entry = JournalEntry()
        formset = JournalLineFormSet(instance=temp_entry, prefix='form')
    
    # إعداد السياق وعرض الصفحة (سواء كان GET أو POST مع أخطاء)
    context = {
        'form': form,
        'formset': formset,
        'accounts': Account.objects.filter(
            is_active=True,
            parent__isnull=False  # استثناء الحسابات الرئيسية (parent is null)
        ).order_by('code'),
        'title': _('Create new journal entry')
    }
    return render(request, 'journal/entry_create.html', context)


@login_required
def test_accounts(request):
    """صفحة اختبار البحث عن الحسابات"""
    context = {
        'accounts': Account.objects.filter(is_active=True).order_by('code'),
    }
    return render(request, 'journal/test_accounts.html', context)


@login_required
@permission_required('journal.change_journalentry', raise_exception=True)
def journal_entry_edit(request, pk):
    """تعديل قيد محاسبي"""
    entry = get_object_or_404(JournalEntry, pk=pk)
    if request.method == 'POST':
        form = JournalEntryForm(request.POST, instance=entry, user=request.user)
        formset = JournalLineFormSet(request.POST, instance=entry, prefix='form')
        try:
            form_valid = form.is_valid()
            formset_valid = formset.is_valid()
            
            if form_valid and formset_valid:
                with transaction.atomic():
                    entry = form.save()
                    formset.instance = entry
                    formset.save()
                    entry.clean()
                    total_debit = entry.lines.aggregate(total=Sum('debit'))['total'] or Decimal('0')
                    entry.total_amount = total_debit
                    entry.save()
                # سجل النشاط
                try:
                    from core.models import AuditLog
                    AuditLog.objects.create(
                        user=request.user,
                        action_type='update',
                        content_type='JournalEntry',
                        object_id=entry.pk,
                        description=f'تم تعديل القيد المحاسبي رقم {entry.entry_number} بإجمالي {entry.total_amount}'
                    )
                except Exception:
                    pass
                messages.success(request, _('Entry modified successfully'))
                return redirect('journal:entry_detail', pk=entry.pk)
            else:
                # إضافة تفاصيل الأخطاء للمستخدم
                error_msg = _('Please correct the errors in the form')
                if not form_valid and form.errors:
                    for field, errors in form.errors.items():
                        messages.error(request, f"{field}: {', '.join(errors)}")
                if not formset_valid:
                    if formset.non_form_errors():
                        for error in formset.non_form_errors():
                            messages.error(request, str(error))
                    for i, form_errors in enumerate(formset.errors):
                        if form_errors:
                            messages.error(request, f"بند {i+1}: {form_errors}")
                if not (form.errors or formset.errors or formset.non_form_errors()):
                    messages.error(request, error_msg)
        except Exception as e:
            messages.error(request, _('An error occurred while modifying the entry: ') + str(e))
    else:
        form = JournalEntryForm(instance=entry, user=request.user)
        formset = JournalLineFormSet(instance=entry, prefix='form')
    context = {
        'form': form,
        'formset': formset,
        'accounts': Account.objects.filter(is_active=True).order_by('code'),
        'title': _('Modify journal entry'),
        'entry': entry
    }
    return render(request, 'journal/entry_create.html', context)

@login_required
def journal_entry_detail(request, pk):
    """تفاصيل القيد المحاسبي"""
    entry = get_object_or_404(JournalEntry, pk=pk)
    lines = entry.lines.all().order_by('id')
    
    # حساب الإجماليات
    totals = lines.aggregate(
        total_debit=Sum('debit'),
        total_credit=Sum('credit')
    )
    
    # سجل النشاط: عرض تفاصيل القيد
    try:
        from core.models import AuditLog
        AuditLog.objects.create(
            user=request.user,
            action_type='view',
            content_type='JournalEntry',
            object_id=entry.pk,
            description=f'عرض تفاصيل القيد رقم {entry.entry_number}'
        )
    except Exception:
        pass

    context = {
        'entry': entry,
        'lines': lines,
        'totals': totals
    }
    return render(request, 'journal/entry_detail.html', context)


@login_required
def trial_balance(request):
    """ميزان المراجعة"""
    from datetime import datetime, date, timedelta
    
    # تواريخ افتراضية (الشهر الحالي)
    today = date.today()
    default_date_from = today.replace(day=1)  # أول يوم في الشهر
    default_date_to = today
    
    # إذا كان هناك بيانات في الطلب، استخدمها، وإلا استخدم القيم الافتراضية
    initial_data = {
        'date_from': request.GET.get('date_from', default_date_from.strftime('%Y-%m-%d')),
        'date_to': request.GET.get('date_to', default_date_to.strftime('%Y-%m-%d')),
        'account_type': request.GET.get('account_type', '')
    }
    
    form = TrialBalanceForm(initial=initial_data)
    accounts_data = []
    totals = {'debit': Decimal('0'), 'credit': Decimal('0')}
    
    # إذا كان الطلب GET وله بيانات، أو استخدم البيانات الافتراضية
    if request.GET or True:  # عرض البيانات دائماً
        # استخدام البيانات من الطلب أو الافتراضية
        date_from_str = request.GET.get('date_from', default_date_from.strftime('%Y-%m-%d'))
        date_to_str = request.GET.get('date_to', default_date_to.strftime('%Y-%m-%d'))
        account_type = request.GET.get('account_type', '')
        
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            date_from = default_date_from
            date_to = default_date_to
        
        # الحصول على الحسابات
        accounts = Account.objects.filter(is_active=True)
        if account_type:
            accounts = accounts.filter(account_type=account_type)
        
        for account in accounts:
            # حساب الحركات في الفترة المحددة
            lines = JournalLine.objects.filter(
                account=account,
                journal_entry__entry_date__gte=date_from,
                journal_entry__entry_date__lte=date_to
            )
            
            if lines.exists():
                line_totals = lines.aggregate(
                    debit_total=Sum('debit'),
                    credit_total=Sum('credit')
                )
                
                debit_total = line_totals['debit_total'] or Decimal('0')
                credit_total = line_totals['credit_total'] or Decimal('0')
                balance = debit_total - credit_total
                
                if balance != 0:  # عرض الحسابات التي لها حركة فقط
                    account_data = {
                        'account': account,
                        'debit_total': debit_total,
                        'credit_total': credit_total,
                        'balance': balance,
                        'debit_balance': balance if balance > 0 else Decimal('0'),
                        'credit_balance': abs(balance) if balance < 0 else Decimal('0')
                    }
                    accounts_data.append(account_data)
                    
                    # إضافة للإجماليات
                    if balance > 0:
                        totals['debit'] += balance
                    else:
                        totals['credit'] += abs(balance)
        
        # تحديث form بالبيانات المستخدمة
        form = TrialBalanceForm(initial={
            'date_from': date_from,
            'date_to': date_to,
            'account_type': account_type
        })
    
    context = {
        'form': form,
        'accounts_data': accounts_data,
        'totals': totals,
        'date_from': date_from if 'date_from' in locals() else default_date_from,
        'date_to': date_to if 'date_to' in locals() else default_date_to,
    }
    return render(request, 'journal/trial_balance.html', context)


@login_required
def get_account_balance(request, account_id):
    """API للحصول على رصيد الحساب"""
    try:
        account = Account.objects.get(pk=account_id)
        balance = account.get_balance()
        return JsonResponse({
            'success': True,
            'balance': float(balance),
            'account_name': account.name,
            'account_code': account.code
        })
    except Account.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': _('Account not found')
        })


@login_required
def accounts_api(request):
    """API للحصول على قائمة الحسابات للاستخدام في JavaScript"""
    search = request.GET.get('search', '')
    accounts = Account.objects.filter(is_active=True)
    
    if search:
        accounts = accounts.filter(
            Q(name__icontains=search) | Q(code__icontains=search)
        )
    
    accounts_data = [
        {
            'id': account.id,
            'code': account.code,
            'name': account.name,
            'type': account.get_account_type_display(),
            'balance': float(account.get_balance())
        }
        for account in accounts[:20]  # حد أقصى 20 نتيجة
    ]
    
    return JsonResponse({'accounts': accounts_data})

@login_required
def journal_entries_by_type(request):
    """عرض القيود المحاسبية حسب النوع"""
    entry_type = request.GET.get('type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    entries = JournalEntry.objects.all()
    
    # تصفية حسب النوع
    if entry_type:
        entries = entries.filter(entry_type=entry_type)
    
    # تصفية حسب التاريخ
    if date_from:
        entries = entries.filter(entry_date__gte=date_from)
    if date_to:
        entries = entries.filter(entry_date__lte=date_to)
    
    entries = entries.order_by('-entry_date', '-created_at')
    
    # إحصائيات
    total_entries = entries.count()
    total_amount = entries.aggregate(total=Sum('total_amount'))['total'] or 0
    
    # الترقيم
    paginator = Paginator(entries, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'entries': page_obj,
        'page_obj': page_obj,
        'selected_type': entry_type,
        'date_from': date_from,
        'date_to': date_to,
        'total_entries': total_entries,
        'total_amount': total_amount,
    }
    return render(request, 'journal/entries_by_type.html', context)

@login_required
def journal_entry_detail_with_lines(request, pk):
    """تفاصيل القيد المحاسبي مع البنود"""
    entry = get_object_or_404(JournalEntry, pk=pk)
    lines = entry.lines.all().order_by('id')
    
    # حساب الإجماليات
    totals = lines.aggregate(
        total_debit=Sum('debit'),
        total_credit=Sum('credit')
    )
    
    context = {
        'entry': entry,
        'lines': lines,
        'totals': totals,
    }
    return render(request, 'journal/entry_detail_with_lines.html', context)

@login_required
def delete_journal_entry(request, pk):
    """حذف قيد محاسبي (للمشرفين العليين فقط)"""
    # التحقق من صلاحيات المستخدم
    if not request.user.has_perm('journal.delete_journalentry'):
        messages.error(request, _('You do not have permission to delete journal entries.'))
        return redirect('journal:entry_list')
    
    entry = get_object_or_404(JournalEntry, pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                entry_number = entry.entry_number
                entry_description = entry.description
                entry_amount = entry.total_amount
                # تسجيل النشاط قبل الحذف
                try:
                    from core.models import AuditLog
                    AuditLog.objects.create(
                        user=request.user,
                        action_type='delete',
                        content_type='JournalEntry',
                        object_id=entry.pk,
                        description=f"تم حذف القيد المحاسبي رقم {entry_number} - المبلغ: {entry_amount} - الوصف: {entry_description[:50]}"
                    )
                except Exception:
                    pass
                # حذف القيد
                entry.delete()
                messages.success(request, _('The Accounting Entry was Deleted Successfully'))
        except Exception as e:
            messages.error(request, _('Error deleting journal entry: ') + str(e))
            # تسجيل خطأ الحذف في سجل الأنشطة
            try:
                from core.models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    action_type='error',
                    content_type='JournalEntry',
                    object_id=entry.pk,
                    description=f'فشل في حذف القيد المحاسبي رقم {entry.entry_number}: {str(e)}'
                )
            except Exception:
                pass
        return redirect('journal:entry_list')
    context = {
        'entry': entry,
        'lines': entry.lines.all(),
    }
    return render(request, 'journal/entry_confirm_delete.html', context)

@login_required
def account_ledger(request, pk):
    """دفتر الحساب (حركات مفصلة)"""
    account = get_object_or_404(Account, pk=pk)
    
    # التصفية بالتاريخ
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    lines = JournalLine.objects.filter(account=account).select_related('journal_entry')
    
    if date_from:
        lines = lines.filter(journal_entry__entry_date__gte=date_from)
    if date_to:
        lines = lines.filter(journal_entry__entry_date__lte=date_to)
    
    lines = lines.order_by('journal_entry__entry_date', 'journal_entry__created_at')
    
    # حساب الرصيد التراكمي
    running_balance = Decimal('0')
    lines_with_balance = []
    
    for line in lines:
        if account.account_type in ['asset', 'expense', 'purchases']:
            # الأصول والمصاريف ترتفع بالمدين
            running_balance += line.debit - line.credit
        else:
            # المطلوبات والإيرادات وحقوق الملكية ترتفع بالدائن
            running_balance += line.credit - line.debit
        
        lines_with_balance.append({
            'line': line,
            'balance': running_balance
        })
    
    # الترقيم
    paginator = Paginator(lines_with_balance, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # إجماليات
    totals = lines.aggregate(
        total_debit=Sum('debit'),
        total_credit=Sum('credit')
    )
    
    context = {
        'account': account,
        'page_obj': page_obj,
        'totals': totals,
        'date_from': date_from,
        'date_to': date_to,
        'final_balance': running_balance,
    }
    return render(request, 'journal/account_ledger.html', context)


@login_required
def year_end_closing(request):
    """إقفال السنة المالية"""
    user = request.user
    has_perm = (
        getattr(user, 'is_superuser', False) or
        getattr(user, 'user_type', None) in ['superadmin', 'admin'] or
        user.has_perm('journal.can_perform_year_end_closing')
    )
    if not has_perm:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    from .models import YearEndClosing
    from django import forms

    class YearEndClosingForm(forms.Form):
        year = forms.IntegerField(
            label=_('Fiscal Year'),
            initial=datetime.now().year,
            min_value=2000,
            max_value=datetime.now().year
        )
        closing_date = forms.DateField(
            label=_('Closing Date'),
            initial=date.today(),
            widget=forms.DateInput(attrs={'type': 'date'})
        )
        confirm = forms.BooleanField(
            label=_('I confirm performing the year-end closing'),
            required=True
        )

    if request.method == 'POST':
        form = YearEndClosingForm(request.POST)
        if form.is_valid():
            year = form.cleaned_data['year']
            closing_date = form.cleaned_data['closing_date']
            
            # التحقق من عدم وجود إقفال سابق لنفس السنة
            if YearEndClosing.objects.filter(year=year).exists():
                messages.error(request, _('Closing has already been performed for this year'))
                return redirect('journal:year_end_closing')
            
            # إنشاء سجل الإقفال
            closing = YearEndClosing.objects.create(
                year=year,
                closing_date=closing_date,
                created_by=request.user
            )
            
            # إجراء الإقفال
            success, message = closing.perform_closing()
            
            if success:
                messages.success(request, message)
                
                # تسجيل النشاط
                try:
                    from core.signals import log_view_activity
                    class ClosingObj:
                        id = closing.id
                        pk = closing.pk
                        def __str__(self):
                            return str(_('Year-End Closing'))
                    log_view_activity(request, 'add', ClosingObj(), 
                                    f"إقفال السنة المالية {year} - صافي الربح: {closing.net_profit}")
                except Exception:
                    pass
                
                return redirect('journal:year_end_closing')
            else:
                messages.error(request, message)
                closing.delete()  # حذف السجل إذا فشل
    else:
        form = YearEndClosingForm()

    # عرض الإقفالات السابقة
    closings = YearEndClosing.objects.all().order_by('-year')[:10]
    
    # حساب صافي الربح المقدر للسنة الحالية
    current_year = datetime.now().year
    estimated_closing = YearEndClosing(year=current_year, closing_date=date.today())
    estimated_profit = estimated_closing.calculate_net_profit()
    
    # الحصول على العملة الأساسية
    from settings.models import Currency
    base_currency = Currency.get_base_currency()
    
    context = {
        'form': form,
        'closings': closings,
        'estimated_profit': estimated_profit,
        'current_year': current_year,
        'base_currency': base_currency,
    }
    
    # تسجيل عرض الصفحة
    try:
        from core.signals import log_view_activity
        class PageObj:
            id = 0
            pk = 0
            def __str__(self):
                return str(_('Year-End Closing Page'))
        log_view_activity(request, 'view', PageObj(), str(_('Viewing Year-End Closing Page')))
    except Exception:
        pass
    
    return render(request, 'journal/year_end_closing.html', context)


@login_required
@permission_required('journal.change_account', raise_exception=True)
def fix_account_hierarchy(request):
    """إصلاح الهرمية المكسورة للحسابات المحاسبية"""
    if request.method == 'POST':
        try:
            fixed_count = Account.fix_broken_hierarchy()
            
            if fixed_count > 0:
                messages.success(
                    request, 
                    _('تم إصلاح {} حساب بنجاح. تم ربط الحسابات الفرعية بحساباتها الأب المناسبة.').format(fixed_count)
                )
            else:
                messages.info(
                    request, 
                    _('لم يتم العثور على حسابات تحتاج إصلاح. جميع الحسابات مرتبطة بشكل صحيح.')
                )
            
            # تسجيل النشاط
            try:
                from core.signals import log_view_activity
                class FixObj:
                    id = 0
                    pk = 0
                    def __str__(self):
                        return str(_('Account Hierarchy Fix'))
                log_view_activity(request, 'change', FixObj(), 
                                f'تم إصلاح {fixed_count} حساب في الهرمية المحاسبية')
            except Exception:
                pass
                
        except Exception as e:
            messages.error(
                request, 
                _('حدث خطأ أثناء إصلاح الهرمية: {}').format(str(e))
            )
    
    return redirect('journal:account_list')
