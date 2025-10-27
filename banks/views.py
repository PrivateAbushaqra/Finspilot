from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.utils import timezone
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, TemplateView, View, DetailView
from django.contrib import messages
from django.urls import reverse_lazy
from django.utils.translation import gettext_lazy as _
from decimal import Decimal, InvalidOperation
from .models import BankAccount, BankTransfer, BankTransaction, BankReconciliation, BankStatement
from settings.models import Currency, CompanySettings
from core.signals import log_user_activity
from journal.services import JournalService

def clean_decimal_input(value):
    """
    Clean decimal input from commas and spaces
    """
    if value is None:
        return '0'
    return str(value).replace(',', '').replace(' ', '').strip()

class BanksViewPermissionMixin:
    """Bank pages view permissions guard (view only)."""
    def user_can_view_banks(self):
        u = getattr(self, 'request', None).user
        # Allow superusers/managers, or those with custom view permissions
        return (
            u.is_authenticated and (
                u.is_superuser or u.has_perm('banks.can_view_banks_account') or u.has_perm('banks.can_edit_banks_account') or u.has_perm('banks.can_delete_banks_account')
            )
        )

    def dispatch(self, request, *args, **kwargs):
        if not self.user_can_view_banks():
            messages.error(request, _('You do not have permission to view bank pages.'))
            return redirect('core:dashboard')
        return super().dispatch(request, *args, **kwargs)

# Temporary placeholder views
class BankAccountListView(LoginRequiredMixin, BanksViewPermissionMixin, TemplateView):
    template_name = 'banks/account_list.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Log activity in activity log
        log_user_activity(
            self.request,
            'ACCESS',
            None,
            _('Bank Account List Accessed')
        )
        
        accounts = BankAccount.objects.all()
        
        # إضافة الرصيد الفعلي لكل حساب
        accounts_with_balance = []
        for account in accounts:
            account.actual_balance = account.calculate_actual_balance()
            accounts_with_balance.append(account)
        
        context['accounts'] = accounts_with_balance
        context['total_accounts'] = accounts.count()
        context['active_accounts'] = accounts.filter(is_active=True).count()
        context['inactive_accounts'] = accounts.filter(is_active=False).count()
        
        # Add active currencies
        context['currencies'] = Currency.get_active_currencies()
        context['base_currency'] = Currency.get_base_currency()
        
        # Calculate balances separately for each currency
        balances_by_currency = {}
        for account in accounts:
            currency_code = account.currency  # CharField
            if currency_code not in balances_by_currency:
                balances_by_currency[currency_code] = 0
            # استخدام الرصيد المحسوب من الحركات بدلاً من الحقل المحفوظ
            actual_balance = account.calculate_actual_balance()
            balances_by_currency[currency_code] += actual_balance
        
        context['balances_by_currency'] = balances_by_currency
        
        # Calculate number of different currencies used in accounts
        context['currencies_count'] = len(balances_by_currency)
        
        # Use base currency from company settings
        from settings.models import CompanySettings
        company_settings = CompanySettings.objects.first()
        if company_settings and company_settings.base_currency:
            base_currency_code = company_settings.base_currency.code
        elif context['base_currency']:
            base_currency_code = context['base_currency'].code
        else:
            base_currency_code = None
        
        context['total_balance'] = balances_by_currency.get(base_currency_code, 0) if base_currency_code else 0
        
        return context

class BankAccountCreateView(LoginRequiredMixin, View):
    template_name = 'banks/account_add.html'
    
    def get(self, request, *args, **kwargs):
        # Prevent view-only users from accessing the add page
        if not (
            request.user.is_superuser
            or request.user.has_perm('banks.add_bankaccount')
            or request.user.has_perm('banks.can_add_banks_account')
        ):
            messages.error(request, _('You do not have permission to add bank accounts.'))
            return redirect('banks:account_list')
        
        # Log activity in activity log
        log_user_activity(
            request,
            'ACCESS',
            None,
            _('Bank Account Add Page Accessed')
        )
        
        context = {
            'currencies': Currency.get_active_currencies(),
            'base_currency': Currency.get_base_currency()
        }
        return render(request, self.template_name, context)
    
    def post(self, request, *args, **kwargs):
        if not (
            request.user.is_superuser
            or request.user.has_perm('banks.add_bankaccount')
            or request.user.has_perm('banks.can_add_banks_account')
        ):
            messages.error(request, _('You do not have permission to add bank accounts.'))
            return redirect('banks:account_list')
        try:
            # Receive data from the form
            name = request.POST.get('name', '').strip()
            bank_name = request.POST.get('bank_name', '').strip()
            account_number = request.POST.get('account_number', '').strip()
            iban = request.POST.get('iban', '').strip()
            swift_code = request.POST.get('swift_code', '').strip()
            balance = request.POST.get('balance', '0')
            currency_code = request.POST.get('currency', '')
            is_active = request.POST.get('is_active') == 'on'
            notes = request.POST.get('notes', '').strip()

            # Validate data
            if not name:
                messages.error(request, 'Account name is required!')
                return render(request, self.template_name)

            if not bank_name:
                messages.error(request, 'Bank name is required!')
                return render(request, self.template_name)

            if not account_number:
                messages.error(request, 'Account number is required!')
                return render(request, self.template_name)

            # Check for duplicate account number
            if BankAccount.objects.filter(account_number=account_number).exists():
                messages.error(request, 'Account number already exists!')
                return render(request, self.template_name)

            # Convert balance to number
            try:
                balance = Decimal(str(balance))
            except (ValueError, InvalidOperation):
                balance = Decimal('0.0')

            # Get currency from database
            currency_obj = Currency.objects.filter(code=currency_code).first()
            if not currency_obj:
                # If currency not found, use the base currency from company settings
                from settings.models import CompanySettings
                company_settings = CompanySettings.objects.first()
                if company_settings and company_settings.base_currency:
                    currency_obj = company_settings.base_currency
                    currency_code = currency_obj.code
                else:
                    currency_obj = Currency.get_base_currency()
                    currency_code = currency_obj.code if currency_obj else ''

            # Create bank account with initial balance journal entry
            from django.db import transaction as db_transaction
            
            with db_transaction.atomic():
                # تعيين علم لمنع Signal من إنشاء قيد مكرر
                # سننشئ القيد هنا في الـ View لضمان التحكم الكامل
                will_create_opening_entry = balance > 0
                
                account = BankAccount(
                    name=name,
                    bank_name=bank_name,
                    account_number=account_number,
                    iban=iban,
                    swift_code=swift_code,
                    balance=balance,
                    initial_balance=balance,  # تعيين الرصيد الافتتاحي
                    currency=currency_code,
                    is_active=is_active,
                    notes=notes,
                    created_by=request.user if request.user.is_authenticated else None
                )
                
                # تعيين علم لمنع Signal من إنشاء قيد مكرر
                if will_create_opening_entry:
                    account._skip_opening_balance_signal = True
                
                account.save()
                
                # إنشاء الحساب الفرعي في journal.Account دائماً
                bank_account_obj = JournalService.get_or_create_bank_account(account)
                
                # إنشاء قيد محاسبي للرصيد الافتتاحي إذا كان أكبر من صفر
                # متوافق مع IFRS - يتم استخدام حساب رأس المال للرصيد الافتتاحي (IAS 1)
                if balance > 0:
                    # التحقق من عدم وجود قيد افتتاحي بالفعل لهذا الحساب
                    from journal.models import JournalEntry
                    existing_entry = JournalEntry.objects.filter(
                        reference_type='bank_initial',
                        reference_id=account.id
                    ).exists()
                    
                    if existing_entry:
                        print(f"⚠ قيد الرصيد الافتتاحي موجود بالفعل للحساب {account.name}، تم تخطي الإنشاء")
                    else:
                        try:
                            from journal.models import Account
                            
                            # استخدام حساب رأس المال للرصيد الافتتاحي (IFRS IAS 1)
                            # الرصيد الافتتاحي يمثل مساهمة في رأس المال
                            capital_account = Account.objects.filter(code='301').first()
                            
                            if bank_account_obj and capital_account:
                                lines_data = [
                                    {
                                        'account_id': bank_account_obj.id,
                                        'debit': balance,
                                        'credit': Decimal('0'),
                                        'description': f'{_("Opening Balance")}: {account.name}'
                                    },
                                    {
                                        'account_id': capital_account.id,
                                        'debit': Decimal('0'),
                                        'credit': balance,
                                        'description': f'{_("Opening Balance - Capital Contribution")} - {account.name}'
                                    }
                                ]
                                
                                journal_entry = JournalService.create_journal_entry(
                                    entry_date=timezone.now().date(),
                                    description=f'{_("Opening Balance - Bank Account")}: {account.name}',
                                    reference_type='bank_initial',
                                    reference_id=account.id,
                                    lines_data=lines_data,
                                    user=request.user
                                )
                                print(f"✓ تم إنشاء قيد الرصيد الافتتاحي {journal_entry.entry_number} للحساب {account.name}")
                        except Exception as e:
                            # لا نريد إيقاف العملية إذا فشل إنشاء القيد
                            print(f"خطأ في إنشاء القيد المحاسبي للرصيد الافتتاحي: {e}")
                            import traceback
                            traceback.print_exc()

            # Log activity in activity log
            log_user_activity(
                request.user,
                'CREATE',
                account,
                f'تم إنشاء حساب بنكي جديد: {account.name} - الرصيد الأولي: {balance} {currency_code}'
            )

            messages.success(request, f'Bank account "{account.name}" created successfully!')
            return redirect('banks:account_list')

        except Exception as e:
            messages.error(request, f'Error occurred while creating account: {str(e)}')
            return render(request, self.template_name)

class BankAccountUpdateView(LoginRequiredMixin, View):
    template_name = 'banks/account_edit.html'
    
    def get(self, request, pk, *args, **kwargs):
        if not (request.user.is_superuser or request.user.has_perm('banks.change_bankaccount') or request.user.has_perm('banks.can_edit_banks_account')):
            messages.error(request, _('You do not have permission to edit bank accounts.'))
            return redirect('banks:account_list')
        account = get_object_or_404(BankAccount, pk=pk)
        # If he has only the custom permission (without change_bankaccount), allow him to edit any account
        # No restriction by creator as per client request
        # Handle balance if it is None
        if account.balance is None:
            account.balance = 0.000
        
        # حساب الرصيد الفعلي من المعاملات
        actual_balance = account.calculate_actual_balance()
        
        # If input data is passed (from post when error appears), use it
        initial = kwargs.get('initial', None)
        context = {
            'account': account,
            'actual_balance': actual_balance,
            'currencies': Currency.get_active_currencies(),
            'base_currency': Currency.get_base_currency(),
        }
        if initial:
            context['initial'] = initial
        return render(request, self.template_name, context)
    
    def post(self, request, pk, *args, **kwargs):
        if not (request.user.is_superuser or request.user.has_perm('banks.change_bankaccount') or request.user.has_perm('banks.can_edit_banks_account')):
            messages.error(request, _('You do not have permission to edit bank accounts.'))
            return redirect('banks:account_list')
        try:
            account = get_object_or_404(BankAccount, pk=pk)
            # Holder of can_edit_banks_account can edit any account even without change permission
            # Receive data from the form
            name = request.POST.get('name', '').strip()
            bank_name = request.POST.get('bank_name', '').strip()
            account_number = request.POST.get('account_number', '').strip()
            iban = request.POST.get('iban', '').strip()
            swift_code = request.POST.get('swift_code', '').strip()
            balance = request.POST.get('balance', '0')
            currency_code = request.POST.get('currency', '')
            is_active = request.POST.get('is_active') == 'on'
            notes = request.POST.get('notes', '').strip()
            # Validate data
            if not name or not bank_name or not account_number:
                if not name:
                    messages.error(request, 'Account name is required!')
                if not bank_name:
                    messages.error(request, 'Bank name is required!')
                if not account_number:
                    messages.error(request, 'Account number is required!')
                initial = {
                    'name': name,
                    'bank_name': bank_name,
                    'account_number': account_number,
                    'iban': iban,
                    'swift_code': swift_code,
                    'balance': balance,
                    'currency': currency_code,
                    'is_active': is_active,
                    'notes': notes,
                }
                return self.get(request, pk, initial=initial)
            # Check for duplicate account number (excluding current account)
            existing_account = BankAccount.objects.filter(
                account_number=account_number
            ).exclude(pk=pk).first()
            if existing_account:
                messages.error(request, 'Account number already exists!')
                initial = {
                    'name': name,
                    'bank_name': bank_name,
                    'account_number': account_number,
                    'iban': iban,
                    'swift_code': swift_code,
                    'balance': balance,
                    'currency': currency_code,
                    'is_active': is_active,
                    'notes': notes,
                }
                return self.get(request, pk, initial=initial)
            # Convert balance to number
            try:
                new_balance = Decimal(str(balance))
            except (ValueError, InvalidOperation):
                new_balance = account.balance  # Keep the current balance in case of error
            
            # حساب الرصيد القديم
            old_balance = account.calculate_actual_balance()
            balance_difference = new_balance - old_balance
            
            # Update account
            account.name = name
            account.bank_name = bank_name
            account.account_number = account_number
            account.iban = iban
            account.swift_code = swift_code
            account.currency = currency_code
            account.is_active = is_active
            account.notes = notes
            
            # معالجة تغيير الرصيد
            if balance_difference != 0:
                # إنشاء حركة بنكية لتعديل الرصيد
                from .models import BankTransaction
                from django.utils import timezone
                from core.utils import get_adjustment_account_code
                
                # الحصول على نوع التعديل من الطلب (إجباري عند تغيير الرصيد - IFRS)
                adjustment_type = request.POST.get('adjustment_type', '').strip()
                
                # التحقق من أن نوع التعديل تم اختياره (IFRS Compliance)
                if not adjustment_type:
                    messages.error(request, _('يجب اختيار نوع التعديل عند تغيير الرصيد الحالي لضمان التوافق مع معايير IFRS'))
                    initial = {
                        'name': name,
                        'bank_name': bank_name,
                        'account_number': account_number,
                        'iban': iban,
                        'swift_code': swift_code,
                        'balance': balance,
                        'currency': currency_code,
                        'is_active': is_active,
                        'notes': notes,
                    }
                    return self.get(request, pk, initial=initial)
                
                # تحديد نوع الحركة
                if balance_difference > 0:
                    transaction_type = 'deposit'
                    description = _('تعديل يدوي للرصيد - زيادة: %(amount)s - نوع: %(type)s') % {
                        'amount': abs(balance_difference),
                        'type': dict(BankTransaction.ADJUSTMENT_TYPES).get(adjustment_type, 'غير محدد')
                    }
                else:
                    transaction_type = 'withdrawal'
                    description = _('تعديل يدوي للرصيد - نقص: %(amount)s - نوع: %(type)s') % {
                        'amount': abs(balance_difference),
                        'type': dict(BankTransaction.ADJUSTMENT_TYPES).get(adjustment_type, 'غير محدد')
                    }
                
                # إنشاء الحركة مع نوع التعديل
                BankTransaction.objects.create(
                    bank=account,
                    date=timezone.now().date(),
                    transaction_type=transaction_type,
                    amount=abs(balance_difference),
                    description=description,
                    reference_number=f'ADJ-{account.id}-{timezone.now().strftime("%Y%m%d%H%M%S")}',
                    adjustment_type=adjustment_type,
                    is_manual_adjustment=True,
                    created_by=request.user
                )
                
                # إنشاء قيد محاسبي للتعديل باستخدام الحساب الصحيح
                try:
                    from accounts.models import Account
                    bank_account_obj = JournalService.get_or_create_bank_account(account)
                    
                    # تحديد الحساب المقابل حسب نوع التعديل (IFRS compliant)
                    adjustment_account_code = get_adjustment_account_code(adjustment_type, is_bank=True)
                    adjustment_account = Account.objects.filter(code=adjustment_account_code).first()
                    
                    if bank_account_obj and adjustment_account:
                        lines_data = []
                        
                        if balance_difference > 0:
                            # زيادة في الرصيد: مدين البنك، دائن الحساب المقابل
                            lines_data = [
                                {
                                    'account_id': bank_account_obj.id,
                                    'debit': abs(balance_difference),
                                    'credit': Decimal('0'),
                                    'description': f'{_("Increase in balance")}: {account.name} ({dict(BankTransaction.ADJUSTMENT_TYPES).get(adjustment_type, "تعديل")})'
                                },
                                {
                                    'account_id': adjustment_account.id,
                                    'debit': Decimal('0'),
                                    'credit': abs(balance_difference),
                                    'description': f'{adjustment_account.name} - {dict(BankTransaction.ADJUSTMENT_TYPES).get(adjustment_type, "تعديل")}'
                                }
                            ]
                        else:
                            # نقصان في الرصيد: دائن البنك، مدين الحساب المقابل
                            lines_data = [
                                {
                                    'account_id': adjustment_account.id,
                                    'debit': abs(balance_difference),
                                    'credit': Decimal('0'),
                                    'description': f'{adjustment_account.name} - {dict(BankTransaction.ADJUSTMENT_TYPES).get(adjustment_type, "تعديل")}'
                                },
                                {
                                    'account_id': bank_account_obj.id,
                                    'debit': Decimal('0'),
                                    'credit': abs(balance_difference),
                                    'description': f'{_("Decrease in balance")}: {account.name} ({dict(BankTransaction.ADJUSTMENT_TYPES).get(adjustment_type, "تعديل")})'
                                }
                            ]
                        
                        journal_entry = JournalService.create_journal_entry(
                            entry_date=timezone.now().date(),
                            description=f'{_("Adjustment of Bank Balance")}: {account.name} - {dict(BankTransaction.ADJUSTMENT_TYPES).get(adjustment_type, "تعديل")}',
                            reference_type='bank_adjustment',
                            reference_id=account.id,
                            lines_data=lines_data,
                            user=request.user
                        )
                except Exception as e:
                    # لا نريد إيقاف العملية إذا فشل إنشاء القيد
                    print(f"خطأ في إنشاء القيد المحاسبي للتعديل: {e}")
                
                # تسجيل في سجل الأنشطة
                from core.models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    action_type='update',
                    content_type='BankAccount',
                    object_id=account.id,
                    description=_('تعديل رصيد الحساب البنكي: %(name)s من %(old)s إلى %(new)s (فرق: %(diff)s)') % {
                        'name': name,
                        'old': old_balance,
                        'new': new_balance,
                        'diff': balance_difference
                    },
                    ip_address=request.META.get('REMOTE_ADDR')
                )
                
                # إعادة حساب الرصيد من المعاملات
                account.sync_balance()
            else:
                # حفظ بدون تغيير الرصيد
                account.save()
            
            messages.success(request, _('تم تحديث الحساب البنكي "%(name)s" بنجاح!') % {'name': account.name})
            return redirect('banks:account_list')
        except Exception as e:
            messages.error(request, f'Error occurred while updating account: {str(e)}')
            initial = {
                'name': request.POST.get('name', ''),
                'bank_name': request.POST.get('bank_name', ''),
                'account_number': request.POST.get('account_number', ''),
                'iban': request.POST.get('iban', ''),
                'swift_code': request.POST.get('swift_code', ''),
                'balance': request.POST.get('balance', '0'),
                'currency': request.POST.get('currency', ''),
                'is_active': request.POST.get('is_active') == 'on',
                'notes': request.POST.get('notes', ''),
            }
            return self.get(request, pk, initial=initial)

class BankAccountDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        from django.db import transaction
        if not (
            request.user.is_superuser
            or request.user.has_perm('banks.delete_bankaccount')
            or request.user.has_perm('banks.can_delete_banks_account')
            or request.user.has_perm('users.can_delete_accounts')
        ):
            messages.error(request, _('You do not have permission to delete bank accounts.'))
            return redirect('banks:account_list')
        try:
            account = BankAccount.objects.get(pk=pk)
            account_name = account.name
            # If no transactions exist on the account, allow direct deletion regardless of balance
            from .models import BankTransaction
            no_transactions = not BankTransaction.objects.filter(bank=account).exists()
            if no_transactions:
                with transaction.atomic():
                    # Delete any related transfers as a precautionary measure
                    BankTransfer.objects.filter(from_account=account).delete()
                    BankTransfer.objects.filter(to_account=account).delete()
                    from cashboxes.models import CashboxTransfer
                    CashboxTransfer.objects.filter(from_bank=account).delete()
                    CashboxTransfer.objects.filter(to_bank=account).delete()
                    # Delete account
                    
                # Log activity in activity log
                from core.signals import log_activity
                log_activity(
                    action_type='delete',
                    obj=None,  # Object deleted
                    description=f'Deleted bank account "{account_name}" (no transactions)',
                    user=request.user
                )
                
                messages.success(request, f'Bank account "{account_name}" deleted because it has no transactions.')
                return redirect('banks:account_list')
            
            # Check that there is no balance in the account
            if account.balance != 0:
                if request.user.is_superuser:
                    messages.error(request, f'Cannot delete account "{account_name}" because balance is not zero ({account.balance:.3f}). You can delete all account transactions first using the eraser button, or transfer the balance to another account.')
                else:
                    messages.error(request, f'Cannot delete account "{account_name}" because balance is not zero ({account.balance:.3f}). The balance must be transferred to another account first.')
                return redirect('banks:account_list')
            
            # Check for transactions related to the account (only if balance is not zero)
            # If balance is zero, account can be deleted even if it has transactions (transfers for example)
            from .models import BankTransaction
            bank_transactions = BankTransaction.objects.filter(bank=account)
            
            # We allow deletion if balance is zero, even if there are transactions
            # Because this means the user has successfully transferred all funds
            if bank_transactions.exists() and account.balance != 0:
                # Save account ID in session to display transactions page
                request.session['delete_account_id'] = account.pk
                request.session['delete_account_name'] = account_name
                if request.user.is_superuser:
                    messages.warning(request, _('Cannot delete account "{}" because it has {} financial transactions and non-zero balance ({}). You can use the eraser button to delete all transactions first.').format(account_name, bank_transactions.count(), account.balance))
                else:
                    messages.warning(request, _('Cannot delete account "{}" because it has {} financial transactions and non-zero balance ({}). You must transfer the balance or delete all transactions first.').format(account_name, bank_transactions.count(), account.balance))
                return redirect('banks:account_transactions', pk=account.pk)
            
            # Check for transfers related to the account (only if balance is not zero)
            # If balance is zero, account can be deleted even if it has transfers
            bank_transfers_from = BankTransfer.objects.filter(from_account=account)
            bank_transfers_to = BankTransfer.objects.filter(to_account=account)
            
            if (bank_transfers_from.exists() or bank_transfers_to.exists()) and account.balance != 0:
                transfers_count = bank_transfers_from.count() + bank_transfers_to.count()
                messages.error(request, f'Cannot delete account "{account_name}" because there are {transfers_count} bank transfers and non-zero balance ({account.balance:.3f}). You must transfer the balance or delete all transfers first.')
                return redirect('banks:account_list')
            
            # Check for transfers between banks and cashboxes (only if balance is not zero)
            from cashboxes.models import CashboxTransfer
            cashbox_transfers_from = CashboxTransfer.objects.filter(from_bank=account)
            cashbox_transfers_to = CashboxTransfer.objects.filter(to_bank=account)
            
            if (cashbox_transfers_from.exists() or cashbox_transfers_to.exists()) and account.balance != 0:
                transfers_count = cashbox_transfers_from.count() + cashbox_transfers_to.count()
                messages.error(request, f'Cannot delete account "{account_name}" because there are {transfers_count} transfers between banks and cashboxes and non-zero balance ({account.balance:.3f}). You must transfer the balance or delete all transfers first.')
                return redirect('banks:account_list')
            
            # If all checks pass, account can be deleted
            with transaction.atomic():
                # If balance is zero, delete transactions and transfers first
                if account.balance == 0:
                    # Delete related transactions
                    from .models import BankTransaction
                    BankTransaction.objects.filter(bank=account).delete()
                    
                    # Delete related transfers
                    BankTransfer.objects.filter(from_account=account).delete()
                    BankTransfer.objects.filter(to_account=account).delete()
                    
                    # Delete related cashbox transfers
                    from cashboxes.models import CashboxTransfer
                    CashboxTransfer.objects.filter(from_bank=account).delete()
                    CashboxTransfer.objects.filter(to_bank=account).delete()
                
                # Delete account
                account.delete()
                
                # Log activity in activity log
                from core.signals import log_activity
                log_activity(
                    action_type='delete',
                    obj=None,  # Object deleted
                    description=f'Deleted bank account "{account_name}"',
                    user=request.user
                )
                
                messages.success(request, f'Bank account "{account_name}" deleted successfully!')
            
        except BankAccount.DoesNotExist:
            messages.error(request, 'Bank account not found!')
        except Exception as e:
            # التعامل مع أخطاء المفاتيح الخارجية المحمية
            error_message = str(e)
            if "Cannot delete some instances" in error_message and "protected foreign keys" in error_message:
                messages.error(request, f'لا يمكن حذف الحساب "{account_name}" لأن هناك بيانات مرتبطة به في النظام. يُنصح بإلغاء تفعيل الحساب بدلاً من حذفه.')
            else:
                messages.error(request, f'Error occurred while deleting account: {error_message}')
        
        return redirect('banks:account_list')

class BankTransferListView(LoginRequiredMixin, BanksViewPermissionMixin, TemplateView):
    template_name = 'banks/transfer_list.html'
    
    def get_context_data(self, **kwargs):
        from cashboxes.models import CashboxTransfer
        from django.db.models import Sum
        from datetime import datetime
        from journal.models import JournalEntry
        
        context = super().get_context_data(**kwargs)
        
        # الحصول على التحويلات البنكية مع القيود المحاسبية
        bank_transfers = BankTransfer.objects.select_related('from_account', 'to_account', 'created_by').all()
        
        # إضافة معلومات القيد المحاسبي لكل تحويل
        bank_transfers_with_entries = []
        for transfer in bank_transfers:
            journal_entry = JournalEntry.objects.filter(
                reference_type='bank_transfer',
                reference_id=transfer.id
            ).first()
            
            bank_transfers_with_entries.append({
                'transfer': transfer,
                'journal_entry': journal_entry
            })
        
        # الحصول على التحويلات بين البنوك والصناديق
        bank_cashbox_transfers = CashboxTransfer.objects.select_related(
            'from_bank', 'to_bank', 'from_cashbox', 'to_cashbox', 'created_by'
        ).filter(transfer_type__in=['bank_to_cashbox', 'cashbox_to_bank'])
        
        # إضافة معلومات القيد المحاسبي للتحويلات بين البنوك والصناديق
        cashbox_transfers_with_entries = []
        for transfer in bank_cashbox_transfers:
            journal_entry = JournalEntry.objects.filter(
                reference_type='cashbox_transfer',
                reference_id=transfer.id
            ).first()
            
            cashbox_transfers_with_entries.append({
                'transfer': transfer,
                'journal_entry': journal_entry
            })
        
        # حساب الإحصائيات
        total_transfers = bank_transfers.count() + bank_cashbox_transfers.count()
        
        # إجمالي المبالغ والرسوم للتحويلات البنكية
        bank_totals = bank_transfers.aggregate(
            total_amount=Sum('amount'),
            total_fees=Sum('fees')
        )
        
        # إجمالي المبالغ والرسوم للتحويلات البنك/صندوق
        cashbox_totals = bank_cashbox_transfers.aggregate(
            total_amount=Sum('amount'),
            total_fees=Sum('fees')
        )
        
        total_amounts = (bank_totals['total_amount'] or 0) + (cashbox_totals['total_amount'] or 0)
        total_fees = (bank_totals['total_fees'] or 0) + (cashbox_totals['total_fees'] or 0)
        
        # تحويلات هذا الشهر
        current_month = datetime.now().month
        current_year = datetime.now().year
        
        this_month_bank = bank_transfers.filter(
            date__month=current_month,
            date__year=current_year
        ).count()
        
        this_month_cashbox = bank_cashbox_transfers.filter(
            date__month=current_month,
            date__year=current_year
        ).count()
        
        this_month_transfers = this_month_bank + this_month_cashbox
        
        context['bank_transfers'] = bank_transfers_with_entries
        context['bank_cashbox_transfers'] = cashbox_transfers_with_entries
        context['total_transfers'] = total_transfers
        context['total_amounts'] = total_amounts
        context['total_fees'] = total_fees
        context['this_month_transfers'] = this_month_transfers
        
        return context

class BankTransferCreateView(LoginRequiredMixin, View):
    template_name = 'banks/transfer_add.html'
    
    def get(self, request, *args, **kwargs):
        if not (request.user.is_superuser or request.user.has_perm('banks.add_banktransfer')):
            messages.error(request, _('You do not have permission to add bank transfers.'))
            return redirect('banks:transfer_list')
        from core.models import DocumentSequence
        from cashboxes.models import Cashbox
        
        context = {
            'accounts': BankAccount.objects.filter(is_active=True),
            'cashboxes': Cashbox.objects.filter(is_active=True),
            'currencies': Currency.get_active_currencies(),
            'base_currency': Currency.get_base_currency()
        }
        
        # التحقق من وجود تسلسل التحويلات البنكية
        try:
            sequence = DocumentSequence.objects.get(document_type='bank_transfer')
            context['transfer_sequence'] = sequence
        except DocumentSequence.DoesNotExist:
            context['transfer_sequence'] = None
            messages.warning(request, _('Warning: Bank transfer number sequence not configured! Please add "Bank Account Transfer" sequence from settings page before creating any transfer.'))
        
        return render(request, self.template_name, context)
    
    def post(self, request, *args, **kwargs):
        # فحص نوع التحويل أولاً
        transfer_type = request.POST.get('transfer_type', 'bank_to_bank')
        
        if transfer_type == 'bank_to_bank':
            return self._handle_bank_to_bank_transfer(request)
        elif transfer_type == 'bank_to_cashbox':
            return self._handle_bank_to_cashbox_transfer(request)
        elif transfer_type == 'cashbox_to_bank':
            return self._handle_cashbox_to_bank_transfer(request)
        else:
            messages.error(request, 'Invalid transfer type!')
            return redirect('banks:transfer_add')
    
    def _handle_bank_to_bank_transfer(self, request):
        """معالجة التحويل بين البنوك"""
        try:
            # Receive data from the form
            date = request.POST.get('date', '').strip()
            from_account_id = request.POST.get('from_account', '').strip()
            to_account_id = request.POST.get('to_account', '').strip()
            amount = request.POST.get('amount', '0')
            fees = request.POST.get('fees', '0')
            exchange_rate = request.POST.get('exchange_rate', '1')
            description = request.POST.get('description', '').strip()
            
            # إعداد السياق لحالات الخطأ
            from cashboxes.models import Cashbox
            context = {
                'accounts': BankAccount.objects.filter(is_active=True),
                'cashboxes': Cashbox.objects.filter(is_active=True),
                'currencies': Currency.get_active_currencies(),
                'base_currency': Currency.get_base_currency()
            }
            
            # Validate data
            if not date:
                messages.error(request, 'Transfer date is required!')
                return render(request, self.template_name, context)
            
            if not from_account_id or not to_account_id:
                messages.error(request, 'Sender and receiver accounts must be selected!')
                return render(request, self.template_name, context)
            
            if from_account_id == to_account_id:
                messages.error(request, 'Cannot transfer from account to itself!')
                return render(request, self.template_name, context)
            
            # التحقق من وجود الحسابات
            try:
                from_account = BankAccount.objects.get(id=from_account_id, is_active=True)
                to_account = BankAccount.objects.get(id=to_account_id, is_active=True)
            except BankAccount.DoesNotExist:
                messages.error(request, 'One of the selected accounts does not exist or is inactive!')
                return render(request, self.template_name, context)
            
            # تحويل المبالغ إلى أرقام
            try:
                amount = Decimal(clean_decimal_input(amount))
                if amount <= 0:
                    raise ValueError
            except (ValueError, InvalidOperation):
                messages.error(request, 'Transfer amount must be a positive number!')
                return render(request, self.template_name, context)
            
            try:
                fees = Decimal(clean_decimal_input(fees))
                if fees < 0:
                    raise ValueError
            except (ValueError, InvalidOperation):
                messages.error(request, 'Transfer fees must be a non-negative number!')
                return render(request, self.template_name, context)
            
            try:
                exchange_rate = Decimal(clean_decimal_input(exchange_rate))
                if exchange_rate <= 0:
                    raise ValueError
            except (ValueError, InvalidOperation):
                messages.error(request, 'Exchange rate must be a positive number!')
                return render(request, self.template_name, context)
            
            # التحقق من كفاية الرصيد
            total_amount = amount + fees
            if from_account.balance < total_amount:
                messages.error(request, f'رصيد الحساب "{from_account.name}" غير كافي! الرصيد الحالي: {from_account.balance:.3f}, المبلغ المطلوب: {total_amount:.3f}')
                return render(request, self.template_name, context)
            
            # إنشاء التحويل البنكي داخل معاملة
            from core.models import DocumentSequence
            from django.db import transaction
            
            with transaction.atomic():
                # الحصول على رقم التحويل من نظام تسلسل المستندات داخل المعاملة
                try:
                    sequence = DocumentSequence.objects.get(document_type='bank_transfer')
                    transfer_number = sequence.get_next_number()
                except DocumentSequence.DoesNotExist:
                    messages.error(request, 'Bank transfer number sequence not configured! Please add "Bank Account Transfer" sequence from settings page.')
                    return render(request, self.template_name, context)
                
                # إنشاء التحويل البنكي
                transfer = BankTransfer.objects.create(
                    transfer_number=transfer_number,
                    date=date,
                    from_account=from_account,
                    to_account=to_account,
                    amount=amount,
                    fees=fees,
                    exchange_rate=exchange_rate,
                    description=description,
                    created_by=request.user
                )
                
                # إنشاء معاملات البنك بدلاً من التعديل المباشر للأرصدة
                
                # إنشاء حركة الخصم من الحساب المرسل
                withdrawal = BankTransaction(
                    bank=from_account,
                    transaction_type='withdrawal',
                    amount=total_amount,
                    description=f'تحويل إلى حساب {to_account.name} - رقم التحويل: {transfer_number}',
                    reference_number=transfer_number,
                    date=date,
                    created_by=request.user
                )
                # علم مؤقت لإعلام إشارات BankTransaction بعدم إنشاء قيد تلقائي
                withdrawal._skip_journal = True
                withdrawal.save()

                # إنشاء حركة الإيداع للحساب المستقبل
                deposit = BankTransaction(
                    bank=to_account,
                    transaction_type='deposit',
                    amount=amount * exchange_rate,
                    description=f'تحويل من حساب {from_account.name} - رقم التحويل: {transfer_number}',
                    reference_number=transfer_number,
                    date=date,
                    created_by=request.user
                )
                deposit._skip_journal = True
                deposit.save()
                
                # لا حاجة لاستدعاء sync_balance هنا - سيتم تلقائياً عند save() للـ BankTransaction
                
                # إنشاء القيد المحاسبي للتحويل
                JournalService.create_bank_transfer_entry(transfer, request.user)
                
                # تسجيل النشاط في سجل الأنشطة
                log_user_activity(
                    request,
                    'CREATE',
                    transfer,
                    f'Bank transfer number {transfer.transfer_number} created from {from_account.name} to {to_account.name} with amount {amount:.3f}'
                )
            
            messages.success(request, f'Bank transfer "{transfer.transfer_number}" created successfully!')
            return redirect('banks:transfer_list')
            
        except Exception as e:
            messages.error(request, f'Error occurred while creating transfer: {str(e)}')
            return render(request, self.template_name, context)

class BankCashboxTransferCreateView(LoginRequiredMixin, View):
    template_name = 'banks/bank_cashbox_transfer_add.html'
    
    def get(self, request, *args, **kwargs):
        from cashboxes.models import Cashbox
        from core.models import DocumentSequence
        
        context = {
            'banks': BankAccount.objects.filter(is_active=True),
            'cashboxes': Cashbox.objects.filter(is_active=True),
            'currencies': Currency.get_active_currencies(),
            'base_currency': Currency.get_base_currency()
        }
        
        # الحصول على رقم التحويل التالي
        try:
            sequence = DocumentSequence.objects.get(document_type='bank_cash_transfer')
            context['transfer_sequence'] = sequence
        except DocumentSequence.DoesNotExist:
            messages.error(request, 'Bank-cashbox transfer number sequence not configured! Please add "Bank-Cashbox Transfer" sequence from settings page.')
        
        return render(request, self.template_name, context)
    
    def post(self, request, *args, **kwargs):
        from cashboxes.models import Cashbox, CashboxTransfer, CashboxTransaction
        from banks.models import BankTransaction
        from core.models import DocumentSequence
        from django.db import transaction
        from datetime import datetime, timedelta
        
        try:
            # حماية من الطلبات المكررة - التحقق من وجود طلب مشابه في آخر 10 ثوان
            current_time = datetime.now()
            recent_time = current_time - timedelta(seconds=10)
            
            # Receive data from the form
            date = request.POST.get('date', '').strip()
            transfer_type = request.POST.get('transfer_type', '').strip()
            bank_id = request.POST.get('bank', '').strip()
            cashbox_id = request.POST.get('cashbox', '').strip()
            amount = request.POST.get('amount', '0')
            fees = request.POST.get('fees', '0')
            exchange_rate = request.POST.get('exchange_rate', '1')
            description = request.POST.get('description', '').strip()
            check_number = request.POST.get('check_number', '').strip()
            check_date_raw = request.POST.get('check_date', '').strip()
            check_date = check_date_raw if check_date_raw else None  # تحويل السلسلة الفارغة إلى None
            check_bank_name = request.POST.get('check_bank_name', '').strip()
            
            # التحقق من وجود تحويل مشابه حديث
            try:
                amount_decimal = Decimal(str(amount))
                recent_transfer = CashboxTransfer.objects.filter(
                    transfer_type=transfer_type,
                    amount=amount_decimal,
                    created_by=request.user,
                    created_at__gte=recent_time
                ).first()
                
                if recent_transfer:
                    messages.warning(request, _('A similar transfer was created recently! Please check the transfer list.'))
                    return redirect('banks:transfer_list')
            except (ValueError, InvalidOperation):
                pass  # سيتم التعامل مع الخطأ لاحقاً
            
            # إعداد السياق لحالات الخطأ
            context = {
                'banks': BankAccount.objects.filter(is_active=True),
                'cashboxes': Cashbox.objects.filter(is_active=True),
                'currencies': Currency.get_active_currencies(),
                'base_currency': Currency.get_base_currency()
            }
            
            # Validate data
            if not date:
                messages.error(request, 'Transfer date is required!')
                return render(request, self.template_name, context)
            
            if not transfer_type or transfer_type not in ['bank_to_cashbox', 'cashbox_to_bank']:
                messages.error(request, 'Transfer type must be selected!')
                return render(request, self.template_name, context)
            
            if not bank_id or not cashbox_id:
                messages.error(request, 'Bank and cashbox must be selected!')
                return render(request, self.template_name, context)
            
            if not check_number:
                messages.error(request, 'Check number is required!')
                return render(request, self.template_name, context)
            
            if not check_date:
                messages.error(request, 'Check date is required!')
                return render(request, self.template_name, context)
            
            if not check_bank_name:
                messages.error(request, 'Check bank name is required!')
                return render(request, self.template_name, context)
            
            # التحقق من وجود البنك والصندوق
            try:
                bank = BankAccount.objects.get(id=bank_id, is_active=True)
                cashbox = Cashbox.objects.get(id=cashbox_id, is_active=True)
            except (BankAccount.DoesNotExist, Cashbox.DoesNotExist):
                messages.error(request, 'البنك أو الصندوق المختار غير موجود أو غير نشط!')
                return render(request, self.template_name, context)
            
            # تحويل المبالغ إلى أرقام
            try:
                amount = Decimal(clean_decimal_input(amount))
                if amount <= 0:
                    raise ValueError
            except (ValueError, InvalidOperation):
                messages.error(request, 'Transfer amount must be a positive number!')
                return render(request, self.template_name, context)
            
            try:
                fees = Decimal(clean_decimal_input(fees))
                if fees < 0:
                    raise ValueError
            except (ValueError, InvalidOperation):
                messages.error(request, 'Transfer fees must be a non-negative number!')
                return render(request, self.template_name, context)
            
            try:
                exchange_rate = Decimal(clean_decimal_input(exchange_rate))
                if exchange_rate <= 0:
                    raise ValueError
            except (ValueError, InvalidOperation):
                messages.error(request, 'Exchange rate must be a positive number!')
                return render(request, self.template_name, context)
            
            # التحقق من كفاية الرصيد
            total_amount = amount + fees
            if transfer_type == 'bank_to_cashbox' and bank.balance < total_amount:
                messages.error(request, f'رصيد البنك "{bank.name}" غير كافي! الرصيد الحالي: {bank.balance:.3f}, المبلغ المطلوب: {total_amount:.3f}')
                return render(request, self.template_name, context)
            elif transfer_type == 'cashbox_to_bank' and cashbox.balance < total_amount:
                messages.error(request, f'رصيد الصندوق "{cashbox.name}" غير كافي! الرصيد الحالي: {cashbox.balance:.3f}, المبلغ المطلوب: {total_amount:.3f}')
                return render(request, self.template_name, context)
            
            # إنشاء التحويل
            with transaction.atomic():
                # الحصول على رقم التحويل من نظام تسلسل المستندات داخل المعاملة
                try:
                    sequence = DocumentSequence.objects.get(document_type='bank_cash_transfer')
                    transfer_number = sequence.get_next_number()
                except DocumentSequence.DoesNotExist:
                    messages.error(request, 'Bank-cashbox transfer number sequence not configured! Please add "Bank-Cashbox Transfer" sequence from settings page.')
                    return render(request, self.template_name, context)
                
                if transfer_type == 'bank_to_cashbox':
                    # إنشاء تحويل من البنك إلى الصندوق
                    transfer = CashboxTransfer.objects.create(
                        transfer_number=transfer_number,
                        transfer_type=transfer_type,
                        date=date,
                        from_bank=bank,
                        to_cashbox=cashbox,
                        amount=amount,
                        fees=fees,
                        exchange_rate=exchange_rate,
                        description=f"{description} - شيك رقم: {check_number} - بنك: {check_bank_name}",
                        created_by=request.user
                    )
                    
                    # إضافة حركة البنك (بدون تعديل الرصيد مباشرة)
                    bank_transaction = BankTransaction(
                        bank=bank,
                        transaction_type='withdrawal',
                        amount=total_amount,
                        description=f'تحويل إلى صندوق {cashbox.name} - شيك: {check_number}',
                        reference_number=transfer_number,
                        date=date,
                        created_by=request.user
                    )
                    # تعيين علم لتجنب إنشاء قيد تلقائي من signal
                    bank_transaction._skip_journal = True
                    bank_transaction.save()
                    
                    # مزامنة رصيد البنك من المعاملات
                    bank.sync_balance()
                    
                    # إضافة حركة الصندوق
                    CashboxTransaction.objects.create(
                        cashbox=cashbox,
                        transaction_type='transfer_in',
                        date=date,
                        amount=amount * exchange_rate,
                        description=f'تحويل من بنك {bank.name} - شيك: {check_number}',
                        related_transfer=transfer,
                        reference_type='transfer',
                        reference_id=transfer.id,
                        created_by=request.user
                    )
                    
                    # تحديث رصيد الصندوق (يمكن الإبقاء عليه لأن الصناديق لا تعتمد على نظام المعاملات)
                    cashbox.balance += (amount * exchange_rate)
                    cashbox.save()
                    
                else:  # cashbox_to_bank
                    # إنشاء تحويل من الصندوق إلى البنك
                    transfer = CashboxTransfer.objects.create(
                        transfer_number=transfer_number,
                        transfer_type=transfer_type,
                        date=date,
                        from_cashbox=cashbox,
                        to_bank=bank,
                        amount=amount,
                        fees=fees,
                        exchange_rate=exchange_rate,
                        description=f"{description} - شيك رقم: {check_number} - بنك: {check_bank_name}",
                        created_by=request.user
                    )
                    
                    # إضافة حركة الصندوق
                    CashboxTransaction.objects.create(
                        cashbox=cashbox,
                        transaction_type='transfer_out',
                        date=date,
                        amount=-total_amount,
                        description=f'تحويل إلى بنك {bank.name} - شيك: {check_number}',
                        related_transfer=transfer,
                        reference_type='transfer',
                        reference_id=transfer.id,
                        created_by=request.user
                    )
                    
                    # إضافة حركة البنك
                    bank_transaction = BankTransaction(
                        bank=bank,
                        transaction_type='deposit',
                        amount=amount * exchange_rate,
                        description=f'تحويل من صندوق {cashbox.name} - شيك: {check_number}',
                        reference_number=transfer_number,
                        date=date,
                        created_by=request.user
                    )
                    # تعيين علم لتجنب إنشاء قيد تلقائي من signal
                    bank_transaction._skip_journal = True
                    bank_transaction.save()
                    
                    # تحديث الأرصدة من خلال حساب المعاملات
                    # تحديث رصيد الصندوق بالطريقة الصحيحة (سيتم تطبيق هذا لاحقاً عند إصلاح الصناديق)
                    cashbox.balance -= total_amount
                    cashbox.save()
                    
                    # تحديث رصيد البنك بناءً على المعاملات الفعلية
                    bank.sync_balance()
                
                # إنشاء قيد محاسبي للتحويل بين البنك والصندوق
                try:
                    from journal.services import JournalService
                    journal_entry = JournalService.create_cashbox_transfer_entry(transfer, request.user)
                    print(f"تم إنشاء القيد المحاسبي للتحويل: {journal_entry.entry_number}")
                except Exception as e:
                    print(f"خطأ في إنشاء القيد المحاسبي: {e}")
                    # يمكن الاستمرار لأن التحويل تم بنجاح
            
            messages.success(request, f'Transfer "{transfer.transfer_number}" created successfully!')
            
            # تسجيل النشاط في سجل التدقيق
            from core.models import AuditLog
            AuditLog.objects.create(
                user=request.user,
                action_type='create',
                content_type='CashboxTransfer',
                object_id=transfer.id,
                description=f'إنشاء تحويل {transfer.transfer_type} رقم {transfer.transfer_number} - المبلغ: {amount} - رسوم: {fees}'
            )
            
            return redirect('banks:transfer_list')
            
        except Exception as e:
            messages.error(request, f'Error occurred while creating transfer: {str(e)}')
            return render(request, self.template_name, context)

class BankAccountDetailView(LoginRequiredMixin, DetailView):
    """عرض تفاصيل الحساب البنكي"""
    model = BankAccount
    template_name = 'banks/account_detail.html'
    context_object_name = 'account'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        account = self.get_object()
        
        # حساب الرصيد الفعلي من الحركات
        actual_balance = account.calculate_actual_balance()
        
        # الحصول على الحركات الأخيرة للحساب
        from .models import BankTransaction
        recent_transactions = BankTransaction.objects.filter(
            bank=account
        ).order_by('-date', '-created_at')[:20]
        
        # الحصول على التحويلات المرتبطة بالحساب
        transfers_from = BankTransfer.objects.filter(from_account=account)[:10]
        transfers_to = BankTransfer.objects.filter(to_account=account)[:10]
        
        # الحصول على التحويلات بين البنك والصناديق
        from cashboxes.models import CashboxTransfer
        cashbox_transfers_from = CashboxTransfer.objects.filter(from_bank=account)[:10]
        cashbox_transfers_to = CashboxTransfer.objects.filter(to_bank=account)[:10]
        
        # إحصائيات الحساب
        from django.db.models import Sum, Count
        from datetime import datetime, timedelta
        
        # إحصائيات الشهر الحالي
        current_month = datetime.now().replace(day=1)
        
        this_month_deposits = BankTransaction.objects.filter(
            bank=account,
            transaction_type='deposit',
            date__gte=current_month
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        this_month_withdrawals = BankTransaction.objects.filter(
            bank=account,
            transaction_type='withdrawal',
            date__gte=current_month
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        context.update({
            'recent_transactions': recent_transactions,
            'transfers_from': transfers_from,
            'transfers_to': transfers_to,
            'cashbox_transfers_from': cashbox_transfers_from,
            'cashbox_transfers_to': cashbox_transfers_to,
            'this_month_deposits': this_month_deposits,
            'this_month_withdrawals': this_month_withdrawals,
            'actual_balance': actual_balance,
            'currencies': Currency.get_active_currencies(),
        })
        
        return context




class BankTransferDetailView(LoginRequiredMixin, DetailView):
    """عرض تفاصيل التحويل البنكي"""
    model = BankTransfer
    template_name = 'banks/transfer_detail.html'
    context_object_name = 'transfer'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # إضافة القيود المحاسبية المرتبطة
        from journal.models import JournalEntry
        from django.db.models import Q
        context['journal_entries'] = JournalEntry.objects.filter(
            Q(bank_transfer=self.object) | 
            Q(reference_type='bank_transfer', reference_id=self.object.id)
        ).select_related('created_by').distinct()
        
        return context


class BankTransferUpdateView(LoginRequiredMixin, UpdateView):
    """تعديل التحويل البنكي"""
    model = BankTransfer
    template_name = 'banks/transfer_edit.html'
    fields = ['date', 'amount', 'fees', 'exchange_rate', 'description']
    success_url = reverse_lazy('banks:transfer_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['accounts'] = BankAccount.objects.filter(is_active=True)
        return context
    
    def form_valid(self, form):
        from django.db import transaction
        from journal.services import JournalService
        from core.signals import log_user_activity
        
        transfer = self.object
        
        with transaction.atomic():
            # حفظ التعديلات الأساسية
            response = super().form_valid(form)
            
            # تحديث معاملات البنك المرتبطة
            # حذف المعاملات القديمة
            BankTransaction.objects.filter(
                reference_number=transfer.transfer_number
            ).delete()
            
            # إنشاء معاملات بنكية جديدة
            total_amount = transfer.amount + transfer.fees
            
            # إنشاء حركة الخصم من الحساب المرسل
            BankTransaction.objects.create(
                bank=transfer.from_account,
                transaction_type='withdrawal',
                amount=total_amount,
                description=f'تحويل إلى حساب {transfer.to_account.name} - رقم التحويل: {transfer.transfer_number}',
                reference_number=transfer.transfer_number,
                date=transfer.date,
                created_by=self.request.user
            )
            
            # إنشاء حركة الإيداع للحساب المستقبل
            BankTransaction.objects.create(
                bank=transfer.to_account,
                transaction_type='deposit',
                amount=transfer.amount * transfer.exchange_rate,
                description=f'تحويل من حساب {transfer.from_account.name} - رقم التحويل: {transfer.transfer_number}',
                reference_number=transfer.transfer_number,
                date=transfer.date,
                created_by=self.request.user
            )
            
            # تحديث القيد المحاسبي
            # حذف القيود القديمة
            from journal.models import JournalEntry
            JournalEntry.objects.filter(
                reference_type='bank_transfer',
                reference_id=transfer.id
            ).delete()
            
            # إنشاء قيد محاسبي جديد
            JournalService.create_bank_transfer_entry(transfer, self.request.user)
            
            # تسجيل النشاط في سجل الأنشطة
            log_user_activity(
                self.request,
                'UPDATE',
                transfer,
                f'تم تحديث التحويل البنكي رقم {transfer.transfer_number} من {transfer.from_account.name} إلى {transfer.to_account.name} بمبلغ {transfer.amount:.3f}'
            )
        
        messages.success(self.request, 'تم تحديث التحويل بنجاح!')
        return response


class CashboxTransferDetailView(LoginRequiredMixin, DetailView):
    """عرض تفاصيل التحويل بين البنك والصندوق"""
    template_name = 'banks/cashbox_transfer_detail.html'
    context_object_name = 'transfer'
    
    def get_object(self):
        from cashboxes.models import CashboxTransfer
        return get_object_or_404(CashboxTransfer, pk=self.kwargs['pk'])


class CashboxTransferUpdateView(LoginRequiredMixin, UpdateView):
    """تعديل التحويل بين البنك والصندوق"""
    template_name = 'banks/cashbox_transfer_edit.html'
    fields = ['date', 'amount', 'fees', 'exchange_rate', 'description']
    success_url = reverse_lazy('banks:transfer_list')
    
    def get_object(self):
        from cashboxes.models import CashboxTransfer
        return get_object_or_404(CashboxTransfer, pk=self.kwargs['pk'])
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from cashboxes.models import Cashbox
        context['banks'] = BankAccount.objects.filter(is_active=True)
        context['cashboxes'] = Cashbox.objects.filter(is_active=True)
        return context
    
    def form_valid(self, form):
        messages.success(self.request, 'تم تحديث التحويل بنجاح!')
        return super().form_valid(form)

class BankTransferDeleteView(LoginRequiredMixin, View):
    """حذف التحويل البنكي مع إعادة ضبط الأرصدة"""
    
    def post(self, request, pk, *args, **kwargs):
        from django.http import JsonResponse
        from django.db import transaction
        
        try:
            # الحصول على التحويل
            transfer = BankTransfer.objects.get(id=pk)
            
            # التحقق من الصلاحيات (يمكن إضافة منطق إضافي هنا)
            if not request.user.is_authenticated:
                return JsonResponse({'error': 'غير مصرح لك بهذا الإجراء'}, status=403)
            
            # حفظ بيانات التحويل قبل الحذف
            transfer_number = transfer.transfer_number
            from_account = transfer.from_account
            to_account = transfer.to_account
            amount = transfer.amount
            fees = transfer.fees
            total_amount = amount + fees
            
            with transaction.atomic():
                # حذف المعاملات البنكية المرتبطة أولاً
                from .models import BankTransaction
                BankTransaction.objects.filter(
                    reference_number=transfer_number
                ).delete()
                
                # حذف القيد المحاسبي المرتبط
                from journal.models import JournalEntry
                JournalEntry.objects.filter(
                    reference_type='bank_transfer',
                    reference_id=transfer.id
                ).delete()
                
                # حذف التحويل
                transfer.delete()
                
                # إعادة مزامنة الأرصدة بناءً على المعاملات المتبقية
                from_account.sync_balance()
                to_account.sync_balance()
            
            messages.success(request, f'تم حذف التحويل "{transfer_number}" بنجاح وتم إعادة ضبط الأرصدة.')
            return JsonResponse({'success': True})
            
        except BankTransfer.DoesNotExist:
            return JsonResponse({'error': 'التحويل غير موجود'}, status=404)
        except Exception as e:
            return JsonResponse({'error': f'Error occurred while deleting transfer: {str(e)}'}, status=500)


class CashboxTransferDeleteView(LoginRequiredMixin, View):
    """حذف التحويل بين البنك والصندوق مع إعادة ضبط الأرصدة"""
    
    def post(self, request, pk, *args, **kwargs):
        from django.http import JsonResponse
        from django.db import transaction
        from cashboxes.models import CashboxTransfer, CashboxTransaction
        from .models import BankTransaction
        
        try:
            # الحصول على التحويل
            transfer = CashboxTransfer.objects.get(id=pk)
            
            # التحقق من الصلاحيات
            if not request.user.is_authenticated:
                return JsonResponse({'error': 'غير مصرح لك بهذا الإجراء'}, status=403)
            
            # حفظ بيانات التحويل قبل الحذف
            transfer_number = transfer.transfer_number
            transfer_type = transfer.transfer_type
            amount = transfer.amount
            fees = transfer.fees
            total_amount = amount + fees
            exchange_rate = transfer.exchange_rate
            
            with transaction.atomic():
                # Delete related transactions أولاً
                BankTransaction.objects.filter(
                    reference_number__icontains=transfer_number.split('-')[0] if '-' in transfer_number else transfer_number
                ).delete()
                
                CashboxTransaction.objects.filter(
                    related_transfer=transfer
                ).delete()
                
                # حذف القيد المحاسبي المرتبط
                from journal.models import JournalEntry
                JournalEntry.objects.filter(
                    reference_type='cashbox_transfer',
                    reference_id=transfer.id
                ).delete()
                
                # حذف التحويل
                transfer.delete()
                
                # إعادة مزامنة الأرصدة
                if transfer_type == 'bank_to_cashbox':
                    transfer.from_bank.sync_balance()
                    # للصناديق نحتاج لإعادة حساب الرصيد يدوياً (إذا لم يكن لديها sync_balance)
                    cashbox = transfer.to_cashbox
                    cashbox.save()  # سيؤدي لإعادة حساب الرصيد إذا كان مطبقاً
                    
                elif transfer_type == 'cashbox_to_bank':
                    transfer.to_bank.sync_balance()
                    # للصناديق نحتاج لإعادة حساب الرصيد يدوياً
                    cashbox = transfer.from_cashbox
                    cashbox.save()  # سيؤدي لإعادة حساب الرصيد إذا كان مطبقاً
            
            messages.success(request, f'تم حذف التحويل "{transfer_number}" بنجاح وتم إعادة ضبط الأرصدة.')
            return JsonResponse({'success': True})
            
        except CashboxTransfer.DoesNotExist:
            return JsonResponse({'error': 'التحويل غير موجود'}, status=404)
        except Exception as e:
            return JsonResponse({'error': f'Error occurred while deleting transfer: {str(e)}'}, status=500)

class BankAccountToggleStatusView(LoginRequiredMixin, View):
    """تفعيل/إلغاء تفعيل الحساب البنكي"""
    def post(self, request, pk, *args, **kwargs):
        try:
            account = BankAccount.objects.get(pk=pk)
            account_name = account.name
            
            # تبديل حالة التفعيل
            account.is_active = not account.is_active
            account.save()
            
            if account.is_active:
                messages.success(request, f'تم تفعيل الحساب البنكي "{account_name}" بنجاح!')
            else:
                messages.success(request, f'تم إلغاء تفعيل الحساب البنكي "{account_name}" بنجاح!')
            
        except BankAccount.DoesNotExist:
            messages.error(request, 'Bank account not found!')
        except Exception as e:
            messages.error(request, f'Error occurred while changing account status: {str(e)}')
        
        return redirect('banks:account_list')

class BankAccountTransactionsView(LoginRequiredMixin, TemplateView):
    """عرض وإدارة حركات الحساب البنكي"""
    template_name = 'banks/account_transactions.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        account_id = self.kwargs.get('pk')
        account = get_object_or_404(BankAccount, pk=account_id)
        
        # تسجيل النشاط في سجل الأنشطة
        log_user_activity(
            self.request,
            'ACCESS',
            account,
            _('Bank Account Transactions Accessed: {}').format(account.name)
        )
        
        # مزامنة الرصيد للتأكد من أنه محدث
        account.sync_balance()
        
        # الحصول على جميع الحركات مرتبة تصاعدياً حسب التاريخ والإنشاء
        from .models import BankTransaction
        transactions = BankTransaction.objects.filter(bank=account).order_by('date', 'created_at')
        
        # حساب الرصيد المتراكم لكل حركة
        from decimal import Decimal
        running_balance = account.initial_balance
        for transaction in transactions:
            # للمعاملات الافتتاحية، الرصيد المتراكم هو قيمة المعاملة نفسها
            if transaction.is_opening_balance:
                transaction.running_balance = transaction.amount
                continue
            # حساب الرصيد المتراكم بناءً على نوع المعاملة
            if transaction.transaction_type == 'deposit':
                running_balance += transaction.amount
            else:  # withdrawal
                running_balance -= transaction.amount
            transaction.running_balance = running_balance
        
        # عكس الترتيب للعرض (الأحدث أولاً)
        transactions = list(reversed(transactions))
        
        # البحث عن transfer_id للمعاملات التي لديها reference_number
        for transaction in transactions:
            if transaction.reference_number:
                from .models import BankTransfer
                from cashboxes.models import CashboxTransfer
                transfer = BankTransfer.objects.filter(transfer_number=transaction.reference_number).first()
                if transfer:
                    transaction.transfer_id = transfer.pk
                    transaction.transfer_type = 'bank_transfer'
                else:
                    # البحث عن CashboxTransfer بطرق مختلفة
                    cashbox_transfer = None
                    
                    # أولاً، البحث بـ transfer_number
                    cashbox_transfer = CashboxTransfer.objects.filter(transfer_number=transaction.reference_number).first()
                    
                    if not cashbox_transfer:
                        # البحث بالحساب والمبلغ والتاريخ
                        cashbox_transfers = CashboxTransfer.objects.filter(
                            date=transaction.date,
                            created_by=transaction.created_by
                        )
                        for ct in cashbox_transfers:
                            expected_amount = ct.amount
                            if ct.fees and transaction.transaction_type == 'withdrawal' and hasattr(ct, 'from_bank') and ct.from_bank == transaction.bank:
                                expected_amount += ct.fees
                            elif ct.fees and transaction.transaction_type == 'deposit' and hasattr(ct, 'to_bank') and ct.to_bank == transaction.bank:
                                expected_amount = ct.amount * ct.exchange_rate
                            
                            if abs(transaction.amount - expected_amount) < 0.01:
                                if ((hasattr(ct, 'from_bank') and ct.from_bank == transaction.bank) or
                                    (hasattr(ct, 'to_bank') and ct.to_bank == transaction.bank)):
                                    cashbox_transfer = ct
                                    break
                    
                    if cashbox_transfer:
                        transaction.transfer_id = cashbox_transfer.pk
                        transaction.transfer_type = 'cashbox_transfer'
                    else:
                        transaction.transfer_id = None
                        transaction.transfer_type = None
            else:
                transaction.transfer_id = None
                transaction.transfer_type = None
        
        # معلومات من الجلسة إذا كان هذا جزء من عملية حذف
        delete_mode = self.request.session.get('delete_account_id') == account.pk
        
        context.update({
            'account': account,
            'transactions': transactions,
            'transactions_count': len(transactions),
            'delete_mode': delete_mode,
            # إظهار زر الحذف النهائي لمن يملك صلاحية الحذف (المخصصة أو الافتراضية أو سوبر)
            # الشرط: لا معاملات على الحساب (حتى لو الرصيد غير صفر)
        'can_delete_account': (
                delete_mode
                and len(transactions) == 0
                and (
                    self.request.user.is_superuser
                    or self.request.user.has_perm('banks.delete_bankaccount')
            or self.request.user.has_perm('banks.can_delete_banks_account')
            or self.request.user.has_perm('users.can_delete_accounts')
                )
            ),
        })
        
        return context

class BankTransactionDeleteView(LoginRequiredMixin, View):
    """حذف حركة بنكية مع إعادة ضبط الرصيد"""
    
    def post(self, request, pk, *args, **kwargs):
        from django.http import JsonResponse
        from django.db import transaction
        from .models import BankTransaction
        
        try:
            # الحصول على المعاملة
            bank_transaction = BankTransaction.objects.get(id=pk)
            bank_account = bank_transaction.bank
            transaction_info = {
                'type': bank_transaction.get_transaction_type_display(),
                'amount': float(bank_transaction.amount),
                'date': bank_transaction.date.strftime('%Y-%m-%d'),
                'description': bank_transaction.description
            }
            
            with transaction.atomic():
                # التحقق إذا كانت المعاملة جزء من تحويل
                reference_number = bank_transaction.reference_number
                
                if reference_number:
                    # البحث عن التحويل المرتبط
                    from banks.models import BankTransfer
                    related_transfer = BankTransfer.objects.filter(transfer_number=reference_number).first()
                    
                    if related_transfer:
                        # إذا وُجد تحويل، احذف جميع المعاملات المرتبطة به
                        affected_accounts = {related_transfer.from_account, related_transfer.to_account}
                        
                        # حذف جميع المعاملات المرتبطة بهذا التحويل
                        BankTransaction.objects.filter(reference_number=reference_number).delete()
                        
                        # حذف التحويل نفسه
                        related_transfer.delete()
                        
                        # إعادة مزامنة جميع الحسابات المتأثرة
                        for account in affected_accounts:
                            account.sync_balance()
                    else:
                        # لا يوجد تحويل مرتبط، ربما معاملة تحويل منفردة
                        # ابحث عن معاملات أخرى بنفس reference_number واحذفها جميعاً
                        related_transactions = BankTransaction.objects.filter(reference_number=reference_number)
                        affected_accounts = set()
                        
                        for rel_trans in related_transactions:
                            affected_accounts.add(rel_trans.bank)
                        
                        # حذف جميع المعاملات المرتبطة
                        related_transactions.delete()
                        
                        # إعادة مزامنة جميع الحسابات المتأثرة
                        for account in affected_accounts:
                            account.sync_balance()
                else:
                    # معاملة عادية غير مرتبطة بتحويل
                    bank_transaction.delete()
                    bank_account.sync_balance()
            
            return JsonResponse({
                'success': True, 
                'message': f'تم حذف المعاملة ({transaction_info["type"]} - {transaction_info["amount"]}) بنجاح وتم إعادة ضبط الرصيد.',
                'new_balance': float(bank_account.balance)
            })
            
        except BankTransaction.DoesNotExist:
            return JsonResponse({'error': 'المعاملة غير موجودة'}, status=404)
        except Exception as e:
            return JsonResponse({'error': f'Error occurred while deleting transaction: {str(e)}'}, status=500)

class BankAccountSuperAdminDeleteView(LoginRequiredMixin, View):
    """حذف مطلق للحسابات البنكية - للـ superadmin فقط"""
    
    def post(self, request, pk, *args, **kwargs):
        from django.db import transaction
        
        # التحقق من أن المستخدم superadmin فقط
        if not request.user.is_superuser:
            messages.error(request, _('Only superadmin can perform absolute deletion.'))
            return redirect('banks:account_list')
        
        try:
            account = BankAccount.objects.get(pk=pk)
            account_name = account.name
            
            with transaction.atomic():
                # حذف جميع المعاملات المرتبطة بالحساب
                from .models import BankTransaction
                BankTransaction.objects.filter(bank=account).delete()
                
                # حذف جميع التحويلات المرتبطة بالحساب
                BankTransfer.objects.filter(from_account=account).delete()
                BankTransfer.objects.filter(to_account=account).delete()
                
                # حذف جميع تحويلات الصناديق المرتبطة بالحساب
                from cashboxes.models import CashboxTransfer
                CashboxTransfer.objects.filter(from_bank=account).delete()
                CashboxTransfer.objects.filter(to_bank=account).delete()
                
                # حذف الحساب نفسه
                account.delete()
                
                # تسجيل النشاط في سجل الأنشطة
                from core.signals import log_activity
                log_activity(
                    action_type='delete',
                    obj=None,  # Object deleted
                    description=f'حذف مطلق للحساب البنكي "{account_name}" مع جميع حركاته وتحويلاته (superadmin)',
                    user=request.user
                )
                
                # مسح بيانات الجلسة
                if 'delete_account_id' in request.session:
                    del request.session['delete_account_id']
                if 'delete_account_name' in request.session:
                    del request.session['delete_account_name']
                
                messages.success(request, f'تم الحذف المطلق للحساب البنكي "{account_name}" مع جميع حركاته وتحويلاته بنجاح!')
            
            return redirect('banks:account_list')
            
        except BankAccount.DoesNotExist:
            messages.error(request, 'Bank account not found!')
            return redirect('banks:account_list')
        except Exception as e:
            messages.error(request, f'Error occurred while permanent deletion: {str(e)}')
            return redirect('banks:account_list')

class BankAccountForceDeleteView(LoginRequiredMixin, View):
    """حذف الحساب البنكي بالقوة بعد إزالة جميع الحركات"""
    
    def post(self, request, pk, *args, **kwargs):
        from django.db import transaction
        try:
            # تحقق صلاحيات الحذف: سوبر أو delete_bankaccount أو الصلاحية المخصصة
            if not (
                request.user.is_superuser
                or request.user.has_perm('banks.delete_bankaccount')
                or request.user.has_perm('banks.can_delete_banks_account')
                or request.user.has_perm('users.can_delete_accounts')
            ):
                messages.error(request, _('You do not have permission to delete bank accounts.'))
                return redirect('banks:account_list')

            account = BankAccount.objects.get(pk=pk)
            account_name = account.name
            
            # Allow direct deletion if no transactions exist, even if balance is not zero
            from .models import BankTransaction
            no_transactions = not BankTransaction.objects.filter(bank=account).exists()
            if no_transactions:
                with transaction.atomic():
                    # Delete related transfers احترازياً
                    BankTransfer.objects.filter(from_account=account).delete()
                    BankTransfer.objects.filter(to_account=account).delete()
                    from cashboxes.models import CashboxTransfer
                    CashboxTransfer.objects.filter(from_bank=account).delete()
                    CashboxTransfer.objects.filter(to_bank=account).delete()
                    # حذف الحساب
                    account.delete()
                    # تنظيف بيانات الجلسة
                    if 'delete_account_id' in request.session:
                        del request.session['delete_account_id']
                    if 'delete_account_name' in request.session:
                        del request.session['delete_account_name']
                messages.success(request, f'Bank account "{account_name}" deleted because it has no transactions.')
                return redirect('banks:account_list')
            
            # التحقق من صلاحية الحذف في حال وجود معاملات: يتطلب رصيد صفر وعدم وجود معاملات
            if account.balance != 0:
                messages.error(request, f'لا يمكن حذف الحساب لأن الرصيد غير صفر: {account.balance}')
                return redirect('banks:account_transactions', pk=pk)
            
            # التحقق من عدم وجود معاملات (في هذا المسار يجب أن لا توجد معاملات للحذف)
            if BankTransaction.objects.filter(bank=account).exists():
                messages.error(request, 'لا يمكن حذف الحساب لأن هناك معاملات مرتبطة به')
                return redirect('banks:account_transactions', pk=pk)

            # في حال تحقق الشروط: لا معاملات ورصيد صفر
            with transaction.atomic():
                BankTransfer.objects.filter(from_account=account).delete()
                BankTransfer.objects.filter(to_account=account).delete()
                from cashboxes.models import CashboxTransfer
                CashboxTransfer.objects.filter(from_bank=account).delete()
                CashboxTransfer.objects.filter(to_bank=account).delete()
                account.delete()
                if 'delete_account_id' in request.session:
                    del request.session['delete_account_id']
                if 'delete_account_name' in request.session:
                    del request.session['delete_account_name']
            messages.success(request, f'تم حذف الحساب البنكي "{account_name}" بنجاح.')
            return redirect('banks:account_list')
            
            # التحقق من التحويلات
            bank_transfers_from = BankTransfer.objects.filter(from_account=account)
            bank_transfers_to = BankTransfer.objects.filter(to_account=account)
            
            if (bank_transfers_from.exists() or bank_transfers_to.exists()) and account.balance != 0:
                transfers_count = bank_transfers_from.count() + bank_transfers_to.count()
                messages.error(request, f'لا يمكن حذف الحساب لأن هناك {transfers_count} تحويل بنكي ورصيد غير صفر ({account.balance:.3f})')
                return redirect('banks:account_transactions', pk=pk)
            
            # التحقق من تحويلات الصناديق
            from cashboxes.models import CashboxTransfer
            cashbox_transfers_from = CashboxTransfer.objects.filter(from_bank=account)
            cashbox_transfers_to = CashboxTransfer.objects.filter(to_bank=account)
            
            if (cashbox_transfers_from.exists() or cashbox_transfers_to.exists()) and account.balance != 0:
                transfers_count = cashbox_transfers_from.count() + cashbox_transfers_to.count()
                messages.error(request, f'لا يمكن حذف الحساب لأن هناك {transfers_count} تحويل بين البنوك والصناديق ورصيد غير صفر ({account.balance:.3f})')
                return redirect('banks:account_transactions', pk=pk)
            
            # إذا مرت جميع الفحوصات، حذف الحساب
            with transaction.atomic():
                # If balance is zero, delete transactions and transfers first
                if account.balance == 0:
                    # Delete related transactions
                    from .models import BankTransaction
                    BankTransaction.objects.filter(bank=account).delete()
                    
                    # Delete related transfers
                    BankTransfer.objects.filter(from_account=account).delete()
                    BankTransfer.objects.filter(to_account=account).delete()
                    
                    # حذف تحويلات الصناديق المرتبطة
                    from cashboxes.models import CashboxTransfer
                    CashboxTransfer.objects.filter(from_bank=account).delete()
                    CashboxTransfer.objects.filter(to_bank=account).delete()
                
                # حذف الحساب
                account.delete()
                
                # تسجيل النشاط في سجل الأنشطة
                from core.signals import log_activity
                log_activity(
                    action_type='delete',
                    obj=None,  # Object deleted
                    description=f'حذف الحساب البنكي "{account_name}" نهائياً',
                    user=request.user
                )
                
                # مسح بيانات الجلسة
                if 'delete_account_id' in request.session:
                    del request.session['delete_account_id']
                if 'delete_account_name' in request.session:
                    del request.session['delete_account_name']
                
                messages.success(request, f'Bank account "{account_name}" deleted successfully!')
            
        except BankAccount.DoesNotExist:
            messages.error(request, 'Bank account not found!')
            return redirect('banks:account_list')
        except Exception as e:
            messages.error(request, f'Error occurred while deleting account: {str(e)}')
            return redirect('banks:account_transactions', pk=pk)

class ClearAccountTransactionsView(LoginRequiredMixin, View):
    """حذف جميع حركات الحساب - للـ Superadmin فقط"""
    
    def post(self, request, pk):
        # التحقق من صلاحية Superadmin
        if not request.user.is_superuser:
            messages.error(request, 'غير مسموح لك بهذا الإجراء!')
            return redirect('banks:account_list')
        
        try:
            account = get_object_or_404(BankAccount, pk=pk)
            
            # حساب العناصر المراد حذفها
            transactions_count = BankTransaction.objects.filter(bank=account).count()
            
            # Delete related transfers بالحساب
            bank_transfers_from = BankTransfer.objects.filter(from_account=account)
            bank_transfers_to = BankTransfer.objects.filter(to_account=account)
            transfers_count = bank_transfers_from.count() + bank_transfers_to.count()
            
            # حذف تحويلات البنك-الصندوق المرتبطة
            from cashboxes.models import CashboxTransfer
            cashbox_transfers_from = CashboxTransfer.objects.filter(from_bank=account)
            cashbox_transfers_to = CashboxTransfer.objects.filter(to_bank=account)
            cashbox_transfers_count = cashbox_transfers_from.count() + cashbox_transfers_to.count()
            
            # تنفيذ الحذف داخل معاملة قاعدة البيانات
            from django.db import transaction
            with transaction.atomic():
                # جمع جميع الحسابات المتأثرة
                affected_accounts = set([account])
                
                # إضافة الحسابات المتأثرة بالتحويلات البنكية
                for transfer in bank_transfers_from:
                    affected_accounts.add(transfer.to_account)
                for transfer in bank_transfers_to:
                    affected_accounts.add(transfer.from_account)
                
                # إضافة الحسابات المتأثرة بتحويلات البنك-الصندوق
                for transfer in cashbox_transfers_from:
                    affected_accounts.add(account)
                for transfer in cashbox_transfers_to:
                    affected_accounts.add(transfer.from_bank)
                
                # حذف جميع حركات الحساب
                BankTransaction.objects.filter(bank=account).delete()
                
                # حذف التحويلات البنكية المرتبطة
                bank_transfers_from.delete()
                bank_transfers_to.delete()
                
                # حذف تحويلات البنك-الصندوق المرتبطة
                cashbox_transfers_from.delete()
                cashbox_transfers_to.delete()
                
                # إعادة مزامنة جميع الحسابات المتأثرة
                for affected_account in affected_accounts:
                    affected_account.sync_balance()
            
            # رسالة النجاح الشاملة
            total_deleted = transactions_count + transfers_count + cashbox_transfers_count
            details = []
            if transactions_count > 0:
                details.append(f'{transactions_count} حركة مالية')
            if transfers_count > 0:
                details.append(f'{transfers_count} تحويل بنكي')
            if cashbox_transfers_count > 0:
                details.append(f'{cashbox_transfers_count} تحويل بنك-صندوق')
            
            details_text = ' و '.join(details) if details else 'لا توجد عناصر'
            
            messages.success(
                request, 
                f'تم حذف {details_text} من حساب "{account.name}" وإعادة تعيين الرصيد إلى {account.initial_balance:.3f}'
            )
            
        except Exception as e:
            messages.error(request, f'Error occurred while deleting movements: {str(e)}')
        
        return redirect('banks:account_list')


# Bank Reconciliation Views
class BankReconciliationListView(LoginRequiredMixin, BanksViewPermissionMixin, ListView):
    model = BankReconciliation
    template_name = 'banks/reconciliation_list.html'
    context_object_name = 'reconciliations'
    paginate_by = 25

    def get_queryset(self):
        queryset = super().get_queryset()
        bank_account = self.request.GET.get('bank_account')
        status = self.request.GET.get('status')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')

        if bank_account:
            queryset = queryset.filter(bank_account_id=bank_account)
        if status:
            queryset = queryset.filter(status=status)
        if date_from:
            queryset = queryset.filter(statement_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(statement_date__lte=date_to)

        return queryset.order_by('-statement_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['bank_accounts'] = BankAccount.objects.filter(is_active=True)
        context['status_choices'] = BankReconciliation._meta.get_field('status').choices

        # تسجيل النشاط
        log_user_activity(
            self.request,
            'ACCESS',
            None,
            _('Bank Reconciliations List Accessed')
        )

        return context


class BankReconciliationCreateView(LoginRequiredMixin, BanksViewPermissionMixin, CreateView):
    model = BankReconciliation
    template_name = 'banks/reconciliation_form.html'
    fields = ['bank_account', 'statement_date', 'statement_balance', 'deposits_in_transit',
             'outstanding_checks', 'bank_charges', 'interest_earned', 'other_adjustments', 'notes']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        bank_account_id = self.request.GET.get('bank_account')
        if bank_account_id:
            try:
                bank_account = BankAccount.objects.get(id=bank_account_id)
                context['bank_account'] = bank_account
                # Calculate book balance
                context['book_balance'] = bank_account.calculate_actual_balance()
            except BankAccount.DoesNotExist:
                pass
        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        # Calculate book balance
        bank_account = form.cleaned_data['bank_account']
        form.instance.book_balance = bank_account.calculate_actual_balance()

        response = super().form_valid(form)

        # تسجيل النشاط
        log_user_activity(
            self.request,
            'CREATE',
            form.instance,
            _('New Bank Reconciliation Created: {}').format(form.instance)
        )

        messages.success(self.request, _('Bank reconciliation created successfully.'))
        return response

    def get_success_url(self):
        return reverse_lazy('banks:reconciliation_detail', kwargs={'pk': self.object.pk})


class BankReconciliationDetailView(LoginRequiredMixin, BanksViewPermissionMixin, DetailView):
    model = BankReconciliation
    template_name = 'banks/reconciliation_detail.html'
    context_object_name = 'reconciliation'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        reconciliation = self.object

        # Get unreconciled statements for this bank account up to statement date
        context['unreconciled_statements'] = BankStatement.objects.filter(
            bank_account=reconciliation.bank_account,
            date__lte=reconciliation.statement_date,
            is_reconciled=False
        ).order_by('date')

        # تسجيل النشاط
        log_user_activity(
            self.request,
            'ACCESS',
            reconciliation,
            _('Bank Reconciliation Details Viewed: {}').format(reconciliation)
        )

        return context


class BankReconciliationUpdateView(LoginRequiredMixin, BanksViewPermissionMixin, UpdateView):
    model = BankReconciliation
    template_name = 'banks/reconciliation_form.html'
    fields = ['statement_balance', 'deposits_in_transit', 'outstanding_checks',
             'bank_charges', 'interest_earned', 'other_adjustments', 'status', 'notes']

    def form_valid(self, form):
        response = super().form_valid(form)

        # تسجيل النشاط
        log_user_activity(
            self.request,
            'UPDATE',
            form.instance,
            _('Bank Reconciliation Updated: {}').format(form.instance)
        )

        messages.success(self.request, _('Bank reconciliation updated successfully.'))
        return response

    def get_success_url(self):
        return reverse_lazy('banks:reconciliation_detail', kwargs={'pk': self.object.pk})


# Bank Statement Views
class BankStatementListView(LoginRequiredMixin, BanksViewPermissionMixin, ListView):
    model = BankStatement
    template_name = 'banks/statement_list.html'
    context_object_name = 'statements'
    paginate_by = 50

    def get_queryset(self):
        queryset = super().get_queryset()
        bank_account = self.request.GET.get('bank_account')
        reconciled = self.request.GET.get('reconciled')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')

        if bank_account:
            queryset = queryset.filter(bank_account_id=bank_account)
        if reconciled is not None:
            if reconciled == 'yes':
                queryset = queryset.filter(is_reconciled=True)
            elif reconciled == 'no':
                queryset = queryset.filter(is_reconciled=False)
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)

        return queryset.order_by('-date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['bank_accounts'] = BankAccount.objects.filter(is_active=True)

        # تسجيل النشاط
        log_user_activity(
            self.request,
            'ACCESS',
            None,
            _('Bank Statements List Accessed')
        )

        return context


class BankStatementCreateView(LoginRequiredMixin, BanksViewPermissionMixin, CreateView):
    model = BankStatement
    template_name = 'banks/statement_form.html'
    fields = ['bank_account', 'date', 'description', 'reference', 'debit', 'credit', 'balance', 'notes']

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)

        # تسجيل النشاط
        log_user_activity(
            self.request,
            'CREATE',
            form.instance,
            _('New Bank Statement Created: {}').format(form.instance)
        )

        messages.success(self.request, _('Bank statement created successfully.'))
        return response

    def get_success_url(self):
        return reverse_lazy('banks:statement_list')


class BankStatementUpdateView(LoginRequiredMixin, BanksViewPermissionMixin, UpdateView):
    model = BankStatement
    template_name = 'banks/statement_form.html'
    fields = ['description', 'reference', 'debit', 'credit', 'balance', 'is_reconciled', 'notes']

    def form_valid(self, form):
        response = super().form_valid(form)

        # تسجيل النشاط
        log_user_activity(
            self.request,
            'UPDATE',
            form.instance,
            _('Bank Statement Updated: {}').format(form.instance)
        )

        messages.success(self.request, _('Bank statement updated successfully.'))
        return response

    def get_success_url(self):
        return reverse_lazy('banks:statement_list')


@login_required
def account_export_xlsx(request, pk):
    """تصدير معاملات الحساب البنكي إلى ملف Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from django.utils.translation import gettext as _
    from core.signals import log_user_activity
    from decimal import Decimal
    
    try:
        account = get_object_or_404(BankAccount, pk=pk)
        
        # تسجيل النشاط في سجل الأنشطة
        log_user_activity(
            request,
            'EXPORT',
            account,
            _('Bank Account Transactions Exported to Excel: {}').format(account.name)
        )
        
        # إنشاء ملف Excel جديد
        wb = Workbook()
        ws = wb.active
        ws.title = _("Bank Account Statement")
        
        # تنسيق العناوين
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        # إضافة معلومات الحساب
        ws['A1'] = _("Bank Account Statement")
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = center_alignment
        
        ws['A3'] = _("Account Name") + ":"
        ws['B3'] = account.name
        
        ws['A4'] = _("Bank Name") + ":"
        ws['B4'] = account.bank_name
        
        ws['A5'] = _("Account Number") + ":"
        ws['B5'] = account.account_number
        
        ws['A6'] = _("Current Balance") + ":"
        ws['B6'] = f"{account.balance} {account.get_currency_symbol()}"
        
        ws['A7'] = _("Transactions Count") + ":"
        transactions_count = BankTransaction.objects.filter(bank=account).count()
        ws['B7'] = transactions_count
        
        ws['A8'] = _("Export Date") + ":"
        ws['B8'] = str(timezone.now().date())
        
        # إضافة عناوين الأعمدة
        headers = [
            _("Date"),
            _("Transaction Type"),
            _("Reference Number"),
            _("Description"),
            _("Amount"),
            _("Balance After Transaction"),
            _("Created By")
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # الحصول على جميع المعاملات
        transactions = BankTransaction.objects.filter(bank=account).order_by('date', 'created_at')
        
        # حساب الرصيد المتراكم
        running_balance = account.initial_balance or Decimal('0')
        for transaction in transactions:
            if transaction.transaction_type == 'deposit':
                running_balance += transaction.amount
            else:  # withdrawal
                running_balance -= transaction.amount
            transaction.running_balance = running_balance
        
        # عكس الترتيب للعرض (الأحدث أولاً)
        transactions = list(reversed(transactions))
        
        # إضافة بيانات المعاملات
        for row_num, transaction in enumerate(transactions, 11):
            ws.cell(row=row_num, column=1).value = transaction.date.strftime('%Y-%m-%d')
            ws.cell(row=row_num, column=2).value = transaction.get_transaction_type_display()
            ws.cell(row=row_num, column=3).value = transaction.reference_number or "-"
            ws.cell(row=row_num, column=4).value = transaction.description
            ws.cell(row=row_num, column=5).value = float(transaction.amount)
            ws.cell(row=row_num, column=6).value = float(transaction.running_balance)
            ws.cell(row=row_num, column=7).value = transaction.created_by.username
        
        # تعديل عرض الأعمدة
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # إنشاء الاستجابة
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"bank_account_statement_{account.name}_{timezone.now().date()}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, _('An error occurred while exporting: {}').format(str(e)))
        return redirect('banks:account_detail', pk=pk)

