from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model

User = get_user_model()
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.db.models import Q, Sum
from django.utils import timezone
from django.core.paginator import Paginator
from decimal import Decimal
from django.views.generic import View
from django.utils.decorators import method_decorator
from django.db.models import ProtectedError
from core.models import AuditLog

from .models import Cashbox, CashboxTransfer, CashboxTransaction
from banks.models import BankAccount


def get_transaction_document_url(transaction):
    """الحصول على رابط صفحة تفاصيل المستند للمعاملة"""
    if transaction.reference_type and transaction.reference_id:
        if transaction.reference_type == 'sales_invoice':
            return f'/ar/sales/invoices/{transaction.reference_id}/'
        elif transaction.reference_type == 'purchase_invoice':
            return f'/ar/purchases/invoices/{transaction.reference_id}/'
        elif transaction.reference_type == 'receipt':
            return f'/ar/receipts/{transaction.reference_id}/'
        elif transaction.reference_type == 'payment':
            return f'/ar/payments/{transaction.reference_id}/'
        elif transaction.reference_type == 'transfer':
            return f'/ar/cashboxes/transfers/{transaction.reference_id}/'
    
    if transaction.related_transfer:
        return f'/ar/cashboxes/transfers/{transaction.related_transfer.id}/'
    
    # محاولة البحث عن المستند من الوصف
    doc_num = get_document_number_from_description(transaction.description)
    if doc_num:
        # البحث في الفواتير
        from sales.models import SalesInvoice
        from purchases.models import PurchaseInvoice
        from receipts.models import PaymentReceipt
        from payments.models import PaymentVoucher
        
        # البحث برقم الفاتورة
        try:
            if 'فاتورة' in transaction.description.lower():
                # فاتورة مبيعات
                invoice = SalesInvoice.objects.filter(invoice_number=doc_num).first()
                if invoice:
                    return f'/ar/sales/invoices/{invoice.id}/'
                
                # فاتورة مشتريات
                invoice = PurchaseInvoice.objects.filter(invoice_number=doc_num).first()
                if invoice:
                    return f'/ar/purchases/invoices/{invoice.id}/'
            
            # البحث في السندات
            receipt = PaymentReceipt.objects.filter(receipt_number=doc_num).first()
            if receipt:
                return f'/ar/receipts/{receipt.id}/'
            
            payment = PaymentVoucher.objects.filter(voucher_number=doc_num).first()
            if payment:
                return f'/ar/payments/{payment.id}/'
                
        except Exception:
            pass
    
    return None


def get_document_number_from_description(description):
    """استخراج رقم المستند من الوصف"""
    import re
    if not description:
        return None
    
    # أنماط شائعة لأرقام المستندات
    patterns = [
        r'فاتورة رقم\s*([A-Za-z0-9\-]+)',  # فاتورة رقم INV-001
        r'رقم\s*([A-Za-z0-9\-]+)',  # رقم INV-001
        r'#([A-Za-z0-9\-]+)',  # #INV-001
        r'([A-Za-z]{2,}-[0-9]+)',  # INV-001, REC-001, PAY-001
        r'([A-Za-z]{3,}[0-9]+)',  # INV001, REC001
        r'سند\s+(قبض|دفع|صرف)\s+([A-Za-z0-9\-]+)',  # سند قبض REC-001
        r'([0-9]{3,}-?[0-9]*)',  # أرقام مثل 001, 001-1, 12345
    ]
    
    for pattern in patterns:
        match = re.search(pattern, description)
        if match:
            if 'سند' in pattern:
                result = match.group(2)
            else:
                result = match.group(1)
            
            # تجنب استخراج التواريخ أو المبالغ
            if len(result) >= 3 and not result.startswith(('20', '19')) and not '.' in result:
                return result
    
    return None


def get_transaction_short_description(transaction):
    """الحصول على وصف مختصر للحركة بناءً على نوع المستند"""
    if transaction.reference_type:
        if transaction.reference_type == 'sales_invoice':
            return 'فاتورة مبيعات-نقدية'
        elif transaction.reference_type == 'purchase_invoice':
            return 'فاتورة مشتريات'
        elif transaction.reference_type == 'receipt':
            return 'إيصال قبض'
        elif transaction.reference_type == 'payment':
            return 'سند صرف'
        elif transaction.reference_type == 'transfer':
            return 'تحويل'
    
    # للحركات التي لا تحتوي على reference_type، نستخدم الوصف الأصلي
    return transaction.description or 'حركة نقدية'


def get_transaction_document_number(transaction):
    """الحصول على رقم المستند البسيط للحركة"""
    if transaction.reference_type and transaction.reference_id:
        # استرجاع رقم المستند الفعلي من النموذج المرتبط
        try:
            if transaction.reference_type == 'payment':
                from payments.models import PaymentVoucher
                voucher = PaymentVoucher.objects.get(id=transaction.reference_id)
                return voucher.voucher_number
            elif transaction.reference_type == 'receipt':
                from receipts.models import Receipt
                receipt = Receipt.objects.get(id=transaction.reference_id)
                return receipt.receipt_number
            elif transaction.reference_type == 'sales_invoice':
                from sales.models import Invoice
                invoice = Invoice.objects.get(id=transaction.reference_id)
                return invoice.invoice_number
            elif transaction.reference_type == 'purchase_invoice':
                from purchases.models import PurchaseInvoice
                invoice = PurchaseInvoice.objects.get(id=transaction.reference_id)
                return invoice.invoice_number
            else:
                return transaction.reference_id
        except Exception:
            return transaction.reference_id
    elif transaction.related_transfer:
        return transaction.related_transfer.transfer_number
    else:
        # محاولة استخراج رقم المستند من الوصف
        doc_num = get_document_number_from_description(transaction.description)
        if doc_num:
            return doc_num
        else:
            return transaction.id


@login_required
def cashbox_list(request):
    """قائمة الصناديق"""
    # التحقق من الصلاحيات
    if not request.user.has_cashboxes_permission():
        messages.error(request, _('You do not have permission to access cashboxes'))
        return redirect('core:dashboard')
    
    from settings.models import Currency
    from banks.models import BankAccount
    from django.utils import timezone
    
    cashboxes = Cashbox.objects.filter(is_active=True).order_by('name')
    
    # مزامنة أرصدة الصناديق
    for cashbox in cashboxes:
        cashbox.sync_balance()
    
    # إحصائيات سريعة
    total_balance = sum(cashbox.balance for cashbox in cashboxes)
    
    # العملات المتاحة
    currencies = Currency.objects.filter(is_active=True).order_by('name')
    base_currency = Currency.get_base_currency()
    
    # البيانات المطلوبة للmodal
    banks = BankAccount.objects.filter(is_active=True).order_by('name')
    
    # الحصول على جميع المستخدمين للاختيار
    users = User.objects.filter(is_active=True).order_by('username')
    
    # الحصول على أرقام التحويل التالية لكل نوع
    from core.models import DocumentSequence
    next_transfer_numbers = {}
    
    # تسلسل التحويلات بين الصناديق
    try:
        sequence = DocumentSequence.objects.get(document_type='cashbox_transfer')
        next_transfer_numbers['cashbox_transfer'] = sequence.peek_next_number()
    except DocumentSequence.DoesNotExist:
        # في حالة عدم وجود تسلسل، استخدم الطريقة القديمة
        prefix = 'CT'
        date_str = timezone.now().strftime('%Y%m%d')
        
        # البحث عن آخر رقم في نفس اليوم
        last_transfer = CashboxTransfer.objects.filter(
            transfer_number__startswith=f'{prefix}{date_str}'
        ).order_by('-transfer_number').first()
        
        if last_transfer:
            last_number = int(last_transfer.transfer_number[-4:])
            new_number = last_number + 1
        else:
            new_number = 1
        
        next_transfer_numbers['cashbox_transfer'] = f'{prefix}{date_str}{new_number:04d}'
    
    # تسلسل التحويلات بين البنوك والصناديق
    try:
        sequence = DocumentSequence.objects.get(document_type='bank_cash_transfer')
        next_transfer_numbers['bank_cash_transfer'] = sequence.peek_next_number()
    except DocumentSequence.DoesNotExist:
        # في حالة عدم وجود تسلسل، استخدم الطريقة القديمة
        prefix = 'BCT'
        date_str = timezone.now().strftime('%Y%m%d')
        
        # البحث عن آخر رقم في نفس اليوم
        last_transfer = CashboxTransfer.objects.filter(
            transfer_number__startswith=f'{prefix}{date_str}'
        ).order_by('-transfer_number').first()
        
        if last_transfer:
            last_number = int(last_transfer.transfer_number[-4:])
            new_number = last_number + 1
        else:
            new_number = 1
        
        next_transfer_numbers['bank_cash_transfer'] = f'{prefix}{date_str}{new_number:04d}'
    
    # استخدم تسلسل التحويل بين البنوك والصناديق كافتراضي
    next_transfer_number = next_transfer_numbers.get('bank_cash_transfer')
    
    context = {
        'cashboxes': cashboxes,
        'total_balance': total_balance,
        'currencies': currencies,
        'base_currency': base_currency,
        'banks': banks,
        'users': users,
        'next_transfer_number': next_transfer_number,
        'next_transfer_numbers': next_transfer_numbers,
        'today': timezone.now().date(),
        'page_title': _('Cashboxes'),
    }
    return render(request, 'cashboxes/cashbox_list.html', context)


@login_required
def cashbox_detail(request, cashbox_id):
    """تفاصيل الصندوق"""
    # التحقق من الصلاحيات
    if not request.user.has_cashboxes_permission():
        messages.error(request, _('You do not have permission to access cashboxes'))
        return redirect('core:dashboard')
    
    from settings.models import Currency
    
    cashbox = get_object_or_404(Cashbox, id=cashbox_id)
    
    # مزامنة رصيد الصندوق مع المعاملات الفعلية
    cashbox.sync_balance()
    
    # الحصول على جميع حركات الصندوق مرتبة تصاعدياً حسب التاريخ والإنشاء
    all_transactions = CashboxTransaction.objects.filter(
        cashbox=cashbox
    ).order_by('date', 'created_at')
    
    # حساب الرصيد المتراكم لكل حركة
    running_balance = Decimal('0')
    for transaction in all_transactions:
        # إضافة المبلغ مباشرة (المبلغ موجب للإيداعات والواردات، سالب للسحوبات والصادرات)
        running_balance += transaction.amount
        transaction.running_balance = running_balance
    
    # عكس الترتيب للعرض (الأحدث أولاً)
    all_transactions = list(reversed(all_transactions))
    
    # التقسيم لصفحات
    paginator = Paginator(all_transactions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # إضافة document_url لكل transaction
    for transaction in page_obj:
        transaction.document_url = get_transaction_document_url(transaction)
        transaction.short_description = get_transaction_short_description(transaction)
        transaction.document_number = get_transaction_document_number(transaction)
    
    # العملات المتاحة
    currencies = Currency.objects.filter(is_active=True).order_by('name')
    
    context = {
        'cashbox': cashbox,
        'transactions': page_obj,
        'currencies': currencies,
        'page_title': f'{_("Cashbox Details")} - {cashbox.name}',
    }
    return render(request, 'cashboxes/cashbox_detail.html', context)


@login_required
def cashbox_create(request):
    """إنشاء صندوق جديد"""
    # التحقق من الصلاحيات - يجب أن لا يكون لديه فقط صلاحية المشاهدة
    if request.user.has_perm('users.can_access_cashboxes') and not (request.user.is_superuser or request.user.user_type in ['superadmin', 'admin']):
        messages.error(request, _('You only have view permission. You cannot create cashboxes.'))
        return redirect('cashboxes:cashbox_list')
    
    if not request.user.has_cashboxes_permission():
        messages.error(request, _('You do not have permission to create cashboxes'))
        return redirect('core:dashboard')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        initial_balance = request.POST.get('initial_balance', 0)
        currency = request.POST.get('currency', '')
        location = request.POST.get('location', '')
        responsible_user_id = request.POST.get('responsible_user')
        
        if not name:
            messages.error(request, _('Cashbox name is required'))
            return redirect('cashboxes:cashbox_list')
        
        try:
            initial_balance_decimal = Decimal(initial_balance or 0)
        except (ValueError, TypeError):
            messages.error(request, _('Initial balance must be a valid number'))
            return redirect('cashboxes:cashbox_list')
        
        # التحقق من responsible_user_id إذا تم تمريره
        responsible_user = None
        if responsible_user_id:
            try:
                responsible_user = User.objects.get(id=responsible_user_id)
            except User.DoesNotExist:
                messages.error(request, _('Selected responsible user does not exist'))
                return redirect('cashboxes:cashbox_list')
        
        try:
            with transaction.atomic():
                # تعيين علم لمنع Signal من إنشاء قيد مكرر إذا كان هناك رصيد افتتاحي
                will_create_opening_entry = initial_balance_decimal > 0
                
                # إنشاء الصندوق برصيد صفر - سيتم تحديثه تلقائياً من المعاملات
                cashbox = Cashbox(
                    name=name,
                    description=description,
                    balance=Decimal('0'),  # برصيد صفر
                    currency=currency,
                    location=location,
                    responsible_user=responsible_user
                )
                
                # تعيين علم لمنع Signal من التدخل عند إنشاء القيد
                if will_create_opening_entry:
                    cashbox._skip_opening_balance_signal = True
                
                cashbox.save()
                
                # إنشاء الحساب الفرعي في journal.Account دائماً
                from journal.services import JournalService
                cashbox_account = JournalService.get_cashbox_account(cashbox)
                
                # إضافة حركة الرصيد الافتتاحي وقيد محاسبي فقط إذا كان أكبر من صفر
                # متوافق مع IFRS - IAS 1 (عرض القوائم المالية)
                if initial_balance_decimal > 0:
                    # التحقق من عدم وجود قيد افتتاحي بالفعل لهذا الصندوق
                    from journal.models import JournalEntry
                    existing_entry = JournalEntry.objects.filter(
                        reference_type='cashbox_initial',
                        reference_id=cashbox.id
                    ).exists()
                    
                    if existing_entry:
                        print(f"⚠ قيد الرصيد الافتتاحي موجود بالفعل للصندوق {cashbox.name}، تم تخطي الإنشاء")
                    else:
                        # إنشاء حركة الرصيد الافتتاحي
                        CashboxTransaction.objects.create(
                            cashbox=cashbox,
                            transaction_type='initial_balance',
                            date=timezone.now().date(),
                            amount=initial_balance_decimal,
                            description=_('Opening Balance'),
                            created_by=request.user
                        )
                        
                        # الحصول على حساب رأس المال
                        # حسب IFRS IAS 1 - الرصيد الافتتاحي يمثل مساهمة في رأس المال
                        from journal.models import Account
                        capital_account = Account.objects.filter(code='301').first()
                        
                        if cashbox_account and capital_account:
                            lines_data = [
                                {
                                    'account_id': cashbox_account.id,
                                    'debit': initial_balance_decimal,
                                    'credit': Decimal('0'),
                                    'description': f'{_("Opening Balance")}: {cashbox.name}'
                                },
                                {
                                    'account_id': capital_account.id,
                                    'debit': Decimal('0'),
                                    'credit': initial_balance_decimal,
                                    'description': f'{_("Opening Balance - Capital Contribution")} - {cashbox.name}'
                                }
                            ]
                            
                            journal_entry = JournalService.create_journal_entry(
                                entry_date=timezone.now().date(),
                                description=f'{_("Opening Balance")}: {cashbox.name}',
                                reference_type='cashbox_initial',
                                reference_id=cashbox.id,
                                lines_data=lines_data,
                                user=request.user
                            )
                            
                            if journal_entry:
                                print(f"✓ تم إنشاء قيد الرصيد الافتتاحي للصندوق {cashbox.name}: {journal_entry.entry_number}")
                        
                        # مزامنة رصيد الصندوق من المعاملات
                        cashbox.sync_balance()
                
                # تسجيل النشاط في سجل الأنشطة
                AuditLog.objects.create(
                    user=request.user,
                    action_type='create',
                    content_type='Cashbox',
                    object_id=cashbox.id,
                    description=f'إنشاء صندوق نقدي جديد: {name}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
                
                messages.success(request, _('Cashbox created successfully'))
                return redirect('cashboxes:cashbox_detail', cashbox_id=cashbox.id)
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"خطأ في إنشاء الصندوق: {e}")
            print(f"تفاصيل الخطأ:\n{error_trace}")
            
            # رسالة خطأ أكثر تفصيلاً
            error_message = str(e) if str(e) else _('An error occurred while creating the cashbox')
            messages.error(request, f'{_("An error occurred while creating the cashbox")}: {error_message}')
    
    return redirect('cashboxes:cashbox_list')


@login_required
def cashbox_edit(request, cashbox_id):
    """تعديل الصندوق"""
    # التحقق من الصلاحيات - يجب أن لا يكون لديه فقط صلاحية المشاهدة
    if request.user.has_perm('users.can_access_cashboxes') and not (request.user.is_superuser or request.user.user_type in ['superadmin', 'admin']):
        messages.error(request, _('You only have view permission. You cannot edit cashboxes.'))
        return redirect('cashboxes:cashbox_list')
    
    if not request.user.has_cashboxes_permission():
        messages.error(request, _('You do not have permission to edit cashboxes'))
        return redirect('core:dashboard')
    
    cashbox = get_object_or_404(Cashbox, id=cashbox_id)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        currency = request.POST.get('currency', '')
        location = request.POST.get('location', '')
        responsible_user_id = request.POST.get('responsible_user')
        initial_balance = request.POST.get('initial_balance')
        
        if not name:
            messages.error(request, _('Cashbox name is required'))
            return redirect('cashboxes:cashbox_list')
        
        # التحقق من الرصيد الافتتاحي
        try:
            new_initial_balance = Decimal(initial_balance or 0)
        except (ValueError, TypeError):
            messages.error(request, _('Initial balance must be a valid number'))
            return redirect('cashboxes:cashbox_list')
        
        # التحقق من responsible_user_id إذا تم تمريره
        responsible_user = None
        if responsible_user_id:
            try:
                responsible_user = User.objects.get(id=responsible_user_id)
            except User.DoesNotExist:
                messages.error(request, _('Selected responsible user does not exist'))
                return redirect('cashboxes:cashbox_list')
        
        try:
            with transaction.atomic():
                # حفظ الرصيد القديم
                old_balance = cashbox.balance
                
                # حساب فرق الرصيد
                balance_diff = new_initial_balance - old_balance
                
                # تحديث بيانات الصندوق (بدون الرصيد)
                cashbox.name = name
                cashbox.description = description
                cashbox.currency = currency
                cashbox.location = location
                cashbox.responsible_user = responsible_user
                cashbox.save()
                
                # إذا كان هناك تغيير في الرصيد
                if balance_diff != 0:
                    # الحصول على نوع التعديل من الطلب (إجباري عند تغيير الرصيد - IFRS)
                    from core.utils import get_adjustment_account_code
                    adjustment_type = request.POST.get('adjustment_type', '').strip()
                    
                    # التحقق من أن نوع التعديل تم اختياره (IFRS Compliance)
                    if not adjustment_type:
                        messages.error(request, _('Adjustment type must be selected when changing balance to ensure IFRS compliance'))
                        return redirect('cashboxes:cashbox_list')
                    
                    # إنشاء حركة للتعديل
                    # حسب IFRS: يجب تسجيل التغيير الفعلي في الرصيد
                    adjustment_description = _('Adjustment of Opening Balance')
                    if balance_diff > 0:
                        adjustment_description += f' - {_("Increase")}: {abs(balance_diff)} - Type: {dict(CashboxTransaction.ADJUSTMENT_TYPES).get(adjustment_type, "Undefined")}'
                        transaction_amount = abs(balance_diff)  # موجب للإيداع
                    else:
                        adjustment_description += f' - {_("Decrease")}: {abs(balance_diff)} - Type: {dict(CashboxTransaction.ADJUSTMENT_TYPES).get(adjustment_type, "Undefined")}'
                        transaction_amount = -abs(balance_diff)  # سالب للسحب
                    
                    CashboxTransaction.objects.create(
                        cashbox=cashbox,
                        transaction_type='deposit' if balance_diff > 0 else 'withdrawal',
                        date=timezone.now().date(),
                        amount=transaction_amount,
                        description=adjustment_description,
                        adjustment_type=adjustment_type,
                        is_manual_adjustment=True,
                        created_by=request.user
                    )
                    
                    # إنشاء قيد محاسبي للتعديل باستخدام الحساب الصحيح
                    from journal.services import JournalService
                    from journal.models import Account
                    cashbox_account = JournalService.get_cashbox_account(cashbox)
                    
                    # تحديد الحساب المقابل حسب نوع التعديل (IFRS compliant)
                    adjustment_account_code = get_adjustment_account_code(adjustment_type, is_bank=False)
                    adjustment_account = Account.objects.filter(code=adjustment_account_code).first()
                    
                    if cashbox_account and adjustment_account:
                        lines_data = []
                        
                        if balance_diff > 0:
                            # زيادة في الرصيد: مدين الصندوق، دائن الحساب المقابل
                            lines_data = [
                                {
                                    'account_id': cashbox_account.id,
                                    'debit': abs(balance_diff),
                                    'credit': Decimal('0'),
                                    'description': f'{_("Increase in balance")}: {cashbox.name} ({dict(CashboxTransaction.ADJUSTMENT_TYPES).get(adjustment_type, "Adjustment")})'
                                },
                                {
                                    'account_id': adjustment_account.id,
                                    'debit': Decimal('0'),
                                    'credit': abs(balance_diff),
                                    'description': f'{adjustment_account.name} - {dict(CashboxTransaction.ADJUSTMENT_TYPES).get(adjustment_type, "Adjustment")}'
                                }
                            ]
                        else:
                            # نقصان في الرصيد: دائن الصندوق، مدين الحساب المقابل
                            lines_data = [
                                {
                                    'account_id': adjustment_account.id,
                                    'debit': abs(balance_diff),
                                    'credit': Decimal('0'),
                                    'description': f'{adjustment_account.name} - {dict(CashboxTransaction.ADJUSTMENT_TYPES).get(adjustment_type, "Adjustment")}'
                                },
                                {
                                    'account_id': cashbox_account.id,
                                    'debit': Decimal('0'),
                                    'credit': abs(balance_diff),
                                    'description': f'{_("Decrease in balance")}: {cashbox.name} ({dict(CashboxTransaction.ADJUSTMENT_TYPES).get(adjustment_type, "Adjustment")})'
                                }
                            ]
                        
                        journal_entry = JournalService.create_journal_entry(
                            entry_date=timezone.now().date(),
                            description=f'{_("Adjustment of Opening Balance")}: {cashbox.name} - {dict(CashboxTransaction.ADJUSTMENT_TYPES).get(adjustment_type, "Adjustment")}',
                            reference_type='cashbox_adjustment',
                            reference_id=cashbox.id,
                            lines_data=lines_data,
                            user=request.user
                        )
                    
                    # إعادة حساب الرصيد من المعاملات
                    cashbox.sync_balance()
                
                # تسجيل النشاط في سجل الأنشطة
                AuditLog.objects.create(
                    user=request.user,
                    action_type='update',
                    content_type='Cashbox',
                    object_id=cashbox_id,
                    description=f'تعديل الصندوق: {name}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
                
                messages.success(request, _('Cashbox updated successfully'))
                return redirect('cashboxes:cashbox_list')
        except Exception as e:
            print(f"خطأ في تحديث الصندوق: {e}")
            messages.error(request, _('An error occurred while updating the cashbox'))
            return redirect('cashboxes:cashbox_list')
    
    return redirect('cashboxes:cashbox_list')


@login_required
def transfer_create(request):
    """إنشاء تحويل"""
    # التحقق من الصلاحيات - يجب أن لا يكون لديه فقط صلاحية المشاهدة
    if request.user.has_perm('users.can_access_cashboxes') and not (request.user.is_superuser or request.user.user_type in ['superadmin', 'admin']):
        messages.error(request, _('You only have view permission. You cannot create transfers.'))
        return redirect('cashboxes:transfer_list')
    
    if not request.user.has_cashboxes_permission():
        messages.error(request, _('You do not have permission to create transfers'))
        return redirect('core:dashboard')
    if request.method == 'POST':
        transfer_type = request.POST.get('transfer_type')
        
        # استخدام تسلسل التحويل بين البنوك والصناديق لجميع التحويلات
        sequence_type = 'bank_cash_transfer'
        
        # التحقق من وجود تسلسل المستندات
        from core.models import DocumentSequence
        try:
            sequence = DocumentSequence.objects.get(document_type=sequence_type)
        except DocumentSequence.DoesNotExist:
            messages.error(request, _(f'{sequence_type.replace("_", " ").title()} sequence must be configured in settings before creating any transfer.'))
            return redirect('cashboxes:transfer_list')
        transfer_type = request.POST.get('transfer_type')
        date_raw = request.POST.get('date', '').strip() if request.POST.get('date') else ''
        date = date_raw if date_raw else None  # تحويل السلسلة الفارغة إلى None
        amount = request.POST.get('amount')
        fees = request.POST.get('fees', 0)
        exchange_rate = request.POST.get('exchange_rate', 1)
        description = request.POST.get('description', '')
        transfer_number = request.POST.get('transfer_number', '').strip()
        
        # معلومات الإيداع (للتحويل من صندوق إلى بنك)
        deposit_document_number = request.POST.get('deposit_document_number', '').strip()
        deposit_type = request.POST.get('deposit_type', '').strip()
        check_number = request.POST.get('check_number', '').strip()
        check_date_raw = request.POST.get('check_date', '').strip() if request.POST.get('check_date') else ''
        check_date = check_date_raw if check_date_raw else None  # تحويل السلسلة الفارغة إلى None
        check_bank_name = request.POST.get('check_bank_name', '').strip()
        
        # معرفات المصدر والهدف
        from_cashbox_id = request.POST.get('from_cashbox')
        to_cashbox_id = request.POST.get('to_cashbox')
        from_bank_id = request.POST.get('from_bank')
        to_bank_id = request.POST.get('to_bank')
        
        if not all([transfer_type, date, amount]):
            messages.error(request, _('All fields are required'))
            return redirect('cashboxes:transfer_list')
        
        # التحقق من معلومات الإيداع للتحويل من صندوق إلى بنك (اختياري)
        if transfer_type == 'cashbox_to_bank':
            if not from_cashbox_id:
                messages.error(request, _('Sender cashbox is required'))
                return redirect('cashboxes:transfer_list')
            if not to_bank_id:
                messages.error(request, _('Receiver bank is required'))
                return redirect('cashboxes:transfer_list')
            if not deposit_type:
                messages.error(request, _('Deposit type is required'))
                return redirect('cashboxes:transfer_list')
            if deposit_type == 'check':
                if not check_number:
                    messages.error(request, _('Check number is required when deposit type is "Check"'))
                    return redirect('cashboxes:transfer_list')
                if not check_date:
                    messages.error(request, _('Check date is required'))
                    return redirect('cashboxes:transfer_list')
                if not check_bank_name:
                    messages.error(request, _('Check bank name is required'))
                    return redirect('cashboxes:transfer_list')
        
        # التحقق من معلومات التحويل من البنك إلى الصندوق
        elif transfer_type == 'bank_to_cashbox':
            if not from_bank_id:
                messages.error(request, _('Sender bank is required'))
                return redirect('cashboxes:transfer_list')
            if not to_cashbox_id:
                messages.error(request, _('Receiver cashbox is required'))
                return redirect('cashboxes:transfer_list')
        
        # التحقق من رقم التحويل إذا تم إدخاله
        if transfer_number:
            # التحقق من عدم تكرار الرقم
            if CashboxTransfer.objects.filter(transfer_number=transfer_number).exists():
                messages.error(request, _('Transfer number already exists! Please use another number.'))
                return redirect('cashboxes:transfer_list')
        
        # حماية من الطلبات المكررة - التحقق من وجود طلب مشابه في آخر 10 ثوان
        from datetime import datetime, timedelta
        current_time = datetime.now()
        recent_time = current_time - timedelta(seconds=10)
        
        try:
            amount_decimal = Decimal(str(amount))
            recent_transfer = CashboxTransfer.objects.filter(
                transfer_type=transfer_type,
                amount=amount_decimal,
                created_by=request.user,
                created_at__gte=recent_time
            ).first()
            
            if recent_transfer:
                messages.warning(request, _('A similar transfer was created recently! Please check the transfer list.'))
                return redirect('cashboxes:transfer_list')
        except (ValueError, TypeError):
            pass  # تجاهل أخطاء التحويل إذا حدثت
        
        try:
            amount = Decimal(amount)
            fees = Decimal(fees or 0)
            exchange_rate = Decimal(exchange_rate or 1)
            
            if amount <= 0:
                messages.error(request, _('Amount must be greater than zero'))
                return redirect('cashboxes:transfer_list')
            
            with transaction.atomic():
                # تحديد رقم التحويل
                if transfer_number:
                    # استخدام الرقم المُدخل
                    final_transfer_number = transfer_number
                else:
                    # معاينة الرقم التالي دون حجزه
                    try:
                        sequence = DocumentSequence.objects.get(document_type=sequence_type)
                        final_transfer_number = sequence.peek_next_number()
                    except DocumentSequence.DoesNotExist:
                        # استخدام الطريقة القديمة
                        from django.utils import timezone
                        if sequence_type == 'bank_cash_transfer':
                            prefix = 'BCT'
                        else:
                            prefix = 'CT'
                        date_str = timezone.now().strftime('%Y%m%d')
                        
                        # البحث عن آخر رقم مُستخدم فعلياً
                        last_transfer = CashboxTransfer.objects.filter(
                            transfer_number__startswith=f'{prefix}{date_str}'
                        ).order_by('-transfer_number').first()
                        
                        if last_transfer:
                            last_number = int(last_transfer.transfer_number[-4:])
                            new_number = last_number + 1
                        else:
                            new_number = 1
                        
                        final_transfer_number = f'{prefix}{date_str}{new_number:04d}'
                
                # إنشاء التحويل
                transfer = CashboxTransfer.objects.create(
                    transfer_number=final_transfer_number,
                    transfer_type=transfer_type,
                    date=date,
                    amount=amount,
                    fees=fees,
                    exchange_rate=exchange_rate,
                    description=description,
                    from_cashbox_id=from_cashbox_id if from_cashbox_id else None,
                    to_cashbox_id=to_cashbox_id if to_cashbox_id else None,
                    from_bank_id=from_bank_id if from_bank_id else None,
                    to_bank_id=to_bank_id if to_bank_id else None,
                    # معلومات الإيداع
                    deposit_document_number=deposit_document_number if transfer_type == 'cashbox_to_bank' else '',
                    deposit_type=deposit_type if transfer_type == 'cashbox_to_bank' else '',
                    check_number=check_number if deposit_type == 'check' else '',
                    check_date=check_date if deposit_type == 'check' and check_date else None,
                    check_bank_name=check_bank_name if deposit_type == 'check' else '',
                    created_by=request.user
                )
                
                # معالجة الحركات حسب نوع التحويل
                if transfer_type == 'cashbox_to_cashbox':
                    from_cashbox = get_object_or_404(Cashbox, id=from_cashbox_id)
                    to_cashbox = get_object_or_404(Cashbox, id=to_cashbox_id)
                    
                    # التحقق من الرصيد
                    if from_cashbox.balance < amount:
                        messages.error(request, _('Insufficient balance in the sender cashbox'))
                        return redirect('cashboxes:transfer_list')
                    
                    # المعاملات سيتم إنشاؤها تلقائياً من خلال signal create_cashbox_transfer_transactions
                
                elif transfer_type == 'cashbox_to_bank':
                    from_cashbox = get_object_or_404(Cashbox, id=from_cashbox_id)
                    to_bank = get_object_or_404(BankAccount, id=to_bank_id)
                    
                    # التحقق من الرصيد
                    if from_cashbox.balance < amount:
                        messages.error(request, _('Insufficient balance in the cashbox'))
                        return redirect('cashboxes:transfer_list')
                    
                    # المعاملات سيتم إنشاؤها تلقائياً من خلال signal create_cashbox_transfer_transactions
                
                elif transfer_type == 'bank_to_cashbox':
                    from_bank = get_object_or_404(BankAccount, id=from_bank_id)
                    to_cashbox = get_object_or_404(Cashbox, id=to_cashbox_id)
                    
                    # التحقق من الرصيد
                    if from_bank.balance < amount:
                        messages.error(request, _('Insufficient balance in the bank'))
                        return redirect('cashboxes:transfer_list')
                    
                    # المعاملات سيتم إنشاؤها تلقائياً من خلال signal create_cashbox_transfer_transactions
                
                messages.success(request, _('Transfer created successfully'))
                
                # تسجيل النشاط في سجل التدقيق
                from core.models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    action_type='create',
                    content_type='CashboxTransfer',
                    object_id=transfer.id,
                    description=f'إنشاء تحويل {transfer.transfer_number} من {transfer.get_from_display_name()} إلى {transfer.get_to_display_name()} - المبلغ: {transfer.amount}'
                )
                
                # تحديث التسلسل بعد نجاح التحويل
                if not transfer_number:  # فقط إذا تم توليد الرقم تلقائياً
                    try:
                        sequence = DocumentSequence.objects.get(document_type=sequence_type)
                        # استخراج الرقم من نهاية الرقم المُستخدم
                        used_number = int(final_transfer_number[len(sequence.prefix):])
                        sequence.advance_to_at_least(used_number)
                    except (DocumentSequence.DoesNotExist, ValueError, IndexError):
                        pass  # تجاهل الأخطاء في تحديث التسلسل
                
                return redirect('cashboxes:transfer_detail', transfer_id=transfer.id)
        
        except Exception as e:
            messages.error(request, _('An error occurred while creating the transfer'))
    
    return redirect('cashboxes:transfer_list')


@login_required
def transfer_list(request):
    """قائمة التحويلات"""
    # التحقق من الصلاحيات
    if not request.user.has_cashboxes_permission():
        messages.error(request, _('You do not have permission to access transfers'))
        return redirect('core:dashboard')
    
    transfers = CashboxTransfer.objects.all().select_related(
        'from_cashbox', 'to_cashbox', 'from_bank', 'to_bank', 'created_by'
    ).order_by('-date', '-created_at')
    
    # فلترة حسب النوع
    transfer_type = request.GET.get('type')
    if transfer_type:
        transfers = transfers.filter(transfer_type=transfer_type)
    
    # فلترة حسب التاريخ
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        transfers = transfers.filter(date__gte=date_from)
    if date_to:
        transfers = transfers.filter(date__lte=date_to)
    
    # التقسيم لصفحات
    paginator = Paginator(transfers, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # البيانات المساعدة
    cashboxes = Cashbox.objects.filter(is_active=True).order_by('name')
    banks = BankAccount.objects.filter(is_active=True).order_by('name')
    
    # الحصول على رقم التحويل التالي
    from core.models import DocumentSequence
    next_transfer_number = None
    sequence_exists = False
    try:
        sequence = DocumentSequence.objects.get(document_type='bank_cash_transfer')
        next_transfer_number = sequence.peek_next_number()
        sequence_exists = True
    except DocumentSequence.DoesNotExist:
        # لا يوجد تسلسل معد
        next_transfer_number = None
        sequence_exists = False
    
    context = {
        'transfers': page_obj,
        'cashboxes': cashboxes,
        'banks': banks,
        'next_transfer_number': next_transfer_number,
        'sequence_exists': sequence_exists,
    # Use English msgid to avoid Arabic leaking in EN UI; AR translation handled in locale files
    'page_title': _('Cashbox Transfers'),
    }
    return render(request, 'cashboxes/transfer_list.html', context)


@login_required
def transfer_detail(request, transfer_id):
    """تفاصيل التحويل"""
    # التحقق من الصلاحيات
    if not request.user.has_cashboxes_permission():
        messages.error(request, _('You do not have permission to access transfers'))
        return redirect('core:dashboard')
    
    transfer = get_object_or_404(CashboxTransfer, id=transfer_id)
    
    # الحصول على الحركات المرتبطة
    related_transactions = CashboxTransaction.objects.filter(
        related_transfer=transfer
    ).select_related('cashbox')
    
    context = {
        'transfer': transfer,
        'related_transactions': related_transactions,
        'page_title': f'{_("Transfer Details")} - {transfer.transfer_number}',
    }
    return render(request, 'cashboxes/transfer_detail.html', context)


@login_required
@require_http_methods(["GET"])
def get_cashbox_balance(request, cashbox_id):
    """الحصول على رصيد الصندوق (Ajax)"""
    try:
        cashbox = get_object_or_404(Cashbox, id=cashbox_id)
        return JsonResponse({
            'balance': float(cashbox.balance),
            'currency': cashbox.currency,
            'currency_symbol': cashbox.get_currency_symbol()
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def get_bank_balance(request, bank_id):
    """الحصول على رصيد البنك (Ajax)"""
    try:
        bank = get_object_or_404(BankAccount, id=bank_id)
        return JsonResponse({
            'balance': float(bank.balance),
            'currency': bank.currency if hasattr(bank, 'currency') else 'SAR'
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@method_decorator(login_required, name='dispatch')
class CashboxTransactionDeleteView(View):
    """حذف معاملة صندوق مع تحديث جميع الحسابات المرتبطة"""
    
    def post(self, request, transaction_id):
        cashbox_transaction = get_object_or_404(CashboxTransaction, id=transaction_id)
        cashbox = cashbox_transaction.cashbox
        
        try:
            with transaction.atomic():
                # إذا كانت المعاملة مرتبطة بتحويل، نحتاج لتحديث جميع الحسابات المرتبطة
                related_transfer = cashbox_transaction.related_transfer
                if related_transfer:
                    # جمع جميع الحسابات التي تحتاج تحديث
                    accounts_to_sync = []
                    cashboxes_to_sync = []
                    
                    # حسب نوع التحويل
                    if related_transfer.transfer_type == 'cashbox_to_cashbox':
                        if related_transfer.from_cashbox:
                            cashboxes_to_sync.append(related_transfer.from_cashbox)
                        if related_transfer.to_cashbox:
                            cashboxes_to_sync.append(related_transfer.to_cashbox)
                    
                    elif related_transfer.transfer_type == 'cashbox_to_bank':
                        if related_transfer.from_cashbox:
                            cashboxes_to_sync.append(related_transfer.from_cashbox)
                        if related_transfer.to_bank:
                            accounts_to_sync.append(related_transfer.to_bank)
                    
                    elif related_transfer.transfer_type == 'bank_to_cashbox':
                        if related_transfer.from_bank:
                            accounts_to_sync.append(related_transfer.from_bank)
                        if related_transfer.to_cashbox:
                            cashboxes_to_sync.append(related_transfer.to_cashbox)
                    
                    # حذف المعاملة
                    cashbox_transaction.delete()
                    
                    # تحديث أرصدة البنوك
                    for account in accounts_to_sync:
                        account.sync_balance()
                    
                    # تحديث أرصدة الصناديق (save() يؤدي لإعادة حساب الرصيد)
                    for cashbox_item in cashboxes_to_sync:
                        cashbox_item.sync_balance()
                
                else:
                    # معاملة عادية غير مرتبطة بتحويل
                    cashbox_transaction.delete()
                    
                    # إعادة حساب رصيد الصندوق
                    cashbox.sync_balance()
                
                messages.success(request, _('Transaction deleted successfully and all related balances have been updated.'))
                
        except Exception as e:
            messages.error(request, _('An error occurred while deleting the transaction: {}').format(str(e)))
        
        return redirect('cashboxes:cashbox_detail', cashbox_id=cashbox.id)


@method_decorator(login_required, name='dispatch')
class CashboxTransferDeleteView(View):
    """حذف تحويل صندوق مع تحديث جميع الحسابات المرتبطة"""
    
    def post(self, request, transfer_id):
        # التحقق من الصلاحيات - يجب أن لا يكون لديه فقط صلاحية المشاهدة
        if request.user.has_perm('users.can_access_cashboxes') and not (request.user.is_superuser or request.user.user_type in ['superadmin', 'admin']):
            messages.error(request, _('You only have view permission. You cannot delete transfers.'))
            return redirect('cashboxes:transfer_list')
        
        if not (request.user.is_superuser or request.user.user_type in ['superadmin', 'admin']):
            messages.error(request, _('Only superadmin can delete transfers'))
            return redirect('cashboxes:transfer_list')
        
        cashbox_transfer = get_object_or_404(CashboxTransfer, id=transfer_id)
        
        try:
            with transaction.atomic():
                # جمع جميع الحسابات والصناديق التي تحتاج تحديث
                accounts_to_sync = []
                cashboxes_to_sync = []
                
                # حسب نوع التحويل
                if cashbox_transfer.transfer_type == 'cashbox_to_cashbox':
                    if cashbox_transfer.from_cashbox:
                        cashboxes_to_sync.append(cashbox_transfer.from_cashbox)
                    if cashbox_transfer.to_cashbox:
                        cashboxes_to_sync.append(cashbox_transfer.to_cashbox)
                
                elif cashbox_transfer.transfer_type == 'cashbox_to_bank':
                    if cashbox_transfer.from_cashbox:
                        cashboxes_to_sync.append(cashbox_transfer.from_cashbox)
                    if cashbox_transfer.to_bank:
                        accounts_to_sync.append(cashbox_transfer.to_bank)
                
                elif cashbox_transfer.transfer_type == 'bank_to_cashbox':
                    if cashbox_transfer.from_bank:
                        accounts_to_sync.append(cashbox_transfer.from_bank)
                    if cashbox_transfer.to_cashbox:
                        cashboxes_to_sync.append(cashbox_transfer.to_cashbox)
                
                # حذف المعاملات المرتبطة بالتحويل
                CashboxTransaction.objects.filter(related_transfer=cashbox_transfer).delete()
                
                # حذف معاملات البنك المرتبطة (إذا وجدت)
                from banks.models import BankTransaction
                transfer_number = cashbox_transfer.transfer_number
                BankTransaction.objects.filter(
                    Q(reference_number__icontains=transfer_number) |
                    Q(reference_number__icontains=transfer_number.split('-')[0] if '-' in transfer_number else transfer_number)
                ).delete()
                
                # حذف التحويل نفسه
                cashbox_transfer.delete()
                
                # تحديث أرصدة البنوك
                for account in accounts_to_sync:
                    account.sync_balance()
                
                # تحديث أرصدة الصناديق
                for cashbox_item in cashboxes_to_sync:
                    cashbox_item.sync_balance()
                
                messages.success(request, _('Transfer deleted successfully and all related balances have been updated.'))
                
        except Exception as e:
            messages.error(request, _('An error occurred while deleting the transfer: {}').format(str(e)))
        
        return redirect('cashboxes:transfer_list')


@method_decorator(login_required, name='dispatch')
class ClearCashboxTransactionsView(View):
    """حذف جميع معاملات الصندوق مع تحديث الأرصدة"""
    
    def post(self, request, cashbox_id):
        # التحقق من الصلاحيات - يجب أن لا يكون لديه فقط صلاحية المشاهدة
        if request.user.has_perm('users.can_access_cashboxes') and not (request.user.is_superuser or request.user.user_type in ['superadmin', 'admin']):
            messages.error(request, _('You only have view permission. You cannot clear cashbox transactions.'))
            return redirect('cashboxes:cashbox_detail', cashbox_id=cashbox_id)
        
        if not (request.user.is_superuser or request.user.is_staff):
            messages.error(request, _('You do not have permission to clear cashbox transactions'))
            return redirect('cashboxes:cashbox_detail', cashbox_id=cashbox_id)
        
        cashbox = get_object_or_404(Cashbox, id=cashbox_id)
        
        try:
            with transaction.atomic():
                # جمع جميع الحسابات والصناديق المرتبطة قبل الحذف
                accounts_to_sync = set()
                cashboxes_to_sync = set()
                
                # البحث عن التحويلات المرتبطة
                transactions = CashboxTransaction.objects.filter(cashbox=cashbox)
                for trans in transactions:
                    if trans.related_transfer:
                        transfer = trans.related_transfer
                        
                        if transfer.transfer_type == 'cashbox_to_cashbox':
                            if transfer.from_cashbox and transfer.from_cashbox != cashbox:
                                cashboxes_to_sync.add(transfer.from_cashbox)
                            if transfer.to_cashbox and transfer.to_cashbox != cashbox:
                                cashboxes_to_sync.add(transfer.to_cashbox)
                        
                        elif transfer.transfer_type == 'cashbox_to_bank':
                            if transfer.to_bank:
                                accounts_to_sync.add(transfer.to_bank)
                        
                        elif transfer.transfer_type == 'bank_to_cashbox':
                            if transfer.from_bank:
                                accounts_to_sync.add(transfer.from_bank)
                
                # حذف جميع المعاملات
                transactions.delete()
                
                # إعادة تعيين رصيد الصندوق إلى صفر
                cashbox.balance = Decimal('0')
                cashbox.save()
                
                # تحديث أرصدة البنوك المرتبطة
                for account in accounts_to_sync:
                    account.sync_balance()
                
                # تحديث أرصدة الصناديق المرتبطة
                for cashbox_item in cashboxes_to_sync:
                    cashbox_item.sync_balance()
                
                messages.success(request, _('All transactions for cashbox "{}" were cleared successfully and all related balances were updated.').format(cashbox.name))
                
        except Exception as e:
            messages.error(request, _('An error occurred while clearing cashbox transactions: {}').format(str(e)))
        
        return redirect('cashboxes:cashbox_detail', cashbox_id=cashbox_id)


@login_required
def cashbox_delete(request, cashbox_id):
    """حذف الصندوق - محسّن للتعامل مع الحماية"""
    # التحقق من الصلاحيات - يجب أن لا يكون لديه فقط صلاحية المشاهدة
    if request.user.has_perm('users.can_access_cashboxes') and not (request.user.is_superuser or request.user.user_type in ['superadmin', 'admin']):
        messages.error(request, _('You only have view permission. You cannot delete cashboxes.'))
        return redirect('cashboxes:cashbox_list')
    
    if not (request.user.is_superuser or request.user.user_type in ['superadmin', 'admin']):
        messages.error(request, _('You do not have permission to delete cashboxes'))
        return redirect('cashboxes:cashbox_list')
    
    cashbox = get_object_or_404(Cashbox, id=cashbox_id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # إذا كان الرصيد صفر ولكن يوجد معاملات، احذف المعاملات الخاصة بالصندوق فقط
                if cashbox.balance == 0 and CashboxTransaction.objects.filter(cashbox=cashbox).exists():
                    # حذف المعاملات الخاصة بالصندوق المُراد حذفه فقط
                    # هذا لن يؤثر على أرصدة الصناديق الأخرى لأنها لم تتأثر
                    CashboxTransaction.objects.filter(cashbox=cashbox).delete()
                    
                    # ملاحظة: التحويلات ستبقى موجودة ولكن بمرجع NULL للصندوق المحذوف
                    # هذا بسبب استخدام on_delete=models.SET_NULL في النموذج
                    # والصناديق الأخرى ستحتفظ بأرصدتها الصحيحة
                
                elif cashbox.balance != 0:
                    messages.error(request, _('Cannot delete the cashbox because the balance is not zero'))
                    return redirect('cashboxes:cashbox_detail', cashbox_id=cashbox_id)
                
                # تسجيل العملية في سجل الأنشطة قبل الحذف
                AuditLog.objects.create(
                    user=request.user,
                    action_type='delete',
                    content_type='Cashbox',
                    object_id=cashbox_id,
                    description=_('Deleted cashbox: {}').format(cashbox.name),
                    ip_address=request.META.get('REMOTE_ADDR')
                )
                
                # الآن احذف الصندوق - التحويلات ستصبح تلقائياً بمرجع NULL
                cashbox_name = cashbox.name
                cashbox.delete()
                messages.success(request, _('Cashbox "{}" was deleted successfully while preserving transfers and balances in other cashboxes').format(cashbox_name))
                return redirect('cashboxes:cashbox_list')
                
        except ProtectedError as e:
            messages.error(request, _('Cannot delete the cashbox due to protected related transactions'))
        except Exception as e:
            messages.error(request, _('An error occurred while deleting the cashbox: {}').format(str(e)))
    
    return redirect('cashboxes:cashbox_detail', cashbox_id=cashbox_id)


@login_required
def cashbox_export_xlsx(request, cashbox_id):
    """تصدير معاملات الصندوق إلى ملف Excel"""
    # التحقق من الصلاحيات
    if not request.user.has_cashboxes_permission():
        messages.error(request, _('You do not have permission to export cashbox data'))
        return redirect('core:dashboard')
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from django.utils.translation import gettext as _
    from core.signals import log_user_activity
    
    try:
        cashbox = get_object_or_404(Cashbox, id=cashbox_id)
        
        # تسجيل النشاط في سجل الأنشطة
        log_user_activity(
            request,
            'EXPORT',
            cashbox,
            _('تم تصدير معاملات الصندوق إلى Excel: {}').format(cashbox.name)
        )
        
        # إنشاء ملف Excel جديد
        wb = Workbook()
        ws = wb.active
        ws.title = _("Cashbox Statement")
        
        # تنسيق العناوين
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        # إضافة معلومات الصندوق
        ws['A1'] = _("Cashbox Statement")
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = center_alignment
        
        ws['A3'] = _("Cashbox Name") + ":"
        ws['B3'] = cashbox.name
        
        ws['A4'] = _("Description") + ":"
        ws['B4'] = cashbox.description or "-"
        
        ws['A5'] = _("Location") + ":"
        ws['B5'] = cashbox.location or "-"
        
        ws['A6'] = _("Manager") + ":"
        ws['B6'] = cashbox.responsible_user.username if cashbox.responsible_user else "-"
        
        ws['A7'] = _("Current Balance") + ":"
        ws['B7'] = f"{cashbox.balance} {cashbox.currency or 'SAR'}"
        
        ws['A8'] = _("Export Date") + ":"
        ws['B8'] = str(timezone.now().date())
        
        # إضافة عناوين الأعمدة
        headers = [
            _("Date"),
            _("Transaction Type"),
            _("Document Number"),
            _("Description"),
            _("Amount"),
            _("Balance After Transaction"),
            _("Created By")
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # الحصول على جميع المعاملات
        transactions = CashboxTransaction.objects.filter(
            cashbox=cashbox
        ).order_by('date', 'created_at')
        
        # حساب الرصيد المتراكم
        running_balance = Decimal('0')
        for transaction in transactions:
            running_balance += transaction.amount
            transaction.running_balance = running_balance
        
        # عكس الترتيب للعرض (الأحدث أولاً)
        transactions = list(reversed(transactions))
        
        # إضافة بيانات المعاملات
        for row_num, transaction in enumerate(transactions, 11):
            ws.cell(row=row_num, column=1).value = transaction.date.strftime('%Y-%m-%d')
            ws.cell(row=row_num, column=2).value = transaction.get_transaction_type_display()
            ws.cell(row=row_num, column=3).value = get_transaction_document_number(transaction)
            ws.cell(row=row_num, column=4).value = transaction.description
            ws.cell(row=row_num, column=5).value = float(transaction.amount)
            ws.cell(row=row_num, column=6).value = float(transaction.running_balance)
            ws.cell(row=row_num, column=7).value = transaction.created_by.username
        
        # تعديل عرض الأعمدة
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # إنشاء الاستجابة
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"cashbox_statement_{cashbox.name}_{timezone.now().date()}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, _('An error occurred while exporting: {}').format(str(e)))
        return redirect('cashboxes:cashbox_detail', cashbox_id=cashbox_id)
