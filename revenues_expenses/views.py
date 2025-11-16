from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.utils.translation import gettext_lazy as _
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.urls import reverse
from datetime import date, datetime
from decimal import Decimal

from .models import RevenueExpenseCategory, RevenueExpenseEntry, RecurringRevenueExpense
from .forms import RevenueExpenseCategoryForm, RevenueExpenseEntryForm, RecurringRevenueExpenseForm
from settings.models import CompanySettings

@login_required
def category_edit(request, category_id):
    """Edit Revenue/Expense Category"""
    category = get_object_or_404(RevenueExpenseCategory, id=category_id)
    # Access control by permission
    if not request.user.has_perm('revenues_expenses.can_edit_categories'):
        messages.error(request, _('You do not have permission to edit revenue/expense category'))
        return redirect('core:dashboard')
    if request.method == 'POST':
        form = RevenueExpenseCategoryForm(request.POST, instance=category)
        if form.is_valid():
            old_name = category.name
            form.save()
            
            # Audit log
            from core.models import AuditLog
            AuditLog.objects.create(
                user=request.user,
                action_type='update',
                content_type='RevenueExpenseCategory',
                object_id=category.id,
                description=_('Edit category: {old_name} -> {new_name}').format(old_name=old_name, new_name=category.name),
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, _('Category updated successfully'))
            return redirect('revenues_expenses:category_list')
        else:
            messages.error(request, _('Please correct the errors in the form'))
    else:
        form = RevenueExpenseCategoryForm(instance=category)
    context = {
        'form': form,
        'category': category,
        'page_title': _('Edit Revenue/Expense Category'),
    }
    return render(request, 'revenues_expenses/category_edit.html', context)

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.utils.translation import gettext_lazy as _
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.urls import reverse
from datetime import date, datetime
from decimal import Decimal

from .models import RevenueExpenseCategory, RevenueExpenseEntry, RecurringRevenueExpense
from .forms import RevenueExpenseCategoryForm, RevenueExpenseEntryForm, RecurringRevenueExpenseForm
from settings.models import CompanySettings


def create_journal_entry_for_revenue_expense(entry, user):
    """Create journal entry for a revenue/expense entry"""
    from journal.models import JournalEntry, JournalLine, Account
    
    # Find appropriate accounting accounts
    if entry.type == 'revenue':
        # For revenues: cash debit, revenue credit
        if entry.payment_method == 'cash':
            debit_account = Account.objects.filter(code='1101').first()  # Cashbox
        elif entry.payment_method == 'bank':
            debit_account = Account.objects.filter(code='1102').first()  # Bank
        else:
            debit_account = Account.objects.filter(code='1101').first()  # Default to cashbox
            
        # Determine revenue account by category
        if 'مبيعات' in entry.category.name:
            credit_account = Account.objects.filter(code='4001').first()  # Sales revenue
        elif 'خدمات' in entry.category.name:
            credit_account = Account.objects.filter(code='4002').first()  # Services revenue
        else:
            credit_account = Account.objects.filter(code='4999').first()  # Other revenues
        
    else:  # expense
        # For expenses: expense debit, cash credit
        # Determine expense account by category
        if 'إدارية' in entry.category.name:
            debit_account = Account.objects.filter(code='5001').first()  # Administrative expenses
        elif 'تشغيل' in entry.category.name:
            debit_account = Account.objects.filter(code='5002').first()  # Operating expenses
        elif 'رواتب' in entry.category.name:
            debit_account = Account.objects.filter(code='5003').first()  # Salaries and wages
        elif 'إيجار' in entry.category.name:
            debit_account = Account.objects.filter(code='5004').first()  # Rent
        elif 'مرافق' in entry.category.name:
            debit_account = Account.objects.filter(code='5005').first()  # Utilities
        else:
            debit_account = Account.objects.filter(code='5999').first()  # Other expenses
        
        if entry.payment_method == 'cash':
            credit_account = Account.objects.filter(code='1101').first()  # Cashbox
        elif entry.payment_method == 'bank':
            credit_account = Account.objects.filter(code='1102').first()  # Bank
        else:
            credit_account = Account.objects.filter(code='1101').first()  # Default to cashbox
    
    if not debit_account or not credit_account:
        raise Exception(_('Appropriate accounting accounts not found'))
    
    # Create the journal entry
    journal_entry = JournalEntry.objects.create(
        entry_date=entry.date,
        description=f'{entry.get_type_display()} - {entry.category.name} - {entry.description}',
        reference_type='revenue_expense',
        reference_id=entry.id,
        total_amount=entry.amount,
        created_by=user
    )
    
    # Create journal lines
    JournalLine.objects.create(
        journal_entry=journal_entry,
        account=debit_account,
        debit=entry.amount,
        credit=Decimal('0'),
        line_description=f'{entry.get_type_display()} - {entry.category.name}'
    )
    
    JournalLine.objects.create(
        journal_entry=journal_entry,
        account=credit_account,
        debit=Decimal('0'),
        credit=entry.amount,
        line_description=f'{entry.get_type_display()} - {entry.category.name}'
    )
    
    return journal_entry


@login_required
def revenue_expense_dashboard(request):
    """{% trans "Revenues and Expenses Dashboard" %}"""
    # Get base currency
    company_settings = CompanySettings.objects.first()
    base_currency = company_settings.base_currency if company_settings else None
    
    # Statistics from the main accounting system (IFRS Compliant)
    current_month = date.today().month
    current_year = date.today().year
    
    from journal.models import JournalEntry, JournalLine, Account
    
    # Compute revenues from journal lines
    monthly_revenues = JournalLine.objects.filter(
        account__account_type='revenue',
        journal_entry__entry_date__month=current_month,
        journal_entry__entry_date__year=current_year,
        credit__gt=0
    ).aggregate(total=Sum('credit'))['total'] or 0
    
    monthly_expenses = JournalLine.objects.filter(
        account__account_type='expense',
        journal_entry__entry_date__month=current_month,
        journal_entry__entry_date__year=current_year,
        debit__gt=0
    ).aggregate(total=Sum('debit'))['total'] or 0
    
    # Yearly statistics
    yearly_revenues = JournalLine.objects.filter(
        account__account_type='revenue',
        journal_entry__entry_date__year=current_year,
        credit__gt=0
    ).aggregate(total=Sum('credit'))['total'] or 0
    
    yearly_expenses = JournalLine.objects.filter(
        account__account_type='expense',
        journal_entry__entry_date__year=current_year,
        debit__gt=0
    ).aggregate(total=Sum('debit'))['total'] or 0
    
    # Total revenues and expenses from all time
    total_revenues = JournalLine.objects.filter(
        account__account_type='revenue',
        credit__gt=0
    ).aggregate(total=Sum('credit'))['total'] or 0
    
    total_expenses = JournalLine.objects.filter(
        account__account_type='expense',
        debit__gt=0
    ).aggregate(total=Sum('debit'))['total'] or 0
    
    # Count of journal entries
    entries_count = JournalEntry.objects.count()
    
    # Recent journal entries
    recent_entries = JournalEntry.objects.select_related('created_by').order_by('-entry_date', '-id')[:10]
    
    # Pending recurring items (from revenues_expenses if present)
    pending_recurring = RecurringRevenueExpense.objects.filter(
        is_active=True,
        next_due_date__lte=date.today()
    ).select_related('category', 'currency').order_by('next_due_date')[:5]
    
    context = {
        'page_title': _('Revenues and Expenses Dashboard'),
        'monthly_revenues': monthly_revenues,
        'monthly_expenses': monthly_expenses,
        'monthly_net': monthly_revenues - monthly_expenses,
        'yearly_revenues': yearly_revenues,
        'yearly_expenses': yearly_expenses,
        'yearly_net': yearly_revenues - yearly_expenses,
        'total_revenues': total_revenues,
        'total_expenses': total_expenses,
        'net_profit': total_revenues - total_expenses,
        'entries_count': entries_count,
        'recent_entries': recent_entries,
        'pending_recurring': pending_recurring,
        'base_currency': base_currency,
    }
    
    return render(request, 'revenues_expenses/dashboard.html', context)


@login_required
def category_list(request):
    """{% trans "Revenue/Expense Categories List" %}"""
    # Access control by permission
    if not request.user.has_perm('revenues_expenses.can_view_categories'):
        messages.error(request, _('You do not have permission to view revenue/expense category'))
        return redirect('core:dashboard')
    categories = RevenueExpenseCategory.objects.filter(is_active=True).select_related('created_by')
    # Filtering
    category_type = request.GET.get('type')
    if category_type:
        categories = categories.filter(type=category_type)
    search = request.GET.get('search')
    if search:
        categories = categories.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )
    # Pagination
    paginator = Paginator(categories, 25)
    page_number = request.GET.get('page')
    categories = paginator.get_page(page_number)
    context = {
        'page_title': _('Revenue and Expense Categories'),
        'categories': categories,
        'category_types': RevenueExpenseCategory.CATEGORY_TYPES,
    }
    return render(request, 'revenues_expenses/category_list.html', context)


@login_required
def category_create(request):
    """{% trans "Add New Category" %}"""
    # Access control by permission
    if not request.user.has_perm('revenues_expenses.can_add_categories'):
        messages.error(request, _('You do not have permission to add revenue/expense category'))
        return redirect('revenues_expenses:category_list')
    if request.method == 'POST':
        form = RevenueExpenseCategoryForm(request.POST)
        if form.is_valid():
            category = form.save(commit=False)
            category.created_by = request.user
            category.save()
            
            # Audit log
            from core.models import AuditLog
            AuditLog.objects.create(
                user=request.user,
                action_type='create',
                content_type='RevenueExpenseCategory',
                object_id=category.id,
                description=_('Create new category: {name}').format(name=category.name),
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, _('Category created successfully'))
            return redirect('revenues_expenses:category_list')
        else:
            messages.error(request, _('Please correct the errors in the form'))
    else:
        form = RevenueExpenseCategoryForm()
    context = {
        'form': form,
        'page_title': _('Add New Category'),
    }
    return render(request, 'revenues_expenses/category_create.html', context)


@login_required
def category_delete(request, category_id):
    """{% trans "Delete Revenue/Expense Category" %}"""
    category = get_object_or_404(RevenueExpenseCategory, id=category_id)
    # Access control by permission
    if not request.user.has_perm('revenues_expenses.can_delete_categories'):
        messages.error(request, _('You do not have permission to delete revenue/expense category'))
        return redirect('core:dashboard')
    if request.method == 'POST':
        category.is_active = False
        category.save()
        messages.success(request, _('Category deleted successfully'))
        return redirect('revenues_expenses:category_list')
    context = {
        'category': category,
        'page_title': _('Confirm Category Deletion'),
    }
    return render(request, 'revenues_expenses/category_confirm_delete.html', context)


@login_required
def entry_create(request):
    """{% trans "Add New Entry" %}"""
    # Access control by permission
    if not request.user.has_perm('revenues_expenses.can_add_entries'):
        messages.error(request, _('You do not have permission to add revenue/expense entry'))
        return redirect('core:dashboard')
    if request.method == 'POST':
        form = RevenueExpenseEntryForm(request.POST)
        if form.is_valid():
            entry = form.save(commit=False)
            entry.created_by = request.user
            entry.save()
            messages.success(request, _('Entry created successfully'))
            return redirect('revenues_expenses:entry_detail', entry_id=entry.id)
        else:
            messages.error(request, _('Please correct the errors in the form'))
    else:
        form = RevenueExpenseEntryForm()
    context = {
        'form': form,
        'page_title': _('Add New Entry'),
    }
    return render(request, 'revenues_expenses/entry_create.html', context)


@login_required
def entry_delete(request, entry_id):
    """Delete Revenue/Expense Entry"""
    entry = get_object_or_404(RevenueExpenseEntry, id=entry_id)
    if not request.user.has_perm('revenues_expenses.can_delete_entries'):
        messages.error(request, _('You do not have permission to delete revenue/expense entry'))
        return redirect('core:dashboard')
    if request.method == 'POST':
        entry.delete()
        messages.success(request, _('Entry deleted successfully'))
        return redirect('revenues_expenses:entry_list')
    context = {
        'entry': entry,
        'page_title': _('Confirm Entry Deletion'),
    }
    return render(request, 'revenues_expenses/entry_confirm_delete.html', context)


@login_required
def entry_edit(request, entry_id):
    """Edit Revenue/Expense Entry"""
    entry = get_object_or_404(RevenueExpenseEntry, id=entry_id)
    if not request.user.has_perm('revenues_expenses.can_edit_entries'):
        messages.error(request, _('You do not have permission to edit revenue/expense entry'))
        return redirect('core:dashboard')
    if request.method == 'POST':
        form = RevenueExpenseEntryForm(request.POST, instance=entry)
        if form.is_valid():
            form.save()
            messages.success(request, _('Entry updated successfully'))
            return redirect('revenues_expenses:entry_detail', entry_id=entry.id)
        else:
            messages.error(request, _('Please correct the errors in the form'))
    else:
        form = RevenueExpenseEntryForm(instance=entry)
    context = {
        'form': form,
        'entry': entry,
        'page_title': _('Edit Revenue/Expense Entry'),
    }
    return render(request, 'revenues_expenses/entry_edit.html', context)


@login_required
def entry_detail(request, entry_id):
    """Entry details"""
    entry = get_object_or_404(RevenueExpenseEntry.objects.select_related('category', 'created_by', 'approved_by', 'currency'), id=entry_id)
    context = {
        'entry': entry,
        'page_title': f'{_("Entry Details")} - {entry.entry_number}',
    }
    return render(request, 'revenues_expenses/entry_detail.html', context)


@login_required
def entry_approve(request, entry_id):
    """Approve Entry"""
    entry = get_object_or_404(RevenueExpenseEntry, id=entry_id)
    if not entry.is_approved:
        entry.is_approved = True
        entry.approved_by = request.user
        entry.approved_at = datetime.now()
        entry.save()
        messages.success(request, _('Entry approved successfully'))
    return redirect('revenues_expenses:entry_detail', entry_id=entry.id)


@login_required
def recurring_list(request):
    """Recurring Revenues and Expenses List"""
    recurring_items = RecurringRevenueExpense.objects.select_related(
        'category', 'created_by', 'currency'
    ).order_by('-created_at')
    
    # Filtering
    category_type = request.GET.get('type')
    if category_type:
        recurring_items = recurring_items.filter(category__type=category_type)
    
    category = request.GET.get('category')
    if category:
        recurring_items = recurring_items.filter(category_id=category)
    
    frequency = request.GET.get('frequency')
    if frequency:
        recurring_items = recurring_items.filter(frequency=frequency)
    
    status = request.GET.get('status')
    if status == 'active':
        recurring_items = recurring_items.filter(is_active=True)
    elif status == 'inactive':
        recurring_items = recurring_items.filter(is_active=False)
    
    search = request.GET.get('search')
    if search:
        recurring_items = recurring_items.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(recurring_items, 25)
    page_number = request.GET.get('page')
    recurring_items = paginator.get_page(page_number)
    
    # Data for filtering
    categories = RevenueExpenseCategory.objects.filter(is_active=True).order_by('name')
    
    # Statistics
    total_recurring = RecurringRevenueExpense.objects.filter(is_active=True).count()
    overdue_count = RecurringRevenueExpense.objects.filter(
        is_active=True,
        next_due_date__lt=date.today()
    ).count()
    due_today_count = RecurringRevenueExpense.objects.filter(
        is_active=True,
        next_due_date=date.today()
    ).count()
    
    context = {
        'page_title': _('Recurring revenues and expenses'),
        'recurring_items': recurring_items,
        'categories': categories,
        'frequency_choices': RecurringRevenueExpense.FREQUENCY_CHOICES,
        'entry_types': RevenueExpenseCategory.CATEGORY_TYPES,
        'total_recurring': total_recurring,
        'overdue_count': overdue_count,
        'due_today_count': due_today_count,
    }
    return render(request, 'revenues_expenses/recurring_list.html', context)


@login_required
def recurring_create(request):
    """Add Recurring Revenue/Expense"""
    if request.method == 'POST':
        form = RecurringRevenueExpenseForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                recurring = form.save(commit=False)
                recurring.created_by = request.user
                
                # Calculate next due date
                from dateutil.relativedelta import relativedelta
                start_date = recurring.start_date
                
                if recurring.frequency == 'daily':
                    recurring.next_due_date = start_date
                elif recurring.frequency == 'weekly':
                    recurring.next_due_date = start_date
                elif recurring.frequency == 'monthly':
                    recurring.next_due_date = start_date
                elif recurring.frequency == 'quarterly':
                    recurring.next_due_date = start_date
                elif recurring.frequency == 'semi_annual':
                    recurring.next_due_date = start_date
                elif recurring.frequency == 'annual':
                    recurring.next_due_date = start_date
                else:
                    recurring.next_due_date = start_date
                
                recurring.save()
                
                # Log activity
                from core.signals import log_activity
                log_activity(
                    action_type='create',
                    obj=recurring,
                    description=_('Add recurring revenue/expense: {name}').format(name=recurring.name),
                    user=request.user
                )
                
                messages.success(request, _('Recurring revenue/expense created successfully'))
                return redirect('revenues_expenses:recurring_list')
        else:
            messages.error(request, _('Please correct the errors in the form'))
    else:
        form = RecurringRevenueExpenseForm()
    
    context = {
        'form': form,
        'page_title': _('Add Recurring Revenue/Expense'),
    }
    return render(request, 'revenues_expenses/recurring_create.html', context)


@login_required
def recurring_detail(request, recurring_id):
    """Recurring Revenue/Expense Details"""
    recurring = get_object_or_404(RecurringRevenueExpense, id=recurring_id)
    
    # Get entries generated from this recurring
    generated_entries = RevenueExpenseEntry.objects.filter(
        description__icontains=f"Generated from recurring #{recurring.id}"
    ).order_by('-date')
    
    context = {
        'page_title': f'{recurring.name} - {_("Recurring Revenue/Expense Details")}',
        'recurring': recurring,
        'generated_entries': generated_entries,
    }
    return render(request, 'revenues_expenses/recurring_detail.html', context)


@login_required
def recurring_edit(request, recurring_id):
    """Edit Recurring Revenue/Expense"""
    recurring = get_object_or_404(RecurringRevenueExpense, id=recurring_id)
    
    if request.method == 'POST':
        form = RecurringRevenueExpenseForm(request.POST, instance=recurring)
        if form.is_valid():
            with transaction.atomic():
                updated_recurring = form.save()
                
                # Log activity
                from core.signals import log_activity
                log_activity(
                    action_type='update',
                    obj=updated_recurring,
                    description=_('Edit recurring revenue/expense: {name}').format(name=updated_recurring.name),
                    user=request.user
                )
                
                messages.success(request, _('Recurring revenue/expense updated successfully'))
                return redirect('revenues_expenses:recurring_detail', recurring_id=recurring.id)
        else:
            messages.error(request, _('Please correct the errors in the form'))
    else:
        form = RecurringRevenueExpenseForm(instance=recurring)
    
    context = {
        'form': form,
        'recurring': recurring,
        'page_title': f'{recurring.name} - {_("Edit Recurring Revenue/Expense")}',
    }
    return render(request, 'revenues_expenses/recurring_edit.html', context)


@login_required
def recurring_delete(request, recurring_id):
    """Delete Recurring Revenue/Expense"""
    recurring = get_object_or_404(RecurringRevenueExpense, id=recurring_id)
    
    if request.method == 'POST':
        with transaction.atomic():
            recurring_name = recurring.name
            
            # Log activity before deletion
            from core.signals import log_activity
            log_activity(
                action_type='delete',
                obj=None,
                description=_('Delete recurring revenue/expense: {name}').format(name=recurring_name),
                user=request.user
            )
            
            recurring.delete()
            
            messages.success(request, _('Recurring revenue/expense "{name}" deleted successfully').format(name=recurring_name))
            return redirect('revenues_expenses:recurring_list')
    
    return redirect('revenues_expenses:recurring_detail', recurring_id=recurring_id)


@login_required
def recurring_toggle_status(request, recurring_id):
    """Toggle Recurring Revenue/Expense Status"""
    recurring = get_object_or_404(RecurringRevenueExpense, id=recurring_id)
    
    if request.method == 'POST':
        with transaction.atomic():
            recurring.is_active = not recurring.is_active
            recurring.save()
            
            if recurring.is_active:
                action_description = _('Activated recurring revenue/expense: {name}').format(name=recurring.name)
                messages.success(request, _('Recurring revenue/expense {name} activated successfully').format(name=recurring.name))
            else:
                action_description = _('Deactivated recurring revenue/expense: {name}').format(name=recurring.name)
                messages.success(request, _('Recurring revenue/expense {name} deactivated successfully').format(name=recurring.name))

            # Audit log
            from core.signals import log_activity
            log_activity(
                action_type='update',
                obj=recurring,
                description=action_description,
                user=request.user
            )
    
    return redirect('revenues_expenses:recurring_detail', recurring_id=recurring_id)


@login_required
def recurring_generate_entry(request, recurring_id):
    """Generate Journal Entry From Recurring Revenue/Expense Manually"""
    recurring = get_object_or_404(RecurringRevenueExpense, id=recurring_id)
    
    if request.method == 'POST':
        with transaction.atomic():
            # Create a new entry from the recurring
            entry = RevenueExpenseEntry.objects.create(
                type=recurring.category.type,
                category=recurring.category,
                amount=recurring.amount,
                currency=recurring.currency,
                description=f"{recurring.description}\n(Generated from recurring #{recurring.id}: {recurring.name})",
                payment_method=recurring.payment_method,
                date=date.today(),
                created_by=request.user
            )
            
            # Update last generated date
            recurring.last_generated = date.today()
            
                # Calculate next due date (recurrence)
            from dateutil.relativedelta import relativedelta
            
            if recurring.frequency == 'daily':
                recurring.next_due_date = recurring.next_due_date + relativedelta(days=1)
            elif recurring.frequency == 'weekly':
                recurring.next_due_date = recurring.next_due_date + relativedelta(weeks=1)
            elif recurring.frequency == 'monthly':
                recurring.next_due_date = recurring.next_due_date + relativedelta(months=1)
            elif recurring.frequency == 'quarterly':
                recurring.next_due_date = recurring.next_due_date + relativedelta(months=3)
            elif recurring.frequency == 'semi_annual':
                recurring.next_due_date = recurring.next_due_date + relativedelta(months=6)
            elif recurring.frequency == 'annual':
                recurring.next_due_date = recurring.next_due_date + relativedelta(years=1)
            
            recurring.save()
            
            # Log activity
            from core.signals import log_activity
            log_activity(
                action_type='create',
                obj=entry,
                description=_('Generated journal entry from recurring: {name} - Entry Number: {num}').format(name=recurring.name, num=entry.entry_number),
                user=request.user
            )
            
            messages.success(request, _('Generated journal entry {num} from recurring revenue/expense successfully').format(num=entry.entry_number))
            return redirect('revenues_expenses:entry_detail', entry_id=entry.id)
    
    return redirect('revenues_expenses:recurring_detail', recurring_id=recurring_id)


@login_required
def entry_detail(request, entry_id):
    """Revenue/Expense Entry Details"""
    entry = get_object_or_404(RevenueExpenseEntry, id=entry_id)
    
    context = {
        'page_title': f'{entry.entry_number} - {_("Entry Details")}',
        'entry': entry,
    }
    return render(request, 'revenues_expenses/entry_detail.html', context)


@login_required
def entry_list(request):
    """Revenue/Expense Entries List"""
    # Permission check
    if not request.user.has_perm('revenues_expenses.can_view_entries'):
        messages.error(request, _('You do not have permission to view revenue and expense entries'))
        return redirect('core:dashboard')
    
    # Fetch entries
    entries = RevenueExpenseEntry.objects.select_related(
        'category', 'created_by', 'approved_by', 'currency'
    ).order_by('-date', '-created_at')
    
    # Filtering
    entry_type = request.GET.get('type')
    if entry_type:
        entries = entries.filter(type=entry_type)
    
    category = request.GET.get('category')
    if category:
        entries = entries.filter(category_id=category)
    
    status = request.GET.get('status')
    if status == 'approved':
        entries = entries.filter(is_approved=True)
    elif status == 'pending':
        entries = entries.filter(is_approved=False)
    
    payment_method = request.GET.get('payment_method')
    if payment_method:
        entries = entries.filter(payment_method=payment_method)
    
    search = request.GET.get('search')
    if search:
        entries = entries.filter(
            Q(entry_number__icontains=search) |
            Q(description__icontains=search) |
            Q(category__name__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(entries, 25)
    page_number = request.GET.get('page')
    entries = paginator.get_page(page_number)
    
    # Data for filtering
    categories = RevenueExpenseCategory.objects.filter(is_active=True).order_by('name')
    
    # Statistics
    total_entries = RevenueExpenseEntry.objects.count()
    approved_count = RevenueExpenseEntry.objects.filter(is_approved=True).count()
    pending_count = RevenueExpenseEntry.objects.filter(is_approved=False).count()
    
    # Totals
    total_revenue = RevenueExpenseEntry.objects.filter(type='revenue').aggregate(
        total=Sum('amount'))['total'] or 0
    total_expense = RevenueExpenseEntry.objects.filter(type='expense').aggregate(
        total=Sum('amount'))['total'] or 0
    
    context = {
        'page_title': _('Revenue and Expense Entries'),
        'entries': entries,
        'categories': categories,
        'entry_types': RevenueExpenseCategory.CATEGORY_TYPES,
        'payment_methods': RevenueExpenseEntry.PAYMENT_METHODS,
        'total_entries': total_entries,
        'approved_count': approved_count,
        'pending_count': pending_count,
        'total_revenue': total_revenue,
        'total_expense': total_expense,
        'net_amount': total_revenue - total_expense,
    }
    return render(request, 'revenues_expenses/entry_list.html', context)


@login_required
def entry_list_export_excel(request):
    """Export Revenue/Expense Entries to Excel"""
    # Permission check
    if not request.user.has_perm('revenues_expenses.can_view_entries'):
        messages.error(request, _('You do not have permission to export revenue and expense entries'))
        return redirect('core:dashboard')
    
    # Fetch entries
    entries = RevenueExpenseEntry.objects.select_related(
        'category', 'created_by', 'approved_by', 'currency'
    ).order_by('-date', '-created_at')
    
    # Filtering
    entry_type = request.GET.get('type')
    if entry_type:
        entries = entries.filter(type=entry_type)
    
    category = request.GET.get('category')
    if category:
        entries = entries.filter(category_id=category)
    
    status = request.GET.get('status')
    if status == 'approved':
        entries = entries.filter(is_approved=True)
    elif status == 'pending':
        entries = entries.filter(is_approved=False)
    
    payment_method = request.GET.get('payment_method')
    if payment_method:
        entries = entries.filter(payment_method=payment_method)
    
    search = request.GET.get('search')
    if search:
        entries = entries.filter(
            Q(entry_number__icontains=search) |
            Q(description__icontains=search) |
            Q(category__name__icontains=search)
        )
    
    # Create Excel file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    
    wb = Workbook()
    ws = wb.active
    ws.title = str(_('Revenue and Expense Entries'))
    
    # Header formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center")
    
    # Add headers
    headers = [
        _('Entry Number'),
        _('Date'),
        _('Type'),
        _('Category'),
        _('Amount'),
        _('Description'),
        _('Approval Status'),
        _('Payment Method'),
        _('Currency'),
        _('Created By'),
        _('Creation Date')
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        ws.column_dimensions[chr(64 + col_num)].width = 15
    
    # Add data
    for row_num, entry in enumerate(entries, 2):
        ws.cell(row=row_num, column=1, value=entry.entry_number or f"#{entry.id}")
        ws.cell(row=row_num, column=2, value=entry.date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=3, value=entry.get_type_display())
        ws.cell(row=row_num, column=4, value=entry.category.name if entry.category else '')
        ws.cell(row=row_num, column=5, value=float(entry.amount))
        ws.cell(row=row_num, column=6, value=entry.description)
        ws.cell(row=row_num, column=7, value=str(_('Approved')) if entry.is_approved else str(_('Pending')))
        ws.cell(row=row_num, column=8, value=entry.get_payment_method_display())
        ws.cell(row=row_num, column=9, value=str(entry.currency) if entry.currency else '')
        ws.cell(row=row_num, column=10, value=str(entry.created_by) if entry.created_by else '')
        ws.cell(row=row_num, column=11, value=entry.created_at.strftime('%Y-%m-%d %H:%M'))
    
    # Build response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"revenues_expenses_entries_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    
    # Log activity
    from core.models import AuditLog
    AuditLog.objects.create(
        user=request.user,
        action_type='export',
        content_type='RevenueExpenseEntry',
        description=_('Export revenue and expense entries to Excel'),
        ip_address=request.META.get('REMOTE_ADDR')
    )
    
    return response


@login_required
def category_list_export_excel(request):
    """Export Revenue/Expense Categories to Excel"""
    # Check permissions
    if not request.user.has_perm('revenues_expenses.can_view_categories'):
        messages.error(request, _('You do not have permission to export revenue and expense categories'))
        return redirect('core:dashboard')
    
    # Fetch categories
    categories = RevenueExpenseCategory.objects.select_related('account', 'created_by').order_by('type', 'name')
    
    # Filtering
    category_type = request.GET.get('type')
    if category_type:
        categories = categories.filter(type=category_type)
    
    status = request.GET.get('status')
    if status == 'active':
        categories = categories.filter(is_active=True)
    elif status == 'inactive':
        categories = categories.filter(is_active=False)
    
    search = request.GET.get('search')
    if search:
        categories = categories.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )
    
    # Create Excel file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    
    wb = Workbook()
    ws = wb.active
    ws.title = str(_('Revenue and Expense Categories'))
    
    # Format headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center")
    
    # Add headers
    headers = [
        _('Category Name'),
        _('Type'),
        _('Accounting Account'),
        _('Description'),
        _('Status'),
        _('Created By'),
        _('Creation Date')
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        ws.column_dimensions[chr(64 + col_num)].width = 20
    
    # Add data
    for row_num, category in enumerate(categories, 2):
        ws.cell(row=row_num, column=1, value=category.name)
        ws.cell(row=row_num, column=2, value=category.get_type_display())
        ws.cell(row=row_num, column=3, value=f"{category.account.code} - {category.account.name}" if category.account else str(_('Not specified')))
        ws.cell(row=row_num, column=4, value=category.description or '')
        ws.cell(row=row_num, column=5, value=str(_('Active')) if category.is_active else str(_('Inactive')))
        ws.cell(row=row_num, column=6, value=str(category.created_by) if category.created_by else '')
        ws.cell(row=row_num, column=7, value=category.created_at.strftime('%Y-%m-%d %H:%M'))
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"revenues_expenses_categories_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    
    # Log activity
    from core.models import AuditLog
    AuditLog.objects.create(
        user=request.user,
        action_type='export',
        content_type='RevenueExpenseCategory',
        description=_('Export revenue and expense categories to Excel'),
        ip_address=request.META.get('REMOTE_ADDR')
    )
    
    return response


@login_required
def dashboard_recent_entries_export_excel(request):
    """Export Recent Entries from Dashboard to Excel"""
    # Check permissions
    if not request.user.has_perm('revenues_expenses.can_view_entries'):
        messages.error(request, _('You do not have permission to export recent entries'))
        return redirect('core:dashboard')
    
    # Get recent entries
    from journal.models import JournalEntry
    recent_entries = JournalEntry.objects.select_related('created_by').order_by('-entry_date', '-id')[:50]  # Increase count for export
    
    # Create Excel file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    
    wb = Workbook()
    ws = wb.active
    ws.title = str(_('Recent Entries'))
    
    # Format headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center")
    
    # Add headers
    headers = [
        _('Entry Number'),
        _('Date'),
        _('Type'),
        _('Reference'),
        _('Total Amount'),
        _('Description'),
        _('Created By'),
        _('Creation Date')
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        ws.column_dimensions[chr(64 + col_num)].width = 20
    
    # Add data
    for row_num, entry in enumerate(recent_entries, 2):
        ws.cell(row=row_num, column=1, value=entry.entry_number or f"#{entry.id}")
        ws.cell(row=row_num, column=2, value=entry.entry_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=3, value=entry.get_entry_type_display())
        ws.cell(row=row_num, column=4, value=entry.reference_type or '')
        ws.cell(row=row_num, column=5, value=float(entry.total_amount))
        ws.cell(row=row_num, column=6, value=entry.description or '')
        ws.cell(row=row_num, column=7, value=str(entry.created_by) if entry.created_by else '')
        ws.cell(row=row_num, column=8, value=entry.created_at.strftime('%Y-%m-%d %H:%M'))
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"dashboard_recent_entries_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    
    # Log activity
    from core.models import AuditLog
    AuditLog.objects.create(
        user=request.user,
        action_type='export',
        content_type='JournalEntry',
        description=_('Export recent entries from dashboard to Excel'),
        ip_address=request.META.get('REMOTE_ADDR')
    )
    
    return response


@login_required
def entry_create(request):
    """Add Revenue/Expense Entry"""
    if request.method == 'POST':
        form = RevenueExpenseEntryForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                entry = form.save(commit=False)
                entry.created_by = request.user
                
                # Process additional options
                auto_approve = request.POST.get('auto_approve') == 'true'
                create_journal = request.POST.get('create_journal') == 'true'
                
                entry.save()
                
                # Auto-approve if selected
                if auto_approve:
                    entry.is_approved = True
                    entry.approved_by = request.user
                    entry.approved_at = datetime.now()
                    entry.save()
                
                # Create journal entry if selected
                if create_journal:
                    try:
                        journal_entry = create_journal_entry_for_revenue_expense(entry, request.user)
                        messages.info(request, _('Journal entry {num} created').format(num=journal_entry.entry_number))
                    except Exception as e:
                        messages.warning(request, _('Entry created successfully but failed to create journal entry: {error}').format(error=str(e)))
                
                # Log activity
                from core.signals import log_activity
                log_activity(
                    action_type='create',
                    obj=entry,
                    description=_('Add new {type} entry: {num} - {desc}').format(type=entry.get_type_display(), num=entry.entry_number, desc=entry.description[:50]),
                    user=request.user
                )
                
                success_message = _('Entry {num} created successfully. Amount: {amount} {currency}').format(num=entry.entry_number, amount=f"{entry.amount:,.3f}", currency=entry.currency.code if entry.currency else "")
                if auto_approve:
                    success_message += ' (' + str(_('Auto-approved')) + ')'
                
                messages.success(request, success_message)
                return redirect('revenues_expenses:entry_detail', entry_id=entry.id)
        else:
            # Add detailed error messages
            for field, errors in form.errors.items():
                for error in errors:
                    field_label = form.fields[field].label if field in form.fields else field
                    messages.error(request, f'{field_label}: {error}')
    else:
        form = RevenueExpenseEntryForm()
        # Set current date as default
        from datetime import date
        form.fields['date'].initial = date.today()
    
    # Fetch statistics for display
    from django.db.models import Sum
    from datetime import date
    today = date.today()
    
    today_revenues = RevenueExpenseEntry.objects.filter(
        type='revenue',
        date=today,
        is_approved=True
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    today_expenses = RevenueExpenseEntry.objects.filter(
        type='expense', 
        date=today,
        is_approved=True
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Fetch active categories for JavaScript
    revenue_categories = RevenueExpenseCategory.objects.filter(
        type='revenue', is_active=True
    ).values('id', 'name').order_by('name')
    
    expense_categories = RevenueExpenseCategory.objects.filter(
        type='expense', is_active=True
    ).values('id', 'name').order_by('name')
    
    context = {
        'form': form,
        'page_title': _('Add Revenue/Expense Entry'),
        'today_revenues': today_revenues,
        'today_expenses': today_expenses,
        'revenue_categories': list(revenue_categories),
        'expense_categories': list(expense_categories),
    }
    return render(request, 'revenues_expenses/entry_create.html', context)


@login_required
def get_categories_by_type(request, entry_type):
    """API to fetch categories by type"""
    categories = RevenueExpenseCategory.objects.filter(
        type=entry_type, 
        is_active=True
    ).values('id', 'name').order_by('name')
    
    return JsonResponse({
        'categories': list(categories)
    })


@login_required
def get_today_stats(request):
    """API to fetch today's statistics"""
    from django.db.models import Sum
    from datetime import date
    today = date.today()
    
    revenues = RevenueExpenseEntry.objects.filter(
        type='revenue',
        date=today,
        is_approved=True
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    expenses = RevenueExpenseEntry.objects.filter(
        type='expense',
        date=today,
        is_approved=True
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    return JsonResponse({
        'revenues': f"{revenues:,.2f}",
        'expenses': f"{expenses:,.2f}",
        'net': f"{revenues - expenses:,.2f}"
    })


@login_required
def entry_approve(request, entry_id):
    """Approve Revenue/Expense Entry"""
    entry = get_object_or_404(RevenueExpenseEntry, id=entry_id)
    
    if request.method == 'POST':
        with transaction.atomic():
            entry.is_approved = True
            entry.approved_by = request.user
            entry.approved_at = datetime.now()
            entry.save()
            
            # Log activity
            from core.signals import log_activity
            log_activity(
                action_type='approve',
                obj=entry,
                description=_('Approve {type} entry: {num}').format(type=entry.get_type_display(), num=entry.entry_number),
                user=request.user
            )
            
            messages.success(request, _('Entry {num} approved successfully').format(num=entry.entry_number))
    
    return redirect('revenues_expenses:entry_detail', entry_id=entry_id)


@login_required
def entry_list(request):
    # Protect access by permission
    if not request.user.has_perm('revenues_expenses.can_view_entries'):
        messages.error(request, _('You do not have permission to view revenue/expense entry'))
        return redirect('core:dashboard')
    """Revenue and Expense Entries List"""
    entries = RevenueExpenseEntry.objects.select_related(
        'category', 'created_by', 'approved_by', 'currency'
    ).order_by('-date', '-created_at')
    
    # Filtering
    entry_type = request.GET.get('type')
    if entry_type:
        entries = entries.filter(type=entry_type)
    
    category = request.GET.get('category')
    if category:
        entries = entries.filter(category_id=category)
    
    status = request.GET.get('status')
    if status == 'approved':
        entries = entries.filter(is_approved=True)
    elif status == 'pending':
        entries = entries.filter(is_approved=False)
    
    search = request.GET.get('search')
    if search:
        entries = entries.filter(
            Q(entry_number__icontains=search) |
            Q(description__icontains=search) |
            Q(reference_number__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(entries, 25)
    page_number = request.GET.get('page')
    entries = paginator.get_page(page_number)
    
    # Categories for filtering
    categories = RevenueExpenseCategory.objects.filter(is_active=True).order_by('name')
    
    context = {
        'page_title': _('Revenue and Expense Entries'),
        'entries': entries,
        'categories': categories,
        'entry_types': RevenueExpenseEntry.ENTRY_TYPES,
    }
    
    return render(request, 'revenues_expenses/entry_list.html', context)


@login_required
def sector_analysis(request):
    """Revenue and Expenses by Sector Analysis"""
    user = request.user
    has_perm = (
        getattr(user, 'is_superuser', False) or
        getattr(user, 'user_type', None) in ['superadmin', 'admin'] or
        user.has_revenues_expenses_permission()
    )
    if not has_perm:
        messages.error(request, _('You do not have permission to view sector analysis'))
        return redirect('revenues_expenses:dashboard')

    from datetime import date, timedelta
    from .models import Sector

    # Date filters
    today = date.today()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sector_filter = request.GET.get('sector')

    if not start_date:
        start_date = today.replace(day=1)  # Start of current month
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()

    if not end_date:
        end_date = today
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    # Analysis by sector
    sectors = Sector.objects.filter(is_active=True)
    if sector_filter:
        sectors = sectors.filter(id=sector_filter)

    analysis_data = []
    total_revenue = Decimal('0')
    total_expense = Decimal('0')

    for sector in sectors:
        # Sector revenue
        sector_revenue = RevenueExpenseEntry.objects.filter(
            sector=sector,
            type='revenue',
            date__gte=start_date,
            date__lte=end_date,
            is_approved=True
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        # Sector expenses
        sector_expense = RevenueExpenseEntry.objects.filter(
            sector=sector,
            type='expense',
            date__gte=start_date,
            date__lte=end_date,
            is_approved=True
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        # Sector net profit
        sector_profit = sector_revenue - sector_expense

        # Calculate profit percentage
        profit_percentage = Decimal('0')
        if sector_revenue > 0:
            profit_percentage = (sector_profit / sector_revenue) * 100

        analysis_data.append({
            'sector': sector,
            'revenue': sector_revenue,
            'expense': sector_expense,
            'profit': sector_profit,
            'profit_percentage': profit_percentage,
        })

        total_revenue += sector_revenue
        total_expense += sector_expense

    total_profit = total_revenue - total_expense

    # Calculate totals percentage
    total_profit_percentage = Decimal('0')
    if total_revenue > 0:
        total_profit_percentage = (total_profit / total_revenue) * 100

    # Log activity
    try:
        from core.signals import log_view_activity
        class ReportObj:
            id = 0
            pk = 0
            def __str__(self):
                return str(_('Revenue and Expenses Analysis by Sector'))
        log_view_activity(request, 'view', ReportObj(), str(_('View sector analysis')))
    except Exception:
        pass

    context = {
        'page_title': _('Revenue and Expenses Analysis by Sector'),
        'analysis_data': analysis_data,
        'start_date': start_date,
        'end_date': end_date,
        'sector_filter': sector_filter,
        'total_revenue': total_revenue,
        'total_expense': total_expense,
        'total_profit': total_profit,
        'total_profit_percentage': total_profit_percentage,
        'sectors': Sector.objects.filter(is_active=True),
    }

    return render(request, 'revenues_expenses/sector_analysis.html', context)

