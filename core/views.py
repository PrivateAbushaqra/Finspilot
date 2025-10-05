from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView, ListView
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count
from django.db import models
from django.utils import timezone as django_timezone
from datetime import datetime, timedelta
from django.utils.translation import gettext_lazy as _
from django.core.exceptions import PermissionDenied
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_GET
import json
import csv
import io
import traceback

# محاولة استيراد openpyxl للتصدير بصيغة Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .models import SystemNotification, CompanySettings, AuditLog
# سيتم استيراد النماذج عند الحاجة لتجنب الاستيراد الدائري


class DashboardView(LoginRequiredMixin, TemplateView):
    """الشاشة الرئيسية"""
    template_name = 'core/dashboard.html'

    def dispatch(self, request, *args, **kwargs):
        """توجيه مستخدمي POS إلى نقطة البيع مباشرة"""
        if request.user.is_authenticated and request.user.user_type == 'pos_user':
            from django.shortcuts import redirect
            from django.urls import reverse
            return redirect(reverse('sales:pos'))
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # جمع dashboard_sections من مجموعات المستخدم
        user_groups = self.request.user.groups.all()
        dashboard_sections = set()
        for group in user_groups:
            sections = group.dashboard_sections.split(',') if group.dashboard_sections else []
            dashboard_sections.update(sections)
        
        # إضافة الأقسام تلقائياً للمستخدمين ذوي الصلاحيات العالية
        if self.request.user.is_superuser or getattr(self.request.user, 'user_type', None) in ['superadmin', 'admin']:
            privileged_sections = ['sales_stats', 'purchases_stats', 'banks_balances', 'quick_links', 'sales_purchases_distribution', 'monthly_performance']
            dashboard_sections.update(privileged_sections)
        
        context['dashboard_sections'] = list(dashboard_sections)
        
        # تواريخ مختلفة للإحصائيات
        today = django_timezone.now().date()
        start_of_month = today.replace(day=1)
        start_of_quarter = datetime(today.year, ((today.month - 1) // 3) * 3 + 1, 1).date()
        start_of_half_year = datetime(today.year, 1 if today.month <= 6 else 7, 1).date()
        start_of_year = datetime(today.year, 1, 1).date()

        # إحصائيات المبيعات
        sales_stats = self.get_sales_statistics(today, start_of_month, start_of_quarter, start_of_half_year, start_of_year)
        context['sales_stats'] = sales_stats

        # إحصائيات المشتريات
        purchase_stats = self.get_purchase_statistics(today, start_of_month, start_of_quarter, start_of_half_year, start_of_year)
        context['purchase_stats'] = purchase_stats

        # أرصدة البنوك والصناديق
        try:
            from banks.models import BankAccount
            from cashboxes.models import Cashbox
            from receipts.models import PaymentReceipt
            from settings.models import Currency, CompanySettings
            
            # العملة الأساسية من إعدادات الشركة (نحتاجها مبكراً)
            company_settings = CompanySettings.get_settings()
            base_currency_code = company_settings.currency if company_settings else 'JOD'
            
            # أرصدة البنوك (الرصيد الفعلي من المعاملات)
            bank_accounts = BankAccount.objects.filter(is_active=True)
            bank_balances = {}
            for account in bank_accounts:
                currency = account.currency  # CharField
                if currency not in bank_balances:
                    bank_balances[currency] = []
                # حساب الرصيد الفعلي من المعاملات
                actual_balance = account.calculate_actual_balance()
                bank_balances[currency].append({
                    'name': account.name,
                    'balance': actual_balance,
                    'account_number': account.account_number
                })
            
            # أرصدة الصناديق (الرصيد الفعلي من المعاملات)
            cashboxes = Cashbox.objects.filter(is_active=True)
            cashbox_balances = {}
            for cashbox in cashboxes:
                currency = cashbox.currency  # CharField
                if currency not in cashbox_balances:
                    cashbox_balances[currency] = []
                # حساب الرصيد الفعلي من المعاملات
                actual_balance = cashbox.calculate_actual_balance()
                cashbox_balances[currency].append({
                    'name': cashbox.name,
                    'balance': actual_balance,
                    'location': cashbox.location
                })
            
            # الشيكات المعلقة (لم يحن موعد استحقاقها ولم يتم تحصيلها)
            pending_checks = PaymentReceipt.objects.filter(
                payment_type='check',
                check_status='pending',
                check_due_date__gt=django_timezone.now().date(),
                is_active=True,
                is_reversed=False
            ).select_related('customer', 'check_cashbox', 'cashbox')
            
            # تجميع الشيكات حسب العملة
            pending_checks_by_currency = {}
            for check in pending_checks:
                # الحصول على العملة من الصندوق المرتبط بالشيك
                currency = base_currency_code  # استخدام العملة الأساسية كافتراضي
                if check.check_cashbox:
                    currency = check.check_cashbox.currency
                elif check.cashbox:
                    currency = check.cashbox.currency
                
                if currency not in pending_checks_by_currency:
                    pending_checks_by_currency[currency] = {
                        'total_amount': 0,
                        'count': 0,
                        'checks': []
                    }
                
                # التأكد من أن المبلغ أكبر من صفر
                if check.amount > 0:
                    pending_checks_by_currency[currency]['total_amount'] += check.amount
                    pending_checks_by_currency[currency]['count'] += 1
                    pending_checks_by_currency[currency]['checks'].append({
                        'receipt_number': check.receipt_number,
                        'customer_name': check.customer.name,
                        'amount': check.amount,
                        'check_number': check.check_number,
                        'due_date': check.check_due_date
                    })
            
            context['bank_balances'] = bank_balances
            context['cashbox_balances'] = cashbox_balances
            context['pending_checks_by_currency'] = pending_checks_by_currency
            context['base_currency_code'] = base_currency_code
            
        except (ImportError, Exception) as e:
            # الحصول على العملة الأساسية في حالة الخطأ
            try:
                company_settings = CompanySettings.get_settings()
                base_currency_code = company_settings.currency if company_settings else 'JOD'
            except:
                base_currency_code = 'JOD'
                
            context['bank_balances'] = {}
            context['cashbox_balances'] = {}
            context['pending_checks_by_currency'] = {}
            context['base_currency_code'] = base_currency_code

        # الإشعارات غير المقروءة
        unread_notifications = SystemNotification.objects.filter(is_read=False)
        if not self.request.user.is_superuser:
            unread_notifications = unread_notifications.filter(user=self.request.user)
        context['unread_notifications_count'] = unread_notifications.count()

        return context

    def get_sales_statistics(self, today, start_of_month, start_of_quarter, start_of_half_year, start_of_year):
        """حساب إحصائيات المبيعات"""
        try:
            from sales.models import SalesInvoice
            
            stats = {}
            
            # إحصائيات اليوم
            today_sales = SalesInvoice.objects.filter(date=today)
            stats['today'] = {
                'count': today_sales.count(),
                'total': today_sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            # إحصائيات الشهر
            month_sales = SalesInvoice.objects.filter(date__gte=start_of_month)
            stats['month'] = {
                'count': month_sales.count(),
                'total': month_sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            # إحصائيات الربع
            quarter_sales = SalesInvoice.objects.filter(date__gte=start_of_quarter)
            stats['quarter'] = {
                'count': quarter_sales.count(),
                'total': quarter_sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            # إحصائيات نصف السنة
            half_year_sales = SalesInvoice.objects.filter(date__gte=start_of_half_year)
            stats['half_year'] = {
                'count': half_year_sales.count(),
                'total': half_year_sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            # إحصائيات السنة
            year_sales = SalesInvoice.objects.filter(date__gte=start_of_year)
            stats['year'] = {
                'count': year_sales.count(),
                'total': year_sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            return stats
            
        except ImportError:
            # في حالة عدم وجود تطبيق المبيعات
            return {
                'today': {'count': 0, 'total': 0},
                'month': {'count': 0, 'total': 0},
                'quarter': {'count': 0, 'total': 0},
                'half_year': {'count': 0, 'total': 0},
                'year': {'count': 0, 'total': 0},
            }

    def get_purchase_statistics(self, today, start_of_month, start_of_quarter, start_of_half_year, start_of_year):
        """حساب إحصائيات المشتريات"""
        try:
            from purchases.models import PurchaseInvoice
            
            stats = {}
            
            # إحصائيات اليوم
            today_purchases = PurchaseInvoice.objects.filter(date=today)
            stats['today'] = {
                'count': today_purchases.count(),
                'total': today_purchases.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            # إحصائيات الشهر
            month_purchases = PurchaseInvoice.objects.filter(date__gte=start_of_month)
            stats['month'] = {
                'count': month_purchases.count(),
                'total': month_purchases.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            # إحصائيات الربع
            quarter_purchases = PurchaseInvoice.objects.filter(date__gte=start_of_quarter)
            stats['quarter'] = {
                'count': quarter_purchases.count(),
                'total': quarter_purchases.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            # إحصائيات نصف السنة
            half_year_purchases = PurchaseInvoice.objects.filter(date__gte=start_of_half_year)
            stats['half_year'] = {
                'count': half_year_purchases.count(),
                'total': half_year_purchases.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            # إحصائيات السنة
            year_purchases = PurchaseInvoice.objects.filter(date__gte=start_of_year)
            stats['year'] = {
                'count': year_purchases.count(),
                'total': year_purchases.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            return stats
            
        except ImportError:
            # في حالة عدم وجود تطبيق المشتريات
            return {
                'today': {'count': 0, 'total': 0},
                'month': {'count': 0, 'total': 0},
                'quarter': {'count': 0, 'total': 0},
                'half_year': {'count': 0, 'total': 0},
                'year': {'count': 0, 'total': 0},
            }
        

class NotificationListView(LoginRequiredMixin, ListView):
    """قائمة الإشعارات"""
    model = SystemNotification
    template_name = 'core/notifications.html'
    context_object_name = 'notifications'
    paginate_by = 20

    def get_queryset(self):
        queryset = SystemNotification.objects.all()
        if not self.request.user.is_superuser:
            queryset = queryset.filter(user=self.request.user)
        return queryset.order_by('-created_at')


@login_required
def mark_notification_read(request, pk):
    """تحديد الإشعار كمقروء"""
    notification = get_object_or_404(SystemNotification, pk=pk)
    
    # التحقق من الصلاحيات
    if not request.user.is_superuser and notification.user != request.user:
        return JsonResponse({'error': 'ليس لديك صلاحية'}, status=403)
    
    notification.is_read = True
    notification.save()
    
    return JsonResponse({'success': True})


def mark_all_notifications_read(request):
    """تحديد جميع الإشعارات كمقروءة"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    notifications = SystemNotification.objects.filter(is_read=False)
    
    # إذا لم يكن مستخدم خارق، فقط إشعاراته
    if not request.user.is_superuser:
        notifications = notifications.filter(user=request.user)
    
    count = notifications.update(is_read=True)
    
    return JsonResponse({
        'success': True, 
        'count': count,
        'message': f'تم وضع علامة قراءة على {count} إشعار'
    })


class TaxReportView(LoginRequiredMixin, TemplateView):
    """تقرير الضرائب للمبيعات والمشتريات ومردوداتهما"""
    template_name = 'core/tax_report.html'
    
    def dispatch(self, request, *args, **kwargs):
        if request.GET.get('export') == 'excel':
            return self.export_excel(request)
        elif request.GET.get('export') == 'pdf':
            return self.export_pdf(request)
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # معالجة الفلاتر
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        tax_rate_filter = self.request.GET.get('tax_rate')
        customer_supplier_filter = self.request.GET.get('customer_supplier')
        document_type_filter = self.request.GET.get('document_type')
        
        # تحويل التواريخ
        if start_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        else:
            start_date_obj = None
        if end_date:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            end_date_obj = None
        
        # جمع جميع المستندات مع الضرائب
        tax_data = []
        
        # فواتير المبيعات
        try:
            from sales.models import SalesInvoice, SalesInvoiceItem
            sales_invoices = SalesInvoice.objects.all().select_related('customer')
            if start_date_obj:
                sales_invoices = sales_invoices.filter(date__gte=start_date_obj)
            if end_date_obj:
                sales_invoices = sales_invoices.filter(date__lte=end_date_obj)
            if customer_supplier_filter:
                sales_invoices = sales_invoices.filter(customer__name__icontains=customer_supplier_filter)
            
            for invoice in sales_invoices:
                items = SalesInvoiceItem.objects.filter(invoice=invoice, tax_rate__gt=0)
                if tax_rate_filter:
                    items = items.filter(tax_rate=float(tax_rate_filter))
                tax_breakdown = {}
                total_tax = 0
                
                for item in items:
                    tax_rate = float(item.tax_rate)
                    tax_amount = float(item.tax_amount)
                    
                    if tax_rate > 0 and tax_amount > 0:
                        if tax_rate not in tax_breakdown:
                            tax_breakdown[tax_rate] = 0
                        tax_breakdown[tax_rate] += tax_amount
                        total_tax += tax_amount
                
                if total_tax > 0:
                    tax_data.append({
                        'document_number': invoice.invoice_number,
                        'document_type': 'فاتورة مبيعات',
                        'customer_supplier': invoice.customer.name if invoice.customer else 'عميل نقدي',
                        'date': invoice.date,
                        'tax_breakdown': tax_breakdown,
                        'total_tax': total_tax,
                        'is_positive': True,  # المبيعات موجبة
                        'amount_before_tax': float(invoice.subtotal) if hasattr(invoice, 'subtotal') else (float(invoice.total_amount) - total_tax)
                    })
        except (ImportError, Exception) as e:
            pass


    
        # فواتير المشتريات
        try:
            from purchases.models import PurchaseInvoice, PurchaseInvoiceItem
            purchase_invoices = PurchaseInvoice.objects.all().select_related('supplier')
            if start_date_obj:
                purchase_invoices = purchase_invoices.filter(date__gte=start_date_obj)
            if end_date_obj:
                purchase_invoices = purchase_invoices.filter(date__lte=end_date_obj)
            if customer_supplier_filter:
                purchase_invoices = purchase_invoices.filter(supplier__name__icontains=customer_supplier_filter)
            
            for invoice in purchase_invoices:
                items = PurchaseInvoiceItem.objects.filter(invoice=invoice, tax_rate__gt=0)
                if tax_rate_filter:
                    items = items.filter(tax_rate=float(tax_rate_filter))
                tax_breakdown = {}
                total_tax = 0
                
                for item in items:
                    tax_rate = float(item.tax_rate)
                    tax_amount = float(item.tax_amount)
                    
                    if tax_rate > 0 and tax_amount > 0:
                        if tax_rate not in tax_breakdown:
                            tax_breakdown[tax_rate] = 0
                        tax_breakdown[tax_rate] += tax_amount
                        total_tax += tax_amount
                
                if total_tax > 0:
                    tax_data.append({
                        'document_number': invoice.invoice_number,
                        'document_type': 'فاتورة مشتريات',
                        'customer_supplier': invoice.supplier.name if invoice.supplier else 'غير محدد',
                        'date': invoice.date,
                        'tax_breakdown': tax_breakdown,
                        'total_tax': -total_tax,  # المشتريات سالبة
                        'is_positive': False,
                        'amount_before_tax': float(invoice.subtotal) if hasattr(invoice, 'subtotal') else (float(invoice.total_amount) - total_tax)
                    })
        except (ImportError, Exception) as e:
            pass
        
        # مردودات المبيعات
        try:
            from sales.models import SalesReturn, SalesReturnItem
            sales_returns = SalesReturn.objects.all().select_related('customer')
            if start_date_obj:
                sales_returns = sales_returns.filter(date__gte=start_date_obj)
            if end_date_obj:
                sales_returns = sales_returns.filter(date__lte=end_date_obj)
            if customer_supplier_filter:
                sales_returns = sales_returns.filter(customer__name__icontains=customer_supplier_filter)
            if document_type_filter and document_type_filter != 'all':
                if document_type_filter == 'sales_return':
                    pass  # لا حاجة لفلتر إضافي
                else:
                    sales_returns = sales_returns.none()  # لا تطابق
            
            for return_invoice in sales_returns:
                items = SalesReturnItem.objects.filter(return_invoice=return_invoice, tax_rate__gt=0)
                if tax_rate_filter:
                    items = items.filter(tax_rate=float(tax_rate_filter))
                tax_breakdown = {}
                total_tax = 0
                
                for item in items:
                    tax_rate = float(item.tax_rate)
                    tax_amount = float(item.tax_amount)
                    
                    if tax_rate > 0 and tax_amount > 0:
                        if tax_rate not in tax_breakdown:
                            tax_breakdown[tax_rate] = 0
                        tax_breakdown[tax_rate] += tax_amount
                        total_tax += tax_amount
                
                if total_tax > 0:
                    tax_data.append({
                        'document_number': return_invoice.return_number,
                        'document_type': 'مردود مبيعات',
                        'customer_supplier': return_invoice.customer.name if return_invoice.customer else 'غير محدد',
                        'date': return_invoice.date,
                        'tax_breakdown': tax_breakdown,
                        'total_tax': -total_tax,  # مردود المبيعات سالب
                        'is_positive': False,
                        'amount_before_tax': float(return_invoice.subtotal) if hasattr(return_invoice, 'subtotal') else (float(return_invoice.total) - total_tax)
                    })
        except (ImportError, Exception) as e:
            pass
        
        # مردودات المشتريات
        try:
            from purchases.models import PurchaseReturn, PurchaseReturnItem
            purchase_returns = PurchaseReturn.objects.all().select_related('original_invoice__supplier')
            if start_date_obj:
                purchase_returns = purchase_returns.filter(date__gte=start_date_obj)
            if end_date_obj:
                purchase_returns = purchase_returns.filter(date__lte=end_date_obj)
            if customer_supplier_filter:
                purchase_returns = purchase_returns.filter(original_invoice__supplier__name__icontains=customer_supplier_filter)
            if document_type_filter and document_type_filter != 'all':
                if document_type_filter == 'purchase_return':
                    pass
                else:
                    purchase_returns = purchase_returns.none()
            
            for return_invoice in purchase_returns:
                items = PurchaseReturnItem.objects.filter(return_invoice=return_invoice, tax_rate__gt=0)
                if tax_rate_filter:
                    items = items.filter(tax_rate=float(tax_rate_filter))
                tax_breakdown = {}
                total_tax = 0
                
                for item in items:
                    tax_rate = float(item.tax_rate)
                    tax_amount = float(item.tax_amount)
                    
                    if tax_rate > 0 and tax_amount > 0:
                        if tax_rate not in tax_breakdown:
                            tax_breakdown[tax_rate] = 0
                        tax_breakdown[tax_rate] += tax_amount
                        total_tax += tax_amount
                
                if total_tax > 0:
                    tax_data.append({
                        'document_number': return_invoice.return_number,
                        'document_type': 'مردود مشتريات',
                        'customer_supplier': return_invoice.supplier.name if hasattr(return_invoice, 'supplier') and return_invoice.supplier else 'غير محدد',
                        'date': return_invoice.date,
                        'tax_breakdown': tax_breakdown,
                        'total_tax': total_tax,  # مردود المشتريات موجب
                        'is_positive': True,
                        'amount_before_tax': float(return_invoice.subtotal) if hasattr(return_invoice, 'subtotal') else (float(return_invoice.total) - total_tax)
                    })
        except (ImportError, Exception) as e:
            pass
        
        # ترتيب البيانات حسب التاريخ (تصاعدي - الأقدم أولاً)
        tax_data.sort(key=lambda x: x['date'])
        
        # جمع جميع أسعار الضرائب المستخدمة
        all_tax_rates = set()
        for item in tax_data:
            all_tax_rates.update(item['tax_breakdown'].keys())
        all_tax_rates = sorted(list(all_tax_rates))
        
        # حساب إجماليات كل عمود ضريبة
        column_totals = {}
        for rate in all_tax_rates:
            column_totals[rate] = 0
            for item in tax_data:
                if rate in item['tax_breakdown']:
                    if item['is_positive']:
                        column_totals[rate] += item['tax_breakdown'][rate]
                    else:
                        column_totals[rate] -= item['tax_breakdown'][rate]
        
        # حساب الإجماليات
        total_positive = sum(item['total_tax'] for item in tax_data if item['is_positive'])
        total_negative = sum(item['total_tax'] for item in tax_data if not item['is_positive'])
        net_tax = total_positive + total_negative
        
        # حساب إجمالي عمود "إجمالي الضريبة"
        grand_total_tax = sum(item['total_tax'] for item in tax_data)
        
        # حساب إجمالي القيم قبل الضريبة
        total_amount_before_tax = sum(item['amount_before_tax'] for item in tax_data)
        
        context.update({
            'tax_data': tax_data,
            'all_tax_rates': all_tax_rates,
            'column_totals': column_totals,
            'total_positive': total_positive,
            'total_negative': total_negative,
            'net_tax': net_tax,
            'grand_total_tax': grand_total_tax,
            'total_amount_before_tax': total_amount_before_tax,
            'start_date': start_date,
            'end_date': end_date,
            'tax_rate_filter': tax_rate_filter,
            'customer_supplier_filter': customer_supplier_filter,
            'document_type_filter': document_type_filter
        })
        
        return context

    def export_excel(self, request):
        """تصدير التقرير إلى Excel"""
        context = self.get_context_data()
        tax_data = context['tax_data']
        all_tax_rates = context['all_tax_rates']
        column_totals = context['column_totals']
        total_positive = context['total_positive']
        total_negative = context['total_negative']
        net_tax = context['net_tax']
        grand_total_tax = context['grand_total_tax']
        total_amount_before_tax = context['total_amount_before_tax']

        if not OPENPYXL_AVAILABLE:
            return HttpResponse("OpenPyXL غير متوفر", status=500)

        wb = Workbook()
        ws = wb.active
        ws.title = _('Tax Report')

        # Headers
        headers = [_('Document Number'), _('Document Type'), _('Customer/Supplier'), _('Date'), _('Amount Before Tax')]
        for rate in all_tax_rates:
            headers.append(f"{_('Tax')} {rate}%")
        headers.append(_('Total Tax'))

        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Data
        row = 2
        for item in tax_data:
            ws.cell(row=row, column=1, value=item['document_number'])
            ws.cell(row=row, column=2, value=item['document_type'])
            ws.cell(row=row, column=3, value=item['customer_supplier'])
            ws.cell(row=row, column=4, value=item['date'].strftime('%Y-%m-%d'))
            ws.cell(row=row, column=5, value=item['amount_before_tax'])
            col = 6
            for rate in all_tax_rates:
                ws.cell(row=row, column=col, value=item['tax_breakdown'].get(rate, 0))
                col += 1
            ws.cell(row=row, column=col, value=item['total_tax'])
            row += 1

        # Totals
        ws.cell(row=row, column=4, value=_('Totals'))
        ws.cell(row=row, column=5, value=total_amount_before_tax)
        col = 6
        for rate in all_tax_rates:
            ws.cell(row=row, column=col, value=column_totals.get(rate, 0))
            col += 1
        ws.cell(row=row, column=col, value=grand_total_tax)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="tax_report_{django_timezone.now().date()}.xlsx"'
        
        wb.save(response)
        
        # تسجيل النشاط
        try:
            from core.signals import log_export_activity
            log_export_activity(request, str(_('Tax Report')), f'tax_report_{django_timezone.now().date()}.xlsx', 'Excel')
        except Exception:
            pass
        
        return response

    def export_pdf(self, request):
        """تصدير التقرير إلى PDF"""
        context = self.get_context_data()
        tax_data = context['tax_data']
        all_tax_rates = context['all_tax_rates']
        column_totals = context['column_totals']
        total_positive = context['total_positive']
        total_negative = context['total_negative']
        net_tax = context['net_tax']
        grand_total_tax = context['grand_total_tax']
        total_amount_before_tax = context['total_amount_before_tax']

        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from io import BytesIO

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
            elements = []

            styles = getSampleStyleSheet()
            elements.append(Paragraph(_('Tax Report'), styles['Heading1']))
            elements.append(Spacer(1, 12))

            # Data for table
            data = [[_('Document Number'), _('Document Type'), _('Customer/Supplier'), _('Date'), _('Amount Before Tax')]]
            for rate in all_tax_rates:
                data[0].append(f"{_('Tax')} {rate}%")
            data[0].append(_('Total Tax'))

            for item in tax_data:
                row = [
                    item['document_number'],
                    item['document_type'],
                    item['customer_supplier'],
                    item['date'].strftime('%Y-%m-%d'),
                    f"{item['amount_before_tax']:.2f}"
                ]
                for rate in all_tax_rates:
                    row.append(f"{item['tax_breakdown'].get(rate, 0):.2f}")
                row.append(f"{item['total_tax']:.2f}")
                data.append(row)

            # Totals row
            totals_row = ['', '', '', _('Totals'), f"{total_amount_before_tax:.2f}"]
            for rate in all_tax_rates:
                totals_row.append(f"{column_totals.get(rate, 0):.2f}")
            totals_row.append(f"{grand_total_tax:.2f}")
            data.append(totals_row)

            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)

            doc.build(elements)
            buffer.seek(0)
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="tax_report_{django_timezone.now().date()}.pdf"'
            
            # تسجيل النشاط
            try:
                from core.signals import log_export_activity
                log_export_activity(request, str(_('Tax Report')), f'tax_report_{django_timezone.now().date()}.pdf', 'PDF')
            except Exception:
                pass
            
            return response
        except ImportError:
            return HttpResponse("ReportLab غير متوفر", status=500)


class ProfitLossReportView(LoginRequiredMixin, TemplateView):
    """تقرير الأرباح والخسائر الشامل"""
    template_name = 'core/profit_loss_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # استلام تواريخ الفلتر
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        # تحديد الفترة الافتراضية: الشهر الحالي
        if not start_date:
            today = django_timezone.now().date()
            start_date = today.replace(day=1).strftime('%Y-%m-%d')  # بداية الشهر الحالي
        if not end_date:
            end_date = django_timezone.now().date().strftime('%Y-%m-%d')   # اليوم الحالي
            
        # تحويل التواريخ لاستخدامها في الاستعلامات
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # حساب عدد الأيام
        days_count = (end_date_obj - start_date_obj).days + 1  # +1 لتضمين اليوم الأخير
        
        # إرسال البيانات للـ template
        context['start_date'] = start_date  # للعرض كنص
        context['end_date'] = end_date      # للعرض كنص
        context['start_date_obj'] = start_date_obj  # للـ timesince filter
        context['end_date_obj'] = end_date_obj      # للـ timesince filter
        context['days_count'] = days_count  # عدد الأيام
        
        # استيراد النماذج المطلوبة
        try:
            from sales.models import SalesInvoice, SalesReturn
            from purchases.models import PurchaseInvoice, PurchaseReturn
        except ImportError:
            context['error'] = _("حدث خطأ في استيراد النماذج المطلوبة")
            return context
            
        # 1. حساب الإيرادات
        
        # فواتير المبيعات
        sales_invoices = SalesInvoice.objects.filter(
            date__gte=start_date_obj,
            date__lte=end_date_obj
        )
        sales_revenue = sales_invoices.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        # مردودات المبيعات (تطرح من الإيرادات)
        sales_returns = SalesReturn.objects.filter(
            date__gte=start_date_obj,
            date__lte=end_date_obj
        )
        sales_returns_total = sales_returns.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        # صافي الإيرادات
        net_revenue = sales_revenue - sales_returns_total
        
        # 2. حساب التكاليف
        
        # فواتير المشتريات
        purchase_invoices = PurchaseInvoice.objects.filter(
            date__gte=start_date_obj,
            date__lte=end_date_obj
        )
        purchase_costs = purchase_invoices.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        # مردودات المشتريات (تطرح من التكاليف)
        purchase_returns = PurchaseReturn.objects.filter(
            date__gte=start_date_obj,
            date__lte=end_date_obj
        )
        purchase_returns_total = purchase_returns.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        # صافي التكاليف
        net_costs = purchase_costs - purchase_returns_total
        
        # 3. حسابات الربحية
        
        # الربح الإجمالي
        gross_profit = net_revenue - net_costs
        
        # هامش الربح الإجمالي (النسبة المئوية)
        gross_margin = (gross_profit / net_revenue * 100) if net_revenue > 0 else 0
        
        # إضافة البيانات للسياق
        context.update({
            # الإيرادات
            'sales_revenue': sales_revenue,
            'sales_returns_total': sales_returns_total,
            'net_revenue': net_revenue,
            
            # التكاليف
            'purchase_costs': purchase_costs,
            'purchase_returns_total': purchase_returns_total,
            'net_costs': net_costs,
            
            # الأرباح
            'gross_profit': gross_profit,
            'gross_margin': gross_margin,
            
            # بيانات إضافية
            'sales_count': sales_invoices.count(),
            'sales_returns_count': sales_returns.count(),
            'purchase_count': purchase_invoices.count(),
            'purchase_returns_count': purchase_returns.count(),
            
            # العملة
            'currency': self.get_base_currency(),
        })
        
        return context
        
    def get_base_currency(self):
        """الحصول على العملة الأساسية"""
        try:
            from settings.models import CompanySettings, Currency
            company_settings = CompanySettings.objects.first()
            if company_settings and company_settings.base_currency:
                return company_settings.base_currency
            return Currency.objects.filter(is_base=True).first() or Currency.objects.first()
        except Exception:
            return None


class AuditLogListView(LoginRequiredMixin, ListView):
    """عرض سجل الأنشطة"""
    model = None  # سيتم تحديده في get_queryset
    template_name = 'core/audit_log.html'
    context_object_name = 'audit_logs'
    paginate_by = 50
    
    def dispatch(self, request, *args, **kwargs):
        """التحقق من الصلاحيات"""
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        
        if not (getattr(request.user, 'is_admin', False) or request.user.has_perm('core.view_audit_log')):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied(_('ليس لديك صلاحية للوصول إلى سجل الأنشطة'))
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        """الحصول على سجل الأنشطة مع الفلاتر"""
        from .models import AuditLog
        queryset = AuditLog.objects.select_related('user').order_by('-timestamp')
        
        # فلتر حسب المستخدم
        user_id = self.request.GET.get('user')
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        
        # فلتر حسب نوع العملية
        action_type = self.request.GET.get('action_type')
        if action_type:
            queryset = queryset.filter(action_type=action_type)
        
        # فلتر حسب نوع المحتوى
        content_type = self.request.GET.get('content_type')
        if content_type:
            queryset = queryset.filter(content_type__icontains=content_type)
        
        # فلتر حسب التاريخ
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        if date_from:
            try:
                date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(timestamp__date__gte=date_from)
            except ValueError:
                pass
        
        if date_to:
            try:
                date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(timestamp__date__lte=date_to)
            except ValueError:
                pass
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # إضافة قائمة المستخدمين للفلتر
        from django.contrib.auth import get_user_model
        User = get_user_model()
        context['users'] = User.objects.filter(is_active=True).order_by('username')
        
        # إضافة أنواع العمليات
        from .models import AuditLog
        context['action_types'] = AuditLog.ACTION_TYPES
        
        # إضافة أنواع المحتوى المختلفة
        context['content_types'] = AuditLog.objects.values_list('content_type', flat=True).distinct()
        
        # إضافة فلاتر البحث الحالية
        context['current_filters'] = {
            'user': self.request.GET.get('user', ''),
            'action_type': self.request.GET.get('action_type', ''),
            'content_type': self.request.GET.get('content_type', ''),
            'date_from': self.request.GET.get('date_from', ''),
            'date_to': self.request.GET.get('date_to', ''),
        }
        
        return context


@require_http_methods(["POST"])
@login_required
def sync_balances_view(request):
    """مزامنة أرصدة البنوك والصناديق"""
    # التحقق من صلاحيات المستخدم
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'غير مصرح لك بهذه العملية'})
    
    try:
        from banks.models import BankAccount
        from cashboxes.models import Cashbox
        
        banks_updated = 0
        cashboxes_updated = 0
        
        # مزامنة أرصدة البنوك
        bank_accounts = BankAccount.objects.filter(is_active=True)
        for account in bank_accounts:
            old_balance = account.balance
            new_balance = account.sync_balance()
            if old_balance != new_balance:
                banks_updated += 1
        
        # مزامنة أرصدة الصناديق
        cashboxes = Cashbox.objects.filter(is_active=True)
        for cashbox in cashboxes:
            old_balance = cashbox.balance
            new_balance = cashbox.sync_balance()
            if old_balance != new_balance:
                cashboxes_updated += 1
        
        return JsonResponse({
            'success': True,
            'banks_updated': banks_updated,
            'cashboxes_updated': cashboxes_updated
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def extend_session_api(request):
    """API لتمديد الجلسة"""
    try:
        # تحديث وقت آخر نشاط في الجلسة
        request.session['last_activity'] = django_timezone.now().isoformat()
        request.session.modified = True
        
        return JsonResponse({
            'success': True,
            'message': 'تم تمديد الجلسة بنجاح',
            'timestamp': django_timezone.now().isoformat()
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'خطأ في تمديد الجلسة: {str(e)}'
        }, status=500)


@require_GET
@login_required
def audit_log_export_excel(request):
    """تصدير سجل الأنشطة إلى ملف Excel"""
    if not (getattr(request.user, 'is_superuser', False) or getattr(request.user, 'user_type', None) == 'superadmin'):
        raise PermissionDenied('غير مصرح لك')
    from .models import AuditLog
    queryset = AuditLog.objects.select_related('user').order_by('-timestamp')
    # تطبيق نفس الفلاتر الموجودة في AuditLogListView
    user_id = request.GET.get('user')
    if user_id:
        queryset = queryset.filter(user_id=user_id)
    action_type = request.GET.get('action_type')
    if action_type:
        queryset = queryset.filter(action_type=action_type)
    content_type = request.GET.get('content_type')
    if content_type:
        queryset = queryset.filter(content_type__icontains=content_type)
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            queryset = queryset.filter(timestamp__date__gte=date_from)
        except ValueError:
            pass
    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            queryset = queryset.filter(timestamp__date__lte=date_to)
        except ValueError:
            pass
    # استيراد متأخر لتفادي اعتماد xlsxwriter عند استيراد views
    try:
        from .export_audit_log_excel import export_audit_log_to_excel
        return export_audit_log_to_excel(queryset)
    except Exception as e:
        # في حال عدم توفر المكتبة، إرجاع CSV بسيط كبديل حتى لا ينكسر النظام
        import csv, io
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['الوقت', 'المستخدم', 'نوع العملية', 'نوع المحتوى', 'الوصف', 'عنوان IP'])
        for log in queryset:
            writer.writerow([
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                getattr(log.user, 'username', ''),
                getattr(log, 'action_type', ''),
                getattr(log, 'content_type', ''),
                getattr(log, 'description', ''),
                getattr(log, 'ip_address', ''),
            ])
        from django.http import HttpResponse
        response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename=audit_log.csv'
        return response

from django.contrib.auth import logout as django_logout
from django.utils.translation import get_language
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

@csrf_exempt
@require_http_methods(["POST", "GET"])  # دعم GET كتحسين إضافي
def logout_alias(request):
    """Alias للتعامل مع /logout/ القادمة من sendBeacon وإعادة التوجيه للمسار الصحيح.
    يحترم i18n بإعادة التوجيه إلى /<lang>/auth/logout/.
    """
    try:
        lang = get_language() or 'ar'
    except Exception:
        lang = 'ar'

    # نفّذ logout فوراً (للطلبات POST القادمة من beacon قد لا تحتوي CSRF)
    if request.user.is_authenticated:
        try:
            django_logout(request)
        except Exception:
            pass

    # إعادة التوجيه لمسار الخروج الرسمي لضمان أي hooks/إشعارات موجودة هناك
    redirect_url = f"/{lang}/auth/logout/"
    return redirect(redirect_url)


class SalesBySalespersonReportView(LoginRequiredMixin, TemplateView):
    """تقرير المبيعات حسب البائع"""
    template_name = 'core/sales_by_salesperson_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # استلام تواريخ الفلتر
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        salesperson_id = self.request.GET.get('salesperson')
        
        # تحديد الفترة الافتراضية: الشهر الحالي
        if not start_date:
            today = django_timezone.now().date()
            start_date = today.replace(day=1).strftime('%Y-%m-%d')  # بداية الشهر الحالي
        if not end_date:
            end_date = django_timezone.now().date().strftime('%Y-%m-%d')   # اليوم الحالي
            
        # تحويل التواريخ لاستخدامها في الاستعلامات
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # استيراد النماذج المطلوبة
        try:
            from sales.models import SalesInvoice
            from django.contrib.auth import get_user_model
            User = get_user_model()
        except ImportError:
            context['error'] = _("حدث خطأ في استيراد النماذج المطلوبة")
            return context
            
        # الحصول على قائمة البائعين (المستخدمين الذين أنشأوا فواتير مبيعات)
        salespeople = User.objects.filter(
            salesinvoice__date__gte=start_date_obj,
            salesinvoice__date__lte=end_date_obj
        ).distinct().order_by('username')
        
        # فلترة الفواتير حسب البائع المحدد
        sales_invoices = SalesInvoice.objects.filter(
            date__gte=start_date_obj,
            date__lte=end_date_obj
        ).select_related('customer', 'created_by')
        
        if salesperson_id:
            sales_invoices = sales_invoices.filter(created_by_id=salesperson_id)
        
        # إحصائيات المبيعات حسب البائع
        salesperson_stats = sales_invoices.values(
            'created_by__username', 
            'created_by__first_name', 
            'created_by__last_name'
        ).annotate(
            total_invoices=Count('id'),
            total_amount=Sum('total_amount'),
            total_tax=Sum('tax_amount'),
            total_discount=Sum('discount_amount')
        ).order_by('-total_amount')
        
        # إرسال البيانات للـ template
        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'salesperson_id': salesperson_id,
            'salespeople': salespeople,
            'sales_invoices': sales_invoices,
            'salesperson_stats': salesperson_stats,
        })
        
        # تسجيل النشاط في سجل المراجعة
        from .signals import log_user_activity
        log_user_activity(
            self.request, 
            'view', 
            self, 
            _('عرض تقرير المبيعات حسب البائع')
        )
        
        return context


# ===== View لإثبات استخدام PostgreSQL =====
@login_required
@require_GET
def database_info_view(request):
    """صفحة تعرض معلومات قاعدة البيانات لإثبات استخدام PostgreSQL"""
    from django.db import connection

    # معلومات قاعدة البيانات
    db_info = {
        'database_type': connection.vendor,
        'database_name': connection.settings_dict.get('NAME', 'غير محدد'),
        'database_host': connection.settings_dict.get('HOST', 'غير محدد'),
        'database_port': connection.settings_dict.get('PORT', 'غير محدد'),
        'database_engine': connection.settings_dict.get('ENGINE', 'غير محدد'),
        'is_postgresql': connection.vendor.lower() == 'postgresql',
        'is_sqlite': connection.vendor.lower() == 'sqlite',
        'connection_status': 'متصل' if connection.ensure_connection() is None else 'غير متصل',
    }

    # معلومات Django
    django_info = {
        'django_version': '4.2.7',
        'python_version': '3.11.9',
        'debug_mode': 'تشغيل' if getattr(connection, '_settings_dict', {}).get('DEBUG', False) else 'إيقاف',
        'render_deployment': 'نعم' if request.META.get('HTTP_HOST', '').endswith('onrender.com') else 'لا',
    }

    context = {
        'db_info': db_info,
        'django_info': django_info,
        'title': 'معلومات قاعدة البيانات - إثبات استخدام PostgreSQL',
    }

    return render(request, 'core/database_info.html', context)


def language_switch_view(request):
    """تبديل اللغة وإعادة التوجيه حسب اللغة المختارة"""
    from django.utils.translation import activate
    from django.shortcuts import redirect

    if request.method == 'POST':
        language = request.POST.get('language', 'ar')
        next_url = request.POST.get('next', None)
    else:
        language = request.GET.get('language', 'ar')
        next_url = request.GET.get('next', None)

    # تفعيل اللغة الجديدة
    activate(language)
    request.session['_language'] = language

    # تسجيل النشاط في سجل المراجعة فقط إذا كان المستخدم مسجل دخول
    if request.user.is_authenticated:
        from .signals import log_user_activity
        log_user_activity(
            request,
            'update',
            None,
            f'تبديل اللغة إلى: {language}'
        )

    # إعادة التوجيه - استخدام next_url إذا كان موجوداً، وإلا حسب اللغة
    if next_url:
        return redirect(next_url)
    elif language == 'en':
        return redirect('/en/')
    else:
        return redirect('/ar/')