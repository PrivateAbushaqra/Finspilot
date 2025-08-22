from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView, ListView
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count
from django.db import models
from django.utils import timezone as django_timezone
from datetime import datetime, timedelta
from django.utils.translation import gettext_lazy as _
from django.core.exceptions import PermissionDenied
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
import csv
import io
import traceback

# محاولة استيراد openpyxl للتصدير بصيغة Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .models import SystemNotification, CompanySettings, AuditLog
# سيتم استيراد النماذج عند الحاجة لتجنب الاستيراد الدائري


class DashboardView(LoginRequiredMixin, TemplateView):
    """الشاشة الرئيسية"""
    template_name = 'core/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # تواريخ مختلفة للإحصائيات
        today = django_timezone.now().date()
        start_of_month = today.replace(day=1)
        start_of_quarter = datetime(today.year, ((today.month - 1) // 3) * 3 + 1, 1).date()
        start_of_half_year = datetime(today.year, 1 if today.month <= 6 else 7, 1).date()
        start_of_year = datetime(today.year, 1, 1).date()

        # إحصائيات المبيعات
        sales_stats = self.get_sales_statistics(today, start_of_month, start_of_quarter, start_of_half_year, start_of_year)
        context['sales_stats'] = sales_stats

        # إحصائيات المشتريات
        purchase_stats = self.get_purchase_statistics(today, start_of_month, start_of_quarter, start_of_half_year, start_of_year)
        context['purchase_stats'] = purchase_stats

        # أرصدة البنوك والصناديق
        try:
            from banks.models import BankAccount
            from cashboxes.models import Cashbox
            from receipts.models import PaymentReceipt
            from settings.models import Currency, CompanySettings
            
            # أرصدة البنوك (الرصيد الفعلي من المعاملات)
            bank_accounts = BankAccount.objects.filter(is_active=True)
            bank_balances = {}
            for account in bank_accounts:
                currency = account.currency  # CharField
                if currency not in bank_balances:
                    bank_balances[currency] = []
                # حساب الرصيد الفعلي من المعاملات
                actual_balance = account.calculate_actual_balance()
                bank_balances[currency].append({
                    'name': account.name,
                    'balance': actual_balance,
                    'account_number': account.account_number
                })
            
            # أرصدة الصناديق (الرصيد الفعلي من المعاملات)
            cashboxes = Cashbox.objects.filter(is_active=True)
            cashbox_balances = {}
            for cashbox in cashboxes:
                currency = cashbox.currency  # CharField
                if currency not in cashbox_balances:
                    cashbox_balances[currency] = []
                # حساب الرصيد الفعلي من المعاملات
                actual_balance = cashbox.calculate_actual_balance()
                cashbox_balances[currency].append({
                    'name': cashbox.name,
                    'balance': actual_balance,
                    'location': cashbox.location
                })
            
            # الشيكات المعلقة (لم يحن موعد استحقاقها ولم يتم تحصيلها)
            pending_checks = PaymentReceipt.objects.filter(
                payment_type='check',
                check_status='pending',
                check_due_date__gt=django_timezone.now().date(),
                is_active=True,
                is_reversed=False
            ).select_related('customer', 'check_cashbox', 'cashbox')
            
            # تجميع الشيكات حسب العملة
            pending_checks_by_currency = {}
            for check in pending_checks:
                # الحصول على العملة من الصندوق المرتبط بالشيك
                currency = 'SAR'  # افتراضي
                if check.check_cashbox:
                    currency = check.check_cashbox.currency
                elif check.cashbox:
                    currency = check.cashbox.currency
                
                if currency not in pending_checks_by_currency:
                    pending_checks_by_currency[currency] = {
                        'total_amount': 0,
                        'count': 0,
                        'checks': []
                    }
                
                # التأكد من أن المبلغ أكبر من صفر
                if check.amount > 0:
                    pending_checks_by_currency[currency]['total_amount'] += check.amount
                    pending_checks_by_currency[currency]['count'] += 1
                    pending_checks_by_currency[currency]['checks'].append({
                        'receipt_number': check.receipt_number,
                        'customer_name': check.customer.name,
                        'amount': check.amount,
                        'check_number': check.check_number,
                        'due_date': check.check_due_date
                    })
            
            # العملة الأساسية من إعدادات الشركة
            company_settings = CompanySettings.objects.first()
            if company_settings and company_settings.base_currency:
                base_currency_code = company_settings.base_currency.code
            else:
                base_currency = Currency.get_base_currency()
                base_currency_code = base_currency.code if base_currency else 'SAR'
            
            context['bank_balances'] = bank_balances
            context['cashbox_balances'] = cashbox_balances
            context['pending_checks_by_currency'] = pending_checks_by_currency
            context['base_currency_code'] = base_currency_code
            
        except (ImportError, Exception) as e:
            context['bank_balances'] = {}
            context['cashbox_balances'] = {}
            context['pending_checks_by_currency'] = {}
            context['base_currency_code'] = 'SAR'

        # الإشعارات غير المقروءة
        unread_notifications = SystemNotification.objects.filter(is_read=False)
        if not self.request.user.is_superuser:
            unread_notifications = unread_notifications.filter(user=self.request.user)
        context['unread_notifications_count'] = unread_notifications.count()

        return context

    def get_sales_statistics(self, today, start_of_month, start_of_quarter, start_of_half_year, start_of_year):
        """حساب إحصائيات المبيعات"""
        try:
            from sales.models import SalesInvoice
            
            stats = {}
            
            # إحصائيات اليوم
            today_sales = SalesInvoice.objects.filter(date=today)
            stats['today'] = {
                'count': today_sales.count(),
                'total': today_sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            # إحصائيات الشهر
            month_sales = SalesInvoice.objects.filter(date__gte=start_of_month)
            stats['month'] = {
                'count': month_sales.count(),
                'total': month_sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            # إحصائيات الربع
            quarter_sales = SalesInvoice.objects.filter(date__gte=start_of_quarter)
            stats['quarter'] = {
                'count': quarter_sales.count(),
                'total': quarter_sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            # إحصائيات نصف السنة
            half_year_sales = SalesInvoice.objects.filter(date__gte=start_of_half_year)
            stats['half_year'] = {
                'count': half_year_sales.count(),
                'total': half_year_sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            # إحصائيات السنة
            year_sales = SalesInvoice.objects.filter(date__gte=start_of_year)
            stats['year'] = {
                'count': year_sales.count(),
                'total': year_sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            return stats
            
        except ImportError:
            # في حالة عدم وجود تطبيق المبيعات
            return {
                'today': {'count': 0, 'total': 0},
                'month': {'count': 0, 'total': 0},
                'quarter': {'count': 0, 'total': 0},
                'half_year': {'count': 0, 'total': 0},
                'year': {'count': 0, 'total': 0},
            }

    def get_purchase_statistics(self, today, start_of_month, start_of_quarter, start_of_half_year, start_of_year):
        """حساب إحصائيات المشتريات"""
        try:
            from purchases.models import PurchaseInvoice
            
            stats = {}
            
            # إحصائيات اليوم
            today_purchases = PurchaseInvoice.objects.filter(date=today)
            stats['today'] = {
                'count': today_purchases.count(),
                'total': today_purchases.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            # إحصائيات الشهر
            month_purchases = PurchaseInvoice.objects.filter(date__gte=start_of_month)
            stats['month'] = {
                'count': month_purchases.count(),
                'total': month_purchases.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            # إحصائيات الربع
            quarter_purchases = PurchaseInvoice.objects.filter(date__gte=start_of_quarter)
            stats['quarter'] = {
                'count': quarter_purchases.count(),
                'total': quarter_purchases.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            # إحصائيات نصف السنة
            half_year_purchases = PurchaseInvoice.objects.filter(date__gte=start_of_half_year)
            stats['half_year'] = {
                'count': half_year_purchases.count(),
                'total': half_year_purchases.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            # إحصائيات السنة
            year_purchases = PurchaseInvoice.objects.filter(date__gte=start_of_year)
            stats['year'] = {
                'count': year_purchases.count(),
                'total': year_purchases.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            }
            
            return stats
            
        except ImportError:
            # في حالة عدم وجود تطبيق المشتريات
            return {
                'today': {'count': 0, 'total': 0},
                'month': {'count': 0, 'total': 0},
                'quarter': {'count': 0, 'total': 0},
                'half_year': {'count': 0, 'total': 0},
                'year': {'count': 0, 'total': 0},
            }
        

class NotificationListView(LoginRequiredMixin, ListView):
    """قائمة الإشعارات"""
    model = SystemNotification
    template_name = 'core/notifications.html'
    context_object_name = 'notifications'
    paginate_by = 20

    def get_queryset(self):
        queryset = SystemNotification.objects.all()
        if not self.request.user.is_superuser:
            queryset = queryset.filter(user=self.request.user)
        return queryset.order_by('-created_at')


@login_required
def mark_notification_read(request, pk):
    """تحديد الإشعار كمقروء"""
    notification = get_object_or_404(SystemNotification, pk=pk)
    
    # التحقق من الصلاحيات
    if not request.user.is_superuser and notification.user != request.user:
        return JsonResponse({'error': 'ليس لديك صلاحية'}, status=403)
    
    notification.is_read = True
    notification.save()
    
    return JsonResponse({'success': True})


class TaxReportView(LoginRequiredMixin, TemplateView):
    """تقرير الضرائب للمبيعات والمشتريات ومردوداتهما"""
    template_name = 'core/tax_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # جمع جميع المستندات مع الضرائب
        tax_data = []
        
        # فواتير المبيعات
        try:
            from sales.models import SalesInvoice, SalesInvoiceItem
            sales_invoices = SalesInvoice.objects.all().select_related('customer')
            
            for invoice in sales_invoices:
                items = SalesInvoiceItem.objects.filter(invoice=invoice, tax_rate__gt=0)
                tax_breakdown = {}
                total_tax = 0
                
                for item in items:
                    tax_rate = float(item.tax_rate)
                    tax_amount = float(item.tax_amount)
                    
                    if tax_rate > 0 and tax_amount > 0:
                        if tax_rate not in tax_breakdown:
                            tax_breakdown[tax_rate] = 0
                        tax_breakdown[tax_rate] += tax_amount
                        total_tax += tax_amount
                
                if total_tax > 0:
                    tax_data.append({
                        'document_number': invoice.invoice_number,
                        'document_type': 'فاتورة مبيعات',
                        'customer_supplier': invoice.customer.name if invoice.customer else 'عميل نقدي',
                        'date': invoice.date,
                        'tax_breakdown': tax_breakdown,
                        'total_tax': total_tax,
                        'is_positive': True,  # المبيعات موجبة
                        'amount_before_tax': float(invoice.subtotal) if hasattr(invoice, 'subtotal') else (float(invoice.total_amount) - total_tax)
                    })
        except (ImportError, Exception) as e:
            pass
        
        # فواتير المشتريات
        try:
            from purchases.models import PurchaseInvoice, PurchaseInvoiceItem
            purchase_invoices = PurchaseInvoice.objects.all().select_related('supplier')
            
            for invoice in purchase_invoices:
                items = PurchaseInvoiceItem.objects.filter(invoice=invoice, tax_rate__gt=0)
                tax_breakdown = {}
                total_tax = 0
                
                for item in items:
                    tax_rate = float(item.tax_rate)
                    tax_amount = float(item.tax_amount)
                    
                    if tax_rate > 0 and tax_amount > 0:
                        if tax_rate not in tax_breakdown:
                            tax_breakdown[tax_rate] = 0
                        tax_breakdown[tax_rate] += tax_amount
                        total_tax += tax_amount
                
                if total_tax > 0:
                    tax_data.append({
                        'document_number': invoice.invoice_number,
                        'document_type': 'فاتورة مشتريات',
                        'customer_supplier': invoice.supplier.name if invoice.supplier else 'غير محدد',
                        'date': invoice.date,
                        'tax_breakdown': tax_breakdown,
                        'total_tax': -total_tax,  # المشتريات سالبة
                        'is_positive': False,
                        'amount_before_tax': float(invoice.subtotal) if hasattr(invoice, 'subtotal') else (float(invoice.total_amount) - total_tax)
                    })
        except (ImportError, Exception) as e:
            pass
        
        # مردودات المبيعات
        try:
            from sales.models import SalesReturn, SalesReturnItem
            sales_returns = SalesReturn.objects.all().select_related('customer')
            
            for return_invoice in sales_returns:
                items = SalesReturnItem.objects.filter(return_invoice=return_invoice, tax_rate__gt=0)
                tax_breakdown = {}
                total_tax = 0
                
                for item in items:
                    tax_rate = float(item.tax_rate)
                    tax_amount = float(item.tax_amount)
                    
                    if tax_rate > 0 and tax_amount > 0:
                        if tax_rate not in tax_breakdown:
                            tax_breakdown[tax_rate] = 0
                        tax_breakdown[tax_rate] += tax_amount
                        total_tax += tax_amount
                
                if total_tax > 0:
                    tax_data.append({
                        'document_number': return_invoice.return_number,
                        'document_type': 'مردود مبيعات',
                        'customer_supplier': return_invoice.customer.name if return_invoice.customer else 'غير محدد',
                        'date': return_invoice.date,
                        'tax_breakdown': tax_breakdown,
                        'total_tax': -total_tax,  # مردود المبيعات سالب
                        'is_positive': False,
                        'amount_before_tax': float(return_invoice.subtotal) if hasattr(return_invoice, 'subtotal') else (float(return_invoice.total) - total_tax)
                    })
        except (ImportError, Exception) as e:
            pass
        
        # مردودات المشتريات
        try:
            from purchases.models import PurchaseReturn, PurchaseReturnItem
            purchase_returns = PurchaseReturn.objects.all().select_related('original_invoice__supplier')
            
            for return_invoice in purchase_returns:
                items = PurchaseReturnItem.objects.filter(return_invoice=return_invoice, tax_rate__gt=0)
                tax_breakdown = {}
                total_tax = 0
                
                for item in items:
                    tax_rate = float(item.tax_rate)
                    tax_amount = float(item.tax_amount)
                    
                    if tax_rate > 0 and tax_amount > 0:
                        if tax_rate not in tax_breakdown:
                            tax_breakdown[tax_rate] = 0
                        tax_breakdown[tax_rate] += tax_amount
                        total_tax += tax_amount
                
                if total_tax > 0:
                    tax_data.append({
                        'document_number': return_invoice.return_number,
                        'document_type': 'مردود مشتريات',
                        'customer_supplier': return_invoice.supplier.name if hasattr(return_invoice, 'supplier') and return_invoice.supplier else 'غير محدد',
                        'date': return_invoice.date,
                        'tax_breakdown': tax_breakdown,
                        'total_tax': total_tax,  # مردود المشتريات موجب
                        'is_positive': True,
                        'amount_before_tax': float(return_invoice.subtotal) if hasattr(return_invoice, 'subtotal') else (float(return_invoice.total) - total_tax)
                    })
        except (ImportError, Exception) as e:
            pass
        
        # ترتيب البيانات حسب التاريخ (تصاعدي - الأقدم أولاً)
        tax_data.sort(key=lambda x: x['date'])
        
        # جمع جميع أسعار الضرائب المستخدمة
        all_tax_rates = set()
        for item in tax_data:
            all_tax_rates.update(item['tax_breakdown'].keys())
        all_tax_rates = sorted(list(all_tax_rates))
        
        # حساب إجماليات كل عمود ضريبة
        column_totals = {}
        for rate in all_tax_rates:
            column_totals[rate] = 0
            for item in tax_data:
                if rate in item['tax_breakdown']:
                    if item['is_positive']:
                        column_totals[rate] += item['tax_breakdown'][rate]
                    else:
                        column_totals[rate] -= item['tax_breakdown'][rate]
        
        # حساب الإجماليات
        total_positive = sum(item['total_tax'] for item in tax_data if item['is_positive'])
        total_negative = sum(item['total_tax'] for item in tax_data if not item['is_positive'])
        net_tax = total_positive + total_negative
        
        # حساب إجمالي عمود "إجمالي الضريبة"
        grand_total_tax = sum(item['total_tax'] for item in tax_data)
        
        # حساب إجمالي القيم قبل الضريبة
        total_amount_before_tax = sum(item['amount_before_tax'] for item in tax_data)
        
        context.update({
            'tax_data': tax_data,
            'all_tax_rates': all_tax_rates,
            'column_totals': column_totals,
            'total_positive': total_positive,
            'total_negative': total_negative,
            'net_tax': net_tax,
            'grand_total_tax': grand_total_tax,
            'total_amount_before_tax': total_amount_before_tax
        })
        
        return context


class ProfitLossReportView(LoginRequiredMixin, TemplateView):
    """تقرير الأرباح والخسائر الشامل"""
    template_name = 'core/profit_loss_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # استلام تواريخ الفلتر
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        # تحديد الفترة الافتراضية: الشهر الحالي
        if not start_date:
            today = django_timezone.now().date()
            start_date = today.replace(day=1).strftime('%Y-%m-%d')  # بداية الشهر الحالي
        if not end_date:
            end_date = django_timezone.now().date().strftime('%Y-%m-%d')   # اليوم الحالي
            
        # تحويل التواريخ لاستخدامها في الاستعلامات
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # حساب عدد الأيام
        days_count = (end_date_obj - start_date_obj).days + 1  # +1 لتضمين اليوم الأخير
        
        # إرسال البيانات للـ template
        context['start_date'] = start_date  # للعرض كنص
        context['end_date'] = end_date      # للعرض كنص
        context['start_date_obj'] = start_date_obj  # للـ timesince filter
        context['end_date_obj'] = end_date_obj      # للـ timesince filter
        context['days_count'] = days_count  # عدد الأيام
        
        # استيراد النماذج المطلوبة
        try:
            from sales.models import SalesInvoice, SalesReturn
            from purchases.models import PurchaseInvoice, PurchaseReturn
        except ImportError:
            context['error'] = _("حدث خطأ في استيراد النماذج المطلوبة")
            return context
            
        # 1. حساب الإيرادات
        
        # فواتير المبيعات
        sales_invoices = SalesInvoice.objects.filter(
            date__gte=start_date_obj,
            date__lte=end_date_obj
        )
        sales_revenue = sales_invoices.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        # مردودات المبيعات (تطرح من الإيرادات)
        sales_returns = SalesReturn.objects.filter(
            date__gte=start_date_obj,
            date__lte=end_date_obj
        )
        sales_returns_total = sales_returns.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        # صافي الإيرادات
        net_revenue = sales_revenue - sales_returns_total
        
        # 2. حساب التكاليف
        
        # فواتير المشتريات
        purchase_invoices = PurchaseInvoice.objects.filter(
            date__gte=start_date_obj,
            date__lte=end_date_obj
        )
        purchase_costs = purchase_invoices.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        # مردودات المشتريات (تطرح من التكاليف)
        purchase_returns = PurchaseReturn.objects.filter(
            date__gte=start_date_obj,
            date__lte=end_date_obj
        )
        purchase_returns_total = purchase_returns.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        # صافي التكاليف
        net_costs = purchase_costs - purchase_returns_total
        
        # 3. حسابات الربحية
        
        # الربح الإجمالي
        gross_profit = net_revenue - net_costs
        
        # هامش الربح الإجمالي (النسبة المئوية)
        gross_margin = (gross_profit / net_revenue * 100) if net_revenue > 0 else 0
        
        # إضافة البيانات للسياق
        context.update({
            # الإيرادات
            'sales_revenue': sales_revenue,
            'sales_returns_total': sales_returns_total,
            'net_revenue': net_revenue,
            
            # التكاليف
            'purchase_costs': purchase_costs,
            'purchase_returns_total': purchase_returns_total,
            'net_costs': net_costs,
            
            # الأرباح
            'gross_profit': gross_profit,
            'gross_margin': gross_margin,
            
            # بيانات إضافية
            'sales_count': sales_invoices.count(),
            'sales_returns_count': sales_returns.count(),
            'purchase_count': purchase_invoices.count(),
            'purchase_returns_count': purchase_returns.count(),
            
            # العملة
            'currency': self.get_base_currency(),
        })
        
        return context
        
    def get_base_currency(self):
        """الحصول على العملة الأساسية"""
        try:
            from settings.models import CompanySettings, Currency
            company_settings = CompanySettings.objects.first()
            if company_settings and company_settings.base_currency:
                return company_settings.base_currency
            return Currency.objects.filter(is_base=True).first() or Currency.objects.first()
        except Exception:
            return None


class AuditLogListView(LoginRequiredMixin, ListView):
    """عرض سجل الأنشطة"""
    model = None  # سيتم تحديده في get_queryset
    template_name = 'core/audit_log.html'
    context_object_name = 'audit_logs'
    paginate_by = 50
    
    def dispatch(self, request, *args, **kwargs):
        """التحقق من الصلاحيات"""
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        
        if not (getattr(request.user, 'is_admin', False) or request.user.has_perm('core.view_audit_log')):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied(_('ليس لديك صلاحية للوصول إلى سجل الأنشطة'))
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        """الحصول على سجل الأنشطة مع الفلاتر"""
        from .models import AuditLog
        queryset = AuditLog.objects.select_related('user').order_by('-timestamp')
        
        # فلتر حسب المستخدم
        user_id = self.request.GET.get('user')
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        
        # فلتر حسب نوع العملية
        action_type = self.request.GET.get('action_type')
        if action_type:
            queryset = queryset.filter(action_type=action_type)
        
        # فلتر حسب نوع المحتوى
        content_type = self.request.GET.get('content_type')
        if content_type:
            queryset = queryset.filter(content_type__icontains=content_type)
        
        # فلتر حسب التاريخ
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        if date_from:
            try:
                date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(timestamp__date__gte=date_from)
            except ValueError:
                pass
        
        if date_to:
            try:
                date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(timestamp__date__lte=date_to)
            except ValueError:
                pass
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # إضافة قائمة المستخدمين للفلتر
        from django.contrib.auth import get_user_model
        User = get_user_model()
        context['users'] = User.objects.filter(is_active=True).order_by('username')
        
        # إضافة أنواع العمليات
        from .models import AuditLog
        context['action_types'] = AuditLog.ACTION_TYPES
        
        # إضافة أنواع المحتوى المختلفة
        context['content_types'] = AuditLog.objects.values_list('content_type', flat=True).distinct()
        
        # إضافة فلاتر البحث الحالية
        context['current_filters'] = {
            'user': self.request.GET.get('user', ''),
            'action_type': self.request.GET.get('action_type', ''),
            'content_type': self.request.GET.get('content_type', ''),
            'date_from': self.request.GET.get('date_from', ''),
            'date_to': self.request.GET.get('date_to', ''),
        }
        
        return context


@require_http_methods(["POST"])
@login_required
def sync_balances_view(request):
    """مزامنة أرصدة البنوك والصناديق"""
    # التحقق من صلاحيات المستخدم
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'غير مصرح لك بهذه العملية'})
    
    try:
        from banks.models import BankAccount
        from cashboxes.models import Cashbox
        
        banks_updated = 0
        cashboxes_updated = 0
        
        # مزامنة أرصدة البنوك
        bank_accounts = BankAccount.objects.filter(is_active=True)
        for account in bank_accounts:
            old_balance = account.balance
            new_balance = account.sync_balance()
            if old_balance != new_balance:
                banks_updated += 1
        
        # مزامنة أرصدة الصناديق
        cashboxes = Cashbox.objects.filter(is_active=True)
        for cashbox in cashboxes:
            old_balance = cashbox.balance
            new_balance = cashbox.sync_balance()
            if old_balance != new_balance:
                cashboxes_updated += 1
        
        return JsonResponse({
            'success': True,
            'banks_updated': banks_updated,
            'cashboxes_updated': cashboxes_updated
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def backup_export_view(request):
    """صفحة تصدير النسخة الاحتياطية الجديدة"""
    if not request.user.is_superuser:
        raise PermissionDenied("يجب أن تكون مدير نظام للوصول لهذه الصفحة")
    
    context = {
        'page_title': _('تصدير النسخة الاحتياطية'),
    }
    return render(request, 'core/backup_export.html', context)


@login_required
def export_backup_data(request):
    """تصدير البيانات بصيغ متعددة - JSON, Excel, CSV"""
    if not request.user.is_superuser:
        return JsonResponse({'error': 'يجب أن تكون مدير نظام'}, status=403)
    
    # الحصول على نوع التصدير المطلوب
    export_format = request.GET.get('format', 'json').lower()
    
    try:
        from django.apps import apps
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        # إعداد البيانات للتصدير
        backup_data = {
            'export_info': {
                'version': '2.0',
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'exported_by': request.user.username,
                'system': 'Finspilot Accounting System',
                'export_format': export_format
            },
            'data': {}
        }
        
        total_records = 0
        tables_exported = []
        excluded_apps = ['auth', 'sessions', 'contenttypes', 'admin']
        
        # تصدير جميع البيانات من جميع التطبيقات
        print("🔄 بدء عملية التصدير...")
        
        # قائمة التطبيقات المهمة التي يجب التأكد من تصديرها
        important_apps = ['banks', 'cashboxes', 'receipts', 'payments', 'purchases', 'sales', 'inventory', 'revenues_expenses', 'products', 'customers']
        
        # إجبار إضافة التطبيقات المهمة حتى لو كانت فارغة
        for app_name in important_apps:
            if app_name not in backup_data['data']:
                backup_data['data'][app_name] = {}
        
        for app_config in apps.get_app_configs():
            app_label = app_config.label
            
            # تجنب تصدير بيانات Django الأساسية
            if app_label in excluded_apps:
                continue
            
            is_important = app_label in important_apps
            print(f"📦 معالجة تطبيق: {app_label} {'🎯 (مهم)' if is_important else ''}")
            
            # تأكد من وجود التطبيق في البيانات
            if app_label not in backup_data['data']:
                backup_data['data'][app_label] = {}
            
            models_in_app = list(app_config.get_models())
            print(f"  📝 عدد النماذج في {app_label}: {len(models_in_app)}")
            
            for model in models_in_app:
                model_name = model._meta.model_name
                print(f"  📋 معالجة نموذج: {app_label}.{model_name}")
                
                try:
                    queryset = model.objects.all()
                    model_data = []
                    record_count = queryset.count()
                    print(f"    📊 عدد السجلات: {record_count}")
                    
                    for obj in queryset:
                        obj_data = {}
                        
                        # تصدير جميع الحقول
                        for field in model._meta.fields:
                            field_name = field.name
                            field_value = getattr(obj, field_name, None)
                            
                            # معالجة أنواع البيانات المختلفة
                            if field_value is None:
                                obj_data[field_name] = None
                            elif hasattr(field, 'choices') and field.choices:
                                obj_data[field_name] = field_value
                            elif field.__class__.__name__ == 'DateTimeField':
                                obj_data[field_name] = field_value.isoformat() if field_value else None
                            elif field.__class__.__name__ == 'DateField':
                                obj_data[field_name] = field_value.isoformat() if field_value else None
                            elif field.__class__.__name__ == 'TimeField':
                                obj_data[field_name] = field_value.isoformat() if field_value else None
                            elif field.__class__.__name__ == 'DecimalField':
                                obj_data[field_name] = float(field_value) if field_value else 0.0
                            elif field.__class__.__name__ == 'JSONField':
                                obj_data[field_name] = field_value
                            elif field.__class__.__name__ in ['ForeignKey', 'OneToOneField']:
                                # استخدام sequence_number للفئات والعملاء بدلاً من ID
                                if field_value:
                                    # تحقق إذا كان النموذج المرتبط يحتوي على sequence_number
                                    if hasattr(field_value, 'sequence_number') and field_value.sequence_number:
                                        obj_data[f"{field_name}_sequence"] = field_value.sequence_number
                                        obj_data[f"{field_name}_name"] = str(field_value)
                                        # احتفظ بال ID أيضاً للتوافق مع النماذج القديمة
                                        obj_data[field_name] = field_value.pk
                                    else:
                                        obj_data[field_name] = field_value.pk
                                else:
                                    obj_data[field_name] = None
                            else:
                                obj_data[field_name] = str(field_value) if field_value is not None else None
                        
                        # تصدير العلاقات many-to-many
                        for field in model._meta.many_to_many:
                            field_name = field.name
                            related_objects = getattr(obj, field_name).all()
                            obj_data[field_name] = [related_obj.pk for related_obj in related_objects]
                            obj_data[f"{field_name}_str"] = [str(related_obj) for related_obj in related_objects]
                        
                        model_data.append(obj_data)
                    
                    if model_data:
                        backup_data['data'][app_label][model_name] = model_data
                        total_records += len(model_data)
                        tables_exported.append(f"{app_label}.{model_name}")
                        print(f"    ✅ تم تصدير {len(model_data)} سجل من {app_label}.{model_name}")
                    else:
                        # إضافة النماذج الفارغة للتطبيقات المهمة
                        if is_important:
                            backup_data['data'][app_label][model_name] = []
                            tables_exported.append(f"{app_label}.{model_name}")
                            print(f"    📝 نموذج فارغ تم تضمينه: {app_label}.{model_name}")
                        else:
                            print(f"    ⚠️  لا توجد بيانات في {app_label}.{model_name}")
                
                except Exception as e:
                    error_msg = f"❌ خطأ في تصدير {app_label}.{model_name}: {str(e)}"
                    print(error_msg)
                    print(f"    📝 تفاصيل الخطأ: {type(e).__name__}")
                    
                    # أضف التطبيق حتى لو كان فارغاً لضمان ظهوره في السجل
                    if app_label not in backup_data['data']:
                        backup_data['data'][app_label] = {}
                    backup_data['data'][app_label][f"{model_name}_error"] = {
                        'error': error_msg,
                        'error_type': type(e).__name__,
                        'model_path': f"{app_label}.{model_name}"
                    }
                    continue
        
        # تصدير سجل الأنشطة بشكل خاص لضمان وجوده
        try:
            audit_logs = AuditLog.objects.all().order_by('-timestamp')
            if 'core' not in backup_data['data']:
                backup_data['data']['core'] = {}
            
            backup_data['data']['core']['audit_logs'] = []
            for log in audit_logs:
                backup_data['data']['core']['audit_logs'].append({
                    'id': log.id,
                    'user_id': log.user_id,
                    'user_str': str(log.user) if log.user else None,
                    'action_type': log.action_type,
                    'content_type': log.content_type,
                    'object_id': log.object_id,
                    'description': log.description,
                    'ip_address': getattr(log, 'ip_address', None),
                    'timestamp': log.timestamp.isoformat() if log.timestamp else None,
                })
            
            if backup_data['data']['core']['audit_logs']:
                total_records += len(backup_data['data']['core']['audit_logs'])
                if 'core.auditlog' not in tables_exported:
                    tables_exported.append('core.auditlog')
                print(f"✅ تم تصدير {len(backup_data['data']['core']['audit_logs'])} سجل من سجل الأنشطة")
        except Exception as e:
            print(f"❌ خطأ في تصدير سجل الأنشطة: {str(e)}")
        
        # إحصائيات التصدير
        backup_data['export_info']['total_records'] = total_records
        backup_data['export_info']['tables_exported'] = tables_exported
        backup_data['export_info']['tables_count'] = len(tables_exported)
        
        # تقرير مفصل عن التطبيقات المهمة
        important_apps_status = {}
        for app_name in important_apps:
            if app_name in backup_data['data']:
                app_tables = list(backup_data['data'][app_name].keys())
                app_records = sum([len(v) for v in backup_data['data'][app_name].values() if isinstance(v, list)])
                important_apps_status[app_name] = {
                    'tables': len(app_tables),
                    'records': app_records,
                    'table_names': app_tables
                }
                print(f"📊 {app_name}: {len(app_tables)} جدول, {app_records} سجل")
            else:
                important_apps_status[app_name] = {'tables': 0, 'records': 0, 'table_names': []}
                print(f"❌ {app_name}: غير موجود!")
        
        backup_data['export_info']['important_apps_status'] = important_apps_status
        
        print(f"📊 إجمالي التصدير: {total_records} سجل من {len(tables_exported)} جدول")
        
        # إنشاء اسم الملف
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # إنشاء الاستجابة حسب نوع التصدير
        if export_format == 'json':
            filename = f"finspilot_backup_{timestamp}.json"
            json_content = json.dumps(backup_data, ensure_ascii=False, indent=2)
            
            response = HttpResponse(
                json_content,
                content_type='application/json; charset=utf-8'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
        elif export_format == 'excel' and OPENPYXL_AVAILABLE:
            filename = f"finspilot_backup_{timestamp}.xlsx"
            response = create_excel_backup(backup_data, filename)
            
        elif export_format == 'csv':
            filename = f"finspilot_backup_{timestamp}.csv"
            response = create_csv_backup(backup_data, filename)
            
        else:
            if export_format == 'excel' and not OPENPYXL_AVAILABLE:
                return JsonResponse({'error': 'مكتبة Excel غير متوفرة'}, status=400)
            else:
                return JsonResponse({'error': 'صيغة تصدير غير مدعومة'}, status=400)
        
        # تسجيل في السجل
        try:
            AuditLog.objects.create(
                user=request.user,
                action_type='export',
                content_type='Backup',
                description=f'تصدير نسخة احتياطية ({export_format.upper()}) - {total_records} سجل'
            )
        except:
            pass
        
        return response
        
    except Exception as e:
        return JsonResponse({'error': f'حدث خطأ: {str(e)}'}, status=500)


def create_excel_backup(backup_data, filename):
    """إنشاء ملف Excel للنسخة الاحتياطية"""
    wb = Workbook()
    wb.remove(wb.active)  # حذف الورقة الافتراضية
    
    # تنسيق الرؤوس
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # ورقة المعلومات العامة
    info_ws = wb.create_sheet("معلومات التصدير")
    info_data = [
        ["المعلومة", "القيمة"],
        ["نسخة النظام", backup_data['export_info']['version']],
        ["تاريخ التصدير", backup_data['export_info']['timestamp']],
        ["المستخدم", backup_data['export_info']['exported_by']],
        ["النظام", backup_data['export_info']['system']],
        ["إجمالي السجلات", backup_data['export_info']['total_records']],
        ["عدد الجداول", backup_data['export_info']['tables_count']],
        ["الجداول المصدرة", ', '.join(backup_data['export_info']['tables_exported'][:10])],
    ]
    
    for row_num, row_data in enumerate(info_data, 1):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = info_ws.cell(row=row_num, column=col_num, value=cell_value)
            if row_num == 1:  # رأس الجدول
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
    
    # تصدير كل تطبيق في ورقة منفصلة
    for app_name, app_data in backup_data['data'].items():
        if not app_data or not isinstance(app_data, dict):
            continue
            
        # إنشاء ورقة للتطبيق
        try:
            sheet_name = app_name[:31]  # حد أقصى 31 حرف للورقة
            ws = wb.create_sheet(sheet_name)
            
            current_row = 1
            
            # تصدير كل نموذج في التطبيق
            for model_name, model_data in app_data.items():
                if not model_data or not isinstance(model_data, list):
                    continue
                
                # عنوان النموذج
                ws.cell(row=current_row, column=1, value=f"=== {model_name.upper()} ===").font = Font(bold=True, size=14)
                current_row += 2
                
                if model_data:
                    # رؤوس الأعمدة
                    headers = list(model_data[0].keys())
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=current_row, column=col_num, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    current_row += 1
                    
                    # البيانات
                    for record in model_data:
                        for col_num, header in enumerate(headers, 1):
                            value = record.get(header, '')
                            # تحويل القيم المعقدة إلى نص
                            if isinstance(value, (list, dict)):
                                value = str(value)
                            elif value is None:
                                value = ''
                            ws.cell(row=current_row, column=col_num, value=value)
                        current_row += 1
                
                current_row += 2  # مسافة بين النماذج
        
        except Exception as e:
            print(f"خطأ في إنشاء ورقة {app_name}: {str(e)}")
            continue
    
    # حفظ الملف
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def create_csv_backup(backup_data, filename):
    """إنشاء ملف CSV للنسخة الاحتياطية"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # معلومات التصدير
    writer.writerow(["=== معلومات التصدير ==="])
    writer.writerow(["نسخة النظام", backup_data['export_info']['version']])
    writer.writerow(["تاريخ التصدير", backup_data['export_info']['timestamp']])
    writer.writerow(["المستخدم", backup_data['export_info']['exported_by']])
    writer.writerow(["النظام", backup_data['export_info']['system']])
    writer.writerow(["إجمالي السجلات", backup_data['export_info']['total_records']])
    writer.writerow(["عدد الجداول", backup_data['export_info']['tables_count']])
    writer.writerow([])
    
    # تصدير جميع التطبيقات والنماذج
    for app_name, app_data in backup_data['data'].items():
        if not app_data or not isinstance(app_data, dict):
            continue
            
        writer.writerow([f"=== {app_name.upper()} ==="])
        writer.writerow([])
        
        for model_name, model_data in app_data.items():
            if not model_data or not isinstance(model_data, list):
                continue
                
            writer.writerow([f"--- {model_name} ---"])
            
            if model_data:
                # رؤوس الأعمدة
                headers = list(model_data[0].keys())
                writer.writerow(headers)
                
                # البيانات
                for record in model_data:
                    row = []
                    for header in headers:
                        value = record.get(header, '')
                        # تحويل القيم المعقدة إلى نص
                        if isinstance(value, (list, dict)):
                            value = str(value)
                        elif value is None:
                            value = ''
                        row.append(value)
                    writer.writerow(row)
            
            writer.writerow([])  # مسافة بين النماذج
        
        writer.writerow([])  # مسافة بين التطبيقات
    
    # إنشاء الاستجابة
    content = output.getvalue()
    output.close()
    
    response = HttpResponse(content, content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def import_backup_data(request):
    """استيراد البيانات من ملف النسخة الاحتياطية"""
    if not request.user.is_superuser:
        return JsonResponse({'error': 'يجب أن تكون مدير نظام'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'يجب استخدام POST'}, status=405)
    
    if 'backup_file' not in request.FILES:
        return JsonResponse({'error': 'لم يتم اختيار ملف'}, status=400)
    
    uploaded_file = request.FILES['backup_file']
    file_extension = uploaded_file.name.lower().split('.')[-1]
    
    try:
        result = {'imported_records': 0, 'errors': [], 'success': True}
        
        if file_extension == 'json':
            result = import_from_json(uploaded_file, request.user)
        elif file_extension == 'xlsx':
            result = import_from_excel(uploaded_file, request.user)
        elif file_extension == 'csv':
            result = import_from_csv(uploaded_file, request.user)
        else:
            return JsonResponse({'error': 'صيغة ملف غير مدعومة'}, status=400)
        
        # تسجيل في السجل
        try:
            AuditLog.objects.create(
                user=request.user,
                action_type='import',
                content_type='Backup',
                description=f'استيراد نسخة احتياطية ({file_extension.upper()}) - {result["imported_records"]} سجل'
            )
        except:
            pass
            
        return JsonResponse(result)
        
    except Exception as e:
        return JsonResponse({'error': f'حدث خطأ: {str(e)}', 'success': False}, status=500)


def import_from_json(uploaded_file, user):
    """استيراد البيانات من ملف JSON"""
    try:
        file_content = uploaded_file.read().decode('utf-8')
        data = json.loads(file_content)
        
        # استخدم دالة الاستيراد المشتركة
        return import_from_json_data(data, user)
        
    except json.JSONDecodeError:
        return {'success': False, 'error': 'ملف JSON تالف أو غير صحيح'}
    except Exception as e:
        return {'success': False, 'error': f'خطأ في معالجة ملف JSON: {str(e)}'}
        
        return {
            'success': True,
            'imported_records': imported_count,
            'errors': errors,
            'message': f'تم استيراد {imported_count} سجل بنجاح'
        }
        
    except json.JSONDecodeError:
        return {'success': False, 'error': 'ملف JSON تالف أو غير صحيح'}
    except Exception as e:
        return {'success': False, 'error': f'خطأ في معالجة ملف JSON: {str(e)}'}


def import_from_excel(uploaded_file, user):
    """استيراد البيانات من ملف Excel"""
    if not OPENPYXL_AVAILABLE:
        return {'success': False, 'error': 'مكتبة Excel غير متوفرة'}
    
    try:
        from openpyxl import load_workbook
        from django.apps import apps
        from django.db import transaction
        
        wb = load_workbook(uploaded_file)
        
        imported_count = 0
        errors = []
        
        # تحقق من وجود ورقة المعلومات
        if 'معلومات النسخة' not in wb.sheetnames:
            return {
                'success': False,
                'error': 'ملف XLSX ليس نسخة احتياطية صحيحة من النظام'
            }
        
        # معالجة كل ورقة في الملف
        for sheet_name in wb.sheetnames:
            # تخطي ورقة المعلومات
            if sheet_name == 'معلومات النسخة':
                continue
            
            try:
                ws = wb[sheet_name]
                
                # التحقق من وجود بيانات
                if ws.max_row < 2:  # لا توجد بيانات (فقط الرأس)
                    continue
                
                # تحويل اسم الورقة إلى app.model
                if '_' not in sheet_name:
                    continue
                
                # البحث عن أول _ للفصل بين app و model
                underscore_pos = sheet_name.find('_')
                if underscore_pos == -1:
                    continue
                
                app_label = sheet_name[:underscore_pos]
                model_name = sheet_name[underscore_pos + 1:]
                
                # الحصول على النموذج
                try:
                    model = apps.get_model(app_label, model_name)
                except:
                    continue
                
                # قراءة أسماء الحقول من الصف الأول
                field_names = []
                for col in range(1, ws.max_column + 1):
                    field_name = ws.cell(row=1, column=col).value
                    if field_name:
                        field_names.append(field_name)
                
                if not field_names:
                    continue
                
                # تخطي الجداول التي قد تسبب مشاكل
                skip_models = ['LogEntry', 'Session', 'ContentType', 'Permission']
                if model_name in skip_models:
                    continue
                
                # قراءة البيانات وإدراجها
                for row_num in range(2, ws.max_row + 1):
                    try:
                        # جمع بيانات الصف
                        row_data = {}
                        has_data = False
                        
                        for col, field_name in enumerate(field_names, 1):
                            cell_value = ws.cell(row=row_num, column=col).value
                            if cell_value is not None and str(cell_value).strip() != '':
                                row_data[field_name] = cell_value
                                has_data = True
                        
                        if not has_data:  # صف فارغ تماماً
                            continue
                        
                        # حذف الـ id لإنشاء سجل جديد
                        if 'id' in row_data:
                            del row_data['id']
                        
                        # إنشاء السجل مع معالجة بسيطة للبيانات
                        if row_data:
                            try:
                                # محاولة إنشاء السجل مباشرة
                                model.objects.create(**row_data)
                                imported_count += 1
                            except Exception:
                                # إذا فشل، جرب مع تنظيف البيانات
                                try:
                                    cleaned_data = {}
                                    for field_name, value in row_data.items():
                                        # تنظيف أساسي للبيانات
                                        if value is not None:
                                            # تحويل التواريخ النصية إلى كائنات تاريخ إذا أمكن
                                            if isinstance(value, str) and ('-' in value or '/' in value):
                                                try:
                                                    from django.utils.dateparse import parse_datetime, parse_date
                                                    if 'T' in value or ':' in value:
                                                        parsed_date = parse_datetime(value)
                                                        if parsed_date:
                                                            cleaned_data[field_name] = parsed_date
                                                            continue
                                                    else:
                                                        parsed_date = parse_date(value)
                                                        if parsed_date:
                                                            cleaned_data[field_name] = parsed_date
                                                            continue
                                                except:
                                                    pass
                                            
                                            cleaned_data[field_name] = value
                                    
                                    # محاولة إنشاء السجل مع البيانات المنظفة
                                    if cleaned_data:
                                        model.objects.create(**cleaned_data)
                                        imported_count += 1
                                except Exception:
                                    # تجاهل السجلات التي لا يمكن استيرادها
                                    continue
                                    
                    except Exception:
                        # تجاهل الأخطاء في الصفوف الفردية
                        continue
                
            except Exception:
                # تجاهل أخطاء الأوراق الفردية
                continue
        
        # تسجيل النشاط
        try:
            from core.models import AuditLog
            AuditLog.objects.create(
                user=user,
                action_type='import',
                content_type='Backup',
                description=f'تم استيراد {imported_count} سجل من ملف XLSX'
            )
        except:
            pass
        
        if imported_count > 0:
            return {
                'success': True,
                'imported_records': imported_count,
                'message': f'تم استيراد {imported_count} سجل بنجاح'
            }
        else:
            return {
                'success': False,
                'error': 'لم يتم العثور على بيانات صالحة للاستيراد في الملف'
            }
        
    except Exception as e:
        return {'success': False, 'error': f'خطأ في معالجة ملف Excel: {str(e)}'}


def import_from_json_data(data, user):
    """استيراد البيانات من JSON data object"""
    try:
        imported_count = 0
        errors = []
        
        # التحقق من صحة البيانات
        if 'data' not in data:
            raise ValueError('ملف JSON غير صحيح - لا يحتوي على قسم البيانات')
        
        # استيراد البيانات من جميع التطبيقات
        from django.apps import apps
        
        # الحصول على جميع النماذج المتاحة
        all_models = {}
        for app_config in apps.get_app_configs():
            app_label = app_config.label
            all_models[app_label] = {}
            for model in app_config.get_models():
                model_name = model._meta.model_name
                all_models[app_label][model_name] = model
        
        # استيراد البيانات لكل تطبيق
        for app_label, app_data in data['data'].items():
            if not isinstance(app_data, dict):
                continue
                
            # استيراد البيانات لكل نموذج في التطبيق
            for model_name, model_data in app_data.items():
                if not isinstance(model_data, list) or not model_data:
                    continue
                
                try:
                    # الحصول على النموذج
                    if app_label in all_models and model_name in all_models[app_label]:
                        Model = all_models[app_label][model_name]
                        
                        for record_data in model_data:
                            try:
                                # تنظيف البيانات من الحقول غير المطلوبة
                                cleaned_data = {}
                                for field_name, field_value in record_data.items():
                                    # تجاهل الحقول المُضافة للعرض فقط
                                    if field_name.endswith('_str') or field_name.endswith('_sequence') or field_name.endswith('_name'):
                                        continue
                                    
                                    # معالجة العلاقات many-to-many
                                    if isinstance(field_value, list) and field_name in [f.name for f in Model._meta.many_to_many]:
                                        continue  # سنتعامل معها لاحقاً
                                    
                                    # معالجة الحقول المطلوبة
                                    try:
                                        field_obj = Model._meta.get_field(field_name)
                                        
                                        # معالجة العلاقات الخارجية
                                        if field_obj.is_relation and field_obj.many_to_one:
                                            if field_value and field_name.endswith('_id'):
                                                # التحقق من وجود sequence_number alternative
                                                sequence_field = f"{field_name[:-3]}_sequence"
                                                if sequence_field in record_data:
                                                    # استخدام sequence_number للبحث
                                                    sequence_value = record_data[sequence_field]
                                                    try:
                                                        related_model = field_obj.related_model
                                                        if hasattr(related_model, 'sequence_number'):
                                                            related_obj = related_model.objects.filter(sequence_number=sequence_value).first()
                                                            if related_obj:
                                                                cleaned_data[field_name] = related_obj.id
                                                        else:
                                                            # استخدام ID التقليدي
                                                            if related_model.objects.filter(id=field_value).exists():
                                                                cleaned_data[field_name] = field_value
                                                    except:
                                                        continue
                                                else:
                                                    # استخدام ID التقليدي
                                                    try:
                                                        related_model = field_obj.related_model
                                                        if related_model.objects.filter(id=field_value).exists():
                                                            cleaned_data[field_name] = field_value
                                                    except:
                                                        continue
                                            continue
                                        
                                        # إذا كان الحقل إجبارياً وفارغاً
                                        if not field_obj.null and field_value is None:
                                            if field_name == 'user_id':
                                                # للسجلات التي تتطلب user_id، استخدم المستخدم الحالي
                                                cleaned_data[field_name] = user.id
                                            else:
                                                continue  # تخطي الحقول الفارغة الإجبارية الأخرى
                                        else:
                                            cleaned_data[field_name] = field_value
                                            
                                    except:
                                        # إذا لم نتمكن من الحصول على معلومات الحقل، أضفه كما هو
                                        cleaned_data[field_name] = field_value
                                
                                # إنشاء أو تحديث السجل
                                # استخدام sequence_number للفئات والعملاء بدلاً من ID
                                if hasattr(Model, 'sequence_number') and 'sequence_number' in record_data:
                                    # البحث بالرقم التسلسلي
                                    obj, created = Model.objects.update_or_create(
                                        sequence_number=record_data.get('sequence_number'),
                                        defaults=cleaned_data
                                    )
                                else:
                                    # استخدام ID التقليدي للنماذج الأخرى
                                    obj, created = Model.objects.update_or_create(
                                        id=record_data.get('id'),
                                        defaults=cleaned_data
                                    )
                                
                                imported_count += 1
                                
                            except Exception as e:
                                errors.append(f"خطأ في استيراد سجل من {app_label}.{model_name}: {str(e)}")
                                continue
                        
                except Exception as e:
                    errors.append(f"خطأ في استيراد {app_label}.{model_name}: {str(e)}")
                    continue
        
        return {
            'success': True,
            'imported_records': imported_count,
            'errors': errors,
            'message': f'تم استيراد {imported_count} سجل بنجاح'
        }
        
    except Exception as e:
        return {'success': False, 'error': f'خطأ في معالجة البيانات: {str(e)}'}


def import_from_csv(uploaded_file, user):
    """استيراد البيانات من ملف CSV"""
    try:
        file_content = uploaded_file.read().decode('utf-8-sig')
        
        # حاول العثور على JSON داخل CSV
        lines = file_content.split('\n')
        for line in lines:
            line = line.strip()
            if line.startswith('{') and line.endswith('}'):
                try:
                    json_data = json.loads(line)
                    return import_from_json_data(json_data, user)
                except:
                    continue
        
        return {
            'success': False,
            'error': 'ملف CSV لا يحتوي على بيانات JSON صالحة'
        }
        
    except Exception as e:
        return {'success': False, 'error': f'خطأ في معالجة ملف CSV: {str(e)}'}
