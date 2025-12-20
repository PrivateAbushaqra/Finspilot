from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.http import JsonResponse, HttpResponse
from django.utils.translation import gettext as _
from django.views.decorators.http import require_http_methods
from datetime import date, datetime
from decimal import Decimal
import json

from core.signals import log_view_activity, log_export_activity
from .services import (
    SalesAnalyticsService, 
    PurchaseAnalyticsService, 
    TaxAnalyticsService, 
    CashFlowAnalyticsService,
    AnalyticsService
)


@login_required
def dashboard(request):
    """Main AI Financial Intelligence dashboard"""
    if not request.user.has_perm('analytics.view_ai_dashboard'):
        raise PermissionDenied(_("You do not have permission to access AI Financial Intelligence"))
    
    # Log activity
    log_view_activity(request, 'view', None, _('Accessed AI Financial Intelligence Dashboard'))
    
    context = {
        'title': _('AI Financial Intelligence'),
        'active_menu': 'analytics',
    }
    return render(request, 'analytics/dashboard.html', context)


def parse_date(date_str):
    """Parse date string to date object"""
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except:
        return None


@login_required
@require_http_methods(["GET", "POST"])
def sales_analytics(request):
    """Sales analytics view"""
    if not request.user.has_perm('analytics.view_sales_analytics'):
        raise PermissionDenied(_("You do not have permission to view sales analytics"))
    
    # Get filter parameters
    period_type = request.GET.get('period_type', 'monthly')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    year = request.GET.get('year')
    month = request.GET.get('month')
    quarter = request.GET.get('quarter')
    
    # Parse dates
    if start_date:
        start_date = parse_date(start_date)
    if end_date:
        end_date = parse_date(end_date)
    if year:
        year = int(year)
    if month:
        month = int(month)
    if quarter:
        quarter = int(quarter)
    
    # Get period dates
    start_date, end_date = AnalyticsService.parse_period(
        period_type, start_date, end_date, year, quarter, month
    )
    
    # Create service and get analytics
    service = SalesAnalyticsService(start_date, end_date, request.user)
    overview = service.get_sales_overview()
    products = service.get_product_analytics()
    categories = service.get_category_analytics()
    customers = service.get_customer_analytics()
    representatives = service.get_sales_representative_analytics()
    insights = service.generate_insights(overview, products, customers)
    
    # Log activity
    log_view_activity(request, 'view', None, _('Viewed Sales Analytics'))
    
    # Check if AJAX request for JSON response
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'overview': {k: str(v) for k, v in overview.items()},
            'products': products,
            'categories': categories,
            'customers': customers,
            'representatives': representatives,
            'insights': insights,
        })
    
    context = {
        'title': _('Sales Analytics'),
        'active_menu': 'analytics',
        'overview': overview,
        'products': products,
        'categories': categories,
        'customers': customers,
        'representatives': representatives,
        'insights': insights,
        'start_date': start_date,
        'end_date': end_date,
        'period_type': period_type,
    }
    return render(request, 'analytics/sales_analytics.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def purchase_analytics(request):
    """Purchase analytics view"""
    if not request.user.has_perm('analytics.view_purchase_analytics'):
        raise PermissionDenied(_("You do not have permission to view purchase analytics"))
    
    # Get filter parameters
    period_type = request.GET.get('period_type', 'monthly')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    year = request.GET.get('year')
    month = request.GET.get('month')
    quarter = request.GET.get('quarter')
    
    # Parse dates
    if start_date:
        start_date = parse_date(start_date)
    if end_date:
        end_date = parse_date(end_date)
    if year:
        year = int(year)
    if month:
        month = int(month)
    if quarter:
        quarter = int(quarter)
    
    # Get period dates
    start_date, end_date = AnalyticsService.parse_period(
        period_type, start_date, end_date, year, quarter, month
    )
    
    # Create service and get analytics
    service = PurchaseAnalyticsService(start_date, end_date)
    overview = service.get_purchase_overview()
    products = service.get_product_analytics()
    categories = service.get_category_analytics()
    suppliers = service.get_supplier_analytics()
    insights = service.generate_insights(overview, suppliers)
    
    # Log activity
    log_view_activity(request, 'view', None, _('Viewed Purchase Analytics'))
    
    # Check if AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'overview': {k: str(v) for k, v in overview.items()},
            'products': products,
            'categories': categories,
            'suppliers': suppliers,
            'insights': insights,
        })
    
    context = {
        'title': _('Purchase Analytics'),
        'active_menu': 'analytics',
        'overview': overview,
        'products': products,
        'categories': categories,
        'suppliers': suppliers,
        'insights': insights,
        'start_date': start_date,
        'end_date': end_date,
        'period_type': period_type,
    }
    return render(request, 'analytics/purchase_analytics.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def tax_analytics(request):
    """Tax analytics view"""
    if not request.user.has_perm('analytics.view_tax_analytics'):
        raise PermissionDenied(_("You do not have permission to view tax analytics"))
    
    # Get filter parameters
    period_type = request.GET.get('period_type', 'monthly')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    year = request.GET.get('year')
    month = request.GET.get('month')
    quarter = request.GET.get('quarter')
    
    # Parse dates
    if start_date:
        start_date = parse_date(start_date)
    if end_date:
        end_date = parse_date(end_date)
    if year:
        year = int(year)
    if month:
        month = int(month)
    if quarter:
        quarter = int(quarter)
    
    # Get period dates
    start_date, end_date = AnalyticsService.parse_period(
        period_type, start_date, end_date, year, quarter, month
    )
    
    # Create service and get analytics
    service = TaxAnalyticsService(start_date, end_date)
    overview = service.get_tax_overview()
    by_document_type = service.get_tax_by_document_type()
    insights = service.generate_insights(overview)
    
    # Log activity
    log_view_activity(request, 'view', None, _('Viewed Tax Analytics'))
    
    # Check if AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'overview': {k: str(v) for k, v in overview.items()},
            'by_document_type': by_document_type,
            'insights': insights,
        })
    
    context = {
        'title': _('Tax Analytics'),
        'active_menu': 'analytics',
        'overview': overview,
        'by_document_type': by_document_type,
        'insights': insights,
        'start_date': start_date,
        'end_date': end_date,
        'period_type': period_type,
    }
    return render(request, 'analytics/tax_analytics.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def cashflow_analytics(request):
    """Cash flow analytics view"""
    if not request.user.has_perm('analytics.view_cashflow_analytics'):
        raise PermissionDenied(_("You do not have permission to view cash flow analytics"))
    
    # Get filter parameters
    period_type = request.GET.get('period_type', 'monthly')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    year = request.GET.get('year')
    month = request.GET.get('month')
    quarter = request.GET.get('quarter')
    
    # Parse dates
    if start_date:
        start_date = parse_date(start_date)
    if end_date:
        end_date = parse_date(end_date)
    if year:
        year = int(year)
    if month:
        month = int(month)
    if quarter:
        quarter = int(quarter)
    
    # Get period dates
    start_date, end_date = AnalyticsService.parse_period(
        period_type, start_date, end_date, year, quarter, month
    )
    
    # Create service and get analytics
    service = CashFlowAnalyticsService(start_date, end_date)
    overview = service.get_cashflow_overview()
    bank_accounts = service.get_bank_account_analytics()
    cashboxes = service.get_cashbox_analytics()
    insights = service.generate_insights(overview)
    
    # Log activity
    log_view_activity(request, 'view', None, _('Viewed Cash Flow Analytics'))
    
    # Check if AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'overview': {k: str(v) for k, v in overview.items()},
            'bank_accounts': bank_accounts,
            'cashboxes': cashboxes,
            'insights': insights,
        })
    
    context = {
        'title': _('Cash Flow Analytics'),
        'active_menu': 'analytics',
        'overview': overview,
        'bank_accounts': bank_accounts,
        'cashboxes': cashboxes,
        'insights': insights,
        'start_date': start_date,
        'end_date': end_date,
        'period_type': period_type,
    }
    return render(request, 'analytics/cashflow_analytics.html', context)


@login_required
def export_analytics(request):
    """Export analytics report"""
    if not request.user.has_perm('analytics.export_analytics_reports'):
        raise PermissionDenied(_("You do not have permission to export analytics reports"))
    
    export_type = request.GET.get('type', 'sales')
    format_type = request.GET.get('format', 'xlsx')
    
    # Get filter parameters (similar to above views)
    period_type = request.GET.get('period_type', 'monthly')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    year = request.GET.get('year')
    month = request.GET.get('month')
    quarter = request.GET.get('quarter')
    
    # Parse dates
    if start_date:
        start_date = parse_date(start_date)
    if end_date:
        end_date = parse_date(end_date)
    if year:
        year = int(year)
    if month:
        month = int(month)
    if quarter:
        quarter = int(quarter)
    
    # Get period dates
    start_date, end_date = AnalyticsService.parse_period(
        period_type, start_date, end_date, year, quarter, month
    )
    
    # Log activity
    log_export_activity(request, 'export', None, 
                       _('Exported %(type)s Analytics Report') % {'type': export_type.title()})
    
    if format_type == 'xlsx':
        return export_to_excel(request, export_type, start_date, end_date)
    elif format_type == 'pdf':
        return export_to_pdf(request, export_type, start_date, end_date)
    else:
        return HttpResponse(_('Invalid format'), status=400)


def export_to_excel(request, export_type, start_date, end_date):
    """Export analytics to Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = export_type.title() + ' Analytics'
    
    # Header style
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # Add title
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = f"{export_type.title()} Analytics Report"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Add period
    ws.merge_cells('A2:D2')
    period_cell = ws['A2']
    period_cell.value = f"Period: {start_date} to {end_date}"
    period_cell.alignment = Alignment(horizontal='center')
    
    row = 4
    
    # Get data based on type
    if export_type == 'sales':
        service = SalesAnalyticsService(start_date, end_date, request.user)
        overview = service.get_sales_overview()
        
        # Overview section
        ws[f'A{row}'] = 'Sales Overview'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        for key, value in overview.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = str(value)
            row += 1
        
        row += 2
        
        # Products
        products = service.get_product_analytics()
        ws[f'A{row}'] = 'Top Products'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        ws[f'A{row}'] = 'Product'
        ws[f'B{row}'] = 'Quantity'
        ws[f'C{row}'] = 'Amount'
        row += 1
        
        for product in products[:10]:
            ws[f'A{row}'] = product.get('product__name', 'N/A')
            ws[f'B{row}'] = product.get('total_quantity', 0)
            ws[f'C{row}'] = product.get('total_amount', 0)
            row += 1
    
    elif export_type == 'purchase':
        service = PurchaseAnalyticsService(start_date, end_date)
        overview = service.get_purchase_overview()
        
        ws[f'A{row}'] = 'Purchase Overview'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        for key, value in overview.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = str(value)
            row += 1
    
    elif export_type == 'tax':
        service = TaxAnalyticsService(start_date, end_date)
        overview = service.get_tax_overview()
        
        ws[f'A{row}'] = 'Tax Overview'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        for key, value in overview.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = str(value)
            row += 1
    
    elif export_type == 'cashflow':
        service = CashFlowAnalyticsService(start_date, end_date)
        overview = service.get_cashflow_overview()
        
        ws[f'A{row}'] = 'Cash Flow Overview'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        for key, value in overview.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = str(value)
            row += 1
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={export_type}_analytics_{start_date}_{end_date}.xlsx'
    wb.save(response)
    
    return response


def export_to_pdf(request, export_type, start_date, end_date):
    """Export analytics to PDF"""
    from django.template.loader import render_to_string
    from xhtml2pdf import pisa
    from io import BytesIO
    
    # Get data based on type
    context = {
        'export_type': export_type,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    if export_type == 'sales':
        service = SalesAnalyticsService(start_date, end_date, request.user)
        context['overview'] = service.get_sales_overview()
        context['products'] = service.get_product_analytics()[:10]
    elif export_type == 'purchase':
        service = PurchaseAnalyticsService(start_date, end_date)
        context['overview'] = service.get_purchase_overview()
        context['products'] = service.get_product_analytics()[:10]
    elif export_type == 'tax':
        service = TaxAnalyticsService(start_date, end_date)
        context['overview'] = service.get_tax_overview()
    elif export_type == 'cashflow':
        service = CashFlowAnalyticsService(start_date, end_date)
        context['overview'] = service.get_cashflow_overview()
    
    # Render template
    html = render_to_string('analytics/export_pdf.html', context)
    
    # Create PDF
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename={export_type}_analytics_{start_date}_{end_date}.pdf'
        return response
    
    return HttpResponse(_('Error generating PDF'), status=500)
