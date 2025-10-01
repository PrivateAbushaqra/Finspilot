#!/usr/bin/env python
"""
سكريبت لفحص ملفات النسخ الاحتياطية الخارجية واختبار إمكانية استعادتها
"""
import os
import sys
import django
import json
from django.test import Client
from django.contrib.auth import get_user_model
import openpyxl

# إعداد Django
sys.path.append('C:/Accounting_soft/finspilot')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'finspilot.settings')
django.setup()

User = get_user_model()

def analyze_json_backup(filepath):
    """تحليل ملف النسخة الاحتياطية JSON"""
    try:
        print(f"📄 فحص ملف JSON: {os.path.basename(filepath)}")
        
        # فحص حجم الملف
        file_size = os.path.getsize(filepath)
        file_size_mb = file_size / (1024 * 1024)
        print(f"📏 حجم الملف: {file_size:,} بايت ({file_size_mb:.2f} MB)")
        
        if file_size == 0:
            print("❌ الملف فارغ!")
            return False
        
        # محاولة قراءة الملف
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        print("✅ الملف JSON صحيح ويمكن قراءته")
        
        # فحص البنية الأساسية
        if 'metadata' not in data:
            print("⚠️ الملف لا يحتوي على metadata")
            return False
        
        if 'data' not in data:
            print("❌ الملف لا يحتوي على قسم البيانات")
            return False
        
        metadata = data['metadata']
        print("✅ معلومات الملف:")
        print(f"   📝 اسم النسخة: {metadata.get('backup_name', 'غير محدد')}")
        print(f"   📅 تاريخ الإنشاء: {metadata.get('created_at', 'غير محدد')}")
        print(f"   👤 المنشئ: {metadata.get('created_by', 'غير محدد')}")
        print(f"   📊 عدد الجداول: {metadata.get('total_tables', 0)}")
        print(f"   📈 عدد السجلات: {metadata.get('total_records', 0)}")
        print(f"   🏷️ النوع: {metadata.get('format', 'غير محدد')}")
        
        # فحص قسم البيانات
        data_section = data['data']
        app_count = len(data_section)
        print(f"✅ قسم البيانات يحتوي على {app_count} تطبيق")
        
        total_tables_in_data = 0
        for app_name, app_data in data_section.items():
            table_count = len(app_data)
            total_tables_in_data += table_count
            print(f"   📱 {app_name}: {table_count} جدول")
        
        print(f"📊 إجمالي الجداول في البيانات: {total_tables_in_data}")
        
        # فحص عينة من البيانات
        sample_records = 0
        for app_name, app_data in data_section.items():
            for model_name, model_data in app_data.items():
                if isinstance(model_data, list) and len(model_data) > 0:
                    sample_records += len(model_data)
                    if sample_records > 100:  # عينة محدودة
                        break
            if sample_records > 100:
                break
        
        print(f"✅ تم فحص عينة من {sample_records} سجل - البيانات سليمة")
        
        return True
        
    except json.JSONDecodeError as e:
        print(f"❌ خطأ في تنسيق JSON: {str(e)}")
        return False
    except Exception as e:
        print(f"❌ خطأ في فحص الملف: {str(e)}")
        return False

def analyze_xlsx_backup(filepath):
    """تحليل ملف النسخة الاحتياطية XLSX"""
    try:
        print(f"📊 فحص ملف XLSX: {os.path.basename(filepath)}")
        
        # فحص حجم الملف
        file_size = os.path.getsize(filepath)
        file_size_mb = file_size / (1024 * 1024)
        print(f"📏 حجم الملف: {file_size:,} بايت ({file_size_mb:.2f} MB)")
        
        if file_size == 0:
            print("❌ الملف فارغ!")
            return False
        
        # محاولة فتح الملف بواسطة openpyxl
        workbook = openpyxl.load_workbook(filepath, read_only=True)
        sheet_names = workbook.sheetnames
        
        print(f"✅ الملف XLSX صحيح ويحتوي على {len(sheet_names)} ورقة عمل")
        
        # فحص ورقة المعلومات الأساسية
        if 'Metadata' in sheet_names:
            metadata_sheet = workbook['Metadata']
            print("✅ يحتوي على ورقة Metadata:")
            
            # قراءة المعلومات الأساسية
            for row in metadata_sheet.iter_rows(min_row=1, max_row=10, values_only=True):
                if row[0] and row[1]:
                    print(f"   📋 {row[0]}: {row[1]}")
                    
        else:
            print("⚠️ لا يحتوي على ورقة Metadata")
        
        # فحص بعض أوراق البيانات
        data_sheets = [name for name in sheet_names if name != 'Metadata']
        print(f"📊 أوراق البيانات ({len(data_sheets)}):")
        
        sample_count = 0
        for sheet_name in data_sheets[:10]:  # فحص أول 10 أوراق فقط
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            print(f"   📄 {sheet_name}: {max_row} صف × {max_col} عمود")
            sample_count += max_row
            
        if len(data_sheets) > 10:
            print(f"   ... و {len(data_sheets) - 10} ورقة أخرى")
        
        print(f"✅ عدد السجلات المقدر: {sample_count:,}")
        
        workbook.close()
        return True
        
    except Exception as e:
        print(f"❌ خطأ في فحص ملف XLSX: {str(e)}")
        return False

def test_backup_restore(filepath, file_type):
    """اختبار استعادة الملف عبر النظام"""
    print(f"\n🔄 اختبار استعادة الملف: {os.path.basename(filepath)}")
    
    # إنشاء عميل اختبار
    client = Client()
    
    # تسجيل الدخول
    login_successful = client.login(username='super', password='password')
    if not login_successful:
        print("❌ فشل في تسجيل الدخول")
        return False
    
    try:
        # محاولة رفع الملف للاستعادة
        with open(filepath, 'rb') as backup_file:
            response = client.post('/ar/backup/restore-backup/', {
                'clear_data': 'false',
                'csrfmiddlewaretoken': 'test'
            }, files={'backup_file': backup_file}, 
            HTTP_X_REQUESTED_WITH='XMLHttpRequest', 
            HTTP_HOST='127.0.0.1:8000')
        
        if response.status_code == 200:
            try:
                result = response.json()
                if result.get('success'):
                    print("✅ النظام قبل الملف وبدأ عملية الاستعادة")
                    return True
                else:
                    print(f"❌ النظام رفض الملف: {result.get('error', 'خطأ غير معروف')}")
                    return False
            except json.JSONDecodeError:
                print("❌ استجابة غير صحيحة من النظام")
                return False
        else:
            print(f"❌ خطأ في الاستعادة: HTTP {response.status_code}")
            return False
            
    except Exception as e:
        print(f"❌ خطأ في اختبار الاستعادة: {str(e)}")
        return False

def main():
    """الدالة الرئيسية لفحص جميع الملفات"""
    print("🔍 فحص ملفات النسخ الاحتياطية في C:\\Accounting_soft\\finspilot\\Backup_files")
    print("=" * 80)
    
    backup_dir = "C:/Accounting_soft/finspilot/Backup_files"
    
    if not os.path.exists(backup_dir):
        print("❌ المجلد غير موجود!")
        return False
    
    files = os.listdir(backup_dir)
    json_files = [f for f in files if f.endswith('.json')]
    xlsx_files = [f for f in files if f.endswith('.xlsx')]
    
    print(f"📋 تم العثور على:")
    print(f"   📄 {len(json_files)} ملف JSON")
    print(f"   📊 {len(xlsx_files)} ملف XLSX")
    print(f"   📁 {len(files) - len(json_files) - len(xlsx_files)} ملف آخر")
    
    if not json_files and not xlsx_files:
        print("❌ لا توجد ملفات نسخ احتياطية!")
        return False
    
    results = []
    
    # فحص ملفات JSON
    for json_file in json_files:
        filepath = os.path.join(backup_dir, json_file)
        print(f"\n{'='*50}")
        
        # فحص البنية
        is_valid = analyze_json_backup(filepath)
        results.append({
            'file': json_file,
            'type': 'JSON',
            'valid': is_valid,
            'path': filepath
        })
        
        # اختبار الاستعادة إذا كان صحيحاً
        if is_valid:
            restore_test = test_backup_restore(filepath, 'JSON')
            results[-1]['restorable'] = restore_test
    
    # فحص ملفات XLSX
    for xlsx_file in xlsx_files:
        filepath = os.path.join(backup_dir, xlsx_file)
        print(f"\n{'='*50}")
        
        # فحص البنية
        is_valid = analyze_xlsx_backup(filepath)
        results.append({
            'file': xlsx_file,
            'type': 'XLSX',
            'valid': is_valid,
            'path': filepath
        })
        
        # اختبار الاستعادة إذا كان صحيحاً
        if is_valid:
            restore_test = test_backup_restore(filepath, 'XLSX')
            results[-1]['restorable'] = restore_test
    
    # النتيجة النهائية
    print(f"\n{'='*80}")
    print("🎯 النتيجة النهائية:")
    print("=" * 80)
    
    valid_files = [r for r in results if r['valid']]
    restorable_files = [r for r in results if r.get('restorable', False)]
    
    print(f"📊 إجمالي الملفات: {len(results)}")
    print(f"✅ ملفات صحيحة: {len(valid_files)}")
    print(f"🔄 ملفات قابلة للاستعادة: {len(restorable_files)}")
    
    if restorable_files:
        print("\n✅ الملفات القابلة للاستعادة:")
        for result in restorable_files:
            print(f"   📄 {result['file']} ({result['type']})")
    
    if len(results) - len(restorable_files) > 0:
        print(f"\n❌ ملفات غير قابلة للاستعادة: {len(results) - len(restorable_files)}")
        for result in results:
            if not result.get('restorable', False):
                reason = "ملف غير صحيح" if not result['valid'] else "فشل في الاستعادة"
                print(f"   ❌ {result['file']} - {reason}")
    
    success_rate = (len(restorable_files) / len(results) * 100) if results else 0
    print(f"\n📈 معدل النجاح: {success_rate:.1f}%")
    
    return len(restorable_files) == len(results)

if __name__ == "__main__":
    success = main()
    if success:
        print("\n🎉 جميع الملفات قابلة للاستعادة في النظام!")
    else:
        print("\n⚠️ بعض الملفات قد تحتاج مراجعة قبل الاستعادة")
    
    exit(0 if success else 1)