from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, TemplateView, View, DetailView
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import datetime, timedelta
from django.http import JsonResponse
from django.contrib import messages
from django.urls import reverse_lazy
from django.utils.translation import gettext as _
from .models import Warehouse, InventoryMovement, WarehouseTransfer
from products.models import Product

class InventoryListView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'inventory/list.html'
    
    def test_func(self):
        return self.request.user.has_inventory_permission()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get sorting parameters
        sort_by = self.request.GET.get('sort', 'product_name')
        sort_direction = self.request.GET.get('dir', 'asc')
        
        # Get filtering parameters
        search = self.request.GET.get('search', '')
        stock_level_filter = self.request.GET.get('stock_level', '')
        
        # Get all warehouses
        warehouses = Warehouse.objects.filter(is_active=True)
        context['total_warehouses'] = warehouses.count()
        
        # Get all products
        products = Product.objects.filter(is_active=True)
        context['total_products'] = products.count()
        
        # Calculate inventory statistics
        total_current_stock = 0
        low_stock_items = []
        
        for product in products:
            # Calculate current stock for this product
            in_movements = InventoryMovement.objects.filter(
                product=product,
                movement_type='in'
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            out_movements = InventoryMovement.objects.filter(
                product=product,
                movement_type='out'
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            current_stock = in_movements - out_movements + product.opening_balance_quantity
            
            # Add to total stock if positive
            if current_stock > 0:
                total_current_stock += current_stock
            
            # Check for low stock (stock <= 10 and >= 0)
            if current_stock <= 10 and current_stock >= 0:
                low_stock_items.append({
                    'product': product,
                    'current_stock': current_stock
                })
        
        context['total_items'] = int(total_current_stock)
        context['low_stock_items'] = len(low_stock_items)
        
        # Get today's movements
        today = timezone.now().date()
        today_movements = InventoryMovement.objects.filter(
            date=today
        ).count()
        context['today_movements'] = today_movements
        
        # Get inventory items with current stock levels
        inventory_items = []
        for product in products:  # Show all products
            # Calculate current stock for this product
            in_movements = InventoryMovement.objects.filter(
                product=product,
                movement_type='in'
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            out_movements = InventoryMovement.objects.filter(
                product=product,
                movement_type='out'
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            current_stock = in_movements - out_movements + product.opening_balance_quantity
            
            # Determine stock level
            if current_stock <= 0:
                stock_level = 'out'
            elif current_stock <= 5:
                stock_level = 'critical'
            elif current_stock <= 20:
                stock_level = 'low'
            else:
                stock_level = 'good'
            
            # --- تعديل احتساب القيمة بناءً على تكلفة الشراء الفعلية المتبقية من فواتير الشراء مع تضمين الرصيد الافتتاحي ---
            from purchases.models import PurchaseInvoiceItem
            from decimal import Decimal
            remaining_qty = Decimal(current_stock)
            value = Decimal('0.0')
            
            # أولاً، استخدم الرصيد الافتتاحي إذا كان موجوداً
            if product.opening_balance_quantity > 0 and remaining_qty > 0:
                used_qty = min(Decimal(product.opening_balance_quantity), remaining_qty)
                value += used_qty * Decimal(product.opening_balance_cost)
                remaining_qty -= used_qty
            
            # ثم، جلب عناصر فواتير الشراء الأقدم فالأحدث
            purchase_items = PurchaseInvoiceItem.objects.filter(product=product).order_by('invoice__date', 'invoice__id')
            for item in purchase_items:
                if remaining_qty <= 0:
                    break
                used_qty = min(item.quantity, remaining_qty)
                value += used_qty * item.unit_price
                remaining_qty -= used_qty
            # إذا بقي كمية لم يتم تغطيتها من فواتير الشراء (حالة بيانات ناقصة)، استخدم سعر التكلفة الافتراضي
            if remaining_qty > 0:
                value += remaining_qty * Decimal(product.cost_price)
            # --- نهاية التعديل ---
            inventory_items.append({
                'product_id': product.id,
                'product_name': product.name,
                'product_code': product.code,
                'quantity': current_stock,
                'value': float(value),
                'sale_price': float(product.sale_price),
                'warehouse_name': 'المستودع الرئيسي',
                'stock_level': stock_level
            })
        
        # Sort inventory items
        if sort_by in ['product_name', 'product_code', 'quantity', 'value', 'stock_level', 'warehouse_name']:
            reverse_sort = sort_direction == 'desc'
            inventory_items = sorted(inventory_items, key=lambda x: x.get(sort_by, ''), reverse=reverse_sort)
        
        # Filter inventory items
        if search:
            inventory_items = [item for item in inventory_items if search.lower() in item['product_name'].lower() or search.lower() in item['product_code'].lower()]
        
        if stock_level_filter:
            inventory_items = [item for item in inventory_items if item['stock_level'] == stock_level_filter]
        
        context['inventory_items'] = inventory_items
        context['current_sort'] = sort_by
        context['sort_direction'] = sort_direction
        context['search'] = search
        context['stock_level_filter'] = stock_level_filter
        
        # Get recent movements
        recent_movements = []
        movements = InventoryMovement.objects.select_related('product').order_by('-created_at')[:5]
        for movement in movements:
            recent_movements.append({
                'product_name': movement.product.name,
                'type': movement.movement_type,
                'quantity': movement.quantity,
                'date': movement.created_at
            })
        
        context['recent_movements'] = recent_movements
        
        # Calculate totals for quick stats
        total_in = InventoryMovement.objects.filter(
            movement_type='in'
        ).aggregate(total=Sum('total_cost'))['total'] or 0
        
        total_out = InventoryMovement.objects.filter(
            movement_type='out'
        ).aggregate(total=Sum('total_cost'))['total'] or 0
        
        # Calculate totals with tax
        total_in_with_tax = 0
        total_out_with_tax = 0
        
        # Calculate incoming with tax
        in_movements = InventoryMovement.objects.filter(movement_type='in').select_related('product')
        for movement in in_movements:
            cost_without_tax = movement.total_cost
            tax_rate = movement.product.tax_rate if movement.product else 0
            from decimal import Decimal
            cost_without_tax_decimal = Decimal(str(cost_without_tax))
            tax_rate_decimal = Decimal(str(tax_rate))
            tax_amount = cost_without_tax_decimal * (tax_rate_decimal / Decimal('100'))
            total_in_with_tax += cost_without_tax_decimal + tax_amount
        
        # Calculate outgoing with tax
        out_movements = InventoryMovement.objects.filter(movement_type='out').select_related('product')
        for movement in out_movements:
            cost_without_tax = movement.total_cost
            tax_rate = movement.product.tax_rate if movement.product else 0
            from decimal import Decimal
            cost_without_tax_decimal = Decimal(str(cost_without_tax))
            tax_rate_decimal = Decimal(str(tax_rate))
            tax_amount = cost_without_tax_decimal * (tax_rate_decimal / Decimal('100'))
            total_out_with_tax += cost_without_tax_decimal + tax_amount
        
        context['total_in_value'] = total_in
        context['total_out_value'] = total_out
        # تعديل جذري: احتساب إجمالي قيمة المخزون كمجموع القيم الفعلية لكل منتج (FIFO)
        context['total_inventory_value'] = sum(item['value'] for item in inventory_items)

        # Add values with tax (تبقى كما هي)
        context['total_in_value_with_tax'] = total_in_with_tax
        context['total_out_value_with_tax'] = total_out_with_tax
        context['total_inventory_value_with_tax'] = total_in_with_tax - total_out_with_tax

        return context

class WarehouseListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Warehouse
    template_name = 'inventory/warehouse_list.html'
    context_object_name = 'warehouses'
    
    def test_func(self):
        return self.request.user.has_inventory_permission()
    
    def get_queryset(self):
        # استثناء المستودع الافتراضي للنظام
        return Warehouse.objects.exclude(code='MAIN')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculate statistics (استثناء المستودع الافتراضي للنظام)
        all_warehouses = Warehouse.objects.exclude(code='MAIN')
        context['total_warehouses'] = all_warehouses.count()
        context['active_warehouses'] = all_warehouses.filter(is_active=True).count()
        context['sub_warehouses'] = all_warehouses.filter(parent__isnull=False).count()
        
        # This month's movements
        current_month = timezone.now().replace(day=1)
        month_movements = InventoryMovement.objects.filter(
            date__gte=current_month
        ).count()
        context['month_movements'] = month_movements
        
        return context

class WarehouseCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Warehouse
    template_name = 'inventory/warehouse_add.html'
    fields = ['name', 'code', 'address', 'parent', 'manager', 'is_active', 'is_default']
    success_url = reverse_lazy('inventory:warehouse_list')
    
    def test_func(self):
        return self.request.user.has_inventory_permission()
    
    def form_valid(self, form):
        # تسجيل النشاط
        from core.signals import log_user_activity
        warehouse = form.save()
        log_user_activity(
            request=self.request,
            action_type='CREATE',
            obj=warehouse,
            description=f'تم إنشاء المستودع: {warehouse.name}'
        )
        messages.success(self.request, 'تم إنشاء المستودع بنجاح')
        return super().form_valid(form)

class WarehouseDeleteView(LoginRequiredMixin, UserPassesTestMixin, View):
    template_name = 'inventory/warehouse_delete.html'
    
    def test_func(self):
        return self.request.user.has_inventory_permission()
    
    def get(self, request, pk, *args, **kwargs):
        warehouse = get_object_or_404(Warehouse, pk=pk)
        
        # فحص البيانات المرتبطة
        related_data = {
            'inventory_movements': 0,
            'transfers_from': 0,
            'transfers_to': 0,
            'active_products': 0
        }
        
        try:
            # فحص حركات المخزون
            movements = InventoryMovement.objects.filter(warehouse=warehouse)
            related_data['inventory_movements'] = movements.count()
            
            # فحص التحويلات من/إلى هذا المستودع
            try:
                transfers_from = WarehouseTransfer.objects.filter(from_warehouse=warehouse)
                transfers_to = WarehouseTransfer.objects.filter(to_warehouse=warehouse)
                related_data['transfers_from'] = transfers_from.count()
                related_data['transfers_to'] = transfers_to.count()
            except:
                pass
            
            # حساب المنتجات النشطة في هذا المستودع
            active_products = set()
            for movement in movements.filter(movement_type='in'):
                active_products.add(movement.product.id)
            related_data['active_products'] = len(active_products)
            
        except Exception as e:
            # في حالة وجود خطأ، لا نمنع عرض الصفحة
            pass
        
        context = {
            'warehouse': warehouse,
            'related_data': related_data,
            'can_delete': not any(related_data.values()),
            'has_sub_warehouses': warehouse.sub_warehouses.exists() if hasattr(warehouse, 'sub_warehouses') else False
        }
        return render(request, self.template_name, context)
    
    def post(self, request, pk, *args, **kwargs):
        warehouse = get_object_or_404(Warehouse, pk=pk)
        
        try:
            # التحقق من تأكيد الحذف
            confirm = request.POST.get('confirm_delete')
            if confirm != 'DELETE':
                messages.error(request, 'يجب كتابة "DELETE" للتأكيد!')
                return redirect('inventory:warehouse_delete', pk=pk)
            
            # التحقق من وجود بيانات مرتبطة
            movements_count = InventoryMovement.objects.filter(warehouse=warehouse).count()
            if movements_count > 0:
                messages.error(
                    request, 
                    f'لا يمكن حذف المستودع "{warehouse.name}" لأنه يحتوي على {movements_count} حركة مخزون. '
                    'يجب حذف جميع الحركات المرتبطة أولاً.'
                )
                return redirect('inventory:warehouse_delete', pk=pk)
            
            # التحقق من المستودعات الفرعية
            if hasattr(warehouse, 'sub_warehouses') and warehouse.sub_warehouses.exists():
                sub_count = warehouse.sub_warehouses.count()
                messages.error(
                    request, 
                    f'لا يمكن حذف المستودع "{warehouse.name}" لأنه يحتوي على {sub_count} مستودع فرعي. '
                    'يجب حذف أو نقل المستودعات الفرعية أولاً.'
                )
                return redirect('inventory:warehouse_delete', pk=pk)
            
            # حفظ البيانات للرسالة
            warehouse_name = warehouse.name
            warehouse_code = warehouse.code
            
            # حذف المستودع
            warehouse.delete()
            
            # رسالة نجاح
            messages.success(
                request, 
                f'تم حذف المستودع "{warehouse_name}" (رمز: {warehouse_code}) بنجاح!\n'
                f'جميع البيانات المرتبطة تم حذفها نهائياً.'
            )
            
            return redirect('inventory:warehouse_list')
            
        except Exception as e:
            messages.error(request, f'حدث خطأ أثناء حذف المستودع: {str(e)}')
            return redirect('inventory:warehouse_delete', pk=pk)

class MovementListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = InventoryMovement
    template_name = 'inventory/movement_list.html'
    context_object_name = 'movements'
    paginate_by = 20
    ordering = ['-date', '-created_at']
    
    def test_func(self):
        return self.request.user.has_inventory_permission()
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Apply filters if provided
        movement_type = self.request.GET.get('movement_type')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        warehouse_id = self.request.GET.get('warehouse')
        
        if movement_type:
            queryset = queryset.filter(movement_type=movement_type)
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)
            
        return queryset.select_related('product', 'warehouse')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Calculate statistics
        all_movements = InventoryMovement.objects.all()
        total_in = all_movements.filter(movement_type='in').aggregate(
            total=Sum('quantity'))['total'] or 0
        total_out = all_movements.filter(movement_type='out').aggregate(
            total=Sum('quantity'))['total'] or 0
        total_transfers = all_movements.filter(movement_type='transfer').count()
        today_movements = all_movements.filter(date=timezone.now().date()).count()
        context.update({
            'total_in': total_in,
            'total_out': total_out,
            'total_transfers': total_transfers,
            'today_movements': today_movements,
            'warehouses': Warehouse.objects.filter(is_active=True),
            'is_superadmin': getattr(self.request.user, 'is_superadmin', False),
        })
        
        # تسجيل النشاط في سجل الأنشطة
        from core.models import AuditLog
        AuditLog.objects.create(
            user=self.request.user,
            action_type='view',
            content_type='inventory_movements_list',
            description=_('عرض قائمة حركات المخزون مع عمود رقم المستند'),
            ip_address=self.request.META.get('REMOTE_ADDR')
        )
        
        return context

class TransferListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = WarehouseTransfer
    template_name = 'inventory/transfer_list.html'
    context_object_name = 'transfers'
    
    def test_func(self):
        return self.request.user.has_inventory_permission()
    
    def get_queryset(self):
        return WarehouseTransfer.objects.select_related('from_warehouse', 'to_warehouse', 'created_by').order_by('-transfer_date', '-id')

class TransferCreateView(LoginRequiredMixin, UserPassesTestMixin, View):
    template_name = 'inventory/transfer_add.html'
    
    def test_func(self):
        return self.request.user.has_inventory_permission()
    
    def get(self, request, *args, **kwargs):
        context = {
            'warehouses': Warehouse.objects.filter(is_active=True).exclude(code='MAIN'),
            'products': Product.objects.filter(is_active=True)
        }
        return render(request, self.template_name, context)
    
    def post(self, request, *args, **kwargs):
        try:
            from_warehouse_id = request.POST.get('from_warehouse')
            to_warehouse_id = request.POST.get('to_warehouse')
            transfer_date = request.POST.get('transfer_date')
            notes = request.POST.get('notes', '')
            
            if from_warehouse_id == to_warehouse_id:
                messages.error(request, 'لا يمكن التحويل من وإلى نفس المستودع!')
                return self.get(request)
            
            from_warehouse = get_object_or_404(Warehouse, id=from_warehouse_id, is_active=True)
            to_warehouse = get_object_or_404(Warehouse, id=to_warehouse_id, is_active=True)
            
            # إنشاء التحويل
            transfer = WarehouseTransfer.objects.create(
                from_warehouse=from_warehouse,
                to_warehouse=to_warehouse,
                transfer_date=transfer_date,
                notes=notes,
                created_by=request.user
            )
            
            # معالجة العناصر
            product_ids = request.POST.getlist('product_id[]')
            quantities = request.POST.getlist('quantity[]')
            
            for i, product_id in enumerate(product_ids):
                if product_id and quantities[i]:
                    product = get_object_or_404(Product, id=product_id)
                    quantity = float(quantities[i])
                    
                    # التحقق من توفر الكمية في المستودع المصدر
                    current_stock = product.get_stock_in_warehouse(from_warehouse)
                    if current_stock < quantity:
                        messages.error(request, f'الكمية المتاحة للمنتج {product.name} في المستودع {from_warehouse.name} هي {current_stock}')
                        transfer.delete()  # حذف التحويل إذا فشل
                        return self.get(request)
                    
                    # إنشاء عنصر التحويل
                    WarehouseTransferItem.objects.create(
                        transfer=transfer,
                        product=product,
                        quantity=quantity
                    )
                    
                    # إنشاء حركات المخزون
                    # حركة صادرة من المستودع المصدر
                    InventoryMovement.objects.create(
                        product=product,
                        warehouse=from_warehouse,
                        movement_type='out',
                        quantity=quantity,
                        unit_cost=product.cost_price,
                        total_cost=quantity * product.cost_price,
                        reference_type='warehouse_transfer',
                        reference_id=transfer.id,
                        notes=f'تحويل إلى {to_warehouse.name}',
                        created_by=request.user,
                        date=transfer.transfer_date
                    )
                    
                    # حركة واردة إلى المستودع الهدف
                    InventoryMovement.objects.create(
                        product=product,
                        warehouse=to_warehouse,
                        movement_type='in',
                        quantity=quantity,
                        unit_cost=product.cost_price,
                        total_cost=quantity * product.cost_price,
                        reference_type='warehouse_transfer',
                        reference_id=transfer.id,
                        notes=f'تحويل من {from_warehouse.name}',
                        created_by=request.user,
                        date=transfer.transfer_date
                    )
            
            # تسجيل النشاط
            AuditLog.objects.create(
                user=request.user,
                action_type='create',
                content_type='WarehouseTransfer',
                object_id=transfer.id,
                description=f'تم إنشاء تحويل مخزون من {from_warehouse.name} إلى {to_warehouse.name}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'تم إنشاء التحويل بنجاح!')
            return redirect('inventory:transfer_list')
            
        except Exception as e:
            messages.error(request, f'حدث خطأ أثناء إنشاء التحويل: {str(e)}')
            return self.get(request)

class LowStockView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/low_stock.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get all products with their current stock levels
        products = Product.objects.filter(is_active=True)
        low_stock_items = []
        
        out_of_stock_count = 0
        critical_stock_count = 0
        low_stock_count = 0
        
        for product in products:
            # Calculate current stock for this product
            in_movements = InventoryMovement.objects.filter(
                product=product,
                movement_type='in'
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            out_movements = InventoryMovement.objects.filter(
                product=product,
                movement_type='out'
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            current_stock = in_movements - out_movements
            min_quantity = 10  # Default minimum quantity
            
            # Determine stock level
            if current_stock <= 0:
                level = 'out'
                out_of_stock_count += 1
            elif current_stock <= 5:
                level = 'critical'
                critical_stock_count += 1
            elif current_stock <= min_quantity:
                level = 'low'
                low_stock_count += 1
            else:
                continue  # Skip products with good stock levels
            
            # Calculate stock percentage
            stock_percentage = min(100, (current_stock / min_quantity) * 100) if min_quantity > 0 else 0
            
            # Get last movement date
            last_movement = InventoryMovement.objects.filter(
                product=product
            ).order_by('-date', '-created_at').first()
            
            low_stock_items.append({
                'product_id': product.id,
                'product_name': product.name,
                'product_code': product.code,
                'current_quantity': current_stock,
                'min_quantity': min_quantity,
                'level': level,
                'stock_percentage': max(0, stock_percentage),
                'warehouse_name': 'المستودع الرئيسي',
                'last_movement_date': last_movement.date if last_movement else None,
            })
        
        # Sort by level (out -> critical -> low)
        level_priority = {'out': 0, 'critical': 1, 'low': 2}
        low_stock_items.sort(key=lambda x: level_priority.get(x['level'], 3))
        
        context.update({
            'low_stock_items': low_stock_items,
            'out_of_stock': out_of_stock_count,
            'critical_stock': critical_stock_count,
            'low_stock': low_stock_count,
            'total_products': products.count(),
        })
        
        return context

class ProductInventoryDetailView(LoginRequiredMixin, TemplateView):
    """عرض تفاصيل المنتج في المخزون"""
    template_name = 'inventory/product_detail.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        product_id = kwargs.get('product_id')
        if not product_id:
            # If no product_id provided, get from URL parameter
            product_id = self.request.GET.get('product_id')
        
        if product_id:
            try:
                product = Product.objects.get(id=product_id, is_active=True)
                
                # Calculate current stock for this product
                in_movements = InventoryMovement.objects.filter(
                    product=product,
                    movement_type='in'
                ).aggregate(total=Sum('quantity'))['total'] or 0
                
                out_movements = InventoryMovement.objects.filter(
                    product=product,
                    movement_type='out'
                ).aggregate(total=Sum('quantity'))['total'] or 0
                
                current_stock = in_movements - out_movements
                
                # Determine stock level
                if current_stock <= 0:
                    stock_level = 'out'
                    stock_level_text = 'نفد المخزون'
                    stock_level_class = 'stock-out'
                elif current_stock <= 5:
                    stock_level = 'critical'
                    stock_level_text = 'مخزون حرج'
                    stock_level_class = 'stock-critical'
                elif current_stock <= 20:
                    stock_level = 'low'
                    stock_level_text = 'مخزون منخفض'
                    stock_level_class = 'stock-low'
                else:
                    stock_level = 'good'
                    stock_level_text = 'مخزون جيد'
                    stock_level_class = 'stock-good'
                
                # Get recent movements for this product
                recent_movements_qs = InventoryMovement.objects.filter(
                    product=product
                ).select_related('warehouse').order_by('-created_at')[:10]
                recent_movements = []
                for movement in recent_movements_qs:
                    recent_movements.append({
                        'movement_type': movement.movement_type,
                        'quantity': movement.quantity,
                        'warehouse': movement.warehouse,
                        'created_at': movement.created_at,
                        'unit_cost': movement.unit_cost,
                        'total_cost': movement.total_cost,
                        'notes': movement.notes,
                    })
                
                # Calculate value statistics
                total_in_value = InventoryMovement.objects.filter(
                    product=product,
                    movement_type='in'
                ).aggregate(total=Sum('total_cost'))['total'] or 0
                
                total_out_value = InventoryMovement.objects.filter(
                    product=product,
                    movement_type='out'
                ).aggregate(total=Sum('total_cost'))['total'] or 0
                
                # Get warehouses where this product exists
                warehouses_with_stock = []
                for warehouse in Warehouse.objects.filter(is_active=True):
                    warehouse_in = InventoryMovement.objects.filter(
                        product=product,
                        warehouse=warehouse,
                        movement_type='in'
                    ).aggregate(total=Sum('quantity'))['total'] or 0
                    
                    warehouse_out = InventoryMovement.objects.filter(
                        product=product,
                        warehouse=warehouse,
                        movement_type='out'
                    ).aggregate(total=Sum('quantity'))['total'] or 0
                    
                    warehouse_stock = warehouse_in - warehouse_out
                    
                    if warehouse_stock > 0:
                        warehouses_with_stock.append({
                            'warehouse': warehouse,
                            'quantity': warehouse_stock
                        })
                
                context.update({
                    'product': product,
                    'current_stock': current_stock,
                    'total_in': in_movements,
                    'total_out': out_movements,
                    'stock_level': stock_level,
                    'stock_level_text': stock_level_text,
                    'stock_level_class': stock_level_class,
                    'recent_movements': recent_movements,
                    'total_in_value': total_in_value,
                    'total_out_value': total_out_value,
                    'current_value': total_in_value - total_out_value,
                    'warehouses_with_stock': warehouses_with_stock,
                })
                
            except Product.DoesNotExist:
                context['error'] = 'المنتج غير موجود'
        else:
            context['error'] = 'يجب تحديد معرف المنتج'
        
        return context

def get_product_inventory_ajax(request):
    """Get product inventory details via AJAX"""
    product_id = request.GET.get('product_id')
    
    if not product_id:
        return JsonResponse({'error': 'معرف المنتج مطلوب'}, status=400)
    
    try:
        product = Product.objects.get(id=product_id, is_active=True)
        
        # Calculate current stock
        in_movements = InventoryMovement.objects.filter(
            product=product,
            movement_type='in'
        ).aggregate(total=Sum('quantity'))['total'] or 0
        
        out_movements = InventoryMovement.objects.filter(
            product=product,
            movement_type='out'
        ).aggregate(total=Sum('quantity'))['total'] or 0
        
        current_stock = in_movements - out_movements
        
        # Determine stock level
        if current_stock <= 0:
            stock_level = 'out'
            stock_level_text = 'نفد المخزون'
        elif current_stock <= 5:
            stock_level = 'critical'
            stock_level_text = 'مخزون حرج'
        elif current_stock <= 20:
            stock_level = 'low'
            stock_level_text = 'مخزون منخفض'
        else:
            stock_level = 'good'
            stock_level_text = 'مخزون جيد'
        
        # Get recent movements
        recent_movements = InventoryMovement.objects.filter(
            product=product
        ).select_related('warehouse').order_by('-created_at')[:5]
        
        movements_data = []
        for movement in recent_movements:
            movements_data.append({
                'type': movement.movement_type,
                'type_text': 'وارد' if movement.movement_type == 'in' else 'صادر' if movement.movement_type == 'out' else 'تحويل',
                'quantity': movement.quantity,
                'warehouse': movement.warehouse.name if movement.warehouse else 'غير محدد',
                'date': movement.created_at.strftime('%Y-%m-%d %H:%M'),
                'notes': movement.notes or ''
            })
        
        data = {
            'product_name': product.name,
            'product_code': product.code,
            'category': product.category.name if product.category else '',
            'sale_price': float(product.sale_price),
            'current_stock': current_stock,
            'total_in': in_movements,
            'total_out': out_movements,
            'stock_level': stock_level,
            'stock_level_text': stock_level_text,
            'recent_movements': movements_data
        }
        
        return JsonResponse(data)
        
    except Product.DoesNotExist:
        return JsonResponse({'error': 'المنتج غير موجود'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

class MovementBulkDeleteView(LoginRequiredMixin, View):
    """حذف جميع حركات المخزون"""
    
    def get(self, request, *args, **kwargs):
        """عرض صفحة تأكيد حذف جميع الحركات"""
        context = {
            'total_movements': InventoryMovement.objects.count(),
            'movements_by_type': {
                'in': InventoryMovement.objects.filter(movement_type='in').count(),
                'out': InventoryMovement.objects.filter(movement_type='out').count(),
                'transfer': InventoryMovement.objects.filter(movement_type='transfer').count(),
                'adjustment': InventoryMovement.objects.filter(movement_type='adjustment').count(),
            }
        }
        return render(request, 'inventory/movement_bulk_delete.html', context)
    
    def post(self, request, *args, **kwargs):
        """تنفيذ حذف جميع حركات المخزون"""
        try:
            # عد الحركات قبل الحذف
            total_count = InventoryMovement.objects.count()
            
            if total_count == 0:
                messages.warning(request, 'لا توجد حركات مخزون للحذف!')
                return redirect('inventory:movement_list')
            
            # التحقق من تأكيد المستخدم
            confirm = request.POST.get('confirm_delete')
            if confirm != 'DELETE_ALL_MOVEMENTS':
                messages.error(request, 'يجب كتابة "DELETE_ALL_MOVEMENTS" للتأكيد!')
                return redirect('inventory:movement_bulk_delete')
            
            # حذف جميع حركات المخزون
            deleted_count, deleted_details = InventoryMovement.objects.all().delete()
            
            # رسالة نجاح
            messages.success(
                request, 
                f'تم حذف {total_count} حركة مخزون بنجاح!\n'
                f'تفاصيل الحذف: {deleted_details}'
            )
            
        except Exception as e:
            messages.error(request, f'حدث خطأ أثناء حذف حركات المخزون: {str(e)}')
        
        return redirect('inventory:movement_list')

class WarehouseDetailView(LoginRequiredMixin, DetailView):
    model = Warehouse
    template_name = 'inventory/warehouse_detail.html'
    context_object_name = 'warehouse'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        warehouse = self.get_object()
        
        # إحصائيات المستودع
        stats = {
            'total_movements': 0,
            'in_movements': 0,
            'out_movements': 0,
            'total_products': 0,
            'active_products': 0
        }
        
        try:
            movements = InventoryMovement.objects.filter(warehouse=warehouse)
            stats['total_movements'] = movements.count()
            stats['in_movements'] = movements.filter(movement_type='in').count()
            stats['out_movements'] = movements.filter(movement_type='out').count()
            
            # حساب المنتجات النشطة
            active_products = set()
            for movement in movements:
                active_products.add(movement.product.id)
            stats['active_products'] = len(active_products)
            
        except Exception as e:
            pass
        
        context['stats'] = stats
        return context


class WarehouseEditView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Warehouse
    fields = ['name', 'code', 'address', 'is_active', 'is_default']
    template_name = 'inventory/warehouse_edit.html'
    success_url = '/ar/inventory/warehouses/'
    
    def test_func(self):
        return self.request.user.has_inventory_permission()
    
    def form_valid(self, form):
        # تسجيل النشاط
        from core.signals import log_user_activity
        warehouse = form.save()
        log_user_activity(
            request=self.request,
            action_type='UPDATE',
            obj=warehouse,
            description=f'تم تحديث المستودع: {warehouse.name}'
        )
        messages.success(self.request, f'تم تحديث المستودع "{form.instance.name}" بنجاح!')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'حدث خطأ أثناء تحديث المستودع. يرجى المحاولة مرة أخرى.')
        return super().form_invalid(form)


class MovementDeleteView(LoginRequiredMixin, UserPassesTestMixin, View):
    """حذف حركة مخزون واحدة"""
    
    def test_func(self):
        return self.request.user.has_inventory_permission()
    
    def post(self, request, pk, *args, **kwargs):
        """حذف حركة المخزون"""
        try:
            # الحصول على الحركة
            movement = get_object_or_404(InventoryMovement, pk=pk)
            movement_number = movement.movement_number
            
            # تسجيل النشاط قبل الحذف
            from core.signals import log_activity
            log_activity(
                user=request.user,
                action_type='DELETE',
                obj=movement,
                description=f'تم حذف حركة مخزون رقم: {movement_number}',
                request=request
            )
            
            # حذف الحركة
            movement.delete()
            
            messages.success(
                request, 
                f'تم حذف حركة المخزون رقم {movement_number} بنجاح'
            )
            
            return JsonResponse({
                'success': True, 
                'message': f'تم حذف حركة المخزون رقم {movement_number} بنجاح'
            })
            
        except InventoryMovement.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'message': 'حركة المخزون غير موجودة'
            }, status=404)
            
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'message': f'حدث خطأ أثناء حذف الحركة: {str(e)}'
            }, status=500)


@login_required
def export_inventory_excel(request):
    """تصدير قائمة المخزون إلى Excel"""
    from django.http import HttpResponse
    from django.utils.translation import gettext_lazy as _
    from core.signals import log_export_activity
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from decimal import Decimal

    # التحقق من الصلاحيات
    if not request.user.has_inventory_permission() and not request.user.is_superuser:
        from django.contrib import messages
        from django.shortcuts import redirect
        messages.error(request, 'ليس لديك صلاحية لعرض المخزون')
        return redirect('/')

    # الحصول على معايير البحث والترتيب من الطلب
    sort_by = request.GET.get('sort', 'product_name')
    sort_direction = request.GET.get('dir', 'asc')
    search = request.GET.get('search', '')
    stock_level_filter = request.GET.get('stock_level', '')

    # الحصول على بيانات المخزون (نفس منطق InventoryListView)
    from products.models import Product

    products = Product.objects.filter(is_active=True)

    inventory_items = []
    for product in products:
        # Calculate current stock for this product
        in_movements = InventoryMovement.objects.filter(
            product=product,
            movement_type='in'
        ).aggregate(total=Sum('quantity'))['total'] or 0

        out_movements = InventoryMovement.objects.filter(
            product=product,
            movement_type='out'
        ).aggregate(total=Sum('quantity'))['total'] or 0

        current_stock = in_movements - out_movements

        # Determine stock level
        if current_stock <= 0:
            stock_level = 'out'
        elif current_stock <= 5:
            stock_level = 'critical'
        elif current_stock <= 20:
            stock_level = 'low'
        else:
            stock_level = 'good'

        # Calculate value based on actual purchase invoice costs
        from purchases.models import PurchaseInvoiceItem
        remaining_qty = Decimal(current_stock)
        value = Decimal('0.0')
        # Get purchase invoice items in oldest to newest order
        purchase_items = PurchaseInvoiceItem.objects.filter(product=product).order_by('invoice__date', 'invoice__id')
        for item in purchase_items:
            if remaining_qty <= 0:
                break
            used_qty = min(item.quantity, remaining_qty)
            value += used_qty * item.unit_price
            remaining_qty -= used_qty
        # If there's remaining quantity not covered by purchase invoices (data gap), use default cost price
        if remaining_qty > 0:
            value += remaining_qty * Decimal(product.cost_price)

        inventory_items.append({
            'product_id': product.id,
            'product_name': product.name,
            'product_code': product.code,
            'quantity': current_stock,
            'value': float(value),
            'sale_price': float(product.sale_price),
            'warehouse_name': 'المستودع الرئيسي',
            'stock_level': stock_level
        })

    # تطبيق الترتيب
    if sort_by in ['product_name', 'product_code', 'quantity', 'value', 'stock_level', 'warehouse_name']:
        reverse_sort = sort_direction == 'desc'
        inventory_items = sorted(inventory_items, key=lambda x: x.get(sort_by, ''), reverse=reverse_sort)

    # تطبيق الفلاتر
    if search:
        inventory_items = [item for item in inventory_items if search.lower() in item['product_name'].lower() or search.lower() in item['product_code'].lower()]

    if stock_level_filter:
        inventory_items = [item for item in inventory_items if item['stock_level'] == stock_level_filter]

    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = str(_('Inventory List'))

    # تنسيق العناوين
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    # كتابة العناوين
    headers = [
        _('Product Name'),
        _('Product Code'),
        _('Warehouse'),
        _('Quantity'),
        _('Value'),
        _('Sale Price'),
        _('Stock Status')
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # كتابة البيانات
    for row_num, item in enumerate(inventory_items, 2):
        ws.cell(row=row_num, column=1, value=item['product_name'])
        ws.cell(row=row_num, column=2, value=item['product_code'])
        ws.cell(row=row_num, column=3, value=item['warehouse_name'])
        ws.cell(row=row_num, column=4, value=item['quantity'])
        ws.cell(row=row_num, column=5, value=item['value'])
        ws.cell(row=row_num, column=6, value=item['sale_price'])

        # تحديد حالة المخزون باللغة العربية
        stock_status = ''
        if item['stock_level'] == 'good':
            stock_status = _('Good Stock')
        elif item['stock_level'] == 'low':
            stock_status = _('Low Stock')
        elif item['stock_level'] == 'critical':
            stock_status = _('Critical Stock')
        else:
            stock_status = _('Out of Stock')

        ws.cell(row=row_num, column=7, value=str(stock_status))

    # تعديل عرض الأعمدة
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # حد أقصى 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # إنشاء الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'inventory_list_{timezone.now().date()}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # حفظ الملف
    wb.save(response)

    # تسجيل النشاط
    log_export_activity(request, str(_('Inventory List')), filename, 'Excel')
